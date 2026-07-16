# test_data 전체 대상 배치 테스트: 각 요청을 Qwen3.6(현재 프롬프트)로 생성 → 휴리스틱 점검.
# 결과를 _batch_results.txt 에 한 줄씩(플러시) 기록. 백그라운드 실행 후 파일로 확인.
import sys, os, re, glob, time, traceback
sys.path.insert(0, os.path.dirname(__file__))
import _qwen_client as q
from openpyxl import load_workbook

HERE = os.path.dirname(__file__)
TD = os.path.join(HERE, "..", "test_data")
REQ_FILE = os.path.join(HERE, "_batch_requests.txt")
OUT_FILE = os.path.join(HERE, "_batch_results.txt")

def build_schema_all():
    out = ["## 현재 파일 스키마"]
    for p in sorted(glob.glob(os.path.join(TD, "*.xlsx"))):
        if os.path.basename(p).startswith("~$"):
            continue
        try:
            wb = load_workbook(p, data_only=True, read_only=True)
        except Exception:
            continue
        out.append("\n### 파일: %s" % os.path.basename(p))
        out.append("시트: " + ", ".join(wb.sheetnames))
        for sn in wb.sheetnames[:12]:
            ws = wb[sn]
            rows = []
            for i, row in enumerate(ws.iter_rows(values_only=True)):
                rows.append(row);
                if i >= 2: break
            if rows:
                ncol = max((len(r) for r in rows), default=0)
                out.append("  [%s] ~%s행x%s열" % (sn, ws.max_row, ncol))
                out.append("    1행: " + " | ".join([("" if v is None else str(v))[:18] for v in rows[0][:16]]))
                for r in rows[1:3]:
                    out.append("    샘플: " + " | ".join([("" if v is None else str(v))[:14] for v in r[:16]]))
        wb.close()
    return "\n".join(out)

FORBIDDEN = re.compile(r"load_workbook|\bws\s*\[|\bimport\s|\bos\.|win32com|\.Select\s*\(|\.Activate\s*\(|ActiveWorkbook|while\s+True|\.Save(?:As)?\s*\(|\.Close\s*\(|\bopen\s*\(|\beval\s*\(")
SIGNSTRIP = re.compile(r"\babs\s*\(|replace\(\s*[\"']-[\"']|re\.sub\(\s*r?[\"']\[\^0-9")
WIDEREAD = re.compile(r"\.read\([^,\n]+,\s*[\"']?[A-Z]{1,3}:[A-Z]{1,3}[\"']?\s*\)")

def loop_write(code):
    lines = code.splitlines()
    for i, ln in enumerate(lines):
        if re.match(r"\s*(for|while)\b", ln):
            indent = len(ln) - len(ln.lstrip())
            for j in range(i+1, min(i+15, len(lines))):
                l2 = lines[j]
                if l2.strip() and (len(l2)-len(l2.lstrip())) <= indent:
                    break
                if re.search(r"ctx\.(write|write_cell|write_formulas)\s*\(", l2):
                    return True
    return False

def extract_code(text):
    m = re.search(r"```(?:python)?\s*(.*?)```", text, re.S)
    return m.group(1) if m else ""

def analyze(req, schema, system_base):
    r = q.chat(system_base + "\n\n" + schema, req, max_tokens=1300, temperature=0.0,
               extra={"chat_template_kwargs": {"enable_thinking": False}})
    content = r["content"] or ""
    code = extract_code(content)
    has_code = bool(code.strip())
    has_transform = bool(re.search(r"def\s+transform\s*\(\s*ctx", code))
    helpers = sorted(set(re.findall(r"ctx(?:\.book\([^)]*\))?\.(\w+)\s*\(", code)) - {"book"})
    flags = []
    if FORBIDDEN.search(code): flags.append("FORBIDDEN")
    if loop_write(code): flags.append("LOOPWRITE")
    if WIDEREAD.search(code): flags.append("WIDEREAD")
    if SIGNSTRIP.search(code): flags.append("SIGNSTRIP")
    if not has_transform:
        status = "ASK/NOCODE"
    elif "FORBIDDEN" in flags:
        status = "FORBIDDEN"
    elif flags:
        status = "WARN"
    else:
        status = "OK"
    return status, helpers, flags, has_code, content[:120]

def main():
    reqs = [l.strip() for l in open(REQ_FILE, encoding="utf-8") if l.strip()]
    schema = build_schema_all()
    system_base = open(os.path.join(HERE, "_prompt_current.txt"), encoding="utf-8").read()
    counts = {}
    with open(OUT_FILE, "w", encoding="utf-8") as f:
        f.write("BATCH %d requests | schema %d자 | prompt %d자\n" % (len(reqs), len(schema), len(system_base)))
        f.flush()
        for idx, req in enumerate(reqs, 1):
            try:
                status, helpers, flags, has_code, head = analyze(req, schema, system_base)
            except Exception as e:
                status, helpers, flags, head = "ERROR", [], [str(e)[:40]], ""
            counts[status] = counts.get(status, 0) + 1
            line = "%3d [%-10s] %-22s %s | %s" % (
                idx, status, ",".join(helpers)[:22], ",".join(flags), req[:46])
            f.write(line + "\n"); f.flush()
        f.write("\n=== SUMMARY ===\n")
        for k in sorted(counts):
            f.write("  %-12s %d\n" % (k, counts[k]))
        f.flush()

if __name__ == "__main__":
    main()
