# -*- coding: utf-8 -*-
"""[0.8.2 녹화 검증] 복붙 캡처로 만든 스텝이 '보통 스텝처럼' 사는가 — 실제 Excel.

검증 요청(2026-08-31)
  1) 기존 파이프라인 수정처럼 동작 가능한가
  2) on/off(켜고 끄기)가 유효한가
  3) 그 스킬을 전체실행해도 문제없이 도는가
  4) 특히 파일 교차 실행 / 시트명 / 파일명 불일치가 없는가
  5) 그 밖에 녹화로 만든 것 특유의 문제

방법: 화면 조작을 흉내내지 않고 **진짜로** Range.Copy → 붙여넣기를 한 뒤
      캡처 API(_capture_copypaste_on_session_impl)를 불러 '녹화가 실제로 뱉는 코드'를 얻는다.
      그 코드를 그대로 전체실행 경로에 태워 결과를 값으로 확인한다.
      (코드를 손으로 지어내면 '녹화가 만든 것'을 검증하는 게 아니다)
"""
import sys, io, tempfile, shutil, time
from pathlib import Path
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")
ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT))
import openpyxl
import pythoncom; pythoncom.CoInitialize()
import win32com.client as win32
import serve_b2b as S

fails = []
notes = []


def check(name, cond, detail=""):
    print(("  PASS  " if cond else "  FAIL  ") + name + (("  -> " + str(detail)[:300]) if not cond else ""))
    if not cond:
        fails.append(name)


work = Path(tempfile.mkdtemp(prefix="b2b_rec_"))
SRC_NAME, DST_NAME = "원본자료.xlsx", "보고서양식.xlsx"


def build():
    a = openpyxl.Workbook(); ws = a.active; ws.title = "회선 현황"
    ws.append(["가입번호", "상태", "요금"])
    for i in range(5):
        ws.append([512102400000 + i, "정지" if i % 2 == 0 else "정상", 1000 + i])
    a.create_sheet("메모")        # 시트가 하나뿐이면 CSV 용 '단일 시트 안전망'이 걸려
    a.save(str(work / SRC_NAME))   # 시트명 검증이 통째로 무의미해진다(실측으로 확인)
    b = openpyxl.Workbook(); wb2 = b.active; wb2.title = "집계"
    wb2["A1"] = "붙여넣을 자리"
    b.create_sheet("표지")
    b.save(str(work / DST_NAME))


build()
app = None
try:
    app = win32.DispatchEx("Excel.Application")
    app.Visible = False
    app.DisplayAlerts = False
    wsrc = app.Workbooks.Open(str(work / SRC_NAME))
    wdst = app.Workbooks.Open(str(work / DST_NAME))
    S.EXCEL_SESSIONS["recSrc"] = {"app": app, "workbook": wsrc, "path": str(work / SRC_NAME),
                                  "openPath": str(wsrc.FullName), "name": SRC_NAME,
                                  "sourcePath": str(work / SRC_NAME), "liveEditable": True}
    S.EXCEL_SESSIONS["recDst"] = {"app": app, "workbook": wdst, "path": str(work / DST_NAME),
                                  "openPath": str(wdst.FullName), "name": DST_NAME,
                                  "sourcePath": str(work / DST_NAME), "liveEditable": True}

    print("[1] 진짜 복붙을 한 뒤 캡처시킨다 (교차파일: 원본자료 → 보고서양식)")
    src_ws = wsrc.Worksheets("회선 현황")
    dst_ws = wdst.Worksheets("집계")
    src_ws.Range("A1:C6").Copy()                 # Ctrl+C 에 해당
    dst_ws.Activate()
    dst_ws.Range("A3").Select()                  # 붙여넣을 자리 선택
    dst_ws.Paste()                               # Ctrl+V 에 해당
    time.sleep(0.3)
    cap = S._capture_copypaste_on_session_impl("recDst", False)
    ok_cap = bool(cap) and bool(cap.get("code"))
    check("캡처가 스텝 코드를 만들어 냈다", ok_cap, cap)
    if not ok_cap:
        raise SystemExit(1)
    code = cap["code"]
    print("      생성 코드:", code.strip().splitlines()[-1][:120])
    check("교차파일이 코드에 드러난다(src_book / dst_book)",
          "src_book=" in code and "dst_book=" in code, code)
    check("파일명이 사용자에게 보이는 이름 그대로",
          SRC_NAME in code and DST_NAME in code, code)
    check("시트명이 원문 그대로", "회선 현황" in code and "집계" in code, code)
    check("대상 파일이 targetFileId 로 고정된다(붙여넣은 쪽)",
          str(cap.get("targetExcelId") or cap.get("excelId") or "recDst") == "recDst", cap.get("targetExcelId"))

    print("[2] 정적 게이트를 통과한다 (보통 스텝과 같은 관문)")
    try:
        S._python_com_static_check(code)
        check("정적 검사 통과", True)
    except Exception as e:
        check("정적 검사 통과", False, e)

    print("[3] 그 코드를 그대로 실행하면 값이 옮겨진다")
    # 붙여넣기 흔적을 지우고 코드만으로 재현되는지 본다
    dst_ws.Range("A3:C10").ClearContents()
    ctx = S.PythonComSkillContext(app, wdst, {"name": DST_NAME}, timeout_s=300)
    g = {"__builtins__": dict(S._PY_SAFE_BUILTINS)}
    exec(compile(code, "<rec>", "exec"), g)
    g["transform"](ctx)
    c2 = S.PythonComSkillContext(app, wdst, {"name": DST_NAME}, timeout_s=300)
    head = c2.read_cell("집계", "A3")
    val = c2.read_cell("집계", "A4")
    check("머리글이 옮겨졌다", head == "가입번호", head)
    check("데이터가 옮겨졌다", str(int(val)) == "512102400000", val)
    check("양식의 기존 칸은 그대로", c2.read_cell("집계", "A1") == "붙여넣을 자리")

    print("[4] 전체실행(실행기 경로)에서 돈다 — 교차파일 그대로")
    dst_ws.Range("A3:C10").ClearContents()
    groups = [{"excelId": "recDst", "steps": [{
        "code": code, "language": "python", "stepIdx": 0, "stepId": "rec1",
        "targetSheetName": "집계",
    }]}]
    out = S.run_full_pipeline_single_instance(groups, reset_excel_ids=["recSrc", "recDst"])
    check("전체실행 성공", bool(out) and out.get("ok") and out.get("applied") == 1, out)
    c3 = S.PythonComSkillContext(app, wdst, {"name": DST_NAME}, timeout_s=300)
    check("전체실행 결과가 라이브에 반영", c3.read_cell("집계", "A3") == "가입번호",
          c3.read_cell("집계", "A3"))
    check("교차 참조한 원본은 안 망가졌다",
          S.PythonComSkillContext(app, wsrc, {"name": SRC_NAME}, timeout_s=300)
           .read_cell("회선 현황", "A1") == "가입번호")

    print("[5] 파일명·시트명 불일치 — 실행기 매핑이 잡아낼 수 있는 형태인가")
    import re
    books = sorted(set(re.findall(r"_book=['\"]([^'\"]+)['\"]", code)))
    sheets = re.findall(r"paste_copied\(\s*['\"]([^'\"]+)['\"]\s*,\s*['\"][^'\"]+['\"]\s*,\s*['\"]([^'\"]+)['\"]", code)
    check("코드에서 파일명 2개를 뽑아낼 수 있다", books == sorted([SRC_NAME, DST_NAME]), books)
    check("코드에서 (원본시트, 대상시트) 쌍을 뽑아낼 수 있다",
          bool(sheets) and sheets[0] == ("회선 현황", "집계"), sheets)
    # 실행기가 '이 스킬이 요구하는 파일'로 무엇을 뽑는지 — 녹화 스텝도 같은 규칙을 타야 한다
    check("파일명이 코드 안에 리터럴로 있다(매핑 치환 대상)",
          ("'%s'" % SRC_NAME) in code or ('"%s"' % SRC_NAME) in code, code[:160])

    print("[6] 시트명이 안 맞으면 조용히 틀리지 않고 사유를 말한다(시트 2개 이상 기준)")
    bad = code.replace("'집계'", "'없는시트'").replace('"집계"', '"없는시트"')
    ctxb = S.PythonComSkillContext(app, wdst, {"name": DST_NAME}, timeout_s=300)
    gb = {"__builtins__": dict(S._PY_SAFE_BUILTINS)}
    exec(compile(bad, "<recbad>", "exec"), gb)
    try:
        gb["transform"](ctxb)
        check("없는 시트 → 오류로 알린다", False, "오류 없이 통과했다(조용한 오답 위험)")
    except Exception as e:
        check("없는 시트 → 사유 있는 오류", "시트" in str(e) or "찾" in str(e), str(e)[:160])

    print("[7] 값만 붙여넣기(values_only) 캡처도 같은 규칙을 따른다")
    # 전체실행이 라이브를 다시 만들어 앞서 잡아 둔 시트 참조가 무효가 된다(실측: OLE 0x800a01a8).
    # 세션에서 워크북을 다시 얻어 시트를 새로 잡는다.
    wsrc2 = S.EXCEL_SESSIONS["recSrc"]["workbook"]
    wdst2 = S.EXCEL_SESSIONS["recDst"]["workbook"]
    src_ws = wsrc2.Worksheets("회선 현황")
    dst_ws = wdst2.Worksheets("집계")
    src_ws.Range("A1:C3").Copy()
    dst_ws.Activate(); dst_ws.Range("E3").Select(); dst_ws.Paste()
    time.sleep(0.3)
    cap2 = S._capture_copypaste_on_session_impl("recDst", True)
    ok2 = bool(cap2) and bool(cap2.get("code"))
    check("값만 붙여넣기 캡처 성공", ok2, cap2)
    if ok2:
        check("values_only=True 가 코드에 있다", "values_only=True" in cap2["code"], cap2["code"])
        try:
            S._python_com_static_check(cap2["code"])
            check("값만 붙여넣기도 정적 검사 통과", True)
        except Exception as e:
            check("값만 붙여넣기도 정적 검사 통과", False, e)
finally:
    for k in ("recSrc", "recDst"):
        S.EXCEL_SESSIONS.pop(k, None)
    try:
        if app is not None:
            for w in list(app.Workbooks):
                try: w.Close(SaveChanges=False)
                except Exception: pass
            app.Quit()
    except Exception:
        pass
    shutil.rmtree(work, ignore_errors=True)

print("")
for n in notes:
    print("  참고:", n)
print("RESULT: ALL PASS" if not fails else "RESULT: %d FAIL -> %s" % (len(fails), fails))
sys.exit(0 if not fails else 1)
