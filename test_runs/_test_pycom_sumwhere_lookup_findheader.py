# [0.5.18] find_header(인접행/병합헤더) + ctx.sum_where(조건집계) + ctx.sum_lookup(다중토큰 매칭) 실파일 검증.
# 원본 미변경(출력은 임시복사본에만 write). 기대값은 openpyxl + serve_b2b 자체 헬퍼로 산출해 비교.
import sys, io, os, glob, tempfile, shutil
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")
from pathlib import Path
sys.path.insert(0, str(Path(__file__).resolve().parents[1]))
import win32com.client as win32
import serve_b2b as s
from openpyxl import load_workbook

KT = r"C:\Users\Admin\Downloads\KT"
OUT = (glob.glob(os.path.join(KT, "output_HCN*.xlsx")) or [None])[0]
IN = (glob.glob(os.path.join(KT, "input_작업파일_03. 관악_03월.xlsx")) or [None])[0]
OUT = next((p for p in glob.glob(os.path.join(KT, "output_HCN*.xlsx")) if not os.path.basename(p).startswith("~$")), None)
IN = next((p for p in glob.glob(os.path.join(KT, "input*관악*.xlsx")) if not os.path.basename(p).startswith("~$")), None)

passed = 0
def check(name, cond, detail=""):
    global passed
    if not cond:
        raise AssertionError(f"FAIL {name}: {detail}")
    passed += 1
    print(" OK ", name)

def expected_sum_where():
    ws = load_workbook(OUT, data_only=True)["SO사업자별요금"]
    tot = 0.0
    for r in range(6, ws.max_row + 1):
        d, g, m = ws.cell(r, 4).value, ws.cell(r, 7).value, ws.cell(r, 13).value
        if s._cond_match(d, "==", "인터넷") and s._cond_match(g, "==", "매 출"):
            num = s._coerce_number(m)
            if num is not None:
                tot += num
    return tot

def expected_lookup():
    wi = load_workbook(IN, data_only=True)["Sheet1"]
    kmap = {}
    for r in range(2, wi.max_row + 1):
        val = s._coerce_number(wi.cell(r, 69).value)   # BQ=69
        if val is None:
            continue
        for tok in s._split_key_tokens(wi.cell(r, 68).value):  # BP=68
            kmap[tok] = kmap.get(tok, 0.0) + val
    wo = load_workbook(OUT, data_only=True)["SO사업자별요금"]
    exp = {}
    for r in range(6, wo.max_row + 1):
        toks = s._split_key_tokens(wo.cell(r, 16).value)   # P=16
        if not toks:
            continue
        ssum, hit = 0.0, False
        for t in toks:
            if t in kmap:
                ssum += kmap[t]; hit = True
        if hit:
            exp[r] = round(ssum, 4)
    return exp, len(kmap)

def main():
    if not OUT or not IN:
        print("실파일 없음 — 건너뜀:", OUT, IN); return
    app = win32.DispatchEx("Excel.Application"); app.Visible = False; app.DisplayAlerts = False
    wbOut = wbIn = wbTmp = None
    try:
        # ── 1) find_header + sum_where : 출력 파일 ReadOnly ──
        wbOut = app.Workbooks.Open(OUT, ReadOnly=True)
        ctx = s.PythonComSkillContext(app, wbOut, {"path": OUT, "workbook": wbOut, "app": app})
        # SBAGENT-19: 국제는 K5. header_row=8 로 줘도 인접행 스캔으로 K(11) 찾아야 함
        col = ctx.find_header("SO사업자별요금", "국제", header_row=8)
        check("find_header('국제', header_row=8) → K열(11) (인접행 구제)", col == 11, str(col))
        check("find_header 정상행(H4:M4 아래 '기본료' 5행)도 정확", ctx.find_header("SO사업자별요금", "기본료", header_row=5) == 8,
              str(ctx.find_header("SO사업자별요금", "기본료", header_row=5)))
        # SBAGENT-35: 조건부 집계(광폭 read 없이)
        exp_w = expected_sum_where()
        got_w = ctx.sum_where("SO사업자별요금", "M", [("D", "인터넷"), ("G", "매 출")], header_row=5)
        check(f"sum_where(D=인터넷 AND G=매출 → M) == openpyxl({exp_w})", abs(got_w - exp_w) < 1.0, f"{got_w} vs {exp_w}")
        check("sum_where 결과가 0 이상(실제 집계됨)", got_w >= 0)
        wbOut.Close(SaveChanges=False); wbOut = None

        # ── 2) sum_lookup : input ReadOnly + output 임시복사본(편집) ──
        tmp = os.path.join(tempfile.gettempdir(), "b2b_lookup_out.xlsx")
        shutil.copyfile(OUT, tmp)
        wbIn = app.Workbooks.Open(IN, ReadOnly=True)
        wbTmp = app.Workbooks.Open(tmp)
        ctx2 = s.PythonComSkillContext(app, wbTmp, {"path": tmp, "workbook": wbTmp, "app": app})
        wsH = wbTmp.Worksheets("SO사업자별요금")
        wsIn = wbIn.Worksheets("Sheet1")
        # 기대값은 COM '라이브' 값으로 산출한다(BQ='기본료변환'은 수식 → openpyxl 캐시와 달라짐).
        in_last = int(wsIn.Cells(wsIn.Rows.Count, 68).End(-4162).Row)
        kmap = {}
        for r in range(2, in_last + 1):
            val = s._coerce_number(wsIn.Cells(r, 69).Value)      # BQ live
            if val is None:
                continue
            for tok in s._split_key_tokens(wsIn.Cells(r, 68).Value):  # BP live
                kmap[tok] = kmap.get(tok, 0.0) + val
        # 출력 P(16) 토큰으로 기대 매칭행/값 산출(COM)
        out_last = max(int(wsH.Cells(wsH.Rows.Count, 16).End(-4162).Row), 179)
        exp_map = {}
        for r in range(6, out_last + 1):
            toks = s._split_key_tokens(wsH.Cells(r, 16).Value)
            hit = [kmap[t] for t in toks if t in kmap]
            if hit:
                exp_map[r] = round(sum(hit), 4)
        rep = ctx2.sum_lookup(f"{os.path.basename(IN)}!Sheet1", "BP", "BQ", "SO사업자별요금", "P", "H",
                              header_row=1, dst_start_row=6)
        got = {r: s._coerce_number(wsH.Cells(r, 8).Value) for r in exp_map}
        print("보고:", rep, "| COM 기대 매칭행:", len(exp_map), "| src_keys:", rep["src_keys"])
        check("sum_lookup 소스 키 수 > 0", rep["src_keys"] > 0, str(rep["src_keys"]))
        check("sum_lookup filled > 0 (VBA 0건 → 해결)", rep["filled"] > 0, str(rep["filled"]))
        check("sum_lookup filled == COM 기대 매칭행", rep["filled"] == len(exp_map), f"{rep['filled']} vs {len(exp_map)}")
        ok_vals = all(got.get(r) is not None and abs(got[r] - ev) < 0.01 for r, ev in exp_map.items())
        check("매칭행 H 값이 전부 기대(COM)와 일치", ok_vals,
              detail=str([(r, got.get(r), ev) for r, ev in list(exp_map.items())[:5]]))

        print(f"\n=== RESULT: {passed} PASS / 0 FAIL ===")
    finally:
        for wb in (wbTmp, wbIn, wbOut):
            try:
                if wb is not None: wb.Close(SaveChanges=False)
            except Exception: pass
        try: app.Quit()
        except Exception: pass

if __name__ == "__main__":
    main()
