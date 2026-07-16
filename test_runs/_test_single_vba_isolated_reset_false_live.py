# -*- coding: utf-8 -*-
# [라이브] 단일 VBA 적용 크래시 수정 검증: 클라가 이제 단일 VBA 도 /api/excel/run-vba-pipeline (격리, 1스텝,
# reset:false=현재 라이브 상태 위)로 보낸다. 백엔드 run_vba_pipeline_on_session(reset=False) 가
#  - 라이브 인스턴스에서 VBA 를 안 돌리고(격리 인스턴스 사용 → RPC 사망/백엔드 크래시 없음)
#  - 현재 라이브 상태(기존 적용 시트)를 보존한 채 새 스텝만 적용
# 하는지 확인. (프로세스가 끝까지 살아 결과 반환 = 크래시 없음의 증거)
import sys, io, tempfile, shutil
from pathlib import Path
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")
sys.path.insert(0, r"C:\Users\Admin\Desktop\KGM_git\B2B_ver0.6.1")
import openpyxl
import pythoncom; pythoncom.CoInitialize()
import win32com.client as win32
import serve_b2b as S

NAME = "book.xlsx"
VBA = ('Sub B2BSkill()\n'
       '  Dim wb As Workbook\n'
       '  Set wb = Workbooks("' + NAME + '")\n'
       '  Dim ws As Worksheet\n'
       '  Set ws = wb.Worksheets.Add(After:=wb.Worksheets(wb.Worksheets.Count))\n'
       '  ws.Name = "새시트"\n'
       '  ws.Range("A1").Value = "added"\n'
       'End Sub')

work = Path(tempfile.mkdtemp(prefix="b2b_singlevba_"))
src = work / NAME       # pristine source (원본 시트만)
_w = openpyxl.Workbook(); _w.active.title = "원본"; _w.active["A1"] = "원본데이터"; _w.save(str(src))
live = work / ("live_" + NAME)
# 라이브 작업복사본 = 원본 + '기존시트'(이전 적용 결과 모사) → reset:false 가 이걸 보존해야 함
_v = openpyxl.load_workbook(str(src)); _v.create_sheet("기존시트")["A1"] = "prev"; _v.save(str(live))

app = None
res = {}
try:
    app = win32.DispatchEx("Excel.Application"); app.Visible = False; app.DisplayAlerts = False
    wb = app.Workbooks.Open(str(live))
    excel_id = "singlevba"
    S.EXCEL_SESSIONS[excel_id] = {"app": app, "workbook": wb, "path": str(live), "openPath": str(wb.FullName),
                                 "name": NAME, "sourcePath": str(src), "liveEditable": True}
    print("적용 전 시트:", [w.Name for w in wb.Worksheets])
    step = {"code": VBA, "language": "vba", "stepIdx": 0, "stepId": "s1", "description": "새시트 추가"}
    # 클라 재라우팅이 보내는 것과 동일: reset=False, 단일 스텝
    out = S.run_vba_pipeline_on_session(excel_id, [step], reset=False)
    print("적용 결과:", {k: out.get(k) for k in ("ok", "applied")})
    lwb = S.EXCEL_SESSIONS[excel_id]["workbook"]
    sheets = [w.Name for w in lwb.Worksheets]
    print("적용 후 시트:", sheets)
    res["원본보존"] = ("원본" in sheets and lwb.Worksheets("원본").Range("A1").Value == "원본데이터")
    res["기존시트보존(reset:false)"] = ("기존시트" in sheets and lwb.Worksheets("기존시트").Range("A1").Value == "prev")
    res["새시트적용"] = ("새시트" in sheets and lwb.Worksheets("새시트").Range("A1").Value == "added")
    res["applied1"] = (out.get("applied") == 1)
    try: del S.EXCEL_SESSIONS[excel_id]
    except Exception: pass
finally:
    try:
        if app is not None: app.Quit()
    except Exception: pass
    shutil.rmtree(work, ignore_errors=True)

checks = [
    ("원본 시트+데이터 보존", res.get("원본보존")),
    ("기존 적용 시트 보존(reset:false=현재 상태 위)", res.get("기존시트보존(reset:false)")),
    ("새 VBA 스텝 적용됨", res.get("새시트적용")),
    ("applied==1, 프로세스 정상 반환(크래시 없음)", res.get("applied1")),
]
fails = 0
print("\n=== 판정 ===")
for n, ok in checks:
    print((" OK  " if ok else "FAIL ") + n)
    if not ok: fails += 1
print(f"\n=== RESULT: {len(checks)-fails}/{len(checks)} PASS ===")
sys.exit(1 if fails else 0)
