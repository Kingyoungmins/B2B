# -*- coding: utf-8 -*-
# [라이브] run_full_pipeline_single_instance — 격리 인스턴스 1개에서 2파일 처리 검증.
#  - 파일 A(원본 시트)·B(데이터 시트)를 각각 라이브 세션으로 띄우고
#  - groups=[{exA:[A에 시트추가]}, {exB:[B에 시트추가]}], resetExcelIds=[exA,exB] 로 1콜 처리
#  - 결과: A 라이브=원본+A결과, B 라이브=데이터+B결과, 원본 데이터 보존, applied==2
# (cross-file Workbooks("파일명") 라우팅 + 파일별 1회 동기화 + 원본 무손상 동시 확인)
import sys, io, tempfile, shutil
from pathlib import Path
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")
sys.path.insert(0, r"C:\Users\Admin\Desktop\KGM_git\B2B_ver0.6.1")
import openpyxl
import pythoncom; pythoncom.CoInitialize()
import win32com.client as win32
import serve_b2b as S

def vba(wbname, sheet_new, val):
    return ('Sub B2BSkill()\n'
            '  Dim wb As Workbook\n'
            '  Set wb = Workbooks("' + wbname + '")\n'
            '  Dim ws As Worksheet\n'
            '  Set ws = wb.Worksheets.Add(After:=wb.Worksheets(wb.Worksheets.Count))\n'
            '  ws.Name = "' + sheet_new + '"\n'
            '  ws.Range("A1").Value = "' + val + '"\n'
            'End Sub')

work = Path(tempfile.mkdtemp(prefix="b2b_fulltest_"))
def build(path, sheet, a1):
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = sheet
    ws["A1"] = a1; ws["A2"] = "row2"; wb.save(str(path))

NA, NB = "fileA.xlsx", "fileB.xlsx"
srcA, srcB = work / NA, work / NB
build(srcA, "원본", "A원본데이터"); build(srcB, "데이터", "B원본데이터")
liveA, liveB = work / ("live_" + NA), work / ("live_" + NB)
shutil.copy2(srcA, liveA); shutil.copy2(srcB, liveB)

app = None
res = {}
try:
    app = win32.DispatchEx("Excel.Application"); app.Visible = False; app.DisplayAlerts = False
    wA = app.Workbooks.Open(str(liveA)); wB = app.Workbooks.Open(str(liveB))
    S.EXCEL_SESSIONS["exA"] = {"app": app, "workbook": wA, "path": str(liveA), "openPath": str(wA.FullName),
                              "name": NA, "sourcePath": str(srcA), "liveEditable": True}
    S.EXCEL_SESSIONS["exB"] = {"app": app, "workbook": wB, "path": str(liveB), "openPath": str(wB.FullName),
                              "name": NB, "sourcePath": str(srcB), "liveEditable": True}
    groups = [
        {"excelId": "exA", "steps": [{"code": vba(NA, "A결과", "fromA"), "language": "vba", "stepIdx": 0, "stepId": "a1"}]},
        {"excelId": "exB", "steps": [{"code": vba(NB, "B결과", "fromB"), "language": "vba", "stepIdx": 1, "stepId": "b1"}]},
    ]
    out = S.run_full_pipeline_single_instance(groups, reset_excel_ids=["exA", "exB"])
    print("결과:", {k: out.get(k) for k in ("ok", "applied")}, "| 스냅샷수:", len(out.get("stepSnapshots") or []))

    lA = S.EXCEL_SESSIONS["exA"]["workbook"]; lB = S.EXCEL_SESSIONS["exB"]["workbook"]
    shA = [w.Name for w in lA.Worksheets]; shB = [w.Name for w in lB.Worksheets]
    print("A 라이브 시트:", shA, "| B 라이브 시트:", shB)
    res["A_원본보존"] = ("원본" in shA and lA.Worksheets("원본").Range("A1").Value == "A원본데이터")
    res["A_결과반영"] = ("A결과" in shA and lA.Worksheets("A결과").Range("A1").Value == "fromA")
    res["B_원본보존"] = ("데이터" in shB and lB.Worksheets("데이터").Range("A1").Value == "B원본데이터")
    res["B_결과반영"] = ("B결과" in shB and lB.Worksheets("B결과").Range("A1").Value == "fromB")
    res["applied2"] = (out.get("applied") == 2)
    for sid in ("exA", "exB"):
        try: del S.EXCEL_SESSIONS[sid]
        except Exception: pass
finally:
    try:
        if app is not None: app.Quit()
    except Exception: pass
    shutil.rmtree(work, ignore_errors=True)

checks = [
    ("A 원본 시트+데이터 보존", res.get("A_원본보존")),
    ("A 결과 시트 반영", res.get("A_결과반영")),
    ("B 원본 시트+데이터 보존", res.get("B_원본보존")),
    ("B 결과 시트 반영", res.get("B_결과반영")),
    ("applied==2", res.get("applied2")),
]
fails = 0
print("\n=== 판정 ===")
for name, ok in checks:
    print((" OK  " if ok else "FAIL ") + name)
    if not ok: fails += 1
print(f"\n=== RESULT: {len(checks)-fails}/{len(checks)} PASS ===")
sys.exit(1 if fails else 0)
