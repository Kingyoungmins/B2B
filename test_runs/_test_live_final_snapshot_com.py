# -*- coding: utf-8 -*-
"""[실측/COM] 새로고침 즉시복원 — 실제 Excel 로 전 구간 재현.

시나리오(사용자가 실제로 하는 것과 동일한 순서):
  1. 파일 2개를 라이브 세션으로 띄운다(작업복사본)
  2. VBA 전체실행(run_full_pipeline_single_instance)에 stateSig 를 주고 돌린다
  3. → 최종상태 사본이 파일마다 남았는가?  ← VBA 는 원래 사본이 아예 없던 경로
  4. 새로고침 = Excel 강제 종료(SaveChanges=False). 원본 파일은 안 바뀌어야 한다
  5. fromStateSig 로 다시 연다 → 적용 결과(추가된 시트/값)가 그대로 살아있는가?
  6. 서명이 다르면(스킬 수정) 사본을 쓰지 않고 원본으로 열리는가?  ← 오복원 방지

실행: python test_runs/_test_live_final_snapshot_com.py   (Excel 필요)
"""
import io
import shutil
import sys
import tempfile
from pathlib import Path

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")
ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT))

import openpyxl
import pythoncom; pythoncom.CoInitialize()
import win32com.client as win32
import serve_b2b as S

SIG = "v1:2:120:deadbeefcafef00d"          # 클라가 보내는 상태 서명 자리(백엔드는 해석하지 않는다)
SIG_OTHER = "v1:2:120:0000000000000000"    # '스킬을 고쳤다' 상황


def vba_add_sheet(wbname, sheet_new, val):
    return ('Sub B2BSkill()\n'
            '  Dim wb As Workbook\n'
            '  Set wb = Workbooks("' + wbname + '")\n'
            '  Dim ws As Worksheet\n'
            '  Set ws = wb.Worksheets.Add(After:=wb.Worksheets(wb.Worksheets.Count))\n'
            '  ws.Name = "' + sheet_new + '"\n'
            '  ws.Range("A1").Value = "' + val + '"\n'
            'End Sub')


work = Path(tempfile.mkdtemp(prefix="b2b_lfs_com_"))
res = {}
app = None


def build(path, sheet, a1):
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = sheet
    ws["A1"] = a1; ws["A2"] = "row2"; wb.save(str(path))


def sheets_of(path):
    wb = openpyxl.load_workbook(str(path), data_only=True)
    out = {ws.title: ws["A1"].value for ws in wb.worksheets}
    wb.close()
    return out


def _session_read_impl(eid, sheet, addr):
    """세션 워크북은 Excel 전용 스레드가 만든 COM 객체다. 다른 스레드에서 직접 만지면
    '다른 스레드용으로 마샬링된 인터페이스' 오류가 난다 — excel_call 로 그 스레드에 태워 읽는다."""
    sess = S.EXCEL_SESSIONS.get(eid)
    if not sess:
        return None, None
    wb = sess["workbook"]
    names = [w.Name for w in wb.Worksheets]
    val = None
    if sheet in names:
        try:
            val = wb.Worksheets(sheet).Range(addr).Value
        except Exception:
            val = None
    return names, val


def session_read(excel_id, sheet="A결과", addr="A1"):
    return S.excel_call(_session_read_impl, excel_id, sheet, addr, timeout=120)


NA, NB = "fileA.xlsx", "fileB.xlsx"
srcA, srcB = work / NA, work / NB
build(srcA, "원본", "A원본데이터")
build(srcB, "데이터", "B원본데이터")
origA_before = srcA.stat().st_mtime_ns, srcA.stat().st_size
origB_before = srcB.stat().st_mtime_ns, srcB.stat().st_size

# 업로드 레코드 등록(실제 앱의 WORKBOOKS 와 같은 모양)
S.WORKBOOKS["wbA"] = {"id": "wbA", "name": NA, "path": str(srcA)}
S.WORKBOOKS["wbB"] = {"id": "wbB", "name": NB, "path": str(srcB)}

try:
    # --- 1) 라이브 세션 2개(작업복사본) ---
    liveA, liveB = work / ("live_" + NA), work / ("live_" + NB)
    shutil.copy2(srcA, liveA); shutil.copy2(srcB, liveB)
    app = win32.DispatchEx("Excel.Application"); app.Visible = False; app.DisplayAlerts = False
    wA = app.Workbooks.Open(str(liveA)); wB = app.Workbooks.Open(str(liveB))
    S.EXCEL_SESSIONS["exA"] = {"app": app, "workbook": wA, "path": str(liveA), "openPath": str(wA.FullName),
                               "name": NA, "sourcePath": str(srcA), "workbookId": "wbA", "liveEditable": True}
    S.EXCEL_SESSIONS["exB"] = {"app": app, "workbook": wB, "path": str(liveB), "openPath": str(wB.FullName),
                               "name": NB, "sourcePath": str(srcB), "workbookId": "wbB", "liveEditable": True}

    # --- 2) VBA 전체실행 + stateSig ---
    groups = [
        {"excelId": "exA", "steps": [{"code": vba_add_sheet(NA, "A결과", "fromA"), "language": "vba", "stepIdx": 0, "stepId": "a1"}]},
        {"excelId": "exB", "steps": [{"code": vba_add_sheet(NB, "B결과", "fromB"), "language": "vba", "stepIdx": 1, "stepId": "b1"}]},
    ]
    out = S.run_full_pipeline_single_instance(groups, reset_excel_ids=["exA", "exB"], state_sig=SIG)
    print("전체실행:", {k: out.get(k) for k in ("ok", "applied")})
    res["applied2"] = (out.get("applied") == 2)

    # 라이브에 반영됐는지(기존 계약 회귀).
    # exA/exB 의 COM 객체는 '이 테스트가 메인 스레드에서' 만든 것이라 여기서 직접 읽어야 한다
    # (excel_call 로 워커 스레드에 태우면 마샬링 오류). 반대로 open_excel_session 이 만든
    # 세션은 워커 소유라 session_read 로 읽는다 — 아래 5~7 단계.
    shA = [w.Name for w in S.EXCEL_SESSIONS["exA"]["workbook"].Worksheets]
    res["라이브반영"] = ("A결과" in shA)

    # --- 3) 최종상태 사본이 남았는가 (VBA 경로!) ---
    snapA = S._find_live_final_snapshot(S.WORKBOOKS["wbA"], SIG)
    snapB = S._find_live_final_snapshot(S.WORKBOOKS["wbB"], SIG)
    print("사본 A:", snapA and snapA["path"])
    print("사본 B:", snapB and snapB["path"])
    res["사본A생성"] = bool(snapA)
    res["사본B생성"] = bool(snapB)
    if snapA:
        sh = sheets_of(Path(snapA["path"]))
        print("사본 A 내용:", sh)
        res["사본A내용"] = (sh.get("A결과") == "fromA" and sh.get("원본") == "A원본데이터")
    res["다른서명은없음"] = (S._find_live_final_snapshot(S.WORKBOOKS["wbA"], SIG_OTHER) is None)

    # --- 4) 새로고침 = Excel 강제 종료 ---
    for sid in ("exA", "exB"):
        S.EXCEL_SESSIONS.pop(sid, None)
    try:
        for w in list(app.Workbooks):
            w.Close(SaveChanges=False)
    except Exception:
        pass
    try:
        app.Quit()
    except Exception:
        pass
    app = None
    res["원본무손상"] = ((srcA.stat().st_mtime_ns, srcA.stat().st_size) == origA_before
                     and (srcB.stat().st_mtime_ns, srcB.stat().st_size) == origB_before)
    print("원본 A 시트(종료 후):", sheets_of(srcA))

    # --- 5) 사본으로 재오픈 (fromStateSig) ---
    opened = S.open_excel_session(str(srcA), name=NA, workbook_id="wbA",
                                  live_editable=True, defer_visible=True, from_state_sig=SIG)
    exid = opened.get("excelId")
    names, val = session_read(exid)
    print("재오픈 시트:", names, "| A결과!A1 =", val)
    res["복원됨"] = ("A결과" in names and val == "fromA")
    res["원본시트도유지"] = ("원본" in names)
    res["레코드path는원본"] = (S.WORKBOOKS["wbA"]["path"] == str(srcA))
    S.close_excel_session(exid)

    # --- 6) 서명이 다르면 원본으로 열려야 한다(오복원 방지) ---
    opened2 = S.open_excel_session(str(srcA), name=NA, workbook_id="wbA",
                                   live_editable=True, defer_visible=True, from_state_sig=SIG_OTHER)
    exid2 = opened2.get("excelId")
    names2, _ = session_read(exid2)
    print("다른 서명으로 재오픈 시트:", names2)
    res["다른서명은원본"] = ("A결과" not in names2 and "원본" in names2)
    S.close_excel_session(exid2)

    # --- 7) 서명 없이 열면 평소처럼 원본 (기존 동작 회귀) ---
    opened3 = S.open_excel_session(str(srcA), name=NA, workbook_id="wbA",
                                   live_editable=True, defer_visible=True)
    exid3 = opened3.get("excelId")
    names3, _ = session_read(exid3)
    print("서명 없이 재오픈 시트:", names3)
    res["서명없으면원본"] = ("A결과" not in names3 and "원본" in names3)
    S.close_excel_session(exid3)

finally:
    for sid in list(S.EXCEL_SESSIONS.keys()):
        try:
            S.close_excel_session(sid)
        except Exception:
            S.EXCEL_SESSIONS.pop(sid, None)
    try:
        if app is not None:
            app.Quit()
    except Exception:
        pass
    try:
        S.cleanup_excel_sessions()
    except Exception:
        pass
    shutil.rmtree(work, ignore_errors=True)

checks = [
    ("전체실행 applied==2", res.get("applied2")),
    ("라이브에 결과 반영(기존 계약)", res.get("라이브반영")),
    ("[핵심] VBA 경로도 최종상태 사본 생성 — A", res.get("사본A생성")),
    ("[핵심] VBA 경로도 최종상태 사본 생성 — B", res.get("사본B생성")),
    ("사본 내용 = 적용 결과 + 원본 데이터", res.get("사본A내용")),
    ("다른 서명으로는 사본이 안 잡힘", res.get("다른서명은없음")),
    ("강제 종료해도 원본 파일 무손상", res.get("원본무손상")),
    ("[핵심] 사본으로 재오픈하면 적용 결과가 살아있음", res.get("복원됨")),
    ("재오픈해도 원본 시트 유지", res.get("원본시트도유지")),
    ("워크북 레코드 path 는 원본 그대로(오염 없음)", res.get("레코드path는원본")),
    ("서명이 다르면 원본으로 열림(오복원 방지)", res.get("다른서명은원본")),
    ("서명 없으면 평소대로 원본(기존 동작)", res.get("서명없으면원본")),
]
fails = sum(1 for _, ok in checks if not ok)
print("\n=== 판정 ===")
for name, ok in checks:
    print((" OK  " if ok else "FAIL ") + name)
print(f"\n=== RESULT: {len(checks)-fails}/{len(checks)} PASS ===")
sys.exit(1 if fails else 0)
