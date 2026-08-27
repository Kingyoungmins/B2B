# -*- coding: utf-8 -*-
"""[적대 검증] ctx.filter_to_range — 얕게 넘어간 곳을 전부 파고든다.

1차 테스트(_test_filter_to_range_com.py)의 구멍
  · 조건을 전부 ctx.column_is 로만 줬다 → **lambda 경로(_as_declarative_filter=None)는 한 번도 안 탔다**
  · 행 '개수'만 봤고 **어떤 행이 붙었는지(값)** 는 안 봤다 — 개수만 맞고 내용이 틀릴 수 있다
  · 값이 여러 개인 필터(Operator=7), 머리글 2줄, include_header, 흩어진 매칭,
    A열이 아닌 시작칸, 원본에 필터가 남는지, 교차파일, COM 예산 — 전부 미검증

여기서 그 전부를 실제 Excel 로 확인한다. 실패의 방향도 함께 본다
(값이 틀리는 조용한 오답 > 오류로 멈추는 것).
"""
import sys, io, tempfile, shutil
from pathlib import Path
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")
ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT))
import openpyxl
import pythoncom; pythoncom.CoInitialize()
import win32com.client as win32
import serve_b2b as S

fails = []


def check(name, cond, detail=""):
    print(("  PASS  " if cond else "  FAIL  ") + name + (("  -> " + str(detail)[:300]) if not cond else ""))
    if not cond:
        fails.append(name)


STATES = ["정지", "정상", "해지", "정상", "정지", "정상", "정지", "해지", "정상", "정지"]


def build(path, header_rows=1, dest_layout="plain"):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "회선 현황"
    if header_rows == 2:
        ws.append(["회선 목록", None, None])          # 제목 줄
    ws.append(["가입번호", "상태", "요금"])
    for i, st in enumerate(STATES):
        ws.append([512102400000 + i, st, 1000 + i])
    f = wb.create_sheet("대상양식")
    f["A1"] = "양식"
    if dest_layout == "offset":
        f["C4"] = "여기부터"
    wb.save(str(path))
    return wb


def open_ctx(app, path):
    book = app.Workbooks.Open(str(path))
    return book, lambda: S.PythonComSkillContext(app, book, {"name": Path(path).name}, timeout_s=300)


def expected_numbers(states=("정지",)):
    return [512102400000 + i for i, s in enumerate(STATES) if s in states]


app = None
work_all = []
try:
    app = win32.DispatchEx("Excel.Application")
    app.Visible = False
    app.DisplayAlerts = False

    # ── [1] lambda 경로 — 1차 테스트가 한 번도 안 탄 길 ─────────────────────────
    print("[1] lambda 조건 — 자동필터가 아닌 값 판정 경로")
    w = Path(tempfile.mkdtemp(prefix="b2b_dl_")); work_all.append(w)
    sp = w / "a.xlsx"; build(sp)
    book, mk = open_ctx(app, sp)
    ctx = mk()
    n = ctx.filter_to_range("회선 현황", lambda r: str(r[1]) == "정지", "대상양식", "A3")
    c2 = mk()
    got = [r[0] for r in c2.read("대상양식", "A3:A%d" % (2 + n))]
    check("lambda 로도 4건", n == 4, n)
    check("lambda: 붙은 값이 정확히 정지 행", [int(x) for x in got] == expected_numbers(), got)
    st = [r[0] for r in c2.read("대상양식", "B3:B%d" % (2 + n))]
    check("lambda: 상태 열이 전부 '정지'", set(st) == {"정지"}, st)
    book.Close(SaveChanges=False)

    # ── [2] 값이 여러 개인 필터(Operator=7) ────────────────────────────────────
    print("[2] 값 여러 개 필터 — Excel Operator=7 분기")
    w = Path(tempfile.mkdtemp(prefix="b2b_multi_")); work_all.append(w)
    sp = w / "b.xlsx"; build(sp)
    book, mk = open_ctx(app, sp)
    ctx = mk()
    col = ctx.find_header("회선 현황", "상태")
    n = ctx.filter_to_range("회선 현황", ctx.column_is(col, ["정지", "해지"]), "대상양식", "A3")
    c2 = mk()
    got = [int(r[0]) for r in c2.read("대상양식", "A3:A%d" % (2 + n))]
    check("정지+해지 = 6건", n == 6, n)
    check("값이 정확히 일치", got == expected_numbers(("정지", "해지")), got)
    book.Close(SaveChanges=False)

    # ── [3] 머리글 2줄 ────────────────────────────────────────────────────────
    print("[3] 머리글이 2줄인 표(header_rows=2)")
    w = Path(tempfile.mkdtemp(prefix="b2b_hr2_")); work_all.append(w)
    sp = w / "c.xlsx"; build(sp, header_rows=2)
    book, mk = open_ctx(app, sp)
    ctx = mk()
    col = ctx.find_header("회선 현황", "상태", header_row=2)
    n = ctx.filter_to_range("회선 현황", ctx.column_is(col, ["정지"]), "대상양식", "A3", header_rows=2)
    c2 = mk()
    got = [int(r[0]) for r in c2.read("대상양식", "A3:A%d" % (2 + n))]
    check("머리글 2줄에서도 4건", n == 4, n)
    check("제목 줄이 데이터로 안 섞인다", got == expected_numbers(), got)
    book.Close(SaveChanges=False)

    # ── [4] include_header=True ───────────────────────────────────────────────
    print("[4] 머리글도 함께 붙이기")
    w = Path(tempfile.mkdtemp(prefix="b2b_ih_")); work_all.append(w)
    sp = w / "d.xlsx"; build(sp)
    book, mk = open_ctx(app, sp)
    ctx = mk()
    col = ctx.find_header("회선 현황", "상태")
    n = ctx.filter_to_range("회선 현황", ctx.column_is(col, ["정지"]), "대상양식", "A3",
                            include_header=True)
    c2 = mk()
    check("머리글이 시작칸에 붙는다", c2.read_cell("대상양식", "A3") == "가입번호",
          c2.read_cell("대상양식", "A3"))
    check("데이터는 그 아래부터", int(c2.read_cell("대상양식", "A4")) == expected_numbers()[0],
          c2.read_cell("대상양식", "A4"))
    check("반환값은 데이터 행 수(머리글 제외)", n == 4, n)
    book.Close(SaveChanges=False)

    # ── [5] A열이 아닌 시작칸 ─────────────────────────────────────────────────
    print("[5] 시작칸이 A열이 아닐 때(C4)")
    w = Path(tempfile.mkdtemp(prefix="b2b_off_")); work_all.append(w)
    sp = w / "e.xlsx"; build(sp, dest_layout="offset")
    book, mk = open_ctx(app, sp)
    ctx = mk()
    col = ctx.find_header("회선 현황", "상태")
    n = ctx.filter_to_range("회선 현황", ctx.column_is(col, ["정지"]), "대상양식", "C5")
    c2 = mk()
    got = [int(r[0]) for r in c2.read("대상양식", "C5:C%d" % (4 + n))]
    check("C5 부터 붙는다", got == expected_numbers(), got)
    check("왼쪽 열(A,B)은 안 건드린다",
          c2.read_cell("대상양식", "A5") is None and c2.read_cell("대상양식", "B5") is None,
          (c2.read_cell("대상양식", "A5"), c2.read_cell("대상양식", "B5")))
    check("바로 위 칸(C4)도 안 건드린다", c2.read_cell("대상양식", "C4") == "여기부터",
          c2.read_cell("대상양식", "C4"))
    book.Close(SaveChanges=False)

    # ── [6] 원본에 필터가 남지 않는다 ─────────────────────────────────────────
    print("[6] 실행 후 원본에 필터/숨은 행이 남지 않는다")
    w = Path(tempfile.mkdtemp(prefix="b2b_af_")); work_all.append(w)
    sp = w / "f.xlsx"; build(sp)
    book, mk = open_ctx(app, sp)
    ctx = mk()
    col = ctx.find_header("회선 현황", "상태")
    ctx.filter_to_range("회선 현황", ctx.column_is(col, ["정지"]), "대상양식", "A3")
    src_ws = book.Worksheets("회선 현황")
    check("자동필터가 꺼져 있다", not bool(src_ws.AutoFilterMode), src_ws.AutoFilterMode)
    hidden = [r for r in range(2, 12) if bool(src_ws.Rows(r).Hidden)]
    check("숨은 행이 없다(원본 화면 무손상)", not hidden, hidden)
    c2 = mk()
    check("원본 데이터가 그대로", c2.used_last_row("회선 현황") == 11, c2.used_last_row("회선 현황"))
    book.Close(SaveChanges=False)

    # ── [7] 매칭 행이 흩어져 있어도 순서·내용이 맞는다 ────────────────────────
    print("[7] 매칭 행이 붙어 있지 않고 흩어져 있을 때(구간 분할)")
    #     STATES 는 이미 정지가 0,4,6,9 로 흩어져 있다 — 두 경로 모두에서 확인
    for label, cond_maker in (("자동필터", lambda c, col: c.column_is(col, ["정지"])),
                              ("lambda", lambda c, col: (lambda r: str(r[1]) == "정지"))):
        w = Path(tempfile.mkdtemp(prefix="b2b_sc_")); work_all.append(w)
        sp = w / "g.xlsx"; build(sp)
        book, mk = open_ctx(app, sp)
        ctx = mk()
        col = ctx.find_header("회선 현황", "상태")
        n = ctx.filter_to_range("회선 현황", cond_maker(ctx, col), "대상양식", "A3")
        c2 = mk()
        got = [int(r[0]) for r in c2.read("대상양식", "A3:A%d" % (2 + n))]
        check("%s: 흩어진 4건이 원래 순서대로" % label, got == expected_numbers(), got)
        book.Close(SaveChanges=False)

    # ── [8] 교차파일 — ctx.book(다른 파일)에 붙이기 ───────────────────────────
    print("[8] 다른 파일의 양식에 붙이기(ctx.book)")
    w = Path(tempfile.mkdtemp(prefix="b2b_x_")); work_all.append(w)
    sp = w / "src.xlsx"; build(sp)
    dp = w / "form.xlsx"
    wb2 = openpyxl.Workbook(); wb2.active.title = "양식"; wb2["양식"]["A1"] = "머리"; wb2.save(str(dp))
    book, mk = open_ctx(app, sp)
    book2 = app.Workbooks.Open(str(dp))
    ctx = mk()
    col = ctx.find_header("회선 현황", "상태")
    other = ctx.book("form.xlsx")
    try:
        n = other.filter_to_range("회선 현황", ctx.column_is(col, ["정지"]), "양식", "A3")
        c2 = mk()
        got = [int(r[0]) for r in c2.book("form.xlsx").read("양식", "A3:A%d" % (2 + n))]
        check("교차파일에도 붙는다", got == expected_numbers(), got)
    except Exception as err:
        # 원본은 다른 워크북이라 못 찾을 수 있다 — 그 경우 '조용히 틀리기'보다 오류가 낫다
        check("교차파일: 조용히 틀리지 않고 사유를 말한다",
              "시트" in str(err) or "워크북" in str(err) or "찾" in str(err), err)
    book2.Close(SaveChanges=False)
    book.Close(SaveChanges=False)

    # ── [9] COM 예산 — 큰 표에서도 한 번의 호출로 끝난다 ──────────────────────
    print("[9] 큰 표에서도 COM 예산을 태우지 않는다")
    w = Path(tempfile.mkdtemp(prefix="b2b_big_")); work_all.append(w)
    sp = w / "big.xlsx"
    wb3 = openpyxl.Workbook(); ws3 = wb3.active; ws3.title = "회선 현황"
    ws3.append(["가입번호", "상태", "요금"])
    for i in range(3000):
        ws3.append([512102400000 + i, "정지" if i % 3 == 0 else "정상", i])
    wb3.create_sheet("대상양식")
    wb3.save(str(sp))
    book, mk = open_ctx(app, sp)
    ctx = mk()
    col = ctx.find_header("회선 현황", "상태")
    n = ctx.filter_to_range("회선 현황", ctx.column_is(col, ["정지"]), "대상양식", "A2")
    used = ctx.summary().get("comCalls")
    check("3000행에서 1000건을 붙인다", n == 1000, n)
    check("COM 호출이 예산(400) 안", used < 400, used)
    print("      (COM 호출 %s회)" % used)
    book.Close(SaveChanges=False)

    # ── [10] 잘못 쓴 경우 — 조용히 틀리지 않고 사유를 말한다 ──────────────────
    print("[10] 잘못 쓴 경우의 반응")
    w = Path(tempfile.mkdtemp(prefix="b2b_err_")); work_all.append(w)
    sp = w / "h.xlsx"; build(sp)
    book, mk = open_ctx(app, sp)
    ctx = mk()
    try:
        ctx.filter_to_range("회선 현황", ctx.column_is(2, ["정지"]), "없는양식", "A3")
        check("없는 대상 시트 → 오류", False, "오류가 안 났다")
    except Exception as err:
        check("없는 대상 시트 → 사유를 말하는 오류", "시트" in str(err) or "찾" in str(err), err)
    ctx = mk()
    try:
        ctx.filter_to_range("회선 현황", lambda r: 1 / 0, "대상양식", "A3")
        check("조건이 터지면 오류", False, "오류가 안 났다")
    except Exception as err:
        check("조건이 터지면 무슨 조건인지 알려준다", "predicate" in str(err) or "조건" in str(err), err)
    book.Close(SaveChanges=False)
finally:
    try:
        if app is not None:
            for wbk in list(app.Workbooks):
                try: wbk.Close(SaveChanges=False)
                except Exception: pass
            app.Quit()
    except Exception:
        pass
    for w in work_all:
        shutil.rmtree(w, ignore_errors=True)

print("")
print("RESULT: ALL PASS" if not fails else "RESULT: %d FAIL -> %s" % (len(fails), fails))
sys.exit(0 if not fails else 1)
