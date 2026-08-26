# -*- coding: utf-8 -*-
"""[설명서 검증] docs/user-guide/AX-Cell_스킬_함수_설명서 에 적은 예시가 실제로 도는지.

왜: 이 문서는 사업팀이 '무엇을 시킬 수 있는가'의 근거로 본다. 문서가 코드와 어긋나면
    없느니만 못하다. 그래서 설명서에 적은 명령들을 진짜 Excel 에 대고 그대로 실행한다.
    (인자 이름·기본값의 일치는 _test_ctx_manual_sync.py 가 따로 잠근다)
"""
import sys, io, tempfile, shutil
from pathlib import Path
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")
sys.path.insert(0, str(Path(__file__).resolve().parent.parent))
import openpyxl
import pythoncom; pythoncom.CoInitialize()
import win32com.client as win32
import serve_b2b as S

fails = []


def check(name, cond, detail=""):
    print(("  PASS  " if cond else "  FAIL  ") + name + (("  -> " + str(detail)[:300]) if not cond else ""))
    if not cond:
        fails.append(name)


work = Path(tempfile.mkdtemp(prefix="b2b_ctxdoc_"))
src = work / "doc.xlsx"
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "데이터"
rows = [["지점", "상품", "금액", "수량", "합계"],
        ["서울", "인터넷", 100, 2, None],
        ["부산", "인터넷", 200, 1, None],
        ["서울", "TV", 300, 3, None],
        ["부산", "TV", 400, 1, None],
        ["서울", "인터넷", 100, 2, None]]
for r in rows:
    ws.append(r)
ws2 = wb.create_sheet("단가표")
for r in [["코드", "단가"], ["인터넷", 11], ["TV", 22]]:
    ws2.append(r)
ws3 = wb.create_sheet("표지")
ws3["A1"] = "2026년 06월 (05월 1일 ~ 05월 31일)"
ws4 = wb.create_sheet("명단")
for r in [["이름"], ["1001/홍길동"], ["1002/김철수"]]:
    ws4.append(r)
wb.save(str(src))

app = None
try:
    app = win32.DispatchEx("Excel.Application")
    app.Visible = False
    app.DisplayAlerts = False
    book = app.Workbooks.Open(str(src))

    # COM 예산(400회)은 '스킬 한 단계' 단위다. 실제 스킬도 단계마다 새 ctx 를 받으므로
    # 여기서도 구획마다 새로 만든다(한 ctx 로 66개를 다 부르면 예산 초과로 죽는다).
    def new_ctx():
        return S.PythonComSkillContext(app, book, {"name": "doc.xlsx"}, timeout_s=600)

    ctx = new_ctx()
    print("[가] 표 살펴보기")
    check("sheets()", "데이터" in ctx.sheets(), ctx.sheets())
    r, c = ctx.used_range("데이터")
    check("used_range() -> (행,열)", r >= 6 and c >= 5, (r, c))
    check("last_row(col=1) = 6", ctx.last_row("데이터", col=1) == 6, ctx.last_row("데이터", col=1))
    check("last_col(row=1) = 5", ctx.last_col("데이터", row=1) == 5, ctx.last_col("데이터", row=1))
    check("used_last_row() = 6", ctx.used_last_row("데이터") == 6, ctx.used_last_row("데이터"))
    check("used_last_col() = 5", ctx.used_last_col("데이터") == 5, ctx.used_last_col("데이터"))
    check("find_header('금액') = 3", ctx.find_header("데이터", "금액") == 3, ctx.find_header("데이터", "금액"))
    check("find_header_row('지점') = 1", ctx.find_header_row("데이터", "지점") == 1)
    empty = ctx.first_empty_col("데이터", after="E")
    check("first_empty_col(after='E') -> 'F'", empty == "F", empty)
    data = ctx.read("데이터", "A1:C3")
    check("read() 2차원", len(data) == 3 and data[1][0] == "서울", data)
    check("read_cell()", ctx.read_cell("데이터", "A2") == "서울", ctx.read_cell("데이터", "A2"))

    ctx = new_ctx()
    print("[나] 값 쓰기·지우기")
    ctx.write("데이터", "G1", [["가", 1], ["나", 2]])
    check("write() 표 모양", ctx.read_cell("데이터", "H2") == 2, ctx.read_cell("데이터", "H2"))
    ctx.write_cell("데이터", "G3", 12345)
    check("write_cell()", ctx.read_cell("데이터", "G3") == 12345)
    ctx.write_formulas("데이터", "E2", [["=C2*D2"], ["=C3*D3"]])
    check("write_formulas()", ctx.read_cell("데이터", "E2") == 200, ctx.read_cell("데이터", "E2"))
    check("has_formulas()", ctx.has_formulas("데이터", "E2:E3") is True)
    mask = ctx.formula_mask("데이터", "E2:E3")
    check("formula_mask() 칸별", mask[0][0] is True, mask)
    check("read_formulas()", "=" in str(ctx.read_formulas("데이터", "E2")[0][0]))
    ctx.clear("데이터", "G1:H3")
    check("clear()", ctx.read_cell("데이터", "G3") is None)
    n = ctx.replace("데이터", "A2:A6", "서울", "서울시")
    check("replace() 바뀐 수 반환", n == 3, n)
    ctx.replace("데이터", "A2:A6", "서울시", "서울")
    ctx.shift_months("표지", "A1", 1)
    check("shift_months(+1) 달·말일 보정",
          ctx.read_cell("표지", "A1") == "2026년 07월 (06월 1일 ~ 06월 30일)",
          ctx.read_cell("표지", "A1"))

    ctx = new_ctx()
    print("[다] 복사")
    ctx.copy("데이터", "A1:C2", "표지", "A5")
    check("copy()", ctx.read_cell("표지", "A6") == "서울", ctx.read_cell("표지", "A6"))
    ctx.copy_values("데이터", "E2:E3", "표지", "E5")
    check("copy_values() 값만", ctx.read_cell("표지", "E5") == 200, ctx.read_cell("표지", "E5"))
    ctx.copy_col("데이터", "C", "J")
    check("copy_col()", ctx.read_cell("데이터", "J2") == 100, ctx.read_cell("데이터", "J2"))
    ctx.clear("데이터", "J1:J10")

    ctx = new_ctx()
    print("[마] 필터")
    ctx.enable_filter("데이터", header_row=1)
    ctx.apply_filter("데이터", "상품", ["TV"])
    ctx.clear_filter("데이터")
    check("enable/apply/clear_filter() 연속 동작", True)
    ctx.filter_to_sheet("데이터", lambda r: r[1] == "TV", "TV목록")
    check("filter_to_sheet() 새 시트", "TV목록" in ctx.sheets(), ctx.sheets())

    ctx = new_ctx()
    print("[바] 합계·피벗")
    total = ctx.sum_column("데이터", "금액", header_row=1)
    check("sum_column('금액') = 1100", total == 1100, total)
    t2 = ctx.sum_where("데이터", "금액", [("상품", "인터넷")], header_row=1)
    check("sum_where(상품=인터넷) = 400", t2 == 400, t2)
    ctx.pivot("데이터", group_by="지점", value="금액", agg="sum", dest_name="지점별요약")
    check("pivot() 새 시트", "지점별요약" in ctx.sheets(), ctx.sheets())
    row = ctx.add_total_row("데이터", ["금액"], label_col="A")
    check("add_total_row() 행번호 반환", isinstance(row, int) and row >= 7, row)
    ctx.delete_rows("데이터", row)

    ctx = new_ctx()
    print("[사] 값 가져오기")
    ctx.write_cell("데이터", "K1", "단가")
    got = ctx.lookup("데이터", "상품", "K", "단가표", "코드", "단가", header_row=1)
    check("lookup() 매칭 수", got == 5, got)
    check("lookup() 값", ctx.read_cell("데이터", "K2") == 11, ctx.read_cell("데이터", "K2"))
    check("normalize() 공백 무시", ctx.normalize(" 가 나 ") == ctx.normalize("가나"),
          (ctx.normalize(" 가 나 "), ctx.normalize("가나")))

    ctx = new_ctx()
    print("[아] 시트")
    ctx.add_sheet("새시트", after="데이터")
    check("add_sheet(after=)", "새시트" in ctx.sheets())
    ctx.rename_sheet("새시트", "이름바꿈")
    check("rename_sheet()", "이름바꿈" in ctx.sheets() and "새시트" not in ctx.sheets())
    ctx.move_sheet("이름바꿈", before="데이터")
    check("move_sheet()", ctx.sheets()[0] == "이름바꿈", ctx.sheets()[:2])
    ctx.copy_sheet("이름바꿈", new_name="복사본")
    check("copy_sheet(new_name=)", "복사본" in ctx.sheets(), ctx.sheets())
    ctx.delete_sheet("복사본")
    ctx.delete_sheet("이름바꿈")
    check("delete_sheet()", "복사본" not in ctx.sheets())

    ctx = new_ctx()
    print("[자] 서식")
    ctx.set_fill("데이터", "A1:C1", "노랑")
    ctx.set_font("데이터", "A1:C1", bold=True, size=12)
    ctx.set_border("데이터", "A1:D1", style="thick", edges="bottom")
    ctx.set_number_format("데이터", "C2:C6", "#,##0")
    check("set_fill/set_font/set_border/set_number_format 연속 동작", True)
    ctx.set_fill("데이터", "A1:C1")     # 색 지우기(None)
    check("set_fill(color 생략) = 채우기 없음", True)

    ctx = new_ctx()
    print("[라][차] 행·열·정리")
    ctx.insert_rows("데이터", 2)
    check("insert_rows()", ctx.read_cell("데이터", "A2") is None)
    ctx.delete_rows("데이터", 2)
    ctx.insert_cols("데이터", "B")
    ctx.delete_cols("데이터", "B")
    check("insert_cols/delete_cols()", ctx.read_cell("데이터", "B1") == "상품", ctx.read_cell("데이터", "B1"))
    ctx.hide_cols("데이터", "J:J")
    ctx.hide_cols("데이터", "J:J", False)
    ctx.merge("표지", "A10:C10")
    ctx.unmerge("표지", "A10:C10")
    check("hide_cols/merge/unmerge 연속 동작", True)
    ctx.split_column("명단", "A", "/", into=["가입번호", "고객명"])
    check("split_column()", ctx.read_cell("명단", "B2") == "1001" or ctx.read_cell("명단", "B2") == 1001,
          ctx.read_cell("명단", "B2"))
    ctx.sort("데이터", "A1:E6", "C", ascending=False)
    check("sort(내림차순)", ctx.read_cell("데이터", "C2") == 400, ctx.read_cell("데이터", "C2"))
    removed = ctx.dedupe("데이터", ["지점", "상품"], header_row=1)
    check("dedupe() 삭제 수 반환", removed == 1, removed)
    left = ctx.delete_rows_where("데이터", lambda r: ctx.normalize(r[0]) == ctx.normalize("부산"))
    check("delete_rows_where() 삭제 수 반환", left == 2, left)

    ctx = new_ctx()
    print("[보강] 설명서가 특별히 약속한 동작들")
    # clear(keep_formulas=True): 수식은 남기고 값만 지운다 — 설명서 '나' 절의 약속
    ctx.write("표지", "H1", [[1], [2]])
    ctx.write_formulas("표지", "H3", [["=H1+H2"]])
    ctx.clear("표지", "H1:H3", keep_formulas=True)
    check("clear(keep_formulas=True) 값만 지우고 수식 유지",
          ctx.read_cell("표지", "H1") is None and ctx.read_formulas("표지", "H3")[0][0].startswith("="),
          (ctx.read_cell("표지", "H1"), ctx.read_formulas("표지", "H3")))
    ctx.clear("표지", "H1:H3")
    # hide_rows / column_is — 다른 테스트에 없어 여기서만 잡힌다
    ctx.hide_rows("표지", "20:21")
    ctx.hide_rows("표지", "20:21", False)
    check("hide_rows() 숨김·해제", True)
    cond = ctx.column_is("B", ["TV"])
    check("column_is() 조건 객체를 만든다", cond is not None, cond)

    ctx = new_ctx()
    ctx.write("정산", "A1", [["항목", "1월", "2월", "합계"]]) if "정산" in ctx.sheets() else None
    if "정산" not in ctx.sheets():
        ctx.add_sheet("정산")
        ctx = new_ctx()
        ctx.write("정산", "A1", [["항목", "1월", "2월", "합계"],
                                ["가", 10, 20, None], ["나", 30, 40, None]])
    ctx = new_ctx()
    ctx.fill_sum_col("정산", "합계", ["1월", "2월"], header_row=1)
    check("fill_sum_col() 합계 수식 채움", ctx.read_cell("정산", "D2") == 30, ctx.read_cell("정산", "D2"))
    ctx = new_ctx()
    ctx.swap_cols("정산", "B", "C")
    check("swap_cols() 인접 두 열 교환 후에도 합계 유지",
          ctx.read_cell("정산", "B1") == "2월" and ctx.read_cell("정산", "D2") == 30,
          (ctx.read_cell("정산", "B1"), ctx.read_cell("정산", "D2")))
    ctx = new_ctx()
    ctx.move_cols("정산", ["합계"], "B", header_row=1)
    check("move_cols() 열을 앞으로 이동", ctx.read_cell("정산", "B1") == "합계", ctx.read_cell("정산", "B1"))
    ctx = new_ctx()
    ctx.move_col_clear("정산", "B", "G", header_row=1)
    check("move_col_clear() 옮기고 원본 비움",
          ctx.read_cell("정산", "G1") == "합계" and ctx.read_cell("정산", "B1") is None,
          (ctx.read_cell("정산", "G1"), ctx.read_cell("정산", "B1")))

    ctx = new_ctx()
    print("[검증] 설명서에 적힌 명령이 하나도 빠짐없이 실제로 있는가")
    import re
    doc = (Path(__file__).resolve().parent.parent / "docs" / "user-guide"
           / "AX-Cell_스킬_함수_설명서_v0.8.0.txt").read_text(encoding="utf-8")
    named = set(re.findall(r"ctx\.([a-z_]+)", doc))
    missing = [n for n in named if not hasattr(ctx, n)]
    check("설명서의 모든 명령이 ctx 에 존재", not missing, missing)
finally:
    try:
        if app is not None:
            for w in list(app.Workbooks):
                try:
                    w.Close(SaveChanges=False)
                except Exception:
                    pass
            app.Quit()
    except Exception:
        pass
    shutil.rmtree(work, ignore_errors=True)

print("")
print("RESULT: ALL PASS" if not fails else "RESULT: %d FAIL -> %s" % (len(fails), fails))
sys.exit(0 if not fails else 1)
