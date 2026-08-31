# -*- coding: utf-8 -*-
"""[제보 2026-08-31 #2] "스킬 추가될 때 창이 내려갔다 올라온다 — VBA일 때".

로그로 짐작하지 않고 **실제 Excel 창이 실제로 내려가는지**를 잰다.
  · 매크로가 도는 순간의 app.Visible 과 IsWindowVisible(app.Hwnd) 를 샘플링
  · 단일 적용 전/후 값과 비교

기대(수정 전): VBA 는 실행 직전 _prepare_vba_macro_run_window_state 가 창을 진짜로 숨기고
              (park + Visible=False + SW_HIDE), 끝나고 _restore_live_window 가 다시 띄운다.
              → 사용자 눈에 '내려갔다 올라옴'. Python 단일 적용에는 이 왕복이 없다.
기대(수정 후): 창을 안 내리고 매크로가 돈다. 거부하는 Office 빌드에서만 그때 숨기고 재시도.
"""
import sys, io, tempfile, shutil
from pathlib import Path

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")
ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT))
import openpyxl
import pythoncom; pythoncom.CoInitialize()
import win32com.client as win32
import win32gui
import serve_b2b as S

fails = []


def check(name, cond, detail=""):
    print(("  PASS  " if cond else "  FAIL  ") + name + (("  -> " + str(detail)[:300]) if not cond else ""))
    if not cond:
        fails.append(name)


VBA = 'Sub B2BSkill()\n    ThisWorkbook.Worksheets(1).Range("A1").Value = "vba-ok"\nEnd Sub\n'
PY = 'def transform(ctx):\n    ctx.write_cell("Sheet", "A2", "py-ok")\n'

work = Path(tempfile.mkdtemp(prefix="b2b_flick_"))
NAME = "깜빡임.xlsx"
wbk = openpyxl.Workbook(); ws = wbk.active; ws.title = "Sheet"
ws["A1"] = "before"; wbk.create_sheet("메모"); wbk.save(str(work / NAME))

app = None
samples = {}


def sample(tag, app):
    """그 시점에 창이 화면에 있는가 — Excel 자체 판단(Visible)과 OS 판단(IsWindowVisible) 둘 다."""
    try:
        vis = bool(app.Visible)
    except Exception:
        vis = None
    try:
        hwnd_vis = bool(win32gui.IsWindowVisible(int(app.Hwnd)))
    except Exception:
        hwnd_vis = None
    samples[tag] = {"appVisible": vis, "hwndVisible": hwnd_vis}
    print("      [%s] app.Visible=%s  IsWindowVisible=%s" % (tag, vis, hwnd_vis))


try:
    app = win32.DispatchEx("Excel.Application")
    app.DisplayAlerts = False
    wb = app.Workbooks.Open(str(work / NAME))
    app.Visible = True                      # 사용자가 보고 있는 라이브 창 상태
    S.EXCEL_SESSIONS["flick"] = {
        "app": app, "workbook": wb, "path": str(work / NAME), "openPath": str(wb.FullName),
        "name": NAME, "sourcePath": str(work / NAME), "liveEditable": True,
        "pid": S._excel_process_id(app),
    }

    print("[1] 매크로가 도는 '그 순간' 창이 화면에 있는가 (VBA 단일 적용)")
    sample("before", app)
    orig_inject = S._inject_and_run_vba

    def spy(a, w, code, entry):
        sample("during-vba", a)             # 실행 직전 = 창 숨김이 이미 걸린 시점
        return orig_inject(a, w, code, entry)

    S._inject_and_run_vba = spy
    try:
        out = S._run_vba_on_session_impl("flick", VBA)
    finally:
        S._inject_and_run_vba = orig_inject
    sample("after", app)

    check("VBA 적용 자체는 성공한다", bool(out) and out.get("ok"), out)
    check("값이 실제로 써졌다",
          str(S.EXCEL_SESSIONS["flick"]["workbook"].Worksheets("Sheet").Range("A1").Value) == "vba-ok")

    b, d, a2 = samples["before"], samples["during-vba"], samples["after"]
    check("적용 전에는 창이 보인다", b["appVisible"] is True and b["hwndVisible"] is True, b)
    check("적용 후에는 창이 돌아온다", a2["appVisible"] is True, a2)
    dropped = (d["appVisible"] is False) or (d["hwndVisible"] is False)
    print("      → 매크로 도는 동안 창이 내려갔나?  %s" % ("예 (증상 재현)" if dropped else "아니오"))
    check("매크로가 도는 동안에도 창이 안 내려간다", not dropped,
          "실행 중 %s — 창이 내려갔다 올라온다" % d)

    print("[2] 대조군 — Python 단일 적용에는 이 왕복이 없다")
    samples.clear()
    sample("py-before", app)
    orig_py = S._run_python_on_session_impl

    def spy_py(excel_id, code, *a, **k):
        sample("during-python", S.EXCEL_SESSIONS[excel_id]["app"])
        return orig_py(excel_id, code, *a, **k)

    S._run_python_on_session_impl = spy_py
    try:
        S._run_python_on_session_impl("flick", PY)
    except Exception as e:
        print("      (python 적용 예외: %s)" % str(e)[:120])
    finally:
        S._run_python_on_session_impl = orig_py
    if "during-python" in samples:
        dp = samples["during-python"]
        check("Python 적용 중에는 창이 그대로 있다",
              dp["appVisible"] is not False and dp["hwndVisible"] is not False, dp)

    print("[5] 사용자가 못 박은 회귀 두 가지 — 회색화면 / 탭이 다른 곳을 봄")
    # 창 왕복을 없앴으니 '되돌리기(_restore_live_window)'가 하던 뒷정리도 같이 빠진다.
    # 그중 라이브에 꼭 필요한 두 가지를 실제로 확인한다.
    #   (가) 매크로가 다른 워크북을 Activate 해도 화면엔 대상 파일만 남는가 (탭 어긋남)
    #   (나) 대상 창이 계속 그려진 채인가 — 빈 회색 프레임이 아닌가 (회색화면)
    other_path = work / "옆파일.xlsx"
    ob = openpyxl.Workbook(); ob.active.title = "옆"; ob.active["A1"] = "x"; ob.save(str(other_path))
    owb = app.Workbooks.Open(str(other_path))
    CROSS = ('Sub B2BSkill()\n'
             '    Workbooks("옆파일.xlsx").Activate\n'          # 녹화 매크로가 실제로 내는 모양
             '    ThisWorkbook.Worksheets(1).Range("A3").Value = "cross-ok"\n'
             'End Sub\n')
    fg_before = win32gui.GetForegroundWindow()
    S._run_vba_on_session_impl("flick", CROSS)
    fg_after = win32gui.GetForegroundWindow()
    live = S.EXCEL_SESSIONS["flick"]["workbook"]
    check("교차 Activate 매크로도 성공", str(live.Worksheets("Sheet").Range("A3").Value) == "cross-ok",
          live.Worksheets("Sheet").Range("A3").Value)
    try:
        active_name = str(app.ActiveWorkbook.Name)
    except Exception:
        active_name = "?"
    check("적용이 끝나면 활성 워크북은 대상 파일이다(탭 어긋남 방지)", active_name == NAME, active_name)
    other_vis = any(bool(owb.Windows(i).Visible) for i in range(1, int(owb.Windows.Count) + 1))
    check("옆 파일 창은 화면에 안 남는다", other_vis is False, other_vis)
    tgt_vis = any(bool(live.Windows(i).Visible) for i in range(1, int(live.Windows.Count) + 1))
    check("대상 파일 창은 계속 보인다(회색 빈 프레임 아님)", tgt_vis is True, tgt_vis)
    check("Excel 앱 창도 계속 보인다", bool(app.Visible) is True, app.Visible)
    print("      포그라운드 %s → %s %s" % (fg_before, fg_after,
                                        "(안 뺏음)" if fg_before == fg_after else "(바뀜)"))
    check("적용이 사용자 포커스를 뺏지 않는다", fg_before == fg_after, (fg_before, fg_after))
    try:
        owb.Close(SaveChanges=False)
    except Exception:
        pass

    print("[4] 거부하는 Office 빌드에서는 숨김 폴백이 그대로 산다 (안전망 확인)")
    S._VBA_WINDOW_HIDE["required"] = False
    samples.clear()
    seen = {"hidden_at_try": []}
    orig_inject = S._inject_and_run_vba
    tries = {"n": 0}

    def spy_blocked(a, w, code, entry):
        tries["n"] += 1
        seen["hidden_at_try"].append(bool(a.Visible) is False)
        if tries["n"] == 1:
            raise RuntimeError("이 통합 문서의 매크로를 실행할 수 없습니다. 모든 매크로를 사용하지 못할 수 있습니다.")
        return orig_inject(a, w, code, entry)

    S._inject_and_run_vba = spy_blocked
    try:
        out2 = S._run_vba_on_session_impl("flick", VBA.replace("vba-ok", "vba-retry"))
    finally:
        S._inject_and_run_vba = orig_inject
    check("거부당하면 재시도한다(1회 → 2회)", tries["n"] == 2, tries)
    check("1회차는 창을 안 내린 상태로 시도한다", seen["hidden_at_try"][:1] == [False], seen)
    check("2회차는 창을 내리고 시도한다", seen["hidden_at_try"][1:2] == [True], seen)
    check("재시도로 결국 성공한다", bool(out2) and out2.get("ok"), out2)
    check("값도 실제로 써졌다",
          str(S.EXCEL_SESSIONS["flick"]["workbook"].Worksheets("Sheet").Range("A1").Value) == "vba-retry")
    check("숨겨야 하는 환경으로 기억한다", S._VBA_WINDOW_HIDE["required"] is True)
    sample("after-retry", app)
    check("숨겼던 경우엔 창을 되돌려 놓는다", samples["after-retry"]["appVisible"] is True,
          samples.get("after-retry"))

    print("[4-b] 한 번 기억한 뒤에는 처음부터 숨긴다(재시도 비용 0 — 종전 동작)")
    seen2 = []
    orig_inject = S._inject_and_run_vba

    def spy_remembered(a, w, code, entry):
        seen2.append(bool(a.Visible) is False)
        return orig_inject(a, w, code, entry)

    S._inject_and_run_vba = spy_remembered
    try:
        S._run_vba_on_session_impl("flick", VBA.replace("vba-ok", "vba-again"))
    finally:
        S._inject_and_run_vba = orig_inject
        S._VBA_WINDOW_HIDE["required"] = False
    check("한 번만 시도하고", len(seen2) == 1, seen2)
    check("그 한 번은 처음부터 숨긴 상태다", seen2[:1] == [True], seen2)

    print("[3] 숨김이 '최후 보루'인지 — 그 앞에 이미 두 단계 우회가 있다")
    src = (ROOT / "serve_b2b.py").read_text(encoding="utf-8", errors="replace")
    check("직접 주입이 막히면 임시 .xlsm 러너로 우회한다",
          "_run_vba_via_runner_with_retry(app, wb, code, entry)" in src)
    check("그래도 막히면 격리 인스턴스 폴백이 있다",
          "VBA-LIVE-BLOCKED -> isolated single-step fallback" in src)
finally:
    S.EXCEL_SESSIONS.pop("flick", None)
    try:
        if app is not None:
            for w in list(app.Workbooks):
                try: w.Close(SaveChanges=False)
                except Exception: pass
            app.Quit()
    except Exception:
        pass
    shutil.rmtree(work, ignore_errors=True)

print("")
print("RESULT: ALL PASS" if not fails else "RESULT: %d FAIL -> %s" % (len(fails), fails))
sys.exit(0 if not fails else 1)
