# 복붙 false-refusal 원인 분리: (a)전체스키마 (b)단일파일 최소스키마 로 #71/#69 재현 비교.
import sys, os, re
sys.path.insert(0, os.path.dirname(__file__))
import _qwen_client as q
import _eval_batch as b
from openpyxl import load_workbook

HERE = os.path.dirname(__file__)
base = open(os.path.join(HERE, "_prompt_current.txt"), encoding="utf-8").read()
full = b.build_schema_all()

def mini_schema():
    p = os.path.join(HERE, "..", "test_data", "v059_복붙캡처.xlsx")
    wb = load_workbook(p, data_only=True, read_only=True)
    out = ["## 현재 파일 스키마", "\n### 파일: v059_복붙캡처.xlsx", "시트: " + ", ".join(wb.sheetnames)]
    for sn in wb.sheetnames:
        ws = wb[sn]
        rows = [r for i, r in enumerate(ws.iter_rows(values_only=True)) if i < 3]
        out.append("  [%s] ~%s행" % (sn, ws.max_row))
        if rows:
            out.append("    1행: " + " | ".join([("" if v is None else str(v))[:18] for v in rows[0][:16]]))
    wb.close()
    return "\n".join(out)

def refuses(content):
    c = content or ""
    has_code = bool(re.search(r"def\s+transform\s*\(\s*ctx", re.search(r"```(?:python)?\s*(.*?)```", c, re.S).group(1) if re.search(r"```", c, re.S) else ""))
    said_notopen = ("목록에 없" in c) or ("업로드" in c and "def transform" not in c)
    return ("REFUSE" if (said_notopen and not has_code) else ("CODE" if has_code else "ASK")), has_code

req71 = "v059_복붙캡처.xlsx의 원본 시트 A1:E6을 대상 시트 A1에 복사해줘"
req69 = "v059_복붙캡처.xlsx의 원본 시트 A:E 전체 열을 G1에 복사해줘"
mini = mini_schema()
for name, req in [("#71", req71), ("#69", req69)]:
    for label, sch in [("full(17개)", full), ("mini(1개)", mini)]:
        r = q.chat(base + "\n\n" + sch, req, max_tokens=900, temperature=0.0,
                   extra={"chat_template_kwargs": {"enable_thinking": False}})
        verdict, hc = refuses(r["content"])
        print("%s  %-10s -> %-7s (code=%s)" % (name, label, verdict, hc))
