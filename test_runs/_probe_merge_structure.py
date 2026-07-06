# 소스(로우데이터)와 대상(콜센터)의 병합/가입번호 구조 확인 — 왜 '첫 행만/중복제거'가 됐는지 근거.
import sys, io, os, glob
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

DOWN = r"C:\Users\Admin\Downloads"
# 실제 파일명 찾기(%20/공백 혼동 방지)
def find(pat):
    hits = glob.glob(os.path.join(DOWN, pat))
    return [h for h in hits if not os.path.basename(h).startswith("~$")]

print("=== Downloads 후보 파일 ===")
for p in find("*농협생명*") + find("531611708899*.xlsx"):
    print("  ", os.path.basename(p))

SRC = (find("531611708899*로우데이터*DSMC_260616.xlsx") or [None])[0]
DST_CANDS = find("LGU*농협생명*콜센터*DSMC_260617.xlsx")
DST = (DST_CANDS or [None])[0]
print("\nSRC =", SRC)
print("DST =", DST)

def dump(path, sheet_hint, key_col, key_range, note):
    print("\n" + "=" * 72)
    print(f"[{note}] {os.path.basename(path)}")
    wb = load_workbook(path, data_only=True)
    print("  시트:", wb.sheetnames)
    ws = wb[sheet_hint] if sheet_hint in wb.sheetnames else wb[wb.sheetnames[0]]
    print(f"  대상시트='{ws.title}'  dims={ws.dimensions}  max_row={ws.max_row} max_col={ws.max_column}")
    merges = list(ws.merged_cells.ranges)
    print(f"  병합 영역 개수: {len(merges)}")
    # key_col(가입번호) 병합만 추려서 몇 행씩 묶였는지
    kc = key_col
    keymerges = [m for m in merges if m.min_col <= kc <= m.max_col]
    print(f"  '{get_column_letter(kc)}열' 관련 병합: {len(keymerges)}개")
    for m in sorted(keymerges, key=lambda x: x.min_row)[:12]:
        v = ws.cell(m.min_row, kc).value
        print(f"     {m.coord}  ({m.max_row-m.min_row+1}행)  값={v!r}")
    return ws

if SRC:
    ws = dump(SRC, "sheet", 2, "B3:B345", "SRC 로우데이터")
    # 소스 B열 가입번호가 반복/블록인지: B3~B40 값 + 병합여부
    print("\n  -- SRC B열(가입번호) B3:B30 값/병합 --")
    for r in range(3, 31):
        c = ws.cell(r, 2)
        mc = any(m.min_row <= r <= m.max_row and m.min_col <= 2 <= m.max_col for m in ws.merged_cells.ranges)
        print(f"     B{r}: {c.value!r}  merged={mc}")

if DST:
    ws = dump(DST, "콜센터", 2, "B6:B89", "DST 콜센터")
    print("\n  -- DST B열 B4:B30 값/병합 --")
    for r in range(4, 31):
        c = ws.cell(r, 2)
        mc = any(m.min_row <= r <= m.max_row and m.min_col <= 2 <= m.max_col for m in ws.merged_cells.ranges)
        print(f"     B{r}: {c.value!r}  merged={mc}")
