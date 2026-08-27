# -*- coding: utf-8 -*-
"""[제보 2026-08-27] 12자리 가입번호를 두 열에서 대조하면 '못 찾음'이 나온다 — 실제 Excel 검증.

제보 시나리오(회선 현황 시트)
  1. 정지회선 가입번호를 V열에 복붙
  2. "C열과 V열에 같은 번호가 있는지 확인"        → 못 찾음
  3. V열 번호 하나를 C열 밑에 복사한 뒤 다시 확인  → 그래도 못 찾음
  4. "중복 확인"                                  → 못 찾음
  5. 가입번호 12자리를 직접 넣어 찾게 하면          → 그중 하나만 찾음

원인(실측 확인): Excel 은 같은 '512102403338' 을 셀마다 숫자로도 글자로도 담는다. 복붙한 열은 글자,
  원래 열은 숫자인 식으로 섞인다. ctx.read 는 숫자 셀을 float 로 주므로 파이썬에선 512102403338.0 이 되고,
  공백/대소문자만 지우던 ctx.normalize 로는 '512102403338.0' != '512102403338' 로 갈렸다.
  오류도 안 나고 결과만 틀리는 조용한 오답이라, 사용자는 프로그램이 '못 찾는다'고만 볼 수 있었다.
  같은 목적의 _norm_key 가 이미 있었지만 copy_key_blocks/sum_lookup 만 쓰고 있었다.

곁다리로 잡은 것: normalize_text 는 `str(value or "")` 라 숫자 0·False 를 빈 문자열로 만든다.
  그대로 값 비교에 쓰면 '0 이 든 셀'과 '빈 셀'이 같은 값이 된다 → 조건 집계·중복 판정이 조용히 틀린다.
  값 비교용 normalize_value_key 는 normalize_text 에 기대지 않고 따로 구현한다.
"""
import sys, io, tempfile, shutil
from pathlib import Path
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")
ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT))
import openpyxl
import pythoncom; pythoncom.CoInitialize()
import win32com.client as win32
import serve_b2b as S

fails = []


def check(name, cond, detail=""):
    print(("  PASS  " if cond else "  FAIL  ") + name + (("  -> " + str(detail)[:250]) if not cond else ""))
    if not cond:
        fails.append(name)


print("[1] 값 비교 정규화의 계약 (Excel 없이)")
for v, want in [(True, "true"), (False, "false"), (None, ""), ("", ""),
                (0, "0"), (0.0, "0"), (512102403338.0, "512102403338"),
                ("512102403338", "512102403338"), (1234.5, "1234.5"),
                (1e16, "10000000000000000"), (" 가 나 ", "가나"), ("ABC", "abc")]:
    got = S.normalize_value_key(v)
    check("normalize_value_key(%r) = %r" % (v, want), got == want, got)
check("숫자 0 과 빈칸을 구분한다(집계·중복 판정의 전제)",
      S.normalize_value_key(0) != S.normalize_value_key(None))
check("숫자 셀과 글자 셀이 같은 값이면 같게 본다",
      S.normalize_value_key(512102403338.0) == S.normalize_value_key("512102403338"))

NUMS = ["512102403338", "512102403339", "512102403340"]
work = Path(tempfile.mkdtemp(prefix="b2b_keymatch_"))
src = work / "회선현황.xlsx"
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "회선 현황"
ws["C1"] = "가입번호"; ws["V1"] = "정지회선"
ws["C2"] = int(NUMS[0])      # 숫자 셀
ws["C3"] = NUMS[1]           # 글자 셀
ws["C4"] = int(NUMS[2])      # 숫자 셀
for i, n in enumerate(NUMS, start=2):
    ws.cell(row=i, column=22).value = n     # V열은 복붙 → 글자
ws["C5"] = 0                 # 0 과 빈칸 구분 확인용
ws["C6"] = None
wb.save(str(src))

app = None
try:
    app = win32.DispatchEx("Excel.Application")
    app.Visible = False
    app.DisplayAlerts = False
    book = app.Workbooks.Open(str(src))
    ctx = S.PythonComSkillContext(app, book, {"name": "회선현황.xlsx"}, timeout_s=300)

    print("[2] 제보 2번 — C열과 V열에 같은 번호가 있는지")
    c = ctx.read("회선 현황", "C2:C4")
    v = ctx.read("회선 현황", "V2:V4")
    types = [type(c[i][0]).__name__ for i in range(3)]
    check("셀 타입이 실제로 섞여 있다(재현 조건 성립)", "float" in types and "str" in types, types)
    matched = sum(1 for i in range(3) if ctx.normalize(c[i][0]) == ctx.normalize(v[i][0]))
    check("3개 모두 같은 번호로 판정", matched == 3, "%d/3" % matched)

    print("[3] 제보 5번 — 12자리 번호를 직접 주고 찾기")
    cvals = [ctx.normalize(r[0]) for r in ctx.read("회선 현황", "C2:C4")]
    found = [n for n in NUMS if ctx.normalize(n) in cvals]
    check("3개 다 찾는다(예전엔 1개만)", len(found) == 3, found)

    print("[4] 제보 4번 — 중복 확인")
    ctx.write_cell("회선 현황", "C7", NUMS[0])      # 글자로 한 번 더(숫자 셀 C2 와 중복)
    col = [ctx.normalize(r[0]) for r in ctx.read("회선 현황", "C2:C7") if ctx.normalize(r[0])]
    dup = [k for k in set(col) if col.count(k) > 1]
    check("숫자 셀과 글자 셀에 걸친 중복을 잡는다", ctx.normalize(NUMS[0]) in dup, dup)

    print("[5] 곁다리 — 0 이 든 셀과 빈 셀을 섞지 않는다")
    z = ctx.read("회선 현황", "C5:C6")
    check("0 과 빈칸이 다른 값", ctx.normalize(z[0][0]) != ctx.normalize(z[1][0]),
          (ctx.normalize(z[0][0]), ctx.normalize(z[1][0])))

    print("[6] 조건 집계(sum_where)도 같은 규칙을 쓴다")
    ctx.write("회선 현황", "Y1", [["코드", "금액"], [NUMS[0], 100], [NUMS[1], 200]])
    ctx.write_cell("회선 현황", "Y2", int(NUMS[0]))      # 숫자 셀로 저장
    t = ctx.sum_where("회선 현황", "Z", [("Y", NUMS[0])], header_row=1)
    check("숫자로 저장된 셀도 글자 조건으로 걸린다", t == 100, t)
finally:
    try:
        if app is not None:
            for w in list(app.Workbooks):
                try: w.Close(SaveChanges=False)
                except Exception: pass
            app.Quit()
    except Exception:
        pass
    shutil.rmtree(work, ignore_errors=True)

print("")
print("RESULT: ALL PASS" if not fails else "RESULT: %d FAIL -> %s" % (len(fails), fails))
sys.exit(0 if not fails else 1)
