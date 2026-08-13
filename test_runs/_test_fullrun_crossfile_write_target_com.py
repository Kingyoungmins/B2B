# -*- coding: utf-8 -*-
"""[스모크 / 제보 재현 2026-08-13] "A 파일에서 B 파일에 피벗을 만들거나 붙여넣는 스텝이
실행기에서 B 로 안 가고 A 안에 만들어진다."

실제 Excel 로 전체실행을 돌려 두 가지를 확인한다.

  1. 그룹이 맞으면 결과도 맞는가
       group=B 인데 소스 시트가 A 에만 있는 피벗 → 피벗 시트는 B 에 생겨야 한다(A 에는 없어야).
       (백엔드는 새 시트를 언제나 '그 그룹의 워크북'에 만든다 — ctx 에 물린 워크북.)
  2. 그룹이 틀리면 정확히 제보된 증상이 나오는가  ← 대조군
       같은 코드를 group=A 로 보내면 피벗이 A 에 생긴다. 즉 이 버그는 백엔드가 아니라
       '어느 파일에 쓸지 정해서 넘겨주는 쪽'의 문제라는 걸 실물로 못박는다.

왜 '피벗만' 티가 나는지도 같이 본다: 기존 시트에 값만 쓰는 스텝은 백엔드가 시트 이름으로
다른 워크북을 찾아가 주기 때문에 그룹이 틀려도 결과가 맞는다.
"""
import io
import shutil
import sys
import tempfile
from pathlib import Path

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")
ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT))

import openpyxl
import pythoncom
pythoncom.CoInitialize()
import win32com.client as win32
import serve_b2b as S

NA, NB = "A.xlsx", "B.xlsx"
work = Path(tempfile.mkdtemp(prefix="b2b_xfile_target_"))

_p = _f = 0


def t(name, cond, got=None):
    global _p, _f
    if cond:
        _p += 1
        print("PASS " + name)
    else:
        _f += 1
        print("FAIL " + name + ("" if got is None else "  got=%r" % (got,)))


def build_a(path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "원본"
    ws.append(["구분", "금액"])
    for row in [["갑", 10], ["을", 20], ["갑", 30]]:
        ws.append(row)
    wb.save(str(path))


def build_b(path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "요약"
    ws["A1"] = "요약시트"
    wb.save(str(path))


def sheets(wb):
    return [str(wb.Worksheets(i + 1).Name) for i in range(int(wb.Worksheets.Count))]


srcA, srcB = work / NA, work / NB
build_a(srcA)
build_b(srcB)

PIVOT = ('def transform(ctx):\n'
         '    ctx.pivot("원본", group_by=["구분"], value="금액", agg="sum", dest_name="집계")\n')
PASTE = ('def transform(ctx):\n'
         '    ctx.copy("A.xlsx!원본", "A1:B2", "요약", "D1")\n')

app = None
try:
    app = win32.DispatchEx("Excel.Application")
    app.Visible = False
    app.DisplayAlerts = False

    def open_session(sid, name, src):
        live = work / ("live_%s_%s" % (sid, name))
        shutil.copy2(src, live)
        wb = app.Workbooks.Open(str(live))
        S.EXCEL_SESSIONS[sid] = {"app": app, "workbook": wb, "path": str(live),
                                 "openPath": str(wb.FullName), "name": name,
                                 "sourcePath": str(src), "liveEditable": True}
        return wb

    # ── 1) 그룹이 맞을 때: 소스는 A, 대상 그룹은 B → 피벗은 B 에 ──────────────
    wA = open_session("exA", NA, srcA)
    wB = open_session("exB", NB, srcB)
    groups = [{"excelId": "exB", "steps": [{"code": PIVOT, "language": "python", "stepIdx": 0, "stepId": "p1"}]}]
    out = S.run_full_pipeline_single_instance(groups, reset_excel_ids=["exA", "exB"])
    t("전체실행 성공", bool(out.get("ok")), out)
    sa, sb = sheets(wA), sheets(wB)
    t("피벗이 대상 파일(B)에 생긴다", "집계" in sb, sb)
    t("소스 파일(A)에는 안 생긴다  ← 제보 증상이 아님", "집계" not in sa, sa)

    # 스텝별 쓰기 증거도 확인 — 이 스텝은 B 만 건드렸으므로 교차 대상이 없어야 한다
    sc = out.get("stepCross") or []
    t("스텝별 쓰기 증거가 실린다", len(sc) == 1, sc)
    if sc:
        t("피벗은 구조 변경이라 추적 불가로 남는다(설계)", sc[0].get("tracked") is False, sc[0])

    for k in ("exA", "exB"):
        S.EXCEL_SESSIONS.pop(k, None)
    wA.Close(SaveChanges=False)
    wB.Close(SaveChanges=False)

    # ── 2) 대조군: 같은 코드를 A 그룹으로 보내면 A 에 생긴다 = 제보된 증상 ──────
    wA2 = open_session("exA2", NA, srcA)
    wB2 = open_session("exB2", NB, srcB)
    groups_wrong = [{"excelId": "exA2", "steps": [{"code": PIVOT, "language": "python", "stepIdx": 0, "stepId": "p2"}]}]
    S.run_full_pipeline_single_instance(groups_wrong, reset_excel_ids=["exA2", "exB2"])
    sa2, sb2 = sheets(wA2), sheets(wB2)
    t("그룹이 A 면 피벗이 A 에 생긴다  ← 이게 제보된 증상", "집계" in sa2, sa2)
    t("그때 B 에는 안 생긴다", "집계" not in sb2, sb2)
    print("   → 같은 코드가 그룹에 따라 갈린다 = 원인은 백엔드가 아니라 '대상을 정해 넘기는 쪽'")

    for k in ("exA2", "exB2"):
        S.EXCEL_SESSIONS.pop(k, None)
    wA2.Close(SaveChanges=False)
    wB2.Close(SaveChanges=False)

    # ── 3) '기존 시트에 값 쓰기'는 그룹이 맞아도 틀려도 결과가 같다 ────────────
    #     (그래서 피벗/새 시트만 티가 났다)
    wA3 = open_session("exA3", NA, srcA)
    wB3 = open_session("exB3", NB, srcB)
    groups_paste = [{"excelId": "exB3", "steps": [{"code": PASTE, "language": "python", "stepIdx": 0, "stepId": "p3"}]}]
    S.run_full_pipeline_single_instance(groups_paste, reset_excel_ids=["exA3", "exB3"])
    pasted = wB3.Worksheets("요약").Range("D1").Value
    t("A 를 읽어 B 의 요약 시트에 붙여넣기 성공", str(pasted or "") == "구분", pasted)
    for k in ("exA3", "exB3"):
        S.EXCEL_SESSIONS.pop(k, None)
    wA3.Close(SaveChanges=False)
    wB3.Close(SaveChanges=False)

finally:
    try:
        if app is not None:
            app.Quit()
    except Exception:
        pass
    shutil.rmtree(work, ignore_errors=True)

print("")
print("%d passed, %d failed" % (_p, _f))
sys.exit(0 if _f == 0 else 1)
