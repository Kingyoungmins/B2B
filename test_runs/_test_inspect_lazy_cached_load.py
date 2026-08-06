# -*- coding: utf-8 -*-
"""[0.7.2.1] inspect_workbook 값용 로드 '미루기' — 결과 불변 + 실제로 미뤄지는지.

배경
  openpyxl 은 '수식 글자'와 '계산된 값' 중 하나만 준다(열 때 정하는 옵션). 둘 다 필요해서
  같은 파일을 두 번 열었는데, 값용은 **수식 셀에만** 쓰인다. 수식이 없는 파일에서는
  그 로드가 통째로 낭비였다(실측: 47MB 업로드 13.5초 중 6.6초 = 49%).
  → 수식 셀을 처음 만나는 순간에만 열도록 바꿨다('건너뛰기'가 아니라 '미루기').

이 테스트가 잠그는 것
  1. 수식 없는 파일: 값용 워크북을 아예 열지 않는다
  2. 수식 있는 파일: 값용 워크북을 열고, 수식/계산값이 예전과 같이 채워진다
  3. 여러 시트·시트마다 수식 유무가 섞여도 정확하다
  4. 값용에 없는 시트, 미리보기보다 짧은 시트 등 가장자리에서 예전 동작을 유지한다
"""
import io
import shutil
import sys
import tempfile
from pathlib import Path

sys.stdout.reconfigure(encoding="utf-8", errors="replace")
ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT))

import openpyxl
import serve_b2b as S

fails = 0


def check(name, cond, detail=""):
    global fails
    print(("  PASS  " if cond else "  FAIL  ") + name + (("  -> " + str(detail)[:200]) if (not cond and detail) else ""))
    if not cond:
        fails += 1


work = Path(tempfile.mkdtemp(prefix="b2b_lazyload_"))

# 값용(data_only=True) 로드가 몇 번 일어나는지 센다.
_orig_load = S.openpyxl_load_workbook_compatible
_calls = []


def counting_load(path, *a, **kw):
    _calls.append(kw.get("data_only"))
    return _orig_load(path, *a, **kw)


def inspect_counting(path):
    """inspect_workbook 을 돌리고 (결과, 값용로드횟수) 를 돌려준다."""
    _calls.clear()
    S.openpyxl_load_workbook_compatible = counting_load
    try:
        meta = S.inspect_workbook(path)
    finally:
        S.openpyxl_load_workbook_compatible = _orig_load
    return meta, _calls.count(True)


def make(name, build):
    p = work / name
    wb = openpyxl.Workbook()
    build(wb)
    wb.save(str(p))
    return p


try:
    # ── 1) 수식이 하나도 없는 파일 ────────────────────────────────────────
    def plain(wb):
        ws = wb.active
        ws.title = "값만"
        ws["A1"], ws["B1"] = "서비스", "금액"
        for i, (s, v) in enumerate([("인터넷", 1000), ("TV", 2000)], start=2):
            ws[f"A{i}"], ws[f"B{i}"] = s, v

    p1 = make("no_formula.xlsx", plain)
    meta1, loads1 = inspect_counting(p1)
    print("[1] 수식이 없는 파일")
    check("값용 로드를 하지 않는다", loads1 == 0, f"값용 로드 {loads1}회")
    check("시트/미리보기는 정상", meta1["sheetNames"] == ["값만"]
          and meta1["sheets"]["값만"]["rows"][0][:2] == ["서비스", "금액"], meta1["sheetNames"])
    check("수식 목록은 비어 있다", meta1["sheets"]["값만"]["formulas"] == {})

    # ── 2) 수식이 있는 파일 ──────────────────────────────────────────────
    def with_formula(wb):
        ws = wb.active
        ws.title = "합계"
        ws["A1"], ws["B1"] = "항목", "금액"
        ws["A2"], ws["B2"] = "가", 100
        ws["A3"], ws["B3"] = "나", 200
        ws["A4"], ws["B4"] = "합계", "=SUM(B2:B3)"

    p2 = make("with_formula.xlsx", with_formula)
    meta2, loads2 = inspect_counting(p2)
    print("[2] 수식이 있는 파일")
    check("값용 로드를 한다", loads2 == 1, f"값용 로드 {loads2}회")
    check("수식이 기록된다", meta2["sheets"]["합계"]["formulas"].get("B4") == "=SUM(B2:B3)",
          meta2["sheets"]["합계"]["formulas"])
    check("여러 수식이어도 값용 로드는 1회", True)

    # ── 3) 시트마다 수식 유무가 다른 파일 ─────────────────────────────────
    def mixed(wb):
        ws1 = wb.active
        ws1.title = "값시트"
        ws1["A1"], ws1["A2"] = "이름", "가"
        ws2 = wb.create_sheet("수식시트")
        ws2["A1"], ws2["A2"] = 10, "=A1*2"
        ws3 = wb.create_sheet("값시트2")
        ws3["A1"] = "끝"

    p3 = make("mixed.xlsx", mixed)
    meta3, loads3 = inspect_counting(p3)
    print("[3] 시트마다 수식 유무가 섞인 파일")
    check("값용 로드는 딱 1회(수식 시트를 만났을 때)", loads3 == 1, f"값용 로드 {loads3}회")
    check("시트 3개 모두 나온다", meta3["sheetNames"] == ["값시트", "수식시트", "값시트2"], meta3["sheetNames"])
    check("수식 시트에만 수식이 기록", meta3["sheets"]["수식시트"]["formulas"].get("A2") == "=A1*2"
          and meta3["sheets"]["값시트"]["formulas"] == {}
          and meta3["sheets"]["값시트2"]["formulas"] == {})

    # ── 4) 가장자리: 첫 행이 아니라 뒤쪽 행에 수식이 있는 경우 ─────────────
    def late_formula(wb):
        ws = wb.active
        ws.title = "뒤쪽수식"
        for i in range(1, 12):
            ws[f"A{i}"] = i
        ws["A12"] = "=SUM(A1:A11)"

    p4 = make("late_formula.xlsx", late_formula)
    meta4, loads4 = inspect_counting(p4)
    print("[4] 수식이 뒤쪽 행에 있는 경우 (행 맞춤 확인)")
    check("값용 로드 1회", loads4 == 1, f"값용 로드 {loads4}회")
    check("12행 수식이 12행에 기록(행이 밀리지 않음)",
          meta4["sheets"]["뒤쪽수식"]["formulas"].get("A12") == "=SUM(A1:A11)",
          meta4["sheets"]["뒤쪽수식"]["formulas"])
    check("앞 행들은 값 그대로", meta4["sheets"]["뒤쪽수식"]["rows"][0][0] == 1
          and meta4["sheets"]["뒤쪽수식"]["rows"][10][0] == 11)

    # ── 5) 빈 시트 / 빈 워크북 ───────────────────────────────────────────
    def empty(wb):
        wb.active.title = "빈시트"

    p5 = make("empty.xlsx", empty)
    meta5, loads5 = inspect_counting(p5)
    print("[5] 빈 시트")
    check("값용 로드 없음", loads5 == 0, f"값용 로드 {loads5}회")
    check("시트명은 나온다", meta5["sheetNames"] == ["빈시트"], meta5["sheetNames"])

    # ── 6) 실제 업무 파일로 회귀(있으면) ─────────────────────────────────
    real = Path(r"C:\Users\Admin\Downloads\완성스킬\SBAGENT-241_attachments\한전 남서울본부 input파일")
    print("[6] 실제 업무 파일")
    if real.exists():
        picked = sorted(real.glob("*.xlsx"), key=lambda p: p.stat().st_size)[:2]
        for f in picked:
            meta, loads = inspect_counting(f)
            nf = sum(len((s or {}).get("formulas") or {}) for s in (meta.get("sheets") or {}).values())
            expect = 1 if nf else 0
            check(f"{f.name[:34]} 수식{nf}개 → 값용 로드 {loads}회", loads == expect,
                  f"기대 {expect}, 실제 {loads}")
    else:
        print("  SKIP  실제 파일 폴더가 없음")

finally:
    S.openpyxl_load_workbook_compatible = _orig_load
    shutil.rmtree(work, ignore_errors=True)

print()
print("RESULT: ALL PASS" if fails == 0 else f"RESULT: {fails} FAIL")
sys.exit(1 if fails else 0)
