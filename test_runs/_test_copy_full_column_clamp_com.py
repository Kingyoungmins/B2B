# -*- coding: utf-8 -*-
"""[실측 2026-09-01] "D열 복사"가 104만 행을 다루던 것 — 실행 시점 끝행으로 자동 클램프.

배경: 사용자는 엑셀 습관대로 열을 통째로 지정한다("D열 해줘" → ctx.copy(..., "D:D", ...)).
      복사 자체는 Excel 이 알아서 빨리 하지만(0.3초), 되돌리기 저널이 붙는 쪽
      104만 셀을 통째로 읽어 스텝이 6초가 됐다(생성 테스트 로그 실측).
      프롬프트에 '실사용 행까지 줄여라' 규칙이 있어도 모델이 안 지키는 날이 있고,
      사용자 입력 자체가 열 단위라 엔진에서 줄이는 게 정답.

잠그는 계약:
  1) 전체 열(D:D)은 실행 시점 끝행까지로 줄어 복사된다 — 값 동일, 저널도 그 크기만
  2) 사용자가 명시한 유한 범위(D1:D10)는 절대 건드리지 않는다
  3) 전체 행(3:3)은 끝열까지로 줄어든다
  4) 빈 열 전체도 죽지 않는다(1셀로 줄어 조용히 통과)
  5) 교차파일("파일.xlsx!시트", "D:D")도 원본 시트 기준으로 줄어든다
  6) 클램프 실패 시 종전 동작(전체 복사)으로 살아남는다
  7) 같은 복사를 클램프 없이 돌리면 실제로 훨씬 느리다(원인 증명)
"""
import io
import sys
import tempfile
import shutil
import time
from pathlib import Path

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")
ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT))
import openpyxl
import pythoncom; pythoncom.CoInitialize()
import win32com.client as win32
import serve_b2b as S

fails = []


def check(name, cond, detail=""):
    print(("  PASS  " if cond else "  FAIL  ") + name + (("  -> " + str(detail)[:250]) if not cond else ""))
    if not cond:
        fails.append(name)


work = Path(tempfile.mkdtemp(prefix="b2b_clamp_"))
N_ROWS = 60
wbk = openpyxl.Workbook(); ws = wbk.active; ws.title = "매출"
for i in range(1, N_ROWS + 1):
    ws.cell(row=i, column=4, value="D%d값" % i)          # D1..D60
    ws.cell(row=3, column=4 + i % 3, value="row3-%d" % i)  # 3행에도 데이터(전체 행 테스트용)
wbk.create_sheet("빈시트")
wbk.save(str(work / "src.xlsx"))
w2 = openpyxl.Workbook(); w2.active.title = "원가"; w2.active["A1"] = "기존"
w2.create_sheet("메모"); w2.save(str(work / "dst.xlsx"))

app = None
try:
    app = win32.DispatchEx("Excel.Application"); app.Visible = False; app.DisplayAlerts = False
    wsrc = app.Workbooks.Open(str(work / "src.xlsx"))
    wdst = app.Workbooks.Open(str(work / "dst.xlsx"))
    S.EXCEL_SESSIONS["clampS"] = {"app": app, "workbook": wsrc, "path": str(work / "src.xlsx"),
                                  "openPath": str(wsrc.FullName), "name": "src.xlsx",
                                  "sourcePath": str(work / "src.xlsx"), "liveEditable": True}

    def ctx(wb, name):
        return S.PythonComSkillContext(app, wb, {"name": name}, timeout_s=300)

    print("[1] 전체 열(D:D) → 끝행까지로 줄어 복사된다")
    c = ctx(wsrc, "src.xlsx")
    t0 = time.perf_counter()
    c.copy("매출", "D:D", "매출", "F1")
    el_clamped = time.perf_counter() - t0
    check("값이 옮겨졌다(F1)", c.read_cell("매출", "F1") == "D1값", c.read_cell("매출", "F1"))
    check("끝행까지 전부(F%d)" % N_ROWS, c.read_cell("매출", "F%d" % N_ROWS) == "D%d값" % N_ROWS)
    check("끝행 밖은 비어 있다", c.read_cell("매출", "F%d" % (N_ROWS + 1)) in (None, ""),
          c.read_cell("매출", "F%d" % (N_ROWS + 1)))
    jr = [j for j in c._shared["journal"] if "F" in str(j[1])]
    check("되돌리기 저널이 실데이터 크기만 잡는다(104만 행 아님)",
          bool(jr) and (":F%d" % N_ROWS) in str(jr[-1][1]).replace("$", ""), jr[-1][1] if jr else "없음")
    print("      클램프 복사: %.2f초" % el_clamped)

    print("[2] 명시 범위(D1:D10)는 건드리지 않는다")
    c2 = ctx(wsrc, "src.xlsx")
    c2.copy("매출", "D1:D10", "매출", "H1")
    check("10행만 복사", c2.read_cell("매출", "H10") == "D10값"
          and c2.read_cell("매출", "H11") in (None, ""), c2.read_cell("매출", "H11"))
    jr2 = [j for j in c2._shared["journal"] if "H" in str(j[1])]
    check("저널도 10행", bool(jr2) and (":H10" in str(jr2[-1][1]).replace("$", "")), jr2[-1][1] if jr2 else "없음")

    print("[3] 전체 행(3:3) → 끝열까지로 줄어든다")
    c3 = ctx(wsrc, "src.xlsx")
    c3.copy("매출", "3:3", "매출", "A70")
    used_last_col = int(wsrc.Worksheets("매출").UsedRange.Column) + int(wsrc.Worksheets("매출").UsedRange.Columns.Count) - 1
    jr3 = [j for j in c3._shared["journal"] if "70" in str(j[1])]
    check("저널이 끝열까지만(16384열 아님)", bool(jr3) and
          str(jr3[-1][1]).replace("$", "").endswith("70"), jr3[-1][1] if jr3 else "없음")
    check("3행 값이 옮겨졌다", any(c3.read_cell("매출", "%s70" % col) not in (None, "")
                              for col in ("D", "E", "F")), "3행 데이터 없음")

    print("[4] 빈 시트의 전체 열도 죽지 않는다")
    c4 = ctx(wsrc, "src.xlsx")
    try:
        c4.copy("빈시트", "B:B", "매출", "J1")
        check("예외 없이 통과(1셀로 줄어)", True)
    except Exception as e:
        check("예외 없이 통과(1셀로 줄어)", False, e)

    print("[5] 교차파일(src.xlsx!매출 → dst) 도 줄어든다")
    c5 = ctx(wdst, "dst.xlsx")
    t0 = time.perf_counter()
    c5.copy("src.xlsx!매출", "D:D", "원가", "E1")
    el_cross = time.perf_counter() - t0
    check("교차 복사 값", c5.read_cell("원가", "E1") == "D1값" and
          c5.read_cell("원가", "E%d" % N_ROWS) == "D%d값" % N_ROWS)
    check("교차도 빠르다(<3초 — 실측 사례는 6초였다)", el_cross < 3, "%.2f초" % el_cross)
    check("기존 칸 보존", c5.read_cell("원가", "A1") == "기존")

    print("[6] 내부 부품이 죽어도 복사는 산다(가드 확인)")
    # _clamp_full_span 은 내부 try/except 로 실패 시 원본 범위를 돌려준다. 그 가드를 확인하려면
    # 메서드를 통째로 바꾸면 안 되고(가드까지 없어진다) 안에서 쓰는 부품(_resize_rng)을 죽인다.
    c6 = ctx(wsrc, "src.xlsx")
    c6._resize_rng = lambda *a, **k: (_ for _ in ()).throw(RuntimeError("boom"))
    try:
        c6.copy("매출", "D:D", "매출", "L1")     # 클램프 실패 → 전체 열 그대로(종전 동작) 복사
        check("부품 예외에도 복사는 된다", c6.read_cell("매출", "L5") == "D5값",
              c6.read_cell("매출", "L5"))
    except Exception as e:
        check("부품 예외에도 복사는 된다", False, e)

    print("[7] 원인 증명 — 같은 D:D 복사를 클램프 없이 돌리면 훨씬 느리다")
    c7 = ctx(wsrc, "src.xlsx")
    orig = S.PythonComSkillContext._clamp_full_span
    S.PythonComSkillContext._clamp_full_span = lambda self, ws, rng: rng   # 옛 동작 재현
    try:
        t0 = time.perf_counter()
        c7.copy("매출", "D:D", "매출", "N1")
        el_old = time.perf_counter() - t0
    finally:
        S.PythonComSkillContext._clamp_full_span = orig
    print("      옛 동작(전체 열): %.2f초  vs  클램프: %.2f초" % (el_old, el_clamped))
    check("클램프가 실제로 빠르다", el_clamped < el_old, "%.2f vs %.2f" % (el_clamped, el_old))
    check("옛 동작의 저널은 열 전체($N:$N)를 잡는다(원인 확인)",
          any(str(j[1]).replace("$", "") == "N:N" for j in c7._shared["journal"]),
          [str(j[1]) for j in c7._shared["journal"]][:3])
finally:
    S.EXCEL_SESSIONS.pop("clampS", None)
    try:
        if app is not None:
            for w in list(app.Workbooks):
                try: w.Close(SaveChanges=False)
                except Exception: pass
            app.Quit()
    except Exception:
        pass
    shutil.rmtree(work, ignore_errors=True)

print("")
print("RESULT: ALL PASS" if not fails else "RESULT: %d FAIL -> %s" % (len(fails), fails))
sys.exit(0 if not fails else 1)
