# 프롬프트 회귀 게이트: 주어진 시스템 프롬프트 + test_data 스키마로 Qwen3.6 에 요청을 보내
# 생성된 코드가 '의도한 헬퍼/패턴'을 쓰는지 정적 채점. 프롬프트 정리 전/후 점수를 비교하는 용도.
# 사용: python _eval_gate.py <prompt.txt 경로> [라벨]
import sys, os, re, glob, json
sys.path.insert(0, os.path.dirname(__file__))
import _qwen_client as q
from openpyxl import load_workbook

TD = os.path.join(os.path.dirname(__file__), "..", "test_data")
FILES = ["input_매출_2026_4월.xlsx", "input_원가_2026_4월.xlsx", "output_청구서_템플릿.xlsx",
         "월증감_요금정산_2026년.xlsx"]

def build_schema():
    out = ["## 현재 파일 스키마"]
    for fn in FILES:
        p = os.path.join(TD, fn)
        if not os.path.exists(p):
            continue
        wb = load_workbook(p, data_only=True, read_only=True)
        out.append("\n### 파일: %s" % fn)
        out.append("시트: " + ", ".join(wb.sheetnames))
        for sn in wb.sheetnames[:12]:
            ws = wb[sn]
            rows = []
            for i, row in enumerate(ws.iter_rows(values_only=True)):
                rows.append(row)
                if i >= 3:
                    break
            ncol = max((len(r) for r in rows), default=0)
            out.append("  [%s] 약 %s행 x %s열" % (sn, ws.max_row, ncol))
            if rows:
                hdr = [("" if v is None else str(v))[:20] for v in rows[0][:20]]
                out.append("    1행: " + " | ".join(hdr))
                for r in rows[1:3]:
                    samp = [("" if v is None else str(v))[:16] for v in r[:20]]
                    out.append("    샘플: " + " | ".join(samp))
        wb.close()
    return "\n".join(out)

# (요청, 채점) — include_any: 하나라도 매칭이면 OK / exclude: 하나라도 매칭이면 FAIL
REQUESTS = [
    {"name": "피벗(회사별집계)",
     "req": "input_매출_2026_4월.xlsx 의 매출 시트에서 회사명별 매출 합계와 거래건수를 피벗으로 '회사별집계' 새 시트에 만들어줘.",
     "any": [r"\.pivot\s*\("]},
    {"name": "필터→새시트(원본보존)",
     "req": "output_청구서_템플릿.xlsx 의 a 시트에서 C열 값이 123인 행만 새 시트 'b'에 가져와. a시트는 원본이니 건들지 마.",
     "any": [r"\.filter_to_sheet\s*\("]},
    {"name": "정렬(열문자 기준)",
     "req": "output_청구서_템플릿.xlsx 의 a 시트를 T열 기준으로 오름차순 정렬해줘.",
     "any": [r"\.sort\s*\("], "exclude": [r"sorted\s*\(", r"\.write\([^)]*sorted"]},
    {"name": "열 삽입(전체열)",
     "req": "output_청구서_템플릿.xlsx 의 회귀_삽입삭제 시트에서 J열 앞에 새 열을 하나 삽입하고 헤더는 '새삽입열'로 해줘.",
     "any": [r"\.insert_cols\s*\("]},
    {"name": "복사(서식·병합·수식 유지)",
     "req": "output_청구서_템플릿.xlsx 의 회귀_병합복붙 시트의 A1:F7을 H1:M7에 병합·서식·수식 그대로 복사해줘.",
     "any": [r"\.copy\s*\(", r"\.copy_values\s*\("]},
    {"name": "값만 복사(수식 아님)",
     "req": "output_청구서_템플릿.xlsx 의 회귀_수식복사 시트의 A1:A5 값을 F1:F5에 적어줘. 수식 말고 보이는 값만 복사해.",
     "any": [r"\.copy_values\s*\(", r"overwrite_formulas\s*=\s*True"], "exclude": [r"(?<!_values)\.copy\s*\("]},
    {"name": "월 +1(shift_months)",
     "req": "output_청구서_템플릿.xlsx 대신 test_data 의 월증감_요금정산_2026년.xlsx 시트의 월 정보를 다음달로 +1 해줘.",
     "any": [r"\.shift_months\s*\("]},
    {"name": "중복 개수(삭제 금지)",
     "req": "output_청구서_템플릿.xlsx 의 회귀_중복계수 시트에서 H열 금액이 같은 값의 개수를 G열에 구해줘. 중복이라고 지우지는 마.",
     "exclude": [r"\.dedupe\s*\(", r"\.delete_rows\s*\("]},
    {"name": "정확매칭 합산(과포함 금지)",
     "req": "output_청구서_템플릿.xlsx 의 회귀_청구항목 시트에서 청구항목이 '기본료', '전국대표 포함한 기본료', '080포함한 기본료'인 행만 합산해줘.",
     # 정확매칭이면 '전국대표 포함한 기본료' 전체 문자열을 코드에 하드코딩한다(부분일치 "기본료"만으론 못 함).
     "any": [r"전국대표 포함한 기본료"],
     "exclude": [r"[\"']기본료[\"']\s+in\b", r"\bin\s+str\("]},
    {"name": "교차파일(원가→요약)",
     "req": "input_원가_2026_4월.xlsx 에서 회사별 원가 합계를 찾아 output_청구서_템플릿.xlsx 의 회사별요약 시트 원가 열에 채워줘.",
     "any": [r"\.book\s*\("]},
    # ── 규칙-찌르기(정리 단계에서 깨지기 쉬운 규칙) ──
    {"name": "음수부호 보존",
     "req": "input_원가_2026_4월.xlsx 의 원가 시트에서 단가 열 합계를 구해 빈 셀에 적어줘. 음수 값이 섞여 있으니 부호는 그대로 두고 더해.",
     "exclude": [r"\babs\s*\(", r"replace\(\s*[\"']-[\"']", r"re\.sub\(\s*r?[\"']\[\^0-9", r"\.strip\(\s*[\"']-"]},
    {"name": "개수 세줘=값(수식 아님)",
     "req": "output_청구서_템플릿.xlsx 의 회귀_중복계수 시트에서 H열 금액이 같은 값의 개수를 G열에 세어서 적어줘.",
     "exclude": [r"COUNTIF", r"write_formulas\s*\("]},
    {"name": "COUNTIF로=수식",
     "req": "output_청구서_템플릿.xlsx 의 회귀_중복계수 시트 G4:G15에 H열 금액별 개수를 COUNTIF 수식으로 넣어줘.",
     "any": [r"COUNTIF", r"write_formulas\s*\("]},
    {"name": "큰 표 정렬(1200행)→ctx.sort",
     "req": "input_매출_2026_4월.xlsx 의 매출 시트를 금액이 큰 순서로 정렬해줘. 전체 행이 같이 움직여야 해.",
     "any": [r"\.sort\s*\("], "exclude": [r"sorted\s*\("]},
    {"name": "마진 수식 채우기",
     "req": "output_청구서_템플릿.xlsx 의 회사별요약 시트 마진 열에 각 행 '매출 - 원가' 수식을 채워줘. 매출/원가 열은 그대로 두고.",
     "any": [r"write_formulas\s*\(", r"fill_sum_col\s*\(", r"f?[\"']=[^\"']*-"]},
]

def extract_code(text):
    m = re.search(r"```(?:python)?\s*(.*?)```", text, re.S)
    return m.group(1) if m else text

def score_one(system, item):
    r = q.chat(system, item["req"], max_tokens=1400, temperature=0.0,
               extra={"chat_template_kwargs": {"enable_thinking": False}})
    code = extract_code(r["content"])
    has_transform = bool(re.search(r"def\s+transform\s*\(\s*ctx\s*\)", code))
    ok = has_transform
    reason = []
    if not has_transform:
        reason.append("no transform")
    for rx in item.get("any", []) or []:
        pass
    if item.get("any"):
        if not any(re.search(rx, code) for rx in item["any"]):
            ok = False; reason.append("기대 헬퍼/패턴 없음")
    for rx in item.get("exclude", []) or []:
        if re.search(rx, code):
            ok = False; reason.append("금지 패턴: %s" % rx)
    return ok, reason, code, r["usage"]

def main():
    prompt_path = sys.argv[1]
    label = sys.argv[2] if len(sys.argv) > 2 else os.path.basename(prompt_path)
    system = open(prompt_path, encoding="utf-8").read() + "\n\n" + build_schema()
    print("=== EVAL: %s ===  (system %d자)" % (label, len(system)))
    passed = 0
    for it in REQUESTS:
        try:
            ok, reason, code, usage = score_one(system, it)
        except Exception as e:
            ok, reason, code = False, ["오류:%s" % e], ""
        passed += 1 if ok else 0
        mark = " OK " if ok else "FAIL"
        print("  %s  %-24s %s" % (mark, it["name"], "" if ok else "← " + ", ".join(reason)))
    print("\n>>> %s: %d/%d PASS" % (label, passed, len(REQUESTS)))

if __name__ == "__main__":
    main()
