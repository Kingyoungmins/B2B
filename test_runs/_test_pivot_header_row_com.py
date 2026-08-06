# -*- coding: utf-8 -*-
"""[실측/COM] 피벗 헤더 인자 문제 — 사용자 제보 2026-08-06 재현.

제보한 상황
  "새로운 시트에 피벗 테이블 생성. 행=서비스, 값=할인 후(합계)" 스킬이 적용 실패.
  실제 오류: pivot() got an unexpected keyword argument 'header_row'

원인 2가지
  ① 같은 뜻인데 이름이 둘이었다 — find_header 는 header_row(단수), pivot 은 header_rows(복수).
     단수 쓰는 헬퍼가 훨씬 많아 그 습관대로 쓰면 pivot 만 터졌다.
  ② 헤더가 2행인데 header_rows 를 안 주면 피벗은 1행만 보고 "필드를 못 찾았다"로 끝났다.
     find_header 는 예전부터 인접 행을 훑어 구제해 주는데 피벗만 안 그랬다.

이 테스트가 잠그는 것
  1. ctx.pivot(..., header_row=2)  → 받아준다(별칭)
  2. ctx.pivot(...)  헤더 인자 없이도 2행 헤더를 자동으로 찾는다
  3. ctx.pivot(..., header_rows=2) → 기존 방식 그대로 동작(회귀)
  4. 정말 없는 옵션 → 원시 TypeError 가 아니라 '쓸 수 있는 옵션'을 알려주는 오류

실행: python test_runs/_test_pivot_header_row_com.py   (Excel 필요)
"""
import io
import shutil
import sys
import tempfile
from pathlib import Path

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")
ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT))

import openpyxl
import pythoncom; pythoncom.CoInitialize()
import win32com.client as win32
import serve_b2b as S

fails = 0


def check(name, cond, detail=""):
    global fails
    print(("  PASS  " if cond else "  FAIL  ") + name + (("  -> " + str(detail)) if (not cond and detail) else ""))
    if not cond:
        fails += 1


work = Path(tempfile.mkdtemp(prefix="b2b_pivot_hdr_"))
SRC = work / "결과_A고객 전화서비스.xlsx"

# 제보 파일과 같은 모양: 1행은 제목, 2행이 헤더, 3행부터 데이터.
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "VIEW"
ws["A1"] = "A고객 전화서비스 정산 결과"          # 제목 행 (헤더 아님)
ws["A2"], ws["B2"], ws["C2"] = "서비스", "할인 전", "할인 후"
rows = [("인터넷", 30000, 25000), ("TV", 20000, 18000), ("인터넷", 30000, 27000),
        ("전화", 10000, 9000), ("TV", 20000, 16000)]
for i, (svc, before, after) in enumerate(rows, start=3):
    ws[f"A{i}"], ws[f"B{i}"], ws[f"C{i}"] = svc, before, after
wb.save(SRC)
EXPECTED = {"인터넷": 52000, "TV": 34000, "전화": 9000}

app = None
results = {}


def read_pivot(sheet_name):
    """만들어진 피벗 시트에서 (서비스 -> 합계) 를 읽는다. 값이 어느 열에 오든 찾아 읽는다."""
    tmp = work / f"out_{sheet_name}.xlsx"
    ctx_wb.SaveCopyAs(str(tmp))
    b = openpyxl.load_workbook(str(tmp), data_only=True)
    if sheet_name not in b.sheetnames:
        b.close()
        return None
    sh = b[sheet_name]
    got = {}
    for row in sh.iter_rows(values_only=True):
        if not row:
            continue
        key = row[0]
        if key is None or str(key).strip() in ("", "서비스", "총합계", "합계", "Grand Total"):
            continue
        num = None
        for v in row[1:]:
            if isinstance(v, (int, float)):
                num = float(v)
                break
        if num is not None:
            got[str(key).strip()] = num
    b.close()
    return got


try:
    app = win32.DispatchEx("Excel.Application")
    app.Visible = False
    app.DisplayAlerts = False
    ctx_wb = app.Workbooks.Open(str(SRC))
    ctx = S.PythonComSkillContext(app, ctx_wb, {"id": "t", "name": SRC.name}, timeout_s=180)

    print("[1] header_row (단수) 로 불러도 받아준다  ← 제보한 그 호출")
    try:
        ctx.pivot("VIEW", group_by="서비스", value="할인 후", agg="sum",
                  dest_name="P_단수", header_row=2)
        results["단수"] = read_pivot("P_단수")
        check("오류 없이 생성", True)
        check("집계값 정확", results["단수"] == EXPECTED, results["단수"])
    except Exception as err:
        check("오류 없이 생성", False, f"{type(err).__name__}: {err}")

    print("[2] 헤더 인자를 아예 안 줘도 2행 헤더를 찾아낸다")
    try:
        ctx.pivot("VIEW", group_by="서비스", value="할인 후", agg="sum", dest_name="P_자동")
        results["자동"] = read_pivot("P_자동")
        check("오류 없이 생성", True)
        check("집계값 정확", results["자동"] == EXPECTED, results["자동"])
    except Exception as err:
        check("오류 없이 생성", False, f"{type(err).__name__}: {err}")

    print("[3] header_rows (복수) 기존 방식도 그대로  ← 회귀")
    try:
        ctx.pivot("VIEW", group_by="서비스", value="할인 후", agg="sum",
                  dest_name="P_복수", header_rows=2)
        results["복수"] = read_pivot("P_복수")
        check("오류 없이 생성", True)
        check("집계값 정확", results["복수"] == EXPECTED, results["복수"])
    except Exception as err:
        check("오류 없이 생성", False, f"{type(err).__name__}: {err}")

    print("[4] find_header 도 두 이름 다 받는다(반대 방향 별칭)")
    try:
        a = ctx.find_header("VIEW", "할인 후", header_row=2)
        b = ctx.find_header("VIEW", "할인 후", header_rows=2)
        check("header_row / header_rows 모두 같은 결과", a == b == 3, f"{a} vs {b}")
    except Exception as err:
        check("header_row / header_rows 모두 같은 결과", False, f"{type(err).__name__}: {err}")

    print("[5] 정말 없는 옵션은 '쓸 수 있는 옵션'을 알려주는 오류")
    try:
        ctx.pivot("VIEW", group_by="서비스", value="할인 후", dest_name="P_없는옵션",
                  없는옵션=1)
        check("오류가 나야 한다", False, "오류 없이 통과함")
    except S.PythonComSkillError as err:
        msg = str(err)
        check("우리 오류로 승격(원시 TypeError 아님)", True)
        check("없는 옵션 이름을 짚어준다", "없는옵션" in msg, msg)
        check("쓸 수 있는 옵션을 알려준다", "header_rows" in msg and "group_by" in msg, msg)
    except Exception as err:
        check("우리 오류로 승격(원시 TypeError 아님)", False, f"{type(err).__name__}: {err}")

    print("[6] 헤더가 1행인 평범한 표는 예전 그대로  ← 회귀")
    ws2 = ctx_wb.Worksheets.Add()
    ws2.Name = "FLAT"
    ws2.Range("A1").Value = "서비스"
    ws2.Range("B1").Value = "할인 후"
    for i, (svc, _b, after) in enumerate(rows, start=2):
        ws2.Range(f"A{i}").Value = svc
        ws2.Range(f"B{i}").Value = after
    try:
        ctx.pivot("FLAT", group_by="서비스", value="할인 후", agg="sum", dest_name="P_평범")
        results["평범"] = read_pivot("P_평범")
        check("집계값 정확", results["평범"] == EXPECTED, results["평범"])
    except Exception as err:
        check("집계값 정확", False, f"{type(err).__name__}: {err}")

finally:
    try:
        if app is not None:
            for w in list(app.Workbooks):
                try:
                    w.Close(SaveChanges=False)
                except Exception:
                    pass
            app.Quit()
    except Exception:
        pass
    shutil.rmtree(work, ignore_errors=True)

print("\n" + ("RESULT: ALL PASS" if fails == 0 else f"RESULT: {fails} FAIL"))
sys.exit(1 if fails else 0)
