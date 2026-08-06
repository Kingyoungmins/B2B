# -*- coding: utf-8 -*-
"""[실측/COM] 되돌리기용 백업(internal) — 빨라지고, 되돌리기는 그대로인가.

배경 (VM 실측 2026-08-06)
  스텝 하나 적용에 10.8초 / 5.8초가 걸렸는데 그중 3.84초 / 2.56초가 '적용 전 백업'이었다.
  쪼개 보니 정작 파일 저장은 17ms 였고, 앞뒤 뒤치다꺼리(시트 보호 해제·재적용, 화면 설정 복구)가
  307ms 로 파일 저장의 18배였다. VM 은 Excel 명령이 느려 이 격차가 더 벌어진다.
  → 되돌리기용 백업은 사용자가 열어 볼 파일이 아니므로 그 뒤치다꺼리를 건너뛴다(internal=True).

이 테스트가 잠그는 것
  1. internal 백업이 일반 저장보다 빠르다
  2. internal 백업 파일이 정상 xlsx 이고 내용이 같다 (보호가 걸려 있어도 읽힌다)
  3. 그 백업으로 되돌리면(replace) 데이터가 정확히 복원된다  ← 기능 핵심
  4. 되돌린 뒤에도 라이브 보호가 걸려 있다(사용자가 직접 못 고치는 상태 유지)
  5. 일반 저장(다운로드용)은 예전처럼 보호 없이 나온다
"""
import io
import shutil
import sys
import tempfile
import time
from pathlib import Path

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")
ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT))

import openpyxl
import pythoncom; pythoncom.CoInitialize()
import serve_b2b as S

fails = 0


def check(name, cond, detail=""):
    global fails
    print(("  PASS  " if cond else "  FAIL  ") + name + (("  -> " + str(detail)[:200]) if (not cond and detail) else ""))
    if not cond:
        fails += 1


work = Path(tempfile.mkdtemp(prefix="b2b_snapint_"))
SRC = work / "샘플.xlsx"
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "데이터"
ws["A1"], ws["B1"] = "항목", "금액"
for i, (a, b) in enumerate([("가", 100), ("나", 200), ("다", 300)], start=2):
    ws[f"A{i}"], ws[f"B{i}"] = a, b
ws["B5"] = "=SUM(B2:B4)"
wb.save(str(SRC))

S.WORKBOOKS["wbT"] = {"id": "wbT", "name": SRC.name, "path": str(SRC)}
excel_id = None
try:
    opened = S.open_excel_session(str(SRC), name=SRC.name, workbook_id="wbT",
                                  live_editable=True, defer_visible=True)
    excel_id = opened["excelId"]

    def read_back(result_id):
        p = Path(S.RESULTS[result_id]["path"])
        b = openpyxl.load_workbook(str(p), data_only=False)
        sh = b["데이터"]
        vals = [[sh.cell(r, c).value for c in (1, 2)] for r in range(1, 6)]
        prot = bool(sh.protection.sheet)
        b.close()
        return vals, prot, p

    print("[1] 속도 — 되돌리기용 백업이 일반 저장보다 빠른가")
    # 예열(첫 호출은 COM 초기화가 섞인다)
    S.save_excel_session(excel_id, internal=True)
    t0 = time.perf_counter(); normal = S.save_excel_session(excel_id); t_normal = (time.perf_counter() - t0) * 1000
    t0 = time.perf_counter(); intern = S.save_excel_session(excel_id, internal=True); t_int = (time.perf_counter() - t0) * 1000
    print(f"      일반 저장(다운로드용) {t_normal:7.0f}ms   /   되돌리기용 백업 {t_int:7.0f}ms")
    check("되돌리기용이 더 빠름", t_int < t_normal, f"{t_int:.0f}ms vs {t_normal:.0f}ms")

    print("[2] 백업 파일이 정상이고 내용이 같은가")
    v_int, prot_int, p_int = read_back(intern["downloadId"])
    v_norm, prot_norm, _ = read_back(normal["downloadId"])
    check("백업 파일이 열린다", p_int.exists() and p_int.stat().st_size > 0)
    check("내용이 일반 저장본과 동일", v_int == v_norm, f"{v_int} vs {v_norm}")
    check("수식도 보존", v_int[4][1] == "=SUM(B2:B4)", v_int[4][1])
    print(f"      (참고: 일반 저장 보호={prot_norm} / 되돌리기용 보호={prot_int})")
    check("일반 저장본은 보호가 풀려 있다(다운로드용 계약 유지)", prot_norm is False, prot_norm)

    print("[3] 되돌리기 — 백업으로 복원하면 데이터가 정확한가  ← 기능 핵심")
    # 백업 후 라이브를 바꾼다 → 되돌리면 원래대로 와야 한다
    def _mutate(eid):
        s = S.EXCEL_SESSIONS[eid]
        w = s["workbook"]
        try:
            S._protect_workbook_for_read_only_mirror(w, False)
        except Exception:
            pass
        w.Worksheets("데이터").Range("B2").Value = 999
        return True

    S.excel_call(_mutate, excel_id, timeout=120)

    def _read_live(eid, addr):
        s = S.EXCEL_SESSIONS[eid]
        return s["workbook"].Worksheets("데이터").Range(addr).Value

    check("바꾼 값이 라이브에 반영됨", S.excel_call(_read_live, excel_id, "B2", timeout=120) == 999)

    # 서버 HTTP 핸들러와 같은 방식으로 resultId → 실제 경로를 먼저 구한다(handle_excel_replace:2004).
    snap_path = S.ensure_result_file(intern["downloadId"])
    check("백업 파일을 결과 저장소에서 찾을 수 있다", snap_path is not None, intern["downloadId"])
    S.replace_excel_session_workbook(excel_id, snap_path, name=snap_path.name,
                                     result_id=intern["downloadId"], read_only_mirror=False)
    restored = S.excel_call(_read_live, excel_id, "B2", timeout=120)
    check("되돌린 뒤 원래 값(100)으로 복원", restored == 100, restored)
    check("다른 값도 그대로", S.excel_call(_read_live, excel_id, "B4", timeout=120) == 300)

    print("[4] AI 도움 검증(verify-step)이 이 백업으로 코드를 실행할 수 있는가  ← 회귀 지점")
    # 라이브는 UserInterfaceOnly 보호라 COM 쓰기가 되지만, 그 설정은 '파일에 저장되지 않는다'.
    # 따라서 보호가 걸린 채 저장된 백업을 다시 열면 COM 쓰기까지 막힌다 → 검증이 통째로 실패한다.
    verify_code = 'def transform(ctx):\n    ctx.write("데이터", "D1", [["검증"]])\n'
    vr = S.verify_step_isolated(intern["downloadId"], verify_code, sheet="데이터")
    check("보호된 백업에서도 검증 코드가 실행됨", bool(vr and vr.get("ok")),
          (vr or {}).get("error") or vr)

    print("[5] 되돌린 뒤에도 라이브 보호가 걸려 있는가")
    def _is_protected(eid):
        s = S.EXCEL_SESSIONS[eid]
        return bool(s["workbook"].Worksheets("데이터").ProtectContents)
    check("시트 보호 유지(사용자 직접 편집 차단)", S.excel_call(_is_protected, excel_id, timeout=120) is True)

finally:
    try:
        if excel_id:
            S.close_excel_session(excel_id)
    except Exception:
        pass
    try:
        S.cleanup_excel_sessions()
    except Exception:
        pass
    shutil.rmtree(work, ignore_errors=True)

print()
print("RESULT: ALL PASS" if fails == 0 else f"RESULT: {fails} FAIL")
sys.exit(1 if fails else 0)
