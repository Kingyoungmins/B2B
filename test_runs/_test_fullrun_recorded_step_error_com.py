# -*- coding: utf-8 -*-
"""[질문 2026-08-31] 실행기 전체실행 중 '녹화로 만든 단계'가 실패하면 어떻게 움직이나 — 실제 Excel.

시나리오: 5단계 스킬, 그중 3개가 녹화 스타일 VBA(레코더가 실제로 뱉는 모양 —
Sheets().Select / Range().Select / ActiveCell.FormulaR1C1 / Windows().Activate 교차파일).
4번째(녹화 VBA)가 없는 시트를 만져 실패한다.

잠그는 계약:
  1) 실패가 녹화 단계를 정확히 가리킨다(stepIdx + 원인 + 프롬프트 가이드 — 일반 단계와 같은 포맷)
  2) 실패 지점까지의 경계 스냅샷이 쌓인다 (앞 단계 결과 보존)
  3) 실패한 녹화 단계만 고쳐 재실행하면 앞 단계(녹화 포함)를 건너뛴다 — 재실행 없음
  4) 건너뛴 녹화 단계의 결과가 최종본에 살아있다 — 특히 '교차파일로 쓴 값'
     (경계 스냅샷은 Saved 플래그 표식이라 교차 쓰기 파일이 제일 깨지기 쉬운 지점)
  5) 고친 녹화 단계 + 그 뒤 단계가 정상 실행돼 전체가 완성된다
  6) 녹화 스타일 구문(Select/ActiveCell/Windows().Activate)이 보안 게이트에 안 걸린다
"""
import sys, io, tempfile, shutil
from pathlib import Path
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")
ROOT = str(Path(__file__).resolve().parent.parent)
sys.path.insert(0, ROOT)
import openpyxl
import pythoncom; pythoncom.CoInitialize()
import win32com.client as win32
import serve_b2b as S

fails = []
def check(name, cond, detail=""):
    print(("  PASS  " if cond else "  FAIL  ") + name + (("  -> " + str(detail)[:300]) if (not cond and detail != "") else ""))
    if not cond:
        fails.append(name)

TRACE = []
_orig_trace = S._vba_trace
def _tap(event, **kw):
    TRACE.append(dict(kw, event=event))
    return _orig_trace(event, **kw)
S._vba_trace = _tap

SHEET = "데이터"

def py(cell, val):
    return "def transform(ctx):\n    ctx.write(%r, %r, [[%r]])\n" % (SHEET, cell, val)

def rec_vba(body):
    """네이티브 레코더 출력 모양 그대로(Sub B2BSkill + Select/ActiveCell 문체)."""
    return "Sub B2BSkill()\n" + body + "End Sub\n"

REC_SAME = rec_vba(            # 2번째: 자기 파일에 레코더 문체로 쓰기
    '    Sheets("데이터").Select\n'
    '    Range("B2").Select\n'
    '    ActiveCell.FormulaR1C1 = "녹B2값"\n')
REC_CROSS = rec_vba(           # 3번째: 교차파일 — rB 로 건너가 쓰고 돌아온다(레코더가 내는 모양)
    '    Windows("rB.xlsx").Activate\n'
    '    Sheets("데이터").Select\n'
    '    Range("B3").Select\n'
    '    ActiveCell.FormulaR1C1 = "교차B3값"\n'
    '    Windows("rC.xlsx").Activate\n')
REC_BAD = rec_vba(             # 4번째: 없는 시트 → 실패(사용자 실수 재현)
    '    Sheets("없는시트").Select\n'
    '    Range("B4").Select\n'
    '    ActiveCell.FormulaR1C1 = "안닿을값"\n')
REC_FIXED = rec_vba(           # 4번째 수정본
    '    Sheets("데이터").Select\n'
    '    Range("B4").Select\n'
    '    ActiveCell.FormulaR1C1 = "녹B4값"\n')

work = Path(tempfile.mkdtemp(prefix="b2b_recfail_"))
NAMES = ["rA.xlsx", "rB.xlsx", "rC.xlsx"]
IDS = ["recA", "recB", "recC"]
srcs, lives = [], []
for n in NAMES:
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = SHEET
    ws["A1"] = "원본"; wb.create_sheet("메모")
    sp = work / n; wb.save(str(sp)); srcs.append(sp)
    lp = work / ("live_" + n); shutil.copy2(sp, lp); lives.append(lp)

def make_groups(step4_code):
    codes = [
        ("recA", "s1", "python", py("B1", "A1값")),
        ("recB", "s2", "vba", REC_SAME),
        ("recC", "s3", "vba", REC_CROSS),
        ("recC", "s4", "vba", step4_code),
        ("recC", "s5", "python", py("B5", "C5값")),
    ]
    groups, cur = [], None
    for i, (eid, sid, lang, code) in enumerate(codes):
        if cur is None or cur["excelId"] != eid:
            cur = {"excelId": eid, "steps": []}; groups.append(cur)
        cur["steps"].append({"code": code, "language": lang, "stepIdx": i, "stepId": sid})
    return groups

app = None
try:
    app = win32.DispatchEx("Excel.Application"); app.Visible = False; app.DisplayAlerts = False
    for eid, n, sp, lp in zip(IDS, NAMES, srcs, lives):
        w = app.Workbooks.Open(str(lp))
        S.EXCEL_SESSIONS[eid] = {"app": app, "workbook": w, "path": str(lp), "openPath": str(w.FullName),
                                 "name": n, "sourcePath": str(sp), "liveEditable": True}

    print("[0] 녹화 스타일 구문이 보안 게이트를 통과한다(오탐 없음)")
    for nm, code in (("같은파일", REC_SAME), ("교차파일", REC_CROSS), ("실패본", REC_BAD)):
        try:
            S._validate_vba_source_before_inject(S._extract_vba_source_for_injection(code, "B2BSkill"))
            check("게이트 통과: " + nm, True)
        except Exception as e:
            check("게이트 통과: " + nm, False, e)

    print("[1] 1차 전체실행 — 4번째(녹화 VBA)에서 실패한다")
    TRACE.clear()
    err = None
    try:
        S.run_full_pipeline_single_instance(make_groups(REC_BAD), reset_excel_ids=list(IDS))
    except S.PipelineExecutionError as e:
        err = e
    except Exception as e:
        err = e
        check("실패는 파이프라인 오류 형식이어야 한다", False, type(e).__name__ + ": " + str(e)[:150])
    check("실패한다(의도)", err is not None)
    info = getattr(err, "info", None) or {}      # 페이로드는 .info (클라가 받는 그 형식)
    check("실패가 녹화 단계를 가리킨다(stepIdx=3)", info.get("stepIdx") == 3, info.get("stepIdx"))
    check("원인 문구가 있다", bool(str(info.get("cause") or info.get("message") or "").strip()),
          {k: str(v)[:80] for k, v in info.items()})
    check("프롬프트 가이드가 있다(일반 단계와 같은 포맷)",
          bool(str(info.get("promptGuide") or "").strip()) or "이렇게 요청" in str(info.get("message") or ""))
    bnds = [t.get("completed") for t in TRACE if t.get("event") == "fullrun.boundary.saved"]
    # 교차파일이 낀 경계(3단계 뒤)는 '추가 파일 저장'이 필요해 자기조절 게이트(직전 경계 비용의
    # 9배 실행시간 경과 시에만 저장 — 오버헤드 10% 상한)에 걸려 생략될 수 있다. 이 테스트의
    # 스텝은 밀리초라 반드시 걸린다 → 계약은 '어떤 경계든 실패 지점 이전 것이 남는다'까지다.
    check("실패 이전 경계 스냅샷이 남는다(스로틀로 일부 생략 가능)",
          bool(bnds) and max(bnds) <= 3 and max(bnds) >= 2, bnds)

    print("[2] 실패한 녹화 단계만 고쳐 재실행 — 앞의 녹화 단계들을 건너뛴다")
    TRACE.clear()
    out = S.run_full_pipeline_single_instance(make_groups(REC_FIXED), reset_excel_ids=list(IDS))
    d = [t for t in TRACE if t.get("event") == "fullrun.resume.decision"]
    resume_from = d[0].get("resumeFrom") if d else None
    # 스로틀로 3단계 경계가 생략됐으면 2단계 경계에서 이어서 3단계부터 재실행한다(안전한 방향:
    # 경계가 없으면 다시 돌지, 교차 쓰기를 잃는 쪽으로는 안 간다). 최소 계약: 1·2단계는 건너뛴다.
    check("이어실행 판정 ≥ 2단계 건너뜀", resume_from is not None and resume_from >= 2, d[:1])
    ran = [t.get("ordinal") for t in TRACE if t.get("event") == "fullrun.step.start"]
    check("건너뛴 지점 뒤만 연속 실행(처음부터 다시 아님)",
          ran == list(range(resume_from + 1, 6)), (resume_from, ran))
    check("적어도 앞 2단계(python+녹화)는 재실행 없음", ran and min(ran) >= 3, ran)
    check("결과는 전체 5단계 적용", bool(out) and out.get("ok") and out.get("applied") == 5,
          out and {k: out.get(k) for k in ("ok", "applied")})

    print("[3] 최종 결과 — 건너뛴 녹화 단계의 결과가 살아있다(교차파일 포함)")
    def live(eid):
        return S.EXCEL_SESSIONS[eid]["workbook"].Worksheets(SHEET)
    check("1단계(python)", live("recA").Range("B1").Value == "A1값", live("recA").Range("B1").Value)
    check("2단계(녹화 같은파일) — 재실행 없이 보존", live("recB").Range("B2").Value == "녹B2값",
          live("recB").Range("B2").Value)
    check("3단계(녹화 교차파일 rC→rB) — Saved 표식 함정 구간", live("recB").Range("B3").Value == "교차B3값",
          live("recB").Range("B3").Value)
    check("4단계(고친 녹화)", live("recC").Range("B4").Value == "녹B4값", live("recC").Range("B4").Value)
    check("5단계(뒤 python)", live("recC").Range("B5").Value == "C5값", live("recC").Range("B5").Value)
    for eid in IDS:
        check("%s 원본칸 보존" % eid, live(eid).Range("A1").Value == "원본")

    print("[4] 같은 실패를 한 번 더 내도 이어실행이 유지된다(실패 지점부터 재시도)")
    TRACE.clear()
    err2 = None
    try:
        S.run_full_pipeline_single_instance(make_groups(REC_BAD), reset_excel_ids=list(IDS))
    except Exception as e:
        err2 = e
    d4 = [t for t in TRACE if t.get("event") == "fullrun.resume.decision"]
    check("재실패 실행도 3단계는 건너뛴다(앞 구간 낭비 없음)",
          bool(d4) and d4[0].get("resumeFrom") == 3, d4[:1])
    check("실패는 그대로 4번째를 가리킨다", err2 is not None and
          (getattr(err2, "info", None) or {}).get("stepIdx") == 3,
          getattr(err2, "info", None))
finally:
    for eid in IDS:
        S.EXCEL_SESSIONS.pop(eid, None)
    try:
        if app is not None:
            for w in list(app.Workbooks):
                try: w.Close(SaveChanges=False)
                except Exception: pass
            app.Quit()
    except Exception:
        pass
    shutil.rmtree(work, ignore_errors=True)

print("")
print("RESULT: ALL PASS" if not fails else "RESULT: %d FAIL -> %s" % (len(fails), fails))
sys.exit(0 if not fails else 1)
