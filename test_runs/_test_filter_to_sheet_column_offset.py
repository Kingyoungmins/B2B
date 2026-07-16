# -*- coding: utf-8 -*-
# [회귀] filter_to_sheet 열 인덱스 어긋남(빈 A열) 수정 검증.
#  - 증상: "원본" 시트가 A열이 비어 B열부터 시작하면 COM UsedRange 가 B부터 시작 → predicate 의 절대
#    열 인덱스(예: E열=r[4])가 한 칸 밀려(F를 봄) 0건 매칭 실패.
#  - 수정: filter_to_sheet 가 선두 빈 열 수만큼 None 패딩 → A열=index0 절대 기준. openpyxl 은 이미 A1
#    기준(무영향)이라, 두 엔진 동작이 일치하는지 함께 본다.
# 검증 대상:
#  (1) LIVE COM  PythonComSkillContext.filter_to_sheet — 실제 실패 경로
#  (2) openpyxl  OpenpyxlSkillContext.filter_to_sheet  — 이미 A1기준(회귀 없음 확인)
import sys, io, tempfile, shutil
from pathlib import Path
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")
sys.path.insert(0, r"C:\Users\Admin\Desktop\KGM_git\B2B_ver0.6.1")
import openpyxl
import serve_b2b as S

KEEP = "전국대표번호 서비스"
fails = 0
def check(name, ok):
    global fails
    print((" OK  " if ok else "FAIL ") + name)
    if not ok: fails += 1

def build_empty_a(path):
    """A열 비움, 데이터는 B부터. E열(절대 5번째)=요금제, 매칭 2건."""
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "원본"
    # B1..F1 헤더 (A1 은 비움)
    ws["B1"] = "고객"; ws["C1"] = "번호"; ws["D1"] = "구분"; ws["E1"] = "요금제"; ws["F1"] = "금액"
    rows = [
        ("홍길동", "010", "개인", KEEP,        1000),
        ("김철수", "011", "개인", "기타요금제", 2000),
        ("이영희", "012", "법인", KEEP,        3000),
    ]
    for i, (b, c, d, e, f) in enumerate(rows, start=2):
        ws.cell(i, 2, b); ws.cell(i, 3, c); ws.cell(i, 4, d); ws.cell(i, 5, e); ws.cell(i, 6, f)
    wb.save(str(path))

def build_normal_a(path):
    """A열부터 데이터. E열(절대 5번째)=요금제, 매칭 2건 — lead=0(회귀 가드)."""
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "원본"
    ws["A1"] = "고객"; ws["B1"] = "번호"; ws["C1"] = "구분"; ws["D1"] = "지역"; ws["E1"] = "요금제"; ws["F1"] = "금액"
    rows = [
        ("홍길동", "010", "개인", "서울", KEEP,        1000),
        ("김철수", "011", "개인", "부산", "기타요금제", 2000),
        ("이영희", "012", "법인", "대구", KEEP,        3000),
    ]
    for i, (a, b, c, d, e, f) in enumerate(rows, start=2):
        ws.cell(i, 1, a); ws.cell(i, 2, b); ws.cell(i, 3, c); ws.cell(i, 4, d); ws.cell(i, 5, e); ws.cell(i, 6, f)
    wb.save(str(path))

# E열 = 절대 0-based index 4. (수정 전 빈A에서는 r[4]가 F를 가리켜 매칭 0건)
def pred(r):
    return (len(r) > 4 and r[4] is not None and S_normalize(r[4]) == S_normalize(KEEP))

# normalize 는 ctx 별로 있지만, 테스트 predicate 는 단순 비교로 충분(값이 동일 문자열).
def S_normalize(v):
    return str(v).strip() if v is not None else ""

work = Path(tempfile.mkdtemp(prefix="b2b_filtcol_"))

# ───────────────────────── (1) LIVE COM ─────────────────────────
import pythoncom; pythoncom.CoInitialize()
import win32com.client as win32
app = None
try:
    app = win32.DispatchEx("Excel.Application"); app.Visible = False; app.DisplayAlerts = False

    # 1a) 빈 A열 — 수정 전이라면 여기서 "조건 만족 행 없음" 예외가 났다.
    p1 = work / "empty_a.xlsx"; build_empty_a(p1)
    wb1 = app.Workbooks.Open(str(p1))
    ctx1 = S.PythonComSkillContext(app, wb1, {})
    ok_match = False; ok_pos = False; ncount = -1
    try:
        ctx1.filter_to_sheet("원본", lambda r: pred(r), "전국대표번호_필터")
        dst = wb1.Worksheets("전국대표번호_필터")
        # 매칭 데이터 행 수(헤더 1 제외) — E열(절대5)에 KEEP 가 들어가야 한다(열 위치 보존).
        last = dst.Cells(dst.Rows.Count, 5).End(-4162).Row  # xlUp=-4162, E열 기준
        vals = [dst.Cells(r, 5).Value for r in range(2, last + 1)]
        ncount = sum(1 for v in vals if S_normalize(v) == S_normalize(KEEP))
        ok_match = (ncount == 2)
        ok_pos = (S_normalize(dst.Cells(1, 5).Value) == "요금제")  # 헤더 E1 = 요금제 (열 위치 보존)
    except Exception as e:
        print("   (COM 빈A 예외:", repr(e)[:140], ")")
    finally:
        try: wb1.Close(SaveChanges=False)
        except Exception: pass
    check(f"COM 빈A: E열 절대인덱스로 2건 매칭(ncount={ncount})", ok_match)
    check("COM 빈A: 결과 E열 위치 보존(헤더=요금제)", ok_pos)

    # 1b) 정상 A열 — lead=0, 기존 동작 유지(회귀 가드).
    p2 = work / "normal_a.xlsx"; build_normal_a(p2)
    wb2 = app.Workbooks.Open(str(p2))
    ctx2 = S.PythonComSkillContext(app, wb2, {})
    ok_n = False; ncount2 = -1
    try:
        ctx2.filter_to_sheet("원본", lambda r: pred(r), "전국대표번호_필터")
        dst2 = wb2.Worksheets("전국대표번호_필터")
        last2 = dst2.Cells(dst2.Rows.Count, 5).End(-4162).Row
        vals2 = [dst2.Cells(r, 5).Value for r in range(2, last2 + 1)]
        ncount2 = sum(1 for v in vals2 if S_normalize(v) == S_normalize(KEEP))
        ok_n = (ncount2 == 2)
    except Exception as e:
        print("   (COM 정상A 예외:", repr(e)[:140], ")")
    finally:
        try: wb2.Close(SaveChanges=False)
        except Exception: pass
    check(f"COM 정상A: 회귀 없음 2건 매칭(ncount={ncount2})", ok_n)
finally:
    try:
        if app is not None: app.Quit()
    except Exception: pass

# ───────────────────────── (2) openpyxl ─────────────────────────
# 이미 UsedRange 가 A1 기준이라 빈 A 에서도 r[4]=E 로 매칭돼야 한다(엔진 간 동작 일치).
ok_opx = False; opx_n = -1
try:
    wbx = openpyxl.Workbook(); wsx = wbx.active; wsx.title = "원본"
    wsx["B1"] = "고객"; wsx["E1"] = "요금제"
    wsx.cell(2, 2, "홍길동"); wsx.cell(2, 5, KEEP)
    wsx.cell(3, 2, "김철수"); wsx.cell(3, 5, "기타요금제")
    wsx.cell(4, 2, "이영희"); wsx.cell(4, 5, KEEP)
    ctxx = S.OpenpyxlSkillContext(output_wb=wbx, input_wbs={})
    ctxx.filter_to_sheet("원본", lambda r: pred(r), "전국대표번호_필터")
    rws = ctxx.rows("전국대표번호_필터")
    opx_n = sum(1 for r in rws[1:] if len(r) > 4 and S_normalize(r[4]) == S_normalize(KEEP))
    ok_opx = (opx_n == 2)
except Exception as e:
    print("   (openpyxl 예외:", repr(e)[:160], ")")
check(f"openpyxl 빈A: 이미 A1기준 → 2건 매칭(opx_n={opx_n})", ok_opx)

shutil.rmtree(work, ignore_errors=True)
print(f"\n=== RESULT: {5 - fails}/5 PASS ===")
sys.exit(1 if fails else 0)
