# -*- coding: utf-8 -*-
# [라이브] run_full_pipeline_single_instance(output_mode="file") — 실행기 파일출력 모드.
#  - 2파일에 VBA 스텝 적용 → 라이브는 '안 건드리고'(무손상) output 폴더에 결과 파일 저장 + outputFiles 반환.
#  - 검증: output/결과_*.xlsx 2개 생성(결과 시트 포함) + 라이브 워크북엔 결과 시트 없음(미반영) + outputFiles 2개.
import sys, io, tempfile, shutil
from pathlib import Path
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")
sys.path.insert(0, r"C:\Users\Admin\Desktop\KGM_git\B2B_ver0.5.15")
import openpyxl
import pythoncom; pythoncom.CoInitialize()
import win32com.client as win32
import serve_b2b as S

NA, NB = "main.xlsx", "other.xlsx"
def vba(wbname, sheet_new, val):
    return ('Sub B2BSkill()\n  Dim wb As Workbook\n  Set wb = Workbooks("' + wbname + '")\n'
            '  Dim ws As Worksheet\n  Set ws = wb.Worksheets.Add(After:=wb.Worksheets(wb.Worksheets.Count))\n'
            '  ws.Name = "' + sheet_new + '"\n  ws.Range("A1").Value = "' + val + '"\nEnd Sub')

work = Path(tempfile.mkdtemp(prefix="b2b_fileout_"))
outdir = work / "out"
S.default_output_dir = (lambda _o=outdir: _o)   # 출력 폴더를 테스트 임시폴더로
def build(p, sheet, a1):
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = sheet; ws["A1"] = a1; wb.save(str(p))
srcA, srcB = work/NA, work/NB
build(srcA, "원본", "A데이터"); build(srcB, "데이터", "B데이터")
liveA, liveB = work/("live_"+NA), work/("live_"+NB)
shutil.copy2(srcA, liveA); shutil.copy2(srcB, liveB)

app = None; res = {}
try:
    app = win32.DispatchEx("Excel.Application"); app.Visible = False; app.DisplayAlerts = False
    wA = app.Workbooks.Open(str(liveA)); wB = app.Workbooks.Open(str(liveB))
    S.EXCEL_SESSIONS["exA"] = {"app": app, "workbook": wA, "path": str(liveA), "openPath": str(wA.FullName),
                              "name": NA, "sourcePath": str(srcA), "liveEditable": True}
    S.EXCEL_SESSIONS["exB"] = {"app": app, "workbook": wB, "path": str(liveB), "openPath": str(wB.FullName),
                              "name": NB, "sourcePath": str(srcB), "liveEditable": True}
    groups = [
        {"excelId": "exA", "steps": [{"code": vba(NA, "A결과", "fromA"), "language": "vba", "stepIdx": 0, "stepId": "a1"}]},
        {"excelId": "exB", "steps": [{"code": vba(NB, "B결과", "fromB"), "language": "vba", "stepIdx": 1, "stepId": "b1"}]},
    ]
    out = S.run_full_pipeline_single_instance(groups, reset_excel_ids=["exA", "exB"], output_mode="file")
    ofs = out.get("outputFiles") or []
    print("결과:", {k: out.get(k) for k in ("ok", "applied")}, "| outputFiles:", len(ofs))
    for o in ofs: print("  -", o.get("name"), "| dlId:", bool(o.get("downloadId")))

    # 라이브는 미반영(결과 시트 없어야)
    lA = S.EXCEL_SESSIONS["exA"]["workbook"]; lB = S.EXCEL_SESSIONS["exB"]["workbook"]
    shA = [w.Name for w in lA.Worksheets]; shB = [w.Name for w in lB.Worksheets]
    print("라이브 A 시트:", shA, "| B 시트:", shB)
    res["applied2"] = (out.get("applied") == 2)
    res["outputFiles2"] = (len(ofs) == 2 and all(o.get("downloadId") for o in ofs))
    res["live_A_미반영"] = ("A결과" not in shA and "원본" in shA)
    res["live_B_미반영"] = ("B결과" not in shB and "데이터" in shB)
    # output 파일이 실제로 디스크에 있고 결과 시트 포함
    files = sorted(outdir.glob("결과_*.xlsx")) if outdir.exists() else []
    res["output_2개"] = (len(files) == 2)
    sheets_in_outputs = set()
    for f in files:
        for s in openpyxl.load_workbook(str(f)).sheetnames: sheets_in_outputs.add(s)
    print("output 파일:", [f.name for f in files], "| 시트합집합:", sorted(sheets_in_outputs))
    res["결과시트_포함"] = ("A결과" in sheets_in_outputs and "B결과" in sheets_in_outputs)
    res["RESULTS_등록"] = all((o.get("downloadId") in S.RESULTS) for o in ofs)
    for sid in ("exA", "exB"):
        try: del S.EXCEL_SESSIONS[sid]
        except Exception: pass
finally:
    try:
        if app is not None: app.Quit()
    except Exception: pass
    shutil.rmtree(work, ignore_errors=True)

checks = [
    ("applied==2", res.get("applied2")),
    ("outputFiles 2개(downloadId 포함)", res.get("outputFiles2")),
    ("라이브 A 미반영(원본만)", res.get("live_A_미반영")),
    ("라이브 B 미반영(데이터만)", res.get("live_B_미반영")),
    ("output 폴더에 결과 2파일", res.get("output_2개")),
    ("결과 시트(A결과/B결과) 파일에 포함", res.get("결과시트_포함")),
    ("RESULTS 에 다운로드 등록", res.get("RESULTS_등록")),
]
fails = 0
print("\n=== 판정 ===")
for n, ok in checks:
    print((" OK  " if ok else "FAIL ") + n)
    if not ok: fails += 1
print(f"\n=== RESULT: {len(checks)-fails}/{len(checks)} PASS ===")
sys.exit(1 if fails else 0)
