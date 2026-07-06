# 가입번호 블록당 '여러 데이터 행'이 실재하는지(=첫 행만 복사시 손실) + 소스/대상 블록 높이 불일치 확인.
import sys, io, os, glob
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")
from openpyxl import load_workbook

DOWN = r"C:\Users\Admin\Downloads"
def find(pat):
    return [h for h in glob.glob(os.path.join(DOWN, pat)) if not os.path.basename(h).startswith("~$")]
SRC = find("531611708899*로우데이터*260616.xlsx")[0]
DST = find("LGU*농협생명*콜센터*260617.xlsx")[0]

def cols_CtoN(ws, r):
    return [ws.cell(r, c).value for c in range(3, 15)]

wsS = load_workbook(SRC, data_only=True)["sheet"]
wsD = load_workbook(DST, data_only=True)["콜센터"]

print("=== SRC 500106220276 블록(B43:B44) C:N 두 행 ===")
for r in (43, 44):
    print(f"  행{r}:", cols_CtoN(wsS, r))

print("\n=== SRC 500011679426 블록(B3:B13, 11행) C:N ===")
for r in range(3, 14):
    print(f"  행{r}:", cols_CtoN(wsS, r))

# SRC C:N 병합 여부(같은 가입번호 안에서 C~N도 행별로 병합돼 있나)
mergesS = list(wsS.merged_cells.ranges)
def merged_at(ws_ranges, r, c):
    return any(m.min_row <= r <= m.max_row and m.min_col <= c <= m.max_col for m in ws_ranges)
print("\n  SRC C43,C44 병합?:", merged_at(mergesS, 43, 3), merged_at(mergesS, 44, 3))
print("  SRC C3 병합영역:", [str(m) for m in mergesS if m.min_row<=3<=m.max_row and m.min_col<=3<=m.max_col][:1])

print("\n=== DST 500106220276 블록(B4:B5) C:N 두 행 ===")
for r in (4, 5):
    print(f"  행{r}:", cols_CtoN(wsD, r))

# 대상에 채워질 가입번호들이 소스에서 각각 몇 행 블록인지 (높이 불일치 규모)
print("\n=== 대상 가입번호 -> 소스 블록 높이 (앞 15개) ===")
# 소스 B열 가입번호 -> 블록높이 매핑
srcblk = {}
for m in mergesS:
    if m.min_col <= 2 <= m.max_col:
        v = wsS.cell(m.min_row, 2).value
        if v is not None:
            srcblk[str(v)] = m.max_row - m.min_row + 1
dstheights = []
for m in sorted([m for m in wsD.merged_cells.ranges if m.min_col<=2<=m.max_col], key=lambda x:x.min_row):
    v = wsD.cell(m.min_row, 2).value
    if v is None: continue
    dh = m.max_row - m.min_row + 1
    sh = srcblk.get(str(v), "없음")
    dstheights.append((str(v), dh, sh))
for v, dh, sh in dstheights[:15]:
    flag = "" if sh == dh else "  <-- 높이 불일치!" if sh != "없음" else "  <-- 소스에 없음"
    print(f"  {v}: 대상 {dh}행 / 소스 {sh}행{flag}")
mismatch = sum(1 for _,dh,sh in dstheights if sh!="없음" and sh!=dh)
missing = sum(1 for _,_,sh in dstheights if sh=="없음")
print(f"\n  대상 가입번호 {len(dstheights)}개 중 소스와 높이 다른 것 {mismatch}개, 소스에 없는 것 {missing}개")
