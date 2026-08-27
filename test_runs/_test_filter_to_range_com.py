# -*- coding: utf-8 -*-
"""[문의 2026-08-27] '필터 → 해당 행 전체 복사 → 기존 양식에 붙여넣기'를 행 수가 변해도 되게.

문의 내용
  복붙 캡처로 필터+복붙을 녹화하는 건 좋은데, 원본에서 가입번호가 늘어나면 캡처된 범위가
  안 맞아 활용성이 떨어진다. '필터를 건다 → 해당 전체 데이터를 복사한다 → 붙여넣는다' 를
  코드로 표현할 수 있나?  (현재는 캡처 시점의 필터 결과 범위로만 고정)

확인한 사실
  · 복붙 캡처는 ctx.paste_copied('시트','A2:N57',...) 처럼 그 순간 좌표를 코드에 박는다 → 지적이 맞다.
  · 기존 ctx.filter_to_sheet 는 범위를 굳히지 않아 행이 늘어도 따라오지만 '새 시트'에만 만든다.
  · 그래서 '이미 있는 양식의 정한 자리'에 붙이는 ctx.filter_to_range 를 새로 만들었다.

이 테스트가 지키는 것
  1) 행이 늘어도 같은 코드가 그대로 동작한다(범위 고정 없음)
  2) 열 위치가 밀려도 동작한다(열은 이름으로 찾는다)
  3) 양식의 다른 열·머리글을 건드리지 않는다
  4) clear_existing 으로 지난달 값이 남지 않는다
  5) 서식이 보존된다
  6) **기존 filter_to_sheet 동작이 하나도 안 변했다**(예전 스킬 호환)
"""
import sys, io, tempfile, shutil
from pathlib import Path
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")
ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT))
import openpyxl
from openpyxl.styles import PatternFill
import pythoncom; pythoncom.CoInitialize()
import win32com.client as win32
import serve_b2b as S

fails = []


def check(name, cond, detail=""):
    print(("  PASS  " if cond else "  FAIL  ") + name + (("  -> " + str(detail)[:250]) if not cond else ""))
    if not cond:
        fails.append(name)


YELLOW = PatternFill("solid", fgColor="FFFF00")


def build(path, n_rows, lead_col=False):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "회선 현황"
    head = (["메모"] if lead_col else []) + ["가입번호", "상태", "요금"]
    ws.append(head)
    for i in range(n_rows):
        row = (["x"] if lead_col else []) + [512102400000 + i, "정지" if i % 2 == 0 else "정상", 1000 + i]
        ws.append(row)
        if i % 2 == 0:                                  # 정지 행에 색 — 서식 보존 확인용
            ws.cell(row=ws.max_row, column=(2 if lead_col else 1)).fill = YELLOW
    f = wb.create_sheet("대상양식")
    f["A1"] = "정지회선 목록"
    f["A2"] = "가입번호"; f["B2"] = "상태"; f["C2"] = "요금"
    f["E2"] = "비고(양식 고유)"                          # 붙여넣기가 침범하면 안 되는 열
    f["E3"] = "건드리지 말 것"
    wb.save(str(path))


def open_ctx(app, path):
    book = app.Workbooks.Open(str(path))
    return book, S.PythonComSkillContext(app, book, {"name": Path(path).name}, timeout_s=300)


def scenario(app, n_rows, lead_col, clear_existing, preset_old=False):
    work = Path(tempfile.mkdtemp(prefix="b2b_ftr_"))
    src = work / "회선현황.xlsx"
    build(src, n_rows, lead_col)
    book, ctx = open_ctx(app, src)
    try:
        if preset_old:                                  # 지난달 값이 남아 있는 상태를 흉내
            ctx.write("대상양식", "A3", [["옛값1", "옛", 1], ["옛값2", "옛", 2], ["옛값3", "옛", 3]])
            ctx = S.PythonComSkillContext(app, book, {"name": "회선현황.xlsx"}, timeout_s=300)
        col = ctx.find_header("회선 현황", "상태")        # 열은 이름으로 — 위치가 밀려도 견딘다
        n = ctx.filter_to_range("회선 현황", ctx.column_is(col, ["정지"]),
                                "대상양식", "A3", clear_existing=clear_existing)
        ctx2 = S.PythonComSkillContext(app, book, {"name": "회선현황.xlsx"}, timeout_s=300)
        vals = [r[0] for r in ctx2.read("대상양식", "A3:A%d" % (2 + max(n, 3) + 2))]
        keep = ctx2.read_cell("대상양식", "E3")
        head_ok = ctx2.read_cell("대상양식", "A2") == "가입번호"
        # 붙여넣기는 원본 표의 열을 '순서대로' 옮긴다 — 앞에 열이 하나 더 있으면 색도 한 칸 밀린다.
        # 그래서 특정 칸을 찍지 말고 붙은 첫 행에서 색이 있는 칸을 찾는다(테스트가 배치에 안 흔들리게).
        color = None
        try:
            sheet_o = book.Worksheets("대상양식")
            colors = [sheet_o.Cells(3, c).Interior.Color for c in range(1, 5)]
            color = next((c for c in colors if c not in (None, 16777215.0, 16777215)), None)
        except Exception:
            pass
        return {"n": n, "vals": vals, "keep": keep, "head_ok": head_ok, "color": color}
    finally:
        try: book.Close(SaveChanges=False)
        except Exception: pass
        shutil.rmtree(work, ignore_errors=True)


app = None
try:
    app = win32.DispatchEx("Excel.Application")
    app.Visible = False
    app.DisplayAlerts = False

    print("[1] 행이 늘어도 같은 코드가 따라온다 (문의의 핵심)")
    a = scenario(app, 10, False, False)
    b = scenario(app, 30, False, False)
    check("10행 → 정지 5건", a["n"] == 5, a["n"])
    check("30행 → 정지 15건 (코드 동일)", b["n"] == 15, b["n"])

    print("[2] 열 위치가 밀려도 동작한다")
    c = scenario(app, 30, True, False)
    check("앞에 열이 하나 생겨도 15건", c["n"] == 15, c["n"])

    print("[3] 양식을 침범하지 않는다")
    check("양식 고유 열(E)이 그대로", c["keep"] == "건드리지 말 것", c["keep"])
    check("양식 머리글(A2)이 그대로", c["head_ok"] is True)
    check("머리글은 다시 붙이지 않는다(기본값)",
          a["vals"] and str(a["vals"][0]) not in ("가입번호",), a["vals"][:2])

    print("[4] 서식이 보존된다")
    check("정지 행의 노란 채우기가 따라온다(붙은 첫 행 어딘가)", c["color"] is not None, c["color"])

    print("[5] 지난달 값이 남지 않는다(clear_existing)")
    d = scenario(app, 10, False, True, preset_old=True)
    check("옛 값이 지워지고 새 값만", not any("옛값" in str(v) for v in d["vals"] if v is not None), d["vals"])
    e = scenario(app, 10, False, False, preset_old=True)
    check("끄면 종전대로 덮어쓰기만 한다(옛 꼬리가 남을 수 있음)", e["n"] == 5, e["n"])

    print("[6] 기존 filter_to_sheet 는 하나도 안 변했다 (예전 스킬 호환)")
    work = Path(tempfile.mkdtemp(prefix="b2b_compat_"))
    srcp = work / "회선현황.xlsx"
    build(srcp, 10, False)
    book, ctx = open_ctx(app, srcp)
    try:
        col = ctx.find_header("회선 현황", "상태")
        out = ctx.filter_to_sheet("회선 현황", ctx.column_is(col, ["정지"]), "정지목록")
        ctx2 = S.PythonComSkillContext(app, book, {"name": "회선현황.xlsx"}, timeout_s=300)
        check("새 시트를 만든다", out == "정지목록" and "정지목록" in ctx2.sheets(), ctx2.sheets())
        check("머리글 + 5건", ctx2.used_last_row("정지목록") == 6, ctx2.used_last_row("정지목록"))
        check("첫 줄은 머리글", ctx2.read_cell("정지목록", "A1") == "가입번호", ctx2.read_cell("정지목록", "A1"))
        # 람다 조건(예전 스킬이 가장 많이 쓰던 형태)도 그대로
        ctx3 = S.PythonComSkillContext(app, book, {"name": "회선현황.xlsx"}, timeout_s=300)
        out2 = ctx3.filter_to_sheet("회선 현황", lambda r: str(r[1]) == "정상", "정상목록")
        ctx4 = S.PythonComSkillContext(app, book, {"name": "회선현황.xlsx"}, timeout_s=300)
        check("람다 조건도 종전대로", out2 == "정상목록" and ctx4.used_last_row("정상목록") == 6,
              ctx4.used_last_row("정상목록"))
    finally:
        try: book.Close(SaveChanges=False)
        except Exception: pass
        shutil.rmtree(work, ignore_errors=True)

    print("[7] 조건에 맞는 행이 0건이어도 죽지 않는다")
    work2 = Path(tempfile.mkdtemp(prefix="b2b_zero_"))
    sp2 = work2 / "z.xlsx"
    build(sp2, 10, False)
    book, ctx = open_ctx(app, sp2)
    try:
        col = ctx.find_header("회선 현황", "상태")
        n0 = ctx.filter_to_range("회선 현황", ctx.column_is(col, ["없는값"]), "대상양식", "A3")
        check("0건이면 0을 돌려주고 오류를 내지 않는다", n0 == 0, n0)
    except Exception as err:
        check("0건이면 0을 돌려주고 오류를 내지 않는다", False, err)
    finally:
        try: book.Close(SaveChanges=False)
        except Exception: pass
        shutil.rmtree(work2, ignore_errors=True)
finally:
    try:
        if app is not None:
            for w in list(app.Workbooks):
                try: w.Close(SaveChanges=False)
                except Exception: pass
            app.Quit()
    except Exception:
        pass

print("")
print("RESULT: ALL PASS" if not fails else "RESULT: %d FAIL -> %s" % (len(fails), fails))
sys.exit(0 if not fails else 1)
