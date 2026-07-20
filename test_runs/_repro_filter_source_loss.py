# -*- coding: utf-8 -*-
# [Bug1 실측 재현] "필터 VBA 스킬 적용 시 원본 시트가 사라지고 필터값만 남음".
# 실제 백엔드 경로(_run_vba_pipeline_on_session_impl, reset=True, 격리실행→SaveCopyAs→
# _copy_source_workbook_into_target)를 그대로 태워 라이브 워크북에 '원본' 시트가 보존되는지 관찰한다.
# 합성 '원본' 시트(A열 비움, 2행 헤더, B:I 데이터, E=요금제) + 사용자 VBA(전국대표번호 필터).
import sys, io, tempfile, shutil
from pathlib import Path
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")
sys.path.insert(0, r"C:\Users\Admin\Desktop\KGM_git\B2B_ver0.6.2")
import openpyxl
import pythoncom; pythoncom.CoInitialize()
import win32com.client as win32
import serve_b2b as S

WBNAME = "가입자별청구내역_20260615_v2_DSMC_260625.xlsx"
VBA = '''Sub B2BSkill()
    Dim wb As Workbook
    Dim wsSrc As Worksheet
    Dim wsDst As Worksheet
    Dim filterRange As Range
    Dim lastRow As Long
    Dim newSheetName As String
    Dim prevCalc As XlCalculation
    Dim raisedNum As Long, raisedSrc As String, raisedDesc As String
    prevCalc = Application.Calculation
    On Error GoTo Cleanup
    Application.ScreenUpdating = False
    Application.Calculation = xlCalculationManual
    Set wb = Workbooks("''' + WBNAME + '''")
    Set wsSrc = wb.Worksheets("원본")
    lastRow = wsSrc.Cells(wsSrc.Rows.Count, "B").End(xlUp).Row
    If lastRow < 2 Then
        Err.Raise vbObjectError + 513, "B2BSkill", "데이터가 없습니다."
    End If
    Set filterRange = wsSrc.Range("B2:I" & lastRow)
    wsSrc.AutoFilterMode = False
    If wsSrc.FilterMode Then wsSrc.ShowAllData
    filterRange.AutoFilter Field:=4, Criteria1:="전국대표번호"
    newSheetName = "전국대표번호_필터"
    Dim sh As Worksheet
    For Each sh In wb.Worksheets
        If sh.Name = newSheetName Then
            sh.Delete
            Exit For
        End If
    Next sh
    Set wsDst = wb.Worksheets.Add(After:=wb.Worksheets(wb.Worksheets.Count))
    wsDst.Name = newSheetName
    filterRange.SpecialCells(xlCellTypeVisible).Copy Destination:=wsDst.Range("A1")
    wsSrc.AutoFilterMode = False
Cleanup:
    Application.Calculation = prevCalc
    Application.ScreenUpdating = True
    Application.CutCopyMode = False
    If Err.Number <> 0 Then
        raisedNum = Err.Number
        raisedSrc = Err.Source
        raisedDesc = Err.Description
        Err.Clear
        Err.Raise raisedNum, raisedSrc, raisedDesc
    End If
End Sub'''

def build_source(path):
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "원본"
    ws["B1"] = "월 청구 내역"                       # 1행 제목
    # 2행 헤더 (B2:I2) — E2(=B:I 4번째)=요금제
    hdr = {"B2":"고객","C2":"번호","D2":"구분","E2":"요금제","F2":"금액","G2":"비고","H2":"x","I2":"y"}
    for k, v in hdr.items(): ws[k] = v
    data = [
        ("홍길동","010","개인","전국대표번호 서비스",1000,"a","x1","y1"),
        ("김철수","011","개인","기타요금제",        2000,"b","x2","y2"),
        ("이영희","012","법인","전국대표번호 서비스",3000,"c","x3","y3"),
        ("박민수","013","법인","데이터쉐어링",      4000,"d","x4","y4"),
    ]
    r = 3
    for b,c,d,e,f,g,h,i in data:
        ws.cell(r,2,b); ws.cell(r,3,c); ws.cell(r,4,d); ws.cell(r,5,e); ws.cell(r,6,f); ws.cell(r,7,g); ws.cell(r,8,h); ws.cell(r,9,i)
        r += 1
    wb.save(str(path))

work = Path(tempfile.mkdtemp(prefix="b2b_reprofilt_"))
src = work / WBNAME            # pristine source (sourcePath)
build_source(src)
livep = work / ("live_" + WBNAME)
shutil.copy2(src, livep)       # 라이브 작업복사본

app = None
res = {}
try:
    app = win32.DispatchEx("Excel.Application"); app.Visible = False; app.DisplayAlerts = False
    live = app.Workbooks.Open(str(livep))
    excel_id = "reprofilt"
    S.EXCEL_SESSIONS[excel_id] = {
        "app": app, "workbook": live, "path": str(livep), "openPath": str(live.FullName),
        "name": WBNAME, "sourcePath": str(src), "liveEditable": True,
    }
    print("적용 전 라이브 시트:", [w.Name for w in live.Worksheets])
    print("적용 전 원본 B3:", live.Worksheets("원본").Cells(3,2).Value, "| E3:", live.Worksheets("원본").Cells(3,5).Value)

    step = {"code": VBA, "language": "vba", "stepIdx": 0, "stepId": "s1", "description": "전국대표번호 필터"}
    try:
        out = S._run_vba_pipeline_on_session_impl(excel_id, [step], reset=True)
        print("적용 결과:", {k: out.get(k) for k in ("ok","applied")})
    except Exception as e:
        print("!! 적용 예외:", repr(e)[:300])

    live2 = S.EXCEL_SESSIONS[excel_id]["workbook"]
    sheets_after = [w.Name for w in live2.Worksheets]
    print("적용 후 라이브 시트:", sheets_after)
    res["원본_present"] = "원본" in sheets_after
    res["필터_present"] = "전국대표번호_필터" in sheets_after
    if res["원본_present"]:
        w0 = live2.Worksheets("원본")
        b3 = w0.Cells(3,2).Value; e3 = w0.Cells(3,5).Value
        lastr = w0.Cells(w0.Rows.Count, 2).End(-4162).Row
        print("적용 후 원본 B3:", b3, "| E3:", e3, "| 원본 마지막행(B):", lastr)
        res["원본_데이터유지"] = (b3 == "홍길동" and int(lastr) >= 6)  # 헤더2 + 데이터4 = 6행
    else:
        res["원본_데이터유지"] = False
    try: del S.EXCEL_SESSIONS[excel_id]
    except Exception: pass
finally:
    try:
        if app is not None: app.Quit()
    except Exception: pass
    shutil.rmtree(work, ignore_errors=True)

print("\n=== 판정 ===")
print("  원본 시트 보존:", "OK" if res.get("원본_present") else "❌ 사라짐(버그 재현!)")
print("  원본 데이터 유지:", "OK" if res.get("원본_데이터유지") else "❌ 손실")
print("  필터 시트 생성:", "OK" if res.get("필터_present") else "(없음)")
reproduced = not (res.get("원본_present") and res.get("원본_데이터유지"))
print("\n버그 재현됨:", "YES (원본 손실)" if reproduced else "NO (원본 보존됨 — 합성케이스로는 재현 안 됨)")
sys.exit(0)
