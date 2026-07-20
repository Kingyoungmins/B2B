# SO사업자별요금 헤더/병합 구조 확인 — SBAGENT-19(국제 헤더 위치), 30(P열 병합/다중토큰), 35(D/G/M) 근거.
import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

PATH = r"C:\Users\Admin\Downloads\KT\output_HCN대사용_영총.사지.LG유플러스 정산내역_2026년03월_LG작성.xlsx"
wb = load_workbook(PATH, data_only=True)
print("시트:", wb.sheetnames)
ws = wb["SO사업자별요금"]
print(f"dims={ws.dimensions} max_row={ws.max_row} max_col={ws.max_column}")

# '국제'가 어느 행/열에 있나 (SBAGENT-19: 코드가 header_row=8 로 find_header('국제') 실패)
print("\n=== '국제' 텍스트 위치 검색 (행1~12) ===")
for r in range(1, 13):
    for c in range(1, ws.max_column + 1):
        v = ws.cell(r, c).value
        if v is not None and "국제" in str(v):
            print(f"  {get_column_letter(c)}{r} = {v!r}")

print("\n=== 헤더 후보 행 5~9, A~P 열 값 ===")
for r in range(5, 10):
    vals = [ws.cell(r, c).value for c in range(1, 17)]
    print(f"  행{r}: " + " | ".join(f"{get_column_letter(i+1)}:{('' if v is None else str(v))[:10]}" for i, v in enumerate(vals)))

# 병합 영역 중 헤더 영역(행<=9)만
hmerges = [m for m in ws.merged_cells.ranges if m.min_row <= 9]
print(f"\n=== 헤더영역(행<=9) 병합 {len(hmerges)}개 (앞 20) ===")
for m in sorted(hmerges, key=lambda x:(x.min_row, x.min_col))[:20]:
    print(f"  {m.coord} = {ws.cell(m.min_row, m.min_col).value!r}")

# P열(16) 병합 + 다중토큰 (SBAGENT-30)
print("\n=== P열(16) 값/병합 샘플 (행 9~30) ===")
pmerges = [m for m in ws.merged_cells.ranges if m.min_col <= 16 <= m.max_col]
def merged_at(r, c):
    return next((m for m in ws.merged_cells.ranges if m.min_row<=r<=m.max_row and m.min_col<=c<=m.max_col), None)
for r in range(9, 31):
    v = ws.cell(r, 16).value
    m = merged_at(r, 16)
    mflag = f" [merge {m.coord}]" if m else ""
    multi = " <다중토큰>" if v is not None and (("\n" in str(v)) or (" " in str(v).strip() and len(str(v).strip())>13)) else ""
    if v is not None or m:
        print(f"  P{r} = {str(v)[:40]!r}{mflag}{multi}")
