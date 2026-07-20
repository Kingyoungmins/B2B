# 실제 HCN 요약 시트에 '합계/총계' 본문 행이 실재하는지 확인 + 클라 감지 로직(_clarifyAoaHasTotalRow) 동치 재현.
import sys, io, re
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")
from openpyxl import load_workbook

PATH = r"C:\Users\Admin\Downloads\KT\output_HCN대사용_영총.사지.LG유플러스 정산내역_2026년03월_LG작성.xlsx"
wb = load_workbook(PATH, data_only=True)
ws = wb["요약"]
print("dims:", ws.dimensions, "max_row:", ws.max_row, "max_col:", ws.max_column)

# 클라 AOA(0-based) 재현: values_only
aoa = [list(r) for r in ws.iter_rows(values_only=True)]

RE = re.compile(r"^(합\s*계|총\s*계|총합계|총\s*합|소\s*계|누\s*계|계|total|sum)$", re.I)
def row_has_number(row):
    for v in row:
        if isinstance(v, (int, float)):
            return True
        if isinstance(v, str):
            n = re.sub(r"[,\s₩원%()]", "", v)
            if n and n != "-":
                try:
                    float(n); return True
                except: pass
    return False

hits = []
for r in range(4, len(aoa)):        # 클라 로직과 동일: r>=4 (1~4행 헤더 제외, 0-based)
    row = aoa[r] or []
    label = None
    for c in range(min(len(row), 3)):
        t = ("" if row[c] is None else str(row[c])).strip()
        if t and RE.match(t):
            label = t; break
    if label and row_has_number(row):
        hits.append((r+1, label, [row[i] for i in range(min(len(row),7))]))

print("\n=== 감지된 총계 행 (엑셀 행번호, 라벨, 앞 7열) ===")
for h in hits:
    print(h)

# F열(6번째, index 5) 미리보기 20~30행
print("\n=== A열/F열 20~30행 ===")
for r in range(19, min(30, len(aoa))):
    row = aoa[r] or []
    a = row[0] if len(row) > 0 else None
    f = row[5] if len(row) > 5 else None
    print(f"  행{r+1}: A={a!r}  F={f!r}")

print("\n_clarifyAoaHasTotalRow ->", bool(hits))

# 실제 AOA 를 JSON 으로 덤프 → node 가 buildSheetStructureDigest 로 재검증.
import json, tempfile, os
def norm(v):
    if isinstance(v, (int, float, str)) or v is None:
        return v
    return str(v)
dump = [[norm(v) for v in (row or [])] for row in aoa]
out = os.path.join(tempfile.gettempdir(), "b2b_summary_aoa.json")
with open(out, "w", encoding="utf-8") as f:
    json.dump(dump, f, ensure_ascii=False)
print("AOA JSON ->", out)
