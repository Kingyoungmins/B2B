# -*- coding: utf-8 -*-
"""[실측/COM] 리셋이 Office 저장 없이도 '재오픈하면 원본' 보장을 지키는가.

배경 (VM 실측 2026-08-12)
  리셋은 워크북을 메모리에서 원본으로 되돌린 뒤 곧바로 wb.Save() 를 했다. 그런데 Save() 는
  Office 의 정식 저장이라 사내 보안 부가기능(MIP)이 끼어들어 기본 라벨을 붙인다 —
  이 저장 직후부터 작업 파일이 암호화됐고(srcLabel none → encrypted), 이후 모든 사본이
  암호화를 물려받아 되돌리기가 느려지고 스킬이 꼬였다. 저장 자체도 리셋 1회당 8~19초.

  → 디스크에 지금 쓰지 않고 표시(diskPristineFrom)만 남기고, 정말로 디스크에서 다시 열어야
    할 때(=COM 참조 사망) 파이썬 파일 복사로 원본을 만들어 준다. 파이썬 복사엔 MIP 가 개입 안 한다.

이 테스트가 잠그는 것
  1. 리셋 후 메모리는 원본 상태다(기존 계약)
  2. 리셋이 작업 파일을 Office 로 저장하지 않는다(파일 수정시각이 그대로)
  3. COM 참조가 죽어 디스크에서 다시 열면 그때 원본이 복구된다  ← SBAGENT-138 이 지키던 것
  4. 복구는 한 번만(표시가 소진된다) — 이후 실행 결과를 원본으로 덮지 않는다
  5. 다운로드용 저장(SaveAs)으로 경로가 바뀌면 표시가 해제된다(결과가 날아가지 않는다)

실행: python test_runs/_test_reset_no_office_save_com.py   (Excel 필요)
"""
import io
import shutil
import sys
import tempfile
import time
from pathlib import Path

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")
ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT))

import openpyxl  # noqa: E402
import pythoncom  # noqa: E402
pythoncom.CoInitialize()
import serve_b2b as S  # noqa: E402

fails = 0


def check(name, cond, detail=""):
    global fails
    print(("  PASS  " if cond else "  FAIL  ") + name + (("  -> " + str(detail)[:200]) if (not cond and detail) else ""))
    if not cond:
        fails += 1


work = Path(tempfile.mkdtemp(prefix="b2b_resetsave_"))
SRC = work / "원본.xlsx"
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "데이터"
ws["A1"] = "원본값"
wb.save(str(SRC))

WORK_COPY = work / "작업본.xlsx"
shutil.copy2(SRC, WORK_COPY)

excel_id = None
try:
    S.WORKBOOKS["wbR"] = {"id": "wbR", "name": SRC.name, "path": str(SRC)}
    opened = S.open_excel_session(str(WORK_COPY), name=SRC.name, workbook_id="wbR",
                                  live_editable=True, defer_visible=True)
    excel_id = opened["excelId"]
    session = S.EXCEL_SESSIONS[excel_id]
    session["sourcePath"] = str(SRC)

    app, live = S.session_workbook(session)
    live.Worksheets("데이터").Range("A1").Value = "스킬이 바꾼값"
    print("[1] 리셋 — 메모리는 원본으로")
    before_mtime = Path(session["path"]).stat().st_mtime
    time.sleep(1.1)   # 수정시각 비교를 위해 1초 이상 벌린다
    S.run_vba_pipeline_on_session(excel_id, [], reset=True)
    app, live = S.session_workbook(session)
    check("메모리 값이 원본으로 복구", live.Worksheets("데이터").Range("A1").Value == "원본값",
          live.Worksheets("데이터").Range("A1").Value)

    print("[2] Office 저장을 하지 않았다  ← 이번 수정")
    after_mtime = Path(session["path"]).stat().st_mtime
    check("작업 파일 수정시각 그대로", after_mtime == before_mtime, f"{before_mtime} → {after_mtime}")
    check("재오픈하면 원본이라는 표시가 남음", bool(session.get("diskPristineFrom")), session.get("diskPristineFrom"))

    print("[3] COM 이 죽어 디스크에서 다시 열면 그때 원본이 복구된다")
    # 디스크의 작업본은 아직 '스킬이 바꾼값' 상태다(저장을 안 했으니 원본 그대로일 수도 있어 명시적으로 오염시킨다)
    live.Close(SaveChanges=False)
    session["workbook"] = None
    dirty = openpyxl.load_workbook(str(session["path"]))
    dirty["데이터"]["A1"] = "디스크에 남은 옛 결과"
    dirty.save(str(session["path"]))
    dirty.close()
    check("디스크가 오염된 상태로 준비됨",
          openpyxl.load_workbook(str(session["path"]))["데이터"]["A1"].value == "디스크에 남은 옛 결과")
    app2, wb2 = S.session_workbook(session)          # 여기서 파이썬 복사로 복구돼야 한다
    check("디스크에서 다시 열었는데 원본", wb2.Worksheets("데이터").Range("A1").Value == "원본값",
          wb2.Worksheets("데이터").Range("A1").Value)
    check("표시는 소진됨(한 번만)", not session.get("diskPristineFrom"), session.get("diskPristineFrom"))

    print("[4] 표시가 소진된 뒤엔 결과를 원본으로 덮지 않는다")
    wb2.Worksheets("데이터").Range("A1").Value = "새 결과"
    wb2.Save()                                        # 이번엔 일부러 저장(사용자 다운로드 등 상황)
    wb2.Close(SaveChanges=False)
    session["workbook"] = None
    app3, wb3 = S.session_workbook(session)
    check("다시 열어도 새 결과가 유지", wb3.Worksheets("데이터").Range("A1").Value == "새 결과",
          wb3.Worksheets("데이터").Range("A1").Value)

    print("[5] 경로가 바뀌는 저장은 표시를 해제한다")
    session["diskPristineFrom"] = str(SRC)
    saved = S.save_excel_session(excel_id, name="다운로드본.xlsx")
    check("SaveAs 후 표시 해제", not session.get("diskPristineFrom"), session.get("diskPristineFrom"))
    check("저장 결과 파일 존재", Path(S.RESULTS[saved["downloadId"]]["path"]).exists())
finally:
    try:
        if excel_id:
            S.close_excel_session(excel_id)
    except Exception:
        pass
    pythoncom.CoUninitialize()

print("")
print("RESULT: ALL PASS" if fails == 0 else "RESULT: %d FAIL" % fails)
sys.exit(0 if fails == 0 else 1)
