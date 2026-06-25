#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""복붙 캡처(0.5.9) 검증용 테스트 엑셀 생성.
- 시트 '원본': 값 + 수식(D=단가*수량, 합계=SUM) 표 → 복사 시 수식 보존 확인용.
- 시트 '대상': 비어 있음(붙여넣기 목적지).
- 시트 '원본2': 교차 시트 복붙용 보조 데이터.
복붙 캡처가 '값/수식/서식'을 그대로 재현하는지(특히 수식이 상대참조로 올바르게 따라가는지) 확인한다."""
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

OUT = Path(__file__).resolve().parent / "v059_복붙캡처.xlsx"

wb = openpyxl.Workbook()

# ---- 시트 '원본' : 값 + 수식 표 ----
ws = wb.active
ws.title = "원본"
hdr_fill = PatternFill("solid", fgColor="D9E1F2")
bold = Font(bold=True)
headers = ["품목", "단가", "수량", "금액", "비고"]
for c, h in enumerate(headers, start=1):
    cell = ws.cell(row=1, column=c, value=h)
    cell.font = bold
    cell.fill = hdr_fill
    cell.alignment = Alignment(horizontal="center")
rows = [
    ("사과", 1000, 5, "과일"),
    ("배", 1500, 3, "과일"),
    ("우유", 2000, 10, "음료"),
    ("빵", 3000, 2, "베이커리"),
]
r = 2
for name, price, qty, memo in rows:
    ws.cell(row=r, column=1, value=name)
    ws.cell(row=r, column=2, value=price)
    ws.cell(row=r, column=3, value=qty)
    ws.cell(row=r, column=4, value=f"=B{r}*C{r}")   # 금액 = 단가*수량 (수식)
    ws.cell(row=r, column=5, value=memo)
    r += 1
# 합계 행(수식)
ws.cell(row=r, column=1, value="합계").font = bold
ws.cell(row=r, column=3, value=f"=SUM(C2:C{r-1})")
ws.cell(row=r, column=4, value=f"=SUM(D2:D{r-1})")
for c in (1, 3, 4):
    ws.cell(row=r, column=c).font = bold
for col, w in (("A", 10), ("B", 10), ("C", 8), ("D", 12), ("E", 12)):
    ws.column_dimensions[col].width = w

# ---- 시트 '대상' : 비어 있음 ----
wsd = wb.create_sheet("대상")
wsd["A1"] = "← 여기(A1)에 원본 A1:E6 을 붙여넣어 테스트"
wsd["A1"].font = Font(italic=True, color="888888")

# ---- 시트 '원본2' : 교차 시트 복붙용 보조 ----
ws2 = wb.create_sheet("원본2")
for c, h in enumerate(["지점", "매출"], start=1):
    ws2.cell(row=1, column=c, value=h).font = bold
for i, (br, sale) in enumerate([("서울", 120), ("부산", 85), ("대구", 60)], start=2):
    ws2.cell(row=i, column=1, value=br)
    ws2.cell(row=i, column=2, value=sale)
ws2.cell(row=5, column=1, value="합계").font = bold
ws2.cell(row=5, column=2, value="=SUM(B2:B4)")

wb.save(str(OUT))
print("created:", OUT)
