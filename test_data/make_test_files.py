"""
ver2.0 기능 테스트용 xlsx 3개 생성.

- input_매출_2026_4월.xlsx
    "매출"     : 회사명/상품/건수/금액, 1200 rows  → 가상 스크롤(item 6)
    "고객정보"  : 회사명/사업자번호/담당자/연락처    → 다중 시트(item 4)
    "월별집계"  : 같은 시트에 표 3개 (상/중/하반기) → 표 다중 인식(item 5)

- input_원가_2026_4월.xlsx
    "원가"     : 회사명/상품/단가/원가     ← 회사명 컬럼이 매출 파일과 겹침 (item 2)

- output_청구서_템플릿.xlsx
    "회사별요약" : 회사명/매출/원가/마진/마진율
                  마진=B-C, 마진율=IFERROR(D/B,0), 마지막 행 SUM/AVERAGE  (item 10)

실행:
  python make_test_files.py
"""
import os
import random
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

random.seed(42)

OUT_DIR = os.path.dirname(os.path.abspath(__file__))

COMPANIES = [
    "ABC통신", "글로벌네트워크", "스마트링크", "코리아텔레콤", "디지털파트너스",
    "프라임모바일", "위즈컴", "미래통신", "네오링크", "테크커넥트",
    "오리진네트", "라이브셀", "퀀텀모바일", "벤티지통신", "솔라테크",
    "유니콘텔", "아크네트", "비전모바일", "포커스링크", "메이저텔레콤",
]
PRODUCTS = ["B2B 데이터(5G)", "B2B 음성", "전용회선", "IoT", "기업메시지", "관제서비스"]

HEADER_FILL = PatternFill("solid", fgColor="2D2240")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
HEADER_ALIGN = Alignment(horizontal="center", vertical="center")
TITLE_FONT = Font(bold=True, size=13, color="2D2240")
SUM_FILL = PatternFill("solid", fgColor="FFE0F2")
SUM_FONT = Font(bold=True, color="7928CA")
THIN = Side(border_style="thin", color="C7C7D9")
BORDER = Border(top=THIN, bottom=THIN, left=THIN, right=THIN)


def style_header(ws, row, cols):
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = HEADER_ALIGN
        cell.border = BORDER


def style_header_range(ws, row, c1, c2):
    for c in range(c1, c2 + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = HEADER_ALIGN
        cell.border = BORDER


def style_data_borders(ws, r1, r2, c1, c2):
    for r in range(r1, r2 + 1):
        for c in range(c1, c2 + 1):
            ws.cell(row=r, column=c).border = BORDER


def autosize(ws, max_w=28):
    for col in ws.columns:
        col_letter = get_column_letter(col[0].column)
        m = 0
        for cell in col:
            v = cell.value
            if v is None:
                continue
            m = max(m, len(str(v)))
        ws.column_dimensions[col_letter].width = min(m + 2, max_w)


def write_row(ws, row, values):
    for col, value in enumerate(values, start=1):
        ws.cell(row=row, column=col, value=value)


def populate_regression_source(ws, title):
    """20개 회귀 테스트에서 공통으로 쓰는 넓은 원본 데이터."""
    ws.cell(row=1, column=1, value=title).font = TITLE_FONT
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=24)

    headers = [
        "삭제테스트A", "삭제테스트B", "조건값", "회사명", "상품", "상태",
        "중복개수", "금액", "수량", "청구항목", "단가", "합계수식",
        "부서", "담당자", "메모", "삭제대상", "복사대상", "우선순위",
        "보조일자", "효력발생인자", "계약상태", "지역", "정렬검증ID", "당월매출",
    ]
    write_row(ws, 3, headers)
    style_header(ws, 3, len(headers))

    bill_items = [
        "기본료", "전국대표 포함한 기본료", "080포함한 기본료", "월구전화 기본료",
        "데이터 기본료", "기본료 할인", "부가서비스", "기본료",
    ]
    amounts = [1000, 1000, 2000, 3000, 3000, 3000, 4500, 4500, 7000, 8200, 8200, 9100]
    effect_keys = [
        "2026-04-20", "2026-01-15", "2026-03-05", "2026-02-10",
        "2026-06-01", "2026-05-17", "2026-01-05", "2026-04-01",
        "2026-03-25", "2026-02-28", "2026-05-01", "2026-06-15",
    ]
    monthly_sales = [900000, 120000, 810000, 300000, 760000, 450000, 980000, 210000, 660000, 530000, 150000, 880000]

    for i in range(12):
        r = 4 + i
        qty = 2 + (i % 5)
        unit = [10000, 12000, 15000, 9000, 8000, 7000][i % 6]
        condition = 123 if i in (1, 4, 7, 10) else [111, 222, 456, 789][i % 4]
        row = [
            f"A삭제{i + 1}", f"B삭제{i + 1}", condition, COMPANIES[i], PRODUCTS[i % len(PRODUCTS)],
            "정상" if i % 3 else "검토", None, amounts[i], qty, bill_items[i % len(bill_items)],
            unit, f"=I{r}*K{r}", f"팀{(i % 4) + 1}", f"담당{i + 1}",
            "행 단위 정렬 검증용", f"삭제값{i + 1}", None,
            ["상", "중", "하", "긴급"][i % 4], f"2026-04-{(i % 20) + 1:02d}",
            effect_keys[i], "유지" if i % 2 else "신규", ["서울", "부산", "대전"][i % 3],
            f"ROW-{i + 1:02d}", monthly_sales[i],
        ]
        write_row(ws, r, row)

    last = 15
    style_data_borders(ws, 3, last, 1, len(headers))
    ws.freeze_panes = "A4"
    autosize(ws, max_w=24)


def make_regression_sheets(wb):
    # 사용자가 예시 프롬프트에서 자주 쓰는 "a시트"를 실제로 제공한다.
    ws_a = wb.create_sheet("a")
    populate_regression_source(ws_a, "■ 회귀 테스트 원본 데이터 (a시트)")

    ws_src = wb.create_sheet("회귀_원본데이터")
    populate_regression_source(ws_src, "■ 회귀 테스트 원본 데이터")

    # 수식 셀 값 복사 / 수식 포함 복사 / 수식 문자열화 방지 테스트.
    ws_formula = wb.create_sheet("회귀_수식복사")
    ws_formula.cell(row=1, column=1, value="■ A1:A5는 수식 셀, F1:F5는 값 복사 대상").font = TITLE_FONT
    for r in range(1, 6):
        ws_formula.cell(row=r, column=2, value=r * 100)
        ws_formula.cell(row=r, column=3, value=r * 7)
        ws_formula.cell(row=r, column=1, value=f"=B{r}+C{r}")
        ws_formula.cell(row=r, column=6, value=None)
    ws_formula.cell(row=8, column=1, value="■ I:K 수식 표").font = TITLE_FONT
    headers = ["기준값", "두배수식", "합계수식"]
    for j, name in enumerate(headers, start=9):
        ws_formula.cell(row=10, column=j, value=name)
    style_header_range(ws_formula, 10, 9, 11)
    for i in range(5):
        r = 11 + i
        ws_formula.cell(row=r, column=9, value=10 + i)
        ws_formula.cell(row=r, column=10, value=f"=I{r}*2")
        ws_formula.cell(row=r, column=11, value=f"=SUM(I{r}:J{r})")
    style_data_borders(ws_formula, 1, 5, 1, 6)
    style_data_borders(ws_formula, 10, 15, 9, 11)
    autosize(ws_formula)

    ws_formula_target = wb.create_sheet("회귀_수식복사_대상")
    ws_formula_target.cell(row=1, column=1, value="■ 수식/서식 복사 대상 시트").font = TITLE_FONT
    for j, name in enumerate(headers, start=9):
        ws_formula_target.cell(row=10, column=j, value=name)
    style_header_range(ws_formula_target, 10, 9, 11)
    style_data_borders(ws_formula_target, 10, 15, 9, 11)
    autosize(ws_formula_target)

    # 병합 셀 복붙 / 서식 유지 테스트.
    ws_merge = wb.create_sheet("회귀_병합복붙")
    ws_merge.cell(row=1, column=1, value="■ 2026년 4월 병합 셀 복사 원본").font = TITLE_FONT
    ws_merge.merge_cells("A1:F1")
    ws_merge.cell(row=2, column=1, value="청구 그룹")
    ws_merge.merge_cells("A2:B2")
    ws_merge.cell(row=2, column=3, value="금액 그룹")
    ws_merge.merge_cells("C2:F2")
    for c in range(1, 7):
        ws_merge.cell(row=2, column=c).fill = SUM_FILL
        ws_merge.cell(row=2, column=c).font = SUM_FONT
        ws_merge.cell(row=2, column=c).alignment = HEADER_ALIGN
        ws_merge.cell(row=2, column=c).border = BORDER
    merge_headers = ["회사명", "항목", "수량", "단가", "공급가", "합계수식"]
    write_row(ws_merge, 3, merge_headers)
    style_header(ws_merge, 3, 6)
    for i in range(4):
        r = 4 + i
        ws_merge.cell(row=r, column=1, value=COMPANIES[i])
        ws_merge.cell(row=r, column=2, value=["기본료", "전국대표 포함한 기본료", "080포함한 기본료", "부가서비스"][i])
        ws_merge.cell(row=r, column=3, value=i + 2)
        ws_merge.cell(row=r, column=4, value=[10000, 12000, 15000, 8000][i])
        ws_merge.cell(row=r, column=5, value=f"=C{r}*D{r}")
        ws_merge.cell(row=r, column=6, value=f"=E{r}*1.1")
    style_data_borders(ws_merge, 1, 7, 1, 6)
    ws_merge.cell(row=1, column=8, value="복붙 대상 영역 H1:M7").font = TITLE_FONT
    autosize(ws_merge)

    # 열/행 삽입, 삭제, 셀 삭제 테스트.
    ws_edit = wb.create_sheet("회귀_삽입삭제")
    ws_edit.cell(row=1, column=1, value="■ 삽입/삭제 테스트 시트").font = TITLE_FONT
    ws_edit.merge_cells("A1:J1")
    edit_headers = ["A값", "B값", "C값", "D값", "E값", "F값", "G값", "H값", "I값", "J열_삽입기준"]
    write_row(ws_edit, 3, edit_headers)
    style_header(ws_edit, 3, len(edit_headers))
    for i in range(8):
        r = 4 + i
        write_row(ws_edit, r, [f"{get_column_letter(c)}{i + 1}" for c in range(1, 11)])
    style_data_borders(ws_edit, 3, 11, 1, 10)
    ws_edit.freeze_panes = "A4"
    autosize(ws_edit)

    # 중복 개수 산출 후 중복 제거 테스트.
    ws_dup = wb.create_sheet("회귀_중복계수")
    ws_dup.cell(row=1, column=1, value="■ H열 금액 중복 개수를 G열에 채우는 테스트").font = TITLE_FONT
    dup_headers = ["ID", "회사명", "비고", "기대값", "공백1", "공백2", "중복개수", "금액"]
    write_row(ws_dup, 3, dup_headers)
    style_header(ws_dup, 3, len(dup_headers))
    dup_amounts = [1000, 1000, 1000, 2000, 2000, 3000, 4000, 4000, 5000, 5000, 5000, 5000]
    expected_counts = {v: dup_amounts.count(v) for v in set(dup_amounts)}
    for i, amount in enumerate(dup_amounts):
        r = 4 + i
        ws_dup.cell(row=r, column=1, value=f"DUP-{i + 1:02d}")
        ws_dup.cell(row=r, column=2, value=COMPANIES[i % len(COMPANIES)])
        ws_dup.cell(row=r, column=4, value=expected_counts[amount])
        ws_dup.cell(row=r, column=7, value=None)
        ws_dup.cell(row=r, column=8, value=amount)
    style_data_borders(ws_dup, 3, 15, 1, 8)
    autosize(ws_dup)

    # 동일 프롬프트 해석 안정성: 정확한 항목만 합산해야 한다.
    ws_item = wb.create_sheet("회귀_청구항목")
    ws_item.cell(row=1, column=1, value="■ 청구항목 합산 테스트").font = TITLE_FONT
    item_headers = ["청구항목", "금액", "포함대상", "비고"]
    write_row(ws_item, 3, item_headers)
    style_header(ws_item, 3, len(item_headers))
    item_rows = [
        ("기본료", 10000, "Y", "포함"),
        ("전국대표 포함한 기본료", 12000, "Y", "포함"),
        ("080포함한 기본료", 15000, "Y", "포함"),
        ("월구전화 기본료", 90000, "N", "제외해야 함"),
        ("기본료 할인", -1000, "N", "제외해야 함"),
        ("데이터 기본료", 3000, "N", "제외해야 함"),
    ]
    for i, row in enumerate(item_rows, start=4):
        write_row(ws_item, i, row)
    ws_item.cell(row=3, column=6, value="기대합계")
    ws_item.cell(row=4, column=6, value=37000)
    ws_item.cell(row=5, column=6, value="월구전화 기본료는 제외")
    style_data_borders(ws_item, 3, 9, 1, 4)
    autosize(ws_item)


# ------------------------------------------------------------
# 1) 매출 파일 — 다중 시트 + 1200행 + 다중 표
# ------------------------------------------------------------
def make_sales():
    wb = Workbook()

    # ---- "매출" 시트 (1200행 — 가상 스크롤 테스트용) ----
    ws1 = wb.active
    ws1.title = "매출"
    headers = ["회사명", "상품", "건수", "금액"]
    ws1.append(headers)
    style_header(ws1, 1, len(headers))
    rows = 1200
    for i in range(rows):
        company = COMPANIES[i % len(COMPANIES)]
        product = PRODUCTS[(i // 7) % len(PRODUCTS)]
        cnt = random.randint(1, 50)
        amt = cnt * random.choice([35000, 50000, 99000, 150000, 250000])
        ws1.append([company, product, cnt, amt])
    style_data_borders(ws1, 2, rows + 1, 1, len(headers))
    ws1.freeze_panes = "A2"
    autosize(ws1)

    # ---- "고객정보" 시트 ----
    ws2 = wb.create_sheet("고객정보")
    h2 = ["회사명", "사업자번호", "담당자", "연락처"]
    ws2.append(h2)
    style_header(ws2, 1, len(h2))
    domains = ["company.kr", "biz.co.kr", "telecom.kr", "net.com"]
    for i, comp in enumerate(COMPANIES):
        biz = f"{random.randint(100,999)}-{random.randint(10,99)}-{random.randint(10000,99999)}"
        manager = random.choice(["김민준", "이지우", "박서연", "최도윤", "정하은", "장유나", "윤시우"])
        phone = f"02-{random.randint(100,999):03d}-{random.randint(1000,9999)}"
        ws2.append([comp, biz, manager, phone])
    style_data_borders(ws2, 2, len(COMPANIES) + 1, 1, len(h2))
    ws2.freeze_panes = "A2"
    autosize(ws2)

    # ---- "월별집계" 시트 — 표 3개 (한 시트 안에 여러 표) ----
    ws3 = wb.create_sheet("월별집계")

    def write_table(ws, start_row, title, months):
        ws.cell(row=start_row, column=1, value=title).font = TITLE_FONT
        # 헤더
        h = ["월", "건수", "금액"]
        for j, name in enumerate(h):
            ws.cell(row=start_row + 1, column=1 + j, value=name)
        style_header(ws, start_row + 1, 3)
        # 데이터
        for i, m in enumerate(months):
            r = start_row + 2 + i
            cnt = random.randint(800, 1500)
            amt = cnt * random.randint(80000, 200000)
            ws.cell(row=r, column=1, value=m)
            ws.cell(row=r, column=2, value=cnt)
            ws.cell(row=r, column=3, value=amt)
        end = start_row + 1 + len(months)
        style_data_borders(ws, start_row + 1, end, 1, 3)
        return end

    end1 = write_table(ws3, 1, "■ 상반기 집계", ["1월", "2월", "3월", "4월", "5월", "6월"])
    end2 = write_table(ws3, end1 + 3, "■ 하반기 집계", ["7월", "8월", "9월", "10월", "11월", "12월"])
    # 분기별 요약 (다른 헤더 — 표 라벨이 다름)
    qstart = end2 + 3
    ws3.cell(row=qstart, column=1, value="■ 분기별 요약").font = TITLE_FONT
    qh = ["분기", "건수", "금액"]
    for j, name in enumerate(qh):
        ws3.cell(row=qstart + 1, column=1 + j, value=name)
    style_header(ws3, qstart + 1, 3)
    for i, q in enumerate(["Q1", "Q2", "Q3", "Q4"]):
        r = qstart + 2 + i
        cnt = random.randint(3000, 5000)
        amt = cnt * random.randint(80000, 200000)
        ws3.cell(row=r, column=1, value=q)
        ws3.cell(row=r, column=2, value=cnt)
        ws3.cell(row=r, column=3, value=amt)
    style_data_borders(ws3, qstart + 1, qstart + 1 + 4, 1, 3)
    autosize(ws3)

    path = os.path.join(OUT_DIR, "input_매출_2026_4월.xlsx")
    wb.save(path)
    return path


# ------------------------------------------------------------
# 2) 원가 파일 — "회사명" 컬럼이 매출 파일과 겹침 (item 2)
# ------------------------------------------------------------
def make_cost():
    wb = Workbook()
    ws = wb.active
    ws.title = "원가"
    headers = ["회사명", "상품", "단가", "원가"]
    ws.append(headers)
    style_header(ws, 1, len(headers))
    for comp in COMPANIES:
        for prod in PRODUCTS:
            unit = random.choice([20000, 35000, 60000, 90000, 140000])
            cost = int(unit * random.uniform(0.55, 0.78))
            ws.append([comp, prod, unit, cost])
    last = ws.max_row
    style_data_borders(ws, 2, last, 1, len(headers))
    ws.freeze_panes = "A2"
    autosize(ws)
    path = os.path.join(OUT_DIR, "input_원가_2026_4월.xlsx")
    wb.save(path)
    return path


# ------------------------------------------------------------
# 3) 청구서 템플릿 — 수식 평가 테스트 (item 10)
# ------------------------------------------------------------
def make_template():
    wb = Workbook()
    ws = wb.active
    ws.title = "회사별요약"

    # 제목
    ws.cell(row=1, column=1, value="■ 2026년 4월 회사별 청구서 요약").font = TITLE_FONT
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=5)

    headers = ["회사명", "매출", "원가", "마진", "마진율"]
    for j, name in enumerate(headers):
        ws.cell(row=3, column=1 + j, value=name)
    style_header(ws, 3, len(headers))

    # 회사별 placeholder 행 (값은 비워두고, 마진/마진율은 수식만 박아둠 — 채팅으로 매출/원가가 채워지면 자동 평가됨)
    n_rows = len(COMPANIES)
    for i, comp in enumerate(COMPANIES):
        r = 4 + i
        ws.cell(row=r, column=1, value=comp)
        # 매출 / 원가는 빈칸 — 사용자가 채팅으로 채울 부분
        ws.cell(row=r, column=2, value=None)
        ws.cell(row=r, column=3, value=None)
        # 마진 = 매출 - 원가
        ws.cell(row=r, column=4, value=f"=B{r}-C{r}")
        # 마진율 = IFERROR(마진/매출, 0)
        ws.cell(row=r, column=5, value=f"=IFERROR(D{r}/B{r},0)")
    last = 3 + n_rows
    style_data_borders(ws, 3, last, 1, len(headers))

    # 합계 행 — SUM/AVERAGE 수식
    sum_row = last + 1
    ws.cell(row=sum_row, column=1, value="합계 / 평균")
    ws.cell(row=sum_row, column=2, value=f"=SUM(B4:B{last})")
    ws.cell(row=sum_row, column=3, value=f"=SUM(C4:C{last})")
    ws.cell(row=sum_row, column=4, value=f"=SUM(D4:D{last})")
    ws.cell(row=sum_row, column=5, value=f"=AVERAGE(E4:E{last})")
    for c in range(1, 6):
        cell = ws.cell(row=sum_row, column=c)
        cell.fill = SUM_FILL
        cell.font = SUM_FONT
        cell.border = BORDER

    ws.freeze_panes = "A4"
    autosize(ws, max_w=22)

    # ---- 빈 시트 한 장 더: "월별실적" — 다중 시트 테스트 ----
    ws2 = wb.create_sheet("월별실적")
    ws2.cell(row=1, column=1, value="■ 월별 누적 실적").font = TITLE_FONT
    h2 = ["월", "건수", "금액", "전월대비"]
    for j, name in enumerate(h2):
        ws2.cell(row=3, column=1 + j, value=name)
    style_header(ws2, 3, len(h2))
    months = ["1월", "2월", "3월", "4월"]
    for i, m in enumerate(months):
        r = 4 + i
        ws2.cell(row=r, column=1, value=m)
        # B,C는 빈칸 (사용자가 채울 부분)
        # D = 이번달 - 이전달; 첫달은 0
        if i == 0:
            ws2.cell(row=r, column=4, value=0)
        else:
            ws2.cell(row=r, column=4, value=f"=C{r}-C{r-1}")
    style_data_borders(ws2, 3, 3 + len(months), 1, len(h2))
    autosize(ws2, max_w=22)

    make_regression_sheets(wb)

    path = os.path.join(OUT_DIR, "output_청구서_템플릿.xlsx")
    wb.save(path)
    return path


def main():
    p1 = make_sales()
    p2 = make_cost()
    p3 = make_template()
    print("생성 완료:")
    for p in (p1, p2, p3):
        sz = os.path.getsize(p)
        print(f"  {os.path.basename(p):<40s} {sz:>9,d} bytes")


if __name__ == "__main__":
    main()
