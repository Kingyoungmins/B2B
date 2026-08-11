#!/usr/bin/env python3
import http.server
import ast
import atexit
import csv
import ctypes
import datetime
import functools
import inspect
import gc
import hashlib
import io
import json
import math
import os
import copy
from pathlib import Path
import queue
import re
import shutil
import socket
import socketserver
import subprocess
import sys
import tempfile
import threading
import time
import zipfile
from urllib.parse import parse_qs, quote, unquote, urlparse
import urllib.error
import urllib.request
import uuid
from functools import total_ordering

try:
    import openpyxl
except Exception:
    openpyxl = None

try:
    import pythoncom
    import win32com.client
    import win32con
    import win32gui
    import win32process
except Exception:
    pythoncom = None
    win32com = None
    win32con = None
    win32gui = None
    win32process = None

try:
    import psutil
except Exception:
    psutil = None


HOST = os.environ.get("B2B_HOST", "127.0.0.1")
PORT = int(os.environ.get("B2B_PORT", "8090"))
VLLM_BASE = os.environ.get(
    "B2B_VLLM_BASE",
    "https://e2e-ns-17786299267796664.mng-1.ip.violet.uplus.co.kr",
).rstrip("/")
PROXY_RETRY_ATTEMPTS = int(os.environ.get("B2B_PROXY_RETRY_ATTEMPTS", "3"))
PROXY_RETRY_BASE_DELAY = float(os.environ.get("B2B_PROXY_RETRY_BASE_DELAY", "0.6"))
BACKEND_DIR = Path(tempfile.gettempdir()) / "b2b_backend_v044"
WORKBOOKS = {}
RESULTS = {}
PIPELINE_PROGRESS = {}  # excelId -> {current, total, ts} : 전체실행 진행률(클라 폴링용, 락 불필요)
DIFFS = {}
PIPELINE_STEP_SNAPSHOTS = {}
# [새로고침 즉시복원 2026-08-04] '스킬을 전부 적용한 뒤의 라이브 상태' 파일 사본.
#   PIPELINE_STEP_SNAPSHOTS 와 다른 점:
#     - 저것은 Python 격리 파이프라인의 '스텝 prefix 이어달리기' 캐시(엔진 내부용, 입력/출력 역할 구분)
#     - 이것은 엔진 무관(Python/VBA 공통) '파일 1개의 최종 상태' — 새로고침 후 재실행 대신 이 파일로 연다
#   키 = sha256(원본파일 지문 + 클라가 준 파이프라인 상태 서명). 값 = {"path", ...}
#   비용: 두 엔진 모두 '이미 디스크에 쓰고 있던 파일'을 지우지 않고 옮겨 담을 뿐이라 추가 COM 저장이 없다.
LIVE_FINAL_SNAPSHOTS = {}
PIPELINE_JOBS = {}
EXCEL_SESSIONS = {}
EXCEL_LOCK = threading.RLock()
EXCEL_QUEUE = None
EXCEL_THREAD = None
LIVE_EXCEL_APP = None  # 라이브 편집 세션들이 공유하는 앱 전용 Excel.Application
LIVE_EXCEL_APP_PID = None  # 위 인스턴스의 프로세스 pid — '죽음' 판정을 COM 예외가 아니라 pid 생존으로 한다


def _note_live_app_reset(reason, **extra):
    """[진단] 공유 라이브 Excel 인스턴스가 리셋/종료되는 순간을 남긴다. 녹화 중(NATIVE_RECORDING.active)
    에 이게 찍히면 = 그 인스턴스가 죽어 녹화가 통째로 유실된 것(실측: 첫 녹화부터 harvested 0).
    무엇이 녹화 중 라이브 앱을 죽였는지 추적하는 결정적 단서."""
    try:
        recording = bool(NATIVE_RECORDING.get("active")) if "NATIVE_RECORDING" in globals() else False
    except Exception:
        recording = False
    try:
        _vba_trace("live_app.reset", reason=reason, duringRecording=recording, **extra)
    except Exception:
        pass
LAST_COPY_SOURCE = {}  # 복사(Ctrl+C) 중 스냅샷한 클립보드 소스 {"source":{book,sheet,range}, "ts":monotonic}
# [최소화 중 미러 유출] 네이티브 호스트가 최소화되어 있는 동안엔 어떤 경로도 미러 창을
# 화면에 띄우면 안 된다(오버레이는 최상위 창이라 호스트가 사라지면 '따로 뜬' 것처럼 보임).
# C#(HandleHostResize)이 /api/excel/host-state 로 갱신하고, 표시 계열 엔드포인트가 이 플래그를
# 존중한다 — hide-all '이후'에 완료되는 열기/위치/복구가 창을 되띄우던 레이스의 단일 차단점.
HOST_MINIMIZED = {"v": False}
                       # — 붙여넣기/탭전환으로 클립보드 Link 가 사라진 뒤 복붙 캡처가 소스를 복구하는 폴백.
COPY_SOURCE_SNAPSHOT_THROTTLE_SECONDS = float(os.environ.get("B2B_COPY_SOURCE_SNAPSHOT_THROTTLE", "5"))
# 단일 Excel 인스턴스(SDI)에서 워크북마다 생기는 최상위 프레임을 "세션별 hwnd"로 직접 제어하는 모드.
# app.Hwnd(=그 순간 활성 프레임 1개) 기반의 기존 동작으로 되돌리려면 B2B_WINMODE=legacy 로 실행.
LIVE_FRAME_MODE = (os.environ.get("B2B_WINMODE") or "frame").strip().lower() != "legacy"
PYTHON_SKILL_APP = None  # 라이브 미러가 없을 때 Python 스킬 실행용으로 재사용하는 숨김 Excel 인스턴스
PYTHON_SKILL_APP_PID = None  # 위 인스턴스의 pid — 강제 정리(force-restart/초기화) 때 COM 없이 종료하기 위해 기록
PYTHON_SKILL_APP_LAST_USED = 0.0
PYTHON_SKILL_APP_IDLE_TTL_SECONDS = float(os.environ.get("B2B_PYTHON_SKILL_APP_IDLE_TTL", "900"))
PYTHON_SKILL_APP_REAP_CHECK_AT = 0.0
# [0.5.2 이식] 이 앱이 DispatchEx 로 띄운 모든 EXCEL.EXE pid. 세션 기록 전에 열기가 실패하면
# 고아 Excel 이 남는데, 강제 정리(초기화/force-restart)가 세션 pid 만 죽이면 영영 안 닫힘 → 전부 추적.
SPAWNED_EXCEL_PIDS = set()
# 진행 중인 비동기 force-restart kill 스레드(종료 시 join 해 고아 방지).
_KILL_INFLIGHT = []
_KILL_INFLIGHT_LOCK = threading.Lock()
EXCEL_LAST_REAP_AT = 0.0
EXCEL_REAP_INTERVAL_SECONDS = float(os.environ.get("B2B_EXCEL_REAP_INTERVAL", "300"))
NATIVE_HOST_PID = int(os.environ.get("B2B_NATIVE_HOST_PID") or "0")
DISABLE_PARENT_WATCH = os.environ.get("B2B_DISABLE_PARENT_WATCH", "").strip().lower() in ("1", "true", "yes")
PARENT_WATCH_GRACE_SECONDS = float(os.environ.get("B2B_PARENT_WATCH_GRACE_SECONDS", "10"))
# 부모(native host) 생존 확인 주기. 1초마다 볼 필요가 없어 30초로 완화 — grace(10s)와 합쳐
# 호스트 사망 후 최악 ~60초 안에 고아 Excel 정리·자가종료가 이뤄지면 충분하다.
PARENT_WATCH_INTERVAL_SECONDS = float(os.environ.get("B2B_PARENT_WATCH_INTERVAL", "30"))
PARENT_WATCH_MISSING_SINCE = 0.0
HEALTH_EXCEL_DIAG_INTERVAL_SECONDS = float(os.environ.get("B2B_HEALTH_EXCEL_DIAG_INTERVAL", "60"))
HEALTH_LAST_EXCEL_DIAG_AT = 0.0
HEALTH_CACHED_EXCEL_DIAG = None
PERF_LOG_INTERVAL_SECONDS = float(os.environ.get("B2B_PERF_LOG_INTERVAL", "60"))
PERF_LAST_LOG_AT = 0.0
RUNTIME_SAMPLER_INTERVAL_SECONDS = float(os.environ.get("B2B_RUNTIME_SAMPLER_INTERVAL", "30"))
RUNTIME_SAMPLER_STARTED = False
# [유휴 무기록] 샘플러가 마지막으로 본 활동 시그니처와 그 시각. 유휴(큐 비고 실행중 작업 없고
# 카운트 불변)면 runtime.sample 을 아예 수집·기록하지 않아 켜두기만 한 상태에서 트레이스가 안 자란다.
RUNTIME_LAST_ACTIVITY_SIG = None
RUNTIME_LAST_ACTIVITY_AT = 0.0
HOUSEKEEPING_INTERVAL_SECONDS = float(os.environ.get("B2B_EXCEL_CLEANUP_INTERVAL", "600"))
HOUSEKEEPING_RUNNING = False
HOUSEKEEPING_LAST_RUN_AT = 0.0
HOUSEKEEPING_LAST_DURATION_MS = 0.0
HOUSEKEEPING_LAST_SKIPPED_REASON = ""
HOUSEKEEPING_RUN_COUNT = 0
HOUSEKEEPING_ERROR = ""
HOUSEKEEPING_GC_LAST_AT = 0.0
HOUSEKEEPING_SNAPSHOT_MAX_BYTES = int(os.environ.get("B2B_PIPELINE_SNAPSHOT_MAX_BYTES", str(256 * 1024 * 1024)))
# [0.5.2 이식] 강제 정리 시 끊어낼 COM 프록시 보관소 — 행 상태 STA 워커로의 Release 마샬링 교착 방지.
_COM_REF_GRAVEYARD = []
PIPELINE_JOBS_LOCK = threading.Lock()
WORKBOOK_CACHE_LOCK = threading.Lock()
NODE_WORKER_LOCK = threading.Lock()
NODE_WORKER = None
NODE_WORKER_SCRIPT_MTIME = None
NODE_WORKER_READY = set()
PREVIEW_ROWS = 500
PREVIEW_COLS = None
MAX_DIFF_CELLS_PER_SHEET = 5000
MAX_PIPELINE_STEP_SNAPSHOTS = 80
MAX_PIPELINE_JOBS = 40
# 큰 출력 워크북은 단계마다 전체를 디스크에 저장(SaveCopyAs)하면 매우 느리다.
# 이 크기를 넘으면 중간 단계 스냅샷을 건너뛰고 "마지막 단계"만 저장한다(동일 파이프라인 재적용은 여전히 즉시).
SNAPSHOT_INTERMEDIATE_MAX_BYTES = 8 * 1024 * 1024
PIPELINE_JOB_TTL_SECONDS = 60 * 60
APP_BUILD_STAMP = "b2b-0.7.2-20260804-assist"
EXCEL_MIRROR_PROTECT_PASSWORD = "b2b_mirror_readonly"


def app_base_dir():
    return Path(getattr(sys, "_MEIPASS", Path(__file__).resolve().parent))


# ===== 버전 확인 =====================================================================
# AX-Cell.exe 가 가진 '파일 버전'(윈도우 파일 속성, 예 0.7.2.0)과 버전 서버(versionTest)의
# version.txt 값을 비교한다. 지금은 확인만 한다 — 다르면 뜨는 안내창은 최종 배포 전에 붙인다.

def _normalize_version_text(text):
    """'0.7.2' / 'v0.7.2' / '0.7.2.0' 을 모두 '0.7.2.0' 으로 맞춘다.
    문자열 그대로 비교하면 '0.7.2' 와 '0.7.2.0' 이 다르다고 나오므로 양쪽 다 여기를 거친다.
    (버전 서버의 normalize_version 과 같은 규칙 — 한쪽만 바꾸면 안 된다)"""
    s = str(text or "").strip().lstrip("vV").strip()
    if not s:
        return ""
    parts = [p for p in s.split(".") if p != ""]
    if not parts or not all(p.isdigit() for p in parts):
        return ""
    parts = (parts + ["0", "0", "0", "0"])[:4]
    try:
        return ".".join(str(int(p)) for p in parts)
    except Exception:
        return ""


def _exe_file_version(exe_path):
    """윈도우 exe 의 파일 버전 리소스를 읽는다(파일 속성 → 자세히 → 파일 버전).
    pywin32 없이 ctypes 만으로 처리 — 백엔드에 새 의존성을 들이지 않기 위해서다."""
    try:
        p = Path(exe_path)
        if not p.exists():
            return ""
        ver_dll = ctypes.WinDLL("version.dll")
        size = ver_dll.GetFileVersionInfoSizeW(ctypes.c_wchar_p(str(p)), None)
        if not size:
            return ""
        buf = ctypes.create_string_buffer(size)
        if not ver_dll.GetFileVersionInfoW(ctypes.c_wchar_p(str(p)), 0, size, buf):
            return ""
        block = ctypes.c_void_p()
        length = ctypes.c_uint()
        if not ver_dll.VerQueryValueW(buf, ctypes.c_wchar_p("\\"),
                                      ctypes.byref(block), ctypes.byref(length)):
            return ""
        # VS_FIXEDFILEINFO: dwFileVersionMS/LS 에 4자리가 16비트씩 들어있다.
        ffi = ctypes.cast(block, ctypes.POINTER(ctypes.c_uint * 4)).contents
        ms, ls = ffi[2], ffi[3]
        return f"{ms >> 16}.{ms & 0xFFFF}.{ls >> 16}.{ls & 0xFFFF}"
    except Exception:
        return ""


def _current_app_version():
    """지금 돌고 있는 AX-Cell 의 버전. 반환 {version, normalized, source}.
      · 배포(프로즌): 실제 exe 의 파일 버전 리소스를 읽는다 — 사용자가 실제로 들고 있는 값
      · 개발(소스 실행): exe 가 없으니 launch_b2b.py 의 CURRENT_VERSION 을 쓴다
    두 경로가 다른 값을 낼 수 있으므로 source 를 같이 돌려줘 화면에서 구분할 수 있게 한다."""
    # 1) 사용자가 실제로 실행한 파일 = AX-Cell.exe 를 먼저 본다.
    #    (백엔드는 B2B_Server.exe 로 도는데, 화면에 띄울 버전은 사용자가 속성 창에서 보는
    #     AX-Cell.exe 의 '파일 버전'이어야 한다. 지금은 gen_version_meta 가 둘 다 같은 값으로
    #     찍지만, 한쪽만 다시 빌드되는 상황에서 표기가 엇갈리지 않게 순서를 못박는다.)
    candidates = []
    exe_dir = None
    try:
        if getattr(sys, "frozen", False):
            exe_dir = Path(sys.executable).resolve().parent
            candidates.append(exe_dir / "AX-Cell.exe")
    except Exception:
        pass
    try:
        candidates.append(app_base_dir() / "AX-Cell.exe")
    except Exception:
        pass
    try:
        if getattr(sys, "frozen", False):
            candidates.append(Path(sys.executable).resolve())   # 폴백: 지금 돌고 있는 exe
    except Exception:
        pass
    seen = set()
    for cand in candidates:
        key = str(cand).lower()
        if key in seen:
            continue
        seen.add(key)
        raw = _exe_file_version(cand)
        if raw:
            return {"version": raw, "normalized": _normalize_version_text(raw), "source": f"exe:{cand.name}"}
    # 2) 소스 실행 — launch_b2b.py 의 CURRENT_VERSION(단일 진실)
    try:
        src = (app_base_dir() / "launch_b2b.py").read_text("utf-8", errors="replace")
        m = re.search(r'^CURRENT_VERSION\s*=\s*["\']([0-9][0-9.]*)["\']', src, re.M)
        if m:
            return {"version": m.group(1), "normalized": _normalize_version_text(m.group(1)),
                    "source": "source:launch_b2b.py"}
    except Exception:
        pass
    return {"version": "", "normalized": "", "source": ""}


def writable_app_dir():
    env_dir = os.environ.get("B2B_WRITABLE_APP_DIR")
    if env_dir:
        return Path(env_dir).expanduser().resolve()
    if getattr(sys, "frozen", False):
        return Path(sys.executable).resolve().parent
    return Path(__file__).resolve().parent


def b2b_logs_dir():
    """트레이스 로그 저장 폴더 — 프로즌/개발 무관하게 항상 %LOCALAPPDATA%\\B2B_logs 로 고정.
    (없으면 만든다. LOCALAPPDATA 없으면 APPDATA→TEMP→홈 순 폴백.)"""
    base = (os.environ.get("LOCALAPPDATA") or os.environ.get("APPDATA")
            or os.environ.get("TEMP") or str(Path.home()))
    d = Path(base) / "B2B_logs"
    try:
        d.mkdir(parents=True, exist_ok=True)
    except Exception:
        pass
    return d


def user_config_dir():
    env_dir = os.environ.get("B2B_USER_CONFIG_DIR")
    if env_dir:
        return Path(env_dir).expanduser().resolve()
    if os.name == "nt":
        base = os.environ.get("LOCALAPPDATA") or os.environ.get("APPDATA")
        if base:
            return Path(base) / "B2B_Billing_Agent"
    return Path.home() / ".b2b_billing_agent"


def logic_backup_settings_path():
    return user_config_dir() / "settings.json"


def _load_user_settings():
    path = logic_backup_settings_path()
    try:
        if path.exists():
            data = json.loads(path.read_text(encoding="utf-8"))
            return data if isinstance(data, dict) else {}
    except Exception:
        pass
    return {}


def _save_user_settings(data):
    path = logic_backup_settings_path()
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(data or {}, ensure_ascii=False, indent=2), encoding="utf-8")


def default_logic_backup_dir():
    return writable_app_dir() / "auto_backup"


def default_output_dir():
    # [실행기 파일출력] 전체실행 결과를 라이브에 반영하는 대신 파일로 저장하는 위치(실행 중인 앱 폴더/output).
    return writable_app_dir() / "output"


def logic_backup_dir():
    # [요청] 자동백업은 '항상 실행 중인 앱 위치'(writable_app_dir()/auto_backup)에 만든다.
    # 예전엔 settings.logicBackupDir 고정 경로를 우선해, 실행 버전이 바뀌어도 백업이 옛 버전 폴더(예: 0.5.13)
    # 에만 쌓였다. 이제 설정 경로는 무시하고 항상 실행 위치를 쓴다 → 0.5.15 를 돌리면 0.5.15/auto_backup 에 생김.
    return default_logic_backup_dir()


def logic_backup_dir_info():
    path = logic_backup_dir()  # 항상 실행 위치/auto_backup
    try:
        path.mkdir(parents=True, exist_ok=True)
    except Exception:
        pass
    return {
        "ok": True,
        "path": str(path),
        "custom": False,
        "defaultPath": str(default_logic_backup_dir()),
        "settingsPath": str(logic_backup_settings_path()),
        "exists": path.exists(),
    }


def set_logic_backup_dir(path):
    chosen = Path(str(path or "")).expanduser().resolve()
    chosen.mkdir(parents=True, exist_ok=True)
    if not chosen.is_dir():
        raise ValueError(f"폴더가 아닙니다: {chosen}")
    probe = chosen / f".b2b_write_test_{uuid.uuid4().hex}"
    try:
        probe.write_text("ok", encoding="utf-8")
    finally:
        try:
            probe.unlink(missing_ok=True)
        except Exception:
            pass
    settings = _load_user_settings()
    settings["logicBackupDir"] = str(chosen)
    _save_user_settings(settings)
    return logic_backup_dir_info()


def choose_logic_backup_dir_dialog():
    try:
        import tkinter as tk
        from tkinter import filedialog
    except Exception as err:
        raise RuntimeError(f"폴더 선택 창을 열 수 없습니다(tkinter 없음): {err}") from err
    root = tk.Tk()
    root.withdraw()
    try:
        try:
            root.attributes("-topmost", True)
        except Exception:
            pass
        initial = str(logic_backup_dir())
        selected = filedialog.askdirectory(
            parent=root,
            title="스킬 자동저장 폴더 선택",
            initialdir=initial if Path(initial).exists() else str(default_logic_backup_dir().parent),
            mustexist=False,
        )
    finally:
        try:
            root.destroy()
        except Exception:
            pass
    if not selected:
        info = logic_backup_dir_info()
        info.update({"ok": False, "cancelled": True})
        return info
    return set_logic_backup_dir(selected)


def reset_logic_backup_dir():
    settings = _load_user_settings()
    settings.pop("logicBackupDir", None)
    _save_user_settings(settings)
    return logic_backup_dir_info()


def node_executable():
    bundled = app_base_dir() / ("node.exe" if os.name == "nt" else "node")
    if bundled.exists():
        return str(bundled)
    found = shutil.which("node")
    return found or None


def hidden_subprocess_kwargs():
    if os.name != "nt":
        return {}
    startupinfo = subprocess.STARTUPINFO()
    startupinfo.dwFlags |= subprocess.STARTF_USESHOWWINDOW
    # [모래시계 수정] CREATE_NO_WINDOW/USESHOWWINDOW 는 콘솔 '창'만 숨기고, Windows 가
    # CreateProcess 때 기본으로 켜는 '앱 시작 중' 피드백 커서(화살표+빙글이)는 못 막는다.
    # 유지관리 루프가 tasklist/taskkill 을 주기적으로 띄울 때마다 유휴 상태에서도 커서가
    # 돌았다 풀렸다 반복돼 보였다. 시작 피드백을 명시적으로 끈다.
    # (subprocess.STARTF_FORCEOFFFEEDBACK 상수는 Python 3.13+ — 구버전은 winbase.h 값 0x80)
    startupinfo.dwFlags |= getattr(subprocess, "STARTF_FORCEOFFFEEDBACK", 0x00000080)
    startupinfo.wShowWindow = 0
    return {
        "startupinfo": startupinfo,
        "creationflags": subprocess.CREATE_NO_WINDOW,
    }


def _kill_pid_quiet(pid):
    # [검은창 제거] taskkill 을 os.system 으로 부르면 cmd.exe 콘솔창이 깜빡 떴다 사라진다(생성기 VBA 적용 시
    # 격리 인스턴스 정리에서 발생). subprocess + CREATE_NO_WINDOW 로 창 없이 종료한다.
    try:
        if not pid:
            return
        subprocess.run(["taskkill", "/F", "/PID", str(int(pid))],
                       stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL,
                       **hidden_subprocess_kwargs())
    except Exception:
        pass


def cleanup_node_worker():
    global NODE_WORKER
    worker = NODE_WORKER
    NODE_WORKER = None
    NODE_WORKER_READY.clear()
    if not worker or worker.poll() is not None:
        return
    try:
        worker.kill()
        worker.wait(timeout=3)
    except Exception:
        try:
            worker.terminate()
        except Exception:
            pass


def cleanup_backend_runtime_files():
    """Delete runtime-only result/snapshot files created under BACKEND_DIR.

    This intentionally keeps uploaded workbook caches and auto_backup files.
    Pre-apply snapshots created by /api/excel/save live in RESULTS and are only
    useful while the process is alive; keeping them after exit just consumes disk.
    """
    backend_root = BACKEND_DIR.resolve()
    removed = 0
    failed = 0

    def _safe_unlink(raw_path):
        nonlocal removed, failed
        if not raw_path:
            return
        try:
            path = Path(raw_path).resolve()
            if path == backend_root or backend_root not in path.parents:
                return
            if path.exists() and path.is_file():
                path.unlink()
                removed += 1
        except Exception:
            failed += 1

    try:
        for item in list(RESULTS.values()):
            _safe_unlink(item.get("path") if isinstance(item, dict) else None)
        RESULTS.clear()
    except Exception:
        failed += 1

    try:
        for key, snapshot in list(PIPELINE_STEP_SNAPSHOTS.items()):
            try:
                _delete_pipeline_snapshot_entry(key, snapshot)
                removed += 1
            except Exception:
                failed += 1
        PIPELINE_STEP_SNAPSHOTS.clear()
    except Exception:
        failed += 1

    try:
        shutil.rmtree(BACKEND_DIR / "pipeline_step_snapshots", ignore_errors=True)
    except Exception:
        failed += 1

    try:
        _perf_trace("runtime.backend_files.cleanup", removed=removed, failed=failed)
    except Exception:
        pass


def _other_b2b_backend_running():
    """현재 프로세스 외 다른 B2B 백엔드(B2B_Server.exe 또는 serve_b2b.py python)가 살아있는지.
    동시 백엔드(이중 실행/강제재시작 잔존)가 있으면 그 백엔드의 작업복사본을 우리가 지우면 안 되므로 가드용."""
    try:
        import psutil as _ps
    except Exception:
        return False
    try:
        me = os.getpid()
        # [onefile 오탐 방지] PyInstaller onefile exe 는 '부트로더 부모' + '앱 자식' 두 프로세스가 모두
        # B2B_Server.exe 다. 자식(우리 코드)이 자기 부모 부트로더를 '다른 백엔드'로 오인하면 other_backend=True
        # 가 되어 시작 정리에서 업로드 작업본({uuid}_*.xlsx)을 영영 건너뛴다. 자기 프로세스 트리(자신·부모·자식)는
        # 제외하고 '진짜 별개 백엔드'만 카운트한다(별도로 두 번 실행하면 그건 트리 밖이라 정상 감지됨).
        skip = {me}
        try:
            skip.add(os.getppid())
        except Exception:
            pass
        try:
            for c in _ps.Process(me).children(recursive=True):
                skip.add(c.pid)
        except Exception:
            pass
        for p in _ps.process_iter(["pid", "name", "cmdline"]):
            try:
                if p.info.get("pid") in skip:
                    continue
                nm = (p.info.get("name") or "").lower()
                if nm == "b2b_server.exe":
                    return True
                if nm.startswith("python") or nm.startswith("pythonw"):
                    cl = " ".join(p.info.get("cmdline") or [])
                    if "serve_b2b" in cl:
                        return True
            except Exception:
                continue
    except Exception:
        return False
    return False


def cleanup_stale_temp_artifacts(min_age_seconds=300, excel_diag_max_age_seconds=86400, mei_max_age_seconds=86400):
    """[디스크 누수 방지] 앱 시작 시 이전 실행(크래시/강제종료 포함)이 남긴 임시 작업물을 정리한다.
    새 프로세스라 활성 세션이 아직 없어 안전하다(종료-시 삭제보다 기능 리스크가 낮고, 매 시작마다 정리되므로
    누적이 한 실행분으로 한계가 잡힌다 = '재활용' 효과).
    - BACKEND_DIR(b2b_backend_v044): 단독 백엔드면 폴더 전체를 무조건 비운다(나이 게이트/예외 없음) — 재시작 시
      파일을 새로 업로드하므로 여기 남은 건 전부 이전 실행의 죽은 작업물. 진짜 스킬 백업은 여기가 아니라
      writable_app_dir()/auto_backup(앱 폴더)에 있어 무관. 다른 백엔드가 살아있을 때만 전이성 stale 만 정리.
    - Temp 루트: b2b_isopipe_*/b2b_replace_*/B2B_ver*_single_*/b2b_*test_* (현재 실행 폴더는 보존).
    - PyInstaller _MEI*: 충분히 오래된 것만(현재 _MEIPASS 제외) — 타 앱 오인삭제 최소화 위해 보수적 나이.
    - Excel 진단 로그(Temp/Diagnostics/EXCEL): 오래된 파일(가장 큰 누적원).
    동시 실행(드묾) 보호: min_age_seconds 이내 최근 항목은 건너뛴다."""
    try:
        now = time.time()
        temp = Path(tempfile.gettempdir())
        try:
            self_exe = str(Path(sys.executable).resolve()).lower()
        except Exception:
            self_exe = ""
        try:
            self_mei = str(Path(getattr(sys, "_MEIPASS", "") or ".").resolve()).lower()
        except Exception:
            self_mei = ""
        stats = {"dirs": 0, "files": 0, "bytes": 0, "failed": 0}
        # [동시 백엔드 보호] 다른 B2B 백엔드가 살아있으면 그 백엔드의 작업복사본/단일임시/_MEI 를 우리가 지우면
        # 그 세션이 깨진다(이중 실행/강제재시작 잔존 시 발생). 이 경우 위험한 정리는 건너뛰고 Excel 진단 로그
        # (항상 안전)만 정리한다. 단독 백엔드일 때만 b2b_backend/single/_MEI 정리(그땐 전부 죽은 잔재라 안전).
        other_backend = _other_b2b_backend_running()

        def _age_ok(p, min_age):
            try:
                return (now - p.stat().st_mtime) >= float(min_age)
            except Exception:
                return False

        def _is_self(p):
            try:
                pr = str(p.resolve()).lower()
            except Exception:
                return False
            # 현재 실행 exe 가 이 폴더 안에 있거나(_single_/onedir), 현재 _MEIPASS 면 보존
            return bool((self_exe and self_exe.startswith(pr)) or (self_mei and self_mei == pr) or (self_mei and self_mei.startswith(pr)))

        def _rm(p):
            if _is_self(p):
                return
            try:
                if p.is_dir():
                    shutil.rmtree(p, ignore_errors=True)
                    stats["dirs"] += 1
                elif p.is_file():
                    try:
                        stats["bytes"] += p.stat().st_size
                    except Exception:
                        pass
                    p.unlink()
                    stats["files"] += 1
            except Exception:
                stats["failed"] += 1

        # 1) BACKEND_DIR 내부 작업물
        # [사용자 지시] 단독 백엔드(정상 재시작)면 여기 든 건 전부 이전 실행의 죽은 작업물이다(재시작 시 파일을
        # 새로 업로드하므로 남길 게 없다) → 나이 게이트/auto_backups 예외 없이 폴더 전체를 무조건 비운다.
        # (진짜 스킬 백업은 여기가 아니라 writable_app_dir()/auto_backup 앱 폴더에 있어 영향 없음.)
        # 예외: 다른 B2B 백엔드가 '실제로' 살아있으면 그 세션이 워크북을 여기서 열고 있어(예: b2b_replace_/uuid_
        # 사본) 통째로 지우면 그 Excel 이 깨진다 → 이 경우만 stale 한 '전이성 작업물'만 보수적으로 정리한다.
        try:
            if BACKEND_DIR.exists():
                if not other_backend:
                    for child in BACKEND_DIR.iterdir():
                        _rm(child)   # 전부 삭제(auto_backups 포함, 나이 무관)
                else:
                    _transient_prefixes = ("prestep_", "result_", "live_", "b2b_replace_", "excel_")
                    for child in BACKEND_DIR.iterdir():
                        if not _age_ok(child, min_age_seconds):
                            continue
                        if any(child.name.startswith(pfx) for pfx in _transient_prefixes):
                            _rm(child)   # 다른 백엔드 생존 시엔 전이성 stale 만(원본사본/기타 보존)
        except Exception:
            pass

        # 2) Temp 루트 B2B 임시 폴더
        for pat in (() if other_backend else ("b2b_isopipe_*", "b2b_replace_*", "b2b_vba_runner_*", "b2b_freshprobe_*", "B2B_ver*_single_*", "b2b_nametest_*", "b2b_realtest_*")):
            try:
                for d in temp.glob(pat):
                    if _age_ok(d, min_age_seconds):
                        _rm(d)
            except Exception:
                pass

        # 2-b) WebView2 사용자데이터(B2B_WebView2/verNNN_<pid>): '죽은 pid' 폴더 삭제(살아있는 인스턴스 보존).
        # NativeHost 도 시작 시 정리하지만(죽은 pid 즉시), 백엔드에서도 백스톱으로 정리한다 — Chromium 캐시가
        # 실행마다 verNNN_<pid> 로 쌓여 수백 MB 누수하던 문제. [수정] 예전엔 'ver044_' 만 글롭해 구버전 폴더
        # (ver043_ 등)가 영영 안 지워졌다 → 'ver*_*' 로 넓혀 모든 버전 마커를 잡는다. dead-pid 검사라 살아있는
        # 인스턴스(현재/다른 백엔드)는 보존되고 다른 백엔드 생존과도 무관하게 안전.
        try:
            wv = temp / "B2B_WebView2"
            if wv.exists():
                for d in wv.glob("ver*_*"):
                    if not d.is_dir():
                        continue
                    # [반쪽 수정 보완] 예전엔 psutil 단독 판정(_ps and _ps.pid_exists)이라, psutil 이 없는
                    # 오프라인 frozen 빌드(build_exe_offline.bat 은 wheelhouse 에 psutil 이 없다)에서는
                    # 모든 폴더가 alive=False 로 떨어져 '지금 쓰고 있는' WebView2 프로필까지 지웠다.
                    # _is_pid_alive 는 psutil → ctypes(OpenProcess) → tasklist 사다리라 psutil 없이도 정확하다.
                    alive = True   # 판정 실패 시엔 지우지 않는다(사용 중 프로필 삭제 방지)
                    try:
                        wv_pid = int(d.name.rsplit("_", 1)[1])
                        alive = _is_pid_alive(wv_pid)
                    except Exception:
                        alive = False   # 폴더명이 pid 형식이 아니면 옛 잔재 → 정리 대상
                    if not alive:
                        _rm(d)
        except Exception:
            pass

        # 3) PyInstaller _MEI* (현재 _MEIPASS 는 _rm 의 _is_self 가 보존)
        # B2B 자신의 onefile 추출본은 datas 로 serve_b2b.py 를 포함한다 → 그 마커로 '우리 것'을 식별해 나이 무관
        # 즉시 삭제한다(크래시/강제종료로 남은 우리 _MEI 잔재를 재실행 때 확실히 청소). 마커가 없는 _MEI(다른
        # PyInstaller 앱의 것)는 오인 삭제 방지를 위해 종전대로 보수적 나이(기본 24h)만. 둘 다 단독 백엔드일 때만.
        try:
            for d in (temp.glob("_MEI*") if not other_backend else []):
                if not d.is_dir():
                    continue
                is_b2b_own = False
                try:
                    is_b2b_own = (d / "serve_b2b.py").exists()
                except Exception:
                    is_b2b_own = False
                if is_b2b_own or _age_ok(d, mei_max_age_seconds):
                    _rm(d)
        except Exception:
            pass

        # 4) Excel 진단 로그(오래된 파일) — 가장 큰 누적원
        try:
            diag = temp / "Diagnostics" / "EXCEL"
            if diag.exists():
                for f in diag.rglob("*"):
                    try:
                        if f.is_file() and (now - f.stat().st_mtime) >= float(excel_diag_max_age_seconds):
                            try:
                                stats["bytes"] += f.stat().st_size
                            except Exception:
                                pass
                            f.unlink()
                            stats["files"] += 1
                    except Exception:
                        stats["failed"] += 1
        except Exception:
            pass

        try:
            _vba_trace("startup.temp_cleanup", removedDirs=stats["dirs"], removedFiles=stats["files"],
                       freedMb=round(stats["bytes"] / (1024 * 1024), 1), failed=stats["failed"],
                       otherBackend=bool(other_backend))
        except Exception:
            pass
        return stats
    except Exception:
        return None


def excel_available():
    return os.name == "nt" and pythoncom is not None and win32com is not None


def cleanup_excel_sessions():
    if not excel_available():
        return
    pids = set()
    try:
        pids.update(int(session.get("pid")) for session in list(EXCEL_SESSIONS.values()) if session.get("pid"))
    except Exception:
        pass
    try:
        # 세션 기록 전에 실패한 고아 인스턴스까지(이 앱이 띄운 pid 전체).
        pids.update(int(p) for p in SPAWNED_EXCEL_PIDS)
    except Exception:
        pass
    _perf_trace("excel.cleanup.start", pids=sorted(pids))
    try:
        excel_call(_cleanup_excel_sessions_impl, timeout=20)
    except Exception:
        # Shutdown must not leave Excel behind.  If COM is busy or the STA worker
        # times out, kill only the Excel processes that this app created.
        try:
            EXCEL_SESSIONS.clear()
        except Exception:
            pass
    # graceful 정리가 성공했어도 추적된 pid 가 살아 있으면(고아) 마저 종료한다.
    for pid in pids:
        if _is_pid_alive(pid):
            _force_kill_pid(pid)
    try:
        SPAWNED_EXCEL_PIDS.clear()
    except Exception:
        pass
    _perf_trace(
        "excel.cleanup.end",
        pids=sorted(pids),
        aliveAfter=[pid for pid in sorted(pids) if _is_pid_alive(pid)],
    )


def ensure_excel_worker():
    global EXCEL_QUEUE, EXCEL_THREAD
    if EXCEL_THREAD and EXCEL_THREAD.is_alive() and EXCEL_QUEUE is not None:
        return
    EXCEL_QUEUE = queue.Queue()

    def worker():
        pythoncom.CoInitializeEx(pythoncom.COINIT_APARTMENTTHREADED)
        try:
            while True:
                try:
                    item = EXCEL_QUEUE.get(timeout=0.05)
                except queue.Empty:
                    _maybe_quit_idle_python_skill_app()
                    pythoncom.PumpWaitingMessages()
                    continue
                if item is None:
                    break
                fn, args, kwargs, done = item
                # 녹화 중이면 B2B 가 직접 하는 Excel 작업(파이프라인 실행 등)이
                # 레코더에 되잡히지 않도록 잡 실행 동안 replaying 플래그를 켠다.
                _rec_set_replaying = None
                try:
                    from record_service import RECORD_SERVICE
                    _rec_set_replaying = RECORD_SERVICE.set_replaying
                except Exception:
                    pass
                try:
                    if _rec_set_replaying:
                        _rec_set_replaying(True)
                    result = fn(*args, **kwargs)
                    pythoncom.PumpWaitingMessages()
                    done.put((True, result))
                except Exception as err:
                    done.put((False, err))
                finally:
                    if _rec_set_replaying:
                        _rec_set_replaying(False)
        finally:
            try:
                _cleanup_excel_sessions_impl()
            finally:
                pythoncom.CoUninitialize()

    EXCEL_THREAD = threading.Thread(target=worker, name="b2b-excel-com", daemon=True)
    EXCEL_THREAD.start()


def excel_call(fn, *args, timeout=60, **kwargs):
    if not excel_available():
        raise RuntimeError("Microsoft Excel COM automation is not available. Excel and pywin32 are required.")
    ensure_excel_worker()
    done = queue.Queue(maxsize=1)
    EXCEL_QUEUE.put((fn, args, kwargs, done))
    try:
        ok, result = done.get(timeout=timeout)
    except queue.Empty:
        raise TimeoutError(
            f"Excel COM 작업이 {timeout}초 안에 끝나지 않았습니다. "
            "컴퓨터 성능에 따라 Excel 작업이 지연될 수 있습니다. 잠시 후 다시 시도해 주세요."
        )
    if ok:
        return result
    raise result


def _cleanup_excel_sessions_impl():
    _quit_python_skill_app()
    sessions = list(EXCEL_SESSIONS.values())
    EXCEL_SESSIONS.clear()
    pids = set()
    apps = []
    app_keys = set()
    for session in sessions:
        pid = session.get("pid")
        if pid:
            pids.add(pid)
        try:
            app, wb = session_workbook(session)
            try:
                key = int(app.Hwnd)
            except Exception:
                key = id(app)
            if key not in app_keys:
                app_keys.add(key)
                apps.append(app)
            _close_companion_workbooks(session, app)
            wb.Close(SaveChanges=False)
        except Exception:
            pass
        for key in ("openTempPath", "workingCopyPath"):
            temp_path = session.get(key)
            if temp_path:
                try:
                    Path(temp_path).unlink(missing_ok=True)
                except Exception:
                    pass
        for cdir in session.get("companionTemps") or []:
            try:
                shutil.rmtree(cdir, ignore_errors=True)
            except Exception:
                pass
    for app in apps:
        try:
            app.DisplayAlerts = False
        except Exception:
            pass
        try:
            app.Quit()
        except Exception:
            pass
    global LIVE_EXCEL_APP
    _note_live_app_reset("close_all_sessions", pids=len(pids))
    LIVE_EXCEL_APP = None
    for pid in pids:
        deadline = time.time() + 1.5
        while time.time() < deadline and _is_pid_alive(pid):
            time.sleep(0.1)
        if _is_pid_alive(pid):
            _force_kill_pid(pid)


def _force_restart_excel_sessions_direct(wait=False):
    """COM 큐를 '우회'하는 응급 복구. 공유 EXCEL.EXE 가 모달/행으로 굳으면 모든 excel_call 이
    타임아웃되고 일반 close-all 조차 같은 큐에 줄을 서서 들어가지 못한다(단일 인스턴스의 단일 장애점).
    여기서는 COM 호출 없이 세션에 저장해 둔 pid 만으로 프로세스를 강제 종료하고 상태를 비운다.
    워커 스레드가 EXCEL_LOCK 을 쥔 채 멈춰 있을 수 있으므로 락은 짧게만 시도하고 실패해도 진행한다
    (프로세스가 죽으면 굳어 있던 COM 호출도 오류로 풀려난다).

    wait=True 는 종료(/api/app/shutdown) 전용: kill 완료까지 이 스레드에서 기다린다.
    백그라운드 kill 인 채로 응답하면 호스트가 응답을 받자마자 서버를 죽여 kill 스레드가
    함께 죽고 EXCEL.EXE 고아가 남는 경합이 있다."""
    # [녹화 보호] 녹화 중(종료 제외)에는 절대 강제 재시작하지 않는다. 녹화는 공유 라이브 Excel 을
    # 매크로 레코더 상태로 두므로 이 사이의 COM 타임아웃/모달을 클라 워치독(noteExcelComTimeout)이
    # '행'으로 오판해 /api/excel/force-restart 로 들어오면, 여기서 공유 인스턴스를 죽여 진행 중 녹화가
    # 통째로 유실된다(실측 2026-07-28: 시작~정지 사이 사망 → 정지 시 새 빈 인스턴스라 harvested=0).
    # 종료(wait=True, /api/app/shutdown)만 예외 — 어차피 프로세스가 내려간다.
    if not wait and "NATIVE_RECORDING" in globals() and NATIVE_RECORDING.get("active"):
        _note_live_app_reset("force_restart_skipped_during_recording", skipped=True)
        return {"ok": True, "skipped": True, "reason": "recording_active"}
    acquired = False
    try:
        acquired = EXCEL_LOCK.acquire(timeout=2)
    except Exception:
        acquired = False
    try:
        sessions = list(EXCEL_SESSIONS.values())
        EXCEL_SESSIONS.clear()
    finally:
        if acquired:
            try:
                EXCEL_LOCK.release()
            except Exception:
                pass
    # [작업 중단 강제화] 진행 중이던 파이프라인 잡을 종료로 마킹 — 강제 재시작 후에도 클라 폴링이
    # status=running 에 영영 매달리지 않게 한다. 굳었던 실행 스레드는 pid kill 로 COM 오류가 나며
    # 깨어나 자체 실패 처리를 하지만, 그와 무관하게 폴링부터 즉시 풀어준다(errorInfo.cancelled
    # 형식은 raise_if_pipeline_cancelled 와 동일 — 프론트는 '사용자 중단'으로 조용히 복귀).
    try:
        with PIPELINE_JOBS_LOCK:
            for _job in PIPELINE_JOBS.values():
                if _job.get("status") not in ("done", "error"):
                    _job["cancelRequested"] = True
                    _job["status"] = "error"
                    _job["errorInfo"] = {"cancelled": True, "stepIdx": -1,
                                         "message": "작업이 강제 중단되었습니다(Excel 세션 재시작)."}
                    _job["updated"] = time.time()
    except Exception:
        pass
    global LIVE_EXCEL_APP, PYTHON_SKILL_APP, PYTHON_SKILL_APP_PID, PYTHON_SKILL_APP_LAST_USED
    # COM 프록시 전역을 그냥 None 으로 떨어뜨리면 마지막 참조 해제(Release)가 이 HTTP 스레드에서
    # 일어난다. 행 상태의 STA 워커로 마샬링되는 Release 는 같이 굳을 수 있으므로(초기화가 영영
    # 안 끝나는 증상) 참조를 graveyard 로 옮겨 Release 자체를 막는다.
    _note_live_app_reset("force_restart_direct")
    if LIVE_EXCEL_APP is not None:
        _COM_REF_GRAVEYARD.append(LIVE_EXCEL_APP)
    if PYTHON_SKILL_APP is not None:
        _COM_REF_GRAVEYARD.append(PYTHON_SKILL_APP)
    LIVE_EXCEL_APP = None
    PYTHON_SKILL_APP = None
    pids = set()
    if PYTHON_SKILL_APP_PID:
        # 숨김 Python 스킬 인스턴스도 같이 정리(놔두면 보이지 않는 EXCEL.EXE 고아로 남는다).
        try:
            pids.add(int(PYTHON_SKILL_APP_PID))
        except Exception:
            pass
    PYTHON_SKILL_APP_PID = None
    PYTHON_SKILL_APP_LAST_USED = 0.0
    for session in sessions:
        pid = session.get("pid")
        if pid:
            try:
                pids.add(int(pid))
            except Exception:
                pass
    # 세션 기록 전에 열기가 실패한 고아 인스턴스까지 포함(이 앱이 띄운 pid 전체).
    try:
        pids.update(int(p) for p in SPAWNED_EXCEL_PIDS)
        SPAWNED_EXCEL_PIDS.clear()
    except Exception:
        pass
    _perf_trace("excel.force_restart.start", pids=sorted(pids))

    def _kill_and_cleanup():
        # taskkill(프로세스당 최대 3초) + 생존 확인 루프는 수 초가 걸릴 수 있다.
        # HTTP 응답을 잡아두면 초기화 버튼이 그 시간만큼 굳으므로 백그라운드에서 수행한다.
        for pid in pids:
            if _is_pid_alive(pid):
                _force_kill_pid(pid)
        deadline = time.time() + 3.0
        while time.time() < deadline and any(_is_pid_alive(p) for p in pids):
            time.sleep(0.2)
        _perf_trace(
            "excel.force_restart.end",
            pids=sorted(pids),
            aliveAfter=[pid for pid in sorted(pids) if _is_pid_alive(pid)],
        )
        # 프로세스 종료 후에야 파일 잠금이 풀리므로 임시 파일 정리는 마지막에.
        for session in sessions:
            for key in ("openTempPath", "workingCopyPath"):
                temp_path = session.get(key)
                if temp_path:
                    try:
                        Path(temp_path).unlink(missing_ok=True)
                    except Exception:
                        pass
            for cdir in session.get("companionTemps") or []:
                try:
                    shutil.rmtree(cdir, ignore_errors=True)
                except Exception:
                    pass

    if wait:
        _kill_and_cleanup()
    else:
        # [종료 경합] 비동기 kill 은 데몬 스레드라, 직후 /api/app/shutdown 이 os._exit 하면 그대로
        # 잘려 아직 taskkill 이 발행되지 않은 EXCEL.EXE 가 고아로 남는다('작업 중단 → 2단계
        # force-restart → 곧바로 X 종료'는 자연스러운 순서라 실제로 겹친다).
        # 종료가 기다릴 수 있도록 진행 중인 kill 스레드를 등록해 둔다.
        t = threading.Thread(target=_kill_and_cleanup, name="b2b-force-restart-kill", daemon=True)
        with _KILL_INFLIGHT_LOCK:
            _KILL_INFLIGHT.append(t)
        t.start()
    return {"ok": True, "killing": len(pids), "sessions": len(sessions)}


def _join_inflight_kills(timeout):
    """진행 중인 비동기 force-restart kill 스레드를 기다린다(종료 경로 전용)."""
    deadline = time.time() + max(0.0, float(timeout or 0))
    with _KILL_INFLIGHT_LOCK:
        threads = [t for t in _KILL_INFLIGHT if t.is_alive()]
    for t in threads:
        remain = deadline - time.time()
        if remain <= 0:
            break
        try:
            t.join(timeout=remain)
        except Exception:
            pass
    with _KILL_INFLIGHT_LOCK:
        _KILL_INFLIGHT[:] = [t for t in _KILL_INFLIGHT if t.is_alive()]
    return not _KILL_INFLIGHT


atexit.register(cleanup_node_worker)
atexit.register(cleanup_backend_runtime_files)
atexit.register(cleanup_excel_sessions)


class PipelineExecutionError(RuntimeError):
    def __init__(self, message, info=None):
        if info is None and isinstance(message, dict):
            info = message
            message = info.get("message") or info.get("error") or "pipeline step failed"
        super().__init__(message)
        self.info = info or {}


def _pipeline_error_guide(message, code=""):
    """난해한 엔진 예외를 (원인, 프롬프트 작성 가이드) 한국어 쌍으로 변환한다.
    사용자가 '왜 났는지' 이해하고 '다음엔 어떻게 요청해야 하는지' 케이스별로 알 수 있게 한다."""
    m = str(message or "")
    ml = m.lower()
    def has(*subs):
        return any(s in ml for s in subs)
    if has("cannot convert") and has("excel", "cell"):
        return ("행(여러 값)을 한 칸(셀)에 통째로 쓰려다 실패했습니다.",
                "한 셀이 아니라 '표/범위'로 써달라고 명확히 하세요. 예: \"안전제일 행들을 새 시트 A1부터 표로 넣어줘\" 또는 \"한 행씩 차례로 추가해줘\".")
    if ("sheet" in ml and "not" in ml and "found" in ml) or ("시트" in m and ("없" in m or "찾지" in m)):
        return ("지정한 시트를 찾지 못했습니다(앞 단계에서 만든 중간 시트를 엉뚱한 곳에서 찾는 경우가 흔합니다).",
                "이전 단계가 만든 시트를 쓸 땐 그 이름을 정확히 적고 \"앞 단계에서 만든 ○○ 시트에서\"라고 하세요. 원본 입력 시트는 @시트[파일/시트]로 콕 집어 지정하세요.")
    if ("column not found" in ml) or ("col" in ml and "not found" in ml) or ("열" in m and ("없" in m or "찾지" in m)) or ("헤더" in m and "찾" in m):
        return ("표의 헤더(열 이름)를 찾지 못했습니다.",
                "열을 정확한 헤더명이나 열 문자로 지정하세요. 예: \"수납금액 열\" 또는 \"C열\". 가장 확실한 방법은 @컬럼[...] / @범위[...] 로 지정하는 것입니다.")
    if has("has no attribute"):
        return ("데이터 값을 객체처럼 잘못 다뤄 생긴 코드 오류입니다(생성된 코드 문제).",
                "한 번 더 생성하거나 작업을 더 작게 쪼개 요청하세요. 예: \"합계만 B30 셀에 직접 써줘\"처럼 단순하게.")
    if has("is not defined") or "name '" in ml:
        return ("코드가 정의되지 않은 것을 사용했습니다(생성된 코드 문제).",
                "재생성하거나 한 번에 한 작업씩 나눠 요청하세요. 여러 작업을 한 문장에 몰아넣지 마세요.")
    if has("index out of range", "list index") or ("out of range" in ml and "subscript" not in ml):
        return ("행/열 범위를 벗어났습니다.",
                "대상 범위를 명시하세요. 예: \"A4:D24 범위만\", \"헤더는 3행\"처럼 위치를 알려주면 안전합니다.")
    if has("division by zero", "zerodivision"):
        return ("0으로 나누는 계산이 있었습니다.",
                "0일 때 처리를 알려주세요. 예: \"분모가 0이면 결과는 0으로\".")
    if has("merged") or "병합" in m:
        return ("병합된 셀에 직접 쓰려다 실패했습니다.",
                "병합 영역은 \"왼쪽 위 셀에만 써줘\"라고 하거나, 병합/서식 변경이 필요하면 코드 첫 줄에 # B2B_ENGINE_FALLBACK: excel-com 을 넣어 Excel 처리를 요청하세요.")
    if has("read-only", "read only", "permission denied") or ("저장" in m and ("실패" in m or "권한" in m)):
        return ("입력 파일은 읽기 전용이라, 입력에 쓰거나 저장하려다 실패했을 수 있습니다.",
                "결과는 출력 파일에 쓰도록 하세요. 입력 파일을 꼭 바꿔야 하면 코드 첫 줄에 # B2B_ENGINE_FALLBACK: excel-com 을 넣어 Excel 처리를 요청하세요.")
    # ── VBA(Excel COM 매크로) 특유 오류 ── (subscript/accessvbom/문법 등은 위 일반 케이스보다 먼저 잡히도록 케이스 조건을 한정)
    code_text = str(code or "")
    if "vba 안전 검사" in m:
        return ("생성된 VBA 코드가 안전 검사를 통과하지 못했습니다.",
                "열 문자는 숫자로 암산하지 말고 BP/BQ처럼 그대로 쓰고, 매칭된 행만 갱신하도록 다시 생성하세요.")
    if re.search(r"\bContinue\s+For\b", code_text, re.IGNORECASE):
        return ("생성된 VBA 코드에 Excel VBA가 지원하지 않는 문법이 있습니다.",
                "Continue For 대신 If Len(...) > 0 Then ... End If 구조로 다시 생성하세요.")
    if re.search(r"\b(?=[A-Za-z0-9_]*bp)(?=[A-Za-z0-9_]*col)[A-Za-z_][A-Za-z0-9_]*\s*=\s*58\b", code_text, re.IGNORECASE) \
            or re.search(r"\b(?=[A-Za-z0-9_]*bq)(?=[A-Za-z0-9_]*col)[A-Za-z_][A-Za-z0-9_]*\s*=\s*59\b", code_text, re.IGNORECASE):
        return ("생성된 VBA 코드가 BP/BQ 열 번호를 잘못 계산했습니다.",
                "BP는 68, BQ는 69입니다. 하지만 숫자를 직접 쓰지 말고 ws.Columns(\"BP\").Column 또는 ws.Cells(r, \"BP\")처럼 열 문자를 그대로 쓰게 다시 생성하세요.")
    if ("매크로를 실행할 수 없습니다" in m) or has("cannot run the macro", "macro may not be available", "macros may be disabled"):
        return ("Excel이 VBA 실행을 거부했습니다.",
                "앱과 Excel 프로세스를 모두 닫은 뒤 다시 실행하세요. 같은 파일에서 채팅 단일 적용은 되는데 전체실행만 실패하면 프로그램 실행 경로 문제입니다.")
    if has("accessvbom") or ("프로젝트에 접근" in m) or ("매크로 설정" in m):
        return ("Excel이 VBA(매크로) 접근을 차단했습니다.",
                "이건 프롬프트로는 해결되지 않습니다 — Excel 옵션 > 보안 센터 > 매크로 설정에서 'VBA 프로젝트 개체 모델에 대한 액세스 신뢰'를 켠 뒤 파일을 다시 여세요.")
    if has("subscript out of range") or ("첨자" in m) or ("적용 범위를 벗어" in m):
        return ("지정한 시트·파일·범위가 없거나 이름이 틀렸습니다(subscript out of range).",
                "@파일[...] / @시트[파일/시트]로 정확한 이름을 지정하세요. 앞 단계에서 만든 시트면 그 이름을 그대로, 새로 만들 거면 \"새 시트 ○○에\"라고 명시하세요.")
    if has("type mismatch") or ("형식이 일치" in m) or ("형식이 맞지" in m):
        return ("값의 형식이 맞지 않습니다(숫자 자리에 글자가 있는 등).",
                "어느 열이 숫자/날짜인지 알려주거나 \"빈칸·문자는 0으로 처리\"처럼 예외 규칙을 함께 적어주세요.")
    if ("변경된 셀이 없" in m) or ("매칭" in m and "없" in m) or ("바뀐" in m and "없" in m):
        return ("코드는 실행됐지만 조건에 맞는 대상이 없어 아무것도 바뀌지 않았습니다.",
                "대상 시트/열/조건이 실제 데이터와 맞는지 확인하세요. 특히 이름 표기(회사명/항목명 등)가 입력·출력에서 다르면 @컬럼/@범위로 정확히 지정하거나 \"비슷한 이름도 매칭해줘\"라고 요청하세요.")
    if ("문법 오류" in m) or has("compile error", "syntax error") or ("컴파일" in m and "오류" in m):
        return ("생성된 VBA 코드에 문법 오류가 있습니다(생성 문제).",
                "한 번 더 생성하거나 작업을 더 단순하게 나눠 요청하세요.")
    if has("1004", "application-defined", "object-defined") or ("개체에서 정의" in m):
        return ("Excel이 그 작업을 거부했습니다(잘못된 범위·시트, 보호, 병합 등 1004 오류).",
                "대상 범위/시트를 @범위·@시트로 명확히 지정하세요. 열/행 삽입·삭제는 \"J열 앞에 한 열\", \"5행 위에 한 행\"처럼 전체 열/행 기준으로 요청하면 안전합니다.")
    return ("작업 중 예기치 못한 오류가 났습니다.",
            "요청을 더 구체적으로 적어주세요 — 대상 파일/시트/열/범위를 @파일·@시트·@컬럼·@범위로 지정하고, 한 번에 한 작업씩 나누면 정확도가 올라갑니다.")


def _extract_pptx_slide_texts(pptx_path):
    """pptx(zip) 에서 슬라이드별 텍스트를 뽑는다(python-pptx 없이 XML 직접). 캡션/제목 보조용."""
    import zipfile, re as _re
    out = {}
    try:
        z = zipfile.ZipFile(pptx_path)
        slides = [n for n in z.namelist() if _re.match(r"ppt/slides/slide\d+\.xml$", n)]
        def snum(n):
            m = _re.search(r"(\d+)", n)
            return int(m.group(1)) if m else 0
        for n in sorted(slides, key=snum):
            xml = z.read(n).decode("utf-8", "ignore")
            texts = _re.findall(r"<a:t>(.*?)</a:t>", xml, _re.S)
            txt = " ".join(t.strip() for t in texts if t.strip())
            out[snum(n)] = txt[:2000]
    except Exception:
        pass
    return out


def render_pptx_to_slides_b64(pptx_path, max_slides=40):
    """PowerPoint COM 으로 슬라이드를 PNG 로 렌더해 base64 로 돌려준다. Excel STA 와 섞이지 않도록
    전용 스레드에서 자체 CoInitialize 로 실행한다(요청마다 새 스레드라 격리 안전)."""
    import threading
    result = {"slides": None, "total": 0, "error": None}

    def worker():
        import os as _os, tempfile as _tf, base64 as _b64, shutil as _sh
        try:
            pythoncom.CoInitialize()
        except Exception as e:
            result["error"] = "CoInitialize 실패: %s" % e
            return
        ppt = None
        pres = None
        tmpdir = _tf.mkdtemp(prefix="b2b_ppt_")
        try:
            ppt = win32com.client.Dispatch("PowerPoint.Application")
            try:
                pres = ppt.Presentations.Open(pptx_path, ReadOnly=True, WithWindow=False)
            except Exception:
                try:
                    ppt.Visible = True
                except Exception:
                    pass
                pres = ppt.Presentations.Open(pptx_path, ReadOnly=True)
            total = int(pres.Slides.Count)
            result["total"] = total
            texts = _extract_pptx_slide_texts(pptx_path)
            slides = []
            for i in range(1, min(int(max_slides), total) + 1):
                png = _os.path.join(tmpdir, "s%03d.png" % i)
                pres.Slides(i).Export(png, "PNG", 1600, 900)
                with open(png, "rb") as fh:
                    slides.append({"index": i, "mime": "image/png",
                                   "imageB64": _b64.b64encode(fh.read()).decode(),
                                   "text": texts.get(i, "")})
            result["slides"] = slides
        except Exception as e:
            result["error"] = str(e)
        finally:
            try:
                if pres is not None:
                    pres.Close()
            except Exception:
                pass
            try:
                if ppt is not None:
                    ppt.Quit()
            except Exception:
                pass
            try:
                _sh.rmtree(tmpdir, ignore_errors=True)
            except Exception:
                pass
            try:
                pythoncom.CoUninitialize()
            except Exception:
                pass

    t = threading.Thread(target=worker, daemon=True)
    t.start()
    t.join(timeout=180)
    if t.is_alive():
        return {"slides": None, "total": 0, "error": "PowerPoint 렌더가 시간 내 끝나지 않았습니다."}
    return result


class B2BHandler(http.server.SimpleHTTPRequestHandler):
    def log_message(self, format, *args):
        if os.environ.get("B2B_LOG_REQUESTS") == "1":
            super().log_message(format, *args)

    def end_headers(self):
        self.send_header("Access-Control-Allow-Origin", "*")
        self.send_header("Access-Control-Allow-Methods", "GET, POST, OPTIONS")
        self.send_header("Access-Control-Allow-Headers", "authorization, content-type, api-key, x-api-key, x-b2b-vllm-base")
        self.send_header("Access-Control-Allow-Private-Network", "true")
        self.send_header("Cache-Control", "no-store")
        super().end_headers()

    def do_OPTIONS(self):
        self.send_response(200)
        self.end_headers()

    def do_GET(self):
        if self.path.split("?")[0] == "/api/backend/health":
            app_dir = app_base_dir()
            def file_info(relative_path):
                path = app_dir / relative_path
                if not path.exists():
                    return None
                return {
                    "path": str(path),
                    "mtime": path.stat().st_mtime,
                }
            self.send_json({
                "ok": True,
                "mode": "python-backend-workbooks",
                "buildStamp": APP_BUILD_STAMP,
                "pid": os.getpid(),
                "cwd": os.getcwd(),
                "serverFile": str(Path(__file__).resolve()),
                "appDir": str(app_dir),
                "openpyxl": bool(openpyxl),
                "excelCom": bool(excel_available()),
                "excel": _health_excel_diagnostics() if excel_available() else None,
                "maintenance": _maintenance_status(),
                "runtime": {
                    "counts": _runtime_counts_snapshot(),
                    "pipelineJobs": _pipeline_job_stats(),
                    "snapshots": _pipeline_snapshot_stats(),
                    "queueSize": _excel_queue_size(),
                },
                "node": bool(node_executable()),
                "nodePath": node_executable(),
                "files": {
                    "index.html": file_info("index.html"),
                    "scripts/config.js": file_info("scripts/config.js"),
                    "scripts/excel-viewer.js": file_info("scripts/excel-viewer.js"),
                    "scripts/excel-mirror.js": file_info("scripts/excel-mirror.js"),
                    "scripts/mentions.js": file_info("scripts/mentions.js"),
                    "scripts/backend-workbooks.js": file_info("scripts/backend-workbooks.js"),
                    "scripts/backend-pipeline-worker.js": file_info("scripts/backend-pipeline-worker.js"),
                },
            })
            return
        if self.path.split("?")[0] == "/api/logic/backup-dir":
            self.send_json(logic_backup_dir_info())
            return
        if self.path.startswith("/api/workbooks/download/"):
            self.handle_backend_download()
            return
        if self.path.startswith("/api/workbooks/source/"):
            self.handle_workbook_source_download()
            return
        if self.path.startswith("/api/pipeline/status/"):
            self.handle_pipeline_status()
            return
        if self.path.startswith("/api/excel/pipeline-progress"):
            self.handle_pipeline_progress()
            return
        if self.path.startswith("/api/diff/"):
            self.handle_cached_diff()
            return
        if self.path == "/api/app/version":
            # [버전 확인] 지금 AX-Cell 의 버전(exe 파일 버전). 최신 버전은 클라가 기존 /v1 프록시로
            # 버전 서버에 물어본다 — 여기서 외부로 나가지 않는다.
            self.send_json({"ok": True, **_current_app_version()})
            return
        if self.path.startswith("/v1/"):
            self.proxy()
            return
        super().do_GET()

    def do_POST(self):
        if self.path.startswith("/api/workbooks/upload"):
            self.handle_workbook_upload()
            return
        if self.path.startswith("/api/assist/attachment"):
            self.handle_assist_attachment()
            return
        if self.path == "/api/workbooks/archive":
            self.handle_workbook_archive()
            return
        if self.path == "/api/workbooks/reinspect":
            self.handle_workbook_reinspect()
            return
        if self.path == "/api/logic/backup":
            self.handle_logic_backup()
            return
        if self.path == "/api/logic/backup-dir/select":
            try:
                self.send_json(choose_logic_backup_dir_dialog())
            except Exception as err:
                self.send_json({"ok": False, "error": str(err)}, status=500)
            return
        if self.path == "/api/logic/backup-dir/reset":
            try:
                self.send_json(reset_logic_backup_dir())
            except Exception as err:
                self.send_json({"ok": False, "error": str(err)}, status=500)
            return
        if self.path == "/api/pipeline/run":
            self.handle_backend_pipeline_run()
            return
        if self.path == "/api/pipeline/start":
            self.handle_backend_pipeline_start()
            return
        if self.path == "/api/pipeline/cancel":
            self.handle_pipeline_cancel()
            return
        if self.path == "/api/client/trace":
            self.handle_client_trace()
            return
        if self.path == "/api/excel/open":
            self.handle_excel_open()
            return
        if self.path == "/api/excel/open-result":
            self.handle_excel_open_result()
            return
        if self.path == "/api/excel/replace":
            self.handle_excel_replace()
            return
        if self.path == "/api/excel/activate":
            self.handle_excel_activate()
            return
        if self.path == "/api/excel/position":
            self.handle_excel_position()
            return
        if self.path == "/api/excel/show-only":
            self.handle_excel_show_only()
            return
        if self.path == "/api/excel/recover":
            self.handle_excel_recover()
            return
        if self.path == "/api/excel/raise":
            self.handle_excel_raise()
            return
        if self.path == "/api/excel/hide":
            self.handle_excel_hide()
            return
        if self.path == "/api/excel/hide-all":
            self.send_json(hide_all_excel_sessions())
            return
        if self.path == "/api/excel/host-state":
            payload = self.read_json_body()
            HOST_MINIMIZED["v"] = bool(payload.get("minimized"))
            _vba_trace("excel.host_state", minimized=HOST_MINIMIZED["v"])
            if HOST_MINIMIZED["v"]:
                # 최소화 통지와 동시에 일괄 숨김(별도 hide-all 호출과 중복돼도 무해).
                try:
                    self.send_json(hide_all_excel_sessions())
                except Exception:
                    self.send_json({"ok": True})
            else:
                self.send_json({"ok": True})
            return
        if self.path == "/api/excel/hide-inactive":
            self.send_json(hide_inactive_excel_sessions())
            return
        if self.path == "/api/excel/save":
            self.handle_excel_save()
            return
        if self.path == "/api/excel/changes":
            self.handle_excel_changes()
            return
        if self.path == "/api/excel/selection":
            self.handle_excel_selection()
            return
        if self.path == "/api/excel/run-vba":
            self.handle_excel_run_vba()
            return
        if self.path == "/api/excel/run-python":
            self.handle_excel_run_python()
            return
        if self.path == "/api/excel/verify-step":
            self.handle_excel_verify_step()
            return
        if self.path == "/api/excel/record/start":
            self.handle_excel_record_start()
            return
        if self.path == "/api/excel/record/stop":
            self.handle_excel_record_stop()
            return
        if self.path == "/api/diag/recent-trace":
            self.handle_diag_recent_trace()
            return
        if self.path == "/api/excel/preview-schema":
            self.handle_excel_preview_schema()
            return
        if self.path == "/api/excel/record/status":
            self.handle_excel_record_status()
            return
        if self.path == "/api/excel/record/verify":
            self.handle_excel_record_verify()
            return
        if self.path == "/api/excel/runner-mode":
            self.handle_excel_runner_mode()
            return
        if self.path == "/api/skill/consolidate":
            self.handle_skill_consolidate()
            return
        if self.path == "/api/excel/run-vba-pipeline":
            self.handle_excel_run_vba_pipeline()
            return
        if self.path == "/api/pipeline/live-final-snapshot":
            self.handle_pipeline_live_final_snapshot()
            return
        if self.path == "/api/excel/run-full-pipeline":
            self.handle_excel_run_full_pipeline()
            return
        if self.path == "/api/excel/capture-copypaste":
            self.handle_excel_capture_copypaste()
            return
        if self.path == "/api/excel/hover-info":
            self.handle_excel_hover_info()
            return
        if self.path == "/api/excel/close":
            self.handle_excel_close()
            return
        if self.path == "/api/excel/close-all-async":
            self.send_json({"ok": True})
            threading.Thread(target=cleanup_excel_sessions, name="b2b-excel-close-all", daemon=True).start()
            return
        if self.path == "/api/app/shutdown":
            # [빠른 종료] 종료 버튼은 '초기화'와 같은 강제 경로를 쓴다 — graceful wb.Close 는
            # 대형 파일에서 건당 수 초 + COM 큐 대기라 창 닫기가 수십 초 굳어 보였다. 세션은
            # 작업복사본 + SaveChanges:=False 라 강제 kill 로 잃는 데이터가 없다('초기화'와 동일 근거).
            # wait=True 로 kill 완료 후 응답해야 호스트의 서버 kill 과 경합하지 않는다(고아 방지).
            # 응답 뒤에는 os._exit 로 즉시 내려간다 — atexit 의 graceful 재정리(cleanup_excel_sessions)가
            # 이미 죽인 Excel 을 상대로 COM 대기를 다시 시작하는 것을 막는다.
            try:
                _force_restart_excel_sessions_direct(wait=True)
            except Exception:
                pass
            # 직전에 '작업 중단 2단계'/초기화가 띄운 비동기 kill 이 아직 돌고 있으면 그것도 기다린다.
            # 그 스레드가 이미 SPAWNED/세션을 비운 뒤라 위 wait=True 는 빈 스냅샷만 보고 즉시 끝나는데,
            # 그대로 os._exit 하면 아직 taskkill 이 안 나간 EXCEL.EXE 가 고아로 남았다.
            try:
                _join_inflight_kills(8)
            except Exception:
                pass
            try:
                cleanup_node_worker()
            except Exception:
                pass
            try:
                cleanup_backend_runtime_files()
            except Exception:
                pass
            self.send_json({"ok": True})

            def _exit_soon():
                time.sleep(0.5)  # 응답이 소켓으로 전달될 여유
                # [등록 중 pid 경합] 위 정리는 SPAWNED_EXCEL_PIDS 스냅샷 시점 기준이다. 그 직후
                # 완료된 DispatchEx(전체실행 격리 인스턴스/companion)가 _track_spawned_excel_app 으로
                # '이미 비운' 집합에 pid 를 넣으면, kill 대상에서 빠진 채 os._exit → 숨김 EXCEL.EXE 가
                # 영구 고아로 남는다(다음 실행은 새 프로세스라 이전 pid 를 모른다). 나가기 직전 한 번 더 쓸어담는다.
                try:
                    stragglers = {int(p) for p in list(SPAWNED_EXCEL_PIDS) if p}
                    for pid in stragglers:
                        if _is_pid_alive(pid):
                            _perf_trace("shutdown.straggler_kill", pid=int(pid))
                            _force_kill_pid(pid)
                except Exception:
                    pass
                os._exit(0)

            threading.Thread(target=_exit_soon, name="b2b-app-shutdown-exit", daemon=True).start()
            return
        if self.path == "/api/excel/force-restart":
            # COM 큐 '우회' 응급 복구(공유 Excel 행/모달 고착 시). excel_call 을 쓰지 않는다.
            try:
                self.send_json(_force_restart_excel_sessions_direct())
            except Exception as err:
                self.send_json({"ok": False, "error": str(err)}, status=500)
            return
        if self.path == "/api/excel/close-all":
            cleanup_excel_sessions()
            self.send_json({"ok": True})
            return
        if self.path == "/api/excel/diagnostics":
            try:
                self.send_json({
                    "ok": True,
                    "excel": _excel_runtime_diagnostics(reap=bool(self.read_json_body().get("reap"))),
                    "maintenance": _maintenance_status(),
                    "runtime": {
                        "counts": _runtime_counts_snapshot(),
                        "pipelineJobs": _pipeline_job_stats(),
                        "snapshots": _pipeline_snapshot_stats(),
                        "queueSize": _excel_queue_size(),
                    },
                })
            except Exception as err:
                self.send_json({"ok": False, "error": str(err)}, status=500)
            return
        if self.path == "/api/diff/current-view":
            self.handle_current_view_diff()
            return
        if self.path.startswith("/v1/"):
            self.proxy()
            return
        self.send_error(404)

    def read_json_body(self):
        length = int(self.headers.get("content-length") or 0)
        if not length:
            return {}
        raw = self.rfile.read(length)
        return json.loads(raw.decode("utf-8"))

    def send_json(self, payload, status=200):
        body = json.dumps(payload, ensure_ascii=False).encode("utf-8")
        try:
            self.send_response(status)
            self.send_header("content-type", "application/json; charset=utf-8")
            self.send_header("content-length", str(len(body)))
            self.end_headers()
            self.wfile.write(body)
        except (BrokenPipeError, ConnectionResetError, ConnectionAbortedError, OSError):
            # NativeHost/WebView can close before cleanup responses are flushed.
            # Treat this as a normal shutdown/reset race instead of printing errors.
            return

    def handle_client_trace(self):
        try:
            payload = self.read_json_body()
            event = re.sub(r"[^a-zA-Z0-9_.:-]+", "_", str(payload.get("event") or "unknown"))[:120]
            fields = payload.get("fields") if isinstance(payload.get("fields"), dict) else {}
            clean = {}
            for key, value in fields.items():
                key = re.sub(r"[^a-zA-Z0-9_.:-]+", "_", str(key))[:80]
                if isinstance(value, (str, int, float, bool)) or value is None:
                    clean[key] = value if not isinstance(value, str) else value[:500]
                else:
                    clean[key] = json.dumps(value, ensure_ascii=False, default=str)[:500]
            _perf_trace("client." + event, **clean)
            self.send_json({"ok": True})
        except Exception as err:
            self.send_json({"ok": False, "error": str(err)}, status=500)

    def handle_current_view_diff(self):
        try:
            payload = self.read_json_body()
            before = payload.get("before") or {}
            after = payload.get("after") or {}
            before_cells = {
                f"{cell.get('r')}:{cell.get('c')}": cell.get("value")
                for cell in before.get("cells", [])
            }
            changes = []
            for cell in after.get("cells", []):
                key = f"{cell.get('r')}:{cell.get('c')}"
                value = cell.get("value")
                if before_cells.get(key) != value:
                    changes.append({
                        "r": cell.get("r"),
                        "c": cell.get("c"),
                        "value": value,
                    })
            self.send_json({
                "ok": True,
                "fileId": after.get("fileId"),
                "sheet": after.get("sheet"),
                "changes": changes,
            })
        except Exception as err:
            self.send_json({"ok": False, "error": str(err)}, status=400)

    def handle_cached_diff(self):
        diff_id = self.path.rsplit("/", 1)[-1]
        diff = DIFFS.get(diff_id)
        if not diff:
            self.send_json({"ok": False, "error": "diff not found"}, status=404)
            return
        self.send_json({"ok": True, **diff})

    def handle_workbook_upload(self):
        parsed = urlparse(self.path)
        qs = parse_qs(parsed.query)
        raw_name = qs.get("name", ["workbook.xlsx"])[0]
        name = Path(unquote(raw_name)).name or "workbook.xlsx"
        if openpyxl is None and not is_csv_path(name) and not excel_available():
            self.send_json({"ok": False, "error": "openpyxl or Microsoft Excel COM is required"}, status=500)
            return
        length = int(self.headers.get("content-length") or 0)
        if length <= 0:
            self.send_json({"ok": False, "error": "empty upload"}, status=400)
            return
        BACKEND_DIR.mkdir(parents=True, exist_ok=True)
        workbook_id = uuid.uuid4().hex
        path = BACKEND_DIR / f"{workbook_id}_{name}"
        # [업로드 계측 0.7.2.1] 현장(VM/저사양 PC)에서 '업로드가 느리다'는 제보가 오면 어디서
        # 시간을 쓰는지 로그만 보고 알 수 있어야 한다. 여기 계측이 없어 매번 추측해야 했다.
        # 저장 위치: %LOCALAPPDATA%\B2B_logs\vba_pipeline_trace.jsonl
        _t_write0 = time.perf_counter()
        with path.open("wb") as f:
            remaining = length
            while remaining > 0:
                chunk = self.rfile.read(min(1024 * 1024, remaining))
                if not chunk:
                    break
                f.write(chunk)
                remaining -= len(chunk)
        _t_write = (time.perf_counter() - _t_write0) * 1000
        _t_inspect0 = time.perf_counter()
        meta = inspect_workbook(path)
        _t_inspect = (time.perf_counter() - _t_inspect0) * 1000
        try:
            _sheets = (meta or {}).get("sheets") or {}
            _vba_trace(
                "upload.done",
                name=name,
                sizeMB=round(length / (1024 * 1024), 2),
                writeMs=round(_t_write),
                inspectMs=round(_t_inspect),
                totalMs=round(_t_write + _t_inspect),
                sheets=len(_sheets),
                formulaCells=sum(len((s or {}).get("formulas") or {}) for s in _sheets.values()),
            )
        except Exception:
            pass
        WORKBOOKS[workbook_id] = {
            "id": workbook_id,
            "name": name,
            "path": str(path),
            "created": time.time(),
            "aoa_cache": None,
            "current_aoa_cache": None,
            "aoa_cache_created": None,
            "aoa_cache_hits": 0,
        }
        self.send_json({"ok": True, "workbookId": workbook_id, "name": name, "meta": meta})

    def handle_assist_attachment(self):
        """[AI 도움 첨부] 첨부 파일을 슬라이드/이미지 base64 로 돌려준다.
        - pptx/ppt: PowerPoint COM 으로 각 슬라이드를 PNG 렌더 + 슬라이드 텍스트.
        - 이미지: 그대로 base64 1장.
        클라이언트가 이 이미지를 비전 모델(dev vLLM 등) content 로 실어 프롬프트를 생성한다."""
        qs = parse_qs(urlparse(self.path).query)
        name = Path(unquote(qs.get("name", ["file"])[0])).name or "file"
        length = int(self.headers.get("content-length") or 0)
        if length <= 0:
            self.send_json({"ok": False, "error": "빈 업로드"}, status=400)
            return
        try:
            max_slides = int(qs.get("maxSlides", ["40"])[0])
        except Exception:
            max_slides = 40
        BACKEND_DIR.mkdir(parents=True, exist_ok=True)
        tmp = BACKEND_DIR / ("attach_%s_%s" % (uuid.uuid4().hex, name))
        with tmp.open("wb") as f:
            remaining = length
            while remaining > 0:
                chunk = self.rfile.read(min(1024 * 1024, remaining))
                if not chunk:
                    break
                f.write(chunk)
                remaining -= len(chunk)
        ext = name.lower().rsplit(".", 1)[-1] if "." in name else ""
        try:
            if ext in ("pptx", "ppt"):
                if win32com is None:
                    self.send_json({"ok": False, "error": "PowerPoint COM(win32com)이 없어 PPT 를 렌더할 수 없습니다."}, status=500)
                    return
                r = render_pptx_to_slides_b64(str(tmp), max_slides=max_slides)
                if r.get("error") or not r.get("slides"):
                    self.send_json({"ok": False, "error": r.get("error") or "슬라이드를 렌더하지 못했습니다."}, status=500)
                    return
                self.send_json({"ok": True, "kind": "pptx", "name": name,
                                "total": r.get("total", 0), "rendered": len(r["slides"]), "slides": r["slides"]})
            elif ext in ("png", "jpg", "jpeg", "gif", "webp", "bmp"):
                import base64 as _b64
                mime = "image/jpeg" if ext in ("jpg", "jpeg") else ("image/%s" % ext)
                b = _b64.b64encode(tmp.read_bytes()).decode()
                self.send_json({"ok": True, "kind": "image", "name": name, "total": 1, "rendered": 1,
                                "slides": [{"index": 1, "mime": mime, "imageB64": b, "text": ""}]})
            else:
                self.send_json({"ok": False, "error": "지원하지 않는 첨부 형식입니다(현재 PPT·이미지 지원). 파일: %s" % name}, status=400)
        finally:
            try:
                tmp.unlink()
            except Exception:
                pass

    def handle_workbook_reinspect(self):
        """업로드 때 시트명을 못 읽은 워크북(meta.requiresExcel)을 다시 검사한다.

        업로드 순간 Excel 이 바쁘거나 spawn 이 실패하면 inspect_workbook 의 COM 재시도(2회)가
        모두 실패해 폴백이 '파일명'을 시트명으로 지어낸다. 그 가짜 목록으로 매핑하면 스킬의
        올바른 시트명이 파일명으로 치환돼 step1 부터 '시트를 찾을 수 없음'이 난다(보안문서 실측).
        실행기 매핑 화면을 열 때 한 번 더 시도하면 Excel 이 한가해져 진짜 이름을 얻는 경우가 많다.
        (DRM/AIP 처럼 끝내 못 여는 파일은 여전히 실패 → 클라가 '스킬 기본값'으로 안전 폴백한다.)
        """
        payload = self.read_json_body() or {}
        workbook_id = str(payload.get("workbookId") or "").strip()
        rec = WORKBOOKS.get(workbook_id)
        if not rec:
            self.send_json({"ok": False, "error": "unknown workbookId"}, status=404)
            return
        path = Path(rec.get("path") or "")
        if not path.exists():
            self.send_json({"ok": False, "error": "workbook file is missing"}, status=404)
            return
        try:
            meta = inspect_workbook(path)
        except Exception as err:
            self.send_json({"ok": False, "error": str(err)}, status=500)
            return
        self.send_json({"ok": True, "workbookId": workbook_id, "name": rec.get("name"), "meta": meta})


    def handle_backend_pipeline_run(self):
        if openpyxl is None:
            self.send_json({"ok": False, "error": "openpyxl is not available"}, status=500)
            return
        if not node_executable():
            self.send_json({"ok": False, "error": "node runtime is not available"}, status=500)
            return
        payload = self.read_json_body()
        try:
            self.send_json(run_backend_pipeline_payload(payload))
        except PipelineExecutionError as err:
            self.send_json({"ok": False, "error": str(err), "errorInfo": err.info}, status=400)
        except Exception as err:
            self.send_json({"ok": False, "error": str(err)}, status=400)

    def handle_backend_pipeline_start(self):
        payload = self.read_json_body()
        job_id = uuid.uuid4().hex
        total_steps = len([s for s in payload.get("pipeline", []) if not (s and s.get("enabled") is False)])
        update_pipeline_job(job_id, {
            "ok": True,
            "jobId": job_id,
            "status": "running",
            "stage": "준비 중",
            "currentStep": 0,
            "totalSteps": total_steps,
            "created": time.time(),
        })

        def worker():
            try:
                result = run_backend_pipeline_payload(payload, job_id=job_id)
                result.update({"ok": True, "jobId": job_id, "status": "done", "stage": "완료"})
                update_pipeline_job(job_id, result)
            except PipelineExecutionError as err:
                update_pipeline_job(job_id, {
                    "ok": False,
                    "jobId": job_id,
                    "status": "error",
                    "stage": "오류",
                    "error": str(err),
                    "errorInfo": err.info,
                })
            except Exception as err:
                update_pipeline_job(job_id, {
                    "ok": False,
                    "jobId": job_id,
                    "status": "error",
                    "stage": "오류",
                    "error": str(err),
                })

        threading.Thread(target=worker, name=f"b2b-pipeline-{job_id[:8]}", daemon=True).start()
        self.send_json({"ok": True, "jobId": job_id, "status": "running"})

    def handle_pipeline_cancel(self):
        # 협조적 취소: 실행 중인 스텝은 끝까지 돌지만, 다음 스텝 시작 시점에 중단된다.
        payload = self.read_json_body()
        job_id = str(payload.get("jobId") or "")
        with PIPELINE_JOBS_LOCK:
            job = PIPELINE_JOBS.get(job_id)
            if job is not None:
                job["cancelRequested"] = True
                job["stage"] = "중단 요청됨"
        self.send_json({"ok": True, "jobId": job_id, "cancelRequested": job is not None})

    def handle_pipeline_progress(self):
        # 전체실행(격리 batch) 진행률 — 락 없이 PIPELINE_PROGRESS 만 읽어 즉시 응답(배치가 EXCEL_LOCK 쥔 동안에도 OK).
        try:
            qs = parse_qs(urlparse(self.path).query)
            excel_id = (qs.get("excelId") or [""])[0]
        except Exception:
            excel_id = ""
        prog = PIPELINE_PROGRESS.get(excel_id) or {}
        self.send_json({
            "ok": True,
            "current": int(prog.get("current") or 0),
            "total": int(prog.get("total") or 0),
            "phase": str(prog.get("phase") or "running"),
            "syncCurrent": int(prog.get("syncCurrent") or 0),
            "syncTotal": int(prog.get("syncTotal") or 0),
        })

    def handle_pipeline_status(self):
        job_id = self.path.rsplit("/", 1)[-1]
        with PIPELINE_JOBS_LOCK:
            prune_pipeline_jobs_locked()
            job = dict(PIPELINE_JOBS.get(job_id) or {})
        if not job:
            self.send_json({"ok": False, "error": "pipeline job not found"}, status=404)
            return
        self.send_json(job)

    def handle_backend_download(self):
        result_id = self.path.rsplit("/", 1)[-1]
        path = ensure_result_file(result_id)
        if not path:
            self.send_error(404)
            return
        data = path.read_bytes()
        self.send_response(200)
        if path.suffix.lower() == ".csv":
            self.send_header("content-type", "text/csv; charset=utf-8")
        else:
            self.send_header("content-type", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        self.send_header("content-disposition", content_disposition_attachment(path.name))
        self.send_header("content-length", str(len(data)))
        self.end_headers()
        self.wfile.write(data)

    def handle_workbook_source_download(self):
        workbook_id = self.path.rsplit("/", 1)[-1]
        wb = recover_workbook_record(workbook_id)
        if not wb:
            self.send_error(404)
            return
        path = Path(wb["path"])
        if not path.exists():
            self.send_error(404)
            return
        data = path.read_bytes()
        self.send_response(200)
        if path.suffix.lower() == ".csv":
            self.send_header("content-type", "text/csv; charset=utf-8")
        else:
            self.send_header("content-type", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        self.send_header("content-disposition", content_disposition_attachment(wb.get("name") or path.name))
        self.send_header("content-length", str(len(data)))
        self.end_headers()
        self.wfile.write(data)

    def handle_workbook_archive(self):
        archive_path = None
        try:
            payload = self.read_json_body()
            archive_path, filename = build_workbook_archive(payload)
            self.send_response(200)
            self.send_header("content-type", "application/zip")
            self.send_header("content-disposition", content_disposition_attachment(filename))
            self.send_header("content-length", str(archive_path.stat().st_size))
            self.end_headers()
            with archive_path.open("rb") as f:
                shutil.copyfileobj(f, self.wfile)
        except Exception as err:
            self.send_json({"ok": False, "error": str(err)}, status=400)
        finally:
            if archive_path:
                try:
                    archive_path.unlink()
                except Exception:
                    pass

    def handle_logic_backup(self):
        try:
            length = int(self.headers.get("content-length") or 0)
            if length <= 0:
                raise ValueError("backup payload is empty")
            raw = self.rfile.read(length)
            # 클라가 한글 파일명을 percent-encode 해서 보낸다(헤더는 latin-1 만 가능) → 복원.
            # ASCII(미인코딩) 값은 unquote 가 그대로 두므로 구버전 클라와도 호환.
            filename = safe_archive_filename(
                unquote(self.headers.get("x-filename") or ""),
                f"logic_auto_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.zip",
            )
            if not filename.lower().endswith(".zip"):
                filename += ".zip"
            target_dir = logic_backup_dir()
            target_dir.mkdir(parents=True, exist_ok=True)
            target_path = target_dir / filename
            if target_path.exists():
                stem = target_path.stem
                suffix = target_path.suffix
                target_path = target_dir / f"{stem}_{uuid.uuid4().hex[:8]}{suffix}"
            target_path.write_bytes(raw)
            cleanup_logic_backups(target_dir)
            self.send_json({
                "ok": True,
                "name": target_path.name,
                "path": str(target_path),
                "reason": self.headers.get("x-backup-reason") or "",
            })
        except Exception as err:
            self.send_json({"ok": False, "error": str(err)}, status=400)

    def handle_excel_open(self):
        payload = self.read_json_body()
        workbook_id = payload.get("workbookId")
        wb_record = recover_workbook_record(workbook_id)
        if not wb_record:
            self.send_json({"ok": False, "error": "workbook not found"}, status=404)
            return
        try:
            self.send_json(self._hide_if_host_minimized(open_excel_session(
                Path(wb_record["path"]),
                name=wb_record.get("name"),
                workbook_id=workbook_id,
                read_only_mirror=bool(payload.get("readOnlyMirror")),
                live_editable=bool(payload.get("liveEditable")),
                left=payload.get("left"),
                top=payload.get("top"),
                width=payload.get("width"),
                height=payload.get("height"),
                client_left=payload.get("clientLeft"),
                client_top=payload.get("clientTop"),
                client_width=payload.get("clientWidth"),
                client_height=payload.get("clientHeight"),
                viewport_width=payload.get("viewportWidth"),
                viewport_height=payload.get("viewportHeight"),
                browser_title=payload.get("browserTitle"),
                native_parent_hwnd=payload.get("nativeParentHwnd"),
                native_host_hwnd=payload.get("nativeHostHwnd"),
                native_overlay=bool(payload.get("nativeOverlay")),
                defer_visible=bool(payload.get("deferVisible")),
                # [새로고침 즉시복원] 있으면 원본 대신 '스킬 적용 끝난 사본'으로 연다(없으면 원본).
                from_state_sig=payload.get("fromStateSig"),
            )))
        except Exception as err:
            self.send_json({"ok": False, "error": str(err)}, status=500)

    def _hide_if_host_minimized(self, data):
        """열기(수 초 소요)가 최소화 '이후'에 끝나면 hide-all 을 이미 지나쳐 창이 화면에 남는다
        (업로드 → 곧바로 최소화 실측). 열기 완료 시점에 재확인해 즉시 숨긴다."""
        try:
            if HOST_MINIMIZED["v"] and isinstance(data, dict) and data.get("excelId"):
                hide_excel_session(data.get("excelId"))
                data["hiddenByHostMinimized"] = True
                _vba_trace("excel.open.hidden_host_minimized", excelId=data.get("excelId"))
        except Exception:
            pass
        return data

    def handle_excel_open_result(self):
        payload = self.read_json_body()
        result_id = payload.get("resultId") or str(payload.get("downloadUrl") or "").rstrip("/").rsplit("/", 1)[-1]
        path = ensure_result_file(result_id)
        if not path:
            self.send_json({"ok": False, "error": "result not found"}, status=404)
            return
        try:
            self.send_json(self._hide_if_host_minimized(open_excel_session(
                path,
                name=path.name,
                result_id=result_id,
                read_only_mirror=bool(payload.get("readOnlyMirror")),
                live_editable=bool(payload.get("liveEditable")),
                left=payload.get("left"),
                top=payload.get("top"),
                width=payload.get("width"),
                height=payload.get("height"),
                client_left=payload.get("clientLeft"),
                client_top=payload.get("clientTop"),
                client_width=payload.get("clientWidth"),
                client_height=payload.get("clientHeight"),
                viewport_width=payload.get("viewportWidth"),
                viewport_height=payload.get("viewportHeight"),
                browser_title=payload.get("browserTitle"),
                native_parent_hwnd=payload.get("nativeParentHwnd"),
                native_host_hwnd=payload.get("nativeHostHwnd"),
                native_overlay=bool(payload.get("nativeOverlay")),
                defer_visible=bool(payload.get("deferVisible")),
                # [새로고침 즉시복원] 있으면 원본 대신 '스킬 적용 끝난 사본'으로 연다(없으면 원본).
                from_state_sig=payload.get("fromStateSig"),
            )))
        except Exception as err:
            self.send_json({"ok": False, "error": str(err)}, status=500)

    def handle_excel_replace(self):
        payload = self.read_json_body()
        result_id = payload.get("resultId") or str(payload.get("downloadUrl") or "").rstrip("/").rsplit("/", 1)[-1]
        path = ensure_result_file(result_id)
        if not path:
            self.send_json({"ok": False, "error": "result not found"}, status=404)
            return
        try:
            read_only_mirror = payload.get("readOnlyMirror") if "readOnlyMirror" in payload else None
            self.send_json(replace_excel_session_workbook(
                payload.get("excelId"),
                path,
                name=path.name,
                result_id=result_id,
                read_only_mirror=read_only_mirror,
            ))
        except Exception as err:
            self.send_json({"ok": False, "error": str(err)}, status=500)

    def handle_excel_capture_copypaste(self):
        """[복붙 캡처] 사용자가 라이브 Excel에서 방금 한 Ctrl+C/Ctrl+V 를 역추적해
        ctx.paste_copied(...) Python 스텝으로 만들어 돌려준다(프론트가 파이프라인에 추가)."""
        payload = self.read_json_body()
        # [녹화중 캡처 차단] 녹화(네이티브/파이썬 양 엔진)는 복붙까지 함께 기록한다 — 녹화 중
        # 캡처를 허용하면 같은 복붙이 캡처 스텝+녹화 VBA 로 이중 주입된다(실측: 스킬 1·2단계 중복).
        # 클라 버튼 잠금과 별개로 서버에서도 거부(구버전 JS 캐시로 버튼이 살아있어도 안전).
        # 게이트는 excel_record_status(양 엔진 정규화)로 — RECORDING_EDIT_UNLOCKED 는 정지 시
        # 재잠금 excel_call 실패로 True 가 남을 수 있어(캡처 영구 차단 위험) 쓰지 않는다.
        _rec_active = False
        try:
            _rec_active = bool(excel_record_status().get("recording"))
        except Exception:
            _rec_active = bool(NATIVE_RECORDING.get("active"))
        if _rec_active:
            _vba_trace("capture.copypaste.reject", excelId=payload.get("excelId"), reason="recording-active")
            self.send_json({"ok": False, "error": "녹화 중에는 복붙 캡처를 쓸 수 없습니다 — 녹화가 복사/붙여넣기까지 함께 기록합니다(정지 후 이용하세요)."})
            return
        try:
            values_only = str(payload.get("valuesOnly", "")).strip().lower() in ("1", "true", "yes", "on")
            result = run_capture_copypaste(payload.get("excelId"), values_only=values_only)
            self.send_json(result)
        except Exception as err:
            # 캡처 실패(복사 없음/선택 영역 문제)는 사용자 안내용이므로 200 + ok:false 로 메시지를 그대로 전달.
            _vba_trace("capture.copypaste.error", excelId=payload.get("excelId"), error=str(err))
            self.send_json({"ok": False, "error": str(err)})

    def handle_excel_run_vba(self):
        payload = self.read_json_body()
        trace_id = uuid.uuid4().hex[:10]
        code = payload.get("code") or payload.get("vba") or ""
        _vba_trace(
            "http.run_vba.request",
            traceId=trace_id,
            excelId=payload.get("excelId"),
            entry=payload.get("entry"),
            restoreWindow=payload.get("restoreWindow"),
            codeLen=len(str(code)),
            codeHash=_trace_hash(code),
            codeHead=_trace_text(code, 350),
        )
        try:
            result = run_vba_on_session(
                payload.get("excelId"),
                code,
                entry=payload.get("entry"),
                restore_window=payload.get("restoreWindow") is not False,
            )
            _vba_trace("http.run_vba.response", traceId=trace_id, ok=True, result=result)
            self.send_json(result)
        except PipelineExecutionError as err:
            _vba_trace("http.run_vba.error", traceId=trace_id, kind="PipelineExecutionError", error=str(err), errorInfo=err.info)
            self.send_json({"ok": False, "error": str(err), "errorInfo": err.info}, status=400)
        except Exception as err:
            _vba_trace("http.run_vba.error", traceId=trace_id, kind=type(err).__name__, error=str(err))
            self.send_json({"ok": False, "error": str(err)}, status=500)

    def handle_excel_run_python(self):
        payload = self.read_json_body()
        try:
            self.send_json(run_python_on_session(
                payload.get("excelId"),
                payload.get("code") or "",
                extended=bool(payload.get("extendedTimeout")),
            ))
        except Exception as err:
            self.send_json({"ok": False, "error": str(err)}, status=500)

    def handle_pipeline_live_final_snapshot(self):
        """[새로고침 즉시복원] 요청한 파일들에 '스킬 전부 적용된 최종 상태' 사본이 있는지 조회.
        요청 {workbookIds:[...], stateSig}. 응답 {ok, ready:bool, have:[workbookId...], missing:[...]}
        ready 는 '전부 있음' — 일부만 있으면 반쪽 복원이 되므로 호출부가 전체 재실행으로 폴백한다."""
        payload = self.read_json_body()
        state_sig = payload.get("stateSig")
        ids = [str(i) for i in (payload.get("workbookIds") or []) if i]
        have, missing = [], []
        for wid in ids:
            rec = WORKBOOKS.get(wid)
            if rec and _find_live_final_snapshot(rec, state_sig):
                have.append(wid)
            else:
                missing.append(wid)
        self.send_json({
            "ok": True,
            "ready": bool(state_sig) and bool(ids) and not missing,
            "have": have,
            "missing": missing,
        })

    def handle_excel_run_vba_pipeline(self):
        payload = self.read_json_body()
        reset = payload.get("reset")
        trace_id = uuid.uuid4().hex[:10]
        steps = payload.get("steps") or []
        step_preview = []
        for idx, st in enumerate(steps[:12]):
            code = (st.get("code") if isinstance(st, dict) else str(st)) or ""
            step_preview.append({
                "idx": idx,
                "stepIdx": st.get("stepIdx") if isinstance(st, dict) else None,
                "stepId": st.get("stepId") if isinstance(st, dict) else None,
                "language": st.get("language") if isinstance(st, dict) else None,
                "description": _trace_text(st.get("description") if isinstance(st, dict) else "", 160),
                "codeLen": len(str(code)),
                "codeHash": _trace_hash(code),
                "codeHead": _trace_text(code, 260),
            })
        _vba_trace(
            "http.run_vba_pipeline.request",
            traceId=trace_id,
            excelId=payload.get("excelId"),
            reset=True if reset is None else bool(reset),
            viewSheet=payload.get("viewSheet"),
            steps=len(steps),
            stepPreview=step_preview,
        )
        try:
            result = run_vba_pipeline_on_session(
                payload.get("excelId"),
                steps,
                reset=True if reset is None else bool(reset),
                entry=payload.get("entry"),
                view_sheet=payload.get("viewSheet"),
            )
            _vba_trace("http.run_vba_pipeline.response", traceId=trace_id, ok=True, result=result)
            self.send_json(result)
        except PipelineExecutionError as err:
            _vba_trace("http.run_vba_pipeline.error", traceId=trace_id, kind="PipelineExecutionError", error=str(err), errorInfo=err.info)
            self.send_json({"ok": False, "error": str(err), "errorInfo": err.info}, status=400)
        except Exception as err:
            _vba_trace("http.run_vba_pipeline.error", traceId=trace_id, kind=type(err).__name__, error=str(err))
            self.send_json({"ok": False, "error": str(err)}, status=500)

    def handle_excel_run_full_pipeline(self):
        # [0.5.15 백그라운드 전체실행] 그룹 전체를 1콜로 받아 격리 인스턴스 1개에서 처리(반복 spawn/sync 제거).
        payload = self.read_json_body()
        trace_id = uuid.uuid4().hex[:10]
        groups = payload.get("groups") or []
        reset_excel_ids = payload.get("resetExcelIds") or []
        total = sum(len((g.get("steps") or [])) for g in groups)
        _vba_trace(
            "http.run_full_pipeline.request",
            traceId=trace_id,
            anchorExcelId=(groups[0].get("excelId") if groups else None),
            groups=len(groups),
            totalSteps=total,
            resetExcelIds=list(reset_excel_ids),
        )
        try:
            result = run_full_pipeline_single_instance(
                groups,
                reset_excel_ids=reset_excel_ids,
                view_sheet=payload.get("viewSheet"),
                entry=payload.get("entry"),
                output_mode=(payload.get("outputMode") or "sync"),
                # 클라가 '원본부터 전체 적용'일 때만 보낸다(부분/이어실행이면 없음 → 사본을 남기지 않는다).
                state_sig=payload.get("stateSig"),
            )
            _vba_trace("http.run_full_pipeline.response", traceId=trace_id, ok=True, applied=result.get("applied"))
            self.send_json(result)
        except PipelineExecutionError as err:
            _vba_trace("http.run_full_pipeline.error", traceId=trace_id, kind="PipelineExecutionError", error=str(err), errorInfo=err.info)
            self.send_json({"ok": False, "error": str(err), "errorInfo": err.info}, status=400)
        except Exception as err:
            _vba_trace("http.run_full_pipeline.error", traceId=trace_id, kind=type(err).__name__, error=str(err))
            self.send_json({"ok": False, "error": str(err)}, status=500)

    def _reject_show_while_host_minimized(self):
        """호스트 최소화 중엔 표시 계열 요청을 조용히 스킵(True 반환 시 호출측은 응답 완료).
        복원 시 C#/JS 가 강제 재배치(force)를 다시 보내므로 스킵해도 상태가 어긋나지 않는다."""
        if HOST_MINIMIZED["v"]:
            self.send_json({"ok": True, "skipped": "host-minimized"})
            return True
        return False

    def handle_excel_activate(self):
        if self._reject_show_while_host_minimized():
            return
        payload = self.read_json_body()
        try:
            self.send_json(activate_excel_session(
                payload.get("excelId"),
                payload.get("sheet"),
                payload.get("range") or payload.get("address"),
            ))
        except Exception as err:
            self.send_json({"ok": False, "error": str(err)}, status=400)

    def handle_excel_position(self):
        if self._reject_show_while_host_minimized():
            return
        payload = self.read_json_body()
        try:
            self.send_json(position_excel_session(
                payload.get("excelId"),
                payload.get("left"),
                payload.get("top"),
                payload.get("width"),
                payload.get("height"),
                client_left=payload.get("clientLeft"),
                client_top=payload.get("clientTop"),
                client_width=payload.get("clientWidth"),
                client_height=payload.get("clientHeight"),
                viewport_width=payload.get("viewportWidth"),
                viewport_height=payload.get("viewportHeight"),
                browser_title=payload.get("browserTitle"),
                native_parent_hwnd=payload.get("nativeParentHwnd"),
                native_host_hwnd=payload.get("nativeHostHwnd"),
                native_overlay=bool(payload.get("nativeOverlay")),
                keep_zorder=bool(payload.get("keepZorder")),
            ))
        except Exception as err:
            self.send_json({"ok": False, "error": str(err)}, status=400)

    def handle_excel_show_only(self):
        if self._reject_show_while_host_minimized():
            return
        payload = self.read_json_body()
        try:
            self.send_json(show_only_excel_session(
                payload.get("excelId"),
                payload.get("left"),
                payload.get("top"),
                payload.get("width"),
                payload.get("height"),
                client_left=payload.get("clientLeft"),
                client_top=payload.get("clientTop"),
                client_width=payload.get("clientWidth"),
                client_height=payload.get("clientHeight"),
                viewport_width=payload.get("viewportWidth"),
                viewport_height=payload.get("viewportHeight"),
                browser_title=payload.get("browserTitle"),
                native_parent_hwnd=payload.get("nativeParentHwnd"),
                native_host_hwnd=payload.get("nativeHostHwnd"),
                native_overlay=bool(payload.get("nativeOverlay")),
                skip_position=bool(payload.get("skipPosition")),
            ))
        except Exception as err:
            self.send_json({"ok": False, "error": str(err)}, status=400)

    def handle_excel_recover(self):
        if self._reject_show_while_host_minimized():
            return
        payload = self.read_json_body()
        try:
            self.send_json(recover_excel_session(
                payload.get("excelId"),
                payload.get("left"),
                payload.get("top"),
                payload.get("width"),
                payload.get("height"),
                client_left=payload.get("clientLeft"),
                client_top=payload.get("clientTop"),
                client_width=payload.get("clientWidth"),
                client_height=payload.get("clientHeight"),
                viewport_width=payload.get("viewportWidth"),
                viewport_height=payload.get("viewportHeight"),
                browser_title=payload.get("browserTitle"),
                native_parent_hwnd=payload.get("nativeParentHwnd"),
                native_host_hwnd=payload.get("nativeHostHwnd"),
                native_overlay=bool(payload.get("nativeOverlay")),
            ))
        except Exception as err:
            self.send_json({"ok": False, "error": str(err)}, status=400)

    def handle_excel_raise(self):
        if self._reject_show_while_host_minimized():
            return
        payload = self.read_json_body()
        try:
            self.send_json(raise_excel_session(payload.get("excelId")))
        except Exception as err:
            self.send_json({"ok": False, "error": str(err)}, status=400)

    def handle_excel_hide(self):
        payload = self.read_json_body()
        try:
            self.send_json(hide_excel_session(payload.get("excelId"), light=bool(payload.get("light"))))
        except Exception as err:
            self.send_json({"ok": False, "error": str(err)}, status=400)

    def handle_excel_verify_step(self):
        # [AI 도움 Tier2] 후보 스텝 코드를 '격리 인스턴스'에서 스냅샷 위에 실행하고 diff 만 돌려준다.
        # 라이브는 절대 건드리지 않는다(불변식). 실패는 데이터로 반환 — 클라가 '미검증 카드'로 폴백한다.
        payload = self.read_json_body()
        try:
            self.send_json(verify_step_isolated(
                payload.get("resultId"),
                payload.get("code") or "",
                sheet=payload.get("sheet"),
            ))
        except Exception as err:
            self.send_json({"ok": False, "error": str(err)}, status=200)

    # ---- 녹화 (기본: 네이티브 매크로 레코더/VBA · 폴백: ixi-Cell-R recorder) ----
    def handle_excel_record_start(self):
        payload = self.read_json_body()
        try:
            engine = str((payload or {}).get("engine") or "vba").strip().lower()
            self.send_json(excel_record_start(engine=engine))
        except Exception as err:
            self.send_json({"ok": False, "error": str(err)}, status=500)

    def handle_excel_record_stop(self):
        self.read_json_body()
        try:
            self.send_json(excel_record_stop())
        except Exception as err:
            self.send_json({"ok": False, "error": str(err)}, status=500)

    def handle_excel_record_status(self):
        self.read_json_body()
        try:
            self.send_json(excel_record_status())
        except Exception as err:
            self.send_json({"ok": False, "error": str(err)}, status=500)

    def handle_excel_preview_schema(self):
        """[AI 도움 라이브 직독] 열린 라이브 세션의 '현재' 시트/그리드(경량 60행 미리보기)를 돌려준다.
        AI 도움 data 도구는 state.inputs 캐시만 읽는데, 교차파일 붙여넣기 대상(컴패니언)은 파이프라인
        liveSchema 가 주 세션만 실어 캐시가 stale 로 남는다(실측: 정산서에 붙여넣었는데 '데이터 없음').
        도구가 조회 직전 이걸 불러 캐시를 라이브로 갱신 → '보이는 것'과 일치한다."""
        payload = self.read_json_body()
        try:
            excel_id = payload.get("excelId")
            only_sheet = str(payload.get("sheet") or "").strip()
            session = get_excel_session(excel_id)

            def _read():
                app, wb = session_workbook(session)
                # [경합 완화] AI 도움은 대개 한 시트만 묻는다. sheet 를 주면 그 시트만 읽어
                # excel_call 워커 점유 시간을 최소화한다(전 시트 UsedRange 읽기 회피 → record-start
                # 등 다른 excel_call 이 그 뒤에 줄서서 '준비 중'이 느려지던 커플링 완화). partial 표시.
                if only_sheet:
                    return _live_preview_schema(wb, only_sheet=only_sheet)
                return _live_preview_schema(wb)

            _t0 = time.perf_counter()
            schema = excel_call(_read, timeout=30)
            _ms = round((time.perf_counter() - _t0) * 1000, 1)
            _vba_trace("assist.preview_schema", excelId=excel_id, sheet=only_sheet or "(all)",
                       sheets=len((schema or {}).get("sheetNames") or []), ms=_ms,
                       partial=bool((schema or {}).get("partial")))
            self.send_json({"ok": True, "schema": schema})
        except Exception as err:
            self.send_json({"ok": False, "error": str(err)})

    def handle_diag_recent_trace(self):
        """[AI 도움 run.trace] 직전 실행의 서버 트레이스 타임라인 — 스텝이 '실제로 어느 워크북에서
        어떤 순서로 돌고 어디서 죽었는지'는 클라 상태(step.error)만으로는 알 수 없다(실측 15:30:
        1조각이 동반본에서 실행된 것은 트레이스에만 남았다). 진단용 화이트리스트 이벤트만 압축해 준다."""
        payload = self.read_json_body()
        try:
            limit = max(10, min(200, int(payload.get("limit") or 80)))
            keep = (
                "http.run_vba_pipeline.request", "http.run_full_pipeline.request",
                "pipeline.impl.start", "pipeline.isolated.target.opened",
                "pipeline.isolated.companion.opened", "pipeline.step.start",
                "pipeline.step.ok", "pipeline.step.error", "pipeline.step.activate_sheet.skip",
                "vba.macro.runtime_error", "fullrun.step.start", "fullrun.step.ok",
                "fullrun.step.error", "fullrun.file.opened", "fullrun.companion.opened",
                "http.run_vba_pipeline.error", "http.run_full_pipeline.error",
                "excel.open.reattach", "excel.open.orphan_close",
            )
            events = []
            try:
                lines = _vba_trace_path().read_text(encoding="utf-8", errors="replace").splitlines()
            except Exception:
                lines = []
            for line in lines[-4000:]:
                try:
                    d = json.loads(line)
                except Exception:
                    continue
                ev = str(d.get("event") or "")
                if ev not in keep:
                    continue
                item = {"ts": d.get("ts"), "event": ev}
                for k in ("description", "errDescription", "errNumber", "error",
                          "targetName", "companionName", "name", "reason", "ordinal", "stepId"):
                    v = d.get(k)
                    if v is not None:
                        item[k] = str(v)[:220]
                # 어느 워크북 컨텍스트에서 돌았는지(있으면 Name 만 압축)
                for k in ("workbook", "targetWorkbook", "liveWorkbook"):
                    v = d.get(k)
                    if isinstance(v, dict) and v.get("Name"):
                        item[k] = str(v.get("Name"))[:120]
                events.append(item)
            self.send_json({"ok": True, "events": events[-limit:], "totalMatched": len(events)})
        except Exception as err:
            self.send_json({"ok": False, "error": str(err)}, status=500)

    def handle_excel_record_verify(self):
        payload = self.read_json_body()
        try:
            self.send_json(excel_record_verify(payload))
        except Exception as err:
            self.send_json({"ok": False, "error": str(err)}, status=500)

    def handle_excel_runner_mode(self):
        # [2A] 실행기(runner) 전면 진입/복귀를 서버에 알린다. suppress=True 면 서버측
        # 라이브 프레임 자동 되띄움을 억제(실행기 위 엑셀 오버레이 방지), False 면 복귀.
        global LIVE_RESTORE_SUPPRESSED
        payload = self.read_json_body() or {}
        try:
            LIVE_RESTORE_SUPPRESSED = bool(payload.get("suppress"))
            self.send_json({"ok": True, "suppress": LIVE_RESTORE_SUPPRESSED})
        except Exception as err:
            self.send_json({"ok": False, "error": str(err)}, status=500)

    def handle_skill_consolidate(self):
        payload = self.read_json_body()
        base = (self.headers.get("x-b2b-vllm-base") or "").strip()
        try:
            self.send_json(skill_consolidate(payload, base))
        except Exception as err:
            # 통합은 보조 기능 — 실패해도 원본 코드를 그대로 돌려줘 재현이 깨지지 않게.
            self.send_json({"ok": True, "consolidated": False,
                            "code": (payload or {}).get("code", ""), "error": str(err)})

    def handle_excel_save(self):
        payload = self.read_json_body()
        try:
            # internal=True 는 되돌리기용 백업(사용자가 여는 파일이 아님) — 보호/화면 복구를 건너뛴다.
            self.send_json(save_excel_session(payload.get("excelId"), payload.get("name"),
                                              internal=bool(payload.get("internal"))))
        except Exception as err:
            self.send_json({"ok": False, "error": str(err)}, status=400)

    def handle_excel_changes(self):
        payload = self.read_json_body()
        try:
            self.send_json(poll_excel_session_changes(payload.get("excelId")))
        except Exception as err:
            self.send_json({"ok": False, "error": str(err)}, status=400)

    def handle_excel_selection(self):
        payload = self.read_json_body()
        try:
            self.send_json(poll_excel_session_selection(payload.get("excelId")))
        except Exception as err:
            self.send_json({"ok": False, "error": str(err)}, status=400)

    def handle_excel_hover_info(self):
        payload = self.read_json_body()
        try:
            self.send_json(get_excel_hover_info(payload.get("excelId")))
        except Exception as err:
            self.send_json({"ok": False, "error": str(err)}, status=400)

    def handle_excel_close(self):
        payload = self.read_json_body()
        try:
            self.send_json(close_excel_session(payload.get("excelId")))
        except Exception as err:
            self.send_json({"ok": False, "error": str(err)}, status=400)

    def proxy(self):
        body = None
        if self.command in {"POST", "PUT", "PATCH"}:
            length = int(self.headers.get("content-length") or 0)
            body = self.rfile.read(length) if length else None

        # 프런트 설정에서 지정한 실제 Violet/vLLM 주소가 있으면 그쪽으로 전달.
        base = VLLM_BASE
        upstream = (self.headers.get("x-b2b-vllm-base") or "").strip()
        if upstream.startswith("http://") or upstream.startswith("https://"):
            base = upstream.rstrip("/")
        target = base + self.path
        headers = {}
        for key in ("authorization", "api-key", "content-type", "accept"):
            value = self.headers.get(key)
            if value:
                headers[key] = value
        if body is not None and "content-type" not in headers:
            headers["content-type"] = "application/json"

        attempts = max(1, PROXY_RETRY_ATTEMPTS)
        last_error = None
        for attempt in range(1, attempts + 1):
            response_started = False
            req = urllib.request.Request(
                target,
                data=body,
                headers=headers,
                method=self.command,
            )
            try:
                with urllib.request.urlopen(req, timeout=300) as resp:
                    self.send_response(resp.status)
                    for key, value in resp.headers.items():
                        if key.lower() in {"connection", "transfer-encoding", "content-encoding", "content-length"}:
                            continue
                        self.send_header(key, value)
                    self.end_headers()
                    response_started = True
                    while True:
                        chunk = resp.read(8192)
                        if not chunk:
                            break
                        self.wfile.write(chunk)
                        self.wfile.flush()
                    return
            except urllib.error.HTTPError as err:
                payload = err.read()
                if err.code in {408, 429, 500, 502, 503, 504} and attempt < attempts:
                    last_error = f"HTTP {err.code}: {payload[:200]!r}"
                    time.sleep(PROXY_RETRY_BASE_DELAY * attempt)
                    continue
                self.send_response(err.code)
                self.send_header("content-type", err.headers.get("content-type", "text/plain"))
                self.end_headers()
                self.wfile.write(payload)
                return
            except (BrokenPipeError, ConnectionResetError):
                return
            except (urllib.error.URLError, TimeoutError, socket.timeout, OSError) as err:
                if response_started:
                    return
                last_error = err
                if attempt < attempts:
                    time.sleep(PROXY_RETRY_BASE_DELAY * attempt)
                    continue
                payload = f"Proxy error after {attempts} attempts: {last_error}".encode("utf-8")
                self.send_response(502)
                self.send_header("content-type", "text/plain; charset=utf-8")
                self.send_header("content-length", str(len(payload)))
                self.end_headers()
                self.wfile.write(payload)
                return
            except Exception as err:
                if response_started:
                    return
                last_error = err
                if attempt < attempts:
                    time.sleep(PROXY_RETRY_BASE_DELAY * attempt)
                    continue
                payload = f"Proxy error after {attempts} attempts: {last_error}".encode("utf-8")
                self.send_response(502)
                self.send_header("content-type", "text/plain; charset=utf-8")
                self.send_header("content-length", str(len(payload)))
                self.end_headers()
                self.wfile.write(payload)
                return


class B2BThreadingTCPServer(socketserver.ThreadingTCPServer):
    allow_reuse_address = True
    daemon_threads = True


def update_pipeline_job(job_id, patch):
    if not job_id:
        return
    with PIPELINE_JOBS_LOCK:
        prune_pipeline_jobs_locked()
        current = PIPELINE_JOBS.get(job_id, {})
        current.update(patch)
        current["updated"] = time.time()
        PIPELINE_JOBS[job_id] = current


def pipeline_job_cancel_requested(job_id):
    if not job_id:
        return False
    with PIPELINE_JOBS_LOCK:
        job = PIPELINE_JOBS.get(job_id)
    return bool(job and job.get("cancelRequested"))


def raise_if_pipeline_cancelled(job_id):
    """협조적 취소 체크포인트 — 스텝 경계에서 호출. 취소 요청이 있으면 cancelled 플래그가
    달린 PipelineExecutionError 를 던져 잡을 '사용자 중단'으로 끝낸다(프론트는 조용히 복귀)."""
    if pipeline_job_cancel_requested(job_id):
        raise PipelineExecutionError({
            "cancelled": True,
            "stepIdx": -1,
            "message": "사용자가 작업을 중단했습니다.",
        })


def prune_pipeline_jobs_locked():
    if not PIPELINE_JOBS:
        return
    now = time.time()
    stale = [
        job_id
        for job_id, job in PIPELINE_JOBS.items()
        if now - float(job.get("updated") or job.get("created") or now) > PIPELINE_JOB_TTL_SECONDS
    ]
    for job_id in stale:
        PIPELINE_JOBS.pop(job_id, None)
    if len(PIPELINE_JOBS) <= MAX_PIPELINE_JOBS:
        return
    ordered = sorted(
        PIPELINE_JOBS.items(),
        key=lambda item: float(item[1].get("updated") or item[1].get("created") or 0),
    )
    for job_id, _job in ordered[: max(0, len(PIPELINE_JOBS) - MAX_PIPELINE_JOBS)]:
        PIPELINE_JOBS.pop(job_id, None)


def content_disposition_attachment(filename):
    safe_ascii = "".join(ch if 32 <= ord(ch) < 127 and ch not in '"\\' else "_" for ch in str(filename))
    encoded = quote(str(filename), safe="")
    return f'attachment; filename="{safe_ascii}"; filename*=UTF-8\'\'{encoded}'


def archive_download_id(url):
    parsed = urlparse(str(url or ""))
    parts = [part for part in parsed.path.split("/") if part]
    if len(parts) >= 4 and parts[-4:-1] == ["api", "workbooks", "download"]:
        return unquote(parts[-1])
    return None


def archive_source_workbook_id(url):
    parsed = urlparse(str(url or ""))
    parts = [part for part in parsed.path.split("/") if part]
    if len(parts) >= 4 and parts[-4:-1] == ["api", "workbooks", "source"]:
        return unquote(parts[-1])
    return None


def safe_archive_filename(name, default_name):
    filename = Path(str(name or default_name)).name.strip() or default_name
    return filename.replace("\\", "_").replace("/", "_").replace("\x00", "_")


def cleanup_logic_backups(target_dir, keep=80):
    try:
        backups = sorted(
            [p for p in Path(target_dir).glob("*.zip") if p.is_file()],
            key=lambda p: p.stat().st_mtime,
            reverse=True,
        )
        for path in backups[keep:]:
            try:
                path.unlink()
            except Exception:
                pass
    except Exception:
        pass


def office_file_signature(path):
    try:
        with Path(path).open("rb") as f:
            return f.read(8)
    except OSError:
        return b""


def excel_zip_file_suffix(path):
    """Return the Excel extension implied by an OPC/ZIP workbook package.

    Excel can SaveCopyAs an HTML-as-.xls workbook into a binary workbook package
    while keeping the caller-provided .xls/.xlsx-looking file name. The package
    starts with PK like .xlsx, but contains xl/workbook.bin and must be opened as
    .xlsb. Treating every PK file as .xlsx causes the "file format or extension is
    not valid" failure seen when saved skills are replayed.
    """
    sig = office_file_signature(path)
    if not (sig.startswith(b"PK\x03\x04") or sig.startswith(b"PK\x05\x06") or sig.startswith(b"PK\x07\x08")):
        return None
    try:
        with zipfile.ZipFile(path) as zf:
            names = set(zf.namelist())
    except Exception:
        return None
    if "[Content_Types].xml" not in names:
        return None
    lower_names = {str(n).lower() for n in names}
    if "xl/workbook.bin" in lower_names:
        return ".xlsb"
    if "xl/workbook.xml" not in lower_names:
        return None
    if "xl/vbaproject.bin" in lower_names:
        return ".xlsm"
    return ".xlsx"


def is_ooxml_zip_file(path):
    return excel_zip_file_suffix(path) in {".xlsx", ".xlsm", ".xltx", ".xltm"}


def _excel_suffix_matches_content(suffix, content_suffix):
    suffix = str(suffix or "").lower()
    content_suffix = str(content_suffix or "").lower()
    if not suffix or not content_suffix:
        return False
    if suffix == content_suffix:
        return True
    if content_suffix == ".xlsx" and suffix == ".xltx":
        return True
    if content_suffix == ".xlsm" and suffix == ".xltm":
        return True
    return False


def is_ole_excel_file(path):
    return office_file_signature(path).startswith(b"\xD0\xCF\x11\xE0")


def sniff_text_excel_suffix(path):
    try:
        raw = Path(path).read_bytes()[:8192]
    except OSError:
        return None
    if not raw or b"\x00" in raw[:512]:
        return None
    lower = raw.lstrip().lower()
    if lower.startswith(b"<!doctype html") or lower.startswith(b"<html") or b"<table" in lower[:4096]:
        return ".html"
    for enc in ("utf-8-sig", "cp949", "euc-kr", "utf-8"):
        try:
            text = raw.decode(enc)
            break
        except UnicodeDecodeError:
            text = ""
    if not text.strip():
        return None
    first_line = text.splitlines()[0] if text.splitlines() else text
    if "\t" in first_line:
        return ".tsv"
    if any(ch in first_line for ch in (",", ";", "|")):
        return ".csv"
    return None


def excel_compatible_open_path(path):
    path = Path(path)
    suffix = path.suffix.lower()
    wanted_suffix = None
    text_suffix = sniff_text_excel_suffix(path)
    if text_suffix and suffix != text_suffix:
        wanted_suffix = text_suffix
    else:
        zip_suffix = excel_zip_file_suffix(path)
        if zip_suffix and not _excel_suffix_matches_content(suffix, zip_suffix):
            wanted_suffix = zip_suffix
    if not wanted_suffix and is_ole_excel_file(path) and suffix != ".xls":
        wanted_suffix = ".xls"
    if not wanted_suffix:
        return path, None
    BACKEND_DIR.mkdir(parents=True, exist_ok=True)
    # [시트명 안정화] 텍스트(HTML/CSV) 위장 파일을 변환해 열면 Excel 이 시트를 '파일명 stem(31자 truncate)'으로
    # 자동명명한다. 매 open 마다 random uuid 면 시트명이 매번 달라져, @멘션 캡처 시점과 VBA 실행 시점의 시트명이
    # 어긋나 "시트를 찾을 수 없음"이 났다. 파일명 앞 31자(= "excel_open_"(11) + 해시 20자)를 '원본 파일명' 기준으로
    # 고정하면 매 open 마다 시트명이 동일해진다. 뒤의 random 으로 파일 경로는 유일 → 동시 오픈 file-lock 회피.
    # (워크북명 전체는 random 때문에 매번 달라도 _alias_open_workbook_name 이 등록명으로 해석한다.)
    stable = hashlib.md5(Path(path).name.encode("utf-8", "ignore")).hexdigest()[:20]
    temp_path = BACKEND_DIR / f"excel_open_{stable}{uuid.uuid4().hex}{wanted_suffix}"
    shutil.copy2(path, temp_path)
    return temp_path, temp_path


def openpyxl_load_workbook_compatible(path, **kwargs):
    if openpyxl is None:
        raise RuntimeError("openpyxl is not available")
    path = Path(path)
    try:
        return openpyxl.load_workbook(path, **kwargs)
    except Exception as first_err:
        if not is_ooxml_zip_file(path):
            raise
        try:
            buffer = io.BytesIO(path.read_bytes())
            wb = openpyxl.load_workbook(buffer, **kwargs)
            wb._b2b_source_buffer = buffer
            return wb
        except Exception:
            raise first_err


def excel_workbooks_open(app, path, read_only=False, intended_name=None):
    open_path, temp_path = excel_compatible_open_path(path)
    try:
        app.DisplayAlerts = False
    except Exception:
        pass
    try:
        app.EnableEvents = False
    except Exception:
        pass
    try:
        app.AskToUpdateLinks = False
    except Exception:
        pass
    try:
        # [핵심 수정] msoAutomationSecurityLow(1) 로 연다(이전엔 ForceDisable(3)).
        # ForceDisable 로 파일을 열면 일부 환경(기업/특정 Office 빌드)에서 그 인스턴스의 '모든 매크로'가
        # 영구 비활성화되어, 이후 주입한 러너 매크로의 Application.Run 이 "매크로를 실행할 수 없습니다"로
        # 실패한다(전체실행이 companion 재오픈 후 100% 실패한 근본원인). run 직전에 Low 로 낮춰도
        # 이미 차단된 상태는 되돌릴 수 없으므로, '열 때부터' Low 로 연다. 업로드 파일은 .xlsx(매크로 없음)이고
        # 러너 .xlsm 은 우리가 만든 것이라 Low 는 안전하다.
        app.AutomationSecurity = 1  # msoAutomationSecurityLow
    except Exception:
        pass

    attempts = [
        {"UpdateLinks": 0, "ReadOnly": bool(read_only), "IgnoreReadOnlyRecommended": True, "Notify": False, "AddToMru": False, "Local": True, "CorruptLoad": 0},
        {"UpdateLinks": 0, "ReadOnly": bool(read_only), "IgnoreReadOnlyRecommended": True, "Notify": False, "AddToMru": False, "Local": True, "CorruptLoad": 1},
        {"UpdateLinks": 0, "ReadOnly": bool(read_only), "IgnoreReadOnlyRecommended": True, "Notify": False, "AddToMru": False, "Local": True, "CorruptLoad": 2},
        {"UpdateLinks": 0, "ReadOnly": bool(read_only), "IgnoreReadOnlyRecommended": True},
    ]
    errors = []
    for kwargs in attempts:
        try:
            wb = app.Workbooks.Open(str(open_path), **kwargs)
            if wb is None:
                raise RuntimeError(f"Workbooks.Open 이 워크북을 반환하지 않았습니다(같은 이름의 파일이 이미 열려 있을 수 있음): {open_path}")
            try:
                # 초기화/응급복구가 EXCEL.EXE 를 강제 종료(taskkill)하는 방식이므로,
                # 다음 Excel 실행에서 '문서 복구' 창이 뜨지 않도록 우리가 여는 모든 워크북을
                # AutoRecover 대상에서 제외한다. 워크북 한정 속성이라 사용자 Excel 설정(레지스트리)은
                # 건드리지 않으며, 작업복사본(폐기 대상)이라 복구 정보 자체가 무의미하다.
                wb.EnableAutoRecover = False
            except Exception:
                pass
            # [포맷 위장 별칭] 변환으로 실제 wb.Name 이 등록명(intended_name/path 의 파일명)과 달라지면
            # 등록명→실제명 별칭을 저장 → VBA 의 Workbooks("등록명") 치환에 사용. 이름이 같으면(일반 파일)
            # _stash 가 스스로 건너뛴다(회귀 0).
            try:
                _stash_workbook_name_alias(app, intended_name or path, str(wb.Name))
            except Exception:
                pass
            return wb, temp_path
        except Exception as err:
            errors.append(err)
    if temp_path:
        try:
            Path(temp_path).unlink(missing_ok=True)
        except Exception:
            pass
    raise errors[-1] if errors else RuntimeError(f"failed to open workbook: {path}")


def unique_archive_name(used, folder, filename):
    filename = safe_archive_filename(filename, "workbook.xlsx")
    base = f"{folder}/{filename}" if folder else filename
    if base not in used:
        used.add(base)
        return base
    path = Path(filename)
    stem = path.stem or "workbook"
    suffix = path.suffix
    index = 2
    while True:
        candidate = f"{folder}/{stem}_{index}{suffix}" if folder else f"{stem}_{index}{suffix}"
        if candidate not in used:
            used.add(candidate)
            return candidate
        index += 1


def resolve_archive_item(item):
    item = item or {}
    name = safe_archive_filename(item.get("name"), "workbook.xlsx")
    path = None

    excel_id = item.get("excelId")
    if excel_id:
        try:
            saved = save_excel_session(excel_id)
            path = ensure_result_file(saved.get("downloadId"))
        except Exception:
            path = None

    if path is None:
        download_id = item.get("downloadId") or archive_download_id(item.get("downloadUrl"))
        if download_id:
            path = ensure_result_file(download_id)

    if path is None:
        source_id = item.get("workbookId") or archive_source_workbook_id(item.get("downloadUrl"))
        if source_id:
            wb = WORKBOOKS.get(source_id)
            if wb:
                path = Path(wb["path"])
                name = safe_archive_filename(item.get("name") or wb.get("name"), path.name)

    if not path or not Path(path).exists():
        raise ValueError(f"download source not found: {name}")
    return Path(path), name


def build_workbook_archive(payload):
    files = payload.get("files") if isinstance(payload, dict) else None
    if not isinstance(files, list) or not files:
        raise ValueError("archive files are empty")

    raw_filename = payload.get("filename") if isinstance(payload, dict) else None
    filename = safe_archive_filename(raw_filename, f"all_files_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.zip")
    if not filename.lower().endswith(".zip"):
        filename += ".zip"

    BACKEND_DIR.mkdir(parents=True, exist_ok=True)
    archive_path = BACKEND_DIR / f"{uuid.uuid4().hex}_{filename}"
    used_names = set()
    with zipfile.ZipFile(archive_path, "w", compression=zipfile.ZIP_STORED, allowZip64=True) as zf:
        for item in files:
            src_path, display_name = resolve_archive_item(item)
            role = str((item or {}).get("role") or "").lower()
            folder = "inputs" if role == "input" else "outputs" if role == "output" else "files"
            arcname = unique_archive_name(used_names, folder, display_name)
            zf.write(src_path, arcname)
    return archive_path, filename


def normalize_text(value):
    return "".join(str(value or "").lower().split())


def normalize_sheet_lookup(value):
    # Sheet names often differ only by spaces/underscores generated by the LLM
    # ("인포콘_올인원_정렬" vs "인포콘올인원_정렬"). Keep this looser
    # matching scoped to sheet lookup only; data value matching remains stricter.
    return re.sub(r"[\s_\-]+", "", str(value or "").lower())


_EPHEMERAL_EXCEL_OPEN_SHEET_RE = re.compile(r"^excel_open_[0-9a-f]{8,}$", re.I)


def _is_ephemeral_excel_open_sheet_name(value):
    text = str(value or "").strip()
    if not text:
        return False
    return bool(_EPHEMERAL_EXCEL_OPEN_SHEET_RE.match(Path(text).stem))


def _resolve_ephemeral_excel_open_sheet_alias(requested, names):
    """Map stale excel_open_<uuid> sheet names from HTML/CSV-compatible opens.

    Some .xls files are actually HTML tables. To make Excel open them reliably we
    copy them to excel_open_<uuid>.html, and Excel then names the only sheet from
    that temporary stem. Stored Python/VBA skills can therefore contain a stale
    excel_open_<old uuid> sheet name after reload. If the workbook has exactly
    one sheet, that stale generated name is not meaningful user intent; the
    single current sheet is the safe target.
    """
    if not _is_ephemeral_excel_open_sheet_name(requested):
        return None
    clean_names = [str(n) for n in (names or []) if str(n or "").strip()]
    if len(clean_names) == 1:
        return clean_names[0]
    return None


def _days_in_month(month, year):
    if month == 2:
        return 29 if (year % 4 == 0 and (year % 100 != 0 or year % 400 == 0)) else 28
    return 31 if month in (1, 3, 5, 7, 8, 10, 12) else 30


# 'YY/YYYY년', 'N월', 'D일' 토큰. 데이터에 공백이 끼어도(예: "2026 년") 매칭되도록 \s* 허용.
_MONTH_SHIFT_PAT = re.compile(r"(?:(\d{2,4})\s*년\s*)?(\d{1,2})\s*월(?:\s*(\d{1,2})\s*일)?")


def _shift_months_in_text(s, delta, current_year=2000):
    """문자열의 모든 'N월'(앞의 'YY/YYYY년', 뒤의 'D일' 포함)을 delta 개월 이동한다.
    12월을 넘으면 연도 +, 말일 보정(없는 날짜는 그 달 말일로 내림, 윤년 고려), 0패딩 폭 보존.
    이 로직을 백엔드 헬퍼(ctx.shift_months)로 두는 이유: LLM 이 VBA 정규식/한글에 공백을 끼워
    매번 깨뜨리던 문제를 결정적으로 제거하기 위함. 매칭 없으면 원문 그대로 반환."""
    s = str(s)
    delta = int(delta)
    out = []
    pos = 0
    ctx_year = 0
    for mo in _MONTH_SHIFT_PAT.finditer(s):
        out.append(s[pos:mo.start()])
        yr = mo.group(1) or ""
        mon = mo.group(2) or ""
        dy = mo.group(3) or ""
        total = (int(mon) - 1) + delta
        new_m = (total % 12) + 1
        y_shift = total // 12
        piece = ""
        if yr:
            ny = int(yr) + y_shift
            ctx_year = ny if len(yr) == 4 else 2000 + ny
            piece += format(ny, "0%dd" % len(yr)) + "년 "
        cy = ctx_year if ctx_year > 0 else current_year
        piece += format(new_m, "0%dd" % len(mon)) + "월"
        if dy:
            nd = int(dy)
            md = _days_in_month(new_m, cy)
            if nd > md:
                nd = md
            piece += " " + format(nd, "0%dd" % len(dy)) + "일"
        out.append(piece)
        pos = mo.end()
    out.append(s[pos:])
    return "".join(out)


# ---- 피벗/크로스탭 집계(순수 함수 — COM 불필요, 단위테스트 가능). ctx.pivot 2D 가 사용. ----
def _pivot_to_num(v):
    if isinstance(v, bool):
        return None
    if isinstance(v, (int, float)):
        return float(v)
    try:
        return float(str(v).replace(",", ""))
    except Exception:
        return None


def _pivot_agg(vals, name):
    name = str(name or "sum").lower()
    nums = [n for n in (_pivot_to_num(v) for v in vals) if n is not None]
    if name == "count":
        return len(vals)
    if name in ("avg", "average", "mean"):
        return (sum(nums) / len(nums)) if nums else 0
    if name == "max":
        return max(nums) if nums else ""
    if name == "min":
        return min(nums) if nums else ""
    return sum(nums)


def _pivot_sort_keys(keys):
    # 숫자로 해석되면 숫자순(월 1,2,…), 아니면 문자순(지점 가나다). 혼재 시 숫자 먼저.
    def _k(x):
        n = _pivot_to_num(x)
        return (0, n) if n is not None else (1, str(x))
    return sorted(keys, key=_k)


def _pivot_crosstab(data, g_i, c_i, v_i, agg, row_label="행"):
    """2D 크로스탭 grid 생성. 반환: [[row_label, col1, col2, ...], [행키, agg, agg, ...], ...].
    행키=group_by 값, 열키=column 값, 셀=value 의 agg. value 없으면 건수(count)."""
    cells = {}
    rkeys, ckeys = [], []
    rseen, cseen = set(), set()
    for r in data:
        r = list(r)
        rk = r[g_i] if (g_i is not None and g_i < len(r)) else ""
        ck = r[c_i] if (c_i is not None and c_i < len(r)) else ""
        if (rk is None or str(rk).strip() == "") and (ck is None or str(ck).strip() == ""):
            continue  # 완전 빈 행 skip
        if rk not in rseen:
            rseen.add(rk); rkeys.append(rk)
        if ck not in cseen:
            cseen.add(ck); ckeys.append(ck)
        v = r[v_i] if (v_i is not None and v_i < len(r)) else 1
        cells.setdefault((rk, ck), []).append(v)
    rkeys = _pivot_sort_keys(rkeys)
    ckeys = _pivot_sort_keys(ckeys)
    out = [[row_label] + list(ckeys)]
    for rk in rkeys:
        out.append([rk] + [(_pivot_agg(cells[(rk, ck)], agg) if (rk, ck) in cells else 0) for ck in ckeys])
    return out, rkeys, ckeys


@total_ordering
class ExcelColumnNumber:
    """1-based Excel column number that is also safe as a Python row-list index."""
    def __init__(self, value):
        self.value = int(value)

    def __int__(self):
        return self.value

    def __index__(self):
        return max(0, self.value - 1)

    def __repr__(self):
        return str(self.value)

    def __str__(self):
        return str(self.value)

    def _coerce(self, other):
        try:
            return int(other)
        except Exception:
            return NotImplemented

    def __eq__(self, other):
        other_value = self._coerce(other)
        if other_value is NotImplemented:
            return False
        return self.value == other_value

    def __lt__(self, other):
        other_value = self._coerce(other)
        if other_value is NotImplemented:
            return NotImplemented
        return self.value < other_value

    def __hash__(self):
        return hash(self.value)

    def __add__(self, other):
        other_value = self._coerce(other)
        if other_value is NotImplemented:
            return NotImplemented
        return self.value + other_value

    def __radd__(self, other):
        return self.__add__(other)

    def __sub__(self, other):
        other_value = self._coerce(other)
        if other_value is NotImplemented:
            return NotImplemented
        return self.value - other_value

    def __rsub__(self, other):
        other_value = self._coerce(other)
        if other_value is NotImplemented:
            return NotImplemented
        return other_value - self.value


def _excel_collection_names(collection):
    names = []
    try:
        count = int(collection.Count)
    except Exception:
        return names
    for idx in range(1, count + 1):
        try:
            item = collection.Item(idx)
            name = getattr(item, "Name", None)
            if name:
                names.append(str(name))
        except Exception:
            continue
    return names


def ensure_result_file(result_id):
    if not result_id:
        return None
    result = RESULTS.get(result_id)
    if not result:
        return None
    if "workerWorkbookId" in result and "sheets" not in result:
        result["sheets"] = export_node_worker_workbook(result["workerWorkbookId"])
    if "path" not in result:
        result_path = BACKEND_DIR / f"{result_id}_{result.get('name') or 'result.xlsx'}"
        write_result_workbook(
            Path(result["template_path"]),
            result_path,
            result.get("sheets") or {},
            result.get("forced_value_cells") or [],
        )
        result["path"] = str(result_path)
    path = Path(result["path"])
    return path if path.exists() else None


# 녹화 편집 모드 — 녹화 중에는 라이브 미러의 편집 잠금(시트 보호·키 차단·우클릭 금지·리본 숨김)을
# 풀어 사용자가 실제 Excel 창에서 입력/서식/병합/필터를 직접 조작할 수 있게 한다. 정지 시 원복.
RECORDING_EDIT_UNLOCKED = False

# [2A] 라이브 프레임 되띄움 억제 — 클라 runner(실행기)가 전면(헤드리스)일 때 True.
# 서버측 자동 되띄움(_present_live_session_frame)이 실행기 위로 엑셀 오버레이를 얹는 것을 막는다.
# 오직 /api/excel/runner-mode 로만 세팅(프로세스 재시작 기본 False), 새 라이브 세션 open 시 False 로 리셋.
LIVE_RESTORE_SUPPRESSED = False


def _recording_edit_unlock_active(app=None):
    """녹화 편집 모드 중이고 대상이 라이브 공유 인스턴스면 True(잠금 적용을 건너뛴다).

    readOnlyMirror 세션은 자체 DispatchEx 인스턴스라 녹화와 무관하게 계속 잠긴다."""
    if not RECORDING_EDIT_UNLOCKED:
        return False
    if app is None:
        return True
    try:
        return _is_live_shared_app(app)
    except Exception:
        return False


def _restore_excel_default_input(app):
    """녹화 편집 모드: 미러 입력 차단 원복(셀 내 편집 + 기본 키 동작)."""
    try:
        app.EditDirectlyInCell = True
    except Exception:
        pass
    for key in ("{F2}", "{DELETE}", "{BACKSPACE}", "^v", "^x", "+{INSERT}", "+{DELETE}"):
        try:
            app.OnKey(key)  # 두 번째 인자 생략 = 기본 동작 복원
        except Exception:
            pass


def _enable_excel_context_menus(app):
    """녹화 편집 모드: 우클릭(컨텍스트) 메뉴 복원 — 병합/셀 서식 진입 경로."""
    try:
        bars = app.CommandBars
        count = bars.Count
    except Exception:
        return
    for idx in range(1, count + 1):
        try:
            bar = bars.Item(idx)
            if bar.Type == 2:  # msoBarTypePopup
                bar.Enabled = True
        except Exception:
            continue


def _set_excel_ribbon_visible(app, visible):
    # [클립보드 보존] SHOW.TOOLBAR(XLM) 실행은 Excel 복사 모드(마퀴)를 취소한다(실측).
    # 이미 원하는 상태면 아무것도 하지 않는다 — 사용자가 복사해 둔 것을 지키기 위해
    # 중복 호출을 no-op 으로 만든다(읽기 실패 시엔 기존 동작 유지).
    # ※ SHOW.TOOLBAR 는 리본을 '보이게' 하는 유일한 경로이기도 하다 — 녹화 중이라고 이걸
    #   통째로 건너뛰면 녹화 시작 시 메뉴바가 안 뜬다(회귀). 그래서 여기선 스킵하지 않고,
    #   교차복붙 마퀴 보존은 재표시 재적용을 recUnlockDone 게이트로 막는 것으로만 처리한다.
    try:
        if bool(app.CommandBars("Ribbon").Visible) == bool(visible):
            return
    except Exception:
        pass
    try:
        app.ExecuteExcel4Macro('SHOW.TOOLBAR("Ribbon",%s)' % ("True" if visible else "False"))
    except Exception:
        pass
    try:
        app.CommandBars("Ribbon").Visible = bool(visible)
    except Exception:
        pass
    if visible:
        # SHOW.TOOLBAR True 로 리본을 다시 켜면 '최소화(탭만 보이고 버튼 본문은 닫힘)'
        # 상태로 복원되는 경우가 있다(사용자: "도구탭이 닫힌채로 안 나온다"). 최소화 여부는
        # GetPressedMso('MinimizeRibbon') 로 정확히 읽어(높이 절대값은 DPI/버전마다 달라
        # 임계값이 부정확), 최소화면 토글로 펼친다. (토글이라 최소화일 때만 호출해야 안전)
        try:
            minimized = bool(app.CommandBars.GetPressedMso("MinimizeRibbon"))
        except Exception:
            # GetPressedMso 미지원 환경 폴백 — 높이 비율 대신 보수적으로 건드리지 않음
            minimized = False
        if minimized:
            try:
                app.CommandBars.ExecuteMso("MinimizeRibbon")
            except Exception:
                pass


def _set_live_sessions_edit_unlock(unlocked):
    """녹화 동안 라이브 엑셀뷰의 편집 잠금을 해제/원복한다(Excel 워커에서 실행).

    잠금 해제: 전 라이브 세션 시트 보호 해제 + 입력키/우클릭 복원 + 리본 표시.
    원복: 전 시트 재보호(녹화 중 만든 새 시트 포함) + 입력 차단 + 리본 숨김."""
    global RECORDING_EDIT_UNLOCKED
    RECORDING_EDIT_UNLOCKED = bool(unlocked)
    apps = {}
    live_workbooks = []  # (app, wb) — 리본 창별 적용용(SDI)
    for session in list(EXCEL_SESSIONS.values()):
        if not session.get("liveEditable"):
            continue
        # 녹화 중 탭 전환 시 재적용을 세션당 1회로 제한하는 플래그(클립보드 보존).
        # 잠금 해제 시 True(이미 해제됨), 원복 시 False 로 리셋.
        session["recUnlockDone"] = bool(unlocked)
        wb = session.get("workbook")
        if wb is not None:
            try:
                _protect_workbook_for_read_only_mirror(wb, not unlocked)
            except Exception:
                pass
        app = session.get("app")
        if app is not None:
            if wb is not None:
                live_workbooks.append((app, wb))
            try:
                apps[int(app.Hwnd)] = app
            except Exception:
                apps[id(app)] = app
    for app in apps.values():
        if unlocked:
            _restore_excel_default_input(app)
            _enable_excel_context_menus(app)
        else:
            _configure_read_only_mirror_input_block(app)
            _disable_excel_context_menus(app)
    # [리본은 창마다] Excel 2013+(SDI)는 워크북 창마다 리본이 따로다 — 앱당 1회
    # SHOW.TOOLBAR 는 '활성 창'에만 적용돼, 두 워크북일 때 다른 워크북은 리본이
    # 안 열렸다(사용자 보고). 각 세션 창을 잠깐 활성화해 창별로 적용한다.
    # (파킹된 창은 화면 밖이라 시각 변화 없음. 시작/정지 시점이라 클립보드 걱정 없음.)
    for app in apps.values():
        _orig_active = None
        try:
            _orig_active = app.ActiveWorkbook
        except Exception:
            pass
        for _app2, _wb2 in live_workbooks:
            if _app2 is not app:
                continue
            try:
                _wb2.Activate()
            except Exception:
                continue
            _set_excel_ribbon_visible(app, bool(unlocked))
        try:
            if _orig_active is not None:
                _orig_active.Activate()
        except Exception:
            pass
            # 녹화 중 확장된 수식 입력줄을 잠금 원복과 함께 1줄로 되돌린다.
            _show_excel_formula_bar(app)
    return {"unlocked": bool(unlocked), "apps": len(apps)}


def _protect_workbook_for_read_only_mirror(wb, enabled=True):
    if enabled:
        try:
            # 녹화 편집 모드 중에는 라이브 워크북을 재보호하지 않는다(정지 시 일괄 복구).
            if _recording_edit_unlock_active(wb.Application):
                return
        except Exception:
            pass
    for idx in range(1, wb.Worksheets.Count + 1):
        ws = wb.Worksheets(idx)
        try:
            if not enabled:
                ws.Unprotect(Password=EXCEL_MIRROR_PROTECT_PASSWORD)
                continue
        except Exception:
            pass
        if enabled:
            try:
                ws.Cells.Locked = True
            except Exception:
                pass
            try:
                ws.Protect(
                    EXCEL_MIRROR_PROTECT_PASSWORD,  # Password
                    True,   # DrawingObjects
                    True,   # Contents
                    True,   # Scenarios
                    True,   # UserInterfaceOnly
                    False,  # AllowFormattingCells
                    True,   # AllowFormattingColumns: column width drag/double-click autofit
                    True,   # AllowFormattingRows: row height drag/double-click autofit
                    False,  # AllowInsertingColumns
                    False,  # AllowInsertingRows
                    False,  # AllowInsertingHyperlinks
                    False,  # AllowDeletingColumns
                    False,  # AllowDeletingRows
                    True,   # AllowSorting
                    True,   # AllowFiltering
                    False,  # AllowUsingPivotTables
                )
            except TypeError:
                ws.Protect(Password=EXCEL_MIRROR_PROTECT_PASSWORD)
            except Exception:
                pass
            try:
                _allow_read_only_mirror_selection(ws)
            except Exception:
                pass


def _allow_read_only_mirror_selection(ws):
    try:
        ws.EnableSelection = 0  # xlNoRestrictions: allow cell/range selection on protected sheets.
    except Exception:
        pass
    try:
        ws.ScrollArea = ""
    except Exception:
        pass


def _configure_read_only_mirror_input_block(app):
    if _recording_edit_unlock_active(app):
        return
    try:
        app.EditDirectlyInCell = False
    except Exception:
        pass
    try:
        app.CellDragAndDrop = True
    except Exception:
        pass
    try:
        app.CutCopyMode = False
    except Exception:
        pass
    for key in ("{F2}", "{DELETE}", "{BACKSPACE}", "^v", "^x", "+{INSERT}", "+{DELETE}"):
        try:
            app.OnKey(key, "")
        except Exception:
            pass


def _set_display_prop_if_changed(obj, name, value):
    """Display* 계열 속성은 '쓰기 자체'가 값 무관하게 복사 마퀴(CutCopyMode)를 취소한다
    (실측 프로브: 같은 값 재대입 7종 전부 킬러, 읽기는 무해). 탭 전환마다 이 속성들을
    재대입하던 것이 '녹화 중 A 복사 → B 탭 전환 → 붙여넣기 불가'(교차파일 Ctrl+V 사망)의
    진짜 원인 — 현재값과 다를 때만 쓴다."""
    try:
        if bool(getattr(obj, name)) == bool(value):
            return
    except Exception:
        pass  # 읽기 실패 → 원래처럼 쓰기 시도
    try:
        setattr(obj, name, value)
    except Exception:
        pass


def _show_excel_formula_bar(app):
    """읽기 전용 미러에서도 실제 Excel처럼 수식 입력줄은 보이게 둔다."""
    _set_display_prop_if_changed(app, "DisplayFormulaBar", True)
    _set_display_prop_if_changed(app, "DisplayStatusBar", True)


def _disable_excel_context_menus(app):
    """오버레이 엑셀에서 마우스 우클릭(컨텍스트) 메뉴를 막는다.
    msoBarTypePopup(2) CommandBar = 우클릭/컨텍스트 메뉴이므로 모두 비활성화한다.
    DispatchEx로 만든 전용 인스턴스라 사용자의 일반 엑셀에는 영향이 없다."""
    if _recording_edit_unlock_active(app):
        return
    try:
        bars = app.CommandBars
        count = bars.Count
    except Exception:
        return
    for idx in range(1, count + 1):
        try:
            bar = bars.Item(idx)
            if bar.Type == 2:  # msoBarTypePopup
                bar.Enabled = False
        except Exception:
            continue


def _configure_excel_grid_window(app, wb=None):
    keep_edit = _recording_edit_unlock_active(app)  # 녹화 편집 모드 — 리본/입력 잠금 유지 금지
    try:
        app.DisplayAlerts = False
        _show_excel_formula_bar(app)
        app.Interactive = True
        app.UserControl = True
        app.EnableEvents = True
        _configure_read_only_mirror_input_block(app)
        if not keep_edit:
            app.ExecuteExcel4Macro('SHOW.TOOLBAR("Ribbon",False)')
    except Exception:
        pass
    if not keep_edit:
        try:
            app.CommandBars("Ribbon").Visible = False
        except Exception:
            pass
    _show_excel_formula_bar(app)
    _disable_excel_context_menus(app)
    try:
        win = app.ActiveWindow
        for _p in ("DisplayHeadings", "DisplayGridlines", "DisplayWorkbookTabs",
                   "DisplayHorizontalScrollBar", "DisplayVerticalScrollBar"):
            _set_display_prop_if_changed(win, _p, True)
    except Exception:
        pass
    if wb is not None:
        try:
            for idx in range(1, wb.Worksheets.Count + 1):
                _allow_read_only_mirror_selection(wb.Worksheets(idx))
        except Exception:
            pass
def _ensure_excel_workbook_view(app, wb=None, make_visible=True, activate=True, maximize_workbook=True, defer_show=False, app_level=True):
    try:
        if make_visible and not defer_show:
            app.Visible = True
    except Exception:
        pass
    try:
        app.Interactive = True
    except Exception:
        pass
    try:
        app.UserControl = True
    except Exception:
        pass
    try:
        app.EnableEvents = True
    except Exception:
        pass
    try:
        if not defer_show:
            app.ScreenUpdating = True
    except Exception:
        pass
    _show_excel_formula_bar(app)
    try:
        if activate and wb is not None:
            wb.Activate()
    except Exception:
        pass
    try:
        if app_level:
            # SDI 공유 인스턴스에서 app.WindowState 는 활성 프레임에만 적용된다.
            # frame 모드 호출자는 app_level=False 로 끄고 wb.Windows(1).WindowState 를 직접 쓴다.
            app.WindowState = -4143  # xlNormal: keep the outer Excel window at the mirror panel size.
    except Exception:
        pass
    try:
        win = wb.Windows(1) if wb is not None else app.ActiveWindow
        if win is not None:
            if not defer_show:
                win.Visible = True
            if maximize_workbook:
                win.WindowState = -4137  # xlMaximized: fill only the workbook area inside Excel.
            # [마퀴 보존] Display* 쓰기는 값 무관하게 복사 모드를 취소(실측) — 다를 때만 쓴다.
            # 이 함수는 탭 전환마다 돌아서(6170), 무조건 대입이 교차파일 Ctrl+V 를 죽였다.
            for _p in ("DisplayHeadings", "DisplayGridlines", "DisplayWorkbookTabs",
                       "DisplayHorizontalScrollBar", "DisplayVerticalScrollBar"):
                _set_display_prop_if_changed(win, _p, True)
    except Exception:
        pass


def _capture_browser_hwnd(title_hint=None):
    if win32gui is None:
        return None
    # [이름 변경] 창 제목이 'B2B 빌링 Agent' → 'AX-Cell' 로 바뀌었다(NativeHost.Text).
    # 이 함수는 그 제목으로 우리 창을 찾으므로 기본값도 함께 바꿔야 한다 — 안 바꾸면 Excel 미러가
    # 붙을 창을 못 찾는다. 구 이름 창(옛 배포본)도 계속 인정한다.
    title_hint = normalize_text(title_hint or "AX-Cell")

    def usable(hwnd):
        try:
            if not win32gui.IsWindow(hwnd) or not win32gui.IsWindowVisible(hwnd):
                return False
            title = win32gui.GetWindowText(hwnd) or ""
            if not title:
                return False
            normalized = normalize_text(title)
            if "excel" in normalized:
                return False
            if title_hint and title_hint in normalized:
                return True
            if "axcell" in normalized.replace("-", ""):      # 새 이름(표기 변형 포함)
                return True
            return "b2b" in normalized and "agent" in normalized   # 구 이름 하위호환
        except Exception:
            return False

    try:
        hwnd = win32gui.GetForegroundWindow()
        if usable(hwnd):
            return int(hwnd)
    except Exception:
        pass

    found = []

    def enum_proc(hwnd, _):
        if usable(hwnd):
            found.append(int(hwnd))
        return True

    try:
        win32gui.EnumWindows(enum_proc, None)
    except Exception:
        return None
    return found[0] if found else None


def _browser_content_target(browser_hwnd):
    if win32gui is None or not browser_hwnd:
        return None
    try:
        if not win32gui.IsWindow(browser_hwnd):
            return None
    except Exception:
        return None
    candidates = []

    def enum_child(hwnd, _):
        try:
            if not win32gui.IsWindowVisible(hwnd):
                return True
            cls = win32gui.GetClassName(hwnd) or ""
            rect = win32gui.GetWindowRect(hwnd)
            width = max(0, rect[2] - rect[0])
            height = max(0, rect[3] - rect[1])
            if width < 300 or height < 200:
                return True
            score = width * height
            if "Chrome_RenderWidgetHostHWND" in cls:
                score *= 4
            candidates.append((score, int(hwnd), rect))
        except Exception:
            pass
        return True

    try:
        win32gui.EnumChildWindows(int(browser_hwnd), enum_child, None)
    except Exception:
        pass
    if candidates:
        candidates.sort(key=lambda item: item[0], reverse=True)
        return candidates[0][1], candidates[0][2]
    try:
        left, top = win32gui.ClientToScreen(int(browser_hwnd), (0, 0))
        client = win32gui.GetClientRect(int(browser_hwnd))
        return int(browser_hwnd), (left, top, left + client[2], top + client[3])
    except Exception:
        return None


def _browser_content_rect(browser_hwnd):
    target = _browser_content_target(browser_hwnd)
    return target[1] if target else None


def _resolve_excel_mirror_rect(left, top, width, height, browser_hwnd=None, client_left=None, client_top=None, client_width=None, client_height=None, viewport_width=None, viewport_height=None):
    screen_left = float(left or 0)
    screen_top = float(top or 0)
    screen_width = float(width or 320)
    screen_height = float(height or 240)
    content_rect = _browser_content_rect(browser_hwnd)
    if content_rect and client_left is not None and client_top is not None:
        content_left, content_top, content_right, content_bottom = content_rect
        content_width = max(1, content_right - content_left)
        content_height = max(1, content_bottom - content_top)
        vw = max(1.0, float(viewport_width or content_width))
        vh = max(1.0, float(viewport_height or content_height))
        scale_x = content_width / vw
        scale_y = content_height / vh
        screen_left = content_left + float(client_left or 0) * scale_x
        screen_top = content_top + float(client_top or 0) * scale_y
        screen_width = float(client_width or width or 320) * scale_x
        screen_height = float(client_height or height or 240) * scale_y
    return (
        int(max(-32000, screen_left)),
        int(max(-32000, screen_top)),
        int(max(320, screen_width)),
        int(max(240, screen_height)),
    )


def _position_excel_window(
    app,
    left,
    top,
    width,
    height,
    browser_hwnd=None,
    native_parent_hwnd=None,
    native_host_hwnd=None,
    native_overlay=False,
    client_left=None,
    client_top=None,
    client_width=None,
    client_height=None,
    viewport_width=None,
    viewport_height=None,
    show=True,
    keep_zorder=False,
    hwnd=None,
    no_activate=False,
):
    left, top, width, height = _resolve_excel_mirror_rect(
        left,
        top,
        width,
        height,
        browser_hwnd=browser_hwnd,
        client_left=client_left,
        client_top=client_top,
        client_width=client_width,
        client_height=client_height,
        viewport_width=viewport_width,
        viewport_height=viewport_height,
    )
    if win32gui is not None and win32con is not None:
        # SDI 공유 인스턴스: app.Hwnd 는 '활성 프레임 1개'라 다중 워크북에서 엉뚱한 창을 만진다.
        # 호출자가 세션 프레임 hwnd 를 넘기면 그 창만 제어한다.
        try:
            hwnd = int(hwnd) if hwnd else int(app.Hwnd)
        except Exception:
            hwnd = int(app.Hwnd)
        style_changed = False  # SetWindowLong 으로 실제 스타일이 바뀐 경우에만 SWP_FRAMECHANGED(전체 리드로우)
        if native_overlay:
            try:
                style = win32gui.GetWindowLong(hwnd, win32con.GWL_STYLE)
                desired_style = style
                desired_style &= ~(
                    win32con.WS_CAPTION |
                    win32con.WS_THICKFRAME |
                    win32con.WS_MINIMIZEBOX |
                    win32con.WS_MAXIMIZEBOX |
                    win32con.WS_SYSMENU |
                    win32con.WS_CHILD |
                    win32con.WS_DISABLED
                )
                desired_style |= win32con.WS_POPUP | win32con.WS_CLIPSIBLINGS | win32con.WS_CLIPCHILDREN
                if show:
                    desired_style |= win32con.WS_VISIBLE
                else:
                    desired_style &= ~win32con.WS_VISIBLE
                if desired_style != style:
                    win32gui.SetWindowLong(hwnd, win32con.GWL_STYLE, desired_style)
                    style_changed = True
                try:
                    ex_style = win32gui.GetWindowLong(hwnd, win32con.GWL_EXSTYLE)
                    desired_ex_style = (ex_style | getattr(win32con, "WS_EX_TOOLWINDOW", 0)) & ~getattr(win32con, "WS_EX_APPWINDOW", 0)
                    if desired_ex_style != ex_style:
                        win32gui.SetWindowLong(hwnd, win32con.GWL_EXSTYLE, desired_ex_style)
                        style_changed = True
                except Exception:
                    pass
                try:
                    # Keep Excel as a real top-level window.  Owning it to the host
                    # makes the grid render in place, but it intermittently breaks
                    # Excel's own mouse activation/selection loop after host resize
                    # or position updates.
                    win32gui.SetWindowLong(hwnd, getattr(win32con, "GWL_HWNDPARENT", -8), 0)
                except Exception:
                    pass
            except Exception:
                pass
        elif native_parent_hwnd:
            try:
                parent_hwnd = int(native_parent_hwnd)
                if win32gui.IsWindow(parent_hwnd):
                    current_parent = win32gui.GetParent(hwnd)
                    style = win32gui.GetWindowLong(hwnd, win32con.GWL_STYLE)
                    desired_style = style
                    desired_style &= ~(
                        win32con.WS_CAPTION |
                        win32con.WS_THICKFRAME |
                        win32con.WS_MINIMIZEBOX |
                        win32con.WS_MAXIMIZEBOX |
                        win32con.WS_SYSMENU |
                        win32con.WS_POPUP |
                        win32con.WS_VISIBLE |
                        win32con.WS_DISABLED
                    )
                    desired_style |= win32con.WS_CHILD | win32con.WS_CLIPSIBLINGS | win32con.WS_CLIPCHILDREN
                    if show:
                        desired_style |= win32con.WS_VISIBLE
                    if desired_style != style:
                        win32gui.SetWindowLong(hwnd, win32con.GWL_STYLE, desired_style)
                        style_changed = True
                    if current_parent != parent_hwnd:
                        win32gui.SetParent(hwnd, parent_hwnd)
                    # SetParent changes Excel into a child window. Coordinates must
                    # be relative to the native panel, not desktop screen coordinates.
                    left = 0
                    top = 0
            except Exception:
                pass
        elif browser_hwnd:
            try:
                parent_hwnd = int(browser_hwnd)
                if win32gui.IsWindow(parent_hwnd):
                    style = win32gui.GetWindowLong(hwnd, win32con.GWL_STYLE)
                    style &= ~(
                        win32con.WS_CAPTION |
                        win32con.WS_THICKFRAME |
                        win32con.WS_MINIMIZEBOX |
                        win32con.WS_MAXIMIZEBOX |
                        win32con.WS_SYSMENU |
                        win32con.WS_POPUP
                    )
                    style |= win32con.WS_CHILD | win32con.WS_VISIBLE
                    win32gui.SetWindowLong(hwnd, win32con.GWL_STYLE, style)
                    style_changed = True
                    win32gui.SetParent(hwnd, parent_hwnd)
                    left, top = win32gui.ScreenToClient(parent_hwnd, (left, top))
            except Exception:
                pass
        # [최대화 배치 무시] SetWindowPos 는 최대화된 창의 좌표/크기를 무시한다 — 미러 좌표로
        # 배치하기 전에 비활성 복원해 둔다(안 하면 프레임이 화면 전체를 덮은 채 남는다).
        _unmaximize_hwnd_no_activate(hwnd)
        # NOACTIVATE 를 항상 적용: 재배치/표시가 배경 미러 창을 활성화해 위로 튀어나오며
        # 회색 플래시를 만들고 활성 창의 포커스를 빼앗는 문제 방지(이 함수는 미러 창 전용).
        flags = win32con.SWP_NOOWNERZORDER | win32con.SWP_NOACTIVATE
        if style_changed:
            flags |= win32con.SWP_FRAMECHANGED  # 스타일이 실제로 바뀐 호출에서만 전체 프레임 리드로우
        if keep_zorder:
            # z-order 를 바꾸지 않고 위치/크기만 변경(비활성 창이 위로 튀어나오는 순회 방지).
            flags |= win32con.SWP_NOZORDER
        if show:
            flags |= win32con.SWP_SHOWWINDOW
        else:
            flags |= win32con.SWP_HIDEWINDOW
        win32gui.SetWindowPos(
            hwnd,
            win32con.HWND_TOP,
            left,
            top,
            width,
            height,
            flags,
        )
        if show:
            try:
                if win32gui.IsIconic(hwnd):
                    win32gui.ShowWindow(hwnd, getattr(win32con, "SW_RESTORE", 9))
                if native_parent_hwnd or native_overlay:
                    win32gui.ShowWindow(hwnd, getattr(win32con, "SW_SHOWNA", 8))
                    _focus_excel_grid_child(hwnd)
                elif no_activate:
                    # SW_SHOWNORMAL 은 활성화를 동반해 호스트 포커스를 뺏는다(다음 UI 클릭이
                    # 창 활성화에 소비되는 '클릭 씹힘'의 원인). 라이브 프레임은 비활성 표시로 충분.
                    win32gui.ShowWindow(hwnd, getattr(win32con, "SW_SHOWNA", 8))
                else:
                    # SW_SHOWNORMAL 은 창을 활성화해 배경 미러가 포커스를 빼앗음 → 비활성 표시(SW_SHOWNA).
                    win32gui.ShowWindow(hwnd, getattr(win32con, "SW_SHOWNA", 8))
            except Exception:
                pass
        return
    # Excel COM uses points, not pixels. This fallback is approximate.
    app.WindowState = -4143
    app.Left = left * 72 / 96
    app.Top = top * 72 / 96
    app.Width = width * 72 / 96
    app.Height = height * 72 / 96


def _focus_excel_grid_child(hwnd):
    if win32gui is None:
        return
    try:
        hwnd = int(hwnd)
    except Exception:
        return
    if not hwnd:
        return
    best = hwnd
    best_score = -1

    def enum_child(child, _param):
        nonlocal best, best_score
        try:
            if not win32gui.IsWindowVisible(child):
                return True
            cls = str(win32gui.GetClassName(child) or "")
            rect = win32gui.GetWindowRect(child)
            width = max(0, int(rect[2]) - int(rect[0]))
            height = max(0, int(rect[3]) - int(rect[1]))
            area = width * height
            if area <= 0:
                return True
            score = area
            if "EXCEL7" in cls.upper():
                score += 10**12
            elif "XLDESK" in cls.upper():
                score += 10**11
            elif "XLMAIN" in cls.upper():
                score += 10**10
            if score > best_score:
                best_score = score
                best = child
        except Exception:
            pass
        return True

    try:
        win32gui.EnumChildWindows(hwnd, enum_child, None)
    except Exception:
        pass
    try:
        user32 = ctypes.windll.user32
        kernel32 = ctypes.windll.kernel32
        target_thread = user32.GetWindowThreadProcessId(int(best), None)
        current_thread = kernel32.GetCurrentThreadId()
        attached = False
        if target_thread and target_thread != current_thread:
            attached = bool(user32.AttachThreadInput(current_thread, target_thread, True))
        try:
            user32.SetFocus(int(best))
        finally:
            if attached:
                user32.AttachThreadInput(current_thread, target_thread, False)
    except Exception:
        try:
            win32gui.SetFocus(best)
        except Exception:
            pass


def _set_window_owner_hwnd(hwnd, owner_hwnd):
    """지정한 최상위 창의 소유자(owner)를 지정/해제한다.
    owner_hwnd가 유효하면 그 창 위에 항상 표시되고(가려지지 않음), None/0이면 소유자 해제.
    SetParent(자식화, WS_CHILD)와 달리 owner 관계는 top-level 창을 유지하므로 셀 입력/선택이 정상 동작."""
    if win32gui is None or win32con is None:
        return False
    try:
        hwnd = int(hwnd)
    except Exception:
        return False
    try:
        if not hwnd or not win32gui.IsWindow(hwnd):
            return False
    except Exception:
        return False
    try:
        owner = int(owner_hwnd) if owner_hwnd else 0
    except Exception:
        owner = 0
    if owner:
        try:
            if not win32gui.IsWindow(owner):
                owner = 0
        except Exception:
            owner = 0
    try:
        win32gui.SetWindowLong(hwnd, getattr(win32con, "GWL_HWNDPARENT", -8), owner)
        return True
    except Exception:
        return False


def _set_excel_window_owner(app, owner_hwnd):
    """(legacy) app.Hwnd 프레임의 owner 지정. frame 모드에서는 세션 프레임 hwnd 에
    _set_window_owner_hwnd 를 직접 쓴다(공유 인스턴스에서 app.Hwnd 는 활성 프레임 1개뿐)."""
    try:
        hwnd = int(app.Hwnd)
    except Exception:
        return False
    return _set_window_owner_hwnd(hwnd, owner_hwnd)


def _style_live_frame(hwnd):
    """라이브 프레임을 작업표시줄/Alt+Tab 목록에서 제외(WS_EX_TOOLWINDOW, WS_EX_APPWINDOW 제거).
    소유(owned) 프레임이 활성일 때 호스트 작업표시줄 버튼이 '활성 그룹'으로 표시되어
    클릭 시 호스트가 최소화되는 혼동과, 화면 밖에 파킹된 프레임으로 Alt+Tab 진입하는
    문제를 함께 막는다. 캡션/프레임 자체는 유지(frameless 는 선택/사이즈를 깨서 폐기됨)."""
    if win32gui is None or win32con is None:
        return
    try:
        hwnd = int(hwnd)
        if not hwnd or not win32gui.IsWindow(hwnd):
            return
        ex_style = win32gui.GetWindowLong(hwnd, win32con.GWL_EXSTYLE)
        desired = (ex_style | getattr(win32con, "WS_EX_TOOLWINDOW", 0x80)) & ~getattr(win32con, "WS_EX_APPWINDOW", 0x40000)
        if desired != ex_style:
            win32gui.SetWindowLong(hwnd, win32con.GWL_EXSTYLE, desired)
            flags = (
                win32con.SWP_NOMOVE | win32con.SWP_NOSIZE | win32con.SWP_NOZORDER |
                win32con.SWP_NOACTIVATE | win32con.SWP_NOOWNERZORDER | win32con.SWP_FRAMECHANGED
            )
            win32gui.SetWindowPos(hwnd, 0, 0, 0, 0, 0, flags)
    except Exception:
        pass


def _show_window_na(hwnd):
    """창을 활성화 없이 표시(SW_SHOWNA). 포커스는 현재 창(호스트)에 그대로 남는다."""
    if win32gui is None or win32con is None:
        return
    try:
        hwnd = int(hwnd)
        if not hwnd or not win32gui.IsWindow(hwnd):
            return
        win32gui.ShowWindow(hwnd, getattr(win32con, "SW_SHOWNA", 8))
    except Exception:
        pass


def _unmaximize_hwnd_no_activate(hwnd):
    """최대화된 창을 '활성화 없이' 보통 크기로 되돌린다.

    [핵심] Windows 의 SetWindowPos 는 **최대화(zoomed) 상태 창의 위치/크기 변경을 무시**한다.
    그래서 새로 열린 Excel 프레임이 최대화 상태면 화면 밖 파킹도, 미러 좌표 배치도 조용히
    실패해 프레임이 화면을 덮은 채 남았다(단계 OFF 되돌리기 직후 '엑셀이 최대화되며 튀어나옴'
    → 최소화/복원 시 회색 오버레이, 사용자 실측 2026-08-04).
    ShowWindow(SW_RESTORE)는 활성화를 동반해 호스트 포커스를 뺏으므로, SetWindowPlacement 에
    SW_SHOWNOACTIVATE 를 넣어 비활성 복원한다. 미러는 어차피 지정 좌표/크기로 배치하므로
    '보통 크기로 되돌리기'가 정상 경로다.
    """
    if win32gui is None:
        return
    try:
        hwnd = int(hwnd)
        if not hwnd or not win32gui.IsWindow(hwnd):
            return
        if not win32gui.IsZoomed(hwnd):
            return
        flags, _show_cmd, minpos, maxpos, normalpos = win32gui.GetWindowPlacement(hwnd)
        win32gui.SetWindowPlacement(hwnd, (flags, 4, minpos, maxpos, normalpos))  # 4 = SW_SHOWNOACTIVATE
    except Exception:
        pass


def _move_hwnd_offscreen(hwnd):
    """프레임을 숨기지 않고 화면 밖(-32000)으로만 이동(WS_VISIBLE 유지).
    SW_HIDE 와 달리 '활성 창 소멸'이 일어나지 않아 OS 가 z-order 의 임의 다음 창
    (무관한 다른 앱일 수 있음)을 끌어올리는 일이 없고, 다시 보일 때 회색 빈 프레임도 안 생긴다."""
    if win32gui is None or win32con is None:
        return
    try:
        hwnd = int(hwnd)
        if not hwnd or not win32gui.IsWindow(hwnd):
            return
        _unmaximize_hwnd_no_activate(hwnd)   # 최대화면 SetWindowPos 가 위치를 무시한다
        flags = (
            getattr(win32con, "SWP_NOACTIVATE", 0x0010) |
            getattr(win32con, "SWP_NOOWNERZORDER", 0x0200) |
            getattr(win32con, "SWP_NOSIZE", 0x0001)
        )
        win32gui.SetWindowPos(hwnd, getattr(win32con, "HWND_BOTTOM", 1), -32000, -32000, 0, 0, flags)
    except Exception:
        pass


def _visible_excel_top_hwnds_for_pids(pids):
    """주어진 pid 들의 '보이는' 최상위 Excel 창(XLMAIN) 목록.

    SDI 모드에서 워크북 프레임과 루트 창은 둘 다 XLMAIN 이다 — 호출부가 이미 파킹한 프레임
    hwnd 를 제외하면 '워크북 없는 루트 창'만 남는다(회색 빈 Excel 의 정체)."""
    out = []
    if win32gui is None or win32process is None:
        return out
    targets = set()
    for p in (pids or []):
        try:
            p = int(p or 0)
            if p:
                targets.add(p)
        except Exception:
            pass
    if not targets:
        return out

    def visit(hwnd, _):
        try:
            if not win32gui.IsWindowVisible(hwnd):
                return True
            _tid, window_pid = win32process.GetWindowThreadProcessId(hwnd)
            if int(window_pid or 0) not in targets:
                return True
            if "XLMAIN" in (win32gui.GetClassName(hwnd) or "").upper():
                out.append(int(hwnd))
        except Exception:
            pass
        return True

    try:
        win32gui.EnumWindows(visit, None)
    except Exception:
        pass
    return out


def _handoff_foreground_to_host(host_hwnd, hwnds):
    """숨기거나 파킹하려는 프레임이 현재 포그라운드면, OS 가 다음 활성 창을 임의로 고르기 전에
    호스트로 포커스를 명시적으로 넘긴다(AttachThreadInput 으로 합법 전환).
    '탭 전환/워크북 열기 때 무관한 다른 앱 창이 최상단으로 튀어나오는' 증상의 직접 방어선."""
    if win32gui is None:
        return False
    try:
        fg = int(win32gui.GetForegroundWindow() or 0)
    except Exception:
        return False
    try:
        targets = {int(h) for h in (hwnds or []) if h}
    except Exception:
        targets = set()
    if not fg or fg not in targets:
        return False
    try:
        host = int(host_hwnd or 0)
        if not host or not win32gui.IsWindow(host):
            return False
    except Exception:
        return False
    try:
        user32 = ctypes.windll.user32
        kernel32 = ctypes.windll.kernel32
        fg_thread = user32.GetWindowThreadProcessId(fg, None)
        cur_thread = kernel32.GetCurrentThreadId()
        attached = False
        if fg_thread and fg_thread != cur_thread:
            attached = bool(user32.AttachThreadInput(cur_thread, fg_thread, True))
        try:
            user32.SetForegroundWindow(host)
        finally:
            if attached:
                user32.AttachThreadInput(cur_thread, fg_thread, False)
        return True
    except Exception:
        return False


def _style_live_overlay_window(app):
    """라이브 창을 프레임리스(제목줄/테두리/최소·최대화 버튼 제거)로 만든다.
    미러 overlay 스타일과 동일하되 owner(GWL_HWNDPARENT)는 건드리지 않는다 — owner 는 따로 지정해 무깜빡임 유지."""
    if win32gui is None or win32con is None:
        return
    try:
        hwnd = int(app.Hwnd)
        style = win32gui.GetWindowLong(hwnd, win32con.GWL_STYLE)
        desired = style & ~(
            win32con.WS_CAPTION |
            win32con.WS_THICKFRAME |
            win32con.WS_MINIMIZEBOX |
            win32con.WS_MAXIMIZEBOX |
            win32con.WS_SYSMENU |
            win32con.WS_CHILD |
            win32con.WS_DISABLED
        )
        desired |= win32con.WS_POPUP | win32con.WS_CLIPSIBLINGS | win32con.WS_CLIPCHILDREN | win32con.WS_VISIBLE
        if desired != style:
            win32gui.SetWindowLong(hwnd, win32con.GWL_STYLE, desired)
        try:
            ex = win32gui.GetWindowLong(hwnd, win32con.GWL_EXSTYLE)
            desired_ex = (ex | getattr(win32con, "WS_EX_TOOLWINDOW", 0)) & ~getattr(win32con, "WS_EX_APPWINDOW", 0)
            if desired_ex != ex:
                win32gui.SetWindowLong(hwnd, win32con.GWL_EXSTYLE, desired_ex)
        except Exception:
            pass
    except Exception:
        pass


def _raise_excel_hwnd(hwnd):
    if win32gui is None or win32con is None:
        return
    try:
        hwnd = int(hwnd)
        if not hwnd or not win32gui.IsWindow(hwnd):
            return
        if win32gui.IsIconic(hwnd):
            win32gui.ShowWindow(hwnd, getattr(win32con, "SW_RESTORE", 9))
        flags = (
            win32con.SWP_NOMOVE |
            win32con.SWP_NOSIZE |
            win32con.SWP_SHOWWINDOW |
            win32con.SWP_NOACTIVATE |
            win32con.SWP_NOOWNERZORDER
        )
        win32gui.SetWindowPos(hwnd, win32con.HWND_TOPMOST, 0, 0, 0, 0, flags)
        win32gui.SetWindowPos(hwnd, win32con.HWND_NOTOPMOST, 0, 0, 0, 0, flags)
    except Exception:
        pass


def _raise_excel_window(app):
    try:
        _raise_excel_hwnd(int(app.Hwnd))
    except Exception:
        pass


def _excel_process_id(app):
    if win32process is None:
        return None
    try:
        hwnd = int(app.Hwnd)
        if not hwnd:
            return None
        _thread_id, pid = win32process.GetWindowThreadProcessId(hwnd)
        return int(pid) if pid else None
    except Exception:
        return None



def _track_spawned_excel_app(app):
    """이 앱이 띄운 Excel 인스턴스의 pid 를 기록한다(고아 정리용).
    DispatchEx 직후에만 호출할 것 — 사용자 개인 Excel(GetActiveObject)을 등록하면 안 된다."""
    try:
        pid = _excel_process_id(app)
        if pid:
            SPAWNED_EXCEL_PIDS.add(int(pid))
            _perf_trace("excel.spawned", pid=int(pid), tracked=sorted(int(p) for p in SPAWNED_EXCEL_PIDS if p))
    except Exception:
        pass

def _force_kill_pid(pid):
    if not pid or os.name != "nt":
        return
    _perf_trace("excel.force_kill", pid=int(pid))
    try:
        subprocess.run(
            ["taskkill", "/PID", str(int(pid)), "/T", "/F"],
            stdout=subprocess.DEVNULL,
            stderr=subprocess.DEVNULL,
            timeout=3,
            **hidden_subprocess_kwargs(),
        )
    except Exception:
        pass


def _is_pid_alive(pid):
    if not pid or os.name != "nt":
        return False
    # [모래시계 수정] 부모(native host) 생존 감시가 이 함수를 1초마다 부른다. 매번 tasklist 를
    # 새 프로세스로 띄우면 비용도 크고 시작 피드백 커서가 유휴 상태에서 계속 깜빡였다.
    # 프로세스를 만들지 않는 확인을 먼저 쓰고, tasklist 는 최후 폴백.
    #
    # [259 오판 수정] 판정은 WaitForSingleObject 로 한다 — 프로세스 핸들은 '종료되면 시그널'되므로
    # 종료코드와 무관하게 정확하다. GetExitCodeProcess/psutil.pid_exists 는 종료코드 259(STILL_ACTIVE)로
    # 죽은 프로세스의 핸들을 누군가(EDR/WerFault/런처) 쥐고 있으면 영영 '살아있음'으로 오판했고,
    # 그러면 부모감시가 발화하지 않아 서버·Excel 이 고아로 남고 reap 은 무의미한 taskkill 을 반복했다.
    try:
        import ctypes
        kernel32 = ctypes.WinDLL("kernel32", use_last_error=True)
        kernel32.OpenProcess.restype = ctypes.c_void_p          # 핸들 절단 방지(기본 c_int)
        kernel32.OpenProcess.argtypes = [ctypes.c_uint32, ctypes.c_int, ctypes.c_uint32]
        kernel32.WaitForSingleObject.restype = ctypes.c_uint32
        kernel32.WaitForSingleObject.argtypes = [ctypes.c_void_p, ctypes.c_uint32]
        kernel32.CloseHandle.argtypes = [ctypes.c_void_p]
        SYNCHRONIZE = 0x00100000
        PROCESS_QUERY_LIMITED_INFORMATION = 0x1000
        ERROR_ACCESS_DENIED = 5
        WAIT_OBJECT_0 = 0x00000000   # 시그널됨 = 이미 종료
        WAIT_TIMEOUT = 0x00000102    # 미시그널 = 실행 중
        handle = kernel32.OpenProcess(SYNCHRONIZE | PROCESS_QUERY_LIMITED_INFORMATION, 0, int(pid))
        if handle:
            try:
                rc = kernel32.WaitForSingleObject(handle, 0)
                if rc == WAIT_TIMEOUT:
                    return True
                if rc == WAIT_OBJECT_0:
                    return False
                # WAIT_FAILED 등은 단정하지 않고 아래 폴백으로 확정
            finally:
                kernel32.CloseHandle(handle)
        elif ctypes.get_last_error() == ERROR_ACCESS_DENIED:
            return True  # 접근 거부 = 프로세스는 존재함
        # 그 외(87 INVALID_PARAMETER=죽은 pid 등)는 단정하지 않고 폴백으로 확정
    except Exception:
        pass
    if psutil is not None:
        try:
            return bool(psutil.pid_exists(int(pid)))
        except Exception:
            pass
    try:
        completed = subprocess.run(
            ["tasklist", "/FI", f"PID eq {int(pid)}", "/NH"],
            stdout=subprocess.PIPE,
            stderr=subprocess.DEVNULL,
            text=True,
            encoding="utf-8",
            errors="ignore",
            timeout=3,
            **hidden_subprocess_kwargs(),
        )
        return str(int(pid)) in (completed.stdout or "")
    except Exception:
        return False


def _health_excel_diagnostics():
    """Health polling should be cheap. Excel process diagnostics are cached and
    refreshed periodically so an idle app does not keep scanning processes."""
    global HEALTH_LAST_EXCEL_DIAG_AT, HEALTH_CACHED_EXCEL_DIAG
    now = time.time()
    if (
        HEALTH_CACHED_EXCEL_DIAG is None
        or now - float(HEALTH_LAST_EXCEL_DIAG_AT or 0) >= HEALTH_EXCEL_DIAG_INTERVAL_SECONDS
    ):
        HEALTH_CACHED_EXCEL_DIAG = _excel_runtime_diagnostics(reap=True)
        HEALTH_LAST_EXCEL_DIAG_AT = now
    return HEALTH_CACHED_EXCEL_DIAG


def _perf_trace_path():
    return b2b_logs_dir() / "runtime_load_trace.jsonl"


def _process_perf_snapshot(pid):
    if not pid:
        return {}
    info = {"pid": int(pid)}
    if psutil is None:
        return info
    try:
        proc = psutil.Process(int(pid))
        mem = proc.memory_info()
        cpu = proc.cpu_times()
        info.update({
            "name": proc.name(),
            "rssMb": round(float(mem.rss) / (1024 * 1024), 1),
            "vmsMb": round(float(mem.vms) / (1024 * 1024), 1),
            "threads": proc.num_threads(),
            "status": proc.status(),
            "cpuUserSeconds": round(float(getattr(cpu, "user", 0.0)), 3),
            "cpuSystemSeconds": round(float(getattr(cpu, "system", 0.0)), 3),
        })
        try:
            info["handles"] = proc.num_handles()
        except Exception:
            pass
    except Exception as err:
        info["error"] = str(err)
    return info


def _perf_trace(event, **fields):
    try:
        payload = {
            "ts": datetime.datetime.now().isoformat(timespec="milliseconds"),
            "pid": os.getpid(),
            "event": event,
        }
        payload.update(fields)
        with _perf_trace_path().open("a", encoding="utf-8") as f:
            f.write(json.dumps(payload, ensure_ascii=False, default=str) + "\n")
    except Exception:
        pass


def _maybe_perf_trace_runtime(reason, diagnostics):
    global PERF_LAST_LOG_AT
    now = time.time()
    if now - float(PERF_LAST_LOG_AT or 0) < PERF_LOG_INTERVAL_SECONDS:
        return
    PERF_LAST_LOG_AT = now
    excel_pids = set()
    try:
        excel_pids.update(int(p) for p in diagnostics.get("sessionPids") or [] if p)
    except Exception:
        pass
    try:
        excel_pids.update(int(row.get("pid")) for row in diagnostics.get("trackedPids") or [] if row.get("pid"))
    except Exception:
        pass
    python_pid = diagnostics.get("pythonSkillPid")
    if python_pid:
        try:
            excel_pids.add(int(python_pid))
        except Exception:
            pass
    _perf_trace(
        "runtime.load",
        reason=reason,
        backend=_process_perf_snapshot(os.getpid()),
        excel=[_process_perf_snapshot(pid) for pid in sorted(excel_pids)],
        diagnostics=diagnostics,
    )


def _excel_queue_size():
    try:
        return int(EXCEL_QUEUE.qsize()) if EXCEL_QUEUE is not None else 0
    except Exception:
        return None


def _pipeline_job_stats():
    now = time.time()
    with PIPELINE_JOBS_LOCK:
        jobs = list(PIPELINE_JOBS.values())
    running = [j for j in jobs if str(j.get("status") or "").lower() == "running"]
    ages = [
        now - float(j.get("created") or j.get("updated") or now)
        for j in jobs
    ]
    return {
        "count": len(jobs),
        "running": len(running),
        "oldestAgeSec": round(max(ages), 1) if ages else 0,
    }


def _pipeline_snapshot_stats():
    snapshots_root = (BACKEND_DIR / "pipeline_step_snapshots").resolve()
    total = 0
    file_count = 0
    missing = 0
    for snapshot in list(PIPELINE_STEP_SNAPSHOTS.values()):
        for raw_path in (snapshot.get("files") or {}).values():
            try:
                path = Path(raw_path)
                if snapshots_root not in path.resolve().parents:
                    continue
                if path.exists():
                    total += path.stat().st_size
                    file_count += 1
                else:
                    missing += 1
            except Exception:
                missing += 1
    return {
        "count": len(PIPELINE_STEP_SNAPSHOTS),
        "files": file_count,
        "missingFiles": missing,
        "bytes": total,
        "mb": round(total / (1024 * 1024), 1),
    }


def _delete_pipeline_snapshot_entry(key, snapshot):
    snapshots_root = (BACKEND_DIR / "pipeline_step_snapshots").resolve()
    dirs = set()
    for raw_path in (snapshot.get("files") or {}).values():
        try:
            path = Path(raw_path).resolve()
            if snapshots_root in path.parents:
                dirs.add(path.parent)
                path.unlink(missing_ok=True)
        except Exception:
            pass
    for directory in sorted(dirs, key=lambda p: len(str(p)), reverse=True):
        try:
            if directory != snapshots_root and snapshots_root in directory.parents:
                shutil.rmtree(directory, ignore_errors=True)
        except Exception:
            pass


def _cleanup_pipeline_snapshots_by_limits():
    before = _pipeline_snapshot_stats()
    ordered = sorted(PIPELINE_STEP_SNAPSHOTS.items(), key=lambda item: item[1].get("created", 0))
    removed = 0
    while ordered:
        stats = _pipeline_snapshot_stats()
        if stats["count"] <= MAX_PIPELINE_STEP_SNAPSHOTS and stats["bytes"] <= HOUSEKEEPING_SNAPSHOT_MAX_BYTES:
            break
        key, snapshot = ordered.pop(0)
        if key in PIPELINE_STEP_SNAPSHOTS:
            PIPELINE_STEP_SNAPSHOTS.pop(key, None)
            _delete_pipeline_snapshot_entry(key, snapshot)
            removed += 1
    after = _pipeline_snapshot_stats()
    return {"removed": removed, "before": before, "after": after}


def _cleanup_stale_copy_source(max_age_seconds=600):
    try:
        ts = float(LAST_COPY_SOURCE.get("ts") or 0)
    except Exception:
        ts = 0
    if ts and time.monotonic() - ts > max_age_seconds:
        LAST_COPY_SOURCE.clear()
        return True
    return False


def _maintenance_status():
    return {
        "cleanupLastRunAt": HOUSEKEEPING_LAST_RUN_AT,
        "cleanupLastDurationMs": HOUSEKEEPING_LAST_DURATION_MS,
        "cleanupLastSkippedReason": HOUSEKEEPING_LAST_SKIPPED_REASON,
        "cleanupRunCount": HOUSEKEEPING_RUN_COUNT,
        "cleanupError": HOUSEKEEPING_ERROR,
        "runtimeSamplerIntervalSeconds": RUNTIME_SAMPLER_INTERVAL_SECONDS,
        "housekeepingIntervalSeconds": HOUSEKEEPING_INTERVAL_SECONDS,
    }


def _runtime_counts_snapshot():
    return {
        "workbooks": len(WORKBOOKS),
        "results": len(RESULTS),
        "diffs": len(DIFFS),
        "pipelineSnapshots": len(PIPELINE_STEP_SNAPSHOTS),
        "pipelineJobs": len(PIPELINE_JOBS),
        "excelSessions": len(EXCEL_SESSIONS),
        "trackedExcelPids": len(SPAWNED_EXCEL_PIDS),
    }


def _sample_lock_contended():
    acquired = False
    try:
        acquired = EXCEL_LOCK.acquire(timeout=0.1)
        return not acquired
    except Exception:
        return True
    finally:
        if acquired:
            try:
                EXCEL_LOCK.release()
            except Exception:
                pass


def _runtime_sampler_once():
    global RUNTIME_LAST_ACTIVITY_SIG, RUNTIME_LAST_ACTIVITY_AT
    # 유휴 판단은 진단 수집(엑셀 pid 조회 등)보다 먼저, 싸게 얻을 수 있는 신호만으로 한다.
    # 실행중 작업/큐가 있으면 항상 기록하고, 그 외에는 카운트 시그니처가 직전과 달라졌을 때만
    # 기록한다(작업이 끝나 카운트가 정착하는 마지막 1회까지는 기록되고 이후 침묵).
    sig = json.dumps({
        "counts": _runtime_counts_snapshot(),
        "queue": _excel_queue_size() or 0,
        "running": _pipeline_job_stats().get("running", 0),
    }, sort_keys=True)
    if not _pipeline_is_busy() and sig == RUNTIME_LAST_ACTIVITY_SIG:
        return
    RUNTIME_LAST_ACTIVITY_SIG = sig
    RUNTIME_LAST_ACTIVITY_AT = time.time()
    diagnostics = _excel_runtime_diagnostics(reap=False, log=False) if excel_available() else None
    excel_pids = set()
    if diagnostics:
        try:
            excel_pids.update(int(p) for p in diagnostics.get("sessionPids") or [] if p)
        except Exception:
            pass
        try:
            excel_pids.update(int(row.get("pid")) for row in diagnostics.get("trackedPids") or [] if row.get("pid"))
        except Exception:
            pass
        if diagnostics.get("pythonSkillPid"):
            try:
                excel_pids.add(int(diagnostics.get("pythonSkillPid")))
            except Exception:
                pass
    _perf_trace(
        "runtime.sample",
        backend=_process_perf_snapshot(os.getpid()),
        excel=[_process_perf_snapshot(pid) for pid in sorted(excel_pids)],
        queueSize=_excel_queue_size(),
        excelLockContended=_sample_lock_contended(),
        counts=_runtime_counts_snapshot(),
        pipelineJobs=_pipeline_job_stats(),
        snapshots=_pipeline_snapshot_stats(),
        diagnostics=diagnostics,
        maintenance=_maintenance_status(),
    )


def _pipeline_is_busy():
    try:
        return _excel_queue_size() not in (None, 0) or _pipeline_job_stats().get("running", 0) > 0
    except Exception:
        return True


def _run_low_risk_housekeeping():
    global HOUSEKEEPING_RUNNING, HOUSEKEEPING_LAST_RUN_AT, HOUSEKEEPING_LAST_DURATION_MS
    global HOUSEKEEPING_LAST_SKIPPED_REASON, HOUSEKEEPING_RUN_COUNT, HOUSEKEEPING_ERROR, HOUSEKEEPING_GC_LAST_AT
    if HOUSEKEEPING_RUNNING:
        HOUSEKEEPING_LAST_SKIPPED_REASON = "already-running"
        _perf_trace("runtime.housekeeping.skip", reason=HOUSEKEEPING_LAST_SKIPPED_REASON)
        return
    if _pipeline_is_busy():
        HOUSEKEEPING_LAST_SKIPPED_REASON = "pipeline-or-queue-busy"
        _perf_trace("runtime.housekeeping.skip", reason=HOUSEKEEPING_LAST_SKIPPED_REASON)
        return
    acquired = False
    try:
        acquired = EXCEL_LOCK.acquire(timeout=0.05)
    except Exception:
        acquired = False
    if not acquired:
        HOUSEKEEPING_LAST_SKIPPED_REASON = "excel-lock-busy"
        _perf_trace("runtime.housekeeping.skip", reason=HOUSEKEEPING_LAST_SKIPPED_REASON)
        return
    try:
        EXCEL_LOCK.release()
    except Exception:
        pass

    started = time.perf_counter()
    HOUSEKEEPING_RUNNING = True
    HOUSEKEEPING_LAST_SKIPPED_REASON = ""
    HOUSEKEEPING_ERROR = ""
    detail = {}
    try:
        detail["copySourceCleared"] = _cleanup_stale_copy_source()
        with PIPELINE_JOBS_LOCK:
            prune_pipeline_jobs_locked()
        detail["snapshots"] = _cleanup_pipeline_snapshots_by_limits()
        detail["excel"] = _excel_runtime_diagnostics(reap=True, log=False) if excel_available() else None
        now = time.time()
        if now - float(HOUSEKEEPING_GC_LAST_AT or 0) >= 300:
            detail["gcCollected"] = gc.collect()
            HOUSEKEEPING_GC_LAST_AT = now
        HOUSEKEEPING_RUN_COUNT += 1
        HOUSEKEEPING_LAST_RUN_AT = time.time()
    except Exception as err:
        HOUSEKEEPING_ERROR = str(err)
        detail["error"] = str(err)
    finally:
        HOUSEKEEPING_LAST_DURATION_MS = round((time.perf_counter() - started) * 1000, 1)
        HOUSEKEEPING_RUNNING = False
        # [유휴 무기록] 아무것도 정리하지 않은 정기 하우스키핑까지 매번 기록하면 유휴 상태에서도
        # 10분마다 트레이스가 자란다. 오류·실제 정리 발생·최근 활동 중 하나일 때만 남긴다.
        did_work = bool(detail.get("copySourceCleared")) or bool((detail.get("snapshots") or {}).get("removed"))
        recent_activity = (time.time() - float(RUNTIME_LAST_ACTIVITY_AT or 0)) < max(60.0, HOUSEKEEPING_INTERVAL_SECONDS)
        if HOUSEKEEPING_ERROR or did_work or recent_activity:
            _perf_trace(
                "runtime.housekeeping",
                durationMs=HOUSEKEEPING_LAST_DURATION_MS,
                skippedReason=HOUSEKEEPING_LAST_SKIPPED_REASON,
                error=HOUSEKEEPING_ERROR,
                detail=detail,
            )


def _runtime_maintenance_loop():
    next_sample = 0.0
    next_parent_watch = 0.0
    next_housekeeping = time.time() + max(60.0, HOUSEKEEPING_INTERVAL_SECONDS)
    while True:
        now = time.time()
        if now >= next_parent_watch:
            _native_parent_watch_once(now)
            next_parent_watch = now + max(1.0, PARENT_WATCH_INTERVAL_SECONDS)
        if now >= next_sample:
            try:
                _runtime_sampler_once()
            except Exception as err:
                _perf_trace("runtime.sample.error", error=str(err))
            next_sample = now + max(5.0, RUNTIME_SAMPLER_INTERVAL_SECONDS)
        if now >= next_housekeeping:
            _run_low_risk_housekeeping()
            next_housekeeping = now + max(60.0, HOUSEKEEPING_INTERVAL_SECONDS)
        time.sleep(1.0)


def _native_parent_watch_once(now):
    """Native host가 사라졌는데 Python 서버만 살아남으면 Excel COM 인스턴스도 고아로 남는다.

    Native host가 서버를 띄운 경우에만 B2B_NATIVE_HOST_PID를 넘기므로, 일반 단독 서버 실행에는
    영향을 주지 않는다. 부모가 사라진 상태가 잠깐의 프로세스 전환이 아니라는 것을 grace로 확인한 뒤
    서버가 직접 app-owned Excel을 정리하고 종료한다.
    """
    global PARENT_WATCH_MISSING_SINCE
    if DISABLE_PARENT_WATCH or not NATIVE_HOST_PID:
        return
    if _is_pid_alive(NATIVE_HOST_PID):
        PARENT_WATCH_MISSING_SINCE = 0.0
        return
    if not PARENT_WATCH_MISSING_SINCE:
        PARENT_WATCH_MISSING_SINCE = now
        _perf_trace("runtime.parent.missing", parentPid=NATIVE_HOST_PID)
        return
    if now - PARENT_WATCH_MISSING_SINCE < max(1.0, PARENT_WATCH_GRACE_SECONDS):
        return
    _perf_trace(
        "runtime.parent.missing.shutdown",
        parentPid=NATIVE_HOST_PID,
        missingForSeconds=round(now - PARENT_WATCH_MISSING_SINCE, 1),
    )
    try:
        cleanup_excel_sessions()
    except Exception as err:
        _perf_trace("runtime.parent.missing.cleanup_error", error=str(err))
    try:
        cleanup_node_worker()
    except Exception:
        pass
    os._exit(0)


def start_runtime_maintenance_threads():
    global RUNTIME_SAMPLER_STARTED
    if RUNTIME_SAMPLER_STARTED:
        return
    RUNTIME_SAMPLER_STARTED = True
    # [시작 유지관리 통합] 진입점이 둘이다: 프로즌 exe(entry=launch_b2b.py)와 직접 실행(python serve_b2b.py).
    # 예전엔 트레이스 리셋/이전 실행 잔재 정리를 serve_b2b.__main__ 에만 뒀는데, 프로즌 exe 는 launch_b2b.py 를
    # 진입점으로 serve_b2b 를 '모듈로 import' 만 하므로 그 __main__ 이 실행되지 않아 정리가 아예 안 돌았다
    # (→ %TEMP%\b2b_backend_v044 에 업로드 작업본 xlsx 누적, 트레이스 로그도 안 비워짐). 두 진입점이 모두
    # 호출하는 이 함수(멱등)로 옮겨 진입점이 갈려도 한 번은 반드시 실행되게 한다.
    try:
        _reset_trace_logs()
    except Exception:
        pass
    try:
        threading.Thread(target=cleanup_stale_temp_artifacts, name="b2b-temp-cleanup", daemon=True).start()
    except Exception:
        pass
    thread = threading.Thread(target=_runtime_maintenance_loop, name="b2b-runtime-maintenance", daemon=True)
    thread.start()


def _excel_runtime_diagnostics(reap=False, log=True):
    """Return lightweight Excel process diagnostics and optionally reap app-owned orphan PIDs.

    Only PIDs recorded in SPAWNED_EXCEL_PIDS are candidates. User-launched Excel
    processes are never touched.
    """
    global EXCEL_LAST_REAP_AT
    lock_unavailable = False
    acquired = False
    try:
        acquired = EXCEL_LOCK.acquire(timeout=0.5)
    except Exception:
        acquired = False
    if not acquired:
        lock_unavailable = True
    try:
        session_pids = {
            int(session.get("pid"))
            for session in list(EXCEL_SESSIONS.values())
            if session.get("pid")
        }
        tracked_pids = {int(p) for p in SPAWNED_EXCEL_PIDS if p}
        python_pid = int(PYTHON_SKILL_APP_PID) if PYTHON_SKILL_APP_PID else None
    finally:
        if acquired:
            try:
                EXCEL_LOCK.release()
            except Exception:
                pass
    protected = set(session_pids)
    if python_pid:
        protected.add(python_pid)

    now = time.time()
    reap_skipped_reason = ""
    should_reap = bool(reap) and (now - float(EXCEL_LAST_REAP_AT or 0)) >= EXCEL_REAP_INTERVAL_SECONDS
    if should_reap and lock_unavailable:
        # If the Excel lock is busy, a pipeline/open operation may be in the
        # middle of registering its protected session PID. Reaping from a stale
        # snapshot can kill that in-flight app-owned Excel instance.
        should_reap = False
        reap_skipped_reason = "lock-unavailable"
    rows = []
    reaped = []
    stale = []
    for pid in sorted(tracked_pids):
        alive = _is_pid_alive(pid)
        protected_pid = pid in protected
        if should_reap and alive and not protected_pid:
            _force_kill_pid(pid)
            reaped.append(pid)
            alive = _is_pid_alive(pid)
        if not alive:
            stale.append(pid)
        rows.append({
            "pid": pid,
            "alive": bool(alive),
            "protected": bool(protected_pid),
            "kind": "python-skill" if pid == python_pid else ("session" if pid in session_pids else "orphan"),
        })
    if should_reap:
        EXCEL_LAST_REAP_AT = now
    if stale or reaped:
        acquired = False
        try:
            acquired = EXCEL_LOCK.acquire(timeout=0.5)
            if acquired:
                for pid in set(stale + reaped):
                    SPAWNED_EXCEL_PIDS.discard(pid)
        except Exception:
            pass
        finally:
            if acquired:
                try:
                    EXCEL_LOCK.release()
                except Exception:
                    pass
    diagnostics = {
        "sessions": len(session_pids),
        "sessionPids": sorted(session_pids),
        "trackedPids": rows,
        "reapedPids": reaped,
        "pythonSkillPid": python_pid,
        "lastReapAt": EXCEL_LAST_REAP_AT,
        "lockUnavailable": lock_unavailable,
        "reapSkippedReason": reap_skipped_reason,
    }
    if log:
        _maybe_perf_trace_runtime("excel-diagnostics", diagnostics)
    return diagnostics


def _b2b_runner_trusted_dir():
    """러너 .xlsm 를 만들 고정 폴더(Excel 신뢰 위치로 등록되는 곳)."""
    d = Path(tempfile.gettempdir()) / "b2b_runner_trusted"
    try:
        d.mkdir(parents=True, exist_ok=True)
    except Exception:
        pass
    return d


def _ensure_runner_trusted_location():
    """러너 폴더를 Excel '신뢰할 수 있는 위치(Trusted Location)'로 등록한다.
    [핵심] 신뢰 위치의 매크로는 Trust Center 매크로 설정(VBAWarnings)·실행 타이밍·인스턴스 상태와
    무관하게 '항상' 실행된다. 전체실행이 어떤 Excel 인스턴스에선 매크로 실행이 막히던(간헐적
    "매크로를 실행할 수 없습니다") 현상을, 러너를 신뢰 위치에서 실행해 결정적으로 우회한다.
    Excel 시작 시점에 읽으므로 인스턴스 DispatchEx 전에 호출한다."""
    base = _b2b_runner_trusted_dir()
    try:
        import winreg
    except Exception:
        return base
    path_str = str(base)
    if not path_str.endswith("\\"):
        path_str += "\\"
    for ver in ("16.0", "15.0", "14.0", "12.0"):
        try:
            sec = r"Software\Microsoft\Office\%s\Excel\Security" % ver
            # 신뢰 위치 전체가 꺼져 있으면 켜고, %TEMP% 하위도 허용.
            try:
                tl = winreg.CreateKey(winreg.HKEY_CURRENT_USER, sec + r"\Trusted Locations")
                try:
                    winreg.SetValueEx(tl, "AllLocationsDisabled", 0, winreg.REG_DWORD, 0)
                finally:
                    winreg.CloseKey(tl)
            except Exception:
                pass
            # 전용 신뢰 위치 등록(고정 키명 — 기존 Location0..N 과 충돌 없음, Excel 은 하위키 전부 열거).
            loc = winreg.CreateKey(winreg.HKEY_CURRENT_USER, sec + r"\Trusted Locations\B2BRunner")
            try:
                winreg.SetValueEx(loc, "Path", 0, winreg.REG_SZ, path_str)
                winreg.SetValueEx(loc, "AllowSubFolders", 0, winreg.REG_DWORD, 1)
                winreg.SetValueEx(loc, "Description", 0, winreg.REG_SZ, "B2B VBA runner (auto)")
            finally:
                winreg.CloseKey(loc)
        except Exception:
            pass
    return base


def _ensure_vbom_access():
    """매크로 주입·실행에 필요한 Trust Center 플래그를 켠다(HKCU).
    - AccessVBOM=1: wb.VBProject 접근 허용(매크로 주입에 필요). 꺼져 있으면 주입 자체가 막힌다.
    - VBAWarnings=1: '모든 매크로 사용'(매크로 실행에 필요). 이게 1이 아니면 일부 인스턴스에서
      주입한 러너 매크로의 Application.Run 이 "매크로를 실행할 수 없습니다"로 막힌다(전체실행 간헐 실패의
      근본원인 — injection 은 되는데 RUN 만 차단). 앱이 만든 임시 러너 + 사용자 .xlsx 라 실행 허용이 안전.
    이 값들은 '이후 새로 띄우는' Excel 인스턴스에 적용되므로 라이브 Excel DispatchEx 직전에 호출한다."""
    try:
        import winreg
    except Exception:
        return False
    enabled = False
    for ver in ("16.0", "15.0", "14.0", "12.0"):
        try:
            key = winreg.CreateKey(
                winreg.HKEY_CURRENT_USER,
                r"Software\Microsoft\Office\%s\Excel\Security" % ver,
            )
            try:
                winreg.SetValueEx(key, "AccessVBOM", 0, winreg.REG_DWORD, 1)
                winreg.SetValueEx(key, "VBAWarnings", 0, winreg.REG_DWORD, 1)  # 1=모든 매크로 사용
                enabled = True
            finally:
                winreg.CloseKey(key)
        except Exception:
            pass
    try:
        _ensure_runner_trusted_location()  # 러너 폴더를 신뢰 위치로 — 매크로 항상 실행 허용(핵심)
    except Exception:
        pass
    return enabled


def _disable_vba_break_on_all_errors():
    """VBE Error Trapping 이 Break on All Errors 면 처리된 오류도 디버거로 진입한다.
    Excel 은 이 값을 시작 시점에 읽으므로 새 Excel 인스턴스 생성 전에 HKCU 값을 Break on Unhandled Errors(0)로 둔다."""
    try:
        import winreg
    except Exception:
        return False
    changed = False
    views = [0]
    for view_name in ("KEY_WOW64_64KEY", "KEY_WOW64_32KEY"):
        view = getattr(winreg, view_name, 0)
        if view and view not in views:
            views.append(view)
    for ver in ("7.1", "7.0"):
        for view in views:
            try:
                key = winreg.CreateKeyEx(
                    winreg.HKEY_CURRENT_USER,
                    r"Software\Microsoft\VBA\%s\Common" % ver,
                    0,
                    winreg.KEY_SET_VALUE | view,
                )
                try:
                    winreg.SetValueEx(key, "BreakOnAllErrors", 0, winreg.REG_DWORD, 0)
                    changed = True
                finally:
                    winreg.CloseKey(key)
            except Exception:
                pass
    return changed
def _is_live_shared_app(app):
    global LIVE_EXCEL_APP
    if app is None or LIVE_EXCEL_APP is None:
        return False
    try:
        return int(app.Hwnd) == int(LIVE_EXCEL_APP.Hwnd)
    except Exception:
        return app is LIVE_EXCEL_APP


def _get_live_excel_app():
    """라이브 편집 워크북을 한 Excel 프로세스 안에 모으기 위한 앱 전용 Excel.Application.

    기존 구조는 파일마다 DispatchEx("Excel.Application")를 호출해 EXCEL.EXE가 여러 개
    생길 수 있었다. 저사양/VDI PC에서는 프로세스 다중화와 창 전환 비용이 커지므로,
    라이브 편집 세션은 이 인스턴스 하나에 여러 Workbook을 여는 방식으로 실험한다.
    사용자가 직접 띄운 Excel을 잡지 않기 위해 GetActiveObject는 쓰지 않고, 앱 전용
    DispatchEx 인스턴스를 최초 1회만 만든다.
    """
    global LIVE_EXCEL_APP, LIVE_EXCEL_APP_PID
    app = LIVE_EXCEL_APP
    if app is not None:
        try:
            _ = app.Workbooks.Count
            return app
        except Exception as probe_err:
            # [죽음 오판 방지] Workbooks.Count 예외 ≠ 프로세스 사망. 사용자가 셀 편집/모달 중이면
            # Excel 이 COM 호출을 일시 거부한다(RPC_E_CALL_REJECTED/RETRYLATER — 정상 동작).
            # 이때 재스폰하면 ① 워크북 0개 빈 EXCEL.EXE 고아가 남고(회색 창, 안 사라짐)
            # ② 진행 중 녹화의 stop 이 새 빈 인스턴스를 조회해 harvested=0 이 된다
            # (실측 2026-07-29 15:34:21 — pid 13124 생존 중인데 get_live_found_dead 로 62860 재스폰).
            # → pid 가 실제로 살아 있으면 기존 프록시를 그대로 반환한다(호출측 작업은 재시도로 회복).
            _alive = False
            try:
                _alive = bool(LIVE_EXCEL_APP_PID) and _is_pid_alive(LIVE_EXCEL_APP_PID)
            except Exception:
                _alive = False
            if _alive:
                try:
                    _vba_trace("live_app.busy_not_dead", pid=int(LIVE_EXCEL_APP_PID),
                               error=str(probe_err)[:160],
                               duringRecording=bool(NATIVE_RECORDING.get("active")) if "NATIVE_RECORDING" in globals() else False)
                except Exception:
                    pass
                return app
            _note_live_app_reset("get_live_found_dead")   # pid 도 죽었다 — 진짜 사망, 새로 만든다(=스폰)
            LIVE_EXCEL_APP = None
            LIVE_EXCEL_APP_PID = None
    _ensure_vbom_access()
    app = win32com.client.DispatchEx("Excel.Application")
    _track_spawned_excel_app(app)  # [0.5.2 이식] 고아 Excel 추적
    app.Visible = False
    app.DisplayAlerts = False
    app.EnableEvents = False
    for attr, value in (("AskToUpdateLinks", False), ("UserControl", True)):
        try:
            setattr(app, attr, value)
        except Exception:
            pass
    LIVE_EXCEL_APP = app
    try:
        LIVE_EXCEL_APP_PID = int(_excel_process_id(app) or 0) or None
    except Exception:
        LIVE_EXCEL_APP_PID = None
    return app


def _excel_grid_hwnds_for_pid(pid):
    """해당 pid 의 XLMAIN 최상위 창 아래 EXCEL7(그리드) 자식 hwnd 목록 — 셀 편집 확정 키 전송 대상."""
    result = []
    if win32gui is None or not pid:
        return result

    def _top(hwnd, _):
        try:
            _tid, wpid = win32process.GetWindowThreadProcessId(hwnd)
            if int(wpid) != int(pid) or win32gui.GetClassName(hwnd) != "XLMAIN":
                return True

            def _child(ch, _2):
                try:
                    if win32gui.GetClassName(ch) == "EXCEL7":
                        result.append(int(ch))
                except Exception:
                    pass
                return True

            try:
                win32gui.EnumChildWindows(hwnd, _child, None)
            except Exception:
                pass
        except Exception:
            pass
        return True

    try:
        win32gui.EnumWindows(_top, None)
    except Exception:
        pass
    return result


def _commit_pending_excel_cell_edit(app, max_wait_s=2.5):
    """[셀 편집 확정] 사용자가 셀 편집(in-cell edit) 중이면 Excel 이 COM 을 거부해
    (RPC_E_CALL_REJECTED) 녹화 시작/정지가 실패한다 — 클라 버튼만 눌리고 레코더는 계속 도는
    이중 상태가 됐다(실측 2026-07-29: '셀 편집 중 녹화 종료 안 됨, 다시 찍어야 함').
    포커스/포그라운드를 훔치지 않고(EXCEL7 그리드에 PostMessage Enter — 다이얼로그 자동확인과
    동일 관용구) 편집을 '확정'시킨 뒤 COM 이 응답할 때까지 짧게 재시도한다.
    정상 상태면 키 전송 0회로 즉시 True(부작용 없음). Enter 확정이므로 입력 중이던 값은 셀에
    반영되고, 그 대입은 매크로 레코더에 정상 기록된다(사용자 의도 보존)."""
    try:
        _ = app.Workbooks.Count
        return True     # COM 정상 응답 — 편집 중 아님, 아무 것도 안 보낸다
    except Exception:
        pass
    pid = LIVE_EXCEL_APP_PID
    if not pid:
        try:
            pid = _excel_process_id(app)   # 편집 중엔 이것도 거부될 수 있음 — 폴백일 뿐
        except Exception:
            pid = None
    deadline = time.time() + float(max_wait_s)
    sent = 0
    while time.time() < deadline:
        for hwnd in _excel_grid_hwnds_for_pid(pid):
            try:
                win32gui.PostMessage(hwnd, win32con.WM_KEYDOWN, win32con.VK_RETURN, 0)
                win32gui.PostMessage(hwnd, win32con.WM_KEYUP, win32con.VK_RETURN, 0)
                sent += 1
            except Exception:
                pass
        time.sleep(0.2)
        try:
            _ = app.Workbooks.Count
            try:
                _vba_trace("record.cell_edit.committed", sentKeys=sent)
            except Exception:
                pass
            return True
        except Exception:
            continue
    try:
        _vba_trace("record.cell_edit.commit_failed", sentKeys=sent)
    except Exception:
        pass
    return False


def _quit_live_excel_app():
    global LIVE_EXCEL_APP
    app = LIVE_EXCEL_APP
    if app is not None:
        _note_live_app_reset("quit_live_excel_app")
    LIVE_EXCEL_APP = None
    if app is None:
        return
    try:
        app.Quit()
    except Exception:
        pass


def _remaining_sessions_for_pid(pid):
    if not pid:
        return []
    try:
        return [
            s for s in EXCEL_SESSIONS.values()
            if s.get("pid") and int(s.get("pid")) == int(pid)
        ]
    except Exception:
        return []


def _open_excel_session_impl(
    path,
    name=None,
    workbook_id=None,
    result_id=None,
    read_only_mirror=False,
    left=None,
    top=None,
    width=None,
    height=None,
    client_left=None,
    client_top=None,
    client_width=None,
    client_height=None,
    viewport_width=None,
    viewport_height=None,
    browser_title=None,
    native_parent_hwnd=None,
    native_host_hwnd=None,
    native_overlay=False,
    live_editable=False,
    defer_visible=False,
    from_state_sig=None,
):
    if not excel_available():
        raise RuntimeError("Microsoft Excel COM automation is not available. Excel and pywin32 are required.")
    path = Path(path)
    if not path.exists():
        raise RuntimeError(f"file not found: {path}")
    source_path = path
    working_copy_path = None
    if live_editable:
        # [2A 안전장치 b] 새 라이브 세션을 열면 미러는 다시 떠야 한다 — 억제 플래그 리셋.
        # [실행기 오버레이 수정 2026-08-04] 단, defer_visible(실행기 헤드리스가 '숨겨서' 여는
        # 오픈)은 리셋하지 않는다 — 토글/수정의 잔여 open 이 실행기 이동 후 도착하면서 억제를
        # 무장해제해, 이후 서버 표시 경로가 실행기 화면 위로 회색 Excel 을 띄웠다(실측).
        if not defer_visible:
            global LIVE_RESTORE_SUPPRESSED
            LIVE_RESTORE_SUPPRESSED = False
        # 리모콘 모델: 원본은 절대 건드리지 않는다 — 항상 작업용 복사본을 만들어
        # 그 위에서 라이브 실행/스킬(VBA) 적용을 한다. 다운로드는 이 복사본을 저장.
        # 워크북 이름이 원본과 동일해야 VBA 의 Workbooks("원본명")/ActiveWorkbook 와 @파일 참조가 일치하고
        # 제목줄도 깔끔하다 → 고유 하위폴더 안에 '원본 이름 그대로' 복사한다.
        BACKEND_DIR.mkdir(parents=True, exist_ok=True)
        clean_name = Path(name).name if name else source_path.name
        live_dir = BACKEND_DIR / f"live_{uuid.uuid4().hex}"
        live_dir.mkdir(parents=True, exist_ok=True)
        working_copy_path = live_dir / clean_name
        # [새로고침 즉시복원 2026-08-04] 복원 요청(from_state_sig)이면 '원본' 대신 '스킬 전부 적용된
        # 최종 상태 사본'을 작업복사본의 원료로 쓴다. 워크북 레코드의 path(원본)는 손대지 않으므로
        # 리셋/되돌리기는 그대로 진짜 원본을 찾는다 — 오염 없음. 사본이 없으면 조용히 원본으로 간다.
        copy_src = source_path
        if from_state_sig and workbook_id:
            _snap = _find_live_final_snapshot(WORKBOOKS.get(workbook_id), from_state_sig)
            if _snap:
                copy_src = Path(_snap["path"])
                _vba_trace("excel.open.from_live_final_snapshot", workbookId=workbook_id, path=str(copy_src))
        shutil.copy2(copy_src, working_copy_path)
        path = working_copy_path
        # VBA 주입(스킬 적용)이 가능하도록, 이 라이브 인스턴스를 띄우기 전에 AccessVBOM 을 켠다.
        _ensure_vbom_access()
        _disable_vba_break_on_all_errors()
    # read_only_mirror(읽기전용 보호 미러)와 live_editable(라이브) 모두 검증된 미러 표시 경로를 그대로 쓴다
    # (owner/프레임리스 커스텀은 선택/사이즈를 깨서 폐기). 라이브 차이는 작업복사본·VBA·저장뿐.
    # 작업복사본을 read-only 로 열어도 VBA/COM 은 메모리에서 수정 가능하고, 다운로드는 SaveCopyAs 로 저장.
    manage_overlay = bool(read_only_mirror) or bool(live_editable)
    # 라이브는 작업복사본을 편집가능(read_only=False)으로 연다(VBA 삽입/리셋 등 인메모리 변경 보장).
    # 사용자 직접 편집은 시트 보호(UserInterfaceOnly)로 막으므로 화면 동작은 미러와 동일.
    open_read_only = bool(read_only_mirror)
    with EXCEL_LOCK:
        browser_hwnd = None if (native_parent_hwnd or native_overlay) else (_capture_browser_hwnd(browser_title) if read_only_mirror else None)
        app = _get_live_excel_app() if live_editable else win32com.client.DispatchEx("Excel.Application")
        _track_spawned_excel_app(app)  # [0.5.2 이식] 고아 Excel 추적
        live_frame_mode = bool(live_editable and LIVE_FRAME_MODE)
        if live_frame_mode:
            # frame 모드: 공유 인스턴스의 글로벌 Visible 토글 금지.
            # app.Visible=False 는 모든 프레임을 동시에 숨겨, 그중 포그라운드였던 창이 '소멸'하며
            # OS 가 z-order 의 무관한 다른 앱 창을 활성화(최상단 점프)하는 원인이 된다.
            # 새 프레임은 아래에서 개별적으로 파킹한다.
            pass
        else:
            app.Visible = False if manage_overlay else True
        app.DisplayAlerts = False
        app.EnableEvents = False
        if manage_overlay:
            try:
                app.ScreenUpdating = False
            except Exception:
                pass
        try:
            app.UserControl = True
        except Exception:
            pass
        try:
            app.AskToUpdateLinks = False
        except Exception:
            pass
        if defer_visible and not live_frame_mode:
            try:
                app.WindowState = -4143  # xlNormal
            except Exception:
                pass
            for attr, value in (("Left", -32000), ("Top", -32000), ("Width", 10), ("Height", 10)):
                try:
                    setattr(app, attr, value)
                except Exception:
                    pass
        wb = None
        open_temp_path = None
        frame_hwnd = None
        try:
            # [동명 세션/고아 자가치유] 공유 라이브 인스턴스에 같은 이름의 워크북이 이미 열려 있으면
            # Excel 이 동명 2개를 거부해 Workbooks.Open 이 실패한다(실측 14:02 — 재현 실패로 클라
            # 매핑은 잊혔는데 워크북은 남아 재오픈이 영원히 실패 = '탭 전환 불능'의 뿌리).
            #  ① 살아있는 세션 소유면: 그 세션을 그대로 재사용(reattach) — 같은 업로드의 라이브 사본.
            #  ② 소유자 없는 고아면: 저장 없이 닫고 새로 연다(작업사본이라 원본 데이터 무손실).
            if live_editable:
                _wanted_name = str(Path(str(path)).name).lower()
                _same_wb = None
                try:
                    for _wb in app.Workbooks:
                        try:
                            if str(_wb.Name).lower() == _wanted_name:
                                _same_wb = _wb
                                break
                        except Exception:
                            continue
                except Exception:
                    _same_wb = None
                if _same_wb is not None:
                    _owner = None
                    for _sid, _sess in list(EXCEL_SESSIONS.items()):
                        try:
                            _swb = _sess.get("workbook")
                            if _swb is not None and str(_swb.Name).lower() == _wanted_name:
                                _owner = _sess
                                break
                        except Exception:
                            continue
                    if _owner is not None:
                        _vba_trace("excel.open.reattach", excelId=_owner.get("id"), name=_wanted_name)
                        try:
                            shutil.rmtree(live_dir, ignore_errors=True)  # 방금 만든 새 작업사본은 불용
                        except Exception:
                            pass
                        if native_host_hwnd:
                            _owner["nativeHostHwnd"] = native_host_hwnd
                        return {
                            "ok": True,
                            "excelId": _owner.get("id"),
                            "name": _owner.get("name"),
                            "path": _owner.get("path"),
                            "sheetNames": _excel_collection_names(_owner["workbook"].Worksheets),
                            "readOnlyMirror": bool(_owner.get("readOnlyMirror")),
                            "liveEditable": bool(_owner.get("liveEditable")),
                            "reattached": True,
                        }
                    _vba_trace("excel.open.orphan_close", name=_wanted_name)
                    try:
                        _same_wb.Close(SaveChanges=False)
                    except Exception as _cerr:
                        _vba_trace("excel.open.orphan_close_fail", name=_wanted_name, error=str(_cerr))
            wb, open_temp_path = excel_workbooks_open(app, path, read_only=open_read_only, intended_name=name or path)
            app_pid = _excel_process_id(app)
            excel_id = uuid.uuid4().hex
            sheets = _excel_collection_names(wb.Worksheets)
            if live_editable:
                # owner 모드: 호스트를 owner 로 둔 일반(프레임 유지) Excel 창.
                # 프레임리스(WS_POPUP)는 owner 와 함께 쓰면 선택/사이즈를 깨므로 쓰지 않는다(검증됨).
                # 리본/수식줄/우클릭/입력키 차단은 앱 레벨이라 선택에 영향 없음.
                if live_frame_mode:
                    # SDI: 이 워크북의 프레임 핸들을 잡아 즉시 화면 밖으로 파킹.
                    # (보이는 공유 인스턴스에 새 프레임이 기본 캐스케이드 위치로 번쩍 뜨는 것과
                    #  그 표시/활성화가 호스트 포커스를 흔드는 것을 모두 차단)
                    frame_hwnd = _workbook_window_hwnd(wb)
                    if frame_hwnd:
                        _move_hwnd_offscreen(frame_hwnd)
                try:
                    wb.Activate()
                except Exception:
                    pass
                _protect_workbook_for_read_only_mirror(wb, True)   # 편집 차단 + 선택 허용
                _configure_excel_grid_window(app, wb)              # 리본/수식줄/우클릭/입력키 차단(앱 레벨)
                if frame_hwnd:
                    try:
                        wb.Windows(1).WindowState = -4143  # xlNormal (이 프레임만)
                    except Exception:
                        pass
                    _set_window_owner_hwnd(frame_hwnd, native_host_hwnd)  # owner (프레임 그대로 유지)
                    _style_live_frame(frame_hwnd)                         # 작업표시줄/Alt+Tab 제외
                else:
                    try:
                        app.WindowState = -4143  # xlNormal
                    except Exception:
                        pass
                    _set_excel_window_owner(app, native_host_hwnd)     # owner (프레임 그대로 유지)
                if defer_visible:
                    if frame_hwnd:
                        # [빈 회색 프레임 수정] 예전엔 여기서 app.Visible=True 를 켰지만(모든 프레임이 파킹됐다는
                        # 가정), 공유 인스턴스가 이미 보이는 상태에서 '다음' 워크북이 열리면 그 새 프레임이 파킹
                        # 되기 전에 잠깐 떠서 회색 빈 프레임으로 남았다(특히 실행기 헤드리스: 표시는 안 하는데
                        # 인스턴스만 켜져 있어 2번째 파일부터 프레임이 샘. 1개일 땐 인스턴스가 계속 숨김이라 안 뜸).
                        # defer_visible 은 '이 창을 지금 보이지 말라'는 뜻이므로 인스턴스 Visible 을 여기서 켜지
                        # 않는다 — 실제 표시가 필요할 때(showOnly → _present_live_session_frame)에서 켠다.
                        pass
                    else:
                        try:
                            app.Visible = False
                        except Exception:
                            pass
                elif frame_hwnd:
                    try:
                        if not bool(app.Visible):
                            app.Visible = True
                    except Exception:
                        pass
                    if width and height:
                        _position_excel_window(
                            app, left, top, width, height,
                            hwnd=frame_hwnd, no_activate=True,
                            client_left=client_left, client_top=client_top,
                            client_width=client_width, client_height=client_height,
                            viewport_width=viewport_width, viewport_height=viewport_height,
                            show=True,
                        )
                    else:
                        _show_window_na(frame_hwnd)
                    try:
                        app.ScreenUpdating = True
                    except Exception:
                        pass
                    _ensure_excel_workbook_view(app, wb, make_visible=False, activate=False, maximize_workbook=False, app_level=False)
                    _set_window_owner_hwnd(frame_hwnd, native_host_hwnd)
                else:
                    if width and height:
                        _position_excel_window(
                            app, left, top, width, height,
                            client_left=client_left, client_top=client_top,
                            client_width=client_width, client_height=client_height,
                            viewport_width=viewport_width, viewport_height=viewport_height,
                            show=False,
                        )
                    try:
                        app.Visible = True
                    except Exception:
                        pass
                    try:
                        app.ScreenUpdating = True
                    except Exception:
                        pass
                    _ensure_excel_workbook_view(app, wb, make_visible=True, activate=False, maximize_workbook=False)
                    if width and height:
                        _position_excel_window(
                            app, left, top, width, height,
                            viewport_width=viewport_width, viewport_height=viewport_height,
                            show=True,
                        )
                    _set_excel_window_owner(app, native_host_hwnd)
            elif read_only_mirror:
                try:
                    wb.Activate()
                except Exception:
                    pass
                _protect_workbook_for_read_only_mirror(wb, True)
                _configure_excel_grid_window(app, wb)
                if width and height:
                    _position_excel_window(
                        app,
                        left,
                        top,
                        width,
                        height,
                        browser_hwnd=browser_hwnd,
                        native_parent_hwnd=None if native_overlay else native_parent_hwnd,
                        native_host_hwnd=native_host_hwnd,
                        native_overlay=bool(native_overlay),
                        client_left=client_left,
                        client_top=client_top,
                        client_width=client_width,
                        client_height=client_height,
                        viewport_width=viewport_width,
                        viewport_height=viewport_height,
                        show=False if (native_parent_hwnd or native_overlay) else True,
                    )
                try:
                    app.Visible = True
                except Exception:
                    pass
                try:
                    app.ScreenUpdating = True
                except Exception:
                    pass
                _ensure_excel_workbook_view(
                    app,
                    wb,
                    make_visible=True,
                    activate=False if (native_parent_hwnd or native_overlay) else True,
                    maximize_workbook=False if (native_overlay or native_parent_hwnd) else True,
                )
                if width and height and (native_parent_hwnd or native_overlay):
                    _position_excel_window(
                        app,
                        left,
                        top,
                        width,
                        height,
                        native_parent_hwnd=None if native_overlay else native_parent_hwnd,
                        native_host_hwnd=native_host_hwnd,
                        native_overlay=bool(native_overlay),
                        viewport_width=viewport_width,
                        viewport_height=viewport_height,
                        show=True,
                    )
                    _ensure_excel_workbook_view(
                        app,
                        wb,
                        make_visible=True,
                        activate=False if (native_parent_hwnd or native_overlay) else True,
                        maximize_workbook=False if (native_overlay or native_parent_hwnd) else True,
                    )
                    _position_excel_window(
                        app,
                        left,
                        top,
                        width,
                        height,
                        native_parent_hwnd=None if native_overlay else native_parent_hwnd,
                        native_host_hwnd=native_host_hwnd,
                        native_overlay=bool(native_overlay),
                        viewport_width=viewport_width,
                        viewport_height=viewport_height,
                        show=True,
                    )
            EXCEL_SESSIONS[excel_id] = {
                "id": excel_id,
                "app": app,
                "workbook": wb,
                "frameHwnd": frame_hwnd,
                "pid": app_pid,
                "path": str(path),
                "openPath": str(wb.FullName),
                "openTempPath": str(open_temp_path) if open_temp_path else "",
                "name": name or path.name,
                "workbookId": workbook_id,
                "resultId": result_id,
                # 라이브(owner 모드)는 readOnlyMirror=False 로 두어 reposition 이 plain 경로(프레임리스 안 함)를 타게 한다.
                # 표시/저장/리셋 분기는 liveEditable 플래그로 구분.
                "readOnlyMirror": bool(read_only_mirror),
                "liveEditable": bool(live_editable),
                "deferredVisible": bool(defer_visible),
                "sourcePath": str(source_path),
                "workingCopyPath": str(working_copy_path) if working_copy_path else "",
                "liveRect": (
                    {"left": int(float(left or 0)), "top": int(float(top or 0)),
                     "width": int(float(width or 0)), "height": int(float(height or 0))}
                    if (live_editable or read_only_mirror) and width and height else None
                ),
                "browserHwnd": browser_hwnd,
                "nativeParentHwnd": None if (native_overlay or live_editable) else native_parent_hwnd,
                "nativeHostHwnd": native_host_hwnd,
                "nativeOverlay": False if live_editable else bool(native_overlay),
                "hidden": bool(defer_visible),
                # [이슈1 수정2] 녹화 중 새로 연 라이브 wb 는 최초 표시 때 편집잠금 재적용
                # (_present_live_session_frame 의 SHOW.TOOLBAR 포함)을 통째로 스킵한다.
                # SHOW.TOOLBAR 가 소스의 복사 마퀴를 죽여 교차파일 Ctrl+V 가 무동작이 되던
                # 버그 차단 — 열 때 이미 편집준비(보호해제)를 마치므로 재적용이 불필요.
                "recUnlockDone": bool(RECORDING_EDIT_UNLOCKED and live_editable),
                "lastNativePositionKey": (
                    f"{'overlay' if native_overlay else native_parent_hwnd}:{int(float(left or 0))}:{int(float(top or 0))}:{int(float(width or 0))}:{int(float(height or 0))}"
                    if read_only_mirror and (native_parent_hwnd or native_overlay) and width and height
                    else ""
                ),
                "created": time.time(),
            }
            # [이슈1 수정2] 녹화 편집모드 중 새 라이브 wb 를 열면 최초 표시 재적용을
            # 스킵(recUnlockDone=True)하므로, 여기서 미리 보호를 풀어 편집 준비를 마친다.
            if RECORDING_EDIT_UNLOCKED and live_editable and wb is not None:
                try:
                    _protect_workbook_for_read_only_mirror(wb, False)
                except Exception:
                    pass
            if not manage_overlay:
                try:
                    app.WindowState = -4143  # xlNormal
                    _safe_activate_excel_app(app)
                except Exception:
                    pass
            return {
                "ok": True,
                "excelId": excel_id,
                "name": name or path.name,
                "path": str(path),
                "sheetNames": sheets,
                "readOnlyMirror": bool(read_only_mirror),
                "liveEditable": bool(live_editable),
            }
        except Exception:
            try:
                if wb is not None:
                    wb.Close(SaveChanges=False)
            except Exception:
                pass
            if not _is_live_shared_app(app):
                try:
                    app.Quit()
                except Exception:
                    pass
            if open_temp_path:
                try:
                    Path(open_temp_path).unlink(missing_ok=True)
                except Exception:
                    pass
            if working_copy_path:
                try:
                    Path(working_copy_path).unlink(missing_ok=True)
                except Exception:
                    pass
            raise


def get_excel_session(excel_id):
    with EXCEL_LOCK:
        session = EXCEL_SESSIONS.get(excel_id)
    if not session:
        raise RuntimeError("Excel session not found")
    return session


def session_workbook(session):
    path = str(Path(session["path"]).resolve()).lower()
    open_path = str(Path(session.get("openPath") or session["path"]).resolve()).lower()
    try:
        app = session.get("app")
        wb = session.get("workbook")
        if wb is not None and str(Path(wb.FullName).resolve()).lower() in {path, open_path}:
            return app or wb.Application, wb
    except Exception:
        pass
    try:
        app = win32com.client.GetActiveObject("Excel.Application")
        for wb in app.Workbooks:
            if str(Path(wb.FullName).resolve()).lower() in {path, open_path}:
                return app, wb
    except Exception:
        pass
    try:
        wb = win32com.client.GetObject(str(session["path"]))
        app = wb.Application
        return app, wb
    except Exception:
        pass
    raise RuntimeError("Excel session is no longer open")


def _activate_excel_session_impl(excel_id, sheet=None, address=None):
    with EXCEL_LOCK:
        session = get_excel_session(excel_id)
        app, wb = session_workbook(session)
        if sheet:
            try:
                ws = wb.Worksheets(str(sheet))
            except Exception:
                names = _excel_collection_names(wb.Worksheets)
                normalized = normalize_text(str(sheet))
                match = next((name for name in names if normalize_text(name) == normalized or normalized in normalize_text(name)), None)
                if not match and len(names) == 1:
                    match = names[0]
                if not match:
                    raise RuntimeError(f"sheet not found: {sheet}")
                ws = wb.Worksheets(match)
            ws.Activate()
        else:
            ws = app.ActiveSheet
        if ws is None:
            names = _excel_collection_names(wb.Worksheets)
            if not names:
                raise RuntimeError("no visible worksheet")
            ws = wb.Worksheets(names[0])
            ws.Activate()
        selected_address = ""
        if address:
            try:
                _ensure_excel_workbook_view(
                    app,
                    wb,
                    make_visible=True,
                    activate=True,
                    maximize_workbook=False if (session.get("nativeOverlay") or session.get("nativeParentHwnd")) else True,
                )
                if session.get("nativeParentHwnd"):
                    _focus_excel_grid_child(app.Hwnd)
                ws.Activate()
                try:
                    wb.Windows(1).Activate()
                except Exception:
                    pass
                target = ws.Range(str(address))
                try:
                    app.Goto(target, True)
                except Exception:
                    target.Select()
                selected_address = str(address or "")
            except Exception as err:
                return {
                    "ok": True,
                    "excelId": excel_id,
                    "sheet": ws.Name,
                    "address": "",
                    "selectSkipped": True,
                    "warning": str(err),
                }
        if not session.get("readOnlyMirror"):
            if session.get("liveEditable") and LIVE_FRAME_MODE:
                # 범위 보기: 포커스는 채팅에 남기고 이 세션 프레임만 보이게/맨 위로.
                # (여기서 활성화하면 사용자의 다음 클릭/타이핑이 창 전환에 씹힌다)
                hwnd = _session_frame_hwnd(session, wb)
                if hwnd:
                    _show_window_na(hwnd)
                    _raise_excel_hwnd(hwnd)
                else:
                    try:
                        _safe_activate_excel_app(app)
                    except Exception:
                        pass
            else:
                try:
                    _safe_activate_excel_app(app)
                except Exception:
                    pass
        return {"ok": True, "excelId": excel_id, "sheet": ws.Name, "address": selected_address}


def _save_excel_session_impl(excel_id, name=None, internal=False):
    """워크북을 파일로 저장한다.

    internal=True 는 '되돌리기용 백업'(스텝 적용 전 스냅샷)이다. 사용자가 열어 볼 파일이 아니라
    나중에 /api/excel/replace 로 다시 열기만 하는 파일이라, 보호 해제/재적용과 화면 설정 복구를
    통째로 건너뛴다.
      · 왜 안전한가: 복원(replace)이 워크북을 새로 열고 스스로 보호를 다시 건다(_replace…:5839).
        보호는 시트에 걸린 플래그일 뿐이라 파일을 열고 읽는 데 지장이 없다.
      · 왜 하는가: VM 실측에서 스텝 적용 시간의 35~45%가 이 백업이었고, 정작 파일 저장은
        17ms 인데 앞뒤 뒤치다꺼리(보호 해제/재적용/화면 설정)가 307ms 였다(파일 저장의 18배).
        VM 은 Excel 명령 하나하나가 느려 이 격차가 더 크게 벌어진다.
    """
    with EXCEL_LOCK:
        _t_all0 = time.perf_counter()
        session = get_excel_session(excel_id)
        app, wb = session_workbook(session)
        read_only_mirror = bool(session.get("readOnlyMirror"))
        live_protected = bool(session.get("liveEditable"))
        _t_unprotect = 0.0
        if (read_only_mirror or live_protected) and not internal:
            # 저장(다운로드)본은 보호 없는 깨끗한 파일이 되도록 먼저 보호 해제.
            _t0 = time.perf_counter()
            try:
                _protect_workbook_for_read_only_mirror(wb, False)
            except Exception:
                pass
            _t_unprotect = (time.perf_counter() - _t0) * 1000
        # [멈춤 방지/측정] 적용-前 스냅샷 저장이 라이브(숨김) Excel 에서 분 단위로 블록되던 4분 구간 — 트레이스에
        # 안 남아 안 보였다. 외부링크 '업데이트?' 모달, 저장-전 재계산, '다른이름저장' 알림이 숨김 창에서 사용자
        # 입력을 기다리며 멈출 수 있어, 저장 동안 모달 억제 + 재계산 없이 저장하고 소요 ms 를 남긴다.
        _save_t0 = time.perf_counter()
        for _attr, _val in (("DisplayAlerts", False), ("AskToUpdateLinks", False), ("EnableEvents", False)):
            try:
                setattr(app, _attr, _val)
            except Exception:
                pass
        _save_calc_prev = None
        try:
            _save_calc_prev = app.Calculation
            app.Calculation = -4135  # xlCalculationManual (저장-전 재계산 회피)
        except Exception:
            _save_calc_prev = None
        if name:
            BACKEND_DIR.mkdir(parents=True, exist_ok=True)
            safe_name = Path(str(name)).name
            if not Path(safe_name).suffix:
                safe_name += Path(session["path"]).suffix or ".xlsx"
            safe_name = _promote_csv_multisheet_name(safe_name, wb)   # CSV 멀티시트 → .xlsx
            result_path = BACKEND_DIR / f"{uuid.uuid4().hex}_{safe_name}"
            if Path(safe_name).suffix.lower() == ".xlsx":
                wb.SaveAs(str(result_path), FileFormat=51)            # 명시 변환(멀티시트 보존)
            else:
                wb.SaveAs(str(result_path))
            session["path"] = str(result_path)
            session["name"] = safe_name
        else:
            BACKEND_DIR.mkdir(parents=True, exist_ok=True)
            _src_path = Path(session["path"])
            safe_name = Path(session.get("name") or _src_path.name).name
            if not Path(safe_name).suffix:
                safe_name += _src_path.suffix or ".xlsx"
            _promoted = _promote_csv_multisheet_name(safe_name, wb)
            result_path = BACKEND_DIR / f"{uuid.uuid4().hex}_{_promoted}"
            if _promoted != safe_name:
                # CSV 멀티시트 → xlsx 변환. SaveCopyAs 는 CSV 포맷을 유지하므로 SaveAs(경로 재지정) 필요 →
                # 멀티시트 파일은 사실상 xlsx 이므로 세션도 xlsx 로 코히어런트 갱신.
                wb.SaveAs(str(result_path), FileFormat=51)
                session["path"] = str(result_path)
                session["name"] = _promoted
                safe_name = _promoted
            else:
                wb.SaveCopyAs(str(result_path))
        try:
            if _save_calc_prev is not None:
                app.Calculation = _save_calc_prev
        except Exception:
            pass
        # [측정] 스냅샷 저장 소요 + 외부링크 수(분 단위 블록의 유력 원인) 기록 → 다음 실행에서 4분이 여기인지 확정.
        _save_link_n = 0
        try:
            _ls = wb.LinkSources(1)  # xlExcelLinks
            _save_link_n = len(_ls) if _ls is not None else 0
        except Exception:
            _save_link_n = -1
        _t_core = (time.perf_counter() - _save_t0) * 1000
        _t_restore0 = time.perf_counter()
        if internal:
            pass          # 되돌리기용 백업 — 보호/화면을 건드리지 않았으니 되돌릴 것도 없다
        elif read_only_mirror:
            try:
                _protect_workbook_for_read_only_mirror(wb, True)
                _configure_excel_grid_window(app, wb)
            except Exception:
                pass
        elif live_protected:
            # 라이브: 저장 후 화면 보호 복구(편집 차단 + 선택 허용). 단 리본/프레임은 평범하게 유지.
            try:
                _protect_workbook_for_read_only_mirror(wb, True)
                _configure_read_only_mirror_input_block(app)
                _disable_excel_context_menus(app)
            except Exception:
                pass
        _t_restore = (time.perf_counter() - _t_restore0) * 1000
        # [측정 구간 확대] 예전엔 SaveCopyAs 주변만 재서, 실제 소요의 상당 부분(보호 해제/재적용,
        # 화면 설정 복구)이 로그에 안 잡혔다 — VM 로그에서 서버 1.2초 vs 클라 3.8초로 벌어진 원인.
        # 이제 핸들러 전 구간을 쪼개서 남긴다.
        try:
            _vba_trace("excel.save.snapshot", excelId=excel_id, name=safe_name,
                       ms=round(_t_core, 1), linkCount=_save_link_n,
                       internal=bool(internal),
                       unprotectMs=round(_t_unprotect, 1),
                       restoreMs=round(_t_restore, 1),
                       totalMs=round((time.perf_counter() - _t_all0) * 1000, 1))
        except Exception:
            pass
        result_id = uuid.uuid4().hex
        RESULTS[result_id] = {
            "path": str(result_path),
            "name": Path(result_path).name,
            "created": time.time(),
        }
        return {
            "ok": True,
            "excelId": excel_id,
            "downloadId": result_id,
            "downloadUrl": f"/api/workbooks/download/{result_id}",
            "name": Path(result_path).name,
        }


def _clean_session_workbook_name(name):
    """스냅샷/결과 파일명에 붙는 우리 접두사(prestep_<32hex>_, <32hex>_)를 반복 제거해 원본 표시명을 복원한다.
    워크북을 '이름'으로 비교/조회하는 VBA(wbIter.Name = "원본", Workbooks("원본"))가 깨지지 않도록 라이브 세션
    워크북은 항상 이 깨끗한 이름으로 연다. 접두사가 복리로 쌓인 경우(prestep_..._prestep_..._원본)도 전부 벗긴다."""
    base = Path(str(name or "")).name
    n = base
    prev = None
    while n != prev:
        prev = n
        n = re.sub(r"^prestep_[0-9a-fA-F]{32}_", "", n)
        n = re.sub(r"^[0-9a-fA-F]{32}_", "", n)
    return n or base


def _resolve_live_identity_name(session_name, incoming_name, path_name):
    """결과/스냅샷을 '같은 라이브 세션'에 교체-로드(result-edit / 스냅샷 복원)할 때, 라이브 워크북이
    유지해야 할 '논리적 정체성 이름'을 고른다.

    실행기 파일출력 전체실행은 결과를 '결과_<원본stem>_<타임스탬프>[_<6hex난수>].xlsx' 로 저장한다(다운로드용
    이름). 이걸 result-edit 로 라이브에 되불러올 때 그 데코명을 그대로 wb.Name 으로 쓰면, @멘션(원본명)·
    VBA(Workbooks("원본"))·`If wb.Name = "원본"` 정확비교가 전부 '파일 못 찾음' 으로 깨진다.
    _clean_session_workbook_name 은 우리 '접두사'(prestep_/<32hex>_)만 벗기므로 '결과_..._난수' 데코는 못 벗긴다.

    replace 는 항상 '기존 세션을 이어서' 콘텐츠만 바꾸는 것이라, 라이브의 정체성은 '원래 세션 이름' 이다.
    따라서 세션에 이미 깨끗한 이름이 있으면 그것을 최우선 보존하고, 없을 때만 들어온 파일명을 접두사 정리해 쓴다.
    스냅샷 복원(prestep_<hex>_원본)도 세션 이름(원본)과 동일 결과라 회귀가 없다."""
    existing = _clean_session_workbook_name(session_name or "")
    if existing:
        return existing
    return _clean_session_workbook_name(incoming_name or path_name or "")


def _replace_excel_session_workbook_impl(excel_id, path, name=None, result_id=None, read_only_mirror=None):
    path = Path(path)
    if not path.exists():
        raise RuntimeError(f"result file not found: {path}")
    _rt0 = time.perf_counter(); _rt = {"mode": "replace"}  # F8 패널용 반영 단계 타이밍
    with EXCEL_LOCK:
        session = get_excel_session(excel_id)
        app, wb = session_workbook(session)
        read_only_mirror = session.get("readOnlyMirror") if read_only_mirror is None else bool(read_only_mirror)
        try:
            active = wb.Application.ActiveSheet
            active_sheet = active.Name if active is not None else None
        except Exception:
            active_sheet = None
        old_temp_path = session.get("openTempPath")
        old_replace_dir = session.get("replaceOpenDir")  # [이름 보존] 직전 replace 가 만든 원본명 사본 디렉토리
        _rt_close = time.perf_counter()
        live_editable = bool(session.get("liveEditable"))
        # [회색 엑셀] 교체 '전' 숨김 상태 — 교체가 숨김 세션을 표시로 승격하면 안 된다.
        # 실측(2026-07-29): 녹화 재현 직전 전세션 스냅샷 복원(replace 루프)이 숨겨져 있던
        # 비활성 파일 프레임까지 띄워 회색/겹침 창이 남았다(표시는 show-only 의 몫).
        was_hidden = bool(session.get("hidden"))
        if read_only_mirror or live_editable:
            try:
                app.ScreenUpdating = False
            except Exception:
                pass
        if read_only_mirror:
            _hide_excel_app_window(app)
        try:
            wb.Close(SaveChanges=False)
        except Exception:
            pass
        _rt["closeMs"] = round((time.perf_counter() - _rt_close) * 1000, 1)
        if old_temp_path:
            try:
                Path(old_temp_path).unlink(missing_ok=True)
            except Exception:
                pass
        if old_replace_dir:
            try:
                shutil.rmtree(old_replace_dir, ignore_errors=True)
            except Exception:
                pass
        # [되돌림 2026-08-04] 여기서 app.Visible=True 를 '조건부'로 바꿨더니(frame 모드 라이브는
        # 표시를 presenter 에 위임) 단계 OFF 직후 Excel 이 최대화된 채 튀어나오는 회귀가 났다
        # (사용자 실측). 새 프레임이 만들어지는 시점에 인스턴스가 보이는 상태여야 프레임 hwnd 가
        # 정상 생성되고 뒤이은 파킹/배치가 먹는다 — 원래 동작으로 되돌린다.
        # 실행기 헤드리스에서 이 창이 새는 문제는 (a) 아래 새 프레임 WindowState 정규화 +
        # (b) _move_hwnd_offscreen/_position_excel_window 의 최대화 해제 로 막는다.
        if not read_only_mirror:
            app.Visible = True
        else:
            _park_excel_app_offscreen(app)
        app.DisplayAlerts = False
        app.EnableEvents = False
        _park_excel_app_offscreen(app) if read_only_mirror else None
        _rt_open = time.perf_counter()
        # [워크북 이름 보존] 스냅샷/결과 파일은 prestep_/uuid 접두사가 붙어, 그대로 열면 wb.Name 이 임시명이 되어
        # 워크북을 '이름'으로 비교/조회하는 VBA(wbIter.Name = "원본", Workbooks("원본"))가 "파일 못 찾음"으로 깨진다
        # (별칭은 Workbooks("name") 호출만 치환할 뿐 .Name 문자열 비교는 못 고침). 전체실행 격리경로가 work/t/원본명
        # 으로 여는 것과 동형으로, 원본 표시명 사본을 만들어 그 사본을 열어 wb.Name 을 원본명으로 유지한다.
        # handle_excel_replace 는 항상 name=path.name(= 스냅샷/결과 파일명, prestep_/uuid 접두사 포함)을 넘긴다.
        # 따라서 name 이 있어도 그대로 쓰면 안 되고 반드시 접두사를 벗겨야 한다(이게 이전 버그 — if name 지름길).
        # 접두사 패턴(prestep_<32hex>_, <32hex>_)은 우리 내부 마커라 사용자 파일명과 충돌하지 않는다.
        # [정체성 보존] 세션의 기존 원본 이름을 최우선 보존한다. 실행기 결과명(결과_<stem>_<ts>[_난수])은
        # 접두사 패턴이 아니라 _clean_session_workbook_name 으로 못 벗겨지므로, 그대로 쓰면 라이브 wb.Name 이
        # 데코명이 되어 @멘션/VBA 가 '파일 못 찾음' 된다. → _resolve_live_identity_name 이 세션 원본명을 보존.
        clean_name = _resolve_live_identity_name(session.get("name"), name, path.name)
        replace_open_path = path
        replace_open_dir = None
        try:
            replace_open_dir = Path(tempfile.mkdtemp(prefix="b2b_replace_", dir=str(BACKEND_DIR)))
            _clean_open = replace_open_dir / clean_name
            shutil.copy2(str(path), str(_clean_open))
            replace_open_path = _clean_open
        except Exception:
            replace_open_path = path  # 사본 실패 시 원본 경로로 폴백(이름은 임시명이지만 최소 동작 보장)
            if replace_open_dir is not None:
                try:
                    shutil.rmtree(replace_open_dir, ignore_errors=True)
                except Exception:
                    pass
                replace_open_dir = None
        new_wb, new_temp_path = excel_workbooks_open(
            app,
            replace_open_path,
            read_only=bool(read_only_mirror),
            intended_name=clean_name,
        )
        _rt["openMs"] = round((time.perf_counter() - _rt_open) * 1000, 1)
        if session.get("liveEditable") and LIVE_FRAME_MODE:
            # 새 SDI 프레임이 기본 위치로 번쩍 뜨지 않게 즉시 파킹(끝의 presenter 가 제자리 표시).
            # [최대화 정규화] open 경로(5039)는 새 프레임에 xlNormal 을 걸어 두는데 replace 에는
            # 그게 빠져 있었다 — 새 프레임이 (직전 창 상태를 물려받아) 최대화로 뜨면 SetWindowPos 가
            # 위치를 무시해 파킹이 실패하고 화면을 덮은 채 남는다(OFF 직후 '엑셀 최대화' 실측).
            try:
                new_wb.Windows(1).WindowState = -4143  # xlNormal (이 프레임만)
            except Exception:
                pass
            _new_frame_hwnd = _workbook_window_hwnd(new_wb)
            if _new_frame_hwnd:
                _move_hwnd_offscreen(_new_frame_hwnd)
        if read_only_mirror:
            _park_excel_app_offscreen(app)
            try:
                new_wb.Activate()
            except Exception:
                pass
            _protect_workbook_for_read_only_mirror(new_wb, True)
            _configure_excel_grid_window(app, new_wb)
            _ensure_excel_workbook_view(
                app,
                new_wb,
                activate=False if (session.get("nativeParentHwnd") or session.get("nativeOverlay")) else True,
                maximize_workbook=False if (session.get("nativeOverlay") or session.get("nativeParentHwnd")) else True,
            )
        session["workbook"] = new_wb
        # 워크북은 원본 표시명 사본(replace_open_path)으로 열었다 → 세션 path/name 도 그 깨끗한 사본 기준으로.
        # session_workbook 가 재오픈(GetObject)해도 깨끗한 이름을 유지해, 이름 비교 VBA 가 안 깨진다.
        session["path"] = str(replace_open_path)
        session["openPath"] = str(new_wb.FullName)
        session["openTempPath"] = str(new_temp_path) if new_temp_path else ""
        session["replaceOpenDir"] = str(replace_open_dir) if replace_open_dir else ""
        session["name"] = clean_name
        session["resultId"] = result_id
        session["readOnlyMirror"] = bool(read_only_mirror)
        session["hidden"] = was_hidden  # [회색 엑셀] 숨김 세션 교체는 숨김 유지(표시 승격 금지)
        session["snapshots"] = {}
        session["appliedStepSigs"] = None  # 워크북이 외부 결과로 교체됨 → 적용 단계 추적 무효화
        sheets = _excel_collection_names(new_wb.Worksheets)
        # [회색 엑셀] Worksheet.Activate 는 그 워크북을 ActiveWorkbook 으로 끌어올린다 —
        # 숨김 세션 교체에서는 건너뛴다(활성/포커스가 비활성 파일로 새는 것 방지).
        if active_sheet and active_sheet in sheets and not (live_editable and was_hidden and LIVE_FRAME_MODE):
            try:
                new_wb.Worksheets(active_sheet).Activate()
            except Exception:
                pass
        if read_only_mirror:
            _ensure_excel_workbook_view(
                app,
                new_wb,
                activate=False if (session.get("nativeParentHwnd") or session.get("nativeOverlay")) else True,
                maximize_workbook=False if (session.get("nativeOverlay") or session.get("nativeParentHwnd")) else True,
            )
            if session.get("nativeParentHwnd") or session.get("nativeOverlay"):
                try:
                    left, top, width, height = [int(v) for v in str(session.get("lastNativePositionKey") or "").split(":")[-4:]]
                    _position_excel_window(
                        app,
                        left,
                        top,
                        width,
                        height,
                        native_parent_hwnd=session.get("nativeParentHwnd") if not session.get("nativeOverlay") else None,
                        native_host_hwnd=session.get("nativeHostHwnd"),
                        native_overlay=bool(session.get("nativeOverlay")),
                        show=True,
                    )
                except Exception:
                    pass
            try:
                app.ScreenUpdating = True
            except Exception:
                pass
        elif live_editable and was_hidden and LIVE_FRAME_MODE:
            # [회색 엑셀 수정] 숨겨져 있던 라이브 세션의 교체는 '내용만' 바꾸고 표시하지 않는다.
            # 예전엔 아래 일반 경로가 무조건 _present_live_session_frame 으로 새 프레임을 띄워,
            # 녹화 재현 직전 전세션 복원에서 비활성 파일 프레임이 화면에 뜨고 남았다(실측).
            # 새 프레임은 오픈 직후 이미 offscreen 파킹돼 있으므로(위 _move_hwnd_offscreen)
            # 그대로 두고, 보호/그리드 설정만 해 둔다 — 이후 탭 전환/show-only 가 표시를 담당.
            # (Activate/Worksheet.Activate 도 하지 않는다 — ActiveWorkbook 이 비활성 파일로 새는
            #  것 방지. 기능 손실 0: 표시 시점에 show-only 가 위치·보기 구성을 다시 잡는다.)
            try:
                _protect_workbook_for_read_only_mirror(new_wb, True)
                _configure_excel_grid_window(app, new_wb)
            except Exception:
                pass
            session["hidden"] = True
            try:
                app.ScreenUpdating = True
            except Exception:
                pass
            _rt["presentMs"] = 0.0
        elif live_editable:
            _rt_show = time.perf_counter()
            # Python(openpyxl) 결과 교체의 라이브 반영 경로. 공유 인스턴스(frame 모드)에서는
            # app 전역(owner/hide/Visible) 조작 금지 — 이 세션의 새 프레임만 제어한다.
            try:
                new_wb.Activate()
            except Exception:
                pass
            try:
                _protect_workbook_for_read_only_mirror(new_wb, True)
                _configure_excel_grid_window(app, new_wb)
            except Exception:
                pass
            presented = None
            if LIVE_FRAME_MODE:
                # 워크북이 교체됐으므로 이전 프레임 핸들을 버리고 새 프레임 기준으로 표시.
                session.pop("frameHwnd", None)
                rect = session.get("liveRect") or {}
                presented = _present_live_session_frame(
                    session, app, new_wb,
                    rect.get("left"), rect.get("top"), rect.get("width"), rect.get("height"),
                    skip_position=not (rect.get("width") and rect.get("height")),
                )
            if presented is None:
                # legacy(비frame) 폴백: 페인트 준비 후 show — 결과 교체 시 회색 프레임 플래시 방지.
                _set_excel_window_owner(app, session.get("nativeHostHwnd"))
                try:
                    _ensure_excel_workbook_view(app, new_wb, make_visible=True, activate=False, maximize_workbook=False, defer_show=True)
                except Exception:
                    pass
                try:
                    _hide_excel_hwnd(app.Hwnd)
                except Exception:
                    pass
                try:
                    app.ScreenUpdating = True
                except Exception:
                    pass
                try:
                    new_wb.Windows(1).Visible = True
                except Exception:
                    pass
                live_rect = session.get("liveRect") or {}
                if live_rect.get("width") and live_rect.get("height"):
                    try:
                        _position_excel_window(
                            app,
                            live_rect.get("left"),
                            live_rect.get("top"),
                            live_rect.get("width"),
                            live_rect.get("height"),
                            native_host_hwnd=session.get("nativeHostHwnd"),
                            show=True,
                        )
                    except Exception:
                        pass
                try:
                    _ensure_excel_workbook_view(app, new_wb, make_visible=True, activate=False, maximize_workbook=False)
                except Exception:
                    pass
                _set_excel_window_owner(app, session.get("nativeHostHwnd"))
            try:
                app.ScreenUpdating = True
            except Exception:
                pass
            _rt["presentMs"] = round((time.perf_counter() - _rt_show) * 1000, 1)
        else:
            presented = None
            if session.get("liveEditable") and LIVE_FRAME_MODE:
                # 워크북이 교체됐으므로 이전 프레임 핸들을 버리고 새 프레임 기준으로 표시.
                session.pop("frameHwnd", None)
                rect = session.get("liveRect") or {}
                presented = _present_live_session_frame(
                    session, app, new_wb,
                    rect.get("left"), rect.get("top"), rect.get("width"), rect.get("height"),
                    skip_position=not (rect.get("width") and rect.get("height")),
                )
            if presented is None:
                try:
                    _safe_activate_excel_app(app)
                except Exception:
                    pass
        return {
            "ok": True,
            "excelId": excel_id,
            # [정체성 보존] 데코된 결과 파일명(name/path.name) 대신, 라이브가 실제로 유지하는 깨끗한 원본명을
            # 돌려준다 → 파일탭/표시명이 결과_..._난수 로 새지 않고, 클라가 @멘션용으로 쓰는 이름과 일치한다.
            "name": clean_name,
            "path": str(path),
            "sheetNames": sheets,
            "readOnlyMirror": bool(read_only_mirror),
            "replaced": True,
            "debugTimings": dict(_rt, totalMs=round((time.perf_counter() - _rt0) * 1000, 1)),
        }


def _position_excel_session_impl(
    excel_id,
    left,
    top,
    width,
    height,
    client_left=None,
    client_top=None,
    client_width=None,
    client_height=None,
    viewport_width=None,
    viewport_height=None,
    browser_title=None,
    native_parent_hwnd=None,
    native_host_hwnd=None,
    native_overlay=False,
    keep_zorder=False,
):
    with EXCEL_LOCK:
        session = get_excel_session(excel_id)
        app, wb = session_workbook(session)
        next_native_parent_hwnd = native_parent_hwnd or session.get("nativeParentHwnd")
        next_native_host_hwnd = native_host_hwnd or session.get("nativeHostHwnd")
        next_native_overlay = bool(native_overlay or session.get("nativeOverlay"))
        native_position_key = None
        if session.get("readOnlyMirror") and (next_native_parent_hwnd or next_native_overlay):
            native_position_key = f"{'overlay' if next_native_overlay else next_native_parent_hwnd}:{int(float(left or 0))}:{int(float(top or 0))}:{int(float(width or 0))}:{int(float(height or 0))}"
            if session.get("lastNativePositionKey") == native_position_key and not session.get("hidden"):
                return {
                    "ok": True,
                    "excelId": excel_id,
                    "left": int(float(left or 0)),
                    "top": int(float(top or 0)),
                    "width": int(float(width or 0)),
                    "height": int(float(height or 0)),
                    "skipped": True,
                }
        if session.get("readOnlyMirror"):
            if native_parent_hwnd:
                session["nativeParentHwnd"] = native_parent_hwnd
            if native_host_hwnd:
                session["nativeHostHwnd"] = native_host_hwnd
            if native_overlay:
                session["nativeOverlay"] = True
                session["nativeParentHwnd"] = None
            browser_hwnd = session.get("browserHwnd")
            try:
                browser_valid = bool(browser_hwnd and win32gui is not None and win32gui.IsWindow(int(browser_hwnd)))
            except Exception:
                browser_valid = False
            if not browser_valid:
                session["browserHwnd"] = _capture_browser_hwnd(browser_title)
        # frame 모드: 공유 인스턴스의 app.Hwnd(활성 프레임)가 아니라 이 세션의 프레임을 직접 이동.
        live_frame_hwnd = _session_frame_hwnd(session, wb) if (session.get("liveEditable") and LIVE_FRAME_MODE) else None
        _position_excel_window(
            app,
            left,
            top,
            width,
            height,
            browser_hwnd=None if (session.get("nativeParentHwnd") or session.get("nativeOverlay")) else (session.get("browserHwnd") if session.get("readOnlyMirror") else None),
            native_parent_hwnd=session.get("nativeParentHwnd") if session.get("readOnlyMirror") and not session.get("nativeOverlay") else None,
            native_host_hwnd=session.get("nativeHostHwnd") if session.get("readOnlyMirror") else None,
            native_overlay=bool(session.get("nativeOverlay")) if session.get("readOnlyMirror") else False,
            client_left=client_left,
            client_top=client_top,
            client_width=client_width,
            client_height=client_height,
            viewport_width=viewport_width,
            viewport_height=viewport_height,
            keep_zorder=keep_zorder,
            hwnd=live_frame_hwnd,
            no_activate=bool(live_frame_hwnd),
        )
        if native_position_key:
            session["lastNativePositionKey"] = native_position_key
        if session.get("liveEditable") and width and height:
            # 리셋(시트 교체) 후 창 복원에 쓰도록 최신 rect 보관.
            session["liveRect"] = {
                "left": int(float(left or 0)), "top": int(float(top or 0)),
                "width": int(float(width or 0)), "height": int(float(height or 0)),
            }
        session["hidden"] = False
        if session.get("readOnlyMirror"):
            _ensure_excel_workbook_view(
                app,
                wb,
                make_visible=True,
                activate=False if (session.get("nativeParentHwnd") or session.get("nativeOverlay")) else True,
                maximize_workbook=False if (session.get("nativeOverlay") or session.get("nativeParentHwnd")) else True,
            )
        elif session.get("liveEditable"):
            # owner 재확인(리사이즈 후 풀릴 수 있음) + 그리드 채움.
            if live_frame_hwnd:
                _set_window_owner_hwnd(live_frame_hwnd, session.get("nativeHostHwnd"))
                try:
                    _ensure_excel_workbook_view(app, wb, make_visible=False, activate=False, maximize_workbook=False, app_level=False)
                except Exception:
                    pass
            else:
                _set_excel_window_owner(app, session.get("nativeHostHwnd"))
                try:
                    _ensure_excel_workbook_view(app, wb, make_visible=True, activate=False, maximize_workbook=False)
                except Exception:
                    pass
        return {
            "ok": True,
            "excelId": excel_id,
            "left": int(float(left or 0)),
            "top": int(float(top or 0)),
            "width": int(float(width or 0)),
            "height": int(float(height or 0)),
        }


def _raise_excel_session_impl(excel_id):
    with EXCEL_LOCK:
        session = get_excel_session(excel_id)
        app, wb = session_workbook(session)
        if session.get("liveEditable") and LIVE_FRAME_MODE:
            hwnd = _session_frame_hwnd(session, wb)
            if hwnd:
                _set_window_owner_hwnd(hwnd, session.get("nativeHostHwnd"))
                _raise_excel_hwnd(hwnd)
                return {"ok": True, "excelId": excel_id}
        if session.get("readOnlyMirror") or session.get("liveEditable"):
            _raise_excel_window(app)
            if session.get("liveEditable"):
                _set_excel_window_owner(app, session.get("nativeHostHwnd"))
            session["hidden"] = False
        return {"ok": True, "excelId": excel_id}


def _workbook_fullname(wb):
    try:
        return str(Path(wb.FullName).resolve()).lower()
    except Exception:
        try:
            return str(wb.FullName).lower()
        except Exception:
            return ""


def _excel_app_hwnd(app):
    try:
        return int(app.Hwnd)
    except Exception:
        return None


def _same_excel_app(app_a, app_b):
    hwnd_a = _excel_app_hwnd(app_a)
    hwnd_b = _excel_app_hwnd(app_b)
    return bool(hwnd_a and hwnd_b and hwnd_a == hwnd_b)


def _session_for_workbook(wb, app=None):
    target = _workbook_fullname(wb)
    if not target:
        return None, None, None
    for sid, session in list(EXCEL_SESSIONS.items()):
        try:
            session_app = session.get("app")
            if app is not None and session_app is not None and not _same_excel_app(app, session_app):
                continue
            session_wb = session.get("workbook")
            if session_wb is not None and _workbook_fullname(session_wb) == target:
                return sid, session, session_wb
        except Exception:
            continue
    return None, None, None


def _active_session_for_app(app, fallback_session=None, fallback_wb=None):
    try:
        active_wb = app.ActiveWorkbook
    except Exception:
        active_wb = None
    sid, session, wb = _session_for_workbook(active_wb, app=app) if active_wb is not None else (None, None, None)
    if sid and session and wb is not None:
        return sid, session, wb
    if fallback_session is not None:
        return fallback_session.get("id"), fallback_session, fallback_wb
    return None, None, None


def _hide_workbook_window(wb):
    try:
        if wb is None or wb.Windows.Count < 1:
            return False
        win = wb.Windows(1)
        try:
            win.Visible = False
            return True
        except Exception:
            pass
        try:
            hwnd = getattr(win, "Hwnd", None)
            if hwnd:
                _hide_excel_hwnd(hwnd)
                return True
        except Exception:
            pass
    except Exception:
        pass
    return False


def _hide_peer_workbook_windows(app, active_excel_id):
    hidden_ids = []
    for sid, other_session in list(EXCEL_SESSIONS.items()):
        if sid == active_excel_id:
            continue
        try:
            other_app = other_session.get("app")
            if other_app is None or not _same_excel_app(app, other_app):
                continue
            other_wb = other_session.get("workbook")
            if other_wb is None:
                continue
            if _hide_workbook_window(other_wb):
                other_session["hidden"] = True
                hidden_ids.append(sid)
        except Exception:
            continue
    return hidden_ids


def _show_workbook_window(app, wb, activate=True):
    try:
        app.Visible = True
    except Exception:
        pass
    try:
        if wb is not None and wb.Windows.Count > 0:
            wb.Windows(1).Visible = True
    except Exception:
        pass
    try:
        if activate and wb is not None:
            wb.Activate()
            wb.Windows(1).Activate()
    except Exception:
        pass
    try:
        _ensure_excel_workbook_view(app, wb, make_visible=True, activate=activate, maximize_workbook=False)
    except Exception:
        pass


def _workbook_window_hwnd(wb):
    """SDI 프레임(이 워크북의 최상위 창) 핸들. 공유 인스턴스에서 app.Hwnd 는
    '그 순간 활성 프레임 1개'라 다중 워크북 제어에 쓰면 엉뚱한 창을 만진다."""
    try:
        if wb is None or wb.Windows.Count < 1:
            return None
        hwnd = int(wb.Windows(1).Hwnd)
        return hwnd or None
    except Exception:
        return None


def _session_frame_hwnd(session, wb=None):
    """세션 워크북의 프레임 핸들(캐시). recover/replace 로 워크북이 바뀌면 자동 재조회."""
    hwnd = session.get("frameHwnd")
    try:
        if hwnd and win32gui is not None and win32gui.IsWindow(int(hwnd)):
            return int(hwnd)
    except Exception:
        pass
    if wb is None:
        wb = session.get("workbook")
    hwnd = _workbook_window_hwnd(wb)
    if hwnd:
        session["frameHwnd"] = hwnd
    return hwnd


def _foreground_session_by_frame():
    """포그라운드 창이 우리 라이브 세션 프레임이면 그 세션을 반환.
    active-sync(엑셀→UI 탭 동기화)는 '사용자가 실제로 클릭해 포그라운드인 미러'만 따라가야
    프로그램적 전환(show-only)과 경합해 탭이 되돌아가는 바운스가 생기지 않는다."""
    if win32gui is None:
        return None, None, None
    try:
        fg = int(win32gui.GetForegroundWindow() or 0)
    except Exception:
        return None, None, None
    if not fg:
        return None, None, None
    for sid, session in list(EXCEL_SESSIONS.items()):
        if not session.get("liveEditable"):
            continue
        try:
            if int(session.get("frameHwnd") or 0) == fg:
                wb = session.get("workbook")
                if wb is not None:
                    return sid, session, wb
        except Exception:
            continue
    return None, None, None


def _hide_peer_session_frames(active_excel_id, host_hwnd=None):
    """frame 모드: 활성 세션 외 라이브 프레임을 전부 화면 밖으로 파킹.
    파킹 대상이 포그라운드면 먼저 호스트로 포커스를 넘긴다(무관 창 점프 방지)."""
    peers = []
    for sid, other in list(EXCEL_SESSIONS.items()):
        if sid == active_excel_id or not other.get("liveEditable"):
            continue
        try:
            hwnd = _session_frame_hwnd(other)
        except Exception:
            hwnd = None
        if hwnd:
            peers.append((sid, other, hwnd))
    if not peers:
        return []
    _handoff_foreground_to_host(host_hwnd, [hwnd for _sid, _other, hwnd in peers])
    hidden_ids = []
    for sid, other, hwnd in peers:
        _move_hwnd_offscreen(hwnd)
        other["hidden"] = True
        hidden_ids.append(sid)
    return hidden_ids


# 네이티브(VBA) 녹화 세션 상태 — {"baseline": [...], "active": True} (HTTP/워커 공용, 단일 세션)
NATIVE_RECORDING = {}


# 녹화/재현은 무조건 VBA(네이티브 매크로 레코더). python 이벤트 캡처(record_service)는
# 이 상수를 True 로 바꿔야만 허용 — 기본은 VBA 전용. 폐쇄망/디버그용 뒷문만 남긴다.
ALLOW_PYTHON_RECORD_ENGINE = os.environ.get("B2B_ALLOW_PYTHON_RECORD", "") == "1"


def excel_record_start(engine="vba"):
    if engine != "vba" and not ALLOW_PYTHON_RECORD_ENGINE:
        engine = "vba"  # 무조건 VBA — python 엔진 요청은 무시
    # 녹화는 사용자가 라이브 Excel 창에서 직접 조작한 것을 캡처한다 —
    # 레코더를 붙이기 전에 편집 잠금(시트 보호·키 차단·우클릭·리본)을 먼저 푼다.
    # [셀 편집 중 시작] 사용자가 셀 입력 중이면 아래 COM 호출들이 전부 거부된다 — 먼저 확정.
    try:
        excel_call(lambda: _commit_pending_excel_cell_edit(_get_live_excel_app()), timeout=15)
    except Exception:
        pass
    excel_call(_set_live_sessions_edit_unlock, True, timeout=60)
    try:
        if engine == "vba":
            # Excel 네이티브 매크로 레코더(MS 매크로 기록기).
            # 녹화 중 부하 ~0, 정지 시 전체 행동이 VBA 스킬 1개로 청킹된다.
            from native_macro_recorder import start_native_recording_impl

            def _start_native():
                return start_native_recording_impl(_get_live_excel_app())

            # [녹화 보호] active 를 레코더 시작(ExecuteMso — 단일 COM 워커를 모달 확인까지 블록) '전에'
            # 올린다. 이 블록 동안에도 배경 폴러(/changes·/selection·/hover-info)가 워커 뒤에 줄서서
            # 타임아웃 → 클라 워치독 → /api/excel/force-restart 로 들어올 수 있는데, 서버 force-restart
            # 가드(NATIVE_RECORDING.active)가 '이 시작 창'까지 자립적으로 커버해 공유 Excel 을 지킨다
            # (active 를 시작 반환 후에 올리면 이 창이 서버측에 뚫려 클라 플래그에만 의존하게 된다).
            # 시작이 실패하면 아래 except 의 NATIVE_RECORDING.clear() 가 되돌린다.
            NATIVE_RECORDING.clear()
            NATIVE_RECORDING["active"] = True
            NATIVE_RECORDING["baseline"] = None
            baseline = excel_call(_start_native, timeout=30)
            try:
                from native_macro_recorder import RECORD_DIAG as _rd
                _vba_trace("record.native.start", editUnlocked=bool(RECORDING_EDIT_UNLOCKED), **dict(_rd))
            except Exception:
                pass
            NATIVE_RECORDING["baseline"] = baseline
        else:
            from record_service import RECORD_SERVICE, marshal_app_stream

            def _marshal_live_app():
                return marshal_app_stream(_get_live_excel_app())

            stream = excel_call(_marshal_live_app, timeout=30)
            RECORD_SERVICE.start(app_stream=stream)
    except Exception:
        NATIVE_RECORDING.clear()
        try:
            excel_call(_set_live_sessions_edit_unlock, False, timeout=60)
        except Exception:
            pass
        raise
    return {"ok": True, "recording": True, "engine": engine, "editUnlocked": True}


def _recorded_vba_hazards(code):
    """녹화 VBA 에서 '절대참조 재현이 불안정한' 패턴을 감지해 사용자 경고를 만든다.

    MS 매크로 레코더는 절대 셀/시트/이름을 하드코딩한다. 새로 만든 시트·피벗은
    실행 환경마다 이름이 달라(예: Sheet1↔Sheet2) 이후 고정 이름 참조가 어긋나
    재현이 깨진다(실사례: PivotFields 실패). 검토 카드에 ⚠ 로 노출해 사용자가
    추가 전에 판단하게 한다(막지는 않음 — 되는 케이스도 있으므로)."""
    text = str(code or "")
    hz = []
    if re.search(r"PivotCaches|PivotTable", text, re.I):
        hz.append(
            "피벗테이블 — 새로 만든 시트/피벗 이름과 원본 범위를 고정으로 참조합니다. "
            "재현 시 시트 이름·필드명이 어긋나면 실패할 수 있어(PivotFields 오류) "
            "피벗은 채팅으로 만드는 편이 안정적입니다.")
    if re.search(r"Sheets\.Add|Worksheets\.Add|Sheets\.Add2", text, re.I):
        hz.append(
            "새 시트 추가 — 시트 이름이 실행 환경마다 달라질 수 있어(Sheet1↔Sheet2 등) "
            "이후 고정 시트 이름 참조가 어긋날 수 있습니다.")
    if re.search(r'Windows\(\s*["\'][^"\']*\.xls', text, re.I) or re.search(r'Workbooks\(\s*["\'][^"\']*\.xls', text, re.I):
        hz.append(
            "다른 워크북 참조 — 재현할 때 그 파일이 함께 열려 있어야 합니다(대상 파일과 함께 여세요).")
    # [보안 게이트 예고] 재생 시 서버가 차단할 동작이 녹화에 들었으면 검토 단계에서 미리 알린다
    # (지금 추가하면 실행에서 반드시 실패 — 사용자가 카드에서 빼거나 다시 녹화하도록).
    _sec = _vba_security_scan(text)
    if _sec:
        hz.append("⛔ 재생 차단 대상 — %s. 이 동작이 든 조각은 실행이 거부되니 빼거나 다시 녹화하세요." % _sec)
    if re.search(r"\bWorkbooks\s*\.\s*Add\b", text, re.I):
        hz.append(
            "새 통합 문서 만들기(Workbooks.Add) — 재현 환경에서 만든 새 파일은 결과로 수집되지 않고 사라집니다. "
            "출력할 파일은 미리 만들어 함께 열어 두세요.")
    return hz


def excel_record_stop():
    try:
        if NATIVE_RECORDING.get("active"):
            from native_macro_recorder import stop_native_recording_impl

            def _stop_native():
                app = _get_live_excel_app()
                # [셀 편집 중 정지] 편집을 확정(Enter)시켜 COM 을 되살린 뒤 수확한다 —
                # 안 하면 in-cell edit 중 정지가 COM 거부로 실패해 녹화를 다시 찍어야 했다.
                _commit_pending_excel_cell_edit(app)
                return stop_native_recording_impl(app, NATIVE_RECORDING.get("baseline"))

            # 정지 시 expected 다이제스트 수확(touched 시트 Value2 해시)이 추가돼 60s 는 빠듯할
            # 수 있다 — 프론트 stop 타임아웃(200s) 안에서 120s 로 여유를 둔다.
            rec = excel_call(_stop_native, timeout=120)
            try:
                from native_macro_recorder import RECORD_DIAG as _rd
                _vba_trace("record.native.stop", hasCode=bool(rec.get("code")), **dict(_rd))
            except Exception:
                pass
            NATIVE_RECORDING.clear()
            steps = []
            if rec.get("code"):
                steps.append({
                    "id": f"rec_vba_{int(time.time() * 1000)}",
                    "language": "vba",
                    "code": rec["code"],
                    "title": f"녹화된 작업 ({rec.get('summary') or '기록'})",
                    "description": f"네이티브 매크로 녹화 — {rec.get('summary') or '기록된 동작'}",
                    "enabled": True,
                    "prompt": f"[녹화됨/VBA] {rec.get('summary') or ''}",
                    "recorded": True,
                    "hazards": _recorded_vba_hazards(rec["code"]),
                })
            result = {"steps": steps, "raw_actions": rec.get("rawLines", 0),
                      "distilled": len(steps), "groups": len(steps), "engine": "vba"}
            # [재현 검증] 네이티브 stop 도 expected(정지 시점 시트 다이제스트)를 실어 준다 —
            # 예전엔 python 엔진 stop 만 실어, 기본(VBA) 경로에서 프론트 검증 블록과
            # /api/excel/record/verify 인프라가 통째로 死코드였다.
            result["expected"] = rec.get("expected") or []
            # [두 워크북 대상 바인딩] 녹화가 실제로 일어난 워크북(새 매크로 모듈이 생긴 곳)의
            # 세션 excelId 를 실어 준다. 클라가 이걸로 targetFileId 를 정확히 바인딩해야
            # UI 탭(state.currentFileId)과 실제 녹화 워크북이 다를 때 재현이 엉뚱한
            # 워크북에 적용되던 문제(예: input 에서 녹화했는데 output:0 로 박힘)를 막는다.
            rec_full = str(rec.get("recordedWorkbookFullName") or "")
            rec_name = str(rec.get("recordedWorkbook") or "")
            result["recordedWorkbook"] = rec_name
            # [3A] 녹화된 활성 시트명 — 실행기 '파일확인'이 필요 시트를 잡는 데 쓴다.
            result["recordedSheet"] = str(rec.get("recordedSheet") or "")
            rec_excel_id = ""
            if rec_full or rec_name:
                # COM(_wb.FullName) 을 HTTP 스레드에서 만지지 않도록, 세션에 저장된
                # 문자열 경로/이름만으로 매칭한다(크로스스레드 COM 접근 회피).
                def _norm(p):
                    try:
                        return str(Path(p).resolve()).lower()
                    except Exception:
                        return str(p or "").lower()
                rec_full_n = _norm(rec_full) if rec_full else ""
                rec_name_l = rec_name.lower()
                with EXCEL_LOCK:
                    _sessions = list(EXCEL_SESSIONS.items())
                _name_hit = ""
                for _sid, _sess in _sessions:
                    try:
                        cand_paths = [
                            _norm(_sess.get("openPath") or ""),
                            _norm(_sess.get("path") or ""),
                            _norm(_sess.get("openTempPath") or ""),
                        ]
                        if rec_full_n and rec_full_n in cand_paths:
                            rec_excel_id = _sid
                            break
                        cand_names = {
                            str(_sess.get("name") or "").lower(),
                            str(Path(_sess.get("openPath") or "").name).lower(),
                            str(Path(_sess.get("path") or "").name).lower(),
                        }
                        if rec_name_l and rec_name_l in cand_names and not _name_hit:
                            _name_hit = _sid  # 경로 매칭 우선, 없으면 이름 폴백
                    except Exception:
                        continue
                if not rec_excel_id and _name_hit:
                    rec_excel_id = _name_hit
            if rec_excel_id:
                result["recordedExcelId"] = rec_excel_id
            # [3A] 스텝에 파일/시트 메타 stamp — 프론트/실행기가 소비해 필요 워크북/시트를
            # 정확히 잡는다. recordedWorkbook 은 raw wb.Name(temp 사본일 수 있음) 대신 위
            # 세션 매칭으로 얻은 세션 표시명으로 정규화한다. 매칭 세션이 없으면(정규화 실패)
            # 두 필드 모두 stamp 하지 않는다(빈 요구가 실행기에 들어가는 것을 방지).
            if rec_excel_id and steps:
                _rec_sess_name = ""
                with EXCEL_LOCK:
                    _rs = EXCEL_SESSIONS.get(rec_excel_id)
                    if _rs:
                        _rec_sess_name = str(_rs.get("name") or "")
                if _rec_sess_name:
                    steps[0]["recordedWorkbook"] = _rec_sess_name
                    _rec_sheet = str(rec.get("recordedSheet") or "")
                    if _rec_sheet:
                        steps[0]["recordedSheet"] = _rec_sheet
        else:
            from record_service import RECORD_SERVICE
            # 정지 시 스냅샷 diff(서식/객체)가 시트 규모에 따라 오래 걸릴 수 있다.
            result = RECORD_SERVICE.stop(timeout=180.0)
    finally:
        # 성공/실패와 무관하게 편집 잠금 원복(녹화 중 만든 새 시트 포함 전 시트 재보호).
        # [진단] '정지 후 편집 가능 상태로 남음' 실측 — 재잠금이 실제로 됐는지/RECORDING_EDIT_UNLOCKED
        # 가 풀렸는지 트레이스한다(실패해도 무해, 다음 재현서 원인 특정).
        _relock_ok = False
        try:
            excel_call(_set_live_sessions_edit_unlock, False, timeout=60)
            _relock_ok = True
        except Exception as _rlerr:
            try: _vba_trace("record.stop.relock_fail", error=str(_rlerr))
            except Exception: pass
        try:
            _vba_trace("record.stop.relock", ok=_relock_ok, editUnlockedAfter=bool(RECORDING_EDIT_UNLOCKED))
        except Exception:
            pass
    return {"ok": True, "recording": False, **result}


def excel_record_status():
    if NATIVE_RECORDING.get("active"):
        return {"ok": True, "recording": True, "engine": "vba"}
    from record_service import RECORD_SERVICE
    return {"ok": True, **RECORD_SERVICE.status()}


def _verify_recorded_expected_live(expected):
    """Excel 워커 — 녹화 정지 시점 기대 상태(expected)와 현재 라이브 시트를 대조.

    양쪽 모두 record_service.sheet_expected_state(같은 정규화·같은 상한)로 계산해
    다이제스트가 직접 비교 가능하다. 재현이 정답과 어긋난 시트를 정확히 지목한다."""
    from record_service import sheet_expected_state
    wbs = {}
    for session in list(EXCEL_SESSIONS.values()):
        wb = session.get("workbook")
        if wb is None:
            continue
        try:
            wbs.setdefault(str(wb.Name), wb)
        except Exception:
            continue
    results = []
    for item in (expected or []):
        book = str(item.get("book") or "")
        sheet = str(item.get("sheet") or "")
        res = {"book": book, "sheet": sheet, "match": None, "reason": ""}
        wb = wbs.get(book)
        if wb is None:
            res["reason"] = "라이브 세션에 이 워크북이 없음(대조 생략)"
            results.append(res)
            continue
        try:
            ws = wb.Worksheets(sheet)
        except Exception:
            res["match"] = False
            res["reason"] = "시트가 없음"
            results.append(res)
            continue
        try:
            cur = sheet_expected_state(ws)
        except Exception as err:
            res["reason"] = f"판독 실패(대조 생략): {err}"
            results.append(res)
            continue
        dims_ok = (cur["rows"] == item.get("rows") and cur["cols"] == item.get("cols"))
        hash_ok = (str(item.get("hash") or "") == str(cur.get("hash") or "")) if item.get("hash") else None
        # 병합 지문 대조 — 구버전 expected(merges 없음)는 건너뛴다(하위호환).
        merge_ok = None
        exp_merges = item.get("merges")
        if isinstance(exp_merges, list) and isinstance(cur.get("merges"), list):
            merge_ok = (sorted(map(str, exp_merges)) == sorted(map(str, cur["merges"])))
        res["match"] = bool(dims_ok and hash_ok is not False and merge_ok is not False)
        if not dims_ok:
            res["reason"] = (f"사용범위 다름: 녹화 {item.get('rows')}x{item.get('cols')}"
                             f" → 재현 {cur['rows']}x{cur['cols']}")
        elif merge_ok is False:
            exp_set = set(map(str, exp_merges))
            got_set = set(map(str, cur["merges"]))
            sample = sorted(got_set - exp_set) or sorted(exp_set - got_set)
            res["reason"] = (f"병합 상태 다름: 녹화 {len(exp_set)}개 → 재현 {len(got_set)}개"
                             + (f" (예: {sample[0]})" if sample else ""))
        elif hash_ok is False:
            hr = int(item.get("hashRows") or 0)
            res["reason"] = "값 다름(상단 %s행 대조)" % hr if hr and hr < int(item.get("rows") or 0) else "값 다름"
        results.append(res)
    checked = [r for r in results if r["match"] is not None]
    return {"ok": True, "results": results,
            "allMatch": bool(checked) and all(r["match"] for r in checked),
            "checked": len(checked)}


def excel_record_verify(payload):
    expected = (payload or {}).get("expected") or []
    if not expected:
        return {"ok": True, "results": [], "allMatch": True, "checked": 0}
    return excel_call(_verify_recorded_expected_live, expected, timeout=120)


_VLLM_PROBE_CACHE = {"ts": 0.0, "base": None, "result": None}
_VLLM_PROBE_TTL = 30.0  # 헬스 프로브 캐시(초) — 매 헬스 호출마다 vLLM 을 두드리지 않게


def _vllm_health_probe(base="", timeout=2.0):
    """vLLM 도달성 프로브 — /v1/models 를 짧은 타임아웃으로 확인(캐시 30s).

    consolidate 가 vLLM 불통 시 조용히 원본 유지하므로, 운영에서 통합이 영구 비활성
    됐는지 헬스로 드러낸다. 반환 {reachable, base, error}."""
    target_base = (base or VLLM_BASE).rstrip("/")
    now = time.time()
    if (_VLLM_PROBE_CACHE["base"] == target_base
            and now - _VLLM_PROBE_CACHE["ts"] < _VLLM_PROBE_TTL
            and _VLLM_PROBE_CACHE["result"] is not None):
        return _VLLM_PROBE_CACHE["result"]
    result = {"reachable": False, "base": target_base, "error": None}
    try:
        key = os.environ.get("B2B_VLLM_KEY", "khkim")
        req = urllib.request.Request(
            target_base + "/v1/models", method="GET",
            headers={"authorization": f"Bearer {key}", "api-key": key})
        with urllib.request.urlopen(req, timeout=timeout) as resp:
            result["reachable"] = 200 <= int(resp.status) < 500
    except Exception as err:
        result["error"] = str(err)[:200]
    _VLLM_PROBE_CACHE.update({"ts": now, "base": target_base, "result": result})
    return result


def _vllm_chat_once(system, user, base, timeout=30):
    """서버측 vLLM 단발 호출 — 프록시와 같은 엔드포인트(enable_thinking=False)."""
    target = (base or VLLM_BASE).rstrip("/") + "/v1/chat/completions"
    body = json.dumps({
        "model": os.environ.get("B2B_VLLM_MODEL", "Qwen/Qwen3.6-27B-FP8"),
        "messages": [{"role": "system", "content": system},
                     {"role": "user", "content": user + "\n\n/no_think"}],
        "temperature": 0, "max_tokens": 1600, "stream": False,
        "chat_template_kwargs": {"enable_thinking": False},
    }, ensure_ascii=False).encode("utf-8")
    key = os.environ.get("B2B_VLLM_KEY", "khkim")
    req = urllib.request.Request(target, data=body, method="POST", headers={
        "content-type": "application/json", "authorization": f"Bearer {key}", "api-key": key})
    with urllib.request.urlopen(req, timeout=timeout) as resp:
        data = json.loads(resp.read().decode("utf-8", "replace"))
    return data.get("choices", [{}])[0].get("message", {}).get("content", "") or ""


def skill_consolidate(payload, base=""):
    """녹화 스킬 코드를 '기존 ctx 헬퍼'로 최대한 통합(등가 게이트 통과 시에만).

    LLM 재작성은 서버가 vLLM 으로 하고, 등가는 ixicellr.replay.equivalence 가 MemCtx
    재현 digest 로 판정한다. 통합 실패/부적격/비등가면 원본을 그대로 돌려준다(무해)."""
    code = (payload or {}).get("code") or ""
    if not code.strip():
        return {"ok": True, "consolidated": False, "reason": "empty_input", "code": code}
    from ixicellr.replay.equivalence import consolidate_via_llm_reason
    res, reason = consolidate_via_llm_reason(code, lambda s, u: _vllm_chat_once(s, u, base))
    if res is None:
        # 사유(reason)를 실어 조용한 실패를 없앤다 — 특히 llm_unreachable 은 운영에서
        # 통합이 영구 비활성됐음을 뜻하므로 UI/로그가 반드시 알아야 한다.
        if reason == "llm_unreachable":
            try:
                _perf_trace("skill.consolidate.llm_unreachable", base=(base or VLLM_BASE))
            except Exception:
                pass
        return {"ok": True, "consolidated": False, "reason": reason, "code": code}
    cand, ca, cb = res
    return {"ok": True, "consolidated": True, "reason": reason,
            "code": cand, "calls": [ca, cb]}


def _present_live_session_frame(
    session, app, wb,
    left, top, width, height,
    client_left=None, client_top=None, client_width=None, client_height=None,
    viewport_width=None, viewport_height=None,
    skip_position=False,
):
    """frame 모드 표시 경로: 대상 프레임만 배치/표시하고 나머지 라이브 프레임은 파킹.
    포커스/활성화는 일절 주지 않는다(SW_SHOWNA/SWP_NOACTIVATE) — 전환 직후 호스트의
    첫 클릭이 '창 활성화'에 소비되어 씹히는 문제를 막는다.
    반환: 파킹한 peer id 리스트. 프레임 핸들을 못 구하면 None(호출자가 legacy 경로 사용)."""
    # [2A 단일 초크포인트] 실행기(runner) 전면 시(LIVE_RESTORE_SUPPRESSED) 라이브 공유
    # 인스턴스의 프레임은 offscreen 유지하고 되띄우지 않는다. restore/replace/show-only 가
    # 모두 이 함수를 거치므로 여기 한 곳이면 충분. 범위는 _is_live_shared_app 로 한정해
    # read-only 격리미러/파일모드 워커엔 영향 0(안전장치 a).
    # 주의: None 을 반환하면 호출자들의 `if ... is None:` legacy 경로가 창을 강제 표시해
    # 억제가 무력화된다 → '파킹한 peer 없음'을 뜻하는 빈 리스트 [] 를 반환해 legacy 표시를
    # 건너뛰고 프레임을 offscreen 그대로 둔다(계약: 반환=파킹 peer 리스트와 일관).
    if LIVE_RESTORE_SUPPRESSED and _is_live_shared_app(app):
        try:
            session["hidden"] = True
        except Exception:
            pass
        return []
    target_hwnd = _session_frame_hwnd(session, wb)
    if not target_hwnd:
        return None
    try:
        if not bool(app.Visible):
            # 모든 프레임이 파킹된 초기 상태에서만 도달 → 시각적 변화 없이 인스턴스만 켠다.
            app.Visible = True
    except Exception:
        pass
    _set_window_owner_hwnd(target_hwnd, session.get("nativeHostHwnd"))
    _style_live_frame(target_hwnd)
    try:
        win = wb.Windows(1)
        if not bool(win.Visible):
            win.Visible = True  # 과거 경로가 COM 숨김을 남겼어도 회색 빈 프레임이 되지 않게 복구
    except Exception:
        pass
    try:
        wb.Windows(1).WindowState = -4143  # xlNormal
    except Exception:
        pass
    do_position = bool(width and height) and not skip_position
    if not do_position and width and height:
        # skipPosition 요청이어도 실제 창이 파킹(-32000)돼 있으면 재배치한다.
        # (클라 위치캐시는 서버측 파킹/숨김을 모를 수 있음 → '영역 밖에 뜨는 창' 방지)
        try:
            rect_now = win32gui.GetWindowRect(target_hwnd) if win32gui is not None else None
            if rect_now and (rect_now[0] <= -30000 or rect_now[1] <= -30000):
                do_position = True
        except Exception:
            pass
    if do_position:
        _position_excel_window(
            app, left, top, width, height,
            hwnd=target_hwnd, no_activate=True,
            client_left=client_left, client_top=client_top,
            client_width=client_width, client_height=client_height,
            viewport_width=viewport_width, viewport_height=viewport_height,
            show=True,
        )
    else:
        _show_window_na(target_hwnd)
        _raise_excel_hwnd(target_hwnd)
    hidden_ids = _hide_peer_session_frames(session.get("id"), host_hwnd=session.get("nativeHostHwnd"))
    try:
        _ensure_excel_workbook_view(app, wb, make_visible=False, activate=False, maximize_workbook=False, app_level=False)
    except Exception:
        pass
    # [녹화 중 리본/편집] 녹화 시작 이후 '다른 탭으로 전환/새 파일 오픈'으로 이 프레임이 표시될 때도
    # 편집 잠금 해제 + 리본 펼침을 다시 적용한다. 예전엔 녹화 시작 시점의 워크북에만 적용돼,
    # 이후 연 워크북은 리본이 닫히고 편집이 잠긴 채였다(사용자 보고).
    # [클립보드 보존] 단, 이 재적용(Unprotect·SHOW.TOOLBAR 등)은 Excel 복사 모드(마퀴)를
    # 취소한다 — 이미 잠금 해제된 워크북이면 통째로 건너뛴다(ProtectContents 읽기는 무해).
    # 안 그러면 '녹화 중 A에서 복사 → B 탭 전환 → 재적용이 복사 취소 → B에 붙여넣기 불가'.
    # [클립보드 실측] 이 재적용에서 복사 모드(마퀴)를 죽이는 건 정확히
    # SHOW.TOOLBAR(XLM) 하나다(프로브로 이등분 — 파킹/Activate/Select/OnKey 는 무해).
    # ProtectContents 휴리스틱은 사용자가 자체 비밀번호로 보호한 시트(우리 암호로
    # Unprotect 불가)가 있으면 항상 '재적용 필요'로 오판해 탭 전환마다 복사를 죽였다.
    # → 세션당 1회 플래그(recUnlockDone)로 바꾼다: 녹화 시작 시 일괄 해제된 세션은
    # 전환 시 no-op, 녹화 중 새로 연 워크북만 최초 표시 때 1회 해제.
    if RECORDING_EDIT_UNLOCKED and _is_live_shared_app(app) and not session.get("recUnlockDone"):
        try:
            _protect_workbook_for_read_only_mirror(wb, False)
        except Exception:
            pass
        try:
            _restore_excel_default_input(app)
            _enable_excel_context_menus(app)
            _set_excel_ribbon_visible(app, True)
        except Exception:
            pass
        session["recUnlockDone"] = True
    session["hidden"] = False
    return hidden_ids


def _show_only_excel_session_impl(
    excel_id,
    left,
    top,
    width,
    height,
    client_left=None,
    client_top=None,
    client_width=None,
    client_height=None,
    viewport_width=None,
    viewport_height=None,
    browser_title=None,
    native_parent_hwnd=None,
    native_host_hwnd=None,
    native_overlay=False,
    skip_position=False,
):
    with EXCEL_LOCK:
        session = get_excel_session(excel_id)
        app, wb = session_workbook(session)
        if native_parent_hwnd:
            session["nativeParentHwnd"] = native_parent_hwnd
        if native_host_hwnd:
            session["nativeHostHwnd"] = native_host_hwnd
        if native_overlay:
            session["nativeOverlay"] = True
            session["nativeParentHwnd"] = None
        browser_hwnd = session.get("browserHwnd")
        try:
            browser_valid = bool(browser_hwnd and win32gui is not None and win32gui.IsWindow(int(browser_hwnd)))
        except Exception:
            browser_valid = False
        if session.get("readOnlyMirror") and not browser_valid:
            session["browserHwnd"] = _capture_browser_hwnd(browser_title)

        hidden_ids = None
        if session.get("liveEditable") and LIVE_FRAME_MODE:
            # frame 모드: 활성화/포커스 없이 대상 프레임만 제자리 표시, 나머지는 파킹.
            hidden_ids = _present_live_session_frame(
                session, app, wb,
                left, top, width, height,
                client_left=client_left, client_top=client_top,
                client_width=client_width, client_height=client_height,
                viewport_width=viewport_width, viewport_height=viewport_height,
                skip_position=bool(skip_position),
            )
        if hidden_ids is None:
            if session.get("liveEditable"):
                _set_excel_window_owner(app, session.get("nativeHostHwnd"))
            _show_workbook_window(app, wb, activate=True)
            hidden_ids = _hide_peer_workbook_windows(app, excel_id)
            _show_workbook_window(app, wb, activate=True)
            if not skip_position:
                _position_excel_window(
                    app,
                    left,
                    top,
                    width,
                    height,
                    browser_hwnd=None if (session.get("nativeParentHwnd") or session.get("nativeOverlay")) else (session.get("browserHwnd") if session.get("readOnlyMirror") else None),
                    native_parent_hwnd=session.get("nativeParentHwnd") if session.get("readOnlyMirror") and not session.get("nativeOverlay") else None,
                    native_host_hwnd=session.get("nativeHostHwnd") if session.get("readOnlyMirror") else None,
                    native_overlay=bool(session.get("nativeOverlay")) if session.get("readOnlyMirror") else False,
                    client_left=client_left,
                    client_top=client_top,
                    client_width=client_width,
                    client_height=client_height,
                    viewport_width=viewport_width,
                    viewport_height=viewport_height,
                    show=True,
                )
            try:
                _safe_activate_excel_app(app)
            except Exception:
                pass
        if width and height:
            session["liveRect"] = {
                "left": int(float(left or 0)), "top": int(float(top or 0)),
                "width": int(float(width or 0)), "height": int(float(height or 0)),
            }
        session["hidden"] = False
        session["lastNativePositionKey"] = ""
        return {
            "ok": True,
            "excelId": excel_id,
            "hidden": len(hidden_ids),
            "hiddenIds": hidden_ids,
            "skipPosition": bool(skip_position),
            "left": int(float(left or 0)),
            "top": int(float(top or 0)),
            "width": int(float(width or 0)),
            "height": int(float(height or 0)),
        }


def _reopen_excel_session_workbook(session):
    if not excel_available():
        raise RuntimeError("Microsoft Excel COM automation is not available. Excel and pywin32 are required.")
    candidates = [
        session.get("workingCopyPath"),
        session.get("openPath"),
        session.get("path"),
        session.get("sourcePath"),
    ]
    path = next((Path(p) for p in candidates if p and Path(p).exists()), None)
    if path is None:
        raise RuntimeError("Excel workbook file for recovery was not found.")
    live_editable = bool(session.get("liveEditable"))
    read_only_mirror = bool(session.get("readOnlyMirror")) and not live_editable
    app = _get_live_excel_app() if live_editable else win32com.client.DispatchEx("Excel.Application")
    _track_spawned_excel_app(app)  # [0.5.2 이식] 고아 Excel 추적
    if not (live_editable and LIVE_FRAME_MODE):
        # frame 모드에서는 글로벌 Visible 토글 금지(다른 라이브 프레임까지 동시에 사라져
        # 포그라운드 소멸 → 무관한 앱 창 활성화). 새 프레임은 아래에서 개별 파킹한다.
        app.Visible = False
    app.DisplayAlerts = False
    try:
        app.EnableEvents = False
    except Exception:
        pass
    try:
        app.AskToUpdateLinks = False
    except Exception:
        pass
    wb, open_temp_path = excel_workbooks_open(
        app,
        path,
        read_only=read_only_mirror,
        intended_name=session.get("name") or path,
    )
    session.pop("frameHwnd", None)  # 이전 워크북의 죽은 프레임 핸들 무효화
    if live_editable and LIVE_FRAME_MODE:
        new_frame_hwnd = _workbook_window_hwnd(wb)
        if new_frame_hwnd:
            _move_hwnd_offscreen(new_frame_hwnd)
            session["frameHwnd"] = new_frame_hwnd
    session["app"] = app
    session["workbook"] = wb
    session["pid"] = _excel_process_id(app)
    session["path"] = str(path)
    session["openPath"] = str(wb.FullName)
    session["openTempPath"] = str(open_temp_path) if open_temp_path else session.get("openTempPath", "")
    session["snapshots"] = {}
    try:
        _protect_workbook_for_read_only_mirror(wb, True)
    except Exception:
        pass
    try:
        _configure_excel_grid_window(app, wb)
    except Exception:
        pass
    return app, wb, True


def _recover_excel_session_impl(
    excel_id,
    left,
    top,
    width,
    height,
    client_left=None,
    client_top=None,
    client_width=None,
    client_height=None,
    viewport_width=None,
    viewport_height=None,
    browser_title=None,
    native_parent_hwnd=None,
    native_host_hwnd=None,
    native_overlay=False,
):
    with EXCEL_LOCK:
        session = get_excel_session(excel_id)
        if native_parent_hwnd:
            session["nativeParentHwnd"] = native_parent_hwnd
        if native_host_hwnd:
            session["nativeHostHwnd"] = native_host_hwnd
        if native_overlay:
            session["nativeOverlay"] = True
            session["nativeParentHwnd"] = None
        reopened = False
        try:
            app, wb = session_workbook(session)
        except Exception:
            app, wb, reopened = _reopen_excel_session_workbook(session)
        _restore_app_state(app)
        try:
            _restore_live_protected_view(app, wb)
        except Exception:
            pass
        if width and height:
            session["liveRect"] = {
                "left": int(float(left or 0)),
                "top": int(float(top or 0)),
                "width": int(float(width or 0)),
                "height": int(float(height or 0)),
            }
        frame_mode = bool(session.get("liveEditable")) and LIVE_FRAME_MODE
        hidden_ids = None
        if frame_mode:
            hidden_ids = _present_live_session_frame(
                session, app, wb,
                left, top, width, height,
                client_left=client_left, client_top=client_top,
                client_width=client_width, client_height=client_height,
                viewport_width=viewport_width, viewport_height=viewport_height,
                skip_position=not (width and height),
            )
        if hidden_ids is None:
            try:
                _restore_live_window(session, app, wb)
            except Exception:
                _show_workbook_window(app, wb, activate=True)
            hidden_ids = _hide_peer_workbook_windows(app, excel_id)
            try:
                _show_workbook_window(app, wb, activate=True)
            except Exception:
                pass
            try:
                _safe_activate_excel_app(app)
            except Exception:
                pass
        session["hidden"] = False
        session["lastNativePositionKey"] = ""
        try:
            sheet = wb.ActiveSheet.Name if frame_mode else app.ActiveSheet.Name
        except Exception:
            sheet = ""
        try:
            if frame_mode:
                # 활성화 없이도 이 세션 창의 선택 범위를 읽는다(Window.RangeSelection).
                address = _excel_address(wb.Windows(1).RangeSelection).replace("$", "")
            else:
                address = _excel_address(app.Selection).replace("$", "")
        except Exception:
            address = ""
        return {
            "ok": True,
            "excelId": excel_id,
            "activeExcelId": excel_id,
            "reopened": bool(reopened),
            "hidden": len(hidden_ids),
            "hiddenIds": hidden_ids,
            "sheet": sheet,
            "address": address,
        }


def _hide_excel_session_impl(excel_id):
    with EXCEL_LOCK:
        session = get_excel_session(excel_id)
        app, wb = session_workbook(session)
        if session.get("liveEditable") and LIVE_FRAME_MODE:
            hwnd = _session_frame_hwnd(session, wb)
            if hwnd:
                # 포그라운드 프레임을 그냥 치우면 OS 가 무관한 다음 창을 활성화한다 → 호스트로 먼저 핸드오프.
                _handoff_foreground_to_host(session.get("nativeHostHwnd"), [hwnd])
                _move_hwnd_offscreen(hwnd)
            else:
                _hide_excel_app_window(app)
        else:
            _hide_excel_app_window(app)
        session["hidden"] = True
        session["lastNativePositionKey"] = ""
        return {"ok": True, "excelId": excel_id, "hidden": True}


def _hide_all_excel_sessions_impl():
    hidden = 0
    with EXCEL_LOCK:
        sessions = list(EXCEL_SESSIONS.values())
    if LIVE_FRAME_MODE:
        # 라이브 프레임 중 하나가 포그라운드면 먼저 호스트로 포커스를 넘긴다(무관 창 점프 방지).
        frame_hwnds = []
        host_hwnd = None
        for session in sessions:
            if not session.get("liveEditable"):
                continue
            hwnd = session.get("frameHwnd")
            if hwnd:
                frame_hwnds.append(hwnd)
            if not host_hwnd:
                host_hwnd = session.get("nativeHostHwnd")
        if frame_hwnds:
            _handoff_foreground_to_host(host_hwnd, frame_hwnds)
    parked_frame_hwnds = set()   # 아래 '루트 창' 정리에서 제외할, 이미 파킹한 워크북 프레임
    live_frame_pids = set()      # 라이브 공유 인스턴스 pid (그 프로세스만 추가 정리 대상)
    for session in sessions:
        try:
            app, wb = session_workbook(session)
            if session.get("liveEditable") and LIVE_FRAME_MODE:
                hwnd = _session_frame_hwnd(session, wb)
                if hwnd:
                    _move_hwnd_offscreen(hwnd)
                    try:
                        parked_frame_hwnds.add(int(hwnd))
                    except Exception:
                        pass
                else:
                    _hide_excel_app_window(app)
                try:
                    _p = int(session.get("pid") or 0)
                    if _p:
                        live_frame_pids.add(_p)
                except Exception:
                    pass
            else:
                _hide_excel_app_window(app)
            session["hidden"] = True
            session["lastNativePositionKey"] = ""
            hidden += 1
        except Exception:
            pass
    # [회색 빈 Excel — 단계 OFF 후 실행기 이동 실측 2026-08-04]
    # 위 루프는 '워크북 프레임'만 파킹한다. 그런데 되돌리기(replace)는 워크북을 닫았다 다시 여는
    # 과정에서 app.Visible=True 를 켜므로(프레임 hwnd 생성에 필요 — 빼면 최대화 회귀), 그 인스턴스의
    # '워크북 없는 루트 창'이 화면에 남는다. 아래 SPAWNED_EXCEL_PIDS 정리는 session_pids 를 건너뛰어
    # (라이브 인스턴스가 세션 소유라) 이 창을 못 덮었다 → 실행기 화면 위 회색 Excel.
    # 안전장치: ① 라이브 프레임 세션의 pid 만 ② 보이는 XLMAIN 창만 ③ 이미 파킹한 프레임은 제외
    #          ④ 숨김(SW_HIDE)/app.Visible 토글이 아니라 '화면 밖 이동'만 — 활성 창 소멸로 인한
    #            무관 창 최상단 점프가 없고, 사용자가 직접 연 Excel(우리 pid 아님)은 건드리지 않는다.
    try:
        for _hwnd in _visible_excel_top_hwnds_for_pids(live_frame_pids):
            if _hwnd in parked_frame_hwnds:
                continue
            _move_hwnd_offscreen(_hwnd)
    except Exception:
        pass
    # [최소화 회색 창] 세션에 속하지 않은 '우리가 띄운' Excel 창(격리 실행 워커, 복원 경로가
    # app.Visible=True 로 띄운 작업사본, Quit 실패 잔존 등)은 위 루프가 못 덮는다 — 호스트가
    # 최소화되면 그 창이 워크북 0개짜리 회색 'Excel' 로 화면에 드러난다(실측: 결과 편집 후 최소화).
    # SPAWNED_EXCEL_PIDS 는 우리가 만든 프로세스만 추적하므로, 사용자가 직접 연 Excel 은 건드리지
    # 않으면서 세션 외 소유 pid 의 창을 전부 숨긴다.
    try:
        session_pids = {int(s.get("pid") or 0) for s in sessions}
        for pid in {int(p) for p in list(SPAWNED_EXCEL_PIDS) if p}:
            if pid in session_pids:
                continue
            _hide_excel_windows_for_pid(pid)
    except Exception:
        pass
    return {"ok": True, "hidden": hidden}


def _foreground_is_excel_window():
    if win32gui is None:
        return False
    try:
        hwnd = win32gui.GetForegroundWindow()
        if not hwnd:
            return False
        cls = (win32gui.GetClassName(hwnd) or "").upper()
        return "XLMAIN" in cls or "XLDESK" in cls or "EXCEL7" in cls
    except Exception:
        return False


def _hide_inactive_excel_sessions_impl():
    # 앱이 포커스를 잃었을 때 호출. 포그라운드가 Excel(사용자가 미러를 클릭)이면 그대로 두고,
    # 그 외(파일 대화상자/다른 앱/최소화)면 미러를 숨겨 Excel 이 위로 튀어나오지 않게 한다.
    if _foreground_is_excel_window():
        return {"ok": True, "hidden": 0, "foregroundExcel": True}
    return _hide_all_excel_sessions_impl()


def _close_excel_session_impl(excel_id):
    # [녹화 보호] 녹화 중에는 공유 라이브 세션을 절대 닫지 않는다. 이 세션의 워크북을 닫으면
    # 공유 LIVE_EXCEL_APP 의 워크북 수가 줄고(마지막이면 앱 quit) 진행 중 녹화가 통째로 유실된다.
    # 실측(2026-07-28): 녹화 중 캐시 트림(trimExcelMirrorSessionCache→/api/excel/close)이
    # 라이브 인스턴스를 죽여, 정지 시 새 빈 인스턴스에서 돌아 harvested=0. registry pop 전에 막는다.
    if NATIVE_RECORDING.get("active"):
        with EXCEL_LOCK:
            _peek = EXCEL_SESSIONS.get(excel_id)
        if _peek is not None:
            _live = False
            try:
                _live = _is_live_shared_app(_peek.get("app")) or bool(_peek.get("liveEditable"))
            except Exception:
                _live = bool(_peek.get("liveEditable"))
            if _live:
                _note_live_app_reset("close_session_skipped_during_recording",
                                     skipped=True, excelId=excel_id)
                return {"ok": True, "closed": False, "keptAliveForRecording": True}
    with EXCEL_LOCK:
        session = EXCEL_SESSIONS.pop(excel_id, None)
    if not session:
        return {"ok": True, "closed": False}
    pid = session.get("pid")
    if LIVE_FRAME_MODE and session.get("liveEditable"):
        # 닫히는 프레임이 포그라운드면 파괴 직전에 호스트로 포커스를 넘긴다(무관 창 점프 방지).
        try:
            hwnd = session.get("frameHwnd")
            if hwnd:
                _handoff_foreground_to_host(session.get("nativeHostHwnd"), [hwnd])
        except Exception:
            pass
    try:
        app, wb = session_workbook(session)
        _close_companion_workbooks(session, app)
        wb.Close(SaveChanges=False)
        if app.Workbooks.Count == 0:
            _hide_excel_app_window(app)
            app.Quit()
            if _is_live_shared_app(app):
                global LIVE_EXCEL_APP
                _note_live_app_reset("close_session_last_workbook")
                LIVE_EXCEL_APP = None
    except Exception:
        pass
    # 공유 Excel 인스턴스에 같은 pid를 쓰는 다른 세션이 남아 있으면 프로세스를 죽이면 안 된다.
    if pid and not _remaining_sessions_for_pid(pid):
        deadline = time.time() + 1.5
        while time.time() < deadline and _is_pid_alive(pid):
            _hide_excel_windows_for_pid(pid)
            time.sleep(0.1)
        if _is_pid_alive(pid):
            _hide_excel_windows_for_pid(pid)
            _force_kill_pid(pid)
    for key in ("openTempPath", "workingCopyPath"):
        temp_path = session.get(key)
        if temp_path:
            try:
                Path(temp_path).unlink(missing_ok=True)
            except Exception:
                pass
    for cdir in session.get("companionTemps") or []:
        try:
            shutil.rmtree(cdir, ignore_errors=True)
        except Exception:
            pass
    return {"ok": True, "closed": True}


VBA_SKILL_ENTRY = "B2BSkill"


def _strip_vba_comment(line):
    in_string = False
    out = []
    i = 0
    while i < len(line):
        ch = line[i]
        if ch == '"':
            out.append(ch)
            if in_string and i + 1 < len(line) and line[i + 1] == '"':
                out.append(line[i + 1])
                i += 2
                continue
            in_string = not in_string
            i += 1
            continue
        if ch == "'" and not in_string:
            break
        out.append(ch)
        i += 1
    return "".join(out)


def _looks_like_long_digit_identifier(value):
    if not isinstance(value, str):
        return False
    s = value.strip()
    if s.startswith("'"):
        s = s[1:].strip()
    return bool(re.fullmatch(r"\d{12,}", s))


def _long_digit_identifier_columns(grid):
    columns = set()
    for row in grid or []:
        if not isinstance(row, (list, tuple)):
            continue
        for idx, value in enumerate(row):
            if _looks_like_long_digit_identifier(value):
                columns.add(idx)
    return sorted(columns)


def _apply_com_text_format_for_long_digit_columns(ws, grid, start_row=1, start_col=1):
    columns = _long_digit_identifier_columns(grid)
    if not columns:
        return
    row_count = max(1, len(grid or []))
    for idx in columns:
        try:
            col = int(start_col) + int(idx)
            rng = ws.Range(ws.Cells(int(start_row), col), ws.Cells(int(start_row) + row_count - 1, col))
            rng.NumberFormat = "@"
        except Exception:
            continue


def _apply_openpyxl_text_format_for_long_digit_columns(raw_ws, grid, start_row=1, start_col=1):
    columns = _long_digit_identifier_columns(grid)
    if not columns:
        return
    row_count = max(1, len(grid or []))
    for idx in columns:
        col = int(start_col) + int(idx)
        for row in range(int(start_row), int(start_row) + row_count):
            try:
                raw_ws.cell(row=row, column=col).number_format = "@"
            except Exception:
                continue


# [VBA 보안 게이트] 녹화/생성 VBA 가 실행 PC 에서 임의 파일을 열고·저장하고·프로세스를 띄우는 것을
# 주입 전에 차단한다. MS 매크로 레코더는 녹화 중 사용자의 파일 열기(Workbooks.Open "C:\녹화PC경로\..."),
# 다른 이름 저장(SaveAs), 닫기까지 전부 절대경로로 기록하는데, sanitize 는 스크롤 제거뿐이라 그대로
# 재생됐다(게다가 실행 직전 AutomationSecurity 를 Low 로 낮춰 열린 파일의 매크로까지 실행됨).
# python 엔진의 AST 금지목록(SaveAs/Close/Quit 등)과 대칭 — VBA 만 무검사였던 비대칭을 닫는다.
# 문자열 리터럴·주석 제거 후 검사해 셀 값("Shell 주유소" 등) 오탐을 막는다.
_VBA_FORBIDDEN_BARE = [
    # (패턴, 사용자 안내) — 문자열/주석 제거된 텍스트에 적용
    (re.compile(r"\bWorkbooks\s*\.\s*Open\b", re.I),
     "파일 열기(Workbooks.Open) — 녹화 중 다른 파일을 열면 실행 PC의 그 경로를 그대로 열게 됩니다. "
     "필요한 파일은 미리 함께 열어두거나 업로드하고 다시 녹화하세요."),
    (re.compile(r"\.\s*Save(?:As|CopyAs)\b", re.I),
     "다른 이름으로 저장(SaveAs/SaveCopyAs) — 실행 PC의 임의 경로에 파일을 쓰게 됩니다. "
     "저장은 실행이 끝나면 자동으로 처리되니 녹화에서 빼세요."),
    (re.compile(r"\.\s*Close\b", re.I),
     "워크북/창 닫기(.Close) — 파이프라인이 관리하는 파일이 닫혀 이후 단계가 깨집니다. "
     "닫기는 녹화에서 빼세요."),
    (re.compile(r"\.\s*PrintOut\b", re.I), "인쇄(PrintOut)"),
    (re.compile(r"\.\s*SendMail\b", re.I), "메일 발송(SendMail)"),
    (re.compile(r"\.\s*Quit\b", re.I), "Excel 종료(.Quit)"),
    (re.compile(r"\bCh(?:Dir|Drive)\b", re.I), "작업 폴더 변경(ChDir/ChDrive)"),
    (re.compile(r"\bShell\b", re.I), "외부 프로그램 실행(Shell)"),
    (re.compile(r"\bKill\b", re.I), "파일 삭제(Kill)"),
    (re.compile(r"\bEnviron\b", re.I), "환경변수 접근(Environ)"),
    (re.compile(r"\bSendKeys\b", re.I), "키 입력 주입(SendKeys)"),
]
_VBA_FORBIDDEN_RAW = [
    # 문자열 안 내용을 봐야 하는 패턴 — 원문에 적용(Scripting.Dictionary 는 허용해야 하므로
    # CreateObject 전면 금지는 안 됨).
    (re.compile(r"CreateObject\s*\(\s*[\"'](?:WScript|Shell\.Application|Scripting\.FileSystemObject)", re.I),
     "시스템 개체 생성(WScript/Shell.Application/FileSystemObject)"),
    (re.compile(r"\bGetObject\s*\(", re.I), "외부 개체 연결(GetObject)"),
    (re.compile(r"FileSystemObject", re.I), "파일시스템 접근(FileSystemObject)"),
]


def _normalize_vba_llm_comment_slips(code):
    """LLM 이 VBA 에 C 계열 주석(//)을 섞는 사고 교정 — 줄머리 // 는 ' 주석으로 변환.

    // 는 VBA 컴파일 오류인데, 컴파일 오류 모듈에 Application.Run 을 하면 숨김 격리
    인스턴스에서 VBE 모달이 떠 영구 블록된다(실측 14:45: '의도 반영' 코드의 // 주석 4줄로
    '녹화 재현 중' 무한 대기 + 중단 불능 → 강제종료). 주석 변환은 의미 무손실이라
    실패 대신 구조한다. 줄 중간 // 는 문자열("http://...") 오탐 위험이 있어 안 건드리고
    _validate_vba_source_before_inject 가 명확한 에러로 거부한다."""
    out = []
    for line in str(code or "").splitlines():
        stripped = line.lstrip()
        if stripped.startswith("//"):
            indent = line[:len(line) - len(stripped)]
            out.append(indent + "'" + stripped[2:])
        else:
            out.append(line)
    return "\n".join(out)


def _vba_strip_strings_and_comments(code):
    """따옴표 문자열("" 이스케이프 포함) → 빈 문자열로, 이후 ' 주석 제거 — 키워드 오탐 방지."""
    text = re.sub(r'"(?:[^"]|"")*"', '""', str(code or ""))
    out = []
    for line in text.splitlines():
        i = line.find("'")
        out.append(line[:i] if i >= 0 else line)
    return "\n".join(out)


def _vba_security_scan(code):
    """금지 구문 발견 시 사용자 안내 문자열 반환, 없으면 None."""
    stripped = _vba_strip_strings_and_comments(code)
    for pat, why in _VBA_FORBIDDEN_BARE:
        if pat.search(stripped):
            return why
    raw = str(code or "")
    for pat, why in _VBA_FORBIDDEN_RAW:
        if pat.search(raw):
            return why
    return None


def _validate_vba_source_before_inject(code):
    """VBE 디버거를 띄우는 명백한 컴파일 오류는 Excel에 주입하기 전에 차단한다."""
    _sec = _vba_security_scan(code)
    if _sec:
        raise RuntimeError(
            "VBA 보안 검사: 이 스킬에 실행할 수 없는 동작이 들어 있습니다 — %s "
            "(녹화라면 해당 동작 없이 다시 녹화해 주세요.)" % _sec)
    # [// 주석 잔존 차단] 줄머리 // 는 normalize 가 ' 로 변환하지만, 줄 중간 // 는 컴파일 오류
    # 그대로다(문자열 제거 후 검사라 "http://..." 오탐 없음). 숨김 인스턴스에서 VBE 모달로
    # 영구 블록되는 대신 즉시 명확한 에러를 낸다.
    if "//" in _vba_strip_strings_and_comments(code):
        raise RuntimeError(
            "VBA 문법 오류: '//' 주석은 VBA 에서 지원되지 않습니다(작은따옴표 ' 를 쓰세요). "
            "컴파일 오류로 실행이 멈추는 것을 막기 위해 실행 전에 차단했습니다.")
    lines = str(code or "").splitlines()
    block_stack = []
    code_text = str(code or "")
    bad_col_assignments = [
        ("BP", 58, 68, re.compile(r"\b(?=[A-Za-z0-9_]*bp)(?=[A-Za-z0-9_]*col)[A-Za-z_][A-Za-z0-9_]*\s*=\s*58\b", re.IGNORECASE)),
        ("BQ", 59, 69, re.compile(r"\b(?=[A-Za-z0-9_]*bq)(?=[A-Za-z0-9_]*col)[A-Za-z_][A-Za-z0-9_]*\s*=\s*59\b", re.IGNORECASE)),
    ]
    for col, wrong, expected, pattern in bad_col_assignments:
        if pattern.search(code_text):
            raise RuntimeError(
                'VBA 안전 검사: %s열 번호를 %d로 하드코딩했습니다. %s열은 %d입니다. '
                'ws.Columns("%s").Column 또는 ws.Cells(r, "%s")처럼 열 문자를 그대로 사용하세요.'
                % (col, wrong, col, expected, col, col)
            )
    multi_value_lookup_shape = (
        re.search(r"\bBP\b", code_text, re.IGNORECASE)
        and re.search(r"\bBQ\b", code_text, re.IGNORECASE)
        and re.search(r"\b(?:token|tokens|Split)\b", code_text, re.IGNORECASE)
        and re.search(r"\b(?:hCol|targetColOut|targetCol|outCol|resultCol)\b|[\"']H[\"']", code_text, re.IGNORECASE)
    )
    if multi_value_lookup_shape:
        for raw in lines:
            line = _strip_vba_comment(raw).strip()
            if re.search(r"\bRange\s*\(", line, re.IGNORECASE) \
                    and re.search(r"\.\s*(?:Value|Value2)\s*=\s*[A-Za-z_][A-Za-z0-9_]*(?:Arr|Array|Values|Data)\b", line, re.IGNORECASE) \
                    and re.search(r"\b(?:hCol|targetColOut|targetCol|outCol|resultCol)\b|Cells\s*\([^)]*,\s*(?:8|[\"']H[\"'])\s*\)|[\"']H", line, re.IGNORECASE):
                raise RuntimeError(
                    "VBA 안전 검사: 다중 가입번호 매칭 결과를 H열 전체 범위에 배열로 다시 쓰고 있습니다. "
                    "매칭된 데이터 행의 H셀만 개별 갱신하고, 미매칭/합계/수식 행은 그대로 두세요."
                )

    def push(kind, line_no, expected):
        block_stack.append((kind, line_no, expected))

    def pop(expected_kind, end_text, line_no):
        if not block_stack or block_stack[-1][0] != expected_kind:
            raise RuntimeError("VBA 문법 오류(%d행): 대응되는 %s 블록이 없습니다." % (line_no, end_text))
        block_stack.pop()

    for idx, raw in enumerate(lines, 1):
        line = _strip_vba_comment(raw).strip()
        if not line or line.endswith("_"):
            continue
        if re.search(r"\bContinue\s+For\b", line, re.IGNORECASE):
            raise RuntimeError(
                "VBA 문법 오류(%d행): Excel VBA는 Continue For를 지원하지 않습니다. "
                "If Len(...) > 0 Then ... End If 또는 GoTo 라벨을 사용하세요." % idx
            )
        if re.search(r"\b[A-Za-z_][A-Za-z0-9_]*\s*\([^()\r\n]*,\s*\)", line):
            raise RuntimeError(
                "VBA 문법 오류(%d행): 함수/배열 호출에서 쉼표 뒤 인수가 비어 있습니다. "
                "예: dataArr(1, ) 같은 코드는 실행할 수 없습니다." % idx
            )
        if re.search(r"\bAs(?:\s+New)?\s*$", line, re.IGNORECASE):
            raise RuntimeError("VBA 문법 오류(%d행): As 뒤의 자료형이 비어 있습니다." % idx)
        # 기호 연산자로 끝나면 미완성 식. 단어 연산자(And/Or/Xor/Mod)는 반드시 \b 단어경계로
        # 검사해야 "Exit For"(끝이 'or'), 변수명 color/vendor/cursor 등을 오탐하지 않는다.
        if re.search(r"(?:,|\+|-|\*|/|&|=|<>|<=|>=|<|>)\s*$", line) \
                or re.search(r"\b(?:And|Or|Xor|Mod)\s*$", line, re.IGNORECASE):
            raise RuntimeError("VBA 문법 오류(%d행): 줄 끝의 식이 완성되지 않았습니다." % idx)
        if re.match(r"^End\s+Sub\b", line, re.IGNORECASE):
            pop("Sub", "Sub", idx)
            continue
        if re.match(r"^End\s+Function\b", line, re.IGNORECASE):
            pop("Function", "Function", idx)
            continue
        if re.match(r"^End\s+If\b", line, re.IGNORECASE):
            pop("If", "If", idx)
            continue
        if re.match(r"^End\s+With\b", line, re.IGNORECASE):
            pop("With", "With", idx)
            continue
        if re.match(r"^End\s+Select\b", line, re.IGNORECASE):
            pop("Select", "Select", idx)
            continue
        if re.match(r"^Next\b", line, re.IGNORECASE):
            pop("For", "For", idx)
            continue
        if re.match(r"^Loop\b", line, re.IGNORECASE):
            pop("Do", "Do", idx)
            continue
        if re.match(r"^(?:(?:Public|Private|Friend|Static)\s+)?Sub\b", line, re.IGNORECASE):
            push("Sub", idx, "End Sub")
            continue
        if re.match(r"^(?:(?:Public|Private|Friend|Static)\s+)?Function\b", line, re.IGNORECASE):
            push("Function", idx, "End Function")
            continue
        if re.match(r"^If\b", line, re.IGNORECASE) and re.search(r"\bThen\s*$", line, re.IGNORECASE):
            push("If", idx, "End If")
            continue
        if re.match(r"^For\b", line, re.IGNORECASE):
            push("For", idx, "Next")
            continue
        if re.match(r"^Do\b", line, re.IGNORECASE):
            push("Do", idx, "Loop")
            continue
        if re.match(r"^With\b", line, re.IGNORECASE):
            push("With", idx, "End With")
            continue
        if re.match(r"^Select\s+Case\b", line, re.IGNORECASE):
            push("Select", idx, "End Select")
            continue
    if block_stack:
        kind, line_no, expected = block_stack[-1]
        raise RuntimeError("VBA 문법 오류(%d행): %s가 없습니다." % (line_no, expected))


def _vba_pipeline_step_info(step, fallback_idx, err):
    _code = (step.get("code") if isinstance(step, dict) else str(step or "")) or ""
    _cause, _guide = _pipeline_error_guide(str(err), _code)
    _msg = f"{_cause}\n💡 이렇게 요청해 보세요: {_guide}\n(자세히: {err})"
    if isinstance(step, dict):
        raw_idx = step.get("stepIdx")
        try:
            step_idx = int(raw_idx)
        except Exception:
            step_idx = int(fallback_idx)
        return {
            "stepIdx": step_idx,
            "stepId": step.get("stepId") or None,
            "description": step.get("description") or "",
            "code": step.get("code") or "",
            "language": step.get("language") or "vba",
            "message": _msg,
            "cause": _cause,
            "promptGuide": _guide,
            "rawError": str(err),
            "stack": "",
        }
    return {
        "stepIdx": int(fallback_idx),
        "stepId": None,
        "description": "",
        "code": str(step or ""),
        "language": "vba",
        "message": _msg,
        "cause": _cause,
        "promptGuide": _guide,
        "rawError": str(err),
        "stack": "",
    }


def _wrap_vba_skill_code(code, entry):
    """사용자 VBA를 내부 Sub로 바꾸고, 런타임 오류를 팝업 대신 상태값으로 전달하는 래퍼를 붙인다."""
    entry = (entry or VBA_SKILL_ENTRY).strip() or VBA_SKILL_ENTRY
    impl_name = "B2B_UserSkill_Impl"
    runner_name = "B2B_RunSkill"
    err_num_name = "B2B_GetLastErrNumber"
    err_desc_name = "B2B_GetLastErrDescription"
    pattern = re.compile(
        r"^(\s*)(?:(Public|Private)\s+)?Sub\s+%s\s*\([^)]*\)"
        % re.escape(entry),
        re.IGNORECASE | re.MULTILINE,
    )
    wrapped_user_code, count = pattern.subn(
        r"\1Private Sub %s()" % impl_name,
        code,
        count=1,
    )
    if count == 0:
        wrapped_user_code = code
        call_line = "%s" % entry
    else:
        call_line = "%s" % impl_name
    declarations = """
Private B2B_LastErrNumber As Long
Private B2B_LastErrDescription As String
"""
    option_match = re.match(r"(?is)^((?:\s*Option\s+[^\r\n]+\r?\n)+)", wrapped_user_code)
    if option_match:
        pos = option_match.end(1)
        wrapped_user_code = wrapped_user_code[:pos] + declarations + wrapped_user_code[pos:]
    else:
        wrapped_user_code = declarations + wrapped_user_code
    wrapper = f"""

Public Sub {runner_name}()
    On Error GoTo B2B_Err
    B2B_LastErrNumber = 0
    B2B_LastErrDescription = vbNullString
    {call_line}
    Exit Sub
B2B_Err:
    B2B_LastErrNumber = Err.Number
    B2B_LastErrDescription = Err.Description
End Sub

Public Function {err_num_name}() As Long
    {err_num_name} = B2B_LastErrNumber
End Function

Public Function {err_desc_name}() As String
    {err_desc_name} = B2B_LastErrDescription
End Function
"""
    return wrapped_user_code + "\n" + wrapper, runner_name, err_num_name, err_desc_name


def _extract_vba_source_for_injection(code, entry=None):
    """Saved skills can contain the assistant reply text around the VBA block.

    Excel's VBE accepts AddFromString even when the module later cannot compile; the
    following Application.Run then reports the misleading "macro cannot run" error.
    Strip markdown/title/reference text here so loaded .zip skills execute the same
    way as freshly generated chat code.
    """
    text = str(code or "").replace("\r\n", "\n").replace("\r", "\n").strip()
    if not text:
        return text

    fence = re.search(r"```(?:vba|vb|visual\s*basic)?\s*\n([\s\S]*?)```", text, re.IGNORECASE)
    if fence:
        text = fence.group(1).strip()

    lines = text.split("\n")
    # Drop obvious non-VBA wrappers that older auto-save/export paths sometimes
    # placed before the real Sub.
    while lines:
        stripped = lines[0].lstrip()
        if not stripped:
            lines.pop(0)
            continue
        if stripped.startswith("//") or stripped.startswith("#"):
            lines.pop(0)
            continue
        if re.match(r"^(?:제목|title|설명|정확\s*참조|\[정확\s*참조\])\s*[:\]]", stripped, re.IGNORECASE):
            lines.pop(0)
            continue
        break
    text = "\n".join(lines).strip()

    entry = (entry or VBA_SKILL_ENTRY).strip() or VBA_SKILL_ENTRY
    entry_re = re.compile(r"^\s*(?:(?:Public|Private|Friend|Static)\s+)?Sub\s+%s\s*\(" % re.escape(entry), re.IGNORECASE | re.MULTILINE)
    first_sub_re = re.compile(r"^\s*(?:(?:Public|Private|Friend|Static)\s+)?Sub\s+\w+\s*\(", re.IGNORECASE | re.MULTILINE)
    option_re = re.compile(r"^\s*Option\s+[^\n]+\n", re.IGNORECASE | re.MULTILINE)

    m = entry_re.search(text) or first_sub_re.search(text)
    if m and m.start() > 0:
        prefix = text[:m.start()]
        # Keep leading Option lines only; discard prose/markdown/reference text.
        options = "".join(mm.group(0) for mm in option_re.finditer(prefix))
        text = (options + text[m.start():]).strip()

    # Remove any trailing markdown/prose after the last End Sub when no helper
    # Function follows. This is deliberately conservative.
    end_matches = list(re.finditer(r"^\s*End\s+Sub\b.*$", text, re.IGNORECASE | re.MULTILINE))
    if end_matches:
        last_end = end_matches[-1]
        tail = text[last_end.end():]
        if tail.strip() and not re.search(r"^\s*(?:(?:Public|Private|Friend|Static)\s+)?Function\s+\w+\s*\(", tail, re.IGNORECASE | re.MULTILINE):
            text = text[:last_end.end()].strip()

    return text


def _vba_macro_ref(wb, module_name, macro_name):
    """Return a workbook-qualified macro reference for Application.Run.

    When multiple workbooks are open, running "Module1.Macro" can resolve against
    the active workbook instead of the workbook that received the temporary
    module. Qualifying with the workbook name makes injected VBA deterministic.
    """
    try:
        wb_name = str(wb.Name)
    except Exception:
        wb_name = ""
    if wb_name:
        wb_name = wb_name.replace("'", "''")
        return "'%s'!%s.%s" % (wb_name, module_name, macro_name)
    return "%s.%s" % (module_name, macro_name)


def _vba_macro_refs(wb, module_name, macro_name):
    refs = []
    def add(value):
        if value and value not in refs:
            refs.append(value)
    try:
        wb_name = str(wb.Name)
    except Exception:
        wb_name = ""
    try:
        wb_fullname = str(wb.FullName)
    except Exception:
        wb_fullname = ""
    if wb_name:
        add("'%s'!%s.%s" % (wb_name.replace("'", "''"), module_name, macro_name))
    if wb_fullname:
        add("'%s'!%s.%s" % (wb_fullname.replace("'", "''"), module_name, macro_name))
    add("%s.%s" % (module_name, macro_name))
    add(macro_name)
    return refs


def _run_vba_macro_any_ref(app, host_wb, module_name, macro_name):
    last_err = None
    refs = _vba_macro_refs(host_wb, module_name, macro_name)
    _vba_trace("vba.macro.refs", host=_trace_workbook_info(host_wb), moduleName=module_name, macro=macro_name, refs=refs)
    for ref in refs:
        try:
            try:
                host_wb.Activate()
            except Exception:
                pass
            result = app.Run(ref)
            _vba_trace("vba.macro.ref.ok", ref=ref, moduleName=module_name, macro=macro_name, result=_trace_text(result, 120))
            return result
        except Exception as err:
            last_err = err
            _vba_trace("vba.macro.ref.fail", ref=ref, moduleName=module_name, macro=macro_name, error=str(err))
            try:
                _diag_vba_log_line("VBA-RUN-REF-FAIL ref=%r err=%r" % (ref, str(err)[:180]))
            except Exception:
                pass
    if last_err is not None:
        raise last_err
    raise RuntimeError("VBA 매크로 참조를 만들 수 없습니다.")


def _workbook_name_lookup_key(value):
    """Normalize workbook names for generated-code lookups.

    Users and browser upload flows sometimes preserve URL-escaped spaces in file
    names (``%20``) while the LLM writes the human-visible space. Treat those as
    the same workbook name, without doing broad fuzzy matching.
    """
    text = str(value or "")
    try:
        text = unquote(text)
    except Exception:
        pass
    text = text.replace("\u00a0", " ")
    text = re.sub(r"\s+", " ", text).strip().lower()
    return text


_GENERATED_WORKBOOK_PREFIX_RE = re.compile(r"^(?:[0-9a-f]{12,}|excel_open_[0-9a-f]{12,}|live_reset_[0-9a-f]{12,})[_-]+", re.I)


def _strip_generated_workbook_prefix(value):
    text = str(value or "").strip()
    if not text:
        return text
    return _GENERATED_WORKBOOK_PREFIX_RE.sub("", text)


# [월/날짜만 다른 저장 스킬 재사용] 4월용으로 저장한 스킬을 5월 파일에 돌릴 때처럼, 파일명의 '바뀌는' 부분
# (날짜·월·년·분기·버전·앞머리 순번)만 제거해 '같은 템플릿'을 같게 본다. 공백/언더바/괄호는 기존 정규화가
# 이미 처리하므로 여기선 '월/날짜류'만 추가로 지운다.
_VOLATILE_NAME_TOKENS = [
    # 구분자 있는 날짜(2026-03-01 / 2026_3_1 / 2026 03 01). 구분자 없는 20260301·202606·260607 은
    # 아래 '날짜 모양' 토큰이 맡는다. 예전엔 구분자가 선택이라 이 정규식이 임의의 6~8자리 숫자
    # (거래처코드 500255 = "5002"+"5"+"5")를 날짜로 먹어치워 식별번호가 통째로 사라졌다.
    (re.compile(r"(?<!\d)(\d{4})[-_.\s]+(\d{1,2})[-_.\s]+(\d{1,2})\s*일?(?!\d)"),
     lambda m: " " if _looks_like_ymd(m.group(1), m.group(2), m.group(3)) else m.group(0)),
    (re.compile(r"\d{2,4}\s*년"), " "),                                        # 2026년 / 26년
    (re.compile(r"\d{1,2}\s*월"), " "),                                        # 3월 / 03월
    (re.compile(r"\d{1,2}\s*분기"), " "),                                      # 1분기
    (re.compile(r"(?<![A-Za-z0-9])v?\d+(?:\.\d+)+", re.I), " "),               # 버전 1.2 / v1.2.3
    # 시각(10_55_33 / 09:23:01) — 실제 배포 파일명이 "..._2026-07-14 10_55_33_DSMC_..." 라
    # 월 재바인딩에 반드시 필요하다. 예전엔 아래 순번 토큰이 이름 중간의 1~3자리를 닥치는 대로
    # 지우면서 '우연히' 시각까지 지워 동작했는데, 그 부작용으로 지점번호 같은 식별자도 함께
    # 사라졌다(다른 실체 파일이 같은 키가 됨) → 시각은 시각으로 정확히 지운다.
    (re.compile(r"(?<!\d)(\d{1,2})[:_.\-](\d{1,2})[:_.\-](\d{1,2})(?!\d)"),
     lambda m: " " if _looks_like_hms(m.group(1), m.group(2), m.group(3)) else m.group(0)),
    # 앞머리 순번 "03." "05." — 이름 '중간'의 번호(지점 105 결산)는 식별자이므로 보존한다.
    (re.compile(r"^\d{1,3}(?=\s*[.\s_\-])"), " "),
    # YYMMDD/YYYYMM/YYYYMMDD '날짜 모양'일 때만 제거. 예전엔 6~8자리 숫자를 무조건 지워
    # 거래처코드·계약번호(500255 vs 610344)까지 같은 키가 돼 서로 다른 파일이 '유일 매칭'으로
    # 조용히 오매칭됐다.
    (re.compile(r"(?<!\d)(?:\d{8}|\d{6})(?!\d)"), lambda m: " " if _looks_like_date_number(m.group(0)) else m.group(0)),
]

# 접미사 토큰은 서로 조합될 수 있어("X (2) - 복사본") 아래 고정점 루프에서 반복 적용한다.
_VOLATILE_SUFFIX_TOKENS = [
    # 브라우저/윈도우 중복 다운로드 접미사 — "파일 (2).xlsx". 항상 구분자 뒤에 붙으므로 구분자를
    # 요구한다: 의미 있는 말미 일련번호("명세서(2)" = 2사업소)까지 흡수하면 안 된다.
    (re.compile(r"[\s_\-]+\(\s*\d{1,3}\s*\)\s*$"), " "),
    # "파일 - 복사본.xlsx" / "file - Copy.xlsx". 구분자를 요구한다: 예전엔 좌측 경계가 없어
    # 'hardcopy' 의 어미까지 잘려 'hard' 와 키가 충돌했다.
    (re.compile(r"[-_\s]+(?:복사본|copy)\s*$", re.I), " "),
]


def _looks_like_hms(hour, minute, second):
    """파일명에 찍히는 시각(10_55_33)인지 — 00~23 / 00~59 / 00~59 만 참."""
    try:
        hh, mm, ss = int(hour), int(minute), int(second)
    except Exception:
        return False
    return 0 <= hh <= 23 and 0 <= mm <= 59 and 0 <= ss <= 59


def _looks_like_ymd(year, month, day):
    """구분자 있는 날짜(2026-03-01)인지 — 연 1900~2199 / 월 1~12 / 일 1~31 만 참."""
    try:
        y, mm, dd = int(year), int(month), int(day)
    except Exception:
        return False
    return 1900 <= y <= 2199 and 1 <= mm <= 12 and 1 <= dd <= 31


def _looks_like_date_number(digits):
    """YYMMDD(260607) / YYYYMM(202606) / YYYYMMDD(20260607) 처럼 날짜로 읽히는 숫자만 True."""
    def _ok(mm, dd=None):
        if not 1 <= mm <= 12:
            return False
        return dd is None or 1 <= dd <= 31
    try:
        if len(digits) == 6:
            if _ok(int(digits[2:4]), int(digits[4:6])):        # YYMMDD
                return True
            year = int(digits[0:4])                            # YYYYMM
            return 1900 <= year <= 2199 and _ok(int(digits[4:6]))
        if len(digits) == 8:
            year = int(digits[0:4])                            # YYYYMMDD
            return 1900 <= year <= 2199 and _ok(int(digits[4:6]), int(digits[6:8]))
    except Exception:
        return False
    return False


def _stable_workbook_key(name):
    """'같은 템플릿, 다른 월/날짜/버전' 파일을 같게 보기 위한 안정 키(소문자·기호제거)."""
    s = str(name or "")
    try:
        s = unquote(s)
    except Exception:
        pass
    s = Path(s).stem                              # 디렉토리/확장자 제거
    s = _strip_generated_workbook_prefix(s)
    s = s.lower()
    s = re.sub(r"[​-‍﻿]", "", s)   # zero-width(다운로드 경로에서 섞여 들어옴)
    for rx, rep in _VOLATILE_NAME_TOKENS:
        s = rx.sub(rep, s)
    # 접미사는 조합될 수 있다("X (2) - 복사본"). 예전엔 목록 순서대로 1회씩만 적용해
    # 복사본 제거 후 드러나는 "(2)" 가 다시 처리되지 않아 흡수가 되다 말다 했다.
    for _ in range(4):
        prev = s
        for rx, rep in _VOLATILE_SUFFIX_TOKENS:
            s = rx.sub(rep, s)
        if s == prev:
            break
    s = re.sub(r"[\s_\-().\[\]]+", "", s)
    return s


def _user_facing_workbook_names(app):
    """사용자에게 보여줄 수 있는(=코드에 그대로 써도 되는) 열린 워크북 이름만.

    내부 작업본/변환본(excel_open_<uuid>.<ext>, <hash>_원본명.xlsx, live_reset_<hash>_…)과
    개인 매크로 통합문서(PERSONAL.XLSB) 같은 애드인은 제외한다. 이 목록은 오류 힌트에 실려
    LLM 재시도 루프로 되먹여지므로, 내부명이 새어 나가면 모델이 그 이름을 코드에 박아
    (사용자 원본명 별칭 체계를 우회) 다음 실행에서 깨진다.
    """
    names = []
    try:
        raw = [str(wb.Name) for wb in app.Workbooks]
    except Exception:
        return names
    for name in raw:
        stem = str(Path(name).stem)
        if re.match(r"^(?:excel_open_|live_reset_)[0-9a-f]{8,}", stem, re.I):
            continue
        if re.match(r"^[0-9a-f]{12,}[_-]", stem):
            continue
        if stem.upper() in ("PERSONAL", "PERSONAL.XLSB"):
            continue
        names.append(name)
    return names


def _match_workbook_by_stable_key(names, requested):
    """열린 워크북 이름들 중 requested 와 '월/날짜 무시 안정 키'가 같은 것이 '정확히 하나'면 그 이름 반환.
    (엉뚱한 파일 오매칭 방지: 유일 매칭일 때만. 키가 너무 짧으면(<4) 매칭 안 함.)"""
    want = _stable_workbook_key(requested)
    if len(want) < 4:
        return None
    hits = [n for n in (names or []) if _stable_workbook_key(n) == want]
    return hits[0] if len(hits) == 1 else None


def _workbook_name_lookup_keys(value):
    """Return conservative lookup keys for workbook-name resolution.

    전체실행/격리/HTML 우회 경로는 실제 열린 파일명이
    ``<hash>_원본명.xlsx`` 또는 ``excel_open_<hash>.html`` 처럼 바뀔 수 있다.
    모델이 생성한 코드는 사용자가 본 원본 파일명을 쓰는 것이 맞으므로, 정확명
    매칭에 더해 확장자 제외명과 해시 접두어 제거명만 추가 후보로 본다. 모호하면
    사용하지 않는다.
    """
    raw = str(value or "").strip()
    if not raw:
        return set()
    values = {raw}
    try:
        values.add(unquote(raw))
    except Exception:
        pass
    try:
        values.add(Path(raw).name)
    except Exception:
        pass
    keys = set()
    for val in list(values):
        if not val:
            continue
        base = Path(str(val)).name
        stem = Path(base).stem
        suffix = Path(base).suffix
        candidates = {base, stem}
        stripped_base = _strip_generated_workbook_prefix(base)
        stripped_stem = _strip_generated_workbook_prefix(stem)
        candidates.update({stripped_base, stripped_stem})
        if suffix and stripped_stem:
            candidates.add(stripped_stem + suffix)
        for candidate in candidates:
            key = _workbook_name_lookup_key(candidate)
            if key:
                keys.add(key)
    return keys


def _resolve_open_workbook_name(app, requested_name):
    """Return the actual open workbook name matching requested_name.

    Exact name wins. Otherwise match only after URL-decoding and whitespace
    normalization. Ambiguous normalized matches are ignored.
    """
    requested = str(requested_name or "")
    if not requested:
        return requested
    try:
        names = [str(wb.Name) for wb in app.Workbooks]
    except Exception:
        return requested
    for name in names:
        if name == requested:
            return name
    req_keys = _workbook_name_lookup_keys(requested)
    matches = [name for name in names if req_keys & _workbook_name_lookup_keys(name)]
    if len(matches) == 1:
        return matches[0]
    # [월/날짜만 다른 저장 스킬 재사용] 안정 키(날짜·월·버전·순번 제거)로 '유일' 매칭이면 그 파일로 바인딩.
    stable = _match_workbook_by_stable_key(names, requested)
    if stable and stable != requested:
        try:
            _vba_trace("workbook.stable_key_match", requested=str(requested), matched=str(stable))
        except Exception:
            pass
        return stable
    return requested


# [포맷 위장 파일 별칭] .xls 로 위장한 HTML/CSV/xlsx 는 excel_compatible_open_path 가 excel_open_<uuid>.<ext> 로
# 변환·리네임해 연다 → 실제 wb.Name 이 등록명(멘션의 500255...xls)과 공유 부분 없이 달라져, 모델이 만든
# Workbooks("500255...xls") 가 'subscript out of range' 로 실패한다. 열 때 '등록명→실제명' 별칭을 앱(pid)별로
# 저장해 두고, VBA 리터럴 치환에서 그 별칭으로 실제명을 풀어준다(시트명은 이미 실제명이라 안 건드림).
_WB_NAME_ALIASES = {}  # pid -> { lookup_key(등록명): set(실제 wb.Name) }
# [SBAGENT-209] 역방향(실제 wb.Name → 등록명). 복붙 캡처가 코드에 '실제 라이브명'(excel_open_<hash>.xls)을
# 박아 저장하면, 그 세션이 죽은 뒤엔 어떤 파일과도 매칭 불가 — 실행기 파일확인에 영원히 못 채우는
# 요구가 뜨고 재생도 깨진다. 캡처는 이 역별칭으로 사용자 원본명을 코드에 쓴다.
_WB_NAME_REVERSE_ALIASES = {}  # pid -> { 실제 wb.Name: 등록명(사용자 파일명) }


def _stash_workbook_name_alias(app, intended_name, actual_name):
    try:
        intended = Path(str(intended_name or "")).name
        actual = str(actual_name or "")
        if not intended or not actual:
            return
        intended_keys = _workbook_name_lookup_keys(intended)
        actual_keys = _workbook_name_lookup_keys(actual)
        if intended_keys & actual_keys:
            return  # 변환 없음(이름 동일) → 별칭 불필요(일반 .xlsx 회귀 0)
        pid = _excel_process_id(app)
        if not pid:
            return
        aliases = _WB_NAME_ALIASES.setdefault(int(pid), {})
        for key in intended_keys:
            aliases.setdefault(key, set()).add(actual)
        _WB_NAME_REVERSE_ALIASES.setdefault(int(pid), {})[actual] = intended
    except Exception:
        pass


def _clear_workbook_name_aliases(app):
    try:
        pid = _excel_process_id(app)
        if pid:
            _WB_NAME_ALIASES.pop(int(pid), None)
            _WB_NAME_REVERSE_ALIASES.pop(int(pid), None)
    except Exception:
        pass


def _user_facing_workbook_name_for_live(app, live_name):
    """라이브 wb.Name → 사용자 파일명(코드에 그대로 저장해도 되는 이름).

    내부 작업본명(excel_open_<hash>.<ext>, live_reset_<hash>_…, <hash>_원본명.xlsx)이면
    ① 열 때 저장한 등록명→실제명 별칭의 역방향으로 원본명을 찾고,
    ② <hash>_원본명 형태면 생성 접두어를 벗긴다. 못 풀면 그대로 반환(현행 유지).
    재생 시엔 반대로 _alias_open_workbook_name 이 사용자명→실제명을 풀므로 왕복이 성립한다."""
    name = str(live_name or "")
    if not name:
        return name
    stem = str(Path(name).stem)
    internal = bool(
        re.match(r"^(?:excel_open_|live_reset_|prestep_)[0-9a-f]{8,}", stem, re.I)
        or re.match(r"^[0-9a-f]{12,}([_-]|$)", stem)
    )
    if not internal:
        return name
    try:
        pid = _excel_process_id(app)
        rev = _WB_NAME_REVERSE_ALIASES.get(int(pid)) if pid else None
        if rev and name in rev:
            return rev[name]
    except Exception:
        pass
    stripped = _GENERATED_WORKBOOK_PREFIX_RE.sub("", name)
    if stripped and stripped != name:
        return stripped
    return name


def _alias_open_workbook_name(app, requested_name):
    # 1) 기존 정확/정규화 매칭 우선(URL/공백 등 기존 동작 보존).
    resolved = _resolve_open_workbook_name(app, requested_name)
    requested = str(requested_name or "")
    if resolved and resolved != requested:
        return resolved
    # 2) 변환 리네임 별칭으로 해결(.xls→excel_open_<uuid>.html 등).
    # [리뷰#6] 별칭 set 은 같은 위장파일을 재오픈할 때마다 새 실제명(uuid 변동)이 쌓여 len>1 이 되면
    # 이전엔 해석이 멈췄다. '현재 실제로 열려 있는' actual 만 남기고 그게 정확히 1개일 때만 사용한다
    # (닫힌 옛 actual 은 자동 배제 → 재오픈 누적에 강건. 서로 다른 파일이 같은 등록명으로 둘 다 열린
    # 진짜 모호 케이스에서만 보류).
    try:
        pid = _excel_process_id(app)
        aliases = _WB_NAME_ALIASES.get(int(pid)) if pid else None
        actuals = set()
        if aliases:
            for key in _workbook_name_lookup_keys(Path(requested).name):
                actuals.update(aliases.get(key, set()))
        if actuals:
            open_names = set()
            try:
                for w in app.Workbooks:
                    try:
                        open_names.add(str(w.Name))
                    except Exception:
                        pass
            except Exception:
                open_names = set()
            open_actuals = [a for a in actuals if a in open_names]
            if len(open_actuals) == 1:
                return open_actuals[0]
    except Exception:
        pass
    return resolved if resolved else requested


def _alias_ephemeral_excel_open_sheet_name(app, requested_name):
    requested = str(requested_name or "").strip()
    # Workbook names produced by compatible-open conversion can also start with
    # excel_open_<hash> and have an extension (.xlsb/.html/.xlsx). Do not let the
    # sheet alias pass rewrite those workbook literals into a sheet name.
    if Path(requested).suffix:
        return requested
    if not _is_ephemeral_excel_open_sheet_name(requested):
        return requested
    candidates = []
    try:
        for wb in app.Workbooks:
            try:
                names = _excel_collection_names(wb.Worksheets)
            except Exception:
                names = []
            for name in names:
                if _is_ephemeral_excel_open_sheet_name(name):
                    candidates.append(str(name))
    except Exception:
        return requested
    unique = []
    seen = set()
    for name in candidates:
        if name not in seen:
            unique.append(name)
            seen.add(name)
    if requested in seen:
        return requested
    return unique[0] if len(unique) == 1 else requested


def _strip_empty_vba_loops(code):
    """본문이 비어 있는 `For Each <var> In <expr> … Next` 루프를 제거한다.
    LLM 이 워크북 찾기 보일러플레이트에 `Dim sh As Worksheet … For Each sh In Application.Workbooks / Next sh`
    같은 '빈 + 타입 안 맞는' 루프를 끼워넣는 경우가 있는데, Workbooks(=Workbook 컬렉션)를 Worksheet 변수로
    순회해 런타임 13(형식 불일치)으로 첫 실행이 실패 → 느린 LLM 에러복구 재생성을 유발했다. 빈 루프는 아무
    동작도 안 하므로 제거해도 무해하고, 이 헛생성으로 인한 실패-재생성 사이클을 없앤다."""
    text = str(code or "")
    if not re.search(r"For\s+Each", text, re.I):
        return text
    # For Each 한 줄 → (빈 줄들) → Next 한 줄 (사이에 실제 문장이 없는 경우만) 제거.
    return re.sub(
        r"(?im)^[ \t]*For\s+Each\s+\w+\s+In\s+[^\r\n]+\r?\n(?:[ \t]*\r?\n)*[ \t]*Next\b[^\r\n]*\r?\n?",
        "",
        text,
    )


def _normalize_vba_workbook_literals(app, code):
    """Patch workbook filename string literals to the actual open workbook name.

    This keeps imported/old skills runnable when the saved code contains
    ``...LGU .xlsx`` but the user's actual file name is ``...LGU%20.xlsx``.
    Only .xls* string literals that resolve to one currently open workbook are
    changed.
    """
    text = str(code or "")

    def repl_workbook(match):
        quote_ch = match.group(1)
        literal = match.group(2)
        actual = _alias_open_workbook_name(app, literal)
        if actual == literal:
            return match.group(0)
        escaped = actual.replace(quote_ch, quote_ch + quote_ch)
        return quote_ch + escaped + quote_ch

    # [엔진 비대칭 수정] 게이트와 정규식이 .xls 계열만 봐서, .csv/.tsv/.html 로 등록된 파일은
    # 별칭(_WB_NAME_ALIASES)이 멀쩡히 있는데도 VBA 리터럴이 치환되지 않았다.
    # 도달 경로: 탭 구분인데 이름이 .csv 면 sniff 가 .tsv 를 반환해 excel_open_<uuid>.tsv 로 리네임
    # 오픈된다 → VBA Workbooks("정산.csv") 는 '첨자가 범위를 벗어났습니다'로 죽는데, 같은 파일이
    # Python ctx.book("정산.csv") 에선 별칭을 타서 성공하는 엔진별 갈림이었다.
    # (별칭이 없으면 repl_workbook 이 원문을 그대로 돌려주므로 확장해도 안전하다.)
    if re.search(r"\.(?:xls[xmb]?|csv|tsv|html?)\b", text, re.I):
        text = re.sub(r'(["\'])([^"\']+\.(?:xls[xmb]?|csv|tsv|html?))\1', repl_workbook, text, flags=re.I)

    if "excel_open_" not in text.lower():
        return text

    def repl_ephemeral_sheet(match):
        quote_ch = match.group(1)
        literal = match.group(2)
        actual = _alias_ephemeral_excel_open_sheet_name(app, literal)
        if actual == literal:
            return match.group(0)
        escaped = actual.replace(quote_ch, quote_ch + quote_ch)
        return quote_ch + escaped + quote_ch

    return re.sub(r'(["\'])(excel_open_[0-9a-f]{8,})\1', repl_ephemeral_sheet, text, flags=re.I)


def _suppress_vba_debug_windows(pid=None):
    """VBE/디버그 다이얼로그가 떠도 사용자에게 보이지 않도록 즉시 닫거나 숨긴다."""
    wg = globals().get("win32gui")
    wp = globals().get("win32process")
    if wg is None or wp is None:
        return
    target_pid = int(pid or 0)
    wm_close = getattr(globals().get("win32con"), "WM_CLOSE", 0x0010)
    sw_hide = getattr(globals().get("win32con"), "SW_HIDE", 0)

    def visit(hwnd, _):
        try:
            _tid, window_pid = wp.GetWindowThreadProcessId(hwnd)
        except Exception:
            return True
        if target_pid and int(window_pid or 0) != target_pid:
            return True
        try:
            title = wg.GetWindowText(hwnd) or ""
            cls = wg.GetClassName(hwnd) or ""
        except Exception:
            return True
        text = (title + " " + cls).lower()
        if "visual basic" not in text and "wndclass_desked" not in text:
            return True
        try:
            if cls == "#32770":
                wg.PostMessage(hwnd, wm_close, 0, 0)
            else:
                wg.ShowWindow(hwnd, sw_hide)
        except Exception:
            pass
        return True

    try:
        wg.EnumWindows(visit, None)
    except Exception:
        pass


def _start_vba_debug_suppressor(pid=None):
    stop = threading.Event()

    def worker():
        while not stop.is_set():
            _suppress_vba_debug_windows(pid)
            stop.wait(0.05)

    thread = threading.Thread(target=worker, name="b2b-vba-debug-suppressor", daemon=True)
    thread.start()
    return stop


def _hide_vba_editor(app):
    """VBE/디버거 창이 사용자 화면으로 올라오지 않게 숨긴다."""
    try:
        vbe = app.VBE
    except Exception:
        return
    try:
        vbe.MainWindow.Visible = False
    except Exception:
        pass
    try:
        count = int(vbe.Windows.Count)
    except Exception:
        count = 0
    for idx in range(1, count + 1):
        try:
            vbe.Windows(idx).Visible = False
        except Exception:
            pass


def _vba_workbook_name(wb):
    try:
        return str(wb.Name or "")
    except Exception:
        return ""


def _vba_string_literal(value):
    return '"' + str(value or "").replace('"', '""') + '"'


def _rewrite_thisworkbook_for_runner_host(code, context_wb):
    """임시 .xlsm 러너에서 실행할 때 ThisWorkbook 은 러너 자신을 가리킨다.
    생성 코드가 ThisWorkbook.Worksheets(...) 를 쓰면 대상 파일이 아니라 러너 파일에 써져
    성공처럼 보이는 no-op 이 된다. 러너 경로에서는 대상 워크북 명시 참조로 바꾼다."""
    text = str(code or "")
    if not re.search(r"\bThisWorkbook\b", text, re.I):
        return text
    return re.sub(
        r"\bThisWorkbook\b",
        "Workbooks(%s)" % _vba_string_literal(_vba_workbook_name(context_wb)),
        text,
        flags=re.I,
    )


def _vba_should_use_runner_host(wb):
    """Return True when the target workbook is not a reliable VBA host."""
    name = _vba_workbook_name(wb).lower()
    if name.endswith(".csv"):
        return True
    try:
        file_format = int(wb.FileFormat)
        # 6=CSV cannot host a VBA project. Normal .xlsx workbooks can still host
        # a temporary in-memory module through VBProject during automation, and
        # that path matches the single-apply behavior more reliably than a temp
        # .xlsm runner in hosted WebView/native full-run contexts.
        if file_format in (6,):
            return True
    except Exception:
        pass
    return False


def _is_vba_macro_run_blocked_error(err):
    text = str(err or "").lower()
    return any(needle in text for needle in [
        "매크로를 실행할 수 없습니다",
        "모든 매크로를 사용하지 못할 수 있습니다",
        "cannot run the macro",
        "macro may not be available",
        "macros may be disabled",
    ])


def _create_vba_runner_workbook(app, context_wb):
    """Create a local temporary .xlsm workbook used only to host injected VBA."""
    # 러너는 '신뢰 위치'로 등록된 폴더 안에 만든다 → 그 매크로는 Trust Center 설정/타이밍과 무관하게 실행됨.
    try:
        _runner_base = str(_b2b_runner_trusted_dir())
    except Exception:
        _runner_base = None
    temp_dir = Path(tempfile.mkdtemp(prefix="b2b_vba_runner_", dir=_runner_base) if _runner_base
                    else tempfile.mkdtemp(prefix="b2b_vba_runner_"))
    temp_path = temp_dir / ("b2b_vba_runner_%s.xlsm" % uuid.uuid4().hex[:8])
    runner = None
    prev_display_alerts = None
    try:
        try:
            prev_display_alerts = app.DisplayAlerts
            app.DisplayAlerts = False
        except Exception:
            pass
        runner = app.Workbooks.Add()
        try:
            runner.SaveAs(str(temp_path), FileFormat=52)  # xlOpenXMLWorkbookMacroEnabled
        except Exception:
            # Unsaved BookN can still host temporary injected modules in most installs.
            pass
        try:
            # 앱 전체실행 컨텍스트에서는 숨겨진 러너 통합문서를 대상으로
            # Application.Run 이 "매크로를 실행할 수 없습니다"로 거부되는 사례가 있었다.
            # 창은 visible 로 두되 화면 밖 작은 normal window 로 치워 사용자 화면에는 거의 노출하지 않는다.
            win = runner.Windows(1)
            win.Visible = True
            try:
                win.WindowState = -4143  # xlNormal
            except Exception:
                pass
            try:
                win.Left = -32000
                win.Top = -32000
                win.Width = 120
                win.Height = 80
            except Exception:
                pass
        except Exception:
            pass
        try:
            context_wb.Activate()
        except Exception:
            pass
        return runner, temp_dir
    except Exception:
        try:
            if runner is not None:
                runner.Close(SaveChanges=False)
        except Exception:
            pass
        shutil.rmtree(temp_dir, ignore_errors=True)
        raise
    finally:
        try:
            if prev_display_alerts is not None:
                app.DisplayAlerts = prev_display_alerts
        except Exception:
            pass


def _close_vba_runner_workbook(app, runner_wb, temp_dir):
    prev_display_alerts = None
    try:
        try:
            prev_display_alerts = app.DisplayAlerts
            app.DisplayAlerts = False
        except Exception:
            pass
        if runner_wb is not None:
            try:
                runner_wb.Close(SaveChanges=False)
            except Exception:
                pass
    finally:
        try:
            if prev_display_alerts is not None:
                app.DisplayAlerts = prev_display_alerts
        except Exception:
            pass
        if temp_dir:
            shutil.rmtree(temp_dir, ignore_errors=True)


def _diag_vba_log_line(msg):
    """[임시 진단] VBA 실행 결과 한 줄을 vba_runner_fail.log 에 남긴다(성공/런타임에러/실행예외 구분)."""
    try:
        import datetime as _dt
        from pathlib import Path as _P
        with open(_P(__file__).resolve().parent / "vba_runner_fail.log", "a", encoding="utf-8") as f:
            f.write("[%s] %s\n" % (_dt.datetime.now().isoformat(), msg))
    except Exception:
        pass


def _vba_trace_path():
    # [트레이스 경로] 프로즌/개발/포터블/단일exe/실행위치와 무관하게 항상 고정된 찾기 쉬운 위치
    # (%LOCALAPPDATA%\B2B_logs)에 쓴다. (프로즌에서 __file__ 이 임시폴더라 repo 옆을 못 쓰는 문제도 해결)
    return b2b_logs_dir() / "vba_pipeline_trace.jsonl"


def _trace_text(value, limit=500):
    text = str(value or "")
    text = text.replace("\r\n", "\n").replace("\r", "\n")
    if len(text) > limit:
        return text[:limit] + "...<truncated %d chars>" % (len(text) - limit)
    return text


def _trace_hash(value):
    return hashlib.sha256(str(value or "").encode("utf-8", errors="replace")).hexdigest()[:16]


def _trace_workbook_info(wb):
    info = {}
    if wb is None:
        return info
    for key in ("Name", "FullName", "Path"):
        try:
            info[key] = str(getattr(wb, key) or "")
        except Exception as err:
            info[key] = "<err %s>" % err
    try:
        info["Worksheets"] = _excel_collection_names(wb.Worksheets)
    except Exception:
        pass
    return info


def _vba_trace(event, **fields):
    """Structured VBA/pipeline trace for field failures.

    File: <repo>/vba_pipeline_trace.jsonl. Each line is a compact JSON event so
    we can compare single apply vs full-run without guessing from UI text.
    """
    try:
        payload = {
            "ts": datetime.datetime.now().isoformat(timespec="milliseconds"),
            "pid": os.getpid(),
            "event": event,
        }
        payload.update(fields)
        with _vba_trace_path().open("a", encoding="utf-8") as f:
            f.write(json.dumps(payload, ensure_ascii=False, default=str) + "\n")
    except Exception:
        pass


def _reset_trace_logs():
    """프로그램 시작 시 이전 실행의 트레이스 로그를 비운다(누적 방지). 저장 위치는 %LOCALAPPDATA%\\B2B_logs."""
    for p in (_perf_trace_path(), _vba_trace_path()):
        try:
            p.parent.mkdir(parents=True, exist_ok=True)
            with open(p, "w", encoding="utf-8"):
                pass   # 내용만 비운다(파일은 남겨 append 경로가 그대로 동작)
        except Exception:
            pass
    # [로그 누적 방지] VBA 러너 실패 진단 로그도 시작 시 비운다 — append 만 해 재시작을 넘어
    # 무한히 컸다(작성 경로와 동일한 __file__ 기준 경로라 프로즌/개발 모두 일치).
    try:
        p = Path(__file__).resolve().parent / "vba_runner_fail.log"
        if p.exists():
            with open(p, "w", encoding="utf-8"):
                pass
    except Exception:
        pass


def _diag_prerun_window_state(app, context_wb):
    """[임시 진단] VBA Application.Run 직전, Excel 앱 프레임 + 대상 워크북 창의
    부모/owner/스타일/가시성을 기록한다. 단일적용 vs 전체실행의 창 상태 차이를 비교하기 위함."""
    try:
        import win32gui as _wg, win32con as _wc, datetime as _dt
        out = ["---- PRE-RUN window state %s ----" % _dt.datetime.now().isoformat()]
        def _desc(tag, h):
            try:
                h = int(h)
                par = _wg.GetParent(h)
                own = _wg.GetWindowLong(h, getattr(_wc, "GWL_HWNDPARENT", -8))
                stl = _wg.GetWindowLong(h, _wc.GWL_STYLE)
                vis = _wg.IsWindowVisible(h)
                out.append("%s hwnd=%s parent=%s owner=%s WS_CHILD=%s visible=%s"
                           % (tag, h, par, own, bool(stl & _wc.WS_CHILD), vis))
            except Exception as e:
                out.append("%s <err %s>" % (tag, e))
        try: _desc("app.Hwnd", app.Hwnd)
        except Exception as e: out.append("app.Hwnd <err %s>" % e)
        try: _desc("ctxwb.Win", context_wb.Windows(1).Hwnd)
        except Exception as e: out.append("ctxwb.Win <err %s>" % e)
        try: out.append("app.Visible=%s ScreenUpdating=%s" % (app.Visible, app.ScreenUpdating))
        except Exception: pass
        from pathlib import Path as _P
        with open(_P(__file__).resolve().parent / "vba_runner_fail.log", "a", encoding="utf-8") as f:
            f.write("\n".join(out) + "\n")
    except Exception:
        pass


def _diag_vba_run_failure(app, host_wb, vbproj, module, module_name, safe_code, err):
    """[임시 진단] 러너 매크로 실행 실패 원인 포착: 컴파일에러 vs 매크로차단 vs 기타.
    log = <repo>/vba_runner_fail.log. 실제 흐름을 절대 깨지 않도록 전부 방어한다."""
    import datetime as _dt
    lines = ["==== VBA RUN FAIL %s ====" % _dt.datetime.now().isoformat()]
    def _s(label, fn):
        try:
            lines.append("%s: %r" % (label, fn()))
        except Exception as e:
            lines.append("%s: <err %s>" % (label, e))
    _s("err", lambda: str(err))
    _s("host_wb.FullName", lambda: str(host_wb.FullName))
    _s("host_wb.Name", lambda: str(host_wb.Name))
    _s("module_name", lambda: module_name)
    _s("app.AutomationSecurity", lambda: app.AutomationSecurity)
    _s("app.Interactive", lambda: app.Interactive)
    _s("app.EnableEvents", lambda: app.EnableEvents)
    _s("app.Visible", lambda: app.Visible)
    _s("app.Version/Build", lambda: "%s/%s" % (app.Version, app.Build))
    _s("open_workbooks", lambda: [str(w.Name) for w in app.Workbooks])
    try:
        cm = module.CodeModule
        injected = cm.Lines(1, cm.CountOfLines)
    except Exception as e:
        injected = "<could not read module: %s>" % e
    # 판별 probe: 같은 워크북에 트리비얼 함수를 넣고 실행 가능한지.
    try:
        pm = vbproj.VBComponents.Add(1)
        pname = pm.Name
        pm.CodeModule.AddFromString("Public Function B2B_Probe999() As Long\r\n B2B_Probe999 = 123\r\nEnd Function\r\n")
        try:
            val = app.Run("'%s'!%s.B2B_Probe999" % (str(host_wb.Name).replace("'", "''"), pname))
            lines.append("DISCRIMINATOR: PROBE_OK(=%r) → 매크로 실행 가능 → 원래 모듈 컴파일에러 의심" % val)
        except Exception as e2:
            lines.append("DISCRIMINATOR: PROBE_FAIL(%s) → 이 워크북/앱에서 매크로 실행 자체가 차단" % e2)
        try:
            vbproj.VBComponents.Remove(pm)
        except Exception:
            pass
    except Exception as e:
        lines.append("DISCRIMINATOR: <probe setup err %s>" % e)
    # 창 상태(임베드 여부) 포착
    try:
        import win32gui as _wg, win32con as _wc
        _h = int(app.Hwnd)
        _s("excel_hwnd", lambda: _h)
        _s("GetParent(excel)", lambda: _wg.GetParent(_h))
        _s("owner(GWL_HWNDPARENT)", lambda: _wg.GetWindowLong(_h, getattr(_wc, "GWL_HWNDPARENT", -8)))
        _s("IsWindowVisible", lambda: _wg.IsWindowVisible(_h))
        _s("style&WS_CHILD", lambda: bool(_wg.GetWindowLong(_h, _wc.GWL_STYLE) & _wc.WS_CHILD))
    except Exception as e:
        lines.append("window-state: <err %s>" % e)
    # 결정적: 완전히 새(비임베드) Excel 인스턴스에서 트리비얼 매크로가 도는가?
    try:
        import win32com.client as _w, uuid as _uuid, tempfile as _tf, os as _os
        fa = _w.DispatchEx("Excel.Application")
        _track_spawned_excel_app(fa)
        fpid = None
        fdir = None
        try:
            fa.Visible = False
            fa.DisplayAlerts = False
            try: fpid = _excel_process_id(fa)
            except Exception: fpid = None
            rb = fa.Workbooks.Add()
            fdir = _tf.mkdtemp(prefix="b2b_freshprobe_")
            ftmp = _os.path.join(fdir, "fp_%s.xlsm" % _uuid.uuid4().hex[:6])
            try: rb.SaveAs(ftmp, FileFormat=52)
            except Exception: pass
            fmod = rb.VBProject.VBComponents.Add(1)
            fmn = fmod.Name
            fmod.CodeModule.AddFromString("Public Function B2B_FP() As Long\r\n B2B_FP = 7\r\nEnd Function\r\n")
            try:
                fv = fa.Run("'%s'!%s.B2B_FP" % (str(rb.Name).replace("'", "''"), fmn))
                lines.append("FRESH_INSTANCE_PROBE: OK(=%r) → 새 인스턴스는 매크로 실행 가능 → 임베드 라이브 인스턴스만 문제" % fv)
            except Exception as e2:
                lines.append("FRESH_INSTANCE_PROBE: FAIL(%s) → 새 인스턴스도 안 됨 → 환경/애드인/프로세스 컨텍스트 문제" % e2)
            try: rb.Close(SaveChanges=False)
            except Exception: pass
        finally:
            try: fa.Quit()
            except Exception: pass
            try:
                if fpid:
                    _force_kill_pid(fpid)
                    SPAWNED_EXCEL_PIDS.discard(int(fpid))
            except Exception: pass
            try:
                if fdir:
                    shutil.rmtree(fdir, ignore_errors=True)
            except Exception: pass
    except Exception as e:
        lines.append("FRESH_INSTANCE_PROBE: <setup err %s>" % e)
    lines.append("---- injected code ----")
    lines.append(injected)
    lines.append("==== END ====\n")
    try:
        from pathlib import Path as _P
        logp = _P(__file__).resolve().parent / "vba_runner_fail.log"
        with open(logp, "a", encoding="utf-8") as f:
            f.write("\n".join(lines) + "\n")
    except Exception:
        pass


def _inject_and_run_vba_in_host(app, host_wb, context_wb, code, entry):
    """Inject/run VBA in host_wb while keeping context_wb as ActiveWorkbook."""
    module = None
    prev_display_alerts = None
    prev_enable_events = None
    prev_enable_cancel_key = None
    prev_auto_security = None
    try:
        try:
            vbproj = host_wb.VBProject
        except Exception as err:
            raise RuntimeError(
                "VBA 프로젝트에 접근할 수 없습니다. Excel 옵션 > 보안 센터 > 매크로 설정에서 "
                "'VBA 프로젝트 개체 모델에 대한 액세스 신뢰'를 켠 뒤 파일을 다시 여세요. (" + str(err) + ")"
            )
        module = vbproj.VBComponents.Add(1)  # 1 = vbext_ct_StdModule
        module_name = module.Name
        if host_wb is not context_wb:
            code = _rewrite_thisworkbook_for_runner_host(code, context_wb)
        safe_code, runner_name, err_num_name, err_desc_name = _wrap_vba_skill_code(code, entry)
        _vba_trace(
            "vba.inject.host.prepare",
            host=_trace_workbook_info(host_wb),
            context=_trace_workbook_info(context_wb),
            moduleName=module_name,
            entry=entry,
            codeLen=len(str(code)),
            codeHash=_trace_hash(code),
            safeCodeLen=len(str(safe_code)),
            safeCodeHash=_trace_hash(safe_code),
            safeCodeHead=_trace_text(safe_code, 420),
        )
        context_name = _vba_workbook_name(context_wb)
        if context_name:
            activate_context = (
                "    On Error Resume Next\n"
                "    Workbooks(%s).Activate\n"
                "    On Error GoTo B2B_Err\n" % _vba_string_literal(context_name)
            )
            safe_code = safe_code.replace(
                "    B2B_LastErrDescription = vbNullString\n",
                "    B2B_LastErrDescription = vbNullString\n" + activate_context,
                1,
            )
        module.CodeModule.AddFromString(safe_code)
        try:
            prev_display_alerts = app.DisplayAlerts
            app.DisplayAlerts = False
        except Exception:
            pass
        try:
            prev_enable_events = app.EnableEvents
            app.EnableEvents = False
        except Exception:
            pass
        try:
            prev_enable_cancel_key = app.EnableCancelKey
            app.EnableCancelKey = 0  # xlDisabled
        except Exception:
            pass
        # [핵심 수정] 매크로 실행 직전 AutomationSecurity 를 Low(1)로 낮춘다.
        # excel_workbooks_open 이 파일을 열 때 AutomationSecurity=ForceDisable(3)로 설정하는데,
        # 일부 환경(기업/특정 Office 빌드)에서는 ForceDisable 이 인스턴스의 '모든 매크로'를 비활성화해
        # 주입한 러너 매크로의 Application.Run 이 "매크로를 실행할 수 없습니다"(-2146827284)로 실패한다.
        # 우리가 만든 임시 러너 + 사용자 자신의 파일이므로 Low 로 낮추는 것은 안전하다. finally 에서 원복.
        try:
            prev_auto_security = app.AutomationSecurity
            app.AutomationSecurity = 1  # msoAutomationSecurityLow
        except Exception:
            pass
        try:
            try:
                context_wb.Activate()
            except Exception:
                pass
            _diag_prerun_window_state(app, context_wb)  # [임시 진단] run 직전 창 상태 기록
            _vba_trace(
                "vba.macro.run.start",
                host=_trace_workbook_info(host_wb),
                context=_trace_workbook_info(context_wb),
                moduleName=module_name,
                macro=runner_name,
                refs=_vba_macro_refs(host_wb, module_name, runner_name),
            )
            _run_vba_macro_any_ref(app, host_wb, module_name, runner_name)
            err_number = 0
            err_description = ""
            try:
                err_number = int(_run_vba_macro_any_ref(app, host_wb, module_name, err_num_name) or 0)
            except Exception:
                err_number = 0
            try:
                err_description = str(_run_vba_macro_any_ref(app, host_wb, module_name, err_desc_name) or "")
            except Exception:
                err_description = ""
            if err_number:
                _diag_vba_log_line("VBA-RUNTIME-ERR num=%s desc=%r" % (err_number, err_description))
                _vba_trace("vba.macro.runtime_error", errNumber=err_number, errDescription=err_description)
                raise RuntimeError("VBA 실행 실패: %s" % (err_description or ("오류 번호 %s" % err_number)))
            _diag_vba_log_line("VBA-OK")
            _vba_trace("vba.macro.run.ok", host=_trace_workbook_info(host_wb), moduleName=module_name, macro=runner_name)
        except Exception as err:
            if str(err).startswith("VBA 실행 실패:"):
                raise
            try:
                _hp = str(host_wb.FullName)
            except Exception:
                _hp = "<?>"
            try:
                _as = app.AutomationSecurity
            except Exception:
                _as = "<?>"
            _diag_vba_log_line("VBA-RUN-EXC host=%r autosec=%s err=%r" % (_hp, _as, str(err)[:200]))
            _vba_trace(
                "vba.macro.run.error",
                host=_trace_workbook_info(host_wb),
                context=_trace_workbook_info(context_wb),
                moduleName=module_name,
                macro=runner_name,
                automationSecurity=_as,
                error=str(err),
            )
            try:
                _diag_vba_run_failure(app, host_wb, vbproj, module, module_name, safe_code, err)
            except Exception:
                pass
            raise RuntimeError("VBA 실행 실패: %s" % err)
    finally:
        try:
            if prev_auto_security is not None:
                app.AutomationSecurity = prev_auto_security
        except Exception:
            pass
        try:
            if prev_enable_cancel_key is not None:
                app.EnableCancelKey = prev_enable_cancel_key
        except Exception:
            pass
        try:
            if prev_enable_events is not None:
                app.EnableEvents = prev_enable_events
        except Exception:
            pass
        try:
            if prev_display_alerts is not None:
                app.DisplayAlerts = prev_display_alerts
        except Exception:
            pass
        if module is not None:
            try:
                vbproj.VBComponents.Remove(module)
            except Exception:
                pass


def _inject_and_run_vba(app, wb, code, entry):
    """VBA 모듈을 임시 추가해 entry Sub를 실행하고, 끝나면 제거한다.

    일반 .xlsx도 자동화 중에는 임시 in-memory VBA 모듈을 직접 주입해 실행할 수 있다.
    따라서 단일 적용/전체실행 모두 우선 대상 워크북을 직접 호스트로 쓴다. CSV처럼
    VBA 프로젝트 호스트가 될 수 없거나 직접 실행이 실제로 차단된 경우에만 임시
    .xlsm 러너를 최후 우회 경로로 사용한다.
    """
    code = code or ""
    if not code.strip():
        return
    original_code = code
    code = _extract_vba_source_for_injection(code, entry)
    code = _strip_empty_vba_loops(code)  # 빈 For Each(워크북 헛순회 등) 제거 → 형식불일치 실패-재생성 사이클 차단
    code = _normalize_vba_llm_comment_slips(code)  # LLM 의 // 주석 → ' 변환(컴파일 오류 모달 영구블록 방지)
    _vba_trace(
        "vba.code.normalized",
        workbook=_trace_workbook_info(wb),
        entry=entry,
        beforeLen=len(str(original_code)),
        beforeHash=_trace_hash(original_code),
        beforeHead=_trace_text(original_code, 420),
        afterLen=len(str(code)),
        afterHash=_trace_hash(code),
        afterHead=_trace_text(code, 420),
        changed=(str(original_code) != str(code)),
    )
    code = _normalize_vba_workbook_literals(app, code)
    _validate_vba_source_before_inject(code)
    excel_pid = _excel_process_id(app)
    _vba_trace(
        "vba.inject.start",
        workbook=_trace_workbook_info(wb),
        excelPid=excel_pid,
        codeLen=len(str(code)),
        codeHash=_trace_hash(code),
        codeHead=_trace_text(code, 420),
    )
    suppressor = _start_vba_debug_suppressor(excel_pid)
    runner_wb = None
    runner_temp = None
    try:
        _disable_vba_break_on_all_errors()
        _hide_vba_editor(app)
        if _vba_should_use_runner_host(wb):
            _run_vba_via_runner_with_retry(app, wb, code, entry)
            return
        try:
            _inject_and_run_vba_in_host(app, wb, wb, code, entry)
            return
        except Exception as err:
            if not _is_vba_macro_run_blocked_error(err):
                raise
            _run_vba_via_runner_with_retry(app, wb, code, entry)
            return
    finally:
        try:
            _hide_vba_editor(app)
            _suppress_vba_debug_windows(excel_pid)
        finally:
            suppressor.set()


def _run_vba_via_runner_with_retry(app, wb, code, entry, attempts=2):
    """임시 .xlsm 러너에서 VBA 를 실행한다. '매크로를 실행할 수 없습니다'(-2146827284)는 일부 환경에서
    Excel 이 파일 오픈/주입 직후 잠깐 '준비 안 됨' 상태일 때 간헐적으로 나는데(전체실행이 될 때/안 될 때가
    갈리던 현상), 잠깐 대기 후 '러너를 새로 만들어' 재시도하면 대개 통과한다. 매크로 차단이 아닌 실제
    오류(스킬 버그 등)는 즉시 전파한다."""
    last_err = None
    for i in range(max(1, attempts)):
        runner_wb = None
        runner_temp = None
        try:
            runner_wb, runner_temp = _create_vba_runner_workbook(app, wb)
            _inject_and_run_vba_in_host(app, runner_wb, wb, code, entry)
            return  # 성공
        except Exception as err:
            last_err = err
            if not _is_vba_macro_run_blocked_error(err):
                raise  # 실제 스킬/실행 오류 → 재시도 무의미, 그대로 전파
        finally:
            if runner_wb is not None:
                _close_vba_runner_workbook(app, runner_wb, runner_temp)
        if i < attempts - 1:
            # 매크로 차단(간헐) → 대기 후 새 러너로 재시도. Excel 이 준비되도록 잠깐 양보.
            try:
                _diag_vba_log_line("RUNNER-RETRY %d/%d (macro-blocked, waiting)" % (i + 1, attempts))
            except Exception:
                pass
            try:
                time.sleep(0.7 * (i + 1))
            except Exception:
                pass
            try:
                _disable_vba_break_on_all_errors()
                _hide_vba_editor(app)
            except Exception:
                pass
    raise last_err if last_err is not None else RuntimeError("VBA 러너 실행 실패")


def _restore_app_state(app):
    """VBA 실행(성공/실패 무관) 후 Application 전역 상태를 결정적으로 정상화한다.

    생성된 VBA 가 Application.Calculation = xlCalculationManual / ScreenUpdating = False 로
    바꿔놓고 (특히 Err.Raise 로) 중단되면, 그 상태가 워크북 인스턴스에 남아 **이후 모든
    수식 재계산이 멈춘다**(품질평가에서 가장 빈번했던 위험). 그래서 호출자의 복원에만
    의존하지 않고 여기서 항상 Automatic/True 로 강제하고 한 번 재계산한다.
    부수효과가 없는 안전한 정상화이므로 모든 실행 경로의 finally 에서 호출한다."""
    try:
        app.Calculation = -4105  # xlCalculationAutomatic (상수 이름이 환경따라 없을 수 있어 값 사용)
    except Exception:
        pass
    try:
        app.ScreenUpdating = True
    except Exception:
        pass
    try:
        app.EnableEvents = True
    except Exception:
        pass
    try:
        app.CutCopyMode = False
    except Exception:
        pass
    # Manual 로 멈춰 있던 동안 갱신 안 된 수식을 한 번 강제 계산(고착 해소).
    try:
        app.Calculate()
    except Exception:
        pass


def _restore_live_protected_view(app, wb):
    """VBA 실행 후 라이브 보기 상태(편집 차단+선택 허용, 리본/우클릭 숨김, 화면갱신)를 복구."""
    try:
        _hide_vba_editor(app)
    except Exception:
        pass
    try:
        _protect_workbook_for_read_only_mirror(wb, True)
    except Exception:
        pass
    try:
        _configure_excel_grid_window(app, wb)  # 리본 숨김 + 입력키 차단 + 우클릭 차단
    except Exception:
        pass
    try:
        app.ScreenUpdating = True
    except Exception:
        pass
    try:
        app.Calculation = -4105  # xlCalculationAutomatic
    except Exception:
        pass


def _same_excel_workbook(a, b):
    if a is None or b is None:
        return False
    for attr in ("FullName", "Name"):
        try:
            av = str(getattr(a, attr) or "").lower()
            bv = str(getattr(b, attr) or "").lower()
            if av and bv and av == bv:
                return True
        except Exception:
            pass
    return False


def _capture_live_view_state(app, wb, session=None):
    """현재 워크북의 활성 시트/선택 주소를 보존한다.
    동반 워크북 오픈, reset, 창 복구가 ActiveSheet 를 마지막 시트로 바꾸는 경우를 막기 위한 상태다."""
    state = {"sheet": "", "address": ""}
    ws = None
    try:
        ws = wb.Windows(1).ActiveSheet
    except Exception:
        ws = None
    if ws is None:
        try:
            active_ws = app.ActiveSheet
            parent = getattr(active_ws, "Parent", None)
            if _same_excel_workbook(parent, wb):
                ws = active_ws
        except Exception:
            ws = None
    if ws is not None:
        try:
            state["sheet"] = str(ws.Name or "")
        except Exception:
            state["sheet"] = ""
    if not state["sheet"] and session:
        remembered = str(session.get("lastSelectionSheet") or "")
        if remembered:
            try:
                names = _excel_collection_names(wb.Worksheets)
                if remembered in names:
                    state["sheet"] = remembered
            except Exception:
                pass
    if not state["sheet"]:
        try:
            names = _excel_collection_names(wb.Worksheets)
            if names:
                state["sheet"] = names[0]
        except Exception:
            pass
    try:
        sel = app.Selection
        sel_ws = getattr(sel, "Worksheet", None)
        if sel_ws is not None and str(getattr(sel_ws, "Name", "") or "") == state["sheet"]:
            state["address"] = _excel_address(sel).replace("$", "")
    except Exception:
        pass
    if not state["address"]:
        try:
            active_cell = app.ActiveCell
            cell_ws = getattr(active_cell, "Worksheet", None)
            if cell_ws is not None and str(getattr(cell_ws, "Name", "") or "") == state["sheet"]:
                state["address"] = _excel_address(active_cell).replace("$", "")
        except Exception:
            pass
    if not state["address"] and session and str(session.get("lastSelectionSheet") or "") == state["sheet"]:
        state["address"] = str(session.get("lastSelectionAddress") or "")
    return state if state.get("sheet") else None


def _restore_live_view_state(app, wb, state, session=None):
    if not state or not state.get("sheet"):
        return
    sheet = str(state.get("sheet") or "")
    try:
        names = _excel_collection_names(wb.Worksheets)
    except Exception:
        names = []
    match = sheet if sheet in names else None
    if not match:
        normalized = normalize_text(sheet)
        match = next((name for name in names if normalize_text(name) == normalized), None)
    if not match:
        return
    try:
        wb.Activate()
    except Exception:
        pass
    try:
        ws = wb.Worksheets(match)
        ws.Activate()
    except Exception:
        return
    address = str(state.get("address") or "")
    if address:
        try:
            ws.Range(address).Select()
        except Exception:
            pass
    if session is not None:
        session["lastSelectionSheet"] = match
        if address:
            session["lastSelectionAddress"] = address


def _restore_live_window(session, app, wb):
    """리셋(_copy_source_workbook_into_target)으로 offscreen park 된 라이브 창을 owner 모드 방식으로
    다시 보이게+제자리로 되돌린다(plain 위치맞춤 + owner + 워크북 뷰 채움)."""
    if session.get("liveEditable") and LIVE_FRAME_MODE:
        rect = session.get("liveRect") or {}
        presented = _present_live_session_frame(
            session, app, wb,
            rect.get("left"), rect.get("top"), rect.get("width"), rect.get("height"),
            skip_position=not (rect.get("width") and rect.get("height")),
        )
        if presented is not None:
            return
    try:
        app.Visible = True
    except Exception:
        pass
    try:
        wb.Activate()
    except Exception:
        pass
    try:
        app.WindowState = -4143  # xlNormal
    except Exception:
        pass
    _set_excel_window_owner(app, session.get("nativeHostHwnd"))
    # show 전에 워크북 뷰/그리기를 먼저 준비(파킹 상태라 화면엔 안 보임) → 회색 프레임 플래시 방지.
    try:
        _ensure_excel_workbook_view(app, wb, make_visible=True, activate=False, maximize_workbook=False)
    except Exception:
        pass
    rect = session.get("liveRect") or {}
    left, top, width, height = rect.get("left"), rect.get("top"), rect.get("width"), rect.get("height")
    if width and height:
        try:
            _position_excel_window(app, left, top, width, height, show=True)  # plain(프레임 유지)
        except Exception:
            pass
    try:
        _ensure_excel_workbook_view(app, wb, make_visible=True, activate=False, maximize_workbook=False)
    except Exception:
        pass
    _hide_non_target_workbook_windows(app, wb)
    _set_excel_window_owner(app, session.get("nativeHostHwnd"))


def _close_companion_workbooks(session, app):
    """이전에 동반 오픈한 워크북을 닫고 임시본 폴더를 정리한다(다음 스냅샷 전에 호출)."""
    for nm in list(session.get("companionNames") or []):
        try:
            for w in list(app.Workbooks):
                if str(w.Name).lower() == str(nm).lower():
                    try:
                        w.Close(SaveChanges=False)
                    except Exception:
                        pass
                    break
        except Exception:
            pass
    session["companionNames"] = []
    for cdir in session.get("companionTemps") or []:
        try:
            shutil.rmtree(cdir, ignore_errors=True)
        except Exception:
            pass
    session["companionTemps"] = []


def _ensure_companion_workbooks(session, excel_id, app, current_wb):
    """다른 라이브 세션들의 '현재(편집 반영된) 상태'를 스냅샷해서 이 인스턴스에 읽기전용으로 동반 오픈한다.
    원본 업로드 파일이 아니라 라이브 임시 워크북의 최신 상태(SaveCopyAs)를 읽으므로,
    사용자가 입력 파일을 먼저 스킬로 수정한 뒤 그 값을 출력 스킬에서 활용할 수 있다.
    매 실행마다 이전 동반본을 닫고 새로 스냅샷한다(항상 최신). VBA 는 Workbooks("파일명") 으로 교차 접근."""
    # [리뷰②] 신선도 스킵: 다른 세션들이 마지막 스냅샷 이후 변하지 않았으면(rev 동일) 통째로 생략.
    # rev 는 각 세션에서 vba/python/파이프라인 실행이 일어날 때마다 +1 된다(아래 run impl 들).
    # 한계: 사용자가 라이브 Excel 창에 직접 타이핑한 변경은 rev 에 안 잡힌다 — 그 경우를 위해
    # B2B_COMPANION_ALWAYS_SNAPSHOT=1 로 기존(매번 스냅샷) 동작을 강제할 수 있다.
    try:
        if os.environ.get("B2B_COMPANION_ALWAYS_SNAPSHOT") != "1":
            want_revs = {}
            for _oid, _o in list(EXCEL_SESSIONS.items()):
                if _oid == excel_id or not _o.get("liveEditable"):
                    continue
                want_revs[_oid] = int(_o.get("rev") or 0)
            prev_revs = session.get("companionRevs")
            temps_ok = all(Path(t).exists() for t in (session.get("companionTemps") or []))
            if prev_revs == want_revs and temps_ok:
                return
            session["companionRevs"] = want_revs
    except Exception:
        pass
    # [필드 찐막] frame(공유 앱) 모드에서 실제 스냅샷/정리가 없으면 꼬리의 Activate/창숨김/
    # ScreenUpdating 토글을 건너뛴다 — 파일별 '첫 적용'에서 창이 내려갔다 뜨던 출렁임의 원인.
    _companion_changed = bool(session.get("companionNames") or session.get("companionTemps"))
    _close_companion_workbooks(session, app)
    try:
        current_name = str(current_wb.Name).lower()
    except Exception:
        current_name = ""
    opened = {current_name} if current_name else set()
    try:
        for existing_wb in app.Workbooks:
            try:
                nm = str(existing_wb.Name).lower()
                if nm:
                    opened.add(nm)
            except Exception:
                pass
    except Exception:
        pass
    names, temps = [], []
    _screen_off = False
    for other_id, other in list(EXCEL_SESSIONS.items()):
        if other_id == excel_id or not other.get("liveEditable"):
            continue
        try:
            _o_app, o_wb = session_workbook(other)
        except Exception:
            continue
        clean = Path(str(other.get("name") or "")).name
        if not clean:
            try:
                clean = Path(str(o_wb.Name)).name
            except Exception:
                clean = ""
        if not clean or clean.lower() in opened:
            continue
        if _is_live_shared_app(app) and _is_live_shared_app(_o_app):
            # 같은 Excel.Application 안에 이미 열린 라이브 워크북이다.
            # 스냅샷 복제본을 또 열면 프로세스 절감 효과가 사라지고 Workbooks("파일명")도 모호해진다.
            try:
                opened.add(str(o_wb.Name).lower())
            except Exception:
                opened.add(clean.lower())
            continue
        if not _screen_off:
            try:
                app.ScreenUpdating = False
                _screen_off = True
            except Exception:
                pass
        _companion_changed = True
        cdir = BACKEND_DIR / f"companion_{uuid.uuid4().hex}"
        try:
            cdir.mkdir(parents=True, exist_ok=True)
            cpath = cdir / clean
            o_wb.SaveCopyAs(str(cpath))   # 라이브 최신 상태(편집 반영)를 스냅샷
            wb2, _t = excel_workbooks_open(app, cpath, read_only=True, intended_name=clean)
            # 동반 워크북 창은 화면에 안 나오게 확실히 숨긴다(Visible=False + 오프스크린 park).
            _hide_workbook_windows(wb2)
            opened.add(clean.lower())
            names.append(clean)
            temps.append(str(cdir))
        except Exception:
            try:
                shutil.rmtree(cdir, ignore_errors=True)
            except Exception:
                pass
    session["companionNames"] = names
    session["companionTemps"] = temps
    # 동반 워크북을 실제로 열고 닫았을 때만 활성/숨김/화면 갱신을 손댄다(불필요한 창 출렁임 방지).
    if _companion_changed:
        try:
            current_wb.Activate()
        except Exception:
            pass
        _hide_non_target_workbook_windows(app, current_wb)
    if _screen_off:
        try:
            app.ScreenUpdating = True
        except Exception:
            pass


def _run_vba_on_session_impl(excel_id, code, entry=None, restore_window=True):
    """라이브 세션에 떠 있는 실제 워크북에 VBA 매크로를 주입해 즉시 실행한다(저지연 리모콘, 단일 단계 append).
    시트는 UserInterfaceOnly=True 보호라 사용자는 직접 편집 못 해도 VBA/COM 은 수정 가능."""
    entry = (str(entry).strip() if entry else "") or VBA_SKILL_ENTRY
    if not (code or "").strip():
        raise RuntimeError("VBA 코드가 비어 있습니다.")
    _t0 = time.perf_counter()
    timings = {"mode": "vba-single", "steps": 1}
    with EXCEL_LOCK:
        _t = time.perf_counter()
        session = get_excel_session(excel_id)
        session["rev"] = int(session.get("rev") or 0) + 1  # [리뷰②] 동반 스냅샷 신선도 추적
        app, wb = session_workbook(session)
        initial_view = _capture_live_view_state(app, wb, session)
        final_view = initial_view
        timings["sessionMs"] = round((time.perf_counter() - _t) * 1000, 2)
        # 교차 파일 접근: 다른 업로드 파일들을 같은 인스턴스에 동반 오픈(읽기전용, 숨김).
        _t = time.perf_counter()
        _ensure_companion_workbooks(session, excel_id, app, wb)
        # VBA 실행 동안에는 시트 보호를 풀어 둔다(보호로 인한 1004 류 실패 방지). 끝나면 다시 보호.
        _t = time.perf_counter()
        try:
            _protect_workbook_for_read_only_mirror(wb, False)
        except Exception:
            pass
        try:
            _t = time.perf_counter()
            _prepare_vba_macro_run_window_state(session, app, wb)
            _inject_and_run_vba(app, wb, code, entry)
            captured = _capture_live_view_state(app, wb, session)
            if captured:
                final_view = captured
            timings["injectRunMs"] = round((time.perf_counter() - _t) * 1000, 2)
        except PipelineExecutionError:
            raise
        except Exception as err:
            if _is_vba_macro_run_blocked_error(err):
                try:
                    _diag_vba_log_line("VBA-LIVE-BLOCKED -> isolated single-step fallback")
                    _t_fb = time.perf_counter()
                    _run_vba_pipeline_on_session_impl(
                        excel_id,
                        [{
                            "stepIdx": 0,
                            "stepId": None,
                            "description": "",
                            "code": code,
                            "language": "vba",
                        }],
                        reset=False,
                        entry=entry,
                        view_sheet=(initial_view or {}).get("sheet") if initial_view else None,
                    )
                    captured = _capture_live_view_state(app, wb, session)
                    if captured:
                        final_view = captured
                    timings["injectRunMs"] = round((time.perf_counter() - _t) * 1000, 2)
                    timings["isolatedFallbackMs"] = round((time.perf_counter() - _t_fb) * 1000, 2)
                    err = None
                except PipelineExecutionError:
                    raise
                except Exception as fallback_err:
                    err = fallback_err
            if err is None:
                pass
            else:
                # 단일 VBA 적용 실패도 파이프라인과 같은 포맷(원인+프롬프트 가이드+세부)으로 통일.
                _cause, _guide = _pipeline_error_guide(str(err), code)
                raise PipelineExecutionError({
                    "stepIdx": 0,
                    "stepId": None,
                    "description": "",
                    "code": code,
                    "language": "vba",
                    "message": _cause + '\n' + "💡 이렇게 요청해 보세요: " + _guide + '\n' + "(자세히: " + str(err) + ")",
                    "cause": _cause,
                    "promptGuide": _guide,
                    "rawError": str(err),
                    "stack": "",
                }) from err
        finally:
            # 성공/실패 무관: Application 전역 상태(Calculation/ScreenUpdating 등)를 먼저 정상화.
            # 생성 VBA 가 Manual 로 두고 죽어도 재계산 고착이 남지 않게 한다.
            _restore_app_state(app)
            try:
                _restore_live_protected_view(app, wb)
            except Exception:
                pass
            # 단일 적용은 바로 결과를 보여야 하므로 복원한다. 전체실행은 step마다 복원하면
            # 빈 Excel 창이 여러 번 튀므로 클라이언트가 마지막에 한 번만 복원한다.
            if restore_window:
                try:
                    _restore_live_window(session, app, wb)
                except Exception:
                    pass
        # [변경없음 검증 제거] VBA 가 예외 없이 끝났으면 성공으로 본다. 예전엔 실행 전후 워크북 지문을
        # 비교해 '변경 0건'이면 실패 처리했으나, 서식만 바꾸는 작업(천단위 콤마 등)·숨김 해제처럼 값 지문에
        # 안 잡히는 정상 작업까지 막아 더 불편해서 가드를 전부 들어냈다.
        result = {"ok": True, "excelId": excel_id, "entry": entry}
        # [새 시트 @멘션 수정] 단일 VBA 적용도 적용 후 경량 스키마를 실어, VBA 로 만든 새 시트가
        # 클라 시트 캐시·@멘션 목록에 즉시 반영되게 한다(단일 Python/격리 파이프라인은 이미 그렇게 함).
        try:
            result["liveSchema"] = _live_preview_schema(wb)
        except Exception:
            pass
        return result


def _detach_live_excel_window(app):
    """라이브 Excel 창이 WebView 에 임베드(WS_CHILD/owner)돼 있으면 독립 top-level 로 분리한다.
    [근거] FRESH_INSTANCE_PROBE 결과: 비임베드 인스턴스는 VBA 매크로(Application.Run)가 정상 실행되지만,
    임베드된 라이브 인스턴스는 reset 의 창 park 이후 매크로 실행 시 COM 이 RPC 사망한다.
    파이프라인 VBA 실행 동안만 분리해 단일적용과 동일한(매크로 실행 가능한) 창 상태로 만든다.
    반환: 복구용 상태 튜플(또는 임베드가 아니면 None)."""
    try:
        import win32gui, win32con
    except Exception:
        return None
    try:
        hwnd = int(app.Hwnd)
        prev_parent = win32gui.GetParent(hwnd)
        owner_idx = getattr(win32con, "GWL_HWNDPARENT", -8)
        prev_owner = win32gui.GetWindowLong(hwnd, owner_idx)
        prev_style = win32gui.GetWindowLong(hwnd, win32con.GWL_STYLE)
        if not prev_parent and not (prev_style & win32con.WS_CHILD) and not prev_owner:
            return None  # 임베드 아님(단일 DispatchEx 등) — 손대지 않음
        # WS_CHILD 제거 + WS_POPUP + 숨김 유지로 독립 top-level 화(화면 깜빡임 없음).
        new_style = (prev_style & ~(win32con.WS_CHILD | win32con.WS_VISIBLE)) | win32con.WS_POPUP
        try:
            win32gui.SetParent(hwnd, 0)
        except Exception:
            pass
        try:
            win32gui.SetWindowLong(hwnd, win32con.GWL_STYLE, new_style)
        except Exception:
            pass
        try:
            win32gui.SetWindowLong(hwnd, owner_idx, 0)
        except Exception:
            pass
        return (hwnd, prev_parent, prev_owner, prev_style)
    except Exception:
        return None


def _reattach_live_excel_window(state):
    """_detach_live_excel_window 로 분리한 창을 원래 부모/owner/스타일로 되돌린다."""
    if not state:
        return
    try:
        import win32gui, win32con
        hwnd, prev_parent, prev_owner, prev_style = state
        owner_idx = getattr(win32con, "GWL_HWNDPARENT", -8)
        try:
            win32gui.SetWindowLong(hwnd, win32con.GWL_STYLE, prev_style)
        except Exception:
            pass
        if prev_parent:
            try:
                win32gui.SetParent(hwnd, prev_parent)
            except Exception:
                pass
        if prev_owner:
            try:
                win32gui.SetWindowLong(hwnd, owner_idx, prev_owner)
            except Exception:
                pass
    except Exception:
        pass


def _setup_isolated_pipeline_instance(session, excel_id, reset, work):
    """격리 실행용 새 Excel 인스턴스를 띄우고 대상+동반 워크북을 '정확한 이름'으로 연다.
    - 대상: reset 이면 source 원본, 아니면 현재 라이브 상태(SaveCopyAs).
    - 동반(다른 라이브 편집 세션): 현재 라이브 상태(SaveCopyAs) — VBA 의 Workbooks("파일명") 교차참조용.
    반환: (fapp, ftarget, fpid). 호출자가 finally 에서 정리한다."""
    live_app0, live_wb0 = session_workbook(session)
    target_name = Path(str(session.get("name") or "")).name
    if not target_name:
        target_name = Path(str(live_wb0.Name)).name
    _ensure_vbom_access()
    _disable_vba_break_on_all_errors()
    fapp = win32com.client.DispatchEx("Excel.Application")
    _track_spawned_excel_app(fapp)
    fpid = None
    try:
        fpid = _excel_process_id(fapp)
    except Exception:
        fpid = None
    _vba_trace(
        "pipeline.isolated.setup.start",
        excelId=excel_id,
        reset=reset,
        work=str(work),
        targetName=target_name,
        sessionName=session.get("name"),
        sessionPath=session.get("path"),
        sourcePath=session.get("sourcePath"),
        liveWorkbook=_trace_workbook_info(live_wb0),
        isolatedPid=fpid,
    )
    for attr, val in (("Visible", False), ("DisplayAlerts", False), ("EnableEvents", False), ("AskToUpdateLinks", False)):
        try:
            setattr(fapp, attr, val)
        except Exception:
            pass
    # 대상 워크북
    tdir = work / "t"
    tdir.mkdir(parents=True, exist_ok=True)
    tpath = tdir / target_name
    if reset:
        src = session.get("sourcePath") or session.get("path")
        shutil.copy2(Path(src), tpath)
    else:
        live_wb0.SaveCopyAs(str(tpath))
    ftarget, _t = excel_workbooks_open(fapp, str(tpath), read_only=False, intended_name=target_name)
    _vba_trace(
        "pipeline.isolated.target.opened",
        excelId=excel_id,
        reset=reset,
        isolatedPid=fpid,
        targetPath=str(tpath),
        targetWorkbook=_trace_workbook_info(ftarget),
    )
    opened = {target_name.lower()}
    companions = []
    # 동반 워크북(다른 라이브 편집 세션의 현재 상태)
    for oid, other in list(EXCEL_SESSIONS.items()):
        if oid == excel_id or not other.get("liveEditable"):
            continue
        try:
            _oa, o_wb = session_workbook(other)
        except Exception:
            continue
        cname = Path(str(other.get("name") or "")).name
        if not cname:
            try:
                cname = Path(str(o_wb.Name)).name
            except Exception:
                cname = ""
        if not cname or cname.lower() in opened:
            continue
        cdir = work / ("c_" + uuid.uuid4().hex[:6])
        cdir.mkdir(parents=True, exist_ok=True)
        cpath = cdir / cname
        try:
            o_wb.SaveCopyAs(str(cpath))
            cwb, _ct = excel_workbooks_open(fapp, str(cpath), read_only=False, intended_name=cname)
            opened.add(cname.lower())
            # [교차파일 유실 수정] 이 동반본을 '쓰기 대상'으로 변형하는 스텝(입력→출력 .Copy,
            # 출력→입력 매칭쓰기 등)이 있으면, 실행 후 변경된 동반본을 그 라이브 세션으로 되돌려써야 한다.
            # 안 하면 라우팅이 어느 세션을 ftarget 으로 잡았든 '다른 파일에 쓴 결과'가 통째로 버려진다
            # (8↔22 가 판마다 깨지던 비결정 버그의 근본원인). _sync_modified_companions_into_live 가 반영.
            companions.append({"excelId": oid, "name": cname, "wb": cwb})
            _vba_trace("pipeline.isolated.companion.opened", excelId=excel_id, isolatedPid=fpid, companionName=cname, companionPath=str(cpath))
        except Exception:
            _vba_trace("pipeline.isolated.companion.error", excelId=excel_id, isolatedPid=fpid, companionName=cname, companionPath=str(cpath))
            pass
    return fapp, ftarget, fpid, companions


def _sync_modified_companions_into_live(companions, excel_id, fpid, work):
    """격리 인스턴스에서 '대상(ftarget)'이 아닌 동반 워크북이 변형됐으면(Saved=False),
    그 변경을 해당 라이브 세션 워크북으로도 되돌려쓴다.
    교차파일 스텝(입력↔출력 쓰기)이 어느 세션을 ftarget 으로 잡았든 결과가 버려지지 않게 한다.
    (기존: ftarget 한 개만 라이브에 반영 → 쓰기 대상이 동반본이면 유실.)
    호출자(EXCEL_LOCK 보유, fapp 워크북 아직 열림) 안에서만 사용한다."""
    for comp in (companions or []):
        cwb = comp.get("wb")
        oid = comp.get("excelId")
        cname = comp.get("name") or ""
        if cwb is None or not oid:
            continue
        try:
            if bool(cwb.Saved):
                continue  # 읽기만 함(변경 없음) → 되돌려쓸 것 없음
        except Exception:
            continue
        other = EXCEL_SESSIONS.get(oid)
        if not other:
            continue
        try:
            oapp, owb = session_workbook(other)
        except Exception:
            continue
        try:
            sdir = Path(work) / ("sync_" + uuid.uuid4().hex[:6])
            sdir.mkdir(parents=True, exist_ok=True)
            spath = sdir / (cname or "companion.xlsx")
            cwb.SaveCopyAs(str(spath))
            other["rev"] = int(other.get("rev") or 0) + 1  # 동반 스냅샷 신선도 무효화(다음 스텝이 최신 읽도록)
            try:
                _protect_workbook_for_read_only_mirror(owb, False)
            except Exception:
                pass
            try:
                oapp.ScreenUpdating = False
            except Exception:
                pass
            _copy_source_workbook_into_target(oapp, owb, str(spath))
            # [리뷰#11] 동기화로 보호를 풀었으니 읽기전용 미러 보호를 다시 건다(primary 경로의
            # _restore_live_protected_view 와 대칭). 안 하면 동반본이 보호 해제된 채 남아 사용자가
            # 읽기전용 미러를 직접 편집 → 라이브가 소스/스킬 상태와 어긋난다.
            try:
                _protect_workbook_for_read_only_mirror(owb, True)
            except Exception:
                pass
            other["appliedStepSigs"] = None
            try:
                _restore_live_window(other, oapp, owb)
            except Exception:
                pass
            _vba_trace("pipeline.companion.synced", excelId=excel_id, isolatedPid=fpid,
                       companionExcelId=oid, companionName=cname)
        except Exception as err:
            _vba_trace("pipeline.companion.sync.error", excelId=excel_id, isolatedPid=fpid,
                       companionExcelId=oid, companionName=cname, error=str(err))


def _run_vba_pipeline_on_session_impl(excel_id, steps, reset=True, entry=None, view_sheet=None):
    """VBA/Python 스킬 파이프라인을 적용한다.
    [핵심] 라이브(임베드/오버레이) Excel 인스턴스는 reset 후 VBA Application.Run 이 간헐적으로 RPC 로
    사망한다(전체실행이 100% 실패한 근본원인 — 진단 로그/FRESH_INSTANCE_PROBE 로 확정). 새(비임베드)
    인스턴스는 항상 정상 실행되므로, 스텝이 있는 파이프라인은 '격리된 새 Excel'에서 실행하고 결과
    워크북만 라이브에 반영(_copy_source_workbook_into_target)한다. 스텝 없는 reset 은 VBA 가 없어
    라이브에서 직접 처리(안전)."""
    entry = (str(entry).strip() if entry else "") or VBA_SKILL_ENTRY
    steps = steps or []
    with EXCEL_LOCK:
        session = get_excel_session(excel_id)
        session["rev"] = int(session.get("rev") or 0) + 1  # [리뷰②] 동반 스냅샷 신선도 추적
        app, wb = session_workbook(session)
        initial_view = _capture_live_view_state(app, wb, session)
        run_steps = [s for s in steps
                     if ((s.get("code") if isinstance(s, dict) else str(s)) or "").strip()]
        _vba_trace(
            "pipeline.impl.start",
            excelId=excel_id,
            reset=reset,
            entry=entry,
            viewSheet=view_sheet,
            sessionName=session.get("name"),
            sessionPath=session.get("path"),
            workbook=_trace_workbook_info(wb),
            steps=len(steps or []),
            runSteps=len(run_steps),
            # [진단 계측 2026-08-11] 클라가 '무엇을 보냈는지'를 백엔드 쪽에도 남긴다.
            # 클라 로그(pipeline.run.request 의 sentIdx)와 대조하면, 꺼진 단계가 실행에
            # 섞여 들어갔는지를 양쪽에서 교차 확인할 수 있다(제보: OFF 인데 시트에 반영).
            stepIdxs=",".join(
                str(s.get("stepIdx")) for s in run_steps if isinstance(s, dict)
            ),
        )
        try:
            if not run_steps:
                # 스텝 없음(리셋 전용/빈 파이프라인): VBA 실행이 없으므로 라이브에서 직접 처리.
                if reset:
                    source = session.get("sourcePath") or session.get("path")
                    try:
                        _protect_workbook_for_read_only_mirror(wb, False)
                    except Exception:
                        pass
                    try:
                        app.ScreenUpdating = False
                    except Exception:
                        pass
                    _copy_source_workbook_into_target(app, wb, source)
                    session["appliedStepSigs"] = None
                    # [SBAGENT-138] 리셋은 워크북을 '메모리'에서만 원본으로 되돌린다. 저사양 PC 는 다음 스텝 콜
                    # 사이에 Excel COM 참조가 죽어 session_workbook 이 '디스크(작업복사본)'에서 재오픈하는데,
                    # 그 파일은 아직 직전 실행 결과(예: Sheet1→06_DAS) 상태라 1단계가 "시트 못 찾음"으로 터졌다
                    # (고사양은 참조가 살아 메모리 복원본을 봐서 정상). → 복원 직후 디스크에 저장해 재오픈해도
                    # 항상 원본을 보장한다(직후 reset:false 격리 SaveCopyAs 도 pristine 을 복사).
                    try:
                        _da_prev = app.DisplayAlerts
                    except Exception:
                        _da_prev = None
                    try:
                        app.DisplayAlerts = False
                    except Exception:
                        pass
                    try:
                        wb.Save()
                        _vba_trace("pipeline.reset.persisted", excelId=excel_id, workbook=_trace_workbook_info(wb))
                    except Exception as _reset_save_err:
                        _vba_trace("pipeline.reset.save.skip", excelId=excel_id, error=str(_reset_save_err))
                    finally:
                        if _da_prev is not None:
                            try:
                                app.DisplayAlerts = _da_prev
                            except Exception:
                                pass
                return {"ok": True, "excelId": excel_id, "applied": 0}

            # 스텝 있음: 격리된 새 인스턴스에서 reset+실행 후 결과를 라이브에 반영.
            work = Path(tempfile.mkdtemp(prefix="b2b_isopipe_"))
            fapp = None
            fpid = None
            companions = []
            step_snapshots = []  # [0.5.14 batch 빠른복구] 스텝 실행 '전' ftarget 스냅샷(downloadId) 목록
            try:
                fapp, ftarget, fpid, companions = _setup_isolated_pipeline_instance(session, excel_id, reset, work)
                try:
                    _protect_workbook_for_read_only_mirror(ftarget, False)
                except Exception:
                    pass
                # [교차파일 쓰기 1004 수정] 동반본도 보호 해제 — 라이브 미러의 UserInterfaceOnly 보호는
                # SaveCopyAs→재오픈을 거치면 '완전 보호'로 바뀐다(Excel 고전 함정). 타깃만 풀고 동반본을
                # 안 풀어서, 교차파일 녹화 재현(정산서 D1 붙여넣기)이 '보호된 시트' 1004 로 즉사했다
                # (실측 14:02 step2). 풀런 경로(9372)는 이미 동반본을 풀고 있어 이 경로만 비대칭이었다.
                for _comp in (companions or []):
                    try:
                        _protect_workbook_for_read_only_mirror(_comp.get("wb"), False)
                    except Exception:
                        pass
                def _activate_step_target_sheet(st):
                    if not isinstance(st, dict):
                        return
                    # [초기 컨텍스트 결정론화] 시트명 유무와 무관하게 먼저 앵커 워크북을 활성화한다.
                    # 격리 인스턴스는 '마지막에 연 동반본'이 활성이라, 선행 Windows.Activate 없는
                    # 녹화 조각이 엉뚱한 워크북에서 실행됐다(실측 15:30: 1조각 복붙 절반 유실).
                    # COM Workbook.Activate 는 복사 마퀴 무해(프로브 실측) — 스텝 자신의
                    # Windows().Activate 가 있으면 그게 다시 대상을 잡으므로 안전한 기본값이다.
                    try:
                        ftarget.Activate()
                    except Exception:
                        pass
                    sheet_name = (
                        st.get("targetSheetName")
                        or st.get("targetSheet")
                        or st.get("viewSheet")
                        or view_sheet
                    )
                    if not sheet_name:
                        return
                    try:
                        try:
                            names = _excel_collection_names(ftarget.Worksheets)
                            alias = _resolve_ephemeral_excel_open_sheet_alias(sheet_name, names)
                            if alias:
                                sheet_name = alias
                        except Exception:
                            pass
                        ftarget.Worksheets(str(sheet_name)).Activate()
                        _vba_trace(
                            "pipeline.step.activate_sheet",
                            excelId=excel_id,
                            isolatedPid=fpid,
                            stepIdx=st.get("stepIdx"),
                            stepId=st.get("stepId"),
                            sheet=str(sheet_name),
                        )
                    except Exception as err:
                        _vba_trace(
                            "pipeline.step.activate_sheet.skip",
                            excelId=excel_id,
                            isolatedPid=fpid,
                            stepIdx=st.get("stepIdx"),
                            stepId=st.get("stepId"),
                            sheet=str(sheet_name),
                            error=str(err),
                        )
                _run_ord = 0
                _run_total = len(run_steps)
                for fallback_idx, st in enumerate(steps):
                    code = (st.get("code") if isinstance(st, dict) else str(st)) or ""
                    if not code.strip():
                        continue
                    _run_ord += 1
                    try:
                        PIPELINE_PROGRESS[excel_id] = {"current": _run_ord, "total": _run_total, "ts": time.time()}
                    except Exception:
                        pass
                    lang = (st.get("language") if isinstance(st, dict) else "") or ""
                    _vba_trace(
                        "pipeline.step.start",
                        excelId=excel_id,
                        isolatedPid=fpid,
                        stepIdx=st.get("stepIdx") if isinstance(st, dict) else None,
                        stepId=st.get("stepId") if isinstance(st, dict) else None,
                        language=lang,
                        description=_trace_text(st.get("description") if isinstance(st, dict) else "", 220),
                        codeLen=len(str(code)),
                        codeHash=_trace_hash(code),
                        codeHead=_trace_text(code, 360),
                    )
                    # [0.5.14 batch 빠른복구] 이 스텝 실행 '전' ftarget 상태를 영속 RESULTS(BACKEND_DIR)에
                    # SaveCopyAs 해 step._preApplySnapshot 용 downloadId 를 만든다(0.5.13 per-step 스냅샷·python
                    # 라이브 경로와 동형). 격리 batch 경로에서도 마지막 단계 OFF/삭제 빠른복구가 13처럼 되게 한다.
                    # 파일은 종료 시 cleanup_backend_runtime_files 가 RESULTS 와 함께 정리(디스크 누수 없음).
                    # 실패해도 파이프라인은 계속 — 그 스텝만 빠른복구 대신 재실행 폴백.
                    try:
                        BACKEND_DIR.mkdir(parents=True, exist_ok=True)
                        _snap_name = Path(str(session.get("name") or "result.xlsx")).name
                        if not Path(_snap_name).suffix:
                            _snap_name += ".xlsx"
                        _snap_path = BACKEND_DIR / ("prestep_%s_%s" % (uuid.uuid4().hex, _snap_name))
                        ftarget.SaveCopyAs(str(_snap_path))
                        _snap_rid = uuid.uuid4().hex
                        RESULTS[_snap_rid] = {
                            "path": str(_snap_path),
                            "name": Path(_snap_path).name,
                            "created": time.time(),
                        }
                        step_snapshots.append({
                            "stepIdx": st.get("stepIdx") if isinstance(st, dict) else None,
                            "stepId": st.get("stepId") if isinstance(st, dict) else None,
                            "downloadId": _snap_rid,
                            "downloadUrl": "/api/workbooks/download/%s" % _snap_rid,
                            "name": Path(_snap_path).name,
                        })
                    except Exception as _snap_err:
                        _vba_trace(
                            "pipeline.step.snapshot.skip",
                            excelId=excel_id,
                            isolatedPid=fpid,
                            stepIdx=st.get("stepIdx") if isinstance(st, dict) else None,
                            stepId=st.get("stepId") if isinstance(st, dict) else None,
                            error=str(_snap_err),
                        )
                    try:
                        _activate_step_target_sheet(st)
                        if str(lang).lower() == "python":
                            _exec_python_com_skill(
                                fapp,
                                ftarget,
                                session,
                                code,
                                skip_static=bool(isinstance(st, dict) and st.get("trustedStatic") is True),
                                timeout_s=_step_extended_timeout_s(st),
                            )
                        else:
                            _inject_and_run_vba(fapp, ftarget, code, entry)
                    except PipelineExecutionError as _pe:
                        # 스텝 내부에서 던진 PipelineExecutionError 에도 직전 스냅샷을 실어 자동복구 이어실행을 살린다.
                        try:
                            if isinstance(getattr(_pe, "info", None), dict):
                                _pe.info["stepSnapshots"] = list(step_snapshots)
                        except Exception:
                            pass
                        raise
                    except Exception as err:
                        info = _vba_pipeline_step_info(st, fallback_idx, err)
                        _vba_trace(
                            "pipeline.step.error",
                            excelId=excel_id,
                            isolatedPid=fpid,
                            stepIdx=info.get("stepIdx"),
                            stepId=info.get("stepId"),
                            language=info.get("language"),
                            description=_trace_text(info.get("description") or "", 220),
                            error=str(err),
                            errorInfo=info,
                        )
                        # [자동복구 이어실행] 실패해도 실패 step 직전까지의 스텝-전 스냅샷을 errorInfo 로 실어 보낸다
                        # → 클라가 step._preApplySnapshot 로 wiring → 자동복구 후 '실패 step 직전'으로 되돌려
                        # 이어실행(중복 적용 없이) 가능. (없으면 "스냅샷 없음"으로 이어실행이 중단됐었음.)
                        try:
                            info["stepSnapshots"] = list(step_snapshots)
                        except Exception:
                            pass
                        raise PipelineExecutionError(info)
                    _vba_trace(
                        "pipeline.step.ok",
                        excelId=excel_id,
                        isolatedPid=fpid,
                        stepIdx=st.get("stepIdx") if isinstance(st, dict) else None,
                        stepId=st.get("stepId") if isinstance(st, dict) else None,
                        language=lang,
                    )
                # 결과 저장 → 라이브 대상 워크북에 시트 교체로 반영(라이브에선 VBA 안 돌리므로 안전).
                # [CSV 무성 데이터 손실] 승격이 다운로드 경로에만 걸려 있어 여기가 뚫려 있었다.
                # data.csv 세션에 스킬이 시트를 추가하면 SaveCopyAs 는 CSV 포맷을 유지해 ActiveSheet
                # 1장만 기록하고, 그 1장짜리 스냅샷을 _copy_source_workbook_into_target 이 되돌리며
                # 라이브의 '모든 시트를 Delete' 한다 → 추가 시트 + 비활성 원본 시트가 통째로 소실.
                # DisplayAlerts=False 라 Excel 의 "CSV는 시트 1장만" 경고도 안 뜨고, 사후 무결성 검사는
                # 스냅샷(1장) ⊆ 라이브(1장) 이라 통과해 구조적으로 못 잡는다.
                result_name = Path(str(session.get("name") or "result.xlsx")).name
                _promoted_name = _promote_csv_multisheet_name(result_name, ftarget)
                rpath = work / ("result_" + _promoted_name)
                if _promoted_name != result_name:
                    # SaveCopyAs 는 현재(CSV) 포맷을 유지하므로 승격엔 SaveAs(경로 재지정)가 필요하다.
                    ftarget.SaveAs(str(rpath), FileFormat=51)
                    # 세션도 xlsx 로 코히어런트 갱신(다운로드 경로 4574-4580 과 동일 정책) —
                    # 멀티시트가 된 이상 이 파일은 사실상 xlsx 다.
                    session["name"] = _promoted_name
                else:
                    ftarget.SaveCopyAs(str(rpath))
                try:
                    _protect_workbook_for_read_only_mirror(wb, False)
                except Exception:
                    pass
                try:
                    app.ScreenUpdating = False
                except Exception:
                    pass
                _copy_source_workbook_into_target(app, wb, str(rpath))
                _vba_trace(
                    "pipeline.result.copied",
                    excelId=excel_id,
                    isolatedPid=fpid,
                    resultPath=str(rpath),
                    liveWorkbook=_trace_workbook_info(wb),
                )
                # [교차파일 유실 수정] 이 스텝이 '다른 파일(동반본)'을 변형했으면 그 변경도
                # 해당 라이브 세션으로 되돌려쓴다. (ftarget 한 개만 반영하던 한계 보완)
                _sync_modified_companions_into_live(companions, excel_id, fpid, work)
                session["appliedStepSigs"] = None
            finally:
                # [리뷰#12] 이 격리 인스턴스(fapp)의 워크북 이름 별칭을 정리한다(pid 단위 누수 방지 +
                # 추후 pid 재사용 시 스테일 엔트리 잔존 방지). fapp 살아있을 때(Quit 전) pid 로 정리.
                try:
                    if fapp is not None:
                        _clear_workbook_name_aliases(fapp)
                except Exception:
                    pass
                try:
                    if fapp is not None:
                        for w in list(fapp.Workbooks):
                            try:
                                w.Close(SaveChanges=False)
                            except Exception:
                                pass
                except Exception:
                    pass
                try:
                    if fapp is not None:
                        fapp.Quit()
                except Exception:
                    pass
                try:
                    if fpid:
                        _kill_pid_quiet(fpid)
                except Exception:
                    pass
                shutil.rmtree(work, ignore_errors=True)
                try:
                    PIPELINE_PROGRESS.pop(excel_id, None)  # 진행률 정리(성공/실패 공통)
                except Exception:
                    pass
            result = {"ok": True, "excelId": excel_id, "applied": len(run_steps)}
            result["stepSnapshots"] = step_snapshots  # [0.5.14] 스텝별 적용-전 스냅샷(클라가 _preApplySnapshot 로 사용)
            # [#5] 격리 적용도 라이브 미러 캐시(시트명/미리보기/차원)를 갱신하도록 경량 스키마를 함께 싣는다.
            # 반영(_copy_source_workbook_into_target) 이후 라이브 wb 기준 → 새로 생긴 시트(추가/복사/피벗)가
            # @멘션 목록에 안 뜨던 문제 해결. Python 단일 경로(_run_python_on_session_impl)와 동일 패턴.
            try:
                result["liveSchema"] = _live_preview_schema(wb)
            except Exception:
                pass
            return result
        finally:
            # 라이브 창/상태 복원(대상 시트/선택 복원 포함).
            # 예전 코드는 sheet.Activate() 만 해서 A1/이전 ActiveCell 로 선택이 유실됐고,
            # 이후 ActiveSheet/Selection 기반 작업과 미러 선택 추적이 서로 어긋났다.
            try:
                if view_sheet:
                    view_state = dict(initial_view or {})
                    view_state["sheet"] = str(view_sheet)
                    view_state["address"] = ""
                    _restore_live_view_state(app, wb, view_state, session)
                else:
                    _restore_live_view_state(app, wb, initial_view, session)
            except Exception:
                pass
            try:
                _ws_target = None
                if view_sheet:
                    try:
                        _ws_target = wb.Worksheets(str(view_sheet))
                    except Exception:
                        _ws_target = None
                if _ws_target is None and initial_view and initial_view.get("sheet"):
                    try:
                        _ws_target = wb.Worksheets(str(initial_view.get("sheet")))
                    except Exception:
                        _ws_target = None
                if _ws_target is not None:
                    _ws_target.Activate()
            except Exception:
                pass
            _restore_app_state(app)
            try:
                _restore_live_protected_view(app, wb)
            except Exception:
                pass
            try:
                _restore_live_window(session, app, wb)
            except Exception:
                pass


def run_vba_on_session(excel_id, code, entry=None, restore_window=True):
    return excel_call(_run_vba_on_session_impl, excel_id, code, entry=entry, restore_window=restore_window, timeout=180)


def run_vba_pipeline_on_session(excel_id, steps, reset=True, entry=None, view_sheet=None):
    return excel_call(_run_vba_pipeline_on_session_impl, excel_id, steps, reset=reset, entry=entry, view_sheet=view_sheet, timeout=PY_UNLIMITED_OUTER_S)


def _run_full_pipeline_single_instance_impl(groups, reset_excel_ids=None, view_sheet=None, entry=None, output_mode="sync", state_sig=None):
    """[0.5.15 백그라운드 전체실행] 격리 인스턴스 '1개'에서 관여 파일 전부를 '원본'부터 열고, 전 그룹·스텝을
    순서대로 실행한 뒤, 변경된 파일만 라이브 세션에 '파일당 1회' 반영한다. 그룹마다 새 인스턴스를 spawn 하고
    파일을 통째 동기화하던 오버헤드(전체실행 '멈춤'의 주원인)를 제거한다 — 실행 중 라이브 뷰는 갱신하지 않고
    (클라가 창을 숨김) 끝에만 반영한다.
      groups: [{"excelId","steps":[payload...]}]  연속 같은-파일 묶음(순서 보존). excelId = 대상 라이브 세션.
      reset_excel_ids: '원본부터' 열어야 할 세션 전부(targets ∪ 교차참조). 모두 pristine 으로 열어 cross-file
                       Workbooks("파일명") 참조가 격리 인스턴스 안에서 그대로 동작한다.
    반환 {ok, applied, stepSnapshots(전역), perFileLiveSchema:{excelId:schema}}. 실패 시 라이브 무손상(반영 전 raise)."""
    entry = (str(entry).strip() if entry else "") or VBA_SKILL_ENTRY
    # 코드 있는 실행 스텝만 남기고 그룹 정규화 + 전역 총합(진행률용)
    norm_groups = []
    total_steps = 0
    for g in (groups or []):
        gsteps = [s for s in (g.get("steps") or [])
                  if ((s.get("code") if isinstance(s, dict) else str(s)) or "").strip()]
        if not gsteps:
            continue
        norm_groups.append({"excelId": g.get("excelId"), "steps": gsteps})
        total_steps += len(gsteps)
    if not norm_groups:
        return {"ok": True, "applied": 0, "stepSnapshots": [], "perFileLiveSchema": {}}

    anchor_excel_id = norm_groups[0]["excelId"]
    # 열어야 할 세션 = 그룹 대상 ∪ reset 대상(교차참조 포함), 최초 등장 순서 유지
    open_ids = []
    for g in norm_groups:
        if g["excelId"] and g["excelId"] not in open_ids:
            open_ids.append(g["excelId"])
    for sid in (reset_excel_ids or []):
        if sid and sid not in open_ids:
            open_ids.append(sid)

    with EXCEL_LOCK:
        sessions = {}        # excelId -> session
        initial_views = {}   # excelId -> view state(끝에 복원)
        for sid in open_ids:
            s = get_excel_session(sid)
            s["rev"] = int(s.get("rev") or 0) + 1
            sessions[sid] = s
            try:
                a0, w0 = session_workbook(s)
                initial_views[sid] = _capture_live_view_state(a0, w0, s)
            except Exception:
                initial_views[sid] = None

        _ensure_vbom_access()
        _disable_vba_break_on_all_errors()
        work = Path(tempfile.mkdtemp(prefix="b2b_fullrun_"))
        fapp = None
        fpid = None
        step_snapshots = []
        byExcel = {}   # excelId -> {"wb","name","session"}
        result = {"ok": True, "applied": 0, "stepSnapshots": step_snapshots, "perFileLiveSchema": {}, "outputFiles": []}
        try:
            fapp = win32com.client.DispatchEx("Excel.Application")
            _track_spawned_excel_app(fapp)
            try:
                fpid = _excel_process_id(fapp)
            except Exception:
                fpid = None
            for attr, val in (("Visible", False), ("DisplayAlerts", False), ("EnableEvents", False), ("AskToUpdateLinks", False)):
                try:
                    setattr(fapp, attr, val)
                except Exception:
                    pass
            _vba_trace("fullrun.setup.start", anchorExcelId=anchor_excel_id, isolatedPid=fpid,
                       openIds=list(open_ids), groups=len(norm_groups), totalSteps=total_steps, work=str(work))

            # 1) 관여 파일 전부 '원본'부터 1회 열기 (intended_name → VBA Workbooks("파일명") 교차참조 동작)
            for sid in open_ids:
                s = sessions[sid]
                name = Path(str(s.get("name") or "")).name
                if not name:
                    try:
                        _a, _w = session_workbook(s)
                        name = Path(str(_w.Name)).name
                    except Exception:
                        name = "book_%s.xlsx" % uuid.uuid4().hex[:6]
                src = s.get("sourcePath") or s.get("path")
                tdir = work / ("t_" + uuid.uuid4().hex[:6])
                tdir.mkdir(parents=True, exist_ok=True)
                tpath = tdir / name
                shutil.copy2(Path(src), tpath)
                fwb, _t = excel_workbooks_open(fapp, str(tpath), read_only=False, intended_name=name)
                try:
                    _protect_workbook_for_read_only_mirror(fwb, False)
                except Exception:
                    pass
                byExcel[sid] = {"wb": fwb, "name": name, "session": s}
                _vba_trace("fullrun.file.opened", anchorExcelId=anchor_excel_id, isolatedPid=fpid,
                           excelId=sid, name=name, targetPath=str(tpath))

            # 1-b) 교차참조용 '동반 워크북': open_ids 에 없는 다른 라이브 세션도 (현재 상태로) 함께 열어 둔다.
            # VBA 의 Workbooks("다른파일.xlsx") / Python ctx.book("…") 교차참조가 격리 인스턴스 안에서 해석되게
            # 한다(구 per-group _setup_isolated_pipeline_instance 의 companion 과 동형). 안 열면 그 파일을 참조하는
            # 스텝이 "워크북이 열려 있지 않습니다"로 실패한다. 변경되면 끝의 동기화에서 함께 라이브로 반영된다.
            _opened_names = set(str(ent["name"]).lower() for ent in byExcel.values())
            for oid, other in list(EXCEL_SESSIONS.items()):
                if oid in byExcel or not other.get("liveEditable"):
                    continue
                try:
                    _oa, o_wb = session_workbook(other)
                except Exception:
                    continue
                cname = Path(str(other.get("name") or "")).name
                if not cname:
                    try:
                        cname = Path(str(o_wb.Name)).name
                    except Exception:
                        cname = ""
                if not cname or cname.lower() in _opened_names:
                    continue
                try:
                    cdir = work / ("c_" + uuid.uuid4().hex[:6])
                    cdir.mkdir(parents=True, exist_ok=True)
                    cpath = cdir / cname
                    o_wb.SaveCopyAs(str(cpath))  # 동반본은 '현재 라이브 상태'로(읽기 소스/교차쓰기 대상)
                    cwb, _ct = excel_workbooks_open(fapp, str(cpath), read_only=False, intended_name=cname)
                    try:
                        _protect_workbook_for_read_only_mirror(cwb, False)
                    except Exception:
                        pass
                    _opened_names.add(cname.lower())
                    sessions[oid] = other
                    byExcel[oid] = {"wb": cwb, "name": cname, "session": other}
                    _vba_trace("fullrun.companion.opened", anchorExcelId=anchor_excel_id, isolatedPid=fpid,
                               excelId=oid, name=cname, companionPath=str(cpath))
                except Exception as _cerr:
                    _vba_trace("fullrun.companion.error", anchorExcelId=anchor_excel_id, isolatedPid=fpid,
                               excelId=oid, name=cname, error=str(_cerr))

            # [적용됨-미반영 수정] 실패 '시점'의 관여 파일 전체 상태(=마지막 완료 스텝까지 반영본)를
            # 파일별로 스냅샷해 오류에 동봉한다. 기존 stepSnapshots 는 '실패 스텝의 파일' 것만 있어,
            # 다파일/교차파일 스킬에서 클라 복원이 그 파일 하나만 되돌리고 나머지는 원본으로 남는데
            # 라벨은 인덱스 기준 '적용됨'이라 UI-라이브 불일치가 났다. 전 파일 스냅샷이면 복원 후
            # 'Step1~(N-1) 적용됨' 표시가 모든 파일에서 실제와 일치한다(companion 포함 — byExcel 전체).
            def _capture_fail_state_snapshots():
                snaps = []
                for _sid, _ent in byExcel.items():
                    try:
                        _p = BACKEND_DIR / ("prestep_%s_%s" % (uuid.uuid4().hex, _ent["name"]))
                        _ent["wb"].SaveCopyAs(str(_p))
                        _rid = uuid.uuid4().hex
                        RESULTS[_rid] = {"path": str(_p), "name": Path(_p).name, "created": time.time()}
                        snaps.append({"excelId": _sid, "downloadId": _rid})
                    except Exception as _fs_err:
                        _vba_trace("fullrun.failstate.snapshot.skip", anchorExcelId=anchor_excel_id,
                                   excelId=_sid, error=str(_fs_err))
                return snaps

            # 2) 전 그룹·스텝 순서대로 실행 (ftarget = 그 그룹 파일의 wb)
            ordinal = 0
            for g in norm_groups:
                gid = g["excelId"]
                if gid not in byExcel:
                    raise RuntimeError("full-run: 그룹 대상 세션이 열리지 않았습니다: %s" % gid)
                ftarget = byExcel[gid]["wb"]
                fsession = byExcel[gid]["session"]
                for fallback_idx, st in enumerate(g["steps"]):
                    code = (st.get("code") if isinstance(st, dict) else str(st)) or ""
                    if not code.strip():
                        continue
                    ordinal += 1
                    try:
                        PIPELINE_PROGRESS[anchor_excel_id] = {"current": ordinal, "total": total_steps, "ts": time.time()}
                    except Exception:
                        pass
                    lang = (st.get("language") if isinstance(st, dict) else "") or ""
                    # 스텝-전 스냅샷(이어실행/빠른복구용) — ftarget 상태를 영속 RESULTS 로
                    try:
                        BACKEND_DIR.mkdir(parents=True, exist_ok=True)
                        _snap_path = BACKEND_DIR / ("prestep_%s_%s" % (uuid.uuid4().hex, byExcel[gid]["name"]))
                        ftarget.SaveCopyAs(str(_snap_path))
                        _snap_rid = uuid.uuid4().hex
                        RESULTS[_snap_rid] = {"path": str(_snap_path), "name": Path(_snap_path).name, "created": time.time()}
                        step_snapshots.append({
                            "excelId": gid,  # 다파일: 이 스냅샷이 속한 라이브 세션(클라 이어실행 정확 매핑)
                            "stepIdx": st.get("stepIdx") if isinstance(st, dict) else None,
                            "stepId": st.get("stepId") if isinstance(st, dict) else None,
                            "downloadId": _snap_rid,
                            "downloadUrl": "/api/workbooks/download/%s" % _snap_rid,
                            "name": Path(_snap_path).name,
                        })
                    except Exception as _snap_err:
                        _vba_trace("fullrun.step.snapshot.skip", anchorExcelId=anchor_excel_id, isolatedPid=fpid,
                                   excelId=gid, error=str(_snap_err))
                    # 대상 시트 활성화(있으면). 시트명이 없어도 앵커 워크북은 먼저 활성화 —
                    # 선행 Windows.Activate 없는 녹화 조각이 '마지막에 연 동반본'에서 실행되는
                    # 초기 컨텍스트 비결정론 차단(실측 15:30, 격리 경로와 대칭).
                    if isinstance(st, dict):
                        try:
                            ftarget.Activate()
                        except Exception:
                            pass
                        _sheet = st.get("targetSheetName") or st.get("targetSheet") or st.get("viewSheet")
                        if _sheet:
                            try:
                                ftarget.Worksheets(str(_sheet)).Activate()
                            except Exception:
                                pass
                    _vba_trace("fullrun.step.start", anchorExcelId=anchor_excel_id, isolatedPid=fpid,
                               excelId=gid, ordinal=ordinal, total=total_steps, language=lang,
                               stepIdx=st.get("stepIdx") if isinstance(st, dict) else None, codeHash=_trace_hash(code))
                    try:
                        if str(lang).lower() == "python":
                            _exec_python_com_skill(fapp, ftarget, fsession, code,
                                                   skip_static=bool(isinstance(st, dict) and st.get("trustedStatic") is True),
                                                   timeout_s=_step_extended_timeout_s(st))
                        else:
                            _inject_and_run_vba(fapp, ftarget, code, entry)
                    except PipelineExecutionError as _pe:
                        try:
                            if isinstance(getattr(_pe, "info", None), dict):
                                _pe.info["stepSnapshots"] = list(step_snapshots)
                                _pe.info["failStateSnapshots"] = _capture_fail_state_snapshots()
                        except Exception:
                            pass
                        raise
                    except Exception as err:
                        info = _vba_pipeline_step_info(st, fallback_idx, err)
                        _vba_trace("fullrun.step.error", anchorExcelId=anchor_excel_id, isolatedPid=fpid,
                                   excelId=gid, stepIdx=info.get("stepIdx"), error=str(err), errorInfo=info)
                        try:
                            info["stepSnapshots"] = list(step_snapshots)
                            info["failStateSnapshots"] = _capture_fail_state_snapshots()
                        except Exception:
                            pass
                        raise PipelineExecutionError(info)
                    _vba_trace("fullrun.step.ok", anchorExcelId=anchor_excel_id, isolatedPid=fpid, excelId=gid, ordinal=ordinal)

            # 3) 변경된 파일만 라이브에 '파일당 1회' 반영 (읽기만 한 파일은 Saved=True → 스킵)
            # [저사양 거짓실패 방지] 이 '최종 동기화'(통째 시트 교체)는 저사양에서 수 분 걸릴 수 있는데, 진행률이
            # 스텝만 추적하면 클라가 'N/N'에서 멈춰보이다 타임아웃으로 거짓 실패한다. 동기화 단계도 진행률에 실어
            # (phase=syncing) 클라가 '결과 반영 중'을 표시하고 활동 중임을 알 수 있게 한다.
            _to_sync = []
            for sid, ent in byExcel.items():
                try:
                    _changed = not bool(ent["wb"].Saved)
                except Exception:
                    _changed = True
                if _changed:
                    _to_sync.append((sid, ent))
            _sync_total = len(_to_sync)
            _is_file_mode = (str(output_mode) == "file")
            for _sync_i, (sid, ent) in enumerate(_to_sync, start=1):
                try:
                    PIPELINE_PROGRESS[anchor_excel_id] = {"current": total_steps, "total": total_steps,
                                                          "phase": ("saving" if _is_file_mode else "syncing"),
                                                          "syncCurrent": _sync_i, "syncTotal": _sync_total, "ts": time.time()}
                except Exception:
                    pass
                fwb = ent["wb"]
                s = ent["session"]
                if _is_file_mode:
                    # [실행기 파일출력] 라이브 동기화(통째 시트 교체) 생략 — 결과를 output 폴더에 저장 + 다운로드용
                    # RESULTS 등록. 라이브 wb 는 안 건드린다(무손상·빠름). 뷰 미반영이라 perFileLiveSchema 도 스킵.
                    try:
                        out_dir = default_output_dir()
                        out_dir.mkdir(parents=True, exist_ok=True)
                        _stem = Path(ent["name"]).stem
                        _suf = Path(ent["name"]).suffix or ".xlsx"
                        # [CSV 새시트 버그] CSV/TSV 는 시트 하나만 담는다 → 스킬이 새 시트(예: 필터→새시트)를 추가했는데
                        # 원본 .csv 로 저장하면 그 포맷을 유지해 시트가 붕괴(기존 시트 덮어써짐). 멀티시트면 .xlsx 로 승격.
                        # SaveCopyAs 는 현재(CSV) 포맷을 유지하므로 승격 시엔 SaveAs FileFormat=51(xlOpenXMLWorkbook)로 변환 저장.
                        try:
                            _multi = int(fwb.Worksheets.Count) > 1
                        except Exception:
                            _multi = False
                        _promote_xlsx = _multi and _suf.lower() in (".csv", ".tsv")
                        if _promote_xlsx:
                            _suf = ".xlsx"
                        _ts = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
                        out_path = out_dir / ("결과_%s_%s%s" % (_stem, _ts, _suf))
                        if out_path.exists():
                            out_path = out_dir / ("결과_%s_%s_%s%s" % (_stem, _ts, uuid.uuid4().hex[:6], _suf))
                        if _promote_xlsx:
                            fwb.SaveAs(str(out_path), FileFormat=51)
                        else:
                            fwb.SaveCopyAs(str(out_path))
                        _rid = uuid.uuid4().hex
                        RESULTS[_rid] = {"path": str(out_path), "name": out_path.name, "created": time.time()}
                        result["outputFiles"].append({
                            "excelId": sid, "name": out_path.name, "path": str(out_path),
                            "downloadId": _rid, "downloadUrl": "/api/workbooks/download/%s" % _rid,
                        })
                        _vba_trace("fullrun.file.saved", anchorExcelId=anchor_excel_id, isolatedPid=fpid,
                                   excelId=sid, name=ent["name"], outputPath=str(out_path))
                    except Exception as _save_err:
                        _vba_trace("fullrun.file.save_error", anchorExcelId=anchor_excel_id, isolatedPid=fpid,
                                   excelId=sid, name=ent["name"], error=str(_save_err))
                    continue
                # output_mode == "sync" (생성기): 기존 라이브 동기화
                try:
                    live_app, live_wb = session_workbook(s)
                except Exception:
                    _vba_trace("fullrun.sync.skip_nolive", anchorExcelId=anchor_excel_id, excelId=sid)
                    continue
                rdir = work / ("r_" + uuid.uuid4().hex[:6])
                rdir.mkdir(parents=True, exist_ok=True)
                rpath = rdir / ("result_" + ent["name"])
                fwb.SaveCopyAs(str(rpath))
                try:
                    _protect_workbook_for_read_only_mirror(live_wb, False)
                except Exception:
                    pass
                try:
                    live_app.ScreenUpdating = False
                except Exception:
                    pass
                _copy_source_workbook_into_target(live_app, live_wb, str(rpath))
                s["appliedStepSigs"] = None
                try:
                    _protect_workbook_for_read_only_mirror(live_wb, True)
                except Exception:
                    pass
                try:
                    result["perFileLiveSchema"][sid] = _live_preview_schema(live_wb)
                except Exception:
                    pass
                _vba_trace("fullrun.file.synced", anchorExcelId=anchor_excel_id, isolatedPid=fpid,
                           excelId=sid, name=ent["name"], resultPath=str(rpath))
                # [새로고침 즉시복원] 방금 라이브에 반영한 그 파일이 곧 '스킬 전부 적용된 최종 상태'다.
                # 어차피 finally 의 rmtree(work) 로 버려질 파일이라 옮겨 담을 뿐 — 추가 COM 저장이 없다.
                # (이 경로가 VBA 전체실행. 예전엔 VBA 만 사본이 없어 새로고침 후 항상 전 스텝 재실행이었다)
                if state_sig:
                    try:
                        _wb_rec = WORKBOOKS.get(s.get("workbookId"))
                        if _wb_rec:
                            _save_live_final_snapshot(_wb_rec, state_sig, rpath, move=True)
                    except Exception as _serr:
                        _warn_excel_nonfatal("fullrun live final snapshot", _serr)

            result["applied"] = ordinal
            return result
        finally:
            # 격리 인스턴스 정리(1회) — 누수 방지(0.5.x 디스크 누수 교훈)
            try:
                if fapp is not None:
                    _clear_workbook_name_aliases(fapp)
            except Exception:
                pass
            try:
                if fapp is not None:
                    for w in list(fapp.Workbooks):
                        try:
                            w.Close(SaveChanges=False)
                        except Exception:
                            pass
            except Exception:
                pass
            try:
                if fapp is not None:
                    fapp.Quit()
            except Exception:
                pass
            try:
                if fpid:
                    _kill_pid_quiet(fpid)
            except Exception:
                pass
            shutil.rmtree(work, ignore_errors=True)
            try:
                PIPELINE_PROGRESS.pop(anchor_excel_id, None)
            except Exception:
                pass
            # 라이브 창/뷰 복원(전 세션 — 동반 워크북 포함)
            for sid in list(sessions.keys()):
                s = sessions.get(sid)
                if not s:
                    continue
                try:
                    la, lw = session_workbook(s)
                except Exception:
                    continue
                try:
                    _restore_live_view_state(la, lw, initial_views.get(sid), s)
                except Exception:
                    pass
                try:
                    _restore_app_state(la)
                except Exception:
                    pass
                try:
                    _restore_live_protected_view(la, lw)
                except Exception:
                    pass
                try:
                    # [#1 헤드리스] 실행기 파일출력(output_mode="file")은 라이브를 안 건드리고 화면에도 안 띄운다.
                    # _restore_live_window 는 app.Visible=True 로 창을 강제로 띄우므로, file 모드에선 건너뛴다
                    # (안 그러면 실행기 전체실행 끝에 업로드한 원본 Excel 이 떠버린다). sync(생성기)는 기존대로 복원.
                    if str(output_mode) != "file":
                        _restore_live_window(s, la, lw)
                except Exception:
                    pass


def run_full_pipeline_single_instance(groups, reset_excel_ids=None, view_sheet=None, entry=None, output_mode="sync", state_sig=None):
    return excel_call(_run_full_pipeline_single_instance_impl, groups,
                      reset_excel_ids=reset_excel_ids, view_sheet=view_sheet, entry=entry,
                      output_mode=output_mode, state_sig=state_sig, timeout=PY_UNLIMITED_OUTER_S)


# ===== 복붙 캡처(녹화): 사용자의 Ctrl+C/Ctrl+V 를 역추적해 스킬 스텝으로 저장 =====

def _r1c1_to_a1(r1c1):
    """Excel 'Link' 포맷의 R1C1 범위 표기를 A1 로 변환. Link 포맷은 항상 R1C1 이다.
    지원하는 모든 선택 형태(실측):
      'R1C1:R6C5' → 'A1:E6'   (일반 사각 범위)
      'R2C2'      → 'B2'       (단일 셀)
      'C1:C5'     → 'A:E'      (전체 열 다중)   ← 전체 열 선택 시 R/행 표기가 없음
      'C3'        → 'C:C'      (전체 열 단일)
      'R1:R6'     → '1:6'      (전체 행 다중)   ← 전체 행 선택 시 C/열 표기가 없음
      'R3'        → '3:3'      (전체 행 단일)
    이전 구현은 매치 실패 토큰을 그대로 반환 → 'C1:C5'(전체열)을 A1 셀범위 C1:C5 로,
    'R1:R6'(전체행)을 무효 주소로 오해해 캡처가 깨졌다."""
    s = str(r1c1).strip()

    def _conv(tok):
        tok = tok.strip()
        m = re.match(r"^R(\d+)C(\d+)$", tok)   # 일반 셀 (R 행 + C 열)
        if m:
            return _col_letter(int(m.group(2))) + m.group(1), "cell"
        m = re.match(r"^C(\d+)$", tok)          # 전체 열 (열 번호만)
        if m:
            return _col_letter(int(m.group(1))), "col"
        m = re.match(r"^R(\d+)$", tok)          # 전체 행 (행 번호만)
        if m:
            return m.group(1), "row"
        return tok, "raw"

    def _area(a):
        a = a.strip()
        if ":" in a:
            x, y = a.split(":", 1)
            cx, _ = _conv(x)
            cy, _ = _conv(y)
            return cx + ":" + cy
        val, kind = _conv(a)
        if kind in ("col", "row"):
            # 단일 전체 열/행은 A1 에서 'C:C' / '3:3' 처럼 시작:끝 동일 표기로 만든다.
            return val + ":" + val
        return val

    # 비연속 다중 영역(Ctrl-클릭) Link 는 'C1:C3,C5:C5' / 'R1C1:R6C5,C8:C10' 처럼 콤마로 묶인다.
    # 각 영역을 개별 변환해 합쳐야 'A:C,E:E' 같은 정상 A1 다중영역이 된다(콤마 무시 시 손상됨).
    return ",".join(_area(a) for a in s.split(","))


def _read_excel_clipboard_source():
    """Windows 클립보드의 Excel 'Link' 포맷에서 복사 소스(워크북/시트/범위)를 역추적한다.
    형식: b'Excel\\x00[BookName]SheetName\\x00R1C1:R2C2\\x00\\x00'. Ctrl+C 직후~붙여넣기 후까지 유지됨.
    반환: {"book","sheet","range"} 또는 None."""
    try:
        import win32clipboard
    except Exception:
        return None
    data = None
    opened = False
    for _ in range(8):  # 다른 프로세스가 잠깐 점유할 수 있어 재시도
        try:
            win32clipboard.OpenClipboard()
            opened = True
        except Exception:
            time.sleep(0.05)
            continue
        try:
            # 포맷을 열거하며 이름 == "Link"(Excel 소스참조) 를 찾는다.
            f = 0
            while True:
                f = win32clipboard.EnumClipboardFormats(f)
                if f == 0:
                    break
                try:
                    nm = win32clipboard.GetClipboardFormatName(f)
                except Exception:
                    nm = ""
                if nm == "Link":
                    try:
                        data = win32clipboard.GetClipboardData(f)
                    except Exception:
                        data = None
                    break
        except Exception:
            data = None
        finally:
            try:
                win32clipboard.CloseClipboard()
            except Exception:
                pass
        break
    if not opened or not data:
        return None
    if isinstance(data, bytes):
        text = None
        for enc in ("cp949", "mbcs", "utf-8", "latin-1"):
            try:
                text = data.decode(enc)
                break
            except Exception:
                continue
        if text is None:
            return None
    else:
        text = str(data)
    parts = text.split("\x00")
    if len(parts) < 3:
        return None
    # parts[1] = '[Book]Sheet' (미저장) 또는 'C:\\경로\\[Book.xlsx]Sheet' (저장된 파일) → [Book]Sheet 를 어디서든 찾는다.
    m = re.search(r"\[(.*?)\]([^\x00]*)$", parts[1])
    if not m:
        return None
    book = m.group(1).strip()
    sheet = m.group(2).strip()
    rng = _r1c1_to_a1(parts[2].strip())
    if not (book and sheet and rng):
        return None
    return {"book": book, "sheet": sheet, "range": rng}


def _maybe_snapshot_copy_source(app):
    """복사(Ctrl+C)로 CutCopyMode 가 켜져 있는 동안 클립보드 소스를 전역 스냅샷에 저장한다.
    교차파일 복붙은 복사(A)→탭전환→붙여넣기(B) 과정에서 클립보드 Link 가 사라져 캡처 시점엔
    소스를 못 읽는 경우가 있다(실측). 폴이 주기적으로 이걸 호출해 '복사 중'에 미리 잡아두면
    캡처가 폴백으로 복구할 수 있다. 폴마다 호출되지만 클립보드 OpenClipboard 는 throttle 로 제한한다."""
    try:
        cut_copy_active = bool(app.CutCopyMode)
        if not cut_copy_active:
            LAST_COPY_SOURCE["cutCopyActive"] = False
            return
    except Exception:
        return
    now = time.monotonic()
    try:
        was_active = bool(LAST_COPY_SOURCE.get("cutCopyActive"))
        last_poll = float(LAST_COPY_SOURCE.get("pollTs") or 0)
    except Exception:
        was_active = False
        last_poll = 0.0
    if was_active and now - last_poll < COPY_SOURCE_SNAPSHOT_THROTTLE_SECONDS:
        return
    LAST_COPY_SOURCE["cutCopyActive"] = True
    LAST_COPY_SOURCE["pollTs"] = now
    try:
        src = _read_excel_clipboard_source()
    except Exception:
        src = None
    if src:
        LAST_COPY_SOURCE["source"] = src
        LAST_COPY_SOURCE["ts"] = now


def _registered_path_for_name(name):
    """워크북 이름 → 업로드/세션 레지스트리의 파일 경로(교차파일 재생 시 소스 자동 열기용).
    활성 세션의 현재 경로(편집본 반영)를 우선하고, 없으면 업로드 레지스트리에서 찾는다."""
    if not name:
        return None
    want = _workbook_name_lookup_key(str(name))
    stem = str(Path(str(name)).stem)
    for sess in list(EXCEL_SESSIONS.values()):
        p = sess.get("path")
        if not p:
            continue
        try:
            if _workbook_name_lookup_key(Path(p).name) == want or str(Path(p).stem) == stem:
                return p
        except Exception:
            continue
    for rec in list(WORKBOOKS.values()):
        p = rec.get("path")
        nm = rec.get("name") or ""
        if not p:
            continue
        try:
            if (_workbook_name_lookup_key(nm) == want
                    or _workbook_name_lookup_key(Path(p).name) == want
                    or str(Path(nm or p).stem) == stem):
                return p
        except Exception:
            continue
    return None


def _capture_copypaste_on_session_impl(excel_id, values_only=False):
    """라이브 세션에서 '방금 한 복붙'을 캡처한다.
    소스 = 클립보드 Link(Ctrl+C 한 범위), 대상 = 현재 Selection(Ctrl+V 로 붙여진 범위).
    반환: 캡처 정보 + 그대로 스킬 스텝으로 쓸 ctx.paste_copied(...) Python 코드."""
    with EXCEL_LOCK:
        session = get_excel_session(excel_id)
        app, wb = session_workbook(session)
        try:
            open_books = [str(w.Name) for w in app.Workbooks]
        except Exception:
            open_books = None
        try:
            session_book = str(wb.Name)
        except Exception:
            session_book = None
        source = _read_excel_clipboard_source()
        # 클립보드 Link 가 이미 사라졌으면(특히 교차파일: 복사→탭전환→붙여넣기 후) 복사 중 폴이
        # 잡아둔 전역 스냅샷으로 복구한다.
        snap_used = False
        if not source:
            snap = LAST_COPY_SOURCE.get("source")
            ts = LAST_COPY_SOURCE.get("ts")
            if snap and ts is not None and (time.monotonic() - ts) < 120:
                source = snap
                snap_used = True
        # 진단 로그: 이 한 줄로 인스턴스 구성(open_books)·클립보드 소스·세션 책을 본다.
        _vba_trace("capture.copypaste.start", excelId=excel_id, sessionBook=session_book,
                   openBooks=open_books, clipboardSource=source, snapUsed=snap_used,
                   valuesOnly=bool(values_only))
        if not source:
            _vba_trace("capture.copypaste.reject", excelId=excel_id, reason="no-clipboard-source")
            raise RuntimeError(
                "복사한 내용을 찾지 못했습니다. 우측 엑셀에서 복사할 범위를 드래그해 Ctrl+C 한 뒤, "
                "붙여넣을 위치에 Ctrl+V 하고 곧바로 [복붙 저장]을 눌러 주세요(복사 후 다른 작업을 하면 사라집니다)."
            )
        # 대상 = 이 세션(excelId) 워크북에 사용자가 붙여넣은 위치.
        # 전역 app.Selection 은 교차파일/오버레이에서 '소스' 워크북을 가리킬 수 있어(붙여넣은 B가 아니라
        # 복사한 A의 선택이 잡힘) 신뢰하지 않는다 — 세션 워크북 자체 창의 RangeSelection 을 직접 읽는다.
        try:
            sel = None
            dst_via = None
            try:
                sel = wb.Windows(1).RangeSelection
                _ = sel.Worksheet  # 유효성 확인
                dst_via = "session-window"
            except Exception:
                sel = app.Selection
                dst_via = "global-selection"
            dst_ws = sel.Worksheet
            dst_sheet = str(dst_ws.Name)
            dst_book = str(dst_ws.Parent.Name)
            top = sel.Cells(1, 1)
            dst_cell = str(top.Address).replace("$", "")  # '$D$1' → 'D1'
            sel_rows = int(sel.Rows.Count)
            sel_cols = int(sel.Columns.Count)
        except Exception as e:
            _vba_trace("capture.copypaste.reject", excelId=excel_id, reason="dest-read-failed", error=str(e))
            raise RuntimeError("붙여넣은 위치(선택 영역)를 읽지 못했습니다: %s" % e)
        # 진단: 전역 app.Selection 이 가리키는 책 — 세션 창 결과와 다르면 오버레이 포커스 이슈 단서.
        try:
            global_sel_book = str(app.Selection.Worksheet.Parent.Name)
        except Exception:
            global_sel_book = None
        _vba_trace("capture.copypaste.dest", excelId=excel_id, via=dst_via,
                   dstBook=dst_book, dstSheet=dst_sheet, dstCell=dst_cell,
                   selRows=sel_rows, selCols=sel_cols, globalSelectionBook=global_sel_book)
        # 비연속 다중 영역(Ctrl-클릭) 복사는 재생이 불안정/모호하므로 거부 — 클립보드 범위 문자열의
        # 콤마로 판정한다(소스가 다른 인스턴스/미오픈이어도 인스턴스 무관하게 검사 가능).
        if "," in str(source.get("range", "")):
            _vba_trace("capture.copypaste.reject", excelId=excel_id, reason="multi-area",
                       range=source.get("range"))
            raise RuntimeError(
                "여러 영역을 한꺼번에 복사한 건 캡처할 수 없어요. 한 영역(연속 범위/열 전체/행 전체)씩 "
                "복사 → 붙여넣기 → [복붙 저장] 해 주세요."
            )
        # 소스 크기는 같은 인스턴스에 열려 있을 때만 검증(경고용). 다른 인스턴스/미오픈이면 검증을
        # 생략하되 저장은 허용한다 — 재생 시 업로드 경로에서 자동으로 열어 복사하므로(교차파일).
        src_rows = src_cols = None
        src_found = False
        try:
            want = _workbook_name_lookup_key(source["book"])
            src_ws = None
            for owb in app.Workbooks:
                if _workbook_name_lookup_key(str(owb.Name)) == want:
                    src_ws = owb.Worksheets(source["sheet"])
                    break
            if src_ws is not None:
                src_found = True
                srng = src_ws.Range(source["range"])
                src_rows = int(srng.Rows.Count)
                src_cols = int(srng.Columns.Count)
        except Exception:
            src_rows = src_cols = None
        same = _workbook_name_lookup_key(source["book"]) == _workbook_name_lookup_key(dst_book)
        _vba_trace("capture.copypaste.source", excelId=excel_id, srcBook=source["book"],
                   srcFoundInInstance=src_found, srcRows=src_rows, srcCols=src_cols,
                   crossFile=(not same))
        # 대상이 단일 셀이면 '붙여넣기 기준점'이므로 정상(ctx.paste_copied 가 소스 크기만큼 자동 확장).
        # 다중 셀을 선택했는데 소스 크기와 다르면 오캡처 의심 → 경고(소스 크기 모르면 검증 생략).
        if not src_rows:
            dims_match = None
        elif sel_rows == 1 and sel_cols == 1:
            dims_match = True
        else:
            dims_match = (src_rows == sel_rows and src_cols == sel_cols)
        # [SBAGENT-209] 코드에 박는 워크북 이름은 반드시 '사용자 파일명'이어야 한다. 위장 파일(.xls=OLE/HTML)은
        # 라이브 wb.Name 이 excel_open_<hash>.xls 라서, 그대로 저장하면 세션이 죽은 뒤 어떤 파일과도 매칭 불가 —
        # 실행기 파일확인에 영원히 못 채우는 요구 행이 뜨고 재생도 깨진다. 역별칭으로 원본명을 되찾아 쓴다.
        src_book_out = _user_facing_workbook_name_for_live(app, source["book"])
        dst_book_out = _user_facing_workbook_name_for_live(app, dst_book)
        if src_book_out != source["book"] or dst_book_out != dst_book:
            _vba_trace("capture.copypaste.display_names", excelId=excel_id,
                       srcLive=source["book"], srcOut=src_book_out,
                       dstLive=dst_book, dstOut=dst_book_out)
        if values_only:
            step_code = (
                "def transform(ctx):\n"
                "    # [복붙 캡처] 사용자가 라이브 Excel에서 직접 복사/붙여넣기한 동작 재현(값만 붙여넣기)\n"
                "    ctx.paste_copied(%r, %r, %r, %r, src_book=%r, dst_book=%r, values_only=True)\n"
                % (source["sheet"], source["range"], dst_sheet, dst_cell, src_book_out, dst_book_out)
            )
        else:
            step_code = (
                "def transform(ctx):\n"
                "    # [복붙 캡처] 사용자가 라이브 Excel에서 직접 복사/붙여넣기한 동작 재현(값+수식+서식 보존)\n"
                "    ctx.paste_copied(%r, %r, %r, %r, src_book=%r, dst_book=%r)\n"
                % (source["sheet"], source["range"], dst_sheet, dst_cell, src_book_out, dst_book_out)
            )
        desc = "%s복붙: %s!%s → %s!%s%s" % (
            "값만 " if values_only else "",
            source["sheet"], source["range"], dst_sheet, dst_cell,
            "" if same else " (교차파일)",
        )
        _vba_trace("capture.copypaste.result", excelId=excel_id, ok=True, description=desc,
                   dimsMatch=dims_match, crossFile=(not same), valuesOnly=bool(values_only))
        return {
            "ok": True,
            "source": source,
            "dest": {"book": dst_book, "sheet": dst_sheet, "cell": dst_cell,
                     "rows": sel_rows, "cols": sel_cols},
            "dimsMatch": dims_match,
            "language": "python",
            "code": step_code,
            "description": desc,
            "valuesOnly": bool(values_only),
        }


def run_capture_copypaste(excel_id, values_only=False):
    return excel_call(_capture_copypaste_on_session_impl, excel_id, bool(values_only), timeout=60)


# =====================================================================
# Python COM 스킬 엔진 (ver0.5.2 4단계 — openpyxl 아님, 라이브 Excel COM bulk 제어)
#
# 강제 구조 4겹:
#   L1 API 표면: 생성코드에 win32com/Application 을 주지 않고 벌크 전용 ctx 만 노출
#      → 셀 단위 COM 루프는 '작성 자체가 불가능'(쓰기 프리미티브가 벌크뿐).
#   L2 AST 정적 게이트: 실행 전 구조 분석(루프 내 ctx 쓰기, 금지 import/빌트인,
#      Select/Activate/ActiveWorkbook, openpyxl 관용구).
#   L3 런타임 가드: COM 호출 예산 + 데드라인(트레이서) + 쓰기 저널(실패 시 정밀 롤백)
#      + finally 의 앱 상태 복구. 저널이 곧 변경 기록이라 별도 풀스냅샷 지문이 불필요.
#   L4 프롬프트: PYTHON_COM_SYSTEM_PROMPT (file-schema.js) — ctx API 레퍼런스 + few-shot.
# =====================================================================

PY_SKILL_ENTRY = "transform"
PY_COM_BUDGET = int(os.environ.get("B2B_PY_COM_BUDGET", "400"))
# [사용자 지시] Python COM 실행 타임아웃/셀 제한을 기본 '무제한(0)'으로 둔다. 60만 행 같은 정상 대용량 작업이
# 데드라인(75초)/셀 상한(6M)에 걸려 안 돌던 문제 → 0=무제한이 기본. (필요하면 env 로 유한값을 다시 줄 수 있다.)
# 안전: 무제한이면 무한 루프 코드가 STA 워커를 영구 정지시킬 수 있으나(앱 재시작 필요) while True 등은 정적검사가,
# import/파일접근은 런타임 샌드박스(_PY_SAFE_BUILTINS/safe_globals)가 막는다.
PY_SKILL_TIMEOUT_S = float(os.environ.get("B2B_PY_SKILL_TIMEOUT", "0"))                    # 0 = 무제한(데드라인 없음)
PY_SKILL_RECOVERY_TIMEOUT_S = float(os.environ.get("B2B_PY_SKILL_RECOVERY_TIMEOUT", "0"))  # 0 = 무제한
PY_READ_MAX_CELLS = int(os.environ.get("B2B_PY_READ_MAX_CELLS", "0"))                      # 0 = 무제한(셀 상한 없음)
# excel_call 등 '바깥' 큐 타임아웃에 쓸 '사실상 무제한' 유한값(무제한 설정 시). 30일.
PY_UNLIMITED_OUTER_S = 30 * 24 * 3600


def _py_skill_deadline(timeout_s=None):
    """Python 스킬 실행 데드라인(monotonic). 유효 타임아웃이 0/음수면 무제한(inf)."""
    eff = float(timeout_s) if timeout_s else PY_SKILL_TIMEOUT_S
    return float("inf") if eff <= 0 else (time.monotonic() + eff)


def _promote_csv_multisheet_name(name, wb):
    """파일명이 .csv/.tsv 인데 워크북에 시트가 2개 이상이면 .xlsx 로 바꾼다.
    CSV 는 시트를 하나만 담으므로, 스킬이 새 시트(예: 필터→새시트)를 추가한 결과를 .csv 로 저장하면
    시트가 붕괴(기존 시트 덮어써짐)한다. 이 경우 .xlsx 로 승격해 저장측이 SaveAs(FileFormat=51) 하게 한다."""
    try:
        p = Path(str(name))
        if p.suffix.lower() in (".csv", ".tsv") and int(wb.Worksheets.Count) > 1:
            stem = p.stem
            # 이중 확장자("a.xlsx.csv")면 stem 이 "a.xlsx" 라 그대로 붙이면 "a.xlsx.xlsx" 가 된다.
            # 이미 스프레드시트 확장자로 끝나면 덧붙이지 않는다(포맷은 어차피 51 로 저장).
            if Path(stem).suffix.lower() in (".xlsx", ".xlsm", ".xlsb", ".xls"):
                return stem
            return stem + ".xlsx"
    except Exception:
        pass
    return str(name)

_PY_SAFE_BUILTINS = {
    "len": len, "range": range, "enumerate": enumerate, "zip": zip,
    "min": min, "max": max, "sum": sum, "sorted": sorted, "reversed": reversed,
    "abs": abs, "round": round, "any": any, "all": all, "isinstance": isinstance,
    "str": str, "int": int, "float": float, "bool": bool,
    "list": list, "dict": dict, "set": set, "tuple": tuple,
    # 열 문자 계산(chr(65+...)/divmod)·문자 코드 변환은 생성 코드가 흔히 쓰는 순수 함수 —
    # 빠져 있으면 "name 'chr' is not defined" 런타임 실패가 난다.
    "chr": chr, "ord": ord, "divmod": divmod, "map": map, "filter": filter,
    "Exception": Exception, "ValueError": ValueError, "TypeError": TypeError,
    "KeyError": KeyError, "IndexError": IndexError, "RuntimeError": RuntimeError,
    "print": (lambda *a, **k: None),  # 출력은 무시(자동 실행 차단 요소 없음)
    "True": True, "False": False, "None": None,
}

_XL_UP = -4162
_XL_TO_LEFT = -4159


class PythonComSkillError(RuntimeError):
    pass


# 서식 헬퍼용 색 이름(한글/영문) → 16진. Excel .Color 는 RGB() 롱값(r + g*256 + b*65536)이다.
_COLOR_NAMES = {
    "빨강": "FF0000", "빨간": "FF0000", "적색": "FF0000", "red": "FF0000",
    "주황": "FFA500", "주황색": "FFA500", "orange": "FFA500",
    "노랑": "FFFF00", "노란": "FFFF00", "황색": "FFFF00", "yellow": "FFFF00",
    "연두": "90EE90", "연두색": "90EE90", "lightgreen": "90EE90",
    "초록": "008000", "녹색": "008000", "green": "008000",
    "하늘": "87CEEB", "하늘색": "87CEEB", "skyblue": "87CEEB",
    "파랑": "0000FF", "파란": "0000FF", "청색": "0000FF", "blue": "0000FF",
    "남색": "000080", "navy": "000080",
    "보라": "800080", "보라색": "800080", "purple": "800080",
    "분홍": "FFC0CB", "핑크": "FFC0CB", "pink": "FFC0CB",
    "갈색": "A52A2A", "brown": "A52A2A",
    "연회색": "D3D3D3", "밝은회색": "D3D3D3", "lightgray": "D3D3D3", "lightgrey": "D3D3D3",
    "회색": "808080", "그레이": "808080", "gray": "808080", "grey": "808080",
    "검정": "000000", "검은": "000000", "흑색": "000000", "black": "000000",
    "흰색": "FFFFFF", "하양": "FFFFFF", "white": "FFFFFF",
}


def _parse_excel_color(c):
    """색 입력을 Excel .Color 롱값으로 변환. '#RRGGBB'/'RRGGBB'/색이름(노랑·red 등)/정수 지원.
    None 또는 빈 값이면 None(호출자가 '채우기 없음' 등으로 처리). 이해 못하면 PythonComSkillError."""
    if c is None:
        return None
    if isinstance(c, bool):
        return None
    if isinstance(c, int):
        return int(c)
    s = str(c).strip()
    if not s:
        return None
    named = _COLOR_NAMES.get(s.lower())
    if named:
        s = named
    s = s.lstrip("#").strip()
    if re.fullmatch(r"[0-9A-Fa-f]{6}", s):
        r = int(s[0:2], 16)
        g = int(s[2:4], 16)
        b = int(s[4:6], 16)
        return r + g * 256 + b * 65536
    raise PythonComSkillError(f"색상 값을 이해하지 못했습니다: {c!r} (예: '#FFFF00' 또는 '노랑')")


class PythonComSkillContext:
    """생성된 Python 스킬에 노출되는 유일한 능력(capability).

    - 모든 읽기/쓰기는 Range.Value2/Formula 벌크 1회 호출로만 수행된다.
    - 쓰기 전 대상 범위의 기존 수식(Formula)을 저널에 백업 → 실패 시 그 범위만 정밀 롤백.
    - 매 연산마다 COM 예산과 데드라인을 검사해 저사양 PC 의 COM 폭주를 차단한다.
    - 대상 워크북은 세션에 고정(pinned)되어 ActiveWorkbook 에 의존하지 않는다.
    """

    def __init__(self, app, wb, session, _shared=None, timeout_s=None):
        self._app = app
        self._wb = wb
        self._session = session
        if _shared is None:
            _shared = {
                "com_calls": 0,
                "deadline": _py_skill_deadline(timeout_s),
                "journal": [],          # (ws_name, address, formulas_2d)
                "structural": [],       # 롤백 불가 구조 변경 설명 목록
                "books": {},
                # [SBAGENT-209 후속] 대상 시트에서 '비어있지 않은 데이터'를 실제로 읽었는가.
                # 조건부 스킬(예: '소계 행이 있으면 삭제')이 이번 파일엔 해당 없어 아무것도 안 바꾼
                # '정상 무변경'과, 엉뚱한 시트를 읽어 아무것도 못 한 '오타겟 무변경'을 구분하는 증거.
                "read_nonempty": False,
                # [0건 매칭 노출] ctx.write 가 기록한 총 셀 수/비어있지 않은 값 수. 전량 빈값이면
                # '조건에 맞는 데이터 0건' 신호 — 실행은 성공이라도 응답에 경고를 실어 조용한 실패를 드러낸다.
                "write_cells_total": 0,
                "write_cells_nonempty": 0,
            }
        self._shared = _shared

    # ---- 내부 가드 ----
    def _tick(self, n=1):
        self._shared["com_calls"] += n
        if self._shared["com_calls"] > PY_COM_BUDGET:
            raise PythonComSkillError(
                f"COM 호출 예산({PY_COM_BUDGET}회)을 초과했습니다. 셀 단위 반복 대신 "
                "ctx.read()/ctx.write() 벌크 호출로 다시 작성하세요."
            )
        if time.monotonic() > self._shared["deadline"]:
            raise PythonComSkillError("Python 스킬 실행 시간이 초과되었습니다.")

    def _ws(self, sheet):
        self._tick(1)
        try:
            ws = self._wb.Worksheets(str(sheet))
            _ = ws.Name
            return ws
        except Exception:
            pass
        try:
            names = _excel_collection_names(self._wb.Worksheets)
            alias = _resolve_ephemeral_excel_open_sheet_alias(sheet, names)
            if alias:
                _vba_trace("python_com.sheet.alias", requested=str(sheet), resolved=alias, reason="ephemeral-excel-open")
                return self._wb.Worksheets(alias)
        except Exception:
            pass
        # 현재 워크북에 없으면 같은 Excel 인스턴스의 다른 열린 워크북에서 찾는다 —
        # 교차파일 스킬이 ctx.book() 없이 시트명만 쓴 경우나, 저장된 스킬의 파일 바인딩이
        # 유실된 채 재실행되는 경우를 구제(안전망). 단 그 시트명이 '정확히 한 워크북'에만
        # 있을 때만 따라간다(여러 곳에 있으면 모호하므로 기존 에러로 둔다).
        try:
            matches = []
            for owb in self._app.Workbooks:
                try:
                    if str(sheet) in _excel_collection_names(owb.Worksheets):
                        matches.append(owb)
                except Exception:
                    continue
            if len(matches) == 1:
                return matches[0].Worksheets(str(sheet))
        except Exception:
            pass
        # 공백/언더스코어/하이픈만 다른 시트명도 매칭(모델이 '2026년'을 '2026 년'으로 공백 끼우는 흔한 케이스).
        # normalize_sheet_lookup 으로 정규화해 '정확히 한 곳'에만 있을 때 따라간다.
        try:
            want = normalize_sheet_lookup(sheet)
            cur = _excel_collection_names(self._wb.Worksheets)
            nm = [n for n in cur if normalize_sheet_lookup(n) == want]
            if len(nm) == 1:
                return self._wb.Worksheets(nm[0])
            cross = []
            for owb in self._app.Workbooks:
                try:
                    for n in _excel_collection_names(owb.Worksheets):
                        if normalize_sheet_lookup(n) == want:
                            cross.append((owb, n))
                except Exception:
                    continue
            if len(cross) == 1:
                return cross[0][0].Worksheets(cross[0][1])
        except Exception:
            pass
        # [월/날짜만 다른 시트명] 시트 이름에 월/날짜가 들어간 경우(예 '2026년04월정산'), 안정 키(월/날짜 제거)로
        # 현재 워크북에서 '유일' 매칭이면 그 시트로. 모호하면 매칭 안 함(아래 오류 유지).
        try:
            cur = _excel_collection_names(self._wb.Worksheets)
            st = _match_workbook_by_stable_key(cur, sheet)
            if st:
                _vba_trace("python_com.sheet.stable_key_match", requested=str(sheet), matched=str(st))
                return self._wb.Worksheets(st)
        except Exception:
            pass
        # [CSV/긴 파일명] Excel 시트명은 31자 제한이라, 파일명(stem)으로 시트가 만들어지는 CSV 등에서
        # 모델이 풀네임(>31자)을 시트명으로 쓰면 실제 시트(31자로 잘린 이름)와 어긋난다.
        # 요청명을 31자로 자른 값(정확/정규화 일치) 또는 접두 관계로 '유일' 매칭이면 그 시트로 따라간다.
        try:
            req = str(sheet)
            cur = _excel_collection_names(self._wb.Worksheets)
            trunc = req[:31]
            cand = [n for n in cur
                    if n == trunc or normalize_sheet_lookup(n) == normalize_sheet_lookup(trunc)]
            if not cand:
                rn = normalize_sheet_lookup(req)

                def _pref(n):
                    nn = normalize_sheet_lookup(n)
                    return bool(rn) and bool(nn) and (rn.startswith(nn) or nn.startswith(rn))

                cand = [n for n in cur if _pref(n)]
            uniq = list(dict.fromkeys(cand))
            if len(uniq) == 1:
                _vba_trace("python_com.sheet.truncated_match", requested=req, matched=uniq[0])
                return self._wb.Worksheets(uniq[0])
        except Exception:
            pass
        # [단일 시트 안전망] 워크북에 시트가 하나뿐이면(대표적으로 CSV) 요청 시트명이 어긋나도 그 시트로.
        # 여러 시트면 모호하므로 아래 오류를 유지한다(엉뚱한 시트에 쓰는 사고 방지).
        try:
            only = _excel_collection_names(self._wb.Worksheets)
            if len(only) == 1:
                _vba_trace("python_com.sheet.single_fallback", requested=str(sheet), matched=only[0])
                return self._wb.Worksheets(only[0])
        except Exception:
            pass
        names = _excel_collection_names(self._wb.Worksheets)
        raise PythonComSkillError(
            f"시트 '{sheet}' 를 찾지 못했습니다. 사용 가능한 시트: {names}"
        )

    @staticmethod
    def _col_num(letters):
        n = 0
        for ch in str(letters).upper():
            n = n * 26 + (ord(ch) - 64)
        return n

    def _resize_rng(self, ws, anchor, rows, cols):
        """anchor 셀에서 rows×cols 명시 범위를 만든다.
        [핵심 버그 수정] 동적 디스패치(DispatchEx)에서 Range.Resize(r, c)가 파라미터 프로퍼티로
        잘못 해석돼 단일 셀(오프셋 위치)을 반환한다 — write 가 '마지막 한 칸'에만 기록되던 원인.
        Resize 를 쓰지 않고 시작 행/열에서 끝 주소를 직접 계산한다."""
        r0 = int(anchor.Row)
        c0 = int(anchor.Column)
        return ws.Range(f"{_col_letter(c0)}{r0}:{_col_letter(c0 + int(cols) - 1)}{r0 + int(rows) - 1}")

    @staticmethod
    def _shaped_matrix(rng, value):
        """_range_matrix 가 빈/None(전부 빈 셀) 결과를 줄 때도 범위 차원을 보존해
        [[None]*cols]*rows 를 돌려준다 — 빈 단일 셀 read 가 [] 로 줄어들어
        호출 코드의 m[0][0] 이 IndexError 나던 문제 방지."""
        m = _range_matrix(value)
        if m:
            return m
        try:
            rows = int(rng.Rows.Count)
            cols = int(rng.Columns.Count)
        except Exception:
            rows = cols = 1
        return [[None] * max(1, cols) for _ in range(max(1, rows))]

    def _rng(self, ws, a1):
        ref = str(a1)
        # 뒤집힌 범위("G1:F100", "B10:B5")는 Excel 이 조용히 정규화해 의도와 다른 폭/높이로
        # 동작한다(예: 6열 복사가 2열 복사가 됨) — 생성 코드의 f-string 실수를 즉시 드러낸다.
        m = re.match(r"^\$?([A-Za-z]{1,3})\$?(\d+)?:\$?([A-Za-z]{1,3})\$?(\d+)?$", ref.strip())
        if m:
            c1, r1, c2, r2 = m.group(1), m.group(2), m.group(3), m.group(4)
            if self._col_num(c1) > self._col_num(c2) or (r1 and r2 and int(r1) > int(r2)):
                raise PythonComSkillError(
                    f"범위 표기가 뒤집혔습니다: '{ref}' — 시작 셀이 끝 셀보다 뒤에 있습니다. "
                    f"'{c2}{r2 or ''}:{c1}{r1 or ''}' 처럼 시작:끝 순서로 고쳐 주세요."
                )
        try:
            return ws.Range(ref)
        except Exception:
            raise PythonComSkillError(f"잘못된 범위 주소입니다: '{a1}' (예: \"B2:D100\")")

    @staticmethod
    def _as_2d(values):
        if not isinstance(values, (list, tuple)) or not values:
            raise PythonComSkillError("write() 값은 2차원 리스트여야 합니다. 예: [[1,2],[3,4]]")
        rows = []
        width = None
        for row in values:
            if not isinstance(row, (list, tuple)):
                raise PythonComSkillError("write() 값은 2차원 리스트여야 합니다(각 행도 리스트).")
            if width is None:
                width = len(row)
            elif len(row) != width:
                raise PythonComSkillError("write() 의 모든 행은 같은 길이여야 합니다.")
            rows.append(tuple("" if v is None else v for v in row))
        if width == 0:
            raise PythonComSkillError("write() 에 빈 행을 전달할 수 없습니다.")
        return tuple(rows), len(rows), width

    def _journal_save(self, ws, rng):
        try:
            address = str(rng.Address)
            formulas = _range_matrix(rng.Formula)
            self._shared["journal"].append((str(ws.Name), address, formulas))
            self._tick(2)
        except Exception:
            # 저널 실패는 실행을 막지 않는다(롤백 불가로만 기록).
            self._shared["structural"].append("journal-save-failed")

    # ---- 조회 ----
    def sheets(self):
        """시트 이름 목록."""
        self._tick(1)
        return _excel_collection_names(self._wb.Worksheets)

    def used_range(self, sheet):
        """(행수, 열수) — 시트의 사용 범위 크기."""
        ws = self._ws(sheet)
        used = ws.UsedRange
        self._tick(2)
        return int(used.Rows.Count), int(used.Columns.Count)

    def last_row(self, sheet, col=1):
        """해당 열 기준 마지막 데이터 행(1-based). 표 끝 합계행 포함 여부는 호출자가 판단."""
        t0 = time.perf_counter()
        ws = self._ws(sheet)
        self._tick(2)
        row = int(ws.Cells(ws.Rows.Count, int(col)).End(_XL_UP).Row)
        try:
            ms = (time.perf_counter() - t0) * 1000
            if ms >= 200:
                _vba_trace("python_com.last_row", sheet=str(sheet), col=int(col), row=row, ms=round(ms, 1))
        except Exception:
            pass
        return row

    def last_col(self, sheet, row=1):
        """해당 행 기준 마지막 데이터 열(1-based)."""
        ws = self._ws(sheet)
        self._tick(2)
        return int(ws.Cells(int(row), ws.Columns.Count).End(_XL_TO_LEFT).Column)

    def used_last_row(self, sheet):
        """시트 '사용 범위' 마지막 행(1-based). 특정 열 기준 last_row(col=N) 은 그 열이 희소/병합이면 표 하단을
        놓쳐 과소산정한다(예: A열이 아래쪽 비어 22 를 주지만 실제 표는 28행). '시트 전체/사용 범위'를 복사·처리할
        땐 이걸 쓴다. UsedRange 와 각 열 End(xlUp) 최대 중 큰 값으로 보수적으로 잡는다(정형 표에 안전)."""
        ws = self._ws(sheet)
        self._tick(2)
        best = 1
        try:
            best = int(ws.UsedRange.Row) + int(ws.UsedRange.Rows.Count) - 1
        except Exception:
            pass
        try:
            last_c = self.used_last_col(sheet)
            rows_n = int(ws.Rows.Count)
            for c in range(1, min(last_c, 64) + 1):   # 상한 64열(정형 표 커버, 성능 보호)
                r = int(ws.Cells(rows_n, c).End(_XL_UP).Row)
                if r > best:
                    best = r
        except Exception:
            pass
        return max(1, best)

    def used_last_col(self, sheet):
        """시트 '사용 범위' 마지막 열(1-based). 특정 행 기준 last_col 이 그 행 병합/빈칸으로 과소산정하는 것 방지."""
        ws = self._ws(sheet)
        self._tick(2)
        try:
            return max(1, int(ws.UsedRange.Column) + int(ws.UsedRange.Columns.Count) - 1)
        except Exception:
            return 1

    def first_empty_col(self, sheet, after=None, header_row=1):
        """'빈 보조열'을 찾아 그 열 '문자'(예 "N")를 돌려준다. after(마지막 데이터 열, 예 "L")를 주면 그 다음
        열부터, 안 주면 데이터가 시작된 뒤부터 오른쪽으로 훑어 '헤더행~사용범위 마지막 행'까지 완전히 빈 첫
        열을 찾는다. "옆 빈 열/빈 보조열에 계산해 넣어줘" 요청에서 데이터 옆 첫 칸이 이미 합계(=SUM) 등으로
        차 있어 그걸 덮어쓰는 사고를 막는다(그 열은 건너뛰고 진짜 빈 열을 고른다).
          예: col = ctx.first_empty_col("SO사업자별요금", after="L")   # -> "N" (M 은 합계열이라 건너뜀)
              ctx.write_cell(sheet, f"{col}8", "행별MAX"); ctx.write(sheet, f"{col}9", out)"""
        ws = self._ws(sheet)
        self._tick(2)
        hr = max(1, int(header_row or 1))
        last_r = max(hr, self.used_last_row(sheet))
        last_c = max(1, self.used_last_col(sheet))

        def _col_empty(c):
            try:
                vals = ws.Range(ws.Cells(hr, c), ws.Cells(last_r, c)).Value
            except Exception:
                return True
            if not isinstance(vals, (tuple, list)):
                return vals in (None, "")
            for row in vals:
                v = row[0] if isinstance(row, (tuple, list)) else row
                if v not in (None, ""):
                    return False
            return True

        start = 1
        if after is not None:
            s = str(after).strip()
            try:
                start = (self._col_index(s) if re.fullmatch(r"[A-Za-z]{1,3}", s) else int(s)) + 1
            except Exception:
                start = 1
        seen_data = after is not None   # after 를 주면 이미 데이터 뒤로 간주(선행 빈 열 무시)
        c = max(1, int(start))
        while c <= last_c + 2:
            if _col_empty(c):
                if seen_data:
                    return _col_letter(c)
            else:
                seen_data = True
            c += 1
        return _col_letter(last_c + 1)

    def find_header(self, sheet, header_text, header_row=1):
        """헤더 행에서 헤더 텍스트로 열 번호(1-based)를 찾는다. 없으면 오류.
        열 번호를 추측/하드코딩하지 말고 반드시 이 함수를 쓸 것.
        지정 행에서 못 찾으면 '인접 행(±)'도 훑는다 — 2행 병합/멀티행 헤더에서 하위 헤더가 다른 행에
        있는 경우(예: 4행 '매 출 액' 병합 아래 5행 '국제')를 구제한다. (사용자가 열문자를 직접 주면
        find_header 로 다시 찾지 말고 그 열을 그대로 쓰는 게 가장 안전하다.)"""
        ws = self._ws(sheet)
        row = int(header_row)
        last_col = self.last_col(sheet, row)
        rng = ws.Range(ws.Cells(row, 1), ws.Cells(row, max(1, last_col)))
        self._tick(2)
        values = _range_matrix(rng.Value2)
        target = str(header_text).strip()
        headers = [str(v).strip() if v is not None else "" for v in (values[0] if values else [])]
        for idx, text in enumerate(headers, start=1):
            if text == target:
                return idx
        # 정규화 비교: 공백·표기 차이("3 월" == "3월", 전각/대소문자)로 LLM 이 헤더에 공백을
        # 끼워 넣어도 매칭한다(정확 매칭 다음, 느슨한 부분포함 전에 둔다).
        ntarget = normalize_text(target)
        if ntarget:
            for idx, text in enumerate(headers, start=1):
                if normalize_text(text) == ntarget:
                    return idx
        # 인접 행 스캔(2행 병합/멀티행 헤더 구제): 지정 행 위/아래를 한 번에 읽어 정확→정규화 순으로,
        # 지정 행에 '가까운 행부터' 매칭한다(오탐 최소화: 부분포함은 인접 행엔 적용하지 않는다).
        top = max(1, row - 4)
        bottom = row + 2
        wcols = max(1, min(int(self.used_last_col(sheet)), 256))
        try:
            wrng = ws.Range(ws.Cells(top, 1), ws.Cells(bottom, wcols))
            self._tick(2)
            wmatrix = _range_matrix(wrng.Value2)
        except Exception:
            wmatrix = []
        order = sorted(range(len(wmatrix)), key=lambda i: abs((top + i) - row))
        for pass_norm in (False, True):
            for i in order:
                if top + i == row:
                    continue   # 지정 행은 위에서 이미 확인함
                cells = wmatrix[i] or []
                for idx, v in enumerate(cells, start=1):
                    t = str(v).strip() if v is not None else ""
                    if not t:
                        continue
                    if not pass_norm and t == target:
                        return idx
                    if pass_norm and ntarget and normalize_text(t) == ntarget:
                        return idx
        # 마지막: 지정 행 부분포함(가장 느슨)
        for idx, text in enumerate(headers, start=1):
            if target and target in text:
                return idx
        raise PythonComSkillError(
            f"'{sheet}' 시트 {row}행 및 인접 행에서 헤더 '{header_text}' 를 찾지 못했습니다. 실제 헤더: {headers}"
        )

    def read(self, sheet, a1_range=None):
        """범위를 2차원 리스트로 한 번에 읽는다(COM 1회). a1_range 생략 시 used range."""
        t0 = time.perf_counter()
        ws = self._ws(sheet)
        rng = self._rng(ws, a1_range) if a1_range else ws.UsedRange
        self._tick(3)
        cells = int(rng.Rows.Count) * int(rng.Columns.Count)
        if PY_READ_MAX_CELLS > 0 and cells > PY_READ_MAX_CELLS:
            raise PythonComSkillError(
                f"읽기 범위가 너무 큽니다({cells:,}셀 > {PY_READ_MAX_CELLS:,}). "
                "Python COM은 단순 작업용으로 보수적으로 제한됩니다. 범위를 더 좁히거나 VBA 경로를 사용하세요."
            )
        out = self._shaped_matrix(rng, rng.Value2)
        self._note_read_evidence(out)
        try:
            ms = (time.perf_counter() - t0) * 1000
            if ms >= 200 or cells >= 100000:
                _vba_trace(
                    "python_com.read",
                    sheet=str(sheet),
                    range=str(a1_range or "UsedRange"),
                    cells=cells,
                    rows=int(rng.Rows.Count),
                    cols=int(rng.Columns.Count),
                    ms=round(ms, 1),
                )
        except Exception:
            pass
        return out

    def read_cell(self, sheet, a1):
        """단일 셀 읽기(write_cell 와 대칭). 스칼라 값 반환(빈 셀은 None).
        모델이 write_cell 대칭으로 read_cell 을 자주 호출하는데 없어서 실패하던 것을 메운다."""
        m = self.read(sheet, a1)
        return m[0][0] if (m and m[0]) else None

    def read_formulas(self, sheet, a1_range):
        """범위의 수식 문자열을 2차원 리스트로 읽는다(수식 없는 셀은 값)."""
        ws = self._ws(sheet)
        rng = self._rng(ws, a1_range)
        self._tick(3)
        cells = int(rng.Rows.Count) * int(rng.Columns.Count)
        if PY_READ_MAX_CELLS > 0 and cells > PY_READ_MAX_CELLS:
            raise PythonComSkillError(
                f"읽기 범위가 너무 큽니다({cells:,}셀 > {PY_READ_MAX_CELLS:,}). "
                "Python COM은 단순 작업용으로 보수적으로 제한됩니다. 범위를 더 좁히거나 VBA 경로를 사용하세요."
            )
        out = self._shaped_matrix(rng, rng.Formula)
        self._note_read_evidence(out)
        return out

    def _note_read_evidence(self, matrix):
        """읽은 매트릭스에 비어있지 않은 값이 하나라도 있으면 '실데이터를 읽었다'는 증거를 남긴다.
        (any 는 첫 값에서 조기 종료 — 큰 읽기에도 추가 비용 미미. 무변경 게이트가 이 증거로
        '조건 미해당 정상 무변경'과 '오타겟 무변경'을 구분한다.)"""
        if self._shared.get("read_nonempty"):
            return
        try:
            if any(v is not None and str(v).strip() != "" for row in (matrix or []) for v in row):
                self._shared["read_nonempty"] = True
        except Exception:
            pass

    def has_formulas(self, sheet, a1_range):
        """범위에 수식이 하나라도 있으면 True."""
        ws = self._ws(sheet)
        rng = self._rng(ws, a1_range)
        self._tick(2)
        has = rng.HasFormula
        return has is not False

    def formula_mask(self, sheet, a1_range):
        """셀별 수식 여부를 2차원 리스트(True/False)로 반환(COM 1회).
        has_formulas 는 범위 '전체'에 대한 단일 bool 이므로 셀별 판단에는 이 메서드를 쓸 것."""
        ws = self._ws(sheet)
        rng = self._rng(ws, a1_range)
        self._tick(2)
        f = self._shaped_matrix(rng, rng.Formula)
        return [[isinstance(v, str) and v.startswith("=") for v in row] for row in f]

    # ---- 쓰기(벌크 전용) ----
    def write(self, sheet, a1_start, values, overwrite_formulas=True):
        """2차원 리스트를 시작 셀 기준으로 한 번에 쓴다(COM 1회).
        0.5.9부터 요청받은 대상 범위는 기본적으로 값으로 덮어쓴다. 수식 보존은
        생성 코드가 데이터 범위/요약 행을 정확히 제외하는 방식으로 처리한다."""
        t0 = time.perf_counter()
        ws = self._ws(sheet)
        data, rows, cols = self._as_2d(values)
        try:
            self._shared["write_cells_total"] += int(rows) * int(cols)
            self._shared["write_cells_nonempty"] += sum(
                1 for _row in data for _v in _row if _v is not None and str(_v).strip() != ""
            )
        except Exception:
            pass
        anchor = self._rng(ws, a1_start)
        rng = self._resize_rng(ws, anchor, rows, cols)
        self._tick(3)
        self._journal_save(ws, rng)
        try:
            _apply_com_text_format_for_long_digit_columns(ws, data, int(anchor.Row), int(anchor.Column))
        except Exception:
            pass
        rng.Value2 = data
        self._tick(1)
        try:
            ms = (time.perf_counter() - t0) * 1000
            cells = int(rows) * int(cols)
            if ms >= 200 or cells >= 100000:
                _vba_trace(
                    "python_com.write",
                    sheet=str(sheet),
                    start=str(a1_start),
                    cells=cells,
                    rows=int(rows),
                    cols=int(cols),
                    ms=round(ms, 1),
                )
        except Exception:
            pass
        return rows * cols

    def write_cell(self, sheet, a1, value, overwrite_formulas=True):
        """단일 셀 쓰기(소량 전용 — 루프에서 반복 호출하면 예산 초과로 차단됨)."""
        return self.write(sheet, a1, [[value]], overwrite_formulas=overwrite_formulas)

    def write_formulas(self, sheet, a1_start, formulas):
        """수식 문자열 2차원 리스트를 한 번에 기록(예: [["=B2-C2"],["=B3-C3"]])."""
        ws = self._ws(sheet)
        data, rows, cols = self._as_2d(formulas)
        anchor = self._rng(ws, a1_start)
        rng = self._resize_rng(ws, anchor, rows, cols)
        self._tick(3)
        self._journal_save(ws, rng)
        rng.Formula = data
        self._tick(1)
        return rows * cols

    def _ctx_and_sheet_from_spec(self, sheet_spec):
        """Resolve "workbook.xlsx!Sheet1" into a context + sheet name.

        LLMs naturally write ctx.copy("file.xlsx!sheet", ...). Before this helper,
        _ws() tried to find a literal sheet with that whole string, failed, and
        recovery often rewrote the task into slow value-array code. Supporting the
        common Excel reference shape keeps cross-file copy inside native COM
        Range.Copy instead of Python read/write loops.
        """
        text = str(sheet_spec or "").strip()
        if not text:
            return self, sheet_spec
        # Excel external refs are often written as [Book.xlsx]Sheet.
        m = re.match(r"^\[([^\]]+)\](.+)$", text)
        if m:
            book_part, sheet_part = m.group(1).strip(), m.group(2).strip()
        elif "!" in text:
            book_part, sheet_part = text.rsplit("!", 1)
            book_part = book_part.strip()
            sheet_part = sheet_part.strip()
        else:
            return self, sheet_spec
        book_part = book_part.strip("'\"[]")
        sheet_part = sheet_part.strip("'\"")
        if not book_part or not sheet_part:
            return self, sheet_spec
        return self.book(book_part), sheet_part

    def copy(self, src_sheet, src_range, dst_sheet, dst_cell):
        """Excel 네이티브 복사(값+수식+서식+병합 보존). '복사/복붙' 요청의 기본 수단."""
        src_ctx, src_sheet_name = self._ctx_and_sheet_from_spec(src_sheet)
        dst_ctx, dst_sheet_name = self._ctx_and_sheet_from_spec(dst_sheet)
        src_ws = src_ctx._ws(src_sheet_name)
        dst_ws = dst_ctx._ws(dst_sheet_name)
        src = src_ctx._rng(src_ws, src_range)
        dst = dst_ctx._rng(dst_ws, dst_cell)
        self._tick(2)
        try:
            dst_target = self._resize_rng(dst.Worksheet, dst, int(src.Rows.Count), int(src.Columns.Count))
            self._tick(2)
            dst_ctx._journal_save(dst_ws, dst_target)
        except Exception:
            self._shared["structural"].append(f"copy:{dst_sheet}!{dst_cell}")
        t0 = time.perf_counter()
        src.Copy(dst)
        self._tick(1)
        try:
            cells = int(src.Rows.Count) * int(src.Columns.Count)
            ms = (time.perf_counter() - t0) * 1000
            if ms >= 200 or cells >= 100000:
                _vba_trace(
                    "python_com.copy",
                    srcSheet=str(src_sheet),
                    srcRange=str(src_range),
                    dstSheet=str(dst_sheet),
                    dstCell=str(dst_cell),
                    cells=cells,
                    ms=round(ms, 1),
                )
        except Exception:
            pass
        try:
            self._app.CutCopyMode = False
        except Exception:
            pass
        return True

    def paste_copied(self, src_sheet, src_range, dst_sheet, dst_cell, src_book=None, dst_book=None, values_only=False):
        """[복붙 캡처 재생] 사용자가 라이브 Excel에서 Ctrl+C/Ctrl+V 한 동작을 그대로 재현한다.
        Excel 네이티브 Range.Copy(Destination=) 로 값+수식+서식+병합을 보존하며, 같은 인스턴스에 열린
        다른 워크북 간 복사(교차파일)도 지원한다. src_book/dst_book 을 주면 그 워크북에서 시트를 찾는다.
        LLM 추측이 아니라 캡처된 실제 좌표로 실행하므로 '값/수식 복붙' 모호성이 없다.
        values_only=True 면 소스 수식의 계산값만 대상에 쓰고, 소스 서식/수식은 복사하지 않는다."""
        # 소스 워크북이 같은 인스턴스에 안 열려 있으면(전체실행/재실행 때 흔함) 업로드 경로에서
        # 읽기전용으로 열어 교차파일 복사를 성립시킨다(작업 후 닫아 누수 방지). dst 는 작업 대상이라
        # 보통 열려 있다.
        opened_src = None
        opened_src_temp = None
        if src_book:
            try:
                src_ctx = self.book(src_book)
                _vba_trace("paste_copied.src", srcBook=src_book, resolved="in-instance")
            except PythonComSkillError:
                p = _registered_path_for_name(src_book)
                if not (p and os.path.exists(p)):
                    _vba_trace("paste_copied.src", srcBook=src_book, resolved="not-open-no-path", path=p)
                    raise
                opened_src, opened_src_temp = excel_workbooks_open(
                    self._app,
                    p,
                    read_only=True,
                    intended_name=src_book,
                )
                self._tick(2)
                src_ctx = PythonComSkillContext(self._app, opened_src, self._session, _shared=self._shared)
                _vba_trace("paste_copied.src", srcBook=src_book, resolved="auto-opened-readonly", path=p)
        else:
            src_ctx = self
        dst_ctx = self.book(dst_book) if dst_book else self
        try:
            src_ws = src_ctx._ws(src_sheet)
            dst_ws = dst_ctx._ws(dst_sheet)
            src = src_ctx._rng(src_ws, src_range)
            dst = dst_ctx._rng(dst_ws, dst_cell)
            self._tick(2)
            src_rows = int(src.Rows.Count)
            src_cols = int(src.Columns.Count)
            sheet_rows = int(src_ws.Rows.Count)
            sheet_cols = int(src_ws.Columns.Count)
            # 전체 열/행 소스('A:E','1:6')는 Rows/Columns.Count 가 시트 전체(1048576/16384)라
            # 대상 전체를 저널하면 수백만 빈 셀을 읽어 멈춘다 → 대상 열/행 ∩ UsedRange(실제 데이터)만
            # 백업해 롤백 가능하게 한다. 전체 열/행 복사는 Excel 이 대상도 전체 열/행으로 자동 확장한다.
            full_col = src_rows >= sheet_rows
            full_row = src_cols >= sheet_cols
            whole = full_col or full_row
            if whole:
                try:
                    c0 = int(dst.Column)
                    r0 = int(dst.Row)
                    if full_col:
                        band = dst_ws.Range(dst_ws.Cells(1, c0), dst_ws.Cells(1, c0 + src_cols - 1)).EntireColumn
                    else:
                        band = dst_ws.Range(dst_ws.Cells(r0, 1), dst_ws.Cells(r0 + src_rows - 1, 1)).EntireRow
                    backup = self._app.Intersect(band, dst_ws.UsedRange)
                    self._tick(2)
                    if backup is not None:
                        dst_ctx._journal_save(dst_ws, backup)
                    else:
                        # 대상에 기존 데이터 없음 → 잃을 것 없음. 그래도 신규 채움은 롤백 불가이므로 표시.
                        self._shared["structural"].append(f"paste_copied(whole):{dst_sheet}!{dst_cell}")
                except Exception:
                    self._shared["structural"].append(f"paste_copied(whole):{dst_sheet}!{dst_cell}")
            else:
                try:
                    dst_target = dst_ctx._resize_rng(dst_ws, dst, src_rows, src_cols)
                    self._tick(2)
                    dst_ctx._journal_save(dst_ws, dst_target)
                except Exception:
                    self._shared["structural"].append(f"paste_copied:{dst_sheet}!{dst_cell}")
            if values_only:
                try:
                    src_ws.Calculate()
                except Exception:
                    try:
                        self._app.Calculate()
                    except Exception:
                        pass
                self._tick(1)
                src.Copy()
                self._tick(1)
                dst.PasteSpecial(Paste=-4163)  # xlPasteValues
            else:
                src.Copy(dst)
            self._tick(1)
            try:
                self._app.CutCopyMode = False
            except Exception:
                pass
            return True
        finally:
            if opened_src is not None:
                try:
                    opened_src.Close(SaveChanges=False)
                except Exception:
                    pass
            if opened_src_temp:
                try:
                    Path(opened_src_temp).unlink(missing_ok=True)
                except Exception:
                    pass

    def set_fill(self, sheet, a1_range, color=None):
        """셀 음영/배경색 설정. color 는 '#RRGGBB'/'노랑'·'red' 같은 색이름/정수. None 이면 '채우기 없음'.
        예: ctx.set_fill("매출", "A1:C1", "노랑") / ctx.set_fill("매출", "A2:A10", "#DDEBF7")."""
        ws = self._ws(sheet)
        rng = self._rng(ws, a1_range)
        self._tick(2)
        col = _parse_excel_color(color)
        try:
            if col is None:
                rng.Interior.ColorIndex = -4142   # xlColorIndexNone (채우기 없음)
            else:
                rng.Interior.Color = int(col)
        except Exception as e:
            raise PythonComSkillError(f"셀 음영 적용 실패({a1_range}): {e}")
        self._shared["structural"].append(f"set_fill:{sheet}!{a1_range}={color}")
        return int(rng.Cells.Count)

    def set_font(self, sheet, a1_range, size=None, bold=None, italic=None, color=None, name=None):
        """글꼴 서식. 지정한 항목만 바꾼다 — size(pt 숫자)/bold/italic(True·False)/color(색)/name(글꼴명).
        예: ctx.set_font("매출","A1:C1", bold=True, size=12) / ctx.set_font("매출","B2", italic=True, color="빨강")."""
        ws = self._ws(sheet)
        rng = self._rng(ws, a1_range)
        self._tick(2)
        f = rng.Font
        try:
            if name is not None:
                f.Name = str(name)
            if size is not None:
                f.Size = float(size)
            if bold is not None:
                f.Bold = bool(bold)
            if italic is not None:
                f.Italic = bool(italic)
            cc = _parse_excel_color(color)
            if cc is not None:
                f.Color = int(cc)
        except Exception as e:
            raise PythonComSkillError(f"글꼴 서식 적용 실패({a1_range}): {e}")
        self._shared["structural"].append(f"set_font:{sheet}!{a1_range}")
        return int(rng.Cells.Count)

    def set_border(self, sheet, a1_range, style="thin", color=None, edges="all"):
        """테두리. style: thin/medium/thick/double/none(지우기). edges: all(각 셀 사방+내부)/outline(바깥 테두리만)/
        top·bottom·left·right(콤마 조합, 한글 위·아래·왼쪽·오른쪽도 가능). 예: ctx.set_border("매출","A1:D20") /
        ctx.set_border("매출","A1:D1", style="thick", edges="bottom")."""
        ws = self._ws(sheet)
        rng = self._rng(ws, a1_range)
        self._tick(2)
        XL_CONTINUOUS, XL_DOUBLE, XL_NONE = 1, -4119, -4142
        XL_THIN, XL_MEDIUM, XL_THICK = 2, -4138, 4
        st = str(style or "thin").lower().strip()
        if st in ("none", "없음", "지우기", "clear", "off"):
            line_style, weight = XL_NONE, XL_THIN
        elif st in ("double", "이중"):
            line_style, weight = XL_DOUBLE, XL_THICK
        elif st in ("medium", "중간"):
            line_style, weight = XL_CONTINUOUS, XL_MEDIUM
        elif st in ("thick", "굵은", "두꺼운"):
            line_style, weight = XL_CONTINUOUS, XL_THICK
        else:   # thin / 얇은 / 기본
            line_style, weight = XL_CONTINUOUS, XL_THIN
        cc = _parse_excel_color(color)
        # xlEdgeLeft/Top/Bottom/Right = 7/8/9/10, xlInsideVertical/Horizontal = 11/12
        edge_ids = {"left": 7, "top": 8, "bottom": 9, "right": 10}
        korean_edge = {"왼쪽": "left", "좌": "left", "위": "top", "상": "top",
                       "아래": "bottom", "하": "bottom", "오른쪽": "right", "우": "right"}

        def _apply_edge(idx):
            try:
                b = rng.Borders(idx)
                b.LineStyle = line_style
                if line_style != XL_NONE:
                    b.Weight = weight
                    if cc is not None:
                        b.Color = int(cc)
            except Exception:
                pass   # 단일 셀에 내부선(11/12) 등 적용 불가 케이스는 조용히 건너뜀
        try:
            spec = str(edges or "all").lower().strip()
            if spec in ("all", "전체", "모두"):
                for idx in (7, 8, 9, 10, 11, 12):
                    _apply_edge(idx)
            elif spec in ("outline", "바깥", "외곽", "테두리", "outside"):
                for idx in (7, 8, 9, 10):
                    _apply_edge(idx)
            else:
                for nm in [x.strip() for x in spec.replace("|", ",").split(",") if x.strip()]:
                    en = korean_edge.get(nm, nm)
                    if en in edge_ids:
                        _apply_edge(edge_ids[en])
        except Exception as e:
            raise PythonComSkillError(f"테두리 적용 실패({a1_range}): {e}")
        self._shared["structural"].append(f"set_border:{sheet}!{a1_range}:{style}:{edges}")
        return int(rng.Cells.Count)

    def clear(self, sheet, a1_range, keep_formulas=False):
        """범위 내용 삭제(서식 유지). keep_formulas=True 면 '수식 셀은 남기고 값(상수) 셀만' 비운다
        ("데이터만 비우고 수식은 유지" 요청용 — 세로/가로 축 실수 없이 결정적). 저널에 백업 후 삭제."""
        ws = self._ws(sheet)
        rng = self._rng(ws, a1_range)
        self._tick(2)
        self._journal_save(ws, rng)
        if keep_formulas:
            # 상수(값)만 지우고 수식은 보존. 대상에 상수가 하나도 없으면 SpecialCells 가 오류 → 조용히 종료.
            try:
                consts = rng.SpecialCells(2)  # xlCellTypeConstants
            except Exception:
                consts = None
            if consts is not None:
                consts.ClearContents()
        else:
            rng.ClearContents()
        self._tick(1)
        return True

    def shift_months(self, sheet, a1_range, delta=1):
        """범위 안 '문자열' 셀의 모든 'N월'(앞 'YY/YYYY년', 뒤 'D일' 포함)을 delta 개월 이동한다.
        예: "2026년 06월 (05월 1일 ~ 05월 31일)" + 1 → "2026년 07월 (06월 1일 ~ 06월 30일)".
        12월 넘김 시 연도 +, 말일 보정(내림, 윤년), 0패딩 폭 보존. 서식/수식/숫자/날짜 셀은 그대로 둠.
        '월 정보 +N / 다음달 / 한 달 뒤' 류 요청은 직접 정규식을 짜지 말고 반드시 이 헬퍼를 쓰세요."""
        ws = self._ws(sheet)
        rng = self._rng(ws, a1_range)
        self._tick(2)
        self._journal_save(ws, rng)
        delta = int(delta)
        nrows = int(rng.Rows.Count)
        ncols = int(rng.Columns.Count)
        data = rng.Value
        changed = 0
        for r in range(nrows):
            for c in range(ncols):
                if nrows == 1 and ncols == 1:
                    v = data
                else:
                    row = data[r] if isinstance(data, (list, tuple)) else data
                    v = row[c] if isinstance(row, (list, tuple)) else row
                if not isinstance(v, str):
                    continue
                shifted = _shift_months_in_text(v, delta)
                if shifted != v:
                    rng.Cells(r + 1, c + 1).Value = shifted
                    changed += 1
        self._tick(1)
        if changed == 0:
            raise PythonComSkillError("월 정보를 가진 문자열 셀이 없습니다(대상 범위/시트를 확인하세요).")
        return changed

    # ---- 구조 변경(저널 롤백 불가 → structural 표시) ----
    def insert_rows(self, sheet, row, count=1):
        ws = self._ws(sheet)
        self._tick(2)
        # 'N:M' 같은 행 범위 문자열은 Excel 이 직접 처리(모델이 자주 이 형태로 호출).
        if isinstance(row, str) and ":" in row:
            spec = row.strip()
            ws.Rows(spec).Insert()
            self._shared["structural"].append(f"insert_rows:{sheet}:{spec}")
            return True
        r = int(row)
        ws.Rows(f"{r}:{r + int(count) - 1}").Insert()
        self._shared["structural"].append(f"insert_rows:{sheet}:{row}+{count}")
        return True

    def insert_cols(self, sheet, col, count=1):
        """전체 열 삽입(병합셀 안전). col 은 'B', 2, 또는 'B:D' 범위 모두 허용."""
        ws = self._ws(sheet)
        self._tick(2)
        # 'B:D' 같은 열 범위 문자열은 Excel 이 직접 처리.
        if isinstance(col, str) and ":" in col:
            spec = col.strip()
            ws.Columns(spec).Insert()
            self._shared["structural"].append(f"insert_cols:{sheet}:{spec}")
            return True
        col_letter = col if isinstance(col, str) else _col_letter(int(col))
        end_letter = col_letter if int(count) <= 1 else _col_letter(self._col_index(col_letter) + int(count) - 1)
        ws.Columns(f"{col_letter}:{end_letter}").Insert()
        self._shared["structural"].append(f"insert_cols:{sheet}:{col_letter}+{count}")
        return True

    @staticmethod
    def _col_index(letter):
        n = 0
        for ch in str(letter).strip().upper():
            if not ("A" <= ch <= "Z"):
                raise PythonComSkillError(f"잘못된 열 문자: {letter}")
            n = n * 26 + (ord(ch) - 64)
        return n

    def delete_rows(self, sheet, row, count=1):
        ws = self._ws(sheet)
        self._tick(2)
        # 'N:M' 같은 행 범위 문자열은 Excel 이 직접 처리.
        if isinstance(row, str) and ":" in row:
            spec = row.strip()
            ws.Rows(spec).Delete()
            self._shared["structural"].append(f"delete_rows:{sheet}:{spec}")
            return True
        r = int(row)
        ws.Rows(f"{r}:{r + int(count) - 1}").Delete()
        self._shared["structural"].append(f"delete_rows:{sheet}:{row}+{count}")
        return True

    def delete_cols(self, sheet, col, count=1):
        """전체 열 삭제. col 은 'Q', 17, 또는 'Q:AU' 범위 모두 허용."""
        ws = self._ws(sheet)
        self._tick(2)
        # 'Q:AU' 같은 열 범위 문자열은 Excel 이 직접 처리(모델이 자주 이 형태로 호출 — '잘못된 열 문자' 방지).
        if isinstance(col, str) and ":" in col:
            spec = col.strip()
            ws.Columns(spec).Delete()
            self._shared["structural"].append(f"delete_cols:{sheet}:{spec}")
            return True
        col_letter = col if isinstance(col, str) else _col_letter(int(col))
        start_idx = self._col_index(col_letter)
        end_letter = _col_letter(start_idx + int(count) - 1)
        ws.Columns(f"{col_letter}:{end_letter}").Delete()
        self._shared["structural"].append(f"delete_cols:{sheet}:{col_letter}+{count}")
        return True

    def clear_filter(self, sheet=None):
        """시트에 걸린 자동필터(AutoFilter)를 해제한다 — 필터 조건을 모두 지워 숨은 행을 복원하고, 헤더의 필터
        드롭다운(화살표)도 제거한다. **"필터 풀어줘/해제해줘/모든 필터 제거/필터 걸린 거 없애줘"** 요청에 이걸 쓴다
        (기능 없다고 거부하지 말 것). 표(ListObject) 안의 필터 조건도 함께 해제한다. sheet 생략 시 활성 시트.
        필터가 하나도 없으면 조용히 통과(오류 아님)."""
        self._tick(2)
        ws = self._ws(sheet) if sheet is not None else self._wb.ActiveSheet
        cleared = False
        # 1) 표(ListObject) 안의 필터 조건 해제(숨은 행 복원)
        try:
            for lo in ws.ListObjects:
                try:
                    af = lo.AutoFilter
                    if af is not None and bool(af.FilterMode):
                        af.ShowAllData()
                        cleared = True
                except Exception:
                    pass
        except Exception:
            pass
        # 2) 시트 자동필터: 활성 조건 먼저 해제(숨은 행 복원) → 필터 드롭다운 자체 제거
        try:
            if bool(ws.FilterMode):
                ws.ShowAllData()
                cleared = True
        except Exception:
            pass
        try:
            if bool(ws.AutoFilterMode):
                ws.AutoFilterMode = False
                cleared = True
        except Exception:
            pass
        self._shared["structural"].append("clear_filter:%s(%s)" % (ws.Name, "ok" if cleared else "none"))
        return ws.Name

    def enable_filter(self, sheet=None, header_row=1):
        """시트 헤더행에 자동필터(필터 드롭다운)를 켠다 — 조건 없이 '필터 기능만' 활성화한다(이미 켜져 있으면 유지).
        "필터 켜줘/활성화/필터 걸어줘(조건 없이)" 요청에 쓴다. 특정 값만 보이게 하려면 ctx.apply_filter 를 쓴다.
        sheet 생략 시 활성 시트."""
        self._tick(2)
        ws = self._ws(sheet) if sheet is not None else self._wb.ActiveSheet
        hr = max(1, int(header_row or 1))
        last_r = max(hr, self.used_last_row(ws.Name))
        last_c = max(1, self.used_last_col(ws.Name))
        try:
            if not bool(ws.AutoFilterMode):
                # 인자 없는 AutoFilter() 토글은 win32com 에서 안 켜지는 경우가 있어, Field=1 로 명시 지정한다
                # (조건 없이 필드만 주면 필터 드롭다운만 켜지고 모든 행은 그대로 보인다).
                ws.Range(ws.Cells(hr, 1), ws.Cells(last_r, last_c)).AutoFilter(Field=1)
        except Exception:
            pass
        self._shared["structural"].append("enable_filter:%s" % ws.Name)
        return ws.Name

    def apply_filter(self, sheet, column, values, header_row=1):
        """'눈으로만' 특정 값만 보이게 필터를 건다 — 데이터를 지우지 않고, 조건에 안 맞는 행을 화면에서 '숨김'
        처리한다(원본 보존; ctx.clear_filter 로 언제든 복원). 행을 실제로 지우거나 새 시트로 빼는 게 아니라,
        사용자가 화면에서 특정 값만 보게 하는 용도다.
          column: 헤더명(권장) 또는 열 문자("C")/번호.  values: 보일 값 하나 또는 여러 개(리스트).
          예: ctx.apply_filter("상품번호별", "상태", ["완료","진행"])   # 상태가 완료/진행인 행만 보임
        (행 삭제·추출이 아니라 '표시 필터'가 필요할 때만. 조건에 맞는 행만 '추출'해 새 시트로 옮기려면
         ctx.filter_to_sheet 를 쓴다.)"""
        self._tick(2)
        ws = self._ws(sheet)
        hr = max(1, int(header_row or 1))
        last_r = max(hr, self.used_last_row(ws.Name))
        last_c = max(1, self.used_last_col(ws.Name))

        def _col1(spec):
            s = str(spec).strip()
            try:
                return int(self.find_header(ws.Name, s, header_row=hr))
            except Exception:
                pass
            if re.fullmatch(r"[A-Za-z]{1,3}", s):
                return self._col_index(s)
            try:
                return int(spec)
            except Exception:
                raise PythonComSkillError("필터 열 '%s' 을 찾지 못했습니다(헤더명 또는 열 문자를 쓰세요)." % spec)

        field = _col1(column)   # 범위를 A열부터 잡으므로 Field = 시트 열번호
        rng = ws.Range(ws.Cells(hr, 1), ws.Cells(last_r, last_c))
        vals = list(values) if isinstance(values, (list, tuple)) else [values]
        vals = [("" if v is None else str(v)) for v in vals]
        # 기존 조건이 걸려 있으면 초기화(중첩/누적 방지)
        try:
            if bool(ws.AutoFilterMode) and bool(ws.FilterMode):
                ws.ShowAllData()
        except Exception:
            pass
        if len(vals) == 1:
            rng.AutoFilter(Field=field, Criteria1=vals[0])
        else:
            rng.AutoFilter(Field=field, Criteria1=vals, Operator=7)  # xlFilterValues (여러 값 허용)
        self._shared["structural"].append("apply_filter:%s[%s]=%s" % (ws.Name, column, ",".join(vals)[:60]))
        return ws.Name

    def filter_to_sheet(self, sheet, predicate, dest_name, header_rows=1, after=None):
        """조건에 맞는 행만 골라 **새 시트(현재 활성 파일)**에 정리한다 — 원본은 그대로 둔다.
        predicate(row) 는 데이터 행(값 리스트, 0-based 인덱스)을 받아 True/False 를 반환.
          ctx.filter_to_sheet("Sheet1", lambda r: r[2] == "안전제일", "안전제일목록")
        "x열에서 y만 필터/추출해 새 시트에 정리" 요청의 기본 수단."""
        ws = self._ws(sheet)
        self._tick(2)
        grid = self.read(sheet)  # used range 전체(헤더+데이터)
        # [0.5.15] UsedRange 는 '첫 사용열'부터 시작한다 → A열이 비면 행 배열 인덱스가 그만큼 밀려서
        # predicate 의 절대 열 인덱스(예: "E열"=r[4])가 한 칸씩 어긋난다(빈 A → r[4]가 F를 가리켜 0건 실패).
        # 선두 빈 열 수만큼 None 패딩해 'A열=index0' 절대 기준으로 맞춘다. openpyxl 엔진은 UsedRange 가 이미
        # A1 기준이라(11836) 패딩 불필요 — 이 보정으로 COM/openpyxl 엔진 동작이 일치한다. 출력도 같은 절대
        # 열 위치(A1부터)로 기록되어 원본 열 정렬이 보존된다.
        try:
            _lead = int(ws.UsedRange.Column) - 1
        except Exception:
            _lead = 0
        if _lead > 0:
            grid = [[None] * _lead + list(r) for r in grid]
        hr = max(0, int(header_rows))
        header = [list(r) for r in grid[:hr]]
        matched = []
        for row in grid[hr:]:
            r = list(row)
            try:
                keep = bool(predicate(r))
            except Exception as err:
                raise PythonComSkillError(f"필터 조건(predicate) 실행 오류: {err}")
            if keep:
                matched.append(r)
        if not matched:
            # 매칭 0건이면 새 시트를 만들지 않는다 — 빈 시트만 남아 재요청 시 '이미 있음'으로
            # 반복 실패하던 문제 방지. 값의 공백/표기 차이는 ctx.normalize 로 맞춰야 한다.
            raise PythonComSkillError(
                "필터 조건을 만족하는 행이 없습니다. 값의 공백·표기 차이(ctx.normalize 사용)나 열/조건을 확인하세요.")
        if str(dest_name) in _excel_collection_names(self._wb.Worksheets):
            raise PythonComSkillError(f"시트 '{dest_name}' 이 이미 있습니다. 다른 이름을 쓰거나 먼저 삭제하세요.")
        self.add_sheet(str(dest_name), after=after)
        out = header + matched
        if out:
            # 새 시트라 수식 충돌 없음 → overwrite_formulas=True 로 그대로 기록.
            self.write(str(dest_name), "A1", out, overwrite_formulas=True)
        self._shared["structural"].append(f"filter_to_sheet:{sheet}->{dest_name}({len(matched)})")
        return dest_name

    def _pivot_value_table(self, sheet, group_by, value=None, agg="sum", dest_name=None, header_rows=1, after=None, column=None):
        """[내부/폴백] 그룹별 집계 '값 표'를 새 시트에 만든다(Python 집계 — 안정적). ctx.pivot 이 진짜
        피벗테이블(native_pivot) 생성에 실패했을 때 이 값-표로 폴백한다. 직접 호출도 가능(값-표 강제).
        group_by/value 는 헤더명(권장) 또는 열 문자("C"). agg 는 sum/count/avg/max/min.
          ctx.pivot("매출", group_by="회사", value="금액", agg="sum", dest_name="회사별합계")  # 1D 그룹요약
        ★ **다중 키·다중 값 지원**: group_by/value/agg 에 '리스트'를 주면 여러 기준으로 묶고 여러 값을 한 번에 집계한다
          (열 순서 = 그룹키들 + 값들). 예: "회사·지점별 매출 합계와 건수 개수":
          ctx.pivot("매출", group_by=["회사","지점"], value=["매출","건수"], agg=["sum","count"], dest_name="요약")
          → 헤더 [회사, 지점, 매출_sum, 건수_count]. value=None 이면 개수(count). agg 를 하나만 주면 모든 값에 적용.
        column 을 주면 **2D 크로스탭**(행=group_by, 열=column, 셀=value 의 agg; 크로스탭은 group_by 하나만):
          ctx.pivot("매출", group_by="지점", column="월", value="매출", agg="sum", dest_name="피벗_결과")
        "~별로 합계/개수/평균", "A·B별로 X합계·Y개수", "행은 X 열은 Y 값은 Z 피벗" 요청의 기본 수단(원본 보존, 손코딩 금지)."""
        self._tick(2)
        grid = self.read(sheet)
        hr = max(1, int(header_rows or 1))
        header = list(grid[hr - 1]) if len(grid) >= hr else []
        data = grid[hr:]

        # [중복 헤더] 같은 헤더가 여러 개면 2번째부터 '헤더명2','헤더명3' 으로 지목 가능(엑셀 피벗과 동일).
        _seen = {}
        _dnames = []
        for h in header:
            hs = str(h).strip()
            if hs in _seen:
                _seen[hs] += 1
                _dnames.append("%s%d" % (hs, _seen[hs]))
            else:
                _seen[hs] = 1
                _dnames.append(hs)

        def _col0(spec):
            s = str(spec).strip()
            for i, h in enumerate(header):
                if str(h).strip() == s:
                    return i
            if s in _dnames:               # '상품명2' 등 중복-리네임 이름 → 그 위치의 열
                return _dnames.index(s)
            # 슬래시-별칭("전화번호/회선번호/ID") — 모델이 후보 이름을 슬래시로 묶어 주는 경우, 각 후보를
            # 헤더에서 시도한다(정확 헤더명이 먼저 매칭되므로 실제 "A/B" 헤더는 위에서 이미 잡힘).
            if "/" in s:
                for alias in (a.strip() for a in s.split("/")):
                    if not alias:
                        continue
                    for i, h in enumerate(header):
                        if str(h).strip() == alias:
                            return i
                    if alias in _dnames:
                        return _dnames.index(alias)
            if re.fullmatch(r"[A-Za-z]{1,3}", s):
                return self._col_index(s) - 1
            try:
                n = int(spec)
                return n - 1 if n >= 1 else n
            except Exception:
                raise PythonComSkillError(f"열 '{spec}' 을 헤더에서 찾지 못했습니다(헤더명 또는 열 문자를 쓰세요).")

        if column is not None:
            # 2D 크로스탭: 행=group_by, 열=column, 셀=value 의 agg. 모델이 손코딩하던 부분을 결정적으로 대체.
            if isinstance(group_by, (list, tuple)):
                raise PythonComSkillError("크로스탭(column 지정)은 group_by 를 하나만 받습니다.")
            g_i = _col0(group_by)
            c_i = _col0(column)
            v_i = _col0(value) if value is not None else None
            row_label = header[g_i] if (g_i is not None and g_i < len(header)) else str(group_by)
            out, rkeys, ckeys = _pivot_crosstab(data, g_i, c_i, v_i, str(agg or "sum"), row_label=row_label)
            name = str(dest_name or "피벗요약")
            if name in _excel_collection_names(self._wb.Worksheets):
                raise PythonComSkillError(f"시트 '{name}' 이 이미 있습니다. 다른 이름을 쓰거나 먼저 삭제하세요.")
            self.add_sheet(name, after=after)
            if out:
                self.write(name, "A1", out, overwrite_formulas=True)
            self._shared["structural"].append(f"pivot_crosstab:{sheet}->{name}({len(rkeys)}x{len(ckeys)})")
            return name

        def _to_num(v):
            if isinstance(v, bool):
                return None
            if isinstance(v, (int, float)):
                return float(v)
            try:
                return float(str(v).replace(",", ""))
            except Exception:
                return None

        group_cols = list(group_by) if isinstance(group_by, (list, tuple)) else [group_by]
        gidx = [_col0(g) for g in group_cols]
        values = list(value) if isinstance(value, (list, tuple)) else [value]
        aggs = list(agg) if isinstance(agg, (list, tuple)) else [agg] * len(values)
        while len(aggs) < len(values):
            aggs.append(aggs[-1] if aggs else "sum")
        aggs = [str(a or "sum").lower() for a in aggs]
        vidx = [(_col0(v) if v is not None else None) for v in values]

        groups = {}
        order = []
        for r in data:
            r = list(r)
            key = tuple((r[i] if (i is not None and i < len(r)) else "") for i in gidx)
            if key not in groups:
                groups[key] = [[] for _ in values]
                order.append(key)
            for pos, vi in enumerate(vidx):
                if vi is None:
                    groups[key][pos].append(1)
                elif vi < len(r):
                    groups[key][pos].append(r[vi])

        def _agg(vals, name):
            nums = [n for n in (_to_num(v) for v in vals) if n is not None]
            if name == "count":
                return len(vals)
            if name in ("avg", "average", "mean"):
                return (sum(nums) / len(nums)) if nums else 0
            if name == "max":
                return max(nums) if nums else ""
            if name == "min":
                return min(nums) if nums else ""
            return sum(nums)

        out_header = []
        for n, i in enumerate(gidx):
            out_header.append(_dnames[i] if (i is not None and i < len(_dnames)) else ("그룹%d" % (n + 1)))
        for v, a in zip(values, aggs):
            out_header.append((str(v) if v is not None else "값") + "_" + ("avg" if a in ("average", "mean") else a))
        out = [out_header]
        for key in order:
            out.append(list(key) + [_agg(groups[key][i], aggs[i]) for i in range(len(values))])

        name = str(dest_name or "피벗요약")
        if name in _excel_collection_names(self._wb.Worksheets):
            raise PythonComSkillError(f"시트 '{name}' 이 이미 있습니다. 다른 이름을 쓰거나 먼저 삭제하세요.")
        self.add_sheet(name, after=after)
        # [앞자리0 보존] 그룹키 열에 '0으로 시작하는 숫자 문자열'(전화번호/가입번호 등)이 있으면, 그 열을 먼저
        # 텍스트(@) 서식으로 지정한 뒤 쓴다 — 그래야 write 시 숫자로 강제 변환돼 앞자리 0 이 사라지는 것을 막는다.
        # (원본 열이 텍스트라 read 가 "010..." 문자열을 돌려준 경우에만 트리거. 원본이 이미 숫자면 보존할 0 이 없음.)
        try:
            n_out = len(out)
            for kc in range(len(gidx)):
                needs_text = any(
                    isinstance(key[kc] if kc < len(key) else "", str)
                    and str(key[kc])[:1] == "0" and str(key[kc])[1:].isdigit()
                    for key in order
                )
                if needs_text and n_out >= 2:
                    cl = _col_letter(kc + 1)
                    self._ws(name).Range(f"{cl}2:{cl}{n_out}").NumberFormat = "@"
        except Exception:
            pass
        if out:
            self.write(name, "A1", out, overwrite_formulas=True)
        self._shared["structural"].append(f"pivot:{sheet}->{name}({len(order)})")
        return name

    def native_pivot(self, sheet, group_by, value=None, agg="sum", dest_name=None, column=None, header_rows=1):
        """엑셀 '진짜 피벗테이블(PivotTable 개체)'을 새 시트에 만든다 — 원본 데이터와 연결돼 '새로 고침'이
        되고 필드 배치가 살아있는 피벗이다(값만 찍는 요약표는 ctx.pivot). 사용자가 "진짜 피벗/새로고침 되는
        피벗/피벗테이블로 만들어줘/슬라이서" 를 원할 때만 쓴다. (COM 특성상 값-표 피벗보다 불안정할 수 있음.)
          group_by: 행 필드(헤더명, 리스트 가능).  value: 값 필드(헤더명; None 이면 개수).  agg: sum/count/avg/max/min.
          column: 주면 열 필드(크로스탭).  dest_name: 새 시트 이름.
          예: ctx.native_pivot("매출", group_by=["회사","지점"], value="금액", agg="sum", dest_name="피벗")"""
        self._tick(2)
        ws = self._ws(sheet)
        hr = max(1, int(header_rows or 1))
        last_r = max(hr, self.used_last_row(ws.Name))
        last_c = max(1, self.used_last_col(ws.Name))

        def _headers_at(row_idx):
            vals = ws.Range(ws.Cells(row_idx, 1), ws.Cells(row_idx, last_c)).Value
            if isinstance(vals, (tuple, list)):
                row0 = vals[0] if (vals and isinstance(vals[0], (tuple, list))) else vals
                return [("" if v is None else str(v)).strip() for v in row0]
            return [("" if vals is None else str(vals)).strip()]

        headers = _headers_at(hr)

        # [헤더 행 자동 보정 2026-08-06] 1행이 제목/공백이고 실제 헤더가 2행인 표가 흔한데, 예전엔
        # header_rows 를 정확히 주지 않으면 "피벗 필드 '서비스' 을 찾지 못했습니다"로 끝났다.
        # ctx.find_header 는 예전부터 인접 행까지 훑어 구제해 주는데(11027) 피벗만 안 그래서
        # "전엔 헤더를 잡았던 것 같은데 왜 안 되지"가 됐다. 여기서도 같은 구제를 해 준다.
        #   찾는 이름 전부를 담은 행이 '딱 하나'일 때만 옮긴다 — 모호하면 손대지 않고 기존 오류로 간다.
        _wanted = []
        for _spec in ([group_by] if not isinstance(group_by, (list, tuple)) else list(group_by)):
            if _spec is not None:
                _wanted.append(str(_spec).strip())
        for _spec in ([value] if not isinstance(value, (list, tuple)) else list(value or [])):
            if _spec is not None:
                _wanted.append(str(_spec).strip())
        if column is not None:
            _wanted.append(str(column).strip())
        _wanted = [w for w in _wanted if w and not re.fullmatch(r"[A-Za-z]{1,3}|\d+", w)]
        if _wanted and not all(w in headers for w in _wanted):
            _hits = []
            for _r in range(1, min(last_r, 10) + 1):
                if _r == hr:
                    continue
                if all(w in _headers_at(_r) for w in _wanted):
                    _hits.append(_r)
            if len(_hits) == 1:
                try:
                    _vba_trace("python_com.native_pivot.header_row_autofix",
                               sheet=str(ws.Name), given=hr, used=_hits[0], fields=_wanted)
                except Exception:
                    pass
                hr = _hits[0]
                last_r = max(hr, last_r)
                headers = _headers_at(hr)

        # [중복 헤더] 엑셀 피벗은 같은 헤더가 2개면 2번째를 '헤더명2'(3번째는 '헤더명3')로 자동 리네임한다
        # (예: 상품명 두 열 → 필드 '상품명', '상품명2'). 필드명 목록을 엑셀과 동일하게 만들어 매칭한다.
        seen = {}
        field_names = []
        for h in headers:
            if h in seen:
                seen[h] += 1
                field_names.append("%s%d" % (h, seen[h]))
            else:
                seen[h] = 1
                field_names.append(h)

        def _fname(spec):
            s = str(spec).strip()
            if s in field_names:                 # '상품명2' 등 엑셀 중복-리네임 필드명
                return s
            if s in headers:                     # 원본 헤더명 → 그 위치의 필드명(중복이면 첫 번째)
                return field_names[headers.index(s)]
            if "/" in s:
                for a in (x.strip() for x in s.split("/")):
                    if a in field_names:
                        return a
                    if a in headers:
                        return field_names[headers.index(a)]
            idx = None
            if re.fullmatch(r"[A-Za-z]{1,3}", s):
                idx = self._col_index(s)
            else:
                try:
                    idx = int(s)
                except Exception:
                    idx = None
            if idx and 1 <= idx <= len(field_names) and field_names[idx - 1]:
                return field_names[idx - 1]     # 열 문자/번호 → 그 위치의 (중복 반영) 필드명
            raise PythonComSkillError("피벗 필드 '%s' 을 헤더에서 찾지 못했습니다(헤더명 또는 열 문자를 쓰세요)." % spec)

        name = str(dest_name or "피벗")[:31]   # Excel 시트명 31자 — 조회(_ws 절단 폴백)와 대칭
        # [항상 진짜 피벗] 같은 스킬을 다시 적용하면 dest 시트가 남아 있는데, 예전엔 여기서 raise 하고
        # ctx.pivot 이 조용히 값-표로 폴백해 "1회차만 진짜 피벗, 2회차부터 가짜"가 됐다(실측 재현).
        # 재실행은 '같은 피벗을 다시 만드는 것'이 의도이므로, 우리가 만든 dest 는 지우고 다시 만든다.
        if name in _excel_collection_names(self._wb.Worksheets):
            self._wb.Worksheets(name).Delete()
        src = ws.Range(ws.Cells(hr, 1), ws.Cells(last_r, last_c))
        self.add_sheet(name)
        pt_ws = self._ws(name)
        XL_DB, XL_ROW, XL_COL = 1, 1, 2
        AGG = {"sum": -4157, "count": -4112, "avg": -4106, "average": -4106, "mean": -4106, "max": -4136, "min": -4139}
        pt_name = "PT_" + uuid.uuid4().hex[:8]
        cache = self._wb.PivotCaches().Create(SourceType=XL_DB, SourceData=src)
        pt = cache.CreatePivotTable(TableDestination=pt_ws.Cells(1, 1), TableName=pt_name)
        groups = list(group_by) if isinstance(group_by, (list, tuple)) else [group_by]
        for g in groups:
            pt.PivotFields(_fname(g)).Orientation = XL_ROW
        if column is not None:
            pt.PivotFields(_fname(column)).Orientation = XL_COL
        values = list(value) if isinstance(value, (list, tuple)) else [value]
        aggs = list(agg) if isinstance(agg, (list, tuple)) else [agg] * len(values)
        while len(aggs) < len(values):
            aggs.append(aggs[-1] if aggs else "sum")
        # [행=값 동일 필드] 엑셀 COM 의 AddDataField 는 '행/열 필드를 값 영역으로 이동'시킨다(한 필드
        # Orientation 은 하나). 그래서 group_by 한 필드를 값(개수 등)으로도 넣으면 그 필드가 행에서 빠져
        # 행 그룹이 통째로 사라지고 피벗이 '전체 총합 1줄'로 붕괴한다(사용자 제보 "개수:MVNO상품명" → 1줄).
        # 엑셀 UI 는 드래그하면 행·값에 동시 배치되므로, AddDataField 직후 그 필드의 행/열 방향을 재지정해
        # 둘 다 유지한다(실측 검증: 재지정 시 값-데이터 필드는 유지되고 행 그룹이 되살아남).
        row_field_names = set(_fname(g) for g in groups)
        col_field_name = _fname(column) if column is not None else None
        for v, a in zip(values, aggs):
            an = str(a or "sum").lower()
            fn = AGG.get(an, -4157)
            if v is None:
                gfn = _fname(groups[0])
                pt.AddDataField(pt.PivotFields(gfn), "개수", -4112)
                pt.PivotFields(gfn).Orientation = XL_ROW          # 개수용으로 옮겨간 행 필드 복구
            else:
                fnm = _fname(v)
                pt.AddDataField(pt.PivotFields(fnm), ("%s_%s" % (fnm, an))[:250], fn)
                if fnm in row_field_names:
                    pt.PivotFields(fnm).Orientation = XL_ROW      # 행 필드를 값에도 넣은 경우 행 유지
                elif col_field_name is not None and fnm == col_field_name:
                    pt.PivotFields(fnm).Orientation = XL_COL      # 열 필드를 값에도 넣은 경우 열 유지
        self._shared["structural"].append("native_pivot:%s->%s" % (ws.Name, name))
        return name

    def pivot(self, sheet, group_by, value=None, agg="sum", dest_name=None, header_rows=1, after=None, column=None):
        """그룹별 집계 피벗을 새 시트에 만든다. **기본은 엑셀 '진짜 피벗테이블(PivotTable 개체)'**(원본과 연결돼
        새로 고침 가능)로 만들고, 그게 실패하면 자동으로 '값 표'로 폴백한다 — 호출부는 ctx.pivot 하나만 쓰면 된다.
        인자: group_by/value/agg 에 '리스트'를 주면 다중 키·다중 값(agg 하나면 모든 값에 적용), value=None 은 개수,
        column 은 열 필드(2D 크로스탭), dest_name 은 새 시트명.
        "~별 합계/개수/평균/요약/피벗" 요청의 기본 수단(원본 보존, 손코딩 집계 금지). 값-표만 강제하려면
        ctx._pivot_value_table 을 직접 호출."""
        eff = (str(dest_name) if dest_name else "피벗요약")[:31]
        try:
            return self.native_pivot(sheet, group_by, value=value, agg=agg, dest_name=eff, column=column, header_rows=header_rows)
        except PythonComSkillError:
            raise                                   # 필드 못 찾음 등 우리 검증 오류는 그대로(원인이 명확)
        except Exception as _e:
            # [항상 진짜 피벗] 예전엔 여기서 값-표로 조용히 폴백해, 같은 명령인데 어떤 땐 진짜 피벗·어떤 땐
            # 가짜 표가 나왔다(사용자 제보). 이제 폴백하지 않고 원인을 말하는 오류로 끝낸다 — 조용한 강등보다
            # 정직한 실패가 낫고, 자동복구/사용자가 원인을 보고 고칠 수 있다.
            try:
                if eff in _excel_collection_names(self._wb.Worksheets):
                    self._wb.Worksheets(eff).Delete()   # 부분 생성분 정리
            except Exception:
                pass
            try:
                _vba_trace("python_com.pivot.native_failed", sheet=str(sheet), dest=eff, error=str(_e)[:200])
            except Exception:
                pass
            msg = str(_e)
            hint = ""
            if "적어도 두 행" in msg or "at least two rows" in msg.lower():
                hint = " 원본에 집계할 데이터 행이 없습니다(헤더만 있는지 확인하세요)."
            raise PythonComSkillError("피벗테이블 생성 실패: %s%s" % (msg[:300], hint))

    def normalize(self, value):
        """텍스트 정규화(공백/표기 차이 제거). 값 비교 보조용."""
        return normalize_text(value)

    def move_cols(self, sheet, columns, before, header_row=1, scan_from=None):
        """여러 열을 헤더+데이터까지 통째로 before 열 앞으로 옮긴다(원본 제거). 인덱스 시프트 자동.
        columns 는 둘 중 하나:
          - 열 목록: 헤더명 또는 "C" 열 문자. 예: ctx.move_cols("S", ["a_항목","b_값"], "J")
          - 조건 함수(헤더명을 받아 True/False): 헤더 행을 스캔해 매칭 열을 자동 선택.
            예: ctx.move_cols("S", lambda h: any(x in str(h) for x in ["a","b","c"]), "J", scan_from="J")
        header_row: 헤더가 몇 행인지(2행 헤더면 2). scan_from: 그 열부터만 스캔(예: "J")."""
        ws = self._ws(sheet)
        self._tick(2)
        hr = max(1, int(header_row or 1))

        def _idx_of(spec):
            s = str(spec).strip()
            if re.fullmatch(r"[A-Za-z]{1,3}", s):
                return self._col_index(s)
            try:
                return int(spec)
            except (TypeError, ValueError):
                pass
            # 헤더명도 허용(columns 인자와 동일). LLM 이 before/scan_from 에 "합계" 같은 헤더명을 넘기는 경우 대비.
            return self.find_header(sheet, s, header_row=hr)

        before_idx = _idx_of(before)

        if callable(columns):
            last_c = self.last_col(sheet, hr)
            grid = self.read(sheet, _col_letter(1) + str(hr) + ":" + _col_letter(max(1, last_c)) + str(hr))
            hdr = list(grid[0]) if grid else []
            sf = _idx_of(scan_from) if scan_from is not None else 1
            src_idx = []
            for i0, h in enumerate(hdr):
                idx1 = i0 + 1
                if idx1 < sf:
                    continue
                try:
                    keep = bool(columns(h))
                except Exception as err:
                    raise PythonComSkillError(f"열 선택 조건(predicate) 실행 오류: {err}")
                if keep:
                    src_idx.append(idx1)
        else:
            cols = columns if isinstance(columns, (list, tuple)) else [columns]
            def _to_idx(spec):
                s = str(spec).strip()
                if re.fullmatch(r"[A-Za-z]{1,3}", s):
                    return self._col_index(s)
                try:
                    return self.find_header(sheet, s, header_row=hr)
                except PythonComSkillError:
                    pass
                try:
                    return int(spec)
                except Exception:
                    raise PythonComSkillError(f"옮길 열 '{spec}' 을 찾지 못했습니다(헤더명 또는 열 문자를 쓰세요).")
            src_idx = sorted(set(_to_idx(c) for c in cols))

        src_idx = sorted(set(src_idx))
        if not src_idx:
            raise PythonComSkillError("옮길 열을 찾지 못했습니다(조건/목록을 확인하세요).")
        n = len(src_idx)
        # before 앞에 n개 삽입 → before_idx 이상의 기존 열이 오른쪽으로 n칸 밀린다.
        self.insert_cols(sheet, _col_letter(before_idx), count=n)
        shifted = [(s + n if s >= before_idx else s) for s in src_idx]
        for k, s in enumerate(shifted):
            src = _col_letter(s)
            self.copy(sheet, src + ":" + src, sheet, _col_letter(before_idx + k) + "1")
        # 원본 삭제: 큰 인덱스부터(역순)라야 남은 원본 인덱스가 안 밀린다.
        for s in sorted(shifted, reverse=True):
            self.delete_cols(sheet, _col_letter(s))
        self._shared["structural"].append(f"move_cols:{sheet}:{src_idx}->{before}")
        return True

    def move_col_clear(self, sheet, src, dst, header_row=None, clear_source=True):
        """한 열의 내용(헤더+데이터+서식+세로병합)을 다른 열로 옮기고 원래 열은 '비운다'(열 구조는 유지 —
        삭제/시프트 없음). '열 이동+원본 비우기' 전용. (열 순서 재배치는 move_cols 를 쓰세요.)

        상단 제목/단위 행의 '가로 병합'(예: A2:F2)을 자동으로 건너뛰고 복사하므로 '병합된 셀에서는 실행할 수
        없습니다'(1004)가 나지 않는다. 대상 열의 기존 병합도 먼저 정리한다. header_row 를 주면 그 행부터,
        없으면 원본 열이 가로 병합에 안 걸리는 첫 행부터 복사한다."""
        ws = self._ws(sheet)
        self._tick(2)
        src_i = self._resolve_col(sheet, src, header_row or 1)
        dst_i = self._resolve_col(sheet, dst, header_row or 1)
        src_l, dst_l = _col_letter(src_i), _col_letter(dst_i)
        last = self.last_row(sheet, col=src_i)
        if last < 1:
            raise PythonComSkillError(f"'{sheet}' {src_l}열에 데이터가 없습니다.")
        # 시작행: 지정 header_row 우선. 없으면 상단에서 원본 열이 '가로(여러 열) 병합'에 걸리는 구간을 건너뛴
        # 첫 행(제목/단위 행 회피). 부분 병합 복사 1004 의 원인을 결정적으로 제거한다.
        if header_row:
            start = max(1, int(header_row))
        else:
            start = 1
            r, cap = 1, min(int(last), 40)
            while r <= cap:
                ma = ws.Cells(r, src_i).MergeArea
                try:
                    wide = int(ma.Columns.Count) > 1
                except Exception:
                    wide = False
                if wide:
                    start = int(ma.Row) + int(ma.Rows.Count)  # 그 가로병합 '바로 아래'
                    r = start
                else:
                    r += 1
        src_rng = f"{src_l}{start}:{src_l}{last}"
        # 대상 열 병합 정리(붙여넣기 충돌 방지) — G 등 대상에 기존 병합이 있어도 안전.
        try:
            ws.Range(f"{dst_l}{start}:{dst_l}{last}").UnMerge()
        except Exception:
            pass
        self.copy(sheet, src_rng, sheet, f"{dst_l}{start}")   # 값+수식+서식+세로병합 보존
        if clear_source:
            self.clear(sheet, src_rng)                        # 원본은 내용만 비움(열 유지, 시프트 없음)
        self._shared["structural"].append(f"move_col_clear:{sheet}:{src_l}->{dst_l}@{start}")
        return True

    def copy_col(self, sheet, src, dst, header_row=None):
        """열 → 열 '복사'(원본 유지). 값+수식+서식+세로병합 보존, 상단 제목의 가로 병합은 자동 회피,
        대상 열 병합은 먼저 정리 → '병합된 셀에서는 실행할 수 없습니다'(1004) 없이 복사. move_col_clear 와
        동일 안전장치이되 원본을 비우지 않는다. '한 열을 다른 열로 (서식째) 복사' 요청용."""
        return self.move_col_clear(sheet, src, dst, header_row=header_row, clear_source=False)

    def copy_values(self, src_sheet, src_range, dst_sheet, dst_cell):
        """'값으로 복사'(계산 결과값 + 서식/숫자서식/테두리/병합 보존, 수식은 넣지 않음).
        ctx.copy 는 수식을 그대로 옮겨 상대참조가 시프트되지만(예: 제목이 =다른시트!A2 이면 J로 옮길 때 =..!J2 로
        어긋남), 이 함수는 소스의 '결과값'만 넣어 참조 시프트가 없다(긴 텍스트/EID 도 숫자서식 동반이라 안전).
        '원문 텍스트 그대로/값으로 복사' 요청용. 서식은 유지하되 수식은 값으로 고정하고 싶을 때."""
        src_ctx, src_name = self._ctx_and_sheet_from_spec(src_sheet)
        dst_ctx, dst_name = self._ctx_and_sheet_from_spec(dst_sheet)
        src_ws = src_ctx._ws(src_name)
        dst_ws = dst_ctx._ws(dst_name)
        src = src_ctx._rng(src_ws, src_range)
        dst = dst_ctx._rng(dst_ws, dst_cell)
        self._tick(2)
        try:
            dst_target = self._resize_rng(dst.Worksheet, dst, int(src.Rows.Count), int(src.Columns.Count))
            dst_ctx._journal_save(dst_ws, dst_target)
        except Exception:
            self._shared["structural"].append(f"copy_values:{dst_sheet}!{dst_cell}")
        src.Copy()
        try:
            dst.PasteSpecial(Paste=-4104)   # xlPasteAll: 서식+테두리+병합+숫자서식(+수식)
            dst.PasteSpecial(Paste=12)      # xlPasteValuesAndNumberFormats: 수식→계산값(참조 시프트 제거, EID 안전)
        finally:
            try:
                self._app.CutCopyMode = False
            except Exception:
                pass
        return True

    def swap_cols(self, sheet, col_a, col_b, header_row=None):
        """인접한 두 열의 위치를 서로 맞바꾼다. Excel 네이티브 Cut/Insert 로 옮겨 **수식 참조가 자동 보정**된다
        (=SUM(D..) 등이 #REF! 로 깨지지 않음 — copy+delete 방식의 한계 해결). 제목처럼 두 열에 걸친 '가로 병합'은
        먼저 임시 해제하고 스왑 후 되돌려 1004 를 피한다. (인접하지 않은 열 재배치는 move_cols 를 쓰세요.)"""
        ws = self._ws(sheet)
        self._tick(2)
        a = self._resolve_col(sheet, col_a, header_row or 1)
        b = self._resolve_col(sheet, col_b, header_row or 1)
        if a == b:
            return True
        lo, hi = min(a, b), max(a, b)
        if hi - lo != 1:
            raise PythonComSkillError("swap_cols 는 인접한 두 열만 지원합니다(떨어진 열은 move_cols 사용).")
        # 두 열에 걸친 '가로(여러 열) 병합'(제목 등) 임시 해제 → 스왑 후 재병합(범위는 스왑해도 동일).
        try:
            cap = min(int(ws.UsedRange.Row) + int(ws.UsedRange.Rows.Count) - 1, 60)
        except Exception:
            cap = 60
        saved, seen = [], set()
        for r in range(1, cap + 1):
            for col in (lo, hi):
                try:
                    ma = ws.Cells(r, col).MergeArea
                    if int(ma.Columns.Count) > 1:
                        addr = str(ma.Address)
                        if addr not in seen:
                            seen.add(addr)
                            saved.append(addr)
                            ma.UnMerge()
                except Exception:
                    pass
        # 스왑: 뒤 열(hi)을 잘라 앞 열(lo) 앞에 삽입(잘라낸 셀 삽입) → 참조 자동 보정.
        ws.Columns(hi).Cut()
        ws.Columns(lo).Insert(-4161)   # Shift:=xlToRight
        for addr in saved:
            try:
                ws.Range(addr).Merge()
            except Exception:
                pass
        try:
            self._app.CutCopyMode = False
        except Exception:
            pass
        self._shared["structural"].append(f"swap_cols:{sheet}:{_col_letter(lo)}<->{_col_letter(hi)}")
        return True

    def fill_sum_col(self, sheet, dest_col, src_cols, header_row=None):
        """합계 열(dest_col)을 원본 열들(src_cols)의 합계 '수식'으로 채운다. dest_col 이 2행 등으로 '세로 병합'된
        표(계정별 그룹)면 **병합 블록 단위로** `=SUM(src_top:src_bottom)+...` 를 병합 top 에 넣는다(골든의 그룹
        합계와 일치). 원본 셀이 숫자가 아닌 **라벨/비데이터 행은 건너뛴다**(단위·소제목을 수식으로 덮지 않음).
        header_row 를 주면 그 헤더 병합 '아래' 데이터부터 처리해 헤더/라벨 행을 안전하게 보존한다."""
        ws = self._ws(sheet)
        self._tick(2)
        dcol = self._resolve_col(sheet, dest_col, header_row or 1)
        src_list = src_cols if isinstance(src_cols, (list, tuple)) else [src_cols]
        scols = [self._resolve_col(sheet, c, header_row or 1) for c in src_list]
        hr = int(header_row) if header_row else 1
        hma = ws.Cells(hr, dcol).MergeArea            # 헤더 셀(병합이면 그 높이만큼) 아래부터 데이터
        start = int(hma.Row) + int(hma.Rows.Count)
        last = self.last_row(sheet, col=scols[0])
        if last < start:
            raise PythonComSkillError(f"'{sheet}' 합계 대상 데이터가 없습니다(시작 {start}, 마지막 {last}).")
        r, n = start, 0
        while r <= last:
            ma = ws.Cells(r, dcol).MergeArea
            top = int(ma.Row)
            bottom = top + int(ma.Rows.Count) - 1
            has_num = False
            for sc in scols:
                for rr in range(top, bottom + 1):
                    v = ws.Cells(rr, sc).Value
                    if isinstance(v, (int, float)) and not isinstance(v, bool):
                        has_num = True
                        break
                if has_num:
                    break
            if has_num:
                parts = [f"SUM({_col_letter(sc)}{top}:{_col_letter(sc)}{bottom})" for sc in scols]
                ws.Cells(top, dcol).Formula = "=" + "+".join(parts)
                n += 1
            r = bottom + 1
        self._shared["structural"].append(f"fill_sum_col:{sheet}:{_col_letter(dcol)}={[_col_letter(c) for c in scols]}")
        return n

    def sum_column(self, sheet, col, header_row=None, exclude_total_rows=True):
        """열(col: 열문자 'F' / 헤더명 '합계' / 열번호 6)의 숫자 값을 더해 **합계 값을 반환**한다.
        write 하지 않음 — 반환값을 ctx.write_cell 로 원하는 셀에 넣으세요.

        exclude_total_rows=True(기본): 표 안에 이미 있는 '합계/총계/소계' 행을 자동 제외해 **이중계산을 막는다.**
          - 총계 행 판정은 라벨 컬럼을 추측하지 말 것: 이 함수가 각 행의 왼쪽 라벨 영역(A~C)을 스캔한다
            (실제로 라벨이 A열에 있는데 코드가 C열을 보다 못 거르는 실수를 원천 차단).
          - 첫 총계 행이 항목 블록의 '끝'이다: 그 아래 꼬리(부가세·검산·단위 표기 등)까지 다 더하는 것을 막기
            위해, 첫 총계 행 '위'의 항목 행들만 더한다. 총계 행이 없으면 데이터 끝까지 더한다.
        exclude_total_rows=False: 데이터 범위 전체(총계 행 포함)를 그대로 더한다.
        header_row 를 주면 그 헤더(병합 포함) '아래'부터 시작한다(헤더/제목 오합산 방지).
          예) total = ctx.sum_column("요약", "합계", header_row=4); ctx.write_cell("요약", "J15", total)"""
        ws = self._ws(sheet)
        self._tick(2)
        tcol = self._resolve_col(sheet, col, header_row or 1)
        if header_row:
            hr = int(header_row)
            hma = ws.Cells(hr, tcol).MergeArea
            start = int(hma.Row) + int(hma.Rows.Count)
        else:
            start = 1
        last = self.last_row(sheet, col=tcol)
        if last < start:
            return 0.0
        L = _col_letter(tcol)
        vals = self.read(sheet, f"{L}{start}:{L}{last}")
        labels = self.read(sheet, f"A{start}:C{last}")
        total = 0.0
        excluded = 0
        for i in range(len(vals)):
            row_labels = labels[i] if (labels and i < len(labels)) else []
            if _is_total_label(row_labels):
                excluded += 1
                if exclude_total_rows:
                    break            # 첫 총계 행 = 항목 블록의 끝(그 아래 꼬리 무시)
            n = _coerce_number(vals[i][0] if vals[i] else None)
            if n is not None:
                total += n
        try:
            _vba_trace("python_com.sum_column", sheet=str(sheet), col=L, start=start, last=last,
                       exclude_total_rows=bool(exclude_total_rows), excluded=excluded, total=total)
        except Exception:
            pass
        return total

    def copy_key_blocks(self, src_sheet, dst_sheet, key_col, first_col, last_col,
                        src_scan=None, dst_scan=None, on_mismatch="skip"):
        """'가입번호'처럼 키가 여러 행 세로병합 블록을 이루는 표에서, 대상의 각 키 블록에 소스의
        같은 키 블록 '전체(모든 행)'를 서식·병합 그대로 복사한다. ★ Range.Find 로 블록 첫 행 1줄만
        복사돼 소계 등 나머지 행이 유실되던 문제를 근본 해결(가입번호=1:N 병합블록 매칭).

        - src_sheet/dst_sheet: 시트명 또는 "파일.xlsx!시트" 교차파일 스펙(ctx 는 대상 파일에 바인딩).
        - key_col: 키(가입번호) 열(문자/헤더명/번호). first_col~last_col: 복사할 데이터 열(키 열 포함 권장).
        - src_scan/dst_scan: 키를 훑을 A1 범위(예 "B3:B345"/"B4:B89"). 생략 시 키 열 사용범위.
        - 소스/대상 모두 키 열이 블록 단위 세로병합(맨 윗셀에만 값)인 표를 가정한다.
        - 블록 높이가 소스≠대상이면 기본 건너뛰고(on_mismatch="skip") 보고한다(아래 블록 침범 방지).
          on_mismatch="src" 면 소스 높이대로 복사(대상 아래 행을 덮을 수 있으니 주의).
        반환: {"copied": 복사수, "missing": [소스에 없는 키...], "height_mismatch": [(키,소스h,대상h)...]}
          예) ctx.copy_key_blocks("531...로우데이터.xlsx!sheet", "콜센터", "B", "B", "N", "B3:B345", "B4:B89")
        """
        src_ctx, src_name = self._ctx_and_sheet_from_spec(src_sheet)
        dst_ctx, dst_name = self._ctx_and_sheet_from_spec(dst_sheet)
        ws_s = src_ctx._ws(src_name)
        ws_d = dst_ctx._ws(dst_name)
        self._tick(3)
        kc_s = int(src_ctx._resolve_col(src_name, key_col, 1))
        fc_s = int(src_ctx._resolve_col(src_name, first_col, 1))
        lc_s = int(src_ctx._resolve_col(src_name, last_col, 1))
        kc_d = int(dst_ctx._resolve_col(dst_name, key_col, 1))
        fc_d = int(dst_ctx._resolve_col(dst_name, first_col, 1))
        lc_d = int(dst_ctx._resolve_col(dst_name, last_col, 1))

        def blocks(ws, kc, scan):
            # 키 열의 (병합)블록을 위→아래로: (정규화키, 블록top행, 블록높이) 목록
            if scan:
                rng = ws.Range(str(scan))
                r0 = int(rng.Row); r1 = r0 + int(rng.Rows.Count) - 1
            else:
                r0 = 1; r1 = int(ws.Cells(ws.Rows.Count, kc).End(_XL_UP).Row)
            out = []
            r = r0
            while r <= r1:
                ma = ws.Cells(r, kc).MergeArea
                top = int(ma.Row); h = int(ma.Rows.Count)
                if top == r:
                    nk = _norm_key(ws.Cells(top, kc).Value)
                    if nk:
                        out.append((nk, top, h))
                    r = top + h
                else:
                    r = max(top + h, r + 1)   # 스캔이 블록 중간에서 시작한 경우 블록 끝으로 점프
            return out

        src_map = {}
        for nk, top, h in blocks(ws_s, kc_s, src_scan):
            if nk not in src_map:            # 소스에 같은 키가 여럿이면 첫 블록(Find 와 동일)
                src_map[nk] = (top, h)

        copied = 0
        missing = []
        mism = []
        for nk, dtop, dh in blocks(ws_d, kc_d, dst_scan):
            info = src_map.get(nk)
            if info is None:
                missing.append(nk)
                continue
            stop, sh = info
            if sh != dh:
                mism.append((nk, sh, dh))
                if on_mismatch == "skip":
                    continue
            use_h = sh if on_mismatch == "src" else min(sh, dh)
            dst_area = ws_d.Range(ws_d.Cells(dtop, fc_d), ws_d.Cells(dtop + use_h - 1, lc_d))
            try:
                dst_area.UnMerge()           # 소스 병합/서식을 그대로 입히기 위해 대상 블록 먼저 병합 해제
            except Exception:
                pass
            src_area = ws_s.Range(ws_s.Cells(stop, fc_s), ws_s.Cells(stop + use_h - 1, lc_s))
            src_area.Copy(ws_d.Cells(dtop, fc_d))   # 값+수식+서식+병합 보존(네이티브 복사)
            copied += 1
        try:
            self._app.CutCopyMode = False
        except Exception:
            pass
        self._shared["structural"].append(f"copy_key_blocks:{dst_name}!{_col_letter(kc_d)} x{copied}")
        try:
            _vba_trace("python_com.copy_key_blocks", src=str(src_name), dst=str(dst_name),
                       copied=copied, missing=len(missing), mismatch=len(mism))
        except Exception:
            pass
        return {"copied": copied, "missing": missing, "height_mismatch": mism}

    def sum_where(self, sheet, value_col, conditions, header_row=None):
        """조건(AND 전부 만족)에 맞는 행의 value_col 숫자를 합산해 값을 반환한다(쓰기 X → 반환값을
        ctx.write_cell 로 넣기). ★ 조건부 집계는 큰 표를 통째로 ctx.read 하지 말고 이 헬퍼를 쓸 것 —
        필요한 열만 좁게 읽어 계산하므로 대용량 표에서도 정적검사('큰 표를 ctx.read...')에 안 걸린다.
        - value_col / 조건 열: 열문자('M')·헤더명·번호 다 됨.
        - conditions: [(열, 값), ...] 또는 [(열, op, 값), ...]. op 생략 시 '=='(정규화 텍스트 일치).
          op ∈ {'==','!=','contains','>','>=','<','<='}. 텍스트는 normalize 로 비교(공백/전각 차이 흡수).
        - header_row 를 주면 그 다음 행부터(헤더/제목 행 제외). 헤더가 5행이면 header_row=5.
          예) t = ctx.sum_where("SO사업자별요금","M",[("D","인터넷"),("G","매 출")],header_row=5)
              ctx.write_cell("SO사업자별요금","AG4", t)"""
        ws = self._ws(sheet)
        self._tick(2)
        vc = int(self._resolve_col(sheet, value_col, header_row or 1))
        norm_conds = []
        for cond in (conditions or []):
            if len(cond) == 2:
                col, op, val = cond[0], "==", cond[1]
            else:
                col, op, val = cond[0], cond[1], cond[2]
            norm_conds.append((int(self._resolve_col(sheet, col, header_row or 1)), str(op), val))
        start = (int(header_row) + 1) if header_row else 1
        last = self.last_row(sheet, col=vc)
        if last < start:
            return 0.0
        need = {vc} | {c for (c, _, _) in norm_conds}
        colvals = {}
        for c in need:                      # 필요한 열만 각각 '좁게' 읽는다(광폭 read 회피)
            L = _col_letter(c)
            colvals[c] = self.read(sheet, f"{L}{start}:{L}{last}")
        total = 0.0
        for i in range(last - start + 1):
            def cell(c):
                col = colvals[c]
                return col[i][0] if (i < len(col) and col[i]) else None
            if all(_cond_match(cell(c), op, val) for (c, op, val) in norm_conds):
                num = _coerce_number(cell(vc))
                if num is not None:
                    total += num
        try:
            _vba_trace("python_com.sum_where", sheet=str(sheet), value=_col_letter(vc),
                       conds=len(norm_conds), start=start, last=last, total=total)
        except Exception:
            pass
        return total

    def sum_lookup(self, src_sheet, src_key_col, src_val_col, dst_sheet, dst_key_col, dst_out_col,
                   header_row=None, dst_start_row=None):
        """키 매칭 합산(교차/동일 파일): src 의 (키→값)을 모은 뒤, dst 각 행의 키에 해당하는 값을 합산해
        dst_out_col 같은 행에 쓴다. ★ dst 키 셀에 여러 토큰(줄바꿈/공백/콤마 구분)이 있으면 각각 분리해
        모두 더한다(한 셀에 가입번호가 여러 개인 KT/HCN 케이스). 전체 문자열로만 매칭해 0건 되던 문제 해결.
        - *_sheet: 시트명 또는 "파일.xlsx!시트"(교차파일). *_col: 열문자/헤더명/번호.
        - header_row: src·dst 공통 헤더행(그 다음 행부터). dst_start_row 로 dst 시작행을 따로 줄 수 있음.
        반환: {"filled": 값 채운 행 수, "src_keys": 소스 키 수}.
          예) ctx.sum_lookup("input.xlsx!Sheet1","BP","BQ","SO사업자별요금","P","H", header_row=1, dst_start_row=6)"""
        src_ctx, src_name = self._ctx_and_sheet_from_spec(src_sheet)
        dst_ctx, dst_name = self._ctx_and_sheet_from_spec(dst_sheet)
        ws_d = dst_ctx._ws(dst_name)
        self._tick(3)
        skc = int(src_ctx._resolve_col(src_name, src_key_col, header_row or 1))
        svc = int(src_ctx._resolve_col(src_name, src_val_col, header_row or 1))
        dkc = int(dst_ctx._resolve_col(dst_name, dst_key_col, header_row or 1))
        doc = int(dst_ctx._resolve_col(dst_name, dst_out_col, header_row or 1))
        s_start = (int(header_row) + 1) if header_row else 1
        # 키 열이 병합/희소면 End(xlUp) 이 표 하단을 놓쳐 과소산정한다 → used_last_row 로 보정.
        s_last = max(src_ctx.last_row(src_name, col=skc), src_ctx.used_last_row(src_name))
        skL, svL = _col_letter(skc), _col_letter(svc)
        keys = src_ctx.read(src_name, f"{skL}{s_start}:{skL}{s_last}") if s_last >= s_start else []
        vals = src_ctx.read(src_name, f"{svL}{s_start}:{svL}{s_last}") if s_last >= s_start else []
        kmap = {}
        for i in range(len(keys)):
            num = _coerce_number(vals[i][0] if (i < len(vals) and vals[i]) else None)
            if num is None:
                continue
            for tok in _split_key_tokens(keys[i][0] if keys[i] else None):
                kmap[tok] = kmap.get(tok, 0.0) + num
        d_start = int(dst_start_row) if dst_start_row else s_start
        d_last = max(dst_ctx.last_row(dst_name, col=dkc), dst_ctx.used_last_row(dst_name))
        dkL = _col_letter(dkc)
        dkeys = dst_ctx.read(dst_name, f"{dkL}{d_start}:{dkL}{d_last}") if d_last >= d_start else []
        filled = 0
        for i in range(len(dkeys)):
            toks = _split_key_tokens(dkeys[i][0] if dkeys[i] else None)
            if not toks:
                continue
            s, hit = 0.0, False
            for tok in toks:
                if tok in kmap:
                    s += kmap[tok]; hit = True
            if hit:
                ws_d.Cells(d_start + i, doc).Value = round(s, 4)
                filled += 1
        self._shared["structural"].append(f"sum_lookup:{dst_name}!{dkL}->{_col_letter(doc)} x{filled}")
        try:
            _vba_trace("python_com.sum_lookup", src=str(src_name), dst=str(dst_name),
                       filled=filled, src_keys=len(kmap))
        except Exception:
            pass
        return {"filled": filled, "src_keys": len(kmap)}

    def add_sheet(self, name, after=None):
        self._tick(3)
        names = _excel_collection_names(self._wb.Worksheets)
        if str(name) in names:
            raise PythonComSkillError(f"시트 '{name}' 이 이미 있습니다. 다른 이름을 쓰거나 먼저 삭제하세요.")
        anchor = self._wb.Worksheets(str(after)) if after else self._wb.Worksheets(self._wb.Worksheets.Count)
        ws = self._wb.Worksheets.Add(After=anchor)
        ws.Name = str(name)
        self._shared["structural"].append(f"add_sheet:{name}")
        return True

    def rename_sheet(self, old_name, new_name):
        """시트 이름만 변경한다(위치·내용 유지). '복사/이동'이 아니라 순수 이름 변경 전용."""
        ws = self._ws(old_name)
        self._tick(1)
        # Excel 시트명 31자 제한 — 초과분은 자른다(시트 '생성' 경로들과 동일 정책).
        # 실행기 매핑이 rename 목적지 리터럴을 긴 실제 시트명으로 치환해도 0x800A03EC 로 죽지 않고,
        # 이후 긴 이름 조회는 _ws 의 31자-절단 폴백이 같은 시트를 찾아준다.
        # [비대칭 수정] 절단만 하고 '금지문자'는 add_sheet 와 달리 그대로 넘겨서, 이름에 / : [ ] * ? \
        # 가 있으면(예: "매출/원가 요약") ws.Name 대입을 COM 이 거부해 '시트 이름 변경 실패'로 죽었다.
        # 생성 경로(add_sheet)와 같은 정규화를 적용해 정책을 일치시킨다.
        eff_new = re.sub(r"[\[\]:*?/\\]", "_", str(new_name))[:31]
        names = _excel_collection_names(self._wb.Worksheets)
        if eff_new != str(old_name) and eff_new in names:
            raise PythonComSkillError(f"시트 '{eff_new}' 이 이미 있습니다. 다른 이름을 쓰세요.")
        try:
            ws.Name = eff_new
        except Exception as e:
            raise PythonComSkillError(f"시트 이름 변경 실패: {e}")
        self._shared["structural"].append(f"rename_sheet:{old_name}->{eff_new}")
        return True

    def delete_sheet(self, name):
        ws = self._ws(name)
        self._tick(1)
        ws.Delete()
        self._shared["structural"].append(f"delete_sheet:{name}")
        return True

    def copy_sheet(self, src_sheet, dst_book=None, new_name=None, before=None, after=None):
        """시트 1장을 통째로 복사한다(서식·수식·값 보존). dst_book 을 주면 다른 파일로 복사(교차 파일).
        비파괴: 원본 시트는 그대로 둔다. '이동'은 복사 후 ctx.delete_sheet(원본) 를 호출하세요.
          ctx.copy_sheet("가시트", dst_book="출력.xlsx")            # 출력 파일 맨 뒤에 복사
          ctx.copy_sheet("요약", new_name="요약본", after="데이터")   # 같은 파일, 이름 지정
        대상이 다른 Excel 인스턴스에 있으면 임시 워크북을 매개로 복사한다. 구조 변경이라 실패 시 롤백되지 않는다."""
        ws = self._ws(src_sheet)
        self._tick(2)
        target_session = None
        target_app = self._app

        def _name_matches(wb_obj, wanted):
            try:
                actual = str(wb_obj.Name)
            except Exception:
                return False
            return (
                actual == wanted
                or str(Path(actual).stem) == str(Path(wanted).stem)
                or _workbook_name_lookup_key(actual) == _workbook_name_lookup_key(wanted)
            )

        if dst_book is not None:
            dst_key = str(dst_book).strip() if isinstance(dst_book, str) else ""
            dst_ctx = self.book(dst_book) if isinstance(dst_book, str) else dst_book
            dst_wb = dst_ctx._wb
            try:
                target_app = dst_wb.Application
            except Exception:
                target_app = self._app
            # [리뷰] 비공유/별도 앱 모드에서는 다른 파일이 읽기전용 스냅샷으로 열릴 수 있다 —
            # 그 경우 같은 이름의 실제 라이브 세션을 찾아 쓰기 대상으로 바꾼다.
            try:
                _dst_ro = bool(dst_wb.ReadOnly)
            except Exception:
                _dst_ro = False
            if _dst_ro and dst_key:
                for _sid, _other in list(EXCEL_SESSIONS.items()):
                    if not _other.get("liveEditable"):
                        continue
                    try:
                        _app, _wb = session_workbook(_other)
                    except Exception:
                        continue
                    try:
                        if _name_matches(_wb, dst_key) and not bool(_wb.ReadOnly):
                            target_session = _other
                            target_app = _app
                            dst_wb = _wb
                            # 뒤이어 같은 코드에서 ctx.book("대상")을 다시 부르면 쓰기 가능한 ctx가 나오게 갱신.
                            self._shared["books"][dst_key] = PythonComSkillContext(
                                target_app, dst_wb, target_session, _shared=self._shared
                            )
                            _dst_ro = False
                            break
                    except Exception:
                        continue
            if _dst_ro:
                raise PythonComSkillError(
                    f"대상 파일이 읽기 전용으로 열려 있어 시트를 복사할 수 없습니다. "
                    f"대상 파일 탭을 한 번 열어 라이브 세션을 만든 뒤 다시 시도해 주세요.")
        else:
            dst_wb = self._wb
        existing = list(_excel_collection_names(dst_wb.Worksheets))
        if new_name is not None and str(new_name) in existing:
            raise PythonComSkillError(f"대상 워크북에 시트 '{new_name}' 이 이미 있습니다. 다른 이름을 쓰세요.")
        # [중요] win32com 에서 Worksheet.Copy(After=...) 키워드 인자는 무동작이라(시트가 안 옮겨지고
        # 새 워크북만 생기거나 아무 일도 안 남), 반드시 positional 로 호출한다: Copy(Before, After).
        def _copy_to_target(copy_ws):
            if before is not None:
                copy_ws.Copy(dst_wb.Worksheets(str(before)))
            elif after is not None:
                copy_ws.Copy(pythoncom.Empty, dst_wb.Worksheets(str(after)))
            else:
                copy_ws.Copy(pythoncom.Empty, dst_wb.Worksheets(int(dst_wb.Worksheets.Count)))

        same_app = True
        try:
            same_app = _same_excel_app(self._app, target_app)
        except Exception:
            same_app = True
        if same_app:
            _copy_to_target(ws)
        else:
            # Excel은 서로 다른 Application 인스턴스 간 Worksheet.Copy 가 불안정하다.
            # 시트 1장짜리 임시 xlsx를 매개로 대상 앱 안에서 다시 Copy 한다.
            tmp_path = BACKEND_DIR / f"sheet_copy_{uuid.uuid4().hex}.xlsx"
            tmp_wb = None
            import_wb = None
            old_alerts_src = old_alerts_dst = None
            try:
                try:
                    old_alerts_src = self._app.DisplayAlerts
                    self._app.DisplayAlerts = False
                except Exception:
                    pass
                ws.Copy()
                tmp_wb = self._app.ActiveWorkbook
                tmp_wb.SaveAs(str(tmp_path), FileFormat=51)
                tmp_wb.Close(SaveChanges=False)
                tmp_wb = None
                try:
                    old_alerts_dst = target_app.DisplayAlerts
                    target_app.DisplayAlerts = False
                except Exception:
                    pass
                import_wb, _ = excel_workbooks_open(target_app, tmp_path, read_only=False)
                _copy_to_target(import_wb.Worksheets(1))
            finally:
                if tmp_wb is not None:
                    try:
                        tmp_wb.Close(SaveChanges=False)
                    except Exception:
                        pass
                if import_wb is not None:
                    try:
                        import_wb.Close(SaveChanges=False)
                    except Exception:
                        pass
                if old_alerts_src is not None:
                    try:
                        self._app.DisplayAlerts = old_alerts_src
                    except Exception:
                        pass
                if old_alerts_dst is not None:
                    try:
                        target_app.DisplayAlerts = old_alerts_dst
                    except Exception:
                        pass
                try:
                    Path(tmp_path).unlink(missing_ok=True)
                except Exception:
                    pass
        # 새로 생긴 시트를 이름 diff 로 식별한다(ActiveSheet 는 헤드리스에서 신뢰할 수 없어
        # 원본을 가리키는 경우가 있다 — 그걸 rename 하면 원본 시트명이 바뀌어 데이터 손실처럼 보인다).
        added = [n for n in _excel_collection_names(dst_wb.Worksheets) if n not in existing]
        if not added:
            raise PythonComSkillError(f"시트 '{src_sheet}' 복사가 적용되지 않았습니다(대상/위치를 확인하세요).")
        if new_name is not None:
            try:
                dst_wb.Worksheets(added[0]).Name = str(new_name)
            except Exception as err:
                raise PythonComSkillError(f"복사된 시트 이름을 '{new_name}' 으로 바꾸지 못했습니다: {err}")
        if target_session is not None:
            try:
                target_session["rev"] = int(target_session.get("rev") or 0) + 1
            except Exception:
                pass
        self._shared["structural"].append(f"copy_sheet:{src_sheet}->{dst_book or 'self'}")
        return True

    def append_same_format_sheets(self, src_books, dest_sheet="통합", src_sheet=None, header_row=None, scan_rows=30):
        """동일 포맷 여러 입력 파일의 표를 현재 워크북 새 시트에 이어붙인다.

        첫 파일은 헤더 포함, 이후 파일은 헤더 다음 행부터 Excel 네이티브 Copy 로 붙인다.
        상단 scan_rows 안에서 비어있지 않은 셀이 가장 많은 행을 헤더로 자동 탐지하므로
        1행이 빈 가입자별청구내역/청구내역 양식도 안전하게 처리한다. 값 배열 재작성 없이
        Range.Copy 를 사용하므로 긴 EID/가입번호, 날짜, 회계 서식이 보존된다.
        """
        books = [src_books] if isinstance(src_books, str) else list(src_books or [])
        if not books:
            raise PythonComSkillError("통합할 입력 파일 목록이 비어 있습니다.")

        def unique_sheet_name(wb, base):
            raw = str(base or "통합").strip() or "통합"
            raw = re.sub(r"[\[\]\:\*\?\/\\]", "_", raw)[:31] or "통합"
            names = set(_excel_collection_names(wb.Worksheets))
            if raw not in names:
                return raw
            n = 2
            while True:
                suffix = f"_{n}"
                candidate = (raw[:31 - len(suffix)] + suffix)[:31]
                if candidate not in names:
                    return candidate
                n += 1

        def first_or_named_ws(ctx, wanted):
            if wanted is None or str(wanted).strip() == "":
                ctx._tick(1)
                return ctx._wb.Worksheets(1)
            return ctx._ws(wanted)

        def row_values(ws, row, last_col):
            rng = ws.Range(ws.Cells(int(row), 1), ws.Cells(int(row), int(last_col)))
            return PythonComSkillContext._shaped_matrix(rng, rng.Value2)[0]

        def detect_bounds(ctx, ws):
            try:
                # UsedRange 는 과거 서식 흔적만으로도 16,384열까지 부풀 수 있다.
                # 실제 값/수식이 있는 마지막 셀을 먼저 찾아 샘플 read 범위를 작게 유지한다.
                last_row_cell = ws.Cells.Find(What="*", LookIn=-4123, SearchOrder=1, SearchDirection=2)
                last_col_cell = ws.Cells.Find(What="*", LookIn=-4123, SearchOrder=2, SearchDirection=2)
                if last_row_cell is not None and last_col_cell is not None:
                    used_last_row = max(1, int(last_row_cell.Row))
                    used_last_col = max(1, int(last_col_cell.Column))
                else:
                    used_last_row = 1
                    used_last_col = 1
            except Exception:
                used = ws.UsedRange
                ctx._tick(2)
                used_last_row = max(1, int(used.Row) + int(used.Rows.Count) - 1)
                used_last_col = max(1, int(used.Column) + int(used.Columns.Count) - 1)
            if header_row is not None:
                hdr = int(header_row)
                vals = row_values(ws, hdr, used_last_col)
            else:
                scan_end = min(used_last_row, max(1, int(scan_rows or 30)))
                sample_rng = ws.Range(ws.Cells(1, 1), ws.Cells(scan_end, used_last_col))
                sample = PythonComSkillContext._shaped_matrix(sample_rng, sample_rng.Value2)
                ctx._tick(2)
                hdr = 1
                best = -1
                vals = []
                for idx, row in enumerate(sample, start=1):
                    count = sum(1 for v in row if v is not None and str(v).strip() != "")
                    if count > best:
                        best = count
                        hdr = idx
                        vals = list(row)
                if best < 1:
                    raise PythonComSkillError(f"'{ws.Parent.Name}/{ws.Name}' 에서 헤더 행을 찾지 못했습니다.")
            header_cols = [i + 1 for i, v in enumerate(vals) if v is not None and str(v).strip() != ""]
            if not header_cols:
                raise PythonComSkillError(f"'{ws.Parent.Name}/{ws.Name}' 헤더 행({hdr})이 비어 있습니다.")
            last_col = max(header_cols)
            last_row = int(hdr)
            for c in header_cols:
                try:
                    r = int(ws.Cells(ws.Rows.Count, int(c)).End(_XL_UP).Row)
                    if r > last_row:
                        last_row = r
                except Exception:
                    continue
            return int(hdr), int(last_row), int(last_col)

        dest_name = unique_sheet_name(self._wb, dest_sheet)
        dst_ws = self._wb.Worksheets.Add(After=self._wb.Worksheets(self._wb.Worksheets.Count))
        dst_ws.Name = dest_name
        self._tick(3)
        dst_next = 1
        copied_rows = 0
        try:
            for i, book_name in enumerate(books):
                src_ctx = self.book(book_name)
                src_ws = first_or_named_ws(src_ctx, src_sheet)
                hdr, last_row, last_col = detect_bounds(src_ctx, src_ws)
                start_row = hdr if i == 0 else hdr + 1
                if last_row < start_row:
                    continue
                src_rng = src_ws.Range(src_ws.Cells(start_row, 1), src_ws.Cells(last_row, last_col))
                rows = int(src_rng.Rows.Count)
                src_rng.Copy(Destination=dst_ws.Cells(dst_next, 1))
                dst_next += rows
                copied_rows += rows
                self._tick(4)
            if copied_rows <= 0:
                raise PythonComSkillError("통합할 데이터가 없습니다.")
        except Exception:
            old_alerts = None
            try:
                old_alerts = self._app.DisplayAlerts
                self._app.DisplayAlerts = False
                dst_ws.Delete()
            except Exception:
                pass
            finally:
                if old_alerts is not None:
                    try:
                        self._app.DisplayAlerts = old_alerts
                    except Exception:
                        pass
            raise
        try:
            self._app.CutCopyMode = False
        except Exception:
            pass
        self._shared["structural"].append(f"append_same_format_sheets:{len(books)}->{dest_name}({copied_rows})")
        return dest_name

    def sort(self, sheet, a1_range, key_col, ascending=True, has_header=True):
        """실제 범위 정렬. key_col 은 범위 내 1-based 열 번호/'B' 열 문자, 또는 이들의 리스트(다중키).
        ascending 도 단일 bool 또는 키별 bool 리스트를 받는다."""
        ws = self._ws(sheet)
        rng = self._rng(ws, a1_range)
        self._tick(3)
        self._journal_save(ws, rng)
        keys = list(key_col) if isinstance(key_col, (list, tuple)) else [key_col]
        if not keys:
            raise PythonComSkillError("정렬 키가 비어 있습니다.")
        asc = list(ascending) if isinstance(ascending, (list, tuple)) else [ascending] * len(keys)
        while len(asc) < len(keys):
            asc.append(asc[-1] if asc else True)
        key_rngs = []
        for k in keys:
            if isinstance(k, str):
                s = k.strip()
                if re.fullmatch(r"[A-Za-z]{1,3}", s):
                    abs_idx = self._col_index(s)
                else:
                    # 열 문자가 아니면 헤더명으로 간주 → find_header 로 절대 열번호 조회.
                    # (LLM 이 key_col 에 "회사" 같은 헤더명을 넘기는 경우가 흔하다.)
                    abs_idx = self.find_header(sheet, s, header_row=1)
                key_idx = abs_idx - int(rng.Column) + 1
                self._tick(1)
            else:
                key_idx = int(k)
            if key_idx < 1 or key_idx > int(rng.Columns.Count):
                raise PythonComSkillError(f"정렬 키({k})가 범위를 벗어났습니다.")
            key_rngs.append(rng.Columns(key_idx))
            self._tick(1)
        # [중요] win32com 에서 rng.Sort(Header=...) 는 Header 인자가 무시돼 헤더행까지 정렬된다
        # (진단 확인). SortFields API(ws.Sort.Header)는 정상 동작하므로 이쪽을 쓴다. 다중키도 자연 지원.
        # 수식 결과 기준 정렬은 저장 캐시가 0/빈값이면 내림차순이 안 먹은 것처럼 보인다.
        # 정렬 직전 워크시트를 계산해 SortFields가 실제 표시값 기준으로 키를 잡게 한다.
        try:
            ws.Calculate()
        except Exception:
            for kr in key_rngs:
                try:
                    kr.Calculate()
                except Exception:
                    pass
        ws.Sort.SortFields.Clear()
        for i, kr in enumerate(key_rngs):
            # win32com 은 Add 의 named/positional Order 인자를 무시할 수 있어(asc 로 고정됨),
            # 반환된 SortField 의 Order 프로퍼티를 직접 설정한다(내림차순 보장).
            sf = ws.Sort.SortFields.Add(kr)
            sf.Order = (1 if asc[i] else 2)
            # 텍스트 식별자(EID/가입번호/전화번호 등)를 Excel이 숫자로 재해석하지 않도록
            # "텍스트를 숫자로 정렬" 옵션을 명시적으로 끈다. 정렬은 값/서식을 다시 쓰지 않고
            # Excel 네이티브 Sort가 행 전체를 이동시키게 둔다.
            try:
                sf.DataOption = 0  # xlSortNormal
            except Exception:
                pass
            self._tick(1)
        ws.Sort.SetRange(rng)
        ws.Sort.Header = (1 if has_header else 2)
        ws.Sort.Apply()
        ws.Sort.SortFields.Clear()
        self._tick(2)
        return True

    # ---- 표시/서식 ----
    def hide_cols(self, sheet, col_range, hidden=True):
        """예: ctx.hide_cols("매출", "B:D")"""
        ws = self._ws(sheet)
        self._tick(2)
        ws.Columns(str(col_range)).Hidden = bool(hidden)
        self._shared["structural"].append(f"hide_cols:{sheet}:{col_range}:{hidden}")
        return True

    def hide_rows(self, sheet, row_range, hidden=True):
        ws = self._ws(sheet)
        self._tick(2)
        ws.Rows(str(row_range)).Hidden = bool(hidden)
        self._shared["structural"].append(f"hide_rows:{sheet}:{row_range}:{hidden}")
        return True

    def merge(self, sheet, a1_range):
        ws = self._ws(sheet)
        rng = self._rng(ws, a1_range)
        self._tick(2)
        self._journal_save(ws, rng)
        rng.Merge()
        self._shared["structural"].append(f"merge:{sheet}:{a1_range}")
        return True

    def unmerge(self, sheet, a1_range):
        ws = self._ws(sheet)
        rng = self._rng(ws, a1_range)
        self._tick(2)
        rng.UnMerge()
        self._shared["structural"].append(f"unmerge:{sheet}:{a1_range}")
        return True

    def set_number_format(self, sheet, a1_range, fmt):
        ws = self._ws(sheet)
        rng = self._rng(ws, a1_range)
        self._tick(2)
        rng.NumberFormat = str(fmt)
        self._shared["structural"].append(f"number_format:{sheet}:{a1_range}")
        return True

    # ---- 데이터 헬퍼(조인/합계/중복/분리/치환) ----
    def _resolve_col(self, sheet, col, header_row=1):
        """열 지정을 1-based 번호로 해석한다. 'A' 같은 열 문자 / 1 같은 번호 / 헤더명 모두 허용."""
        if isinstance(col, bool):
            raise PythonComSkillError("열 지정이 잘못되었습니다.")
        if isinstance(col, (int, float)):
            return int(col)
        s = str(col).strip()
        if re.fullmatch(r"[A-Za-z]{1,3}", s):
            return self._col_index(s)
        return self.find_header(sheet, s, header_row=int(header_row))

    def lookup(self, sheet, key_col, into_col, table_sheet, table_key_col, table_val_col, header_row=1, default=None):
        """VLOOKUP/조인: sheet 의 key_col 값을 table_sheet 의 table_key_col 에서 찾아 그 행의 table_val_col 값을
        sheet 의 into_col 에 채운다(단가표→청구내역 단가 매칭 등). 키 비교는 normalize(공백/전각/대소문자/숫자-텍스트
        차이 무시)로 안전. 열은 'A' 문자/번호/헤더명 모두 허용. 미매칭은 default(없으면 빈칸). 반환: 매칭된 행 수."""
        kcol = self._resolve_col(sheet, key_col, header_row)
        icol = self._resolve_col(sheet, into_col, header_row)
        tkcol = self._resolve_col(table_sheet, table_key_col, header_row)
        tvcol = self._resolve_col(table_sheet, table_val_col, header_row)
        hr = int(header_row)
        t_last = self.last_row(table_sheet, tkcol)
        table_map = {}
        if t_last > hr:
            tkl, tvl = _col_letter(tkcol), _col_letter(tvcol)
            keys = [r[0] for r in self.read(table_sheet, "%s%d:%s%d" % (tkl, hr + 1, tkl, t_last))]
            vals = [r[0] for r in self.read(table_sheet, "%s%d:%s%d" % (tvl, hr + 1, tvl, t_last))]
            for k, v in zip(keys, vals):
                nk = self.normalize(k)
                if nk != "" and nk not in table_map:
                    table_map[nk] = v
        last = self.last_row(sheet, kcol)
        if last <= hr:
            return 0
        kcl, icl = _col_letter(kcol), _col_letter(icol)
        src_keys = [r[0] for r in self.read(sheet, "%s%d:%s%d" % (kcl, hr + 1, kcl, last))]
        out, matched = [], 0
        for k in src_keys:
            nk = self.normalize(k)
            if nk in table_map:
                out.append([table_map[nk]]); matched += 1
            else:
                out.append([default if default is not None else ""])
        self.write(sheet, "%s%d" % (icl, hr + 1), out)
        return matched

    def match_fill(self, source, target, columns, key=None,
                   source_header_row=1, header_row=1, rows=None,
                   aliases=None, allow_partial=False):
        """소스 표(예: 피벗)의 행을 대상 시트의 '키 열(구분명)'과 이름 매칭해서, 지정한 값 열들을 대상의
        해당 열에 '값만' 채운다. 이름이 완전히 일치하지 않아도 (정확→공백무시→기호무시→부분포함) 순서로
        자동 매칭하고, 확실히 못 맞춘 대상 이름은 '후보'와 함께 오류로 알려 한 번에 확정하게 한다.
        '피벗/요약값을 다른 시트에 이름 맞춰 붙여넣기/채우기' 요청의 기본 수단(손코딩 매칭 루프 금지).

        source/target : 시트명. 다른 파일이면 "파일.xlsx!시트" 형식(예: "input_...001.xlsx!MVNO상품명별요약").
        columns : {소스 값열: 대상 값열} 매핑(헤더명 또는 "B" 열문자). 예:
                  {"MVNO상품명_count":"건수", "수납금액_sum":"고객납부금액", "가입자당단가_도매대가_sum":"청구금액"}
        key : (소스 키열, 대상 키열). 생략 시 둘 다 A열. 헤더명/열문자/번호 허용.
        source_header_row : 소스 헤더 행(피벗 값표는 보통 1). header_row : 대상 헤더 행(예: 4).
        rows : 대상 데이터 행 (start, end). 생략 시 header_row+1 부터 키열 마지막 행까지(합계/소계 행은 자동 제외).
        aliases : {대상이름: 소스이름} 강제 매핑 — 리포트에 뜬 못 맞춘 이름을 확정할 때 넣어 재실행.
        allow_partial : True 면 못 맞춘 행은 건너뛰고 맞춘 것만 채운다(오류 없이). 기본 False.
        반환: {"matched": n, "unmatched": [대상이름...], "rows": (start,end)}."""
        import difflib as _difflib

        def _nlite(s):
            return normalize_text(s)  # 소문자 + 모든 공백 제거

        def _nhard(s):
            # 소문자화 후 한글/영숫자만 남긴다(괄호·밑줄·점·공백 제거). "안전제일(망개통용)" == "안전제일_망개통용".
            return re.sub(r"[^0-9a-z가-힣]", "", str(s or "").lower())

        def _is_summary(s):
            n = _nhard(s)
            # 단독 '계/합/합계계' 같은 짧은 총계 라벨(부분포함으로 오탐 안 나게 '정확' 판정)
            if n in ("계", "합", "합계", "소계", "총계", "누계", "총합", "합계계"):
                return True
            return any(w in n for w in ("합계", "소계", "총계", "누계", "부가세", "vat", "total", "subtotal", "grand"))

        def _num(v):
            try:
                return float(v)
            except (TypeError, ValueError):
                return None

        def _row_num(tok):
            m = re.search(r"(\d+)", str(tok if tok is not None else ""))
            return int(m.group(1)) if m else None

        def _parse_rows(rw):
            """rows 인자를 (start, end) 로 관대하게 해석. None/튜플(int·'A5')/문자열('A5:E11','5:11','A5')·
            열 정보는 무시하고 '행 번호'만 뽑는다. 끝을 못 정하면 None(→ 자동 last_row)."""
            if rw is None:
                return None
            if isinstance(rw, (list, tuple)):
                nums = [n for n in (_row_num(x) for x in rw) if n]
                if len(nums) >= 2:
                    return (min(nums), max(nums))
                if len(nums) == 1:
                    return (nums[0], None)
                return None
            s = str(rw).strip()
            if ":" in s:
                a, b = s.split(":", 1)
                ra, rb = _row_num(a), _row_num(b)
                if ra and rb:
                    return (min(ra, rb), max(ra, rb))
                if ra:
                    return (ra, None)
            n = _row_num(s)
            return (n, None) if n else None

        _AGG_SUF = ("합계", "총합", "소계", "총계", "sum", "count", "개수", "건수",
                    "평균", "average", "avg", "mean", "max", "min", "최대값", "최소값", "최대", "최소")

        def _agg_base(s):
            # 집계 접미사(합계/sum/개수/count 등)를 떼어 '기준 이름'만 남긴다("수납금액 합계"≈"수납금액_sum").
            n = _nhard(s)
            for suf in _AGG_SUF:
                if len(n) > len(suf) and n.endswith(suf):
                    return n[:-len(suf)]
            return n

        def _resolve_val_col(rc, sh, spec, hr):
            """값 열 해석: 열문자/번호/헤더(정확+부분포함) → 실패 시 집계접미사 무시 퍼지(수납금액 합계=수납금액_sum)."""
            if spec is None:
                raise PythonComSkillError("match_fill: 값 열 스펙이 비어 있습니다(소스/대상 열을 지정하세요).")
            try:
                return rc._resolve_col(sh, spec, hr)
            except PythonComSkillError:
                pass
            base = _agg_base(spec)
            if base:
                last_c = rc.last_col(sh, hr)
                row = rc.read(sh, "%s%d:%s%d" % (_col_letter(1), hr, _col_letter(max(1, last_c)), hr))
                hdrs = row[0] if row else []
                cands = []
                for i, h in enumerate(hdrs, start=1):
                    hb = _agg_base(h)
                    if hb and (hb == base or base in hb or hb in base):
                        cands.append((abs(len(hb) - len(base)), i))
                if cands:
                    cands.sort()
                    if len(cands) == 1 or cands[0][0] < cands[1][0]:
                        return cands[0][1]
            raise PythonComSkillError("match_fill: '%s' 시트에서 값 열 '%s' 을 찾지 못했습니다(열문자나 정확 헤더명을 쓰세요)." % (sh, spec))

        def _resolve_key_col(rc, sh, spec, hr):
            """키(구분명/이름) 열 해석. 값 열과 달리 '집계 열(_count/_sum/개수/합계)'에 부분매칭하면 안 된다
            — 피벗은 group_by 이름이 데이터필드 헤더(예: 'MVNO상품명_count')에 들어가 그 숫자 열을 잘못 잡는다.
            우선순위: 열문자/번호 → 정확/정규화 헤더 → 피벗 '행 레이블' 열(A) → 비집계 부분포함(유일) → A열."""
            if isinstance(spec, bool):
                raise PythonComSkillError("match_fill: 키 열 지정이 잘못되었습니다.")
            if isinstance(spec, (int, float)):
                return int(spec)
            s = str(spec).strip()
            if re.fullmatch(r"[A-Za-z]{1,3}", s):
                return rc._col_index(s)
            last_c = rc.last_col(sh, hr)
            hrow = rc.read(sh, "%s%d:%s%d" % (_col_letter(1), hr, _col_letter(max(1, last_c)), hr))
            hdr = hrow[0] if hrow else []
            # 1) 정확 / 정규화 일치
            for i, h in enumerate(hdr, 1):
                if h is not None and str(h).strip() == s:
                    return i
            ns = _nlite(s)
            for i, h in enumerate(hdr, 1):
                if h is not None and _nlite(h) == ns:
                    return i
            # 2) 피벗 '행 레이블'(Row Labels) 열 → A열
            c1 = _nlite(hdr[0]) if hdr else ""
            if c1 in ("행레이블", "rowlabels", "레이블", "labels") or "행레이블" in c1:
                return 1
            # 3) 비집계 열 중 부분포함 유일 매칭('_count'/'_sum' 등 집계 열은 후보에서 제외)
            cands = []
            for i, h in enumerate(hdr, 1):
                if h is None:
                    continue
                ht = _nhard(h)
                if not ht or ht != _agg_base(h):   # 접미사가 떨어지는 = 집계 열 → 키 후보 아님
                    continue
                if s in str(h) or str(h).strip() in s:
                    cands.append(i)
            if len(cands) == 1:
                return cands[0]
            # 4) 못 정하면 A열(피벗/요약표의 구분명은 대개 좌측 첫 열)
            return 1

        src_ctx, src_sheet = self._ctx_and_sheet_from_spec(source)
        tgt_ctx, tgt_sheet = self._ctx_and_sheet_from_spec(target)
        s_hr = max(1, int(source_header_row or 1))
        t_hr = max(1, int(header_row or 1))

        # 키 열 해석 — key 생략/None/튜플의 None 요소는 모두 첫 열(A, 구분명)로 기본 처리.
        if key is None:
            s_key_spec, t_key_spec = "A", "A"
        elif isinstance(key, (list, tuple)):
            s_key_spec = key[0] if len(key) > 0 and key[0] is not None else "A"
            t_key_spec = key[1] if len(key) > 1 and key[1] is not None else "A"
        else:
            s_key_spec = t_key_spec = key
        s_key = _resolve_key_col(src_ctx, src_sheet, s_key_spec, s_hr)
        t_key = _resolve_key_col(tgt_ctx, tgt_sheet, t_key_spec, t_hr)

        # 값 열 매핑 해석: [(소스열idx, 대상열idx)] — dict 또는 [[소스,대상],...] 리스트 허용, 집계접미사 무시 퍼지.
        if isinstance(columns, dict) and columns:
            col_items = list(columns.items())
        elif isinstance(columns, (list, tuple)) and columns and all(
                isinstance(x, (list, tuple)) and len(x) >= 2 for x in columns):
            col_items = [(x[0], x[1]) for x in columns]
        else:
            raise PythonComSkillError("match_fill: columns 는 {소스열: 대상열} 매핑(또는 [[소스,대상],...])이어야 합니다.")
        pairs = []
        for s_spec, t_spec in col_items:
            pairs.append((_resolve_val_col(src_ctx, src_sheet, s_spec, s_hr),
                          _resolve_val_col(tgt_ctx, tgt_sheet, t_spec, t_hr)))

        # 소스 읽기(키 + 값 열들, 헤더 다음 행부터 마지막 행까지)
        s_last = src_ctx.last_row(src_sheet, s_key)
        if s_last <= s_hr:
            raise PythonComSkillError("match_fill: 소스 '%s' 에 매칭할 데이터 행이 없습니다." % src_sheet)
        s_cols = [s_key] + [sc for sc, _ in pairs]
        s_lo, s_hi = min(s_cols), max(s_cols)
        s_block = src_ctx.read(src_sheet, "%s%d:%s%d" % (_col_letter(s_lo), s_hr + 1, _col_letter(s_hi), s_last))

        # 소스 인덱스 구성
        src_names = []            # [(raw, nlite, nhard)]
        src_vals = []             # [{대상열idx: 값}]
        exact_map, nlite_map, nhard_map = {}, {}, {}
        for ri, row in enumerate(s_block):
            raw = row[s_key - s_lo]
            name = "" if raw is None else str(raw).strip()
            if name == "":
                src_names.append(None); src_vals.append(None); continue
            vals = {t_idx: row[s_idx - s_lo] for s_idx, t_idx in pairs}
            src_names.append((name, _nlite(name), _nhard(name)))
            src_vals.append(vals)
            exact_map.setdefault(name, ri)
            nlite_map.setdefault(_nlite(name), ri)
            nhard_map.setdefault(_nhard(name), ri)

        alias_map = {}
        for tk, sk in (aliases or {}).items():
            alias_map[_nlite(tk)] = _nlite(sk)

        def _match_src(tname):
            """대상 이름 → 소스 행 인덱스(또는 None). aliases→정확→공백무시→기호무시→부분포함(유일 최선)."""
            nl, nh = _nlite(tname), _nhard(tname)
            if nl in alias_map:
                a = alias_map[nl]
                if a in nlite_map:
                    return nlite_map[a]
                for ri, ent in enumerate(src_names):
                    if ent and (a in ent[1] or ent[1] in a):
                        return ri
            if tname.strip() in exact_map:
                return exact_map[tname.strip()]
            if nl in nlite_map:
                return nlite_map[nl]
            if nh and nh in nhard_map:
                return nhard_map[nh]
            # 부분포함(양방향) — 유일한 '최소 잉여' 후보만 채택(모호하면 미매칭)
            if len(nh) >= 2:
                cands = []
                for ri, ent in enumerate(src_names):
                    if not ent:
                        continue
                    sh = ent[2]
                    if len(sh) >= 2 and (nh in sh or sh in nh):
                        cands.append((abs(len(sh) - len(nh)), ri))
                if cands:
                    cands.sort()
                    if len(cands) == 1 or cands[0][0] < cands[1][0]:
                        return cands[0][1]
            return None

        def _suggest(tname):
            pool = [ent[0] for ent in src_names if ent]
            near = _difflib.get_close_matches(tname, pool, n=1, cutoff=0.3)
            return near[0] if near else None

        def _combined_parts(tname):
            # 결합행("올인원+올인원2.0", "A / B", "A 및 B")은 각 부분이 모두 매칭되면 그 소스 인덱스 목록을 반환.
            # 부분 하나라도 못 맞추면 None(→ 통짜 매칭/미매칭 리포트로 폴백). 오탐 방지로 '전부 매칭'만 인정.
            parts = [p.strip() for p in re.split(r"\s*[+＋/／]\s*|\s+및\s+", tname) if p.strip()]
            if len(parts) < 2:
                return None
            idxs = []
            for p in parts:
                pri = _match_src(p)
                if pri is None:
                    return None
                idxs.append(pri)
            return idxs

        # 대상 행 범위 — rows 는 (start,end)/'A5:E11'/'5:11' 등 관대하게 해석(열 정보 무시). 끝 미정이면 last_row.
        parsed = _parse_rows(rows)
        if parsed:
            r0, r1 = parsed
            if r0 is None:
                r0 = t_hr + 1
            if r1 is None:
                r1 = tgt_ctx.last_row(tgt_sheet, t_key)
        else:
            r0 = t_hr + 1
            r1 = tgt_ctx.last_row(tgt_sheet, t_key)
        if r1 < r0:
            raise PythonComSkillError("match_fill: 대상 '%s' 에 채울 데이터 행이 없습니다." % tgt_sheet)
        t_keys = [r[0] for r in tgt_ctx.read(tgt_sheet, "%s%d:%s%d" % (_col_letter(t_key), r0, _col_letter(t_key), r1))]

        matched, unmatched, fills = {}, [], {t_idx: {} for _, t_idx in pairs}  # fills: 대상열idx -> {row: 값}
        for off, raw in enumerate(t_keys):
            row = r0 + off
            tname = "" if raw is None else str(raw).strip()
            if tname == "" or _is_summary(tname):
                continue  # 빈 행/합계·소계·'계' 등 요약 행은 건드리지 않는다
            ri = _match_src(tname)
            if ri is not None:
                matched[row] = ri
                for t_idx, v in src_vals[ri].items():
                    fills[t_idx][row] = v
                continue
            # 통짜 매칭 실패 → 결합행("A+B")이면 각 부분 값 합산
            combo = _combined_parts(tname)
            if combo:
                matched[row] = combo
                for t_idx in fills:
                    tot, anynum = 0.0, False
                    for cri in combo:
                        nv = _num(src_vals[cri].get(t_idx))
                        if nv is not None:
                            tot += nv
                            anynum = True
                    fills[t_idx][row] = tot if anynum else None
                continue
            unmatched.append((tname, _suggest(tname)))

        if unmatched and not allow_partial:
            lines = []
            for nm, cand in unmatched:
                lines.append("  · '%s'%s" % (nm, (" — 혹시 '%s'?" % cand) if cand else " — 후보 없음"))
            raise PythonComSkillError(
                "대상 '%s' 의 다음 이름을 소스 '%s' 에서 확실히 매칭하지 못했습니다:\n%s\n"
                "확정하려면 aliases={'대상이름':'소스이름', ...} 로 다시 실행하세요"
                "(맞춘 것만 우선 채우려면 allow_partial=True)."
                % (tgt_sheet, src_sheet, "\n".join(lines))
            )

        # 쓰기: 대상 값열별로 '매칭된 행'만(연속 구간 묶어서). 키/요약/미매칭 셀은 안 건드림.
        for t_idx, rowvals in fills.items():
            if not rowvals:
                continue
            cl = _col_letter(t_idx)
            rs = sorted(rowvals)
            i = 0
            while i < len(rs):
                j = i
                while j + 1 < len(rs) and rs[j + 1] == rs[j] + 1:
                    j += 1
                run = rs[i:j + 1]
                tgt_ctx.write(tgt_sheet, "%s%d" % (cl, run[0]), [[rowvals[r]] for r in run])
                i = j + 1

        self._shared["structural"].append(
            "match_fill:%s->%s:%d matched" % (src_sheet, tgt_sheet, len(matched)))
        return {"matched": len(matched), "unmatched": [nm for nm, _ in unmatched], "rows": (r0, r1)}

    def add_total_row(self, sheet, sum_cols, label_col=None, label="합계", header_row=1):
        """표 끝(마지막 데이터행 바로 아래)에 합계 행을 만든다. sum_cols(열 리스트/단일)에 =SUM(데이터범위) 수식을
        넣고, label_col 이 있으면 그 셀에 label 을 쓴다. 열은 'A'/번호/헤더명 허용. 반환: 합계행 번호."""
        cols = sum_cols if isinstance(sum_cols, (list, tuple)) else [sum_cols]
        hr = int(header_row)
        idx = [self._resolve_col(sheet, c, hr) for c in cols]
        last = hr
        for c in idx:
            last = max(last, self.last_row(sheet, c))
        if last <= hr:
            raise PythonComSkillError("합계를 만들 데이터 행이 없습니다.")
        total_row = last + 1
        if label_col is not None:
            lcl = _col_letter(self._resolve_col(sheet, label_col, hr))
            self.write(sheet, "%s%d" % (lcl, total_row), [[label]])
        for c in idx:
            cl = _col_letter(c)
            self.write_formulas(sheet, "%s%d" % (cl, total_row),
                                [["=SUM(%s%d:%s%d)" % (cl, hr + 1, cl, last)]])
        return total_row

    def dedupe(self, sheet, key_cols, header_row=1, keep="first"):
        """key_cols(열 리스트/단일) 조합이 같은 중복 행을 삭제한다. keep='first'면 처음 것, 'last'면 마지막 것을
        남긴다. 비교는 normalize 기준. 헤더 행은 보존. 반환: 삭제된 행 수.
        (행 단위 삭제라 중복이 매우 많은 대용량은 느림 — 그 경우 VBA AutoFilter 경로 권장)"""
        cols = key_cols if isinstance(key_cols, (list, tuple)) else [key_cols]
        hr = int(header_row)
        idx = [self._resolve_col(sheet, c, hr) for c in cols]
        last = hr
        for c in idx:
            last = max(last, self.last_row(sheet, c))
        if last <= hr:
            return 0
        keycells = []
        for c in idx:
            cl = _col_letter(c)
            keycells.append([r[0] for r in self.read(sheet, "%s%d:%s%d" % (cl, hr + 1, cl, last))])
        n = last - hr
        seen, dup_rows = set(), []
        order = range(n) if str(keep) != "last" else range(n - 1, -1, -1)
        for i in order:
            key = tuple(self.normalize(keycells[j][i]) for j in range(len(idx)))
            if all(k == "" for k in key):
                continue
            if key in seen:
                dup_rows.append(hr + 1 + i)
            else:
                seen.add(key)
        removed = 0
        for r in sorted(set(dup_rows), reverse=True):
            self.delete_rows(sheet, r, 1)
            removed += 1
        return removed

    def split_column(self, sheet, col, delimiter, into=None, header_row=1):
        """col 셀을 delimiter 로 나눠 col 바로 오른쪽의 새 열들에 기록한다(예: "1001/홍길동" → 가입번호 / 고객명).
        into 가 있으면 그 개수만큼 새 열을 만들고 헤더로 쓴다(없으면 가장 많은 조각 수만큼). 원본 col 은 보존.
        열은 'A'/번호/헤더명 허용. 반환: 처리한 데이터 행 수."""
        ccol = self._resolve_col(sheet, col, header_row)
        hr = int(header_row)
        last = self.last_row(sheet, ccol)
        if last <= hr:
            return 0
        cl = _col_letter(ccol)
        src = [r[0] for r in self.read(sheet, "%s%d:%s%d" % (cl, hr + 1, cl, last))]
        d = str(delimiter)
        parts = [(str("" if v is None else v).split(d) if d != "" else [str("" if v is None else v)]) for v in src]
        width = len(into) if into else max((len(p) for p in parts), default=1)
        if width < 1:
            width = 1
        self.insert_cols(sheet, ccol + 1, width)
        for j in range(width):
            tcl = _col_letter(ccol + 1 + j)
            if into and j < len(into):
                self.write(sheet, "%s%d" % (tcl, hr), [[into[j]]])
            self.write(sheet, "%s%d" % (tcl, hr + 1), [[(p[j] if j < len(p) else "")] for p in parts])
        return len(src)

    def replace(self, sheet, a1_range, find, repl, match_entire=False):
        """범위 안 셀에서 find 를 repl 로 바꾼다(부분 치환, match_entire=True면 셀 전체 일치만). 수식 셀은 보존.
        반환: 바뀐 셀 수."""
        ws = self._ws(sheet)
        rng = self._rng(ws, a1_range)
        self._tick(2)
        vals = _range_matrix(rng.Value2)
        formulas = _range_matrix(rng.Formula)
        find_s = str(find)
        repl_s = "" if repl is None else str(repl)
        changed = 0
        out = []
        for ri, row in enumerate(vals):
            orow = []
            for ci, v in enumerate(row):
                fcell = formulas[ri][ci] if (ri < len(formulas) and ci < len(formulas[ri])) else None
                if isinstance(fcell, str) and fcell.startswith("="):
                    orow.append(fcell)  # 수식 그대로 보존
                    continue
                if v is None:
                    orow.append("")
                    continue
                sv = str(v)
                if match_entire:
                    if sv == find_s:
                        orow.append(repl_s); changed += 1
                    else:
                        orow.append(v)
                else:
                    if find_s != "" and find_s in sv:
                        orow.append(sv.replace(find_s, repl_s)); changed += 1
                    else:
                        orow.append(v)
            out.append(orow)
        if changed:
            self._journal_save(ws, rng)
            rng.Formula = out  # 값셀만 치환, 수식셀은 위에서 원래 수식 유지
            self._tick(1)
        return changed

    # ---- 교차 파일 ----
    def book(self, workbook_name):
        """같은 Excel 인스턴스에 열린 다른 업로드 파일을 대상으로 하는 ctx.
        예: out = ctx.book("output_검증파일.xlsx"); out.write(...)"""
        key = str(workbook_name).strip()
        if key in self._shared["books"]:
            return self._shared["books"][key]
        self._tick(2)
        target = None
        requested_keys = _workbook_name_lookup_keys(key)
        # [정확명 우선] 예전엔 정확명·stem·별칭키를 한 루프에서 같이 보고 '첫 히트'에 break 했다.
        # 그래서 '정산.csv' 를 요청해도 컬렉션 순서상 앞에 있는 '정산.xlsx'(stem 동일, 확장자만 다름)가
        # 가로채, 정확히 그 이름으로 열려 있는 파일을 두고 엉뚱한 워크북에 썼다.
        # 정확명 → 별칭키 → stem 순으로 '패스를 나눠' 본다(_resolve_open_workbook_name 과 같은 원칙).
        try:
            books = list(self._app.Workbooks)
        except Exception:
            books = []

        def _pick(pred):
            for wb in books:
                try:
                    if pred(str(wb.Name)):
                        return wb
                except Exception:
                    continue
            return None

        target = _pick(lambda n: n == key)
        if target is None:
            target = _pick(lambda n: bool(_workbook_name_lookup_keys(n) & requested_keys))
        if target is None:
            # stem 매칭(확장자 무시)은 가장 느슨하므로 '유일할 때만' 따라간다 —
            # 확장자만 다른 동명 파일이 둘 이상이면 모호하므로 조용히 아무거나 잡지 않는다.
            stem_hits = []
            for wb in books:
                try:
                    if str(Path(str(wb.Name)).stem) == str(Path(key).stem):
                        stem_hits.append(wb)
                except Exception:
                    continue
            if len(stem_hits) == 1:
                target = stem_hits[0]
        if target is None:
            # 공백/_/- 만 다른 파일명도 매칭(모델이 한글 파일명에 공백을 끼우는 케이스: '기업DW추출'→'기업 DW 추출').
            # 정규화 후 '정확히 한 워크북'에만 맞을 때만 따라간다(모호하면 기존 에러).
            try:
                want = normalize_sheet_lookup(Path(key).stem)
                nm = []
                for wb in self._app.Workbooks:
                    try:
                        if normalize_sheet_lookup(Path(str(wb.Name)).stem) == want:
                            nm.append(wb)
                    except Exception:
                        continue
                if len(nm) == 1:
                    target = nm[0]
            except Exception:
                pass
        if target is None:
            # [리뷰#8] 포맷 위장 파일(.xls=HTML/CSV)은 excel_open_<uuid> 로 리네임돼 열려 위 매칭이 전부 실패한다.
            # VBA(_normalize_vba_workbook_literals)와 동일하게 등록명→실제명 별칭으로 풀어 그 워크북을 찾는다
            # (VBA 는 되는데 Python ctx.book 만 못 찾던 경로 의존 비대칭 해소).
            try:
                aliased = _alias_open_workbook_name(self._app, key)
                if aliased and aliased != key:
                    for wb in self._app.Workbooks:
                        try:
                            if str(wb.Name) == aliased:
                                target = wb
                                break
                        except Exception:
                            continue
            except Exception:
                pass
        if target is None:
            # [월/날짜만 다른 저장 스킬 재사용] 4월용 스킬을 5월 파일에 돌릴 때: 날짜·월·버전·순번을 뺀
            # 안정 키로 '유일' 매칭이면 그 파일로 바인딩(모호하면 매칭 안 함 → 아래 오류 유지).
            try:
                names = [str(wb.Name) for wb in self._app.Workbooks]
                stable = _match_workbook_by_stable_key(names, key)
                if stable:
                    for wb in self._app.Workbooks:
                        try:
                            if str(wb.Name) == stable:
                                target = wb
                                _vba_trace("python_com.book.stable_key_match", requested=str(key), matched=str(stable))
                                break
                        except Exception:
                            continue
            except Exception:
                pass
        if target is None:
            # [지원성] 어떤 파일이 열려 있는지 함께 알려줘야 이름 불일치(중복다운로드 접미사·본부별
            # 파일명 차이 등)를 사용자가/지원팀이 바로 식별할 수 있다.
            # 단, 내부 작업본 이름(excel_open_<uuid>, <hash>_원본명, PERSONAL.XLSB 등)은 보여주면 안 된다 —
            # 이 에러는 적용 게이트의 LLM 재시도 루프로 되먹임되므로, 모델이 그 내부명을 그대로 코드에
            # 박아버리면(별칭 체계 우회) 재실행 때 uuid 가 달라져 반드시 깨진다.
            open_names = []
            try:
                open_names = _user_facing_workbook_names(self._app)
            except Exception:
                pass
            hint = ""
            if open_names:
                shown = ", ".join(f"'{n}'" for n in open_names[:6])
                more = f" 외 {len(open_names) - 6}개" if len(open_names) > 6 else ""
                hint = f" 현재 열린 파일: {shown}{more}."
            raise PythonComSkillError(
                f"워크북 '{workbook_name}' 이 열려 있지 않습니다. 업로드된 파일명을 그대로 쓰세요.{hint}"
            )
        sub = PythonComSkillContext(self._app, target, self._session, _shared=self._shared)
        self._shared["books"][key] = sub
        return sub

    # ---- 마무리/롤백 ----
    def _changed(self):
        return bool(self._shared["journal"]) or bool(self._shared["structural"])

    def _rollback(self):
        """실패 시 저널 역순 복원(쓰기 범위만 정밀 원복). 구조 변경은 롤백 불가."""
        restored = 0
        for ws_name, address, formulas in reversed(self._shared["journal"]):
            try:
                ws = self._wb.Worksheets(ws_name)
                data = tuple(tuple("" if v is None else v for v in row) for row in formulas)
                if data:
                    ws.Range(address).Formula = data
                    restored += 1
            except Exception:
                continue
        return restored, bool(self._shared["structural"])

    def summary(self):
        return {
            "comCalls": self._shared["com_calls"],
            "writes": len(self._shared["journal"]),
            "structural": list(self._shared["structural"]),
        }


def _python_com_static_check(code):
    """실행 전 AST 정적 게이트. 위반은 사람이 읽을 수 있는 한국어 사유로 모아 한 번에 반환."""
    import ast as _ast
    failures = []
    code_text = str(code or "")
    # Field regression: Python COM can hang the UI on multi-file, multi-token
    # lookup/aggregation jobs even when the generated code is syntactically
    # "bulk-ish". This class should be generated/executed as VBA.
    if (re.search(r"\bctx\s*\.\s*book\s*\(", code_text, re.I)
            and re.search(r"\b(?:split|re\s*\.\s*split)\s*\(", code_text, re.I)
            and re.search(r"(?:BP|BQ|P:P|H:H|token|tokens|account|key|가입)", code_text, re.I)
            and re.search(r"\b(?:sum|total|amount|fee)\b", code_text, re.I)
            and re.search(r"\bctx\s*\.\s*(?:write|write_cell)\s*\(", code_text, re.I)):
        failures.append(
            "다중 토큰 매칭/합산/쓰기 작업은 Python COM으로 실행하지 마세요. "
            "현장 멈춤 재현 패턴이므로 VBA(Scripting.Dictionary + 배열 처리)로 작성해야 합니다."
        )
    if (re.search(r"\bctx\s*\.\s*read\s*\(", code_text, re.I)
            and re.search(r"(?:\bsorted\s*\(|\.\s*sort\s*\()", code_text, re.I)
            and re.search(r"\bctx\s*\.\s*(?:write|write_cell)\s*\(", code_text, re.I)):
        failures.append(
            "정렬을 ctx.read → Python sorted/list.sort → ctx.write 로 구현하면 헤더/행 관계가 깨지고 "
            "긴 숫자 식별자가 8.90E+31 형태로 손실될 수 있습니다. ctx.sort(...)를 사용하세요."
        )
    # [소수점 쪼개기 차단] '연속 숫자만' findall + 콤마 join 은 '20.0' → '20','0' → "20, 0" 오답
    # (실측: 한화테크윈 DSMC ':' 뒤 숫자 — 프롬프트 규칙을 모델이 반복 위반해 게이트로 승격).
    # ''.join(숫자 이어붙이기) 같은 정상 패턴은 콤마 join 이 아니라서 통과한다.
    if (re.search(r"\bfind(?:all|iter)\s*\(\s*r?['\"](?:\\d\+|\[0-9\]\+)['\"]", code_text)
            and re.search(r"['\"],\s?['\"]\s*\.\s*join\s*\(", code_text)):
        failures.append(
            "re.findall 의 숫자 패턴이 '연속 숫자만'(\\d+)이라 '20.0' 같은 소수점 값을 '20'과 '0'으로 "
            "쪼개 콤마 나열합니다. 소수점 포함 r'\\d+(?:\\.\\d+)?' 패턴을 쓰거나 구분자로 자른 조각을 "
            "통째로 기입하고, 매칭이 1개면 join 나열 대신 그 값 하나만 쓰세요."
        )

    def _col_to_index(col):
        n = 0
        for ch in str(col or "").upper():
            if "A" <= ch <= "Z":
                n = n * 26 + (ord(ch) - 64)
        return n

    def _a1_cells_estimate(a1):
        s = str(a1 or "").replace("$", "").strip()
        m = re.match(r"^([A-Z]{1,3})(\d+)?\s*:\s*([A-Z]{1,3})(\d+)?$", s, re.I)
        if not m:
            return None
        c1, c2 = _col_to_index(m.group(1)), _col_to_index(m.group(3))
        if not c1 or not c2:
            return None
        cols = abs(c2 - c1) + 1
        if not m.group(2) and not m.group(4):
            return float("inf")
        if not m.group(2) or not m.group(4):
            return None
        r1, r2 = int(m.group(2)), int(m.group(4))
        return (abs(r2 - r1) + 1) * cols

    def _dynamic_range_text_is_wide(a1):
        s = str(a1 or "").replace("$", "").strip()
        # 폭을 알 수 없는 동적 열(전체 열/열문자 계산/사용범위)은 항상 '넓음'으로 본다.
        if re.search(r"last_col|col_letter|UsedRange", s, re.I):
            return True
        m = re.match(r"^([A-Z]{1,3})\d+\s*:\s*([A-Z]{1,3})(?:\d+|\{[^}]+\})$", s, re.I)
        if not m:
            return False
        # [오탐 완화] 시작·끝 '행'이 명시된(끝행이 변수 {n} 여도) '좁은 열 스팬(≤8열)' 읽기는 값-요약/이름매칭
        # 같은 소형 작업의 정상 패턴이다. 예전엔 열이 2개만 돼도 무조건 막아, A2:D{last} 같은 8행짜리 값
        # 붙여넣기가 차단되고 불필요하게 VBA 로 넘어갔다. 폭이 넓은(>8열) 동적 읽기만 '대용량 위험'으로 막는다.
        # (전체 열 A:D 는 위 _a1_cells_estimate 가 inf 로 잡아 risky_read 로 막으므로 여기 영향 없음.)
        return abs(_col_to_index(m.group(2)) - _col_to_index(m.group(1))) + 1 > 8

    has_read_call = bool(re.search(r"\b(?:ctx|[A-Za-z_]\w*)\s*\.\s*read\s*\(", code_text))
    has_data_move = bool(re.search(
        r"\b(?:ctx|[A-Za-z_]\w*)\s*\.\s*(?:write|write_cell|write_formulas|copy|copy_sheet|filter_to_sheet|sort)\s*\(",
        code_text,
        re.I,
    ))
    has_python_transform = bool(re.search(
        r"\bfor\s+\w+\s+in\s+\w+|\bsorted\s*\(|\.\s*sort\s*\(|\.\s*append\s*\(|\bfilter\s*\(|\blambda\b",
        code_text,
        re.I,
    ))
    if has_read_call and has_data_move and has_python_transform:
        risky_read = False
        for m in re.finditer(r"\.\s*read\s*\(\s*[^,\n]+,\s*[fF]?([\"'])([^\"']+)\1", code_text, re.I):
            cells = _a1_cells_estimate(m.group(2))
            if cells == float("inf") or (cells is not None and cells >= 200000):
                risky_read = True
                break
        dynamic_wide_read = False
        for m in re.finditer(r"\.\s*read\s*\(\s*[^,\n]+,\s*f([\"'])([^\"']+)\1", code_text, re.I):
            if _dynamic_range_text_is_wide(m.group(2)):
                dynamic_wide_read = True
                break
        if not dynamic_wide_read:
            for m in re.finditer(
                r"\b(?:rng|range_|a1|src_range|read_range)\w*\s*=\s*f([\"'])([^\"']+)\1",
                code_text,
                re.I,
            ):
                if _dynamic_range_text_is_wide(m.group(2)):
                    dynamic_wide_read = True
                    break
        if not dynamic_wide_read:
            dynamic_wide_read = bool(
                re.search(r"\.\s*read\s*\(\s*[^,\n]+,\s*(?:rng|range_|a1|src_range|read_range)\w*\b", code_text, re.I)
                and re.search(r"\blast_col\b|col_letter|UsedRange", code_text, re.I)
            )
        if risky_read or dynamic_wide_read:
            failures.append(
                "큰 표를 ctx.read 로 Python 리스트에 올려 가공한 뒤 다시 쓰거나 복사하지 마세요. "
                "대용량 파일에서 WebView/COM 응답이 멈추고 긴 숫자·날짜·서식이 손실될 수 있습니다. "
                "복사/이어붙이기는 ctx.copy 또는 ctx.append_same_format_sheets, 정렬은 ctx.sort, "
                "필터 새 시트는 VBA AutoFilter/전용 헬퍼를 사용하세요."
            )
    try:
        tree = _ast.parse(code)
    except SyntaxError as err:
        raise PythonComSkillError(f"Python 문법 오류: {err}")

    has_entry = False
    forbidden_names = {
        "open", "eval", "exec", "compile", "__import__", "input", "globals", "locals",
        "vars", "getattr", "setattr", "delattr", "exit", "quit", "breakpoint", "help",
    }
    forbidden_attrs = {"Select", "Activate", "ActiveWorkbook", "ActiveSheet", "Application", "Quit",
                       "Save", "SaveAs", "SaveCopyAs", "Close"}
    # copy/clear/delete_rows/delete_cols 는 셀이 아니라 범위·구조 단위라 여러 열 재배치·중복 제거·
    # 양식 비우기 등에서 루프 반복이 정당하다(셀 단위 남용은 런타임 COM 예산 PY_COM_BUDGET 로 차단).
    # 그래서 루프 차단 대상에서 제외한다.
    write_ops = {"write", "write_cell", "write_formulas", "insert_rows",
                 "insert_cols", "merge", "unmerge", "sort"}

    loop_stack = []
    # 루프 내 쓰기 금지는 'ctx 계열' 수신자에만 적용한다 — 일반 리스트/딕셔너리의
    # .copy()/.sort()/.clear() 는 메모리 연산이라 무해한데, 수신자 확인 없이 이름만 보면
    # for r in rows: out.append(r.copy()) 같은 흔한 관용구가 전부 오탐으로 차단된다.
    ctx_aliases = {"ctx"}

    def _is_ctx_receiver(value):
        if isinstance(value, _ast.Name):
            return value.id in ctx_aliases
        if isinstance(value, _ast.Call):
            f = value.func
            return isinstance(f, _ast.Attribute) and f.attr == "book" and _is_ctx_receiver(f.value)
        return False

    class _Checker(_ast.NodeVisitor):
        def visit_Assign(self, node):
            # book = ctx.book("다른파일.xlsx") 별칭 추적(별칭의 루프 내 쓰기도 잡기 위함).
            if isinstance(node.value, _ast.Call):
                f = node.value.func
                if isinstance(f, _ast.Attribute) and f.attr == "book" and _is_ctx_receiver(f.value):
                    for tgt in node.targets:
                        if isinstance(tgt, _ast.Name):
                            ctx_aliases.add(tgt.id)
            self.generic_visit(node)

        def visit_Import(self, node):
            failures.append("import 는 사용할 수 없습니다(re/datetime/math 는 이미 주어져 있음).")

        def visit_ImportFrom(self, node):
            failures.append("import 는 사용할 수 없습니다(re/datetime/math 는 이미 주어져 있음).")

        def visit_FunctionDef(self, node):
            nonlocal has_entry
            if node.name == PY_SKILL_ENTRY:
                has_entry = True
            self.generic_visit(node)

        def visit_While(self, node):
            if isinstance(node.test, _ast.Constant) and node.test.value is True:
                failures.append("while True 무한 루프는 금지입니다.")
            loop_stack.append("while")
            self.generic_visit(node)
            loop_stack.pop()

        def visit_For(self, node):
            loop_stack.append("for")
            self.generic_visit(node)
            loop_stack.pop()

        def visit_Call(self, node):
            func = node.func
            if isinstance(func, _ast.Name) and func.id in forbidden_names:
                failures.append(f"{func.id}() 는 사용할 수 없습니다.")
            if isinstance(func, _ast.Attribute):
                if func.attr in forbidden_attrs:
                    failures.append(f".{func.attr} 는 사용할 수 없습니다(ctx API 만 사용).")
                if loop_stack and func.attr in write_ops and _is_ctx_receiver(func.value):
                    failures.append(
                        f"루프 안에서 ctx.{func.attr}() 를 반복 호출하면 안 됩니다. "
                        "데이터를 메모리(리스트)에서 모두 계산한 뒤 ctx.write() 한 번으로 쓰세요."
                    )
            self.generic_visit(node)

        def visit_Attribute(self, node):
            if isinstance(node.value, _ast.Name) and node.value.id in {"win32com", "openpyxl", "os", "sys"}:
                failures.append(f"{node.value.id} 모듈은 사용할 수 없습니다(ctx API 만 사용).")
            self.generic_visit(node)

        def visit_Subscript(self, node):
            # openpyxl 관용구 ws["A1"] 차단(이 빌드의 Python 은 openpyxl 이 아님).
            if isinstance(node.value, _ast.Name) and node.value.id in {"ws", "sheet", "worksheet"}:
                failures.append(
                    'ws["A1"] 식 openpyxl 관용구는 지원되지 않습니다. '
                    "ctx.read()/ctx.write() 벌크 API 를 사용하세요."
                )
            self.generic_visit(node)

    _Checker().visit(tree)
    if not has_entry:
        failures.append(f"def {PY_SKILL_ENTRY}(ctx): 진입 함수가 필요합니다.")
    if failures:
        # 중복 사유 정리
        unique = list(dict.fromkeys(failures))
        raise PythonComSkillError("정적 검사 위반:\n- " + "\n- ".join(unique))


def _step_extended_timeout_s(st):
    """스텝 dict 이 extendedTimeout=True(VBA→Python 복구/강제 대용량)면 확장 데드라인(초)을, 아니면 None 을
    돌려준다. 파이프라인 스텝 실행 경로에서 이 스텝만 75초 데드라인 대신 복구용 데드라인을 쓰게 한다."""
    try:
        if isinstance(st, dict) and st.get("extendedTimeout") is True:
            return PY_SKILL_RECOVERY_TIMEOUT_S
    except Exception:
        pass
    return None


def _exec_python_com_skill(app, wb, session, code, skip_static=False, timeout_s=None):
    """샌드박스 exec + 데드라인 트레이서로 생성 Python 스킬을 실행한다.
    반환: ctx.summary(). 실패 시 저널 롤백 후 PythonComSkillError 재전파.
    timeout_s: 지정 시 ctx/트레이서 데드라인을 이 값(초)으로 확장(복구/강제 대용량 경로)."""
    if not skip_static:
        _python_com_static_check(code)
    ctx = PythonComSkillContext(app, wb, session, timeout_s=timeout_s)
    safe_globals = {
        "__builtins__": dict(_PY_SAFE_BUILTINS),
        "re": re,
        "datetime": datetime,
        "math": math,
    }
    try:
        exec(compile(code, "<b2b_python_skill>", "exec"), safe_globals)
    except PythonComSkillError:
        raise
    except Exception as err:
        raise PythonComSkillError(f"Python 스킬 정의 중 오류: {err}")
    fn = safe_globals.get(PY_SKILL_ENTRY)
    if not callable(fn):
        raise PythonComSkillError(f"def {PY_SKILL_ENTRY}(ctx): 함수를 찾지 못했습니다.")

    deadline = _py_skill_deadline(timeout_s)
    counter = {"n": 0}

    def _tracer(frame, event, arg):
        counter["n"] += 1
        if counter["n"] % 20000 == 0 and time.monotonic() > deadline:
            raise PythonComSkillError("Python 스킬 실행 시간이 초과되었습니다(무한 루프 의심).")
        return _tracer

    sys.settrace(_tracer)
    try:
        fn(ctx)
    except PythonComSkillError as err:
        restored, structural = ctx._rollback()
        if structural:
            err = PythonComSkillError(
                str(err) + " (쓰기 변경은 원복했지만 행/열/시트 구조 변경은 원복하지 못했습니다 — "
                "실행 버튼으로 리셋 재적용을 권장합니다)"
            )
        raise err
    except Exception as err:
        restored, structural = ctx._rollback()
        note = " (쓰기 변경은 원복됨)" if restored or not structural else ""
        if structural:
            note = " (쓰기 변경은 원복했지만 구조 변경은 원복하지 못했습니다 — 리셋 재적용 권장)"
        raise PythonComSkillError(f"Python 스킬 실행 오류: {err}{note}")
    finally:
        sys.settrace(None)

    if not ctx._changed():
        # [SBAGENT-209 후속] 조건부 스킬의 '정상 무변경'은 성공이다.
        # 실측: '소계 행이 있으면 삭제' 스킬이 새 달 파일(소계 없음)에서 아무것도 안 바꾸자
        # 이 게이트가 실패로 처리해 전체실행이 Step 8 에서 죽었다. 대상 시트에서 실데이터를
        # 읽은 증거(read_nonempty)가 있으면 오타겟이 아니라 '조건 미해당' — 성공으로 통과시킨다.
        # 아무것도 못 읽고 무변경이면(엉뚱한 시트/빈 시트) 기존대로 실패 유지('적용됨' 거짓 방지).
        if ctx._shared.get("read_nonempty"):
            _vba_trace("python_com.no_change_ok", comCalls=ctx._shared.get("com_calls"))
            summary = ctx.summary()
            summary["noChange"] = True
            return summary
        raise PythonComSkillError(
            "스킬이 실행됐지만 워크북에 아무 변경도 없습니다(대상 시트/범위/조건을 확인하세요). "
            "'적용됨'으로 잘못 보고되지 않도록 실패로 처리했습니다."
        )
    summary = ctx.summary()
    try:
        _w_total = int(ctx._shared.get("write_cells_total") or 0)
        _w_filled = int(ctx._shared.get("write_cells_nonempty") or 0)
        if _w_total > 0 and _w_filled == 0:
            # [조용한 0건 매칭] 실행은 성공했지만 기록한 값이 전부 빈값 — 대개 구분자/조건 표기가
            # 실제 데이터와 달라 매칭 0건인 경우(예: ' : ' 리터럴 매칭 vs 실제 셀 '03:20.0').
            # 의도적 비우기일 수도 있으니 실패로 만들지 않고 성공 응답에 경고만 실어 드러낸다.
            summary["emptyWrites"] = True
            summary["warning"] = (
                f"적용은 됐지만 기록된 {_w_total:,}칸이 전부 빈값입니다. 조건에 맞는 데이터가 "
                "0건일 수 있습니다 — 구분자·조건 표기(공백 등)가 실제 셀 값과 일치하는지 확인하세요."
            )
            _vba_trace("python_com.all_empty_writes", cells=_w_total)
    except Exception:
        pass
    return summary


def _run_python_on_session_impl(excel_id, code, skip_static=False, timeout_s=None):
    """라이브 세션에 떠 있는 실제 워크북에 Python COM 스킬을 실행한다(VBA 경로와 동일한 외피:
    동반 워크북 보장 → 보호 해제 → 실행 → 앱 상태/보호/창 복구). 변경 검출은 ctx 저널이 담당.
    skip_static/timeout_s: 복구·강제 대용량 경로에서 정적검사 우회 + 데드라인 확장."""
    if not (code or "").strip():
        raise RuntimeError("Python 코드가 비어 있습니다.")
    run_t0 = time.perf_counter()
    try:
        code_hash = hashlib.sha1((code or "").encode("utf-8", "ignore")).hexdigest()[:12]
    except Exception:
        code_hash = ""
    _vba_trace("python_com.run.start", excelId=excel_id, codeHash=code_hash)
    with EXCEL_LOCK:
        session = get_excel_session(excel_id)
        session["rev"] = int(session.get("rev") or 0) + 1  # [리뷰②] 동반 스냅샷 신선도 추적
        app, wb = session_workbook(session)
        _ensure_companion_workbooks(session, excel_id, app, wb)
        try:
            _protect_workbook_for_read_only_mirror(wb, False)
        except Exception:
            pass
        try:
            app.ScreenUpdating = False
        except Exception:
            pass
        try:
            summary = _exec_python_com_skill(app, wb, session, code, skip_static=skip_static, timeout_s=timeout_s)
        finally:
            _restore_app_state(app)
            try:
                _restore_live_protected_view(app, wb)
            except Exception:
                pass
            try:
                _restore_live_window(session, app, wb)
            except Exception:
                pass
        result = {"ok": True, "excelId": excel_id, "engine": "python-com", **summary}
        # [#5] 구조 변경(열삭제·시트추가 등)이 있었으면 경량 미리보기 스키마를 함께 실어,
        # 클라가 대상 파일 캐시를 갱신해 다음 단계 생성이 옛 구조를 보지 않게 한다.
        if summary.get("structural"):
            try:
                result["liveSchema"] = _live_preview_schema(wb)
            except Exception:
                pass
        _vba_trace(
            "python_com.run.ok",
            excelId=excel_id,
            codeHash=code_hash,
            ms=round((time.perf_counter() - run_t0) * 1000, 1),
            comCalls=summary.get("comCalls"),
            writes=summary.get("writes"),
            structural=summary.get("structural"),
        )
        return result


# [AI 도움 Tier2] 격리 검증 — 스냅샷(스텝 직전 상태) 위에 후보 코드를 '보이지 않는 별도 인스턴스'에서
# 실행하고 대상 시트의 before/after diff 만 돌려준다. 라이브 세션은 절대 건드리지 않는다.
VERIFY_TIMEOUT_S = float(os.environ.get("B2B_VERIFY_TIMEOUT", "60"))
VERIFY_DIFF_CELL_CAP = 200


def _verify_capture_sheet_aoa(wb, sheet_name):
    """대상 시트(없으면 전 시트)의 UsedRange 값을 {시트명: 2차원리스트} 로 캡처(diff 입력용)."""
    out = {}
    try:
        targets = []
        if sheet_name:
            try:
                targets = [wb.Worksheets(sheet_name)]
            except Exception:
                targets = []
        if not targets:
            targets = [wb.Worksheets(i + 1) for i in range(int(wb.Worksheets.Count))]
        for ws in targets:
            try:
                nm = str(ws.Name)
                used = ws.UsedRange
                mat = _range_matrix(used.Value) if used is not None else []
                # 스칼라 정규화(날짜/COM 객체 → 문자열)
                out[nm] = [[_com_scalar(c) for c in row] for row in mat]
            except Exception:
                continue
    except Exception:
        pass
    return out


def _verify_step_isolated_impl(result_id, code, sheet_name=None):
    if not (code or "").strip():
        return {"ok": False, "error": "코드가 비어 있습니다.", "verifiable": False}
    # 교차파일(ctx.book) 후보는 동반 워크북이 없어 격리 검증이 무의미 → 검증 불가로 정직히 반환.
    if re.search(r"\bctx\s*\.\s*book\s*\(", str(code)):
        return {"ok": False, "error": "교차파일 참조 스텝은 격리 검증을 지원하지 않습니다.", "verifiable": False}
    path = ensure_result_file(result_id)
    if not path:
        return {"ok": False, "error": "검증용 스냅샷을 찾을 수 없습니다(스텝을 한 번도 적용하지 않았을 수 있습니다).", "verifiable": False}
    app = None
    wb = None
    pid = None
    t0 = time.perf_counter()
    with EXCEL_LOCK:
        try:
            app = win32com.client.DispatchEx("Excel.Application")
            _track_spawned_excel_app(app)   # 고아 방지: 앱 reaper 가 pid 로 잡게 즉시 등록
            try:
                pid = _excel_process_id(app)
            except Exception:
                pid = None
            try:
                app.Visible = False
                app.DisplayAlerts = False
                app.ScreenUpdating = False
                app.EnableEvents = False
            except Exception:
                pass
            wb = app.Workbooks.Open(str(path), ReadOnly=False)
            # [보호 해제 0.7.2.1 / 2026-08-06] 검증용 격리 사본은 보호가 걸려 있으면 안 된다.
            #   라이브는 시트를 UserInterfaceOnly 로 보호해 화면 편집만 막고 COM 쓰기는 허용하는데,
            #   **그 UserInterfaceOnly 는 파일에 저장되지 않는다**(엑셀 규격). 그래서 보호가 걸린 채
            #   저장된 사본을 여기서 다시 열면 COM 쓰기까지 막혀
            #   "변경하려는 셀 또는 차트가 보호된 시트에 있습니다" 로 검증이 통째로 실패한다(실측).
            #   여긴 버려지는 격리 사본이라 보호를 유지할 이유가 없다 → 열자마자 푼다.
            try:
                _protect_workbook_for_read_only_mirror(wb, False)
            except Exception:
                pass
            before = _verify_capture_sheet_aoa(wb, sheet_name)
            # 격리 인스턴스라 정적검사는 클라가 이미 통과시킨 코드 — skip_static 로 중복 우회.
            session = {"rev": 0, "excelId": "verify", "companionNames": [], "companionTemps": []}
            _exec_python_com_skill(app, wb, session, code, skip_static=True, timeout_s=VERIFY_TIMEOUT_S)
            after = _verify_capture_sheet_aoa(wb, sheet_name)
            diff = compute_workbook_diff(before, after)
            # 카드 표시용 '쓰이는 값' 샘플. compute_sheet_diff 의 cell 은 {r,c,value}(after 값)이고
            # r/c 는 UsedRange 기준 상대좌표라 절대주소를 단정하지 않는다 — 값 예시만 보여준다.
            sample = []
            try:
                for snm, sd in (diff.get("sheets") or {}).items():
                    for cell in (sd.get("cells") or []):
                        v = cell.get("value")
                        if v is None or str(v).strip() == "":
                            continue
                        sample.append({"sheet": snm, "value": str(v)[:60]})
                        if len(sample) >= 12:
                            break
                    if len(sample) >= 12:
                        break
            except Exception:
                pass
            _vba_trace("assist.verify.ok", changed=diff.get("changedCount"),
                       ms=round((time.perf_counter() - t0) * 1000, 1))
            return {"ok": True, "verifiable": True,
                    "changedCount": diff.get("changedCount", 0),
                    "truncated": bool(diff.get("truncated")),
                    "sample": sample[:VERIFY_DIFF_CELL_CAP]}
        except PythonComSkillError as err:
            return {"ok": False, "verifiable": True, "error": str(err)[:600]}
        except Exception as err:
            _vba_trace("assist.verify.fail", error=str(err)[:600])
            return {"ok": False, "verifiable": True, "error": str(err)[:600]}
        finally:
            # 삼중 정리(고아 EXCEL.EXE 방지): wb.Close → app.Quit → pid taskkill.
            try:
                if wb is not None:
                    wb.Close(SaveChanges=False)
            except Exception:
                pass
            try:
                if app is not None:
                    app.Quit()
            except Exception:
                pass
            try:
                if pid:
                    _kill_pid_quiet(pid)
                    SPAWNED_EXCEL_PIDS.discard(int(pid))
            except Exception:
                pass


def verify_step_isolated(result_id, code, sheet=None):
    # 외부 큐 타임아웃은 내부 데드라인보다 넉넉히(+30s). 실패는 예외가 아니라 데이터로 반환한다.
    try:
        return excel_call(_verify_step_isolated_impl, result_id, code, sheet_name=sheet,
                          timeout=max(45, int(VERIFY_TIMEOUT_S) + 30))
    except TimeoutError:
        return {"ok": False, "verifiable": True, "error": "검증 시간이 초과됐습니다."}
    except Exception as err:
        return {"ok": False, "verifiable": True, "error": str(err)[:400]}


def run_python_on_session(excel_id, code, extended=False):
    # extended=True: VBA 실패→복구 Python / 원본 Python 강제적용 경로. 정적검사 우회 + 데드라인 확장으로
    # 대용량 작업이 다시 VBA 로 튕기거나 75초에 잘리지 않고 끝까지 실행된다. 외부 excel_call 타임아웃은
    # 내부 데드라인보다 크게 잡아(+60s), 내부 데드라인이 먼저 발동하도록 한다(정상 예외 경로).
    deadline_s = PY_SKILL_RECOVERY_TIMEOUT_S if extended else None
    eff = float(deadline_s) if deadline_s else PY_SKILL_TIMEOUT_S   # 내부 데드라인(0=무제한)
    if eff <= 0:
        timeout = PY_UNLIMITED_OUTER_S            # 무제한 → 바깥 큐 타임아웃도 사실상 무제한(30일)
    else:
        timeout = max(45, int(eff) + 60)          # 내부 데드라인보다 크게 잡아 내부가 먼저 발동
    try:
        return excel_call(_run_python_on_session_impl, excel_id, code,
                          skip_static=bool(extended), timeout_s=deadline_s, timeout=timeout)
    except TimeoutError as err:
        _vba_trace("python_com.run.timeout", excelId=excel_id, timeout=timeout, error=str(err)[:1000])
        try:
            _force_restart_excel_sessions_direct()
        except Exception:
            pass
        raise RuntimeError(
            "Python COM 스킬 실행이 제한 시간을 넘겨 Excel 세션을 정리했습니다. "
            "전체 열/셀 단위 반복이나 큰 조건 루프는 VBA 또는 벌크 read/write 로 다시 생성해야 합니다. "
            + str(err)
        )
    except Exception as err:
        _vba_trace("python_com.run.fail", excelId=excel_id, error=str(err)[:1000])
        raise




def _com_scalar(value):
    if value is None:
        return ""
    if isinstance(value, (str, int, float, bool)):
        return value
    if hasattr(value, "isoformat"):
        return value.isoformat()
    return str(value)


def _range_matrix(value):
    if value is None:
        return []
    if not isinstance(value, tuple):
        return [[value]]
    if value and not isinstance(value[0], tuple):
        return [list(value)]
    return [list(row) for row in value]


def _excel_address(obj):
    address = getattr(obj, "Address", "")
    if callable(address):
        try:
            return address(False, False)
        except TypeError:
            return address
    return str(address or "")


def _col_letter(n):
    # 1-based 열 번호 → 엑셀 열 문자(A, B, ..., AA ...). COM 호출 없이 Python 으로 계산.
    s = ""
    n = int(n)
    while n > 0:
        n, r = divmod(n - 1, 26)
        s = chr(65 + r) + s
    return s or "A"


# '합계/총계/소계/누계/계' 총계 행 판정(라벨이 어느 열에 있든 왼쪽 라벨 영역을 스캔).
# 클라 buildSheetStructureDigest 의 TOTAL 정규식과 동일 규칙 — 둘이 어긋나면 안 됨.
_TOTAL_LABEL_RE = re.compile(r"^(합\s*계|총\s*계|총합계|총\s*합|소\s*계|누\s*계|계|total|sum)$", re.I)


def _is_total_label(cells):
    """행의 라벨 셀들(보통 A~C) 중 하나라도 '합계/총계/소계/누계/계' 면 True."""
    for v in (cells or []):
        if v is None:
            continue
        t = str(v).strip()
        if t and _TOTAL_LABEL_RE.match(t):
            return True
    return False


def _coerce_number(v):
    """셀 값을 숫자로. bool·라벨·빈칸은 None. 콤마·통화·괄호(음수) 표기 허용."""
    if isinstance(v, bool):
        return None
    if isinstance(v, (int, float)):
        return float(v)
    if isinstance(v, str):
        s = v.strip().replace(",", "").replace(" ", "").replace("₩", "").replace("원", "").replace("%", "")
        neg = s.startswith("(") and s.endswith(")")
        if neg:
            s = "-" + s[1:-1]
        if s and s not in ("-", "+"):
            try:
                return float(s)
            except Exception:
                return None
    return None


def _norm_key(v):
    """블록 키(가입번호 등) 정규화 — 숫자로 저장돼 있어도 텍스트 키와 매칭되도록 정수문자열로."""
    if v is None or isinstance(v, bool):
        return ""
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    return str(v).strip()


def _cond_match(cell, op, target):
    """sum_where 조건 비교. 비교연산자는 숫자로, 그 외는 normalize 텍스트로."""
    op = str(op or "==")
    if op in (">", ">=", "<", "<="):
        a = _coerce_number(cell)
        b = _coerce_number(target)
        if a is None or b is None:
            return False
        return {">": a > b, ">=": a >= b, "<": a < b, "<=": a <= b}[op]
    cs = normalize_text("" if cell is None else str(cell))
    ts = normalize_text("" if target is None else str(target))
    if op == "contains":
        return bool(ts) and ts in cs
    if op == "!=":
        return cs != ts
    return cs == ts   # ==(기본)


def _split_key_tokens(v):
    """한 셀 안 다중 키(가입번호 등)를 분리 — 줄바꿈/공백/콤마/세미콜론/슬래시 구분. 각 토큰은 _norm_key."""
    if v is None:
        return []
    s = str(v).strip()
    if not s:
        return []
    out = []
    for p in re.split(r"[\s,;/]+", s):
        p = p.strip()
        if p:
            out.append(_norm_key(p))
    return out


# 스냅샷이 읽을 최대 범위(거대/부풀린 UsedRange 방어). 변경 감지용이라 이 정도면 충분.
_SNAPSHOT_MAX_ROWS = 20000
_SNAPSHOT_MAX_COLS = 256


def _live_preview_schema(wb, max_rows=60, max_cols=_SNAPSHOT_MAX_COLS, only_sheet=None):
    """라이브 적용 후 클라 스키마 캐시 갱신용 경량 미리보기(시트명 + 상위 N행 AoA + 차원).
    구조 변경(열삭제/시트추가 등) 뒤 다음 단계 생성이 옛 구조(삭제된 열 등)를 보지 않게 한다.
    Value2 라 날짜도 숫자(serial)로 와 JSON 직렬화가 안전하다(클라 ctx.read 와 동일).
    only_sheet 지정 시 그 한 시트만 읽고 partial=True 로 표시(캐시 병합만, 타 시트 삭제 금지)."""
    all_names = list(_excel_collection_names(wb.Worksheets))
    partial = False
    if only_sheet:
        match = next((n for n in all_names if str(n) == str(only_sheet)), None)
        if match is not None:
            names = [match]
            partial = True
        else:
            names = all_names  # 못 찾으면 전체(안전)
    else:
        names = all_names
    sheets, dims = {}, {}
    for nm in names:
        try:
            ws = wb.Worksheets(str(nm))
            used = ws.UsedRange
            total_rows = int(used.Row) + int(used.Rows.Count) - 1
            total_cols = int(used.Column) + int(used.Columns.Count) - 1
            nrows = max(0, min(total_rows, max_rows))
            ncols = max(0, min(total_cols, max_cols))
            if nrows < 1 or ncols < 1:
                sheets[nm] = []
            else:
                rng = ws.Range(ws.Cells(1, 1), ws.Cells(nrows, ncols))
                sheets[nm] = _range_matrix(rng.Value2)
            dims[nm] = {"maxRow": max(0, total_rows), "maxCol": max(0, total_cols)}
        except Exception:
            sheets[nm] = []
            dims[nm] = {"maxRow": 0, "maxCol": 0}
    # partial(단일 시트) 이면 전체 시트명도 함께 줘 클라가 시트 목록은 온전히 유지하되
    # 그리드는 읽은 시트만 병합하도록 한다(applyLiveSchemaToFileCache 가 partial 존중).
    out = {"sheetNames": names, "sheets": sheets, "dims": dims}
    if partial:
        out["partial"] = True
        out["allSheetNames"] = all_names
    return out


def _sheet_snapshot(ws):
    used = ws.UsedRange
    start_row = int(used.Row)
    start_col = int(used.Column)
    # 범위를 한정해 읽는다(전체 UsedRange 가 시트 끝까지 부풀어 있어도 안전).
    n_rows = min(int(used.Rows.Count), _SNAPSHOT_MAX_ROWS)
    n_cols = min(int(used.Columns.Count), _SNAPSHOT_MAX_COLS)
    if n_rows <= 0 or n_cols <= 0:
        return {}
    rng = ws.Range(ws.Cells(start_row, start_col), ws.Cells(start_row + n_rows - 1, start_col + n_cols - 1))
    values = _range_matrix(rng.Value)       # 1회 COM 호출
    formulas = _range_matrix(rng.Formula)   # 1회 COM 호출
    cells = {}
    for r_offset, row in enumerate(values):
        formula_row = formulas[r_offset] if r_offset < len(formulas) else []
        row_num = start_row + r_offset
        for c_offset, value in enumerate(row):
            formula = formula_row[c_offset] if c_offset < len(formula_row) else value
            col_num = start_col + c_offset
            address = f"{_col_letter(col_num)}{row_num}"  # ← Python 계산(셀당 COM 호출 제거)
            formula_text = _com_scalar(formula)
            value_text = _com_scalar(value)
            is_formula = isinstance(formula_text, str) and formula_text.startswith("=")
            cells[address] = {
                "address": address,
                "row": row_num,
                "col": col_num,
                "value": value_text,
                "formula": formula_text if is_formula else "",
                "key": formula_text if is_formula else value_text,
            }
    return cells


def _active_sheet_snapshot(wb, prefer_workbook=False):
    ws = None
    if prefer_workbook:
        # frame 모드: 전환 시 COM Activate 를 하지 않으므로 Application.ActiveSheet 는
        # 다른 워크북을 가리킬 수 있다. 이 워크북 자체의 활성 시트를 우선 사용.
        try:
            ws = wb.ActiveSheet
        except Exception:
            ws = None
    if ws is None:
        ws = wb.Application.ActiveSheet
    try:
        if ws is not None and _workbook_fullname(ws.Parent) != _workbook_fullname(wb):
            ws = None
    except Exception:
        pass
    if ws is None:
        names = _excel_collection_names(wb.Worksheets)
        if not names:
            raise RuntimeError("no visible worksheet")
        ws = wb.Worksheets(names[0])
        ws.Activate()
    return ws.Name, _sheet_snapshot(ws)


def _left_mouse_button_down():
    if os.name != "nt":
        return False
    try:
        return bool(ctypes.windll.user32.GetAsyncKeyState(0x01) & 0x8000)
    except Exception:
        return False


def _active_sheet_name(wb, prefer_workbook=False):
    """활성 시트 '이름만' — 풀스냅샷 없이. 라이브 폴링 경량 경로용."""
    ws = None
    if prefer_workbook:
        try:
            ws = wb.ActiveSheet
        except Exception:
            ws = None
    if ws is None:
        try:
            ws = wb.Application.ActiveSheet
            if ws is not None and _workbook_fullname(ws.Parent) != _workbook_fullname(wb):
                ws = None
        except Exception:
            ws = None
    if ws is None:
        names = _excel_collection_names(wb.Worksheets)
        return names[0] if names else ""
    try:
        return str(ws.Name)
    except Exception:
        return ""


def refresh_excel_session_snapshots(session, wb):
    snapshots = session.setdefault("snapshots", {})
    for name in _excel_collection_names(wb.Worksheets):
        try:
            ws = wb.Worksheets(name)
            snapshots[ws.Name] = _sheet_snapshot(ws)
        except Exception:
            continue


def _poll_excel_session_changes_impl(excel_id):
    with EXCEL_LOCK:
        session = get_excel_session(excel_id)
        app, wb = session_workbook(session)
        # 복사 중(CutCopyMode)이면 클립보드 소스를 스냅샷 — 붙여넣기/탭전환 후 캡처 폴백용.
        _maybe_snapshot_copy_source(app)
        if session.get("liveEditable") and LIVE_FRAME_MODE:
            # active-sync: 사용자가 실제로 클릭해 '포그라운드'인 미러 프레임만 따라간다.
            # ActiveWorkbook 기반 판정은 프로그램적 전환(show-only)과 경합해
            # 방금 누른 탭이 이전 탭으로 되돌아가는 바운스를 만들었다.
            fg_sid, fg_session, fg_wb = _foreground_session_by_frame()
            if fg_sid and fg_sid != excel_id and fg_wb is not None:
                session = fg_session
                wb = fg_wb
                excel_id = fg_sid
        else:
            active_excel_id, active_session, active_wb = _active_session_for_app(app, fallback_session=session, fallback_wb=wb)
            if active_session is not None and active_wb is not None:
                session = active_session
                wb = active_wb
                excel_id = active_excel_id or excel_id
        if session.get("readOnlyMirror"):
            if _left_mouse_button_down():
                return {
                    "ok": True,
                    "excelId": excel_id,
                    "activeExcelId": excel_id,
                    "sheet": session.get("lastSelectionSheet") or "",
                    "address": session.get("lastSelectionAddress") or "",
                    "changes": [],
                    "readOnlyMirror": True,
                    "mouseActive": True,
                }
            ws = wb.Application.ActiveSheet
            if ws is None:
                names = _excel_collection_names(wb.Worksheets)
                if not names:
                    raise RuntimeError("no visible worksheet")
                ws = wb.Worksheets(names[0])
                ws.Activate()
            active_address = ""
            try:
                active_address = _excel_address(app.Selection).replace("$", "")
            except Exception:
                try:
                    active_address = _excel_address(app.ActiveCell).replace("$", "")
                except Exception:
                    pass
            session["lastSelectionSheet"] = ws.Name
            session["lastSelectionAddress"] = active_address
            return {
                "ok": True,
                "excelId": excel_id,
                "activeExcelId": excel_id,
                "sheet": ws.Name,
                "address": active_address,
                "changes": [],
                "readOnlyMirror": True,
            }
        frame_mode = bool(session.get("liveEditable")) and LIVE_FRAME_MODE
        if session.get("liveEditable"):
            # 라이브 미러: 실제 Excel 창이 곧 화면이므로 셀 단위 변경 감지(풀스냅샷 diff)가 필요 없다.
            # 시트가 UserInterfaceOnly 보호라 직접 편집도 불가하고, changes 의 유일한 소비처는
            # 읽기전용 미러 경고 문구였다. 특히 거대 파일(수십만 행)에서는 이 스냅샷이 폴마다
            # 수 초씩 걸려 COM 큐를 도배 → 탭 전환 지연과 stale 응답에 의한 '탭 회귀'의 주범.
            sheet_name = _active_sheet_name(wb, prefer_workbook=frame_mode)
            active_address = ""
            try:
                if frame_mode:
                    # 활성화 없이도 이 세션 창의 선택 범위를 읽는다(app.Selection 은 다른 워크북일 수 있음).
                    active_address = _excel_address(wb.Windows(1).RangeSelection).replace("$", "")
                else:
                    active_address = _excel_address(app.Selection).replace("$", "")
            except Exception:
                try:
                    active_address = _excel_address(app.Selection).replace("$", "")
                except Exception:
                    pass
            session["lastSelectionSheet"] = sheet_name
            session["lastSelectionAddress"] = active_address
            return {
                "ok": True,
                "excelId": excel_id,
                "activeExcelId": excel_id,
                "sheet": sheet_name,
                "address": active_address,
                "changes": [],
                "liveSelectionOnly": True,
            }
        sheet_name, snapshot = _active_sheet_snapshot(wb)
        snapshots = session.setdefault("snapshots", {})
        previous = snapshots.get(sheet_name)
        snapshots[sheet_name] = snapshot
        active_address = ""
        try:
            active_address = _excel_address(app.Selection).replace("$", "")
        except Exception:
            pass
        if previous is None:
            return {
                "ok": True,
                "excelId": excel_id,
                "activeExcelId": excel_id,
                "sheet": sheet_name,
                "address": active_address,
                "changes": [],
                "baseline": True,
            }
        changes = []
        addresses = set(previous.keys()) | set(snapshot.keys())
        for address in sorted(addresses, key=lambda a: (snapshot.get(a) or previous.get(a) or {}).get("row", 0)):
            before = previous.get(address)
            after = snapshot.get(address)
            before_key = before.get("key") if before else ""
            after_key = after.get("key") if after else ""
            if before_key == after_key:
                continue
            current = after or {"address": address, "row": before.get("row"), "col": before.get("col"), "value": "", "formula": ""}
            changes.append({
                "sheet": sheet_name,
                "address": address,
                "r": int(current.get("row") or 1) - 1,
                "c": int(current.get("col") or 1) - 1,
                "value": current.get("value", ""),
                "formula": current.get("formula", ""),
            })
            if len(changes) >= 200:
                break
        return {
            "ok": True,
            "excelId": excel_id,
            "activeExcelId": excel_id,
            "sheet": sheet_name,
            "address": active_address,
            "changes": changes,
            "truncated": len(changes) >= 200,
        }


def _range_formula_info(rng):
    if rng is None:
        return None
    try:
        cell = rng.Cells(1, 1)
    except Exception:
        cell = rng
    try:
        formula = _com_scalar(cell.Formula)
    except Exception:
        formula = ""
    try:
        value = _com_scalar(cell.Value)
    except Exception:
        value = ""
    try:
        address = _excel_address(cell).replace("$", "")
    except Exception:
        address = ""
    try:
        sheet = cell.Worksheet.Name
    except Exception:
        sheet = ""
    has_formula = isinstance(formula, str) and formula.startswith("=")
    return {
        "sheet": sheet,
        "address": address,
        "formula": formula if has_formula else "",
        "value": value,
        "hasFormula": has_formula,
    }


def _excel_range_from_cursor(app):
    if win32gui is None:
        return None
    try:
        x, y = win32gui.GetCursorPos()
    except Exception:
        return None
    try:
        return app.ActiveWindow.RangeFromPoint(int(x), int(y))
    except Exception:
        return None


def _get_excel_hover_info_impl(excel_id):
    with EXCEL_LOCK:
        session = get_excel_session(excel_id)
        app, wb = session_workbook(session)
        info = None
        try:
            if session.get("liveEditable") and LIVE_FRAME_MODE:
                info = _range_formula_info(wb.Windows(1).RangeSelection)
        except Exception:
            info = None
        if info is None:
            try:
                info = _range_formula_info(app.Selection)
            except Exception:
                info = None
        if info is None:
            info = _range_formula_info(_excel_range_from_cursor(app))
        info = info or {"sheet": "", "address": "", "formula": "", "value": "", "hasFormula": False}
        try:
            app.StatusBar = (
                f"{info.get('sheet')}!{info.get('address')}  {info.get('formula')}"
                if info.get("hasFormula") else False
            )
        except Exception:
            pass
        return {"ok": True, "excelId": excel_id, **info}


def poll_excel_session_changes(excel_id):
    return excel_call(_poll_excel_session_changes_impl, excel_id, timeout=60)


def _read_excel_session_selection_impl(excel_id):
    """[0.5.17] 현재 탭의 선택(Selection)만 가볍게 읽는다 — active-sync(포그라운드/탭 따라가기)·복사소스
    스냅샷·셀 diff 없이 최소 COM 만. 행/열/셀 선택이 채팅에 빨리 뜨도록 하는 전용 경량 폴 대상.
    무거운 /api/excel/changes(2200ms) 는 그대로 두어 탭 전환 로직 회귀 위험이 없다."""
    with EXCEL_LOCK:
        session = get_excel_session(excel_id)
        app, wb = session_workbook(session)
        frame_mode = bool(session.get("liveEditable")) and LIVE_FRAME_MODE
        sheet_name = _active_sheet_name(wb, prefer_workbook=frame_mode)
        addr = ""
        try:
            if frame_mode:
                # 활성화 없이도 이 세션 창의 선택을 읽는다(app.Selection 은 다른 워크북일 수 있음).
                addr = _excel_address(wb.Windows(1).RangeSelection).replace("$", "")
            else:
                addr = _excel_address(app.Selection).replace("$", "")
        except Exception:
            try:
                addr = _excel_address(app.Selection).replace("$", "")
            except Exception:
                pass
        return {"ok": True, "excelId": excel_id, "sheet": sheet_name, "address": addr}


def poll_excel_session_selection(excel_id):
    return excel_call(_read_excel_session_selection_impl, excel_id, timeout=30)


def get_excel_hover_info(excel_id):
    return excel_call(_get_excel_hover_info_impl, excel_id, timeout=60)


def open_excel_session(
    path,
    name=None,
    workbook_id=None,
    result_id=None,
    read_only_mirror=False,
    left=None,
    top=None,
    width=None,
    height=None,
    client_left=None,
    client_top=None,
    client_width=None,
    client_height=None,
    viewport_width=None,
    viewport_height=None,
    browser_title=None,
    native_parent_hwnd=None,
    native_host_hwnd=None,
    native_overlay=False,
    live_editable=False,
    defer_visible=False,
    from_state_sig=None,
):
    return excel_call(
        _open_excel_session_impl,
        path,
        name=name,
        workbook_id=workbook_id,
        result_id=result_id,
        read_only_mirror=read_only_mirror,
        left=left,
        top=top,
        width=width,
        height=height,
        client_left=client_left,
        client_top=client_top,
        client_width=client_width,
        client_height=client_height,
        viewport_width=viewport_width,
        viewport_height=viewport_height,
        browser_title=browser_title,
        native_parent_hwnd=native_parent_hwnd,
        native_host_hwnd=native_host_hwnd,
        native_overlay=native_overlay,
        live_editable=live_editable,
        defer_visible=defer_visible,
        from_state_sig=from_state_sig,
        timeout=180,  # 느린 PC/대용량 파일에서 Excel 열기가 길어질 수 있음
    )


def activate_excel_session(excel_id, sheet=None, address=None):
    return excel_call(_activate_excel_session_impl, excel_id, sheet=sheet, address=address)


def save_excel_session(excel_id, name=None, internal=False):
    return excel_call(_save_excel_session_impl, excel_id, name=name, internal=internal)


def close_excel_session(excel_id):
    return excel_call(_close_excel_session_impl, excel_id)


def replace_excel_session_workbook(excel_id, path, name=None, result_id=None, read_only_mirror=None):
    return excel_call(_replace_excel_session_workbook_impl, excel_id, path, name=name, result_id=result_id, read_only_mirror=read_only_mirror)


def position_excel_session(
    excel_id,
    left,
    top,
    width,
    height,
    client_left=None,
    client_top=None,
    client_width=None,
    client_height=None,
    viewport_width=None,
    viewport_height=None,
    browser_title=None,
    native_parent_hwnd=None,
    native_host_hwnd=None,
    native_overlay=False,
    keep_zorder=False,
):
    return excel_call(
        _position_excel_session_impl,
        excel_id,
        left,
        top,
        width,
        height,
        client_left=client_left,
        client_top=client_top,
        client_width=client_width,
        client_height=client_height,
        viewport_width=viewport_width,
        viewport_height=viewport_height,
        browser_title=browser_title,
        native_parent_hwnd=native_parent_hwnd,
        native_host_hwnd=native_host_hwnd,
        native_overlay=native_overlay,
        keep_zorder=keep_zorder,
        timeout=60,
    )


def raise_excel_session(excel_id):
    return excel_call(_raise_excel_session_impl, excel_id, timeout=60)


def show_only_excel_session(
    excel_id,
    left,
    top,
    width,
    height,
    client_left=None,
    client_top=None,
    client_width=None,
    client_height=None,
    viewport_width=None,
    viewport_height=None,
    browser_title=None,
    native_parent_hwnd=None,
    native_host_hwnd=None,
    native_overlay=False,
    skip_position=False,
):
    return excel_call(
        _show_only_excel_session_impl,
        excel_id,
        left,
        top,
        width,
        height,
        client_left=client_left,
        client_top=client_top,
        client_width=client_width,
        client_height=client_height,
        viewport_width=viewport_width,
        viewport_height=viewport_height,
        browser_title=browser_title,
        native_parent_hwnd=native_parent_hwnd,
        native_host_hwnd=native_host_hwnd,
        native_overlay=native_overlay,
        skip_position=skip_position,
        timeout=60,
    )


def recover_excel_session(
    excel_id,
    left,
    top,
    width,
    height,
    client_left=None,
    client_top=None,
    client_width=None,
    client_height=None,
    viewport_width=None,
    viewport_height=None,
    browser_title=None,
    native_parent_hwnd=None,
    native_host_hwnd=None,
    native_overlay=False,
):
    return excel_call(
        _recover_excel_session_impl,
        excel_id,
        left,
        top,
        width,
        height,
        client_left=client_left,
        client_top=client_top,
        client_width=client_width,
        client_height=client_height,
        viewport_width=viewport_width,
        viewport_height=viewport_height,
        browser_title=browser_title,
        native_parent_hwnd=native_parent_hwnd,
        native_host_hwnd=native_host_hwnd,
        native_overlay=native_overlay,
        timeout=90,
    )


def hide_excel_session(excel_id):
    return excel_call(_hide_excel_session_impl, excel_id, timeout=60)


def hide_all_excel_sessions():
    return excel_call(_hide_all_excel_sessions_impl, timeout=60)


def hide_inactive_excel_sessions():
    return excel_call(_hide_inactive_excel_sessions_impl, timeout=60)


def is_python_pipeline_step(step):
    if not step or step.get("enabled") is False:
        return False
    code = normalize_python_pipeline_code(str(step.get("code") or ""))
    language = str(step.get("language") or "").lower()
    return (
        language in ("python", "py")
        or re.search(r"^\s*def\s+transform\s*\(\s*ctx\s*\)\s*:", code, re.M) is not None
    )


_LEGACY_PY_DIALECT_RE = re.compile(
    r"(^[ \t]*#[ \t]*B2B_ENGINE[ \t]*:[ \t]*openpyxl)"
    r"|(^[ \t]*#[ \t]*B2B_ENGINE_FALLBACK[ \t]*:[ \t]*excel-com)"
    r"|\bopenpyxl\b|\bload_workbook[ \t]*\("
    r"|\bctx\.(?:rows|rows_with_index|iter_rows|sheet|input|workbook|write_grid|set_range|col|display_rows|display_value|value|cell)[ \t]*\("
    r"|\bctx\.workbook\b|\bws\.cell[ \t]*\(|\.iter_rows[ \t]*\(",
    re.M | re.I)


def python_step_uses_legacy_dialect(code):
    """[혼합 호환] 구버전 openpyxl/excel-com 방언인가 — True 면 ExcelSkillContext(레거시 ctx)로,
    False 면 라이브와 동일한 COM-bulk ctx(_exec_python_com_skill)로 실행한다."""
    return bool(_LEGACY_PY_DIALECT_RE.search(str(code or "")))


def is_vba_pipeline_step(step):
    if not step or step.get("enabled") is False:
        return False
    lang = str(step.get("language") or "").lower()
    if lang == "vba":
        return True
    if is_python_pipeline_step(step):
        return False
    code = str(step.get("code") or "")
    return re.search(r"^[ \t]*(?:Public[ \t]+|Private[ \t]+)?Sub[ \t]+\w+[ \t]*\(", code, re.M) is not None


def _pipeline_payload_has_vba(payload):
    steps = [s for s in (payload.get("pipeline") or []) if not (s and s.get("enabled") is False)]
    return any(is_vba_pipeline_step(s) for s in steps)


def normalize_python_pipeline_code(code):
    text = str(code or "").replace("\ufeff", "")
    fence = re.search(r"```(?:python|py)?\s*\n([\s\S]*?)```", text, re.I)
    if fence:
        text = fence.group(1)
    lines = text.replace("\r\n", "\n").replace("\r", "\n").split("\n")
    normalized = []
    seen_python = False
    for line in lines:
        stripped = line.strip()
        if not seen_python:
            if not stripped:
                continue
            if stripped.startswith("```"):
                continue
            if stripped.startswith("//"):
                continue
            if re.match(r"^(제목|설명|설명문|title|description)\s*[:：]", stripped, re.I):
                continue
        if re.match(r"^(def|import|from|class|@)\b", stripped):
            seen_python = True
        if stripped.startswith("//"):
            line = line[:len(line) - len(line.lstrip())] + "#" + stripped[2:]
        normalized.append(line)
    text = "\n".join(normalized).strip()
    # LLMs often write workbook.sheets["SheetName"] by analogy with dict-like
    # workbook objects. Our proxies expose .sheet(name) / .sheets(name). Normalize
    # the subscript form so otherwise valid generated skills do not fail at runtime.
    text = re.sub(r"(\b[\w\.]+\s*)\.sheets\s*\[\s*([^\]\n]+?)\s*\]", r"\1.sheet(\2)", text)
    return text + "\n"


def pipeline_has_python(payload):
    active_steps = [step for step in (payload.get("pipeline") or []) if not (step and step.get("enabled") is False)]
    if any(is_python_pipeline_step(step) for step in active_steps):
        return True
    current = payload.get("current") or {}
    return bool(current.get("outputExcelId")) and not active_steps


def _excel_names(collection):
    return _excel_collection_names(collection)


_EXCEL_NO_CELL_VALUE = object()


class ExcelCellProxy:
    def __init__(self, cell):
        object.__setattr__(self, "_cell", cell)

    def __getattr__(self, name):
        return getattr(object.__getattribute__(self, "_cell"), name)

    @property
    def value(self):
        return self._cell.Value

    @value.setter
    def value(self, v):
        self._cell.Value = v


class ExcelWorksheetProxy:
    """COM Worksheet wrapper with small openpyxl-style aliases for fallback runs."""
    def __init__(self, worksheet):
        object.__setattr__(self, "_worksheet", worksheet)

    def __getattr__(self, name):
        return getattr(object.__getattribute__(self, "_worksheet"), name)

    def __setattr__(self, key, value):
        if key == "_worksheet":
            object.__setattr__(self, key, value)
        elif key == "Name":
            self._worksheet.Name = value
        else:
            setattr(self._worksheet, key, value)

    @property
    def raw(self):
        return self._worksheet

    @property
    def Name(self):
        return self._worksheet.Name

    @property
    def Parent(self):
        return self._worksheet.Parent

    @property
    def max_row(self):
        try:
            return int(self._worksheet.UsedRange.Row) + int(self._worksheet.UsedRange.Rows.Count) - 1
        except Exception:
            return 1

    @property
    def max_column(self):
        try:
            return int(self._worksheet.UsedRange.Column) + int(self._worksheet.UsedRange.Columns.Count) - 1
        except Exception:
            return 1

    def cell(self, row=None, column=None, value=_EXCEL_NO_CELL_VALUE):
        c = self._worksheet.Cells(int(row), int(column))
        if value is not _EXCEL_NO_CELL_VALUE:
            c.Value = value
        return ExcelCellProxy(c)

    def insert_cols(self, idx, amount=1):
        idx = int(idx)
        amount = max(1, int(amount or 1))
        rng = self._worksheet.Range(self._worksheet.Columns(idx), self._worksheet.Columns(idx + amount - 1))
        rng.Insert(Shift=-4161)  # xlShiftToRight

    def insert_rows(self, idx, amount=1):
        idx = int(idx)
        amount = max(1, int(amount or 1))
        rng = self._worksheet.Range(self._worksheet.Rows(idx), self._worksheet.Rows(idx + amount - 1))
        rng.Insert(Shift=-4121)  # xlShiftDown

    def delete_cols(self, idx, amount=1):
        idx = int(idx)
        amount = max(1, int(amount or 1))
        rng = self._worksheet.Range(self._worksheet.Columns(idx), self._worksheet.Columns(idx + amount - 1))
        rng.Delete()

    def delete_rows(self, idx, amount=1):
        idx = int(idx)
        amount = max(1, int(amount or 1))
        rng = self._worksheet.Range(self._worksheet.Rows(idx), self._worksheet.Rows(idx + amount - 1))
        rng.Delete()

    def append(self, values):
        row = self.max_row + 1
        values = list(values or [])
        if not values:
            return
        rng = self._worksheet.Range(
            self._worksheet.Cells(row, 1),
            self._worksheet.Cells(row, len(values)),
        )
        rng.Value = [values]

    def clear(self):
        self._worksheet.Cells.Clear()
        return self


class ExcelWorksheetsProxy:
    def __init__(self, ctx, workbook, collection):
        self._ctx = ctx
        self._workbook = workbook
        self._collection = collection

    def __call__(self, key):
        if isinstance(key, str):
            sheet_name = self._ctx._find_sheet_name(self._workbook, key, allow_single=False)
            if sheet_name:
                return self._ctx.sheet(sheet_name, workbook=self._workbook)
        return self._collection(key)

    def Item(self, key):
        return self.__call__(key)

    def __getitem__(self, key):
        return self.__call__(key)

    def __iter__(self):
        for idx in range(1, self._collection.Count + 1):
            yield self._collection.Item(idx)

    def __len__(self):
        return int(self._collection.Count)

    @property
    def names(self):
        return _excel_collection_names(self._collection)

    def add(self, name=None):
        ws = self._collection.Add()
        if name:
            ws.Name = str(name)
        return ExcelWorksheetProxy(ws)

    def __getattr__(self, name):
        return getattr(self._collection, name)


class ExcelWorkbookProxy:
    def __init__(self, ctx, workbook, name=None):
        self._ctx = ctx
        self._workbook = workbook
        self.name = name or ""

    @property
    def raw(self):
        return self._workbook

    @property
    def Worksheets(self):
        return ExcelWorksheetsProxy(self._ctx, self._workbook, self._workbook.Worksheets)

    def __getattr__(self, name):
        return getattr(self._workbook, name)

    def sheet(self, name=None):
        return self._ctx.sheet(name, workbook=self)

    @property
    def sheets(self):
        return self.Worksheets

    def sheet_like(self, name=None):
        return self._ctx.sheet_like(name, workbook=self)

    def range(self, sheet_or_name, address):
        return self._ctx.range(sheet_or_name, address, workbook=self)

    def rows(self, sheet_or_name=None):
        return self._ctx.rows(sheet_or_name, workbook=self)

    def iter_rows(self, sheet_or_name=None, start_row=1):
        return self._ctx.iter_rows(sheet_or_name, workbook=self, start_row=start_row)

    def rows_with_index(self, sheet_or_name=None, start_row=1):
        return self._ctx.iter_rows(sheet_or_name, workbook=self, start_row=start_row)

    def display_rows(self, sheet_or_name=None):
        return self._ctx.display_rows(sheet_or_name, workbook=self)

    def value(self, sheet_or_name, row, col):
        return self._ctx.value(sheet_or_name, row, col, workbook=self)

    def display_value(self, sheet_or_name, row, col):
        return self._ctx.display_value(sheet_or_name, row, col, workbook=self)

    def col(self, sheet_or_name, header, header_rows=20):
        return self._ctx.col(sheet_or_name, header, workbook=self, header_rows=header_rows)

    def header_row(self, sheet_or_name=None, header_rows=20):
        return self._ctx.header_row(sheet_or_name, workbook=self, header_rows=header_rows)


class ExcelSkillContext:
    def __init__(self, app, output_wb, input_wbs, output_name=None, active_file_id=None, active_sheet=None):
        self.excel = app
        self._workbook = output_wb
        self.output_name = output_name or "output"
        self.workbook = ExcelWorkbookProxy(self, output_wb, self.output_name)
        self.output = self.workbook
        self.last_output_sheet = None
        self.last_output_address = None
        self.inputs = {
            name: ExcelWorkbookProxy(self, wb, name)
            for name, wb in (input_wbs or {}).items()
        }
        self.active_file_id = str(active_file_id or "")
        self.active_sheet_name = str(active_sheet or "")
        self.active_workbook = self._workbook_for_file_id(self.active_file_id) or self.workbook

    def _unwrap_workbook(self, wb):
        return wb.raw if isinstance(wb, ExcelWorkbookProxy) else wb

    def _workbook_for_file_id(self, file_id):
        file_id = str(file_id or "")
        if not file_id:
            return None
        if file_id == "output" or file_id.startswith("output:"):
            return self.workbook
        if file_id.startswith("input:"):
            hint = file_id[6:]
            try:
                return self.workbook_like(hint)
            except Exception:
                return self.inputs.get(hint)
        return None

    def _default_workbook(self):
        return self.active_workbook or self.workbook

    def _is_output_workbook(self, wb):
        wb = self._unwrap_workbook(wb)
        try:
            return str(Path(wb.FullName).resolve()).lower() == str(Path(self._workbook.FullName).resolve()).lower()
        except Exception:
            return wb is self._workbook

    def normalize(self, value):
        return normalize_text(value)

    def workbook_like(self, hint=None):
        if not hint:
            return self.workbook
        norm = self.normalize(hint)
        candidates = [(name, wb) for name, wb in self.inputs.items()]
        candidates.append((self.output_name, self.workbook))
        candidates.append(("output", self.workbook))
        for name, wb in candidates:
            if self.normalize(name) == norm:
                return wb
        for name, wb in candidates:
            if norm in self.normalize(name) or self.normalize(name) in norm:
                return wb
        for _, wb in candidates:
            if self._find_sheet_name(wb, hint, allow_single=False):
                return wb
        if len(self.inputs) == 1:
            return next(iter(self.inputs.values()))
        raise RuntimeError(f"workbook not found: {hint}")

    def input(self, hint=None):
        if hint is None:
            if len(self.inputs) == 1:
                return next(iter(self.inputs.values()))
            raise RuntimeError("input workbook hint is required when multiple input files exist")
        return self.workbook_like(hint)

    def _find_sheet_name(self, wb, name=None, allow_single=True):
        wb = self._unwrap_workbook(wb)
        names = _excel_names(wb.Worksheets)
        if not names:
            return None
        if not name:
            try:
                return wb.ActiveSheet.Name
            except Exception:
                return names[0]
        alias = _resolve_ephemeral_excel_open_sheet_alias(name, names)
        if alias:
            return alias
        norm = self.normalize(name)
        sheet_norm_loose = normalize_sheet_lookup(name)
        for sheet_name in names:
            if self.normalize(sheet_name) == norm:
                return sheet_name
        for sheet_name in names:
            sheet_norm = self.normalize(sheet_name)
            if norm in sheet_norm or sheet_norm in norm:
                return sheet_name
        for sheet_name in names:
            candidate = normalize_sheet_lookup(sheet_name)
            if sheet_norm_loose and (candidate == sheet_norm_loose or sheet_norm_loose in candidate or candidate in sheet_norm_loose):
                return sheet_name
        if allow_single and len(names) == 1:
            return names[0]
        return None

    def sheet(self, name=None, workbook=None):
        default_wb = workbook or self._default_workbook()
        wb = self._unwrap_workbook(default_wb)
        lookup_name = self.active_sheet_name if (name is None and workbook is None and self.active_sheet_name) else name
        sheet_name = self._find_sheet_name(wb, lookup_name)
        if not sheet_name and workbook is None:
            matches = []
            candidates = [self.workbook] + list(self.inputs.values())
            seen = set()
            for candidate in candidates:
                raw_candidate = self._unwrap_workbook(candidate)
                try:
                    key = str(Path(raw_candidate.FullName).resolve()).lower()
                except Exception:
                    key = str(id(raw_candidate))
                if key in seen:
                    continue
                seen.add(key)
                input_sheet = self._find_sheet_name(raw_candidate, lookup_name, allow_single=False)
                if input_sheet:
                    matches.append((raw_candidate, input_sheet))
            if len(matches) == 1:
                wb, sheet_name = matches[0]
        if not sheet_name and workbook is None and lookup_name != name:
            sheet_name = self._find_sheet_name(wb, name)
        if not sheet_name and workbook is None:
            matches = []
            for input_wb in self.inputs.values():
                raw_input = self._unwrap_workbook(input_wb)
                input_sheet = self._find_sheet_name(raw_input, name, allow_single=False)
                if input_sheet:
                    matches.append((raw_input, input_sheet))
            if len(matches) == 1:
                wb, sheet_name = matches[0]
        if not sheet_name:
            raise RuntimeError(f"sheet not found: {name}")
        ws = ExcelWorksheetProxy(wb.Worksheets(sheet_name))
        if self._is_output_workbook(wb):
            self.last_output_sheet = ws.Name
        return ws

    def sheet_like(self, name=None, workbook=None):
        return self.sheet(name, workbook)

    def input_sheet(self, sheet_hint=None, file_hint=None):
        workbooks = []
        if file_hint:
            workbooks.append(self.workbook_like(file_hint))
        else:
            workbooks.extend(self.inputs.values())
        for wb in workbooks:
            sheet_name = self._find_sheet_name(wb, sheet_hint, allow_single=True)
            if sheet_name:
                return wb.Worksheets(sheet_name)
        raise RuntimeError(f"input sheet not found: {sheet_hint}")

    def range(self, sheet_or_name, address, workbook=None):
        ws = sheet_or_name if hasattr(sheet_or_name, "Range") else self.sheet(sheet_or_name, workbook)
        try:
            if self._is_output_workbook(ws.Parent):
                self.last_output_sheet = ws.Name
                self.last_output_address = str(address)
        except Exception:
            pass
        return ws.Range(str(address))

    def rows(self, sheet_or_name, workbook=None):
        ws = sheet_or_name if hasattr(sheet_or_name, "UsedRange") else self.sheet(sheet_or_name, workbook)
        values = ws.UsedRange.Value
        if values is None:
            return []
        if not isinstance(values, tuple):
            return [[values]]
        if values and not isinstance(values[0], tuple):
            return [list(values)]
        return [list(row) for row in values]

    def iter_rows(self, sheet_or_name, workbook=None, start_row=1):
        ws = self._ws_of(sheet_or_name, workbook)
        rows = self.rows(ws)
        try:
            base_row = int(ws.UsedRange.Row)
        except Exception:
            base_row = 1
        min_row = max(int(start_row or base_row), base_row)
        for offset, row in enumerate(rows):
            excel_row = base_row + offset
            if excel_row >= min_row:
                yield excel_row, row

    def rows_with_index(self, sheet_or_name, workbook=None, start_row=1):
        return self.iter_rows(sheet_or_name, workbook=workbook, start_row=start_row)

    def value(self, sheet_or_name, row, col, workbook=None):
        ws = self._ws_of(sheet_or_name, workbook)
        return ws.Cells(int(row), int(col)).Value

    def display_value(self, sheet_or_name, row, col, workbook=None):
        return self.value(sheet_or_name, row, col, workbook=workbook)

    def display_rows(self, sheet_or_name, workbook=None):
        return self.rows(sheet_or_name, workbook=workbook)

    def col(self, sheet_or_name, header, workbook=None, header_rows=20):
        rows = self.rows(sheet_or_name, workbook)
        target = self.normalize(header)
        for r_idx, row in enumerate(rows[:header_rows], start=1):
            for c_idx, value in enumerate(row, start=1):
                if self.normalize(value) == target:
                    return ExcelColumnNumber(c_idx)
        for r_idx, row in enumerate(rows[:header_rows], start=1):
            for c_idx, value in enumerate(row, start=1):
                if target and target in self.normalize(value):
                    return ExcelColumnNumber(c_idx)
        raise RuntimeError(f"column not found: {header}")

    def header_row(self, sheet_or_name, workbook=None, header_rows=20):
        rows = self.rows(sheet_or_name, workbook)
        best_idx = 1
        best_score = -1
        for idx, row in enumerate(rows[:header_rows], start=1):
            score = sum(1 for value in row if value not in (None, ""))
            if score > best_score:
                best_idx, best_score = idx, score
        return best_idx

    def data_start_row(self, sheet_or_name, workbook=None, header_rows=20):
        return self.header_row(sheet_or_name, workbook, header_rows) + 1

    # ---- 정렬 / 필터 / 피벗 헬퍼 (자주 쓰는 작업을 안정적으로) ----
    def _ws_of(self, sheet_or_name, workbook=None):
        return sheet_or_name if hasattr(sheet_or_name, "UsedRange") else self.sheet(sheet_or_name, workbook)

    def _col0(self, rows, name_or_idx, header_rows=20):
        # 행 리스트 기준 0-based 열 인덱스. 정수는 1-based 로 간주.
        if isinstance(name_or_idx, (int, ExcelColumnNumber)):
            return max(0, int(name_or_idx) - 1)
        target = self.normalize(name_or_idx)
        scan = rows[:header_rows] if header_rows else rows
        for row in scan:
            for i, v in enumerate(row or []):
                if self.normalize(v) == target:
                    return i
        for row in scan:
            for i, v in enumerate(row or []):
                if target and target in self.normalize(v):
                    return i
        return None

    def add_sheet(self, name, workbook=None):
        wb = self._unwrap_workbook(workbook or self._default_workbook())
        base = re.sub(r"[\[\]:*?/\\]", "_", str(name) or "Sheet")[:31] or "Sheet"
        existing = {self.normalize(n) for n in _excel_names(wb.Worksheets)}
        final = base
        idx = 1
        while self.normalize(final) in existing:
            idx += 1
            suffix = "_" + str(idx)
            final = (base[: max(1, 31 - len(suffix))] + suffix)
        ws = wb.Worksheets.Add(After=wb.Worksheets(wb.Worksheets.Count))
        ws.Name = final
        if self._is_output_workbook(wb):
            self.last_output_sheet = ws.Name
        return ws

    def _write_grid(self, ws, grid, start_row=1, start_col=1):
        if not grid:
            return ws
        ncol = max((len(r or []) for r in grid), default=0)
        if ncol <= 0:
            return ws
        norm = [list(r or []) + [None] * (ncol - len(r or [])) for r in grid]
        rng = ws.Range(ws.Cells(start_row, start_col), ws.Cells(start_row + len(norm) - 1, start_col + ncol - 1))
        _apply_com_text_format_for_long_digit_columns(ws, norm, start_row, start_col)
        rng.Value = norm
        return ws

    # ---- 벌크 입출력(성능): 범위 전체를 한 번에 읽고/쓴다(COM 호출 최소화). ----
    def write_grid(self, ws, grid, start_row=1, start_col=1):
        """2D 리스트(grid)를 start_row/start_col 부터 한 번의 COM 호출로 쓴다."""
        ws = self._ws_of(ws)
        self._write_grid(ws, grid, start_row, start_col)
        if self._is_output_workbook(ws.Parent):
            self.last_output_sheet = ws.Name
        return ws

    def set_range(self, sheet_or_name, address, grid, workbook=None):
        """주소(예 'A2' 또는 'A2:F100')의 좌상단부터 2D 리스트를 한 번에 쓴다(1회 COM 호출)."""
        ws = self._ws_of(sheet_or_name, workbook)
        if not grid:
            return ws
        ncol = max((len(r or []) for r in grid), default=0)
        if ncol <= 0:
            return ws
        norm = [list(r or []) + [None] * (ncol - len(r or [])) for r in grid]
        anchor = ws.Range(str(address)).Cells(1, 1)
        r0, c0 = int(anchor.Row), int(anchor.Column)
        rng = ws.Range(ws.Cells(r0, c0), ws.Cells(r0 + len(norm) - 1, c0 + ncol - 1))
        _apply_com_text_format_for_long_digit_columns(ws, norm, r0, c0)
        rng.Value = norm
        if self._is_output_workbook(ws.Parent):
            self.last_output_sheet = ws.Name
            self.last_output_address = str(address)
        return ws

    @staticmethod
    def _num(v):
        if isinstance(v, bool):
            return None
        if isinstance(v, (int, float)):
            return float(v)
        if isinstance(v, str):
            s = v.strip().replace(",", "")
            try:
                return float(s)
            except ValueError:
                return None
        return None

    def sort(self, sheet_or_name, by, ascending=True, header=True, workbook=None):
        # COM 경로 정렬은 반드시 Excel 네이티브 SortFields 를 사용한다.
        # 값 배열로 읽어 Python 에서 정렬한 뒤 다시 쓰면 긴 숫자 식별자(EID 등)가
        # 8.90E+31 같은 숫자로 재해석되거나 날짜/회계 서식이 떨어진다. SortFields 는
        # 셀 값·수식·서식을 행과 함께 이동시키므로 기존 Excel 동작과 가장 가깝다.
        ws = self._ws_of(sheet_or_name, workbook)
        used = ws.UsedRange
        rows = self.rows(ws)
        keys = list(by) if isinstance(by, (list, tuple)) else [by]
        rels = []
        for k in keys:
            rel = self._col0(rows, k)
            if rel is None:
                raise RuntimeError("sort: column not found: %r" % (k,))
            rels.append(rel)
        asc_list = list(ascending) if isinstance(ascending, (list, tuple)) else [ascending] * len(keys)
        while len(asc_list) < len(keys):
            asc_list.append(asc_list[-1] if asc_list else True)
        try:
            ws.Calculate()
        except Exception:
            pass
        ws.Sort.SortFields.Clear()
        for i, rel in enumerate(rels):
            kr = used.Columns(int(rel) + 1)
            sf = ws.Sort.SortFields.Add(kr)
            sf.Order = (1 if bool(asc_list[i]) else 2)
            try:
                sf.DataOption = 0  # xlSortNormal: 긴 텍스트 숫자를 숫자로 재해석하지 않음
            except Exception:
                pass
        ws.Sort.SetRange(used)
        ws.Sort.Header = (1 if header else 2)
        ws.Sort.Apply()
        ws.Sort.SortFields.Clear()
        if self._is_output_workbook(ws.Parent):
            self.last_output_sheet = ws.Name
        return ws

    def filter_to_sheet(self, sheet_or_name, predicate, dest_name, header_rows=1, workbook=None):
        # AutoFilter 대신 헤더 + 조건에 맞는 행을 새 시트로 복사(읽기전용 미러에서 안정적).
        ws = self._ws_of(sheet_or_name, workbook)
        rows = self.rows(ws)
        # [0.5.15] COM UsedRange 는 '첫 사용열'부터 시작 → A열이 비면 행 인덱스가 밀려 predicate 의 절대 열
        # 인덱스가 어긋난다. 선두 빈 열 수만큼 None 패딩해 'A열=index0' 으로 맞춘다(openpyxl 은 이미 A1 기준
        # 이라 lead=0 → 무영향). 두 엔진의 filter_to_sheet 열 인덱싱을 일치시킨다.
        try:
            _lead = int(ws.UsedRange.Column) - 1
        except Exception:
            _lead = 0
        if _lead > 0:
            rows = [[None] * _lead + list(r) for r in rows]
        hr = max(0, int(header_rows or 0))
        header = rows[:hr]
        matched = []
        body = rows[hr:]
        total = len(body)
        report = getattr(self, "_progress", None)
        report = report if (total > 20000 and callable(report)) else None
        for k, r in enumerate(body):
            try:
                if predicate(r):
                    matched.append(r)
            except Exception:
                continue
            if report is not None and (k % 20000 == 0):
                try: report("거르는 중", k, total)
                except Exception: pass
        dest_wb = workbook
        if dest_wb is None:
            try:
                dest_wb = ws.Parent
            except Exception:
                dest_wb = self._default_workbook()
        dest = self.add_sheet(dest_name, workbook=dest_wb)
        self._write_grid(dest, list(header) + matched)
        return dest

    def _merge_pivot_grid_into_base(self, workbook, dest_name, grid):
        name = str(dest_name or "")
        if not name or name.endswith("_피벗") or "_" not in name or not grid or len(grid[0]) < 2:
            return
        prefix = name.rsplit("_", 1)[0]
        base_names = [prefix + "_피벗", prefix + "_pivot"]
        base_ws = None
        for base_name in base_names:
            try:
                base_ws = self.sheet(base_name, workbook=workbook)
                break
            except Exception:
                base_ws = None
        if base_ws is None:
            return
        try:
            base_sheet_name = base_ws.Name
        except Exception:
            base_sheet_name = ""
        if self.normalize(base_sheet_name) == self.normalize(name):
            return
        base_rows = self.rows(base_ws)
        if not base_rows:
            return
        base_header = list(base_rows[0] or [])
        src_header = list(grid[0] or [])
        add_cols = []
        for src_idx, label in enumerate(src_header[1:], start=1):
            if not any(self.normalize(label) == self.normalize(h) for h in base_header):
                add_cols.append((src_idx, label))
        if not add_cols:
            return
        out = [base_header + [label for _, label in add_cols]]
        key_to_values = {}
        for row in grid[1:]:
            if not row:
                continue
            key_to_values[self.normalize(row[0])] = row
        for row in base_rows[1:]:
            cur = list(row or [])
            key = self.normalize(cur[0] if cur else "")
            src = key_to_values.get(key)
            cur += [(src[i] if src is not None and i < len(src) else None) for i, _ in add_cols]
            out.append(cur)
        existing_keys = {self.normalize((row or [""])[0]) for row in base_rows[1:]}
        for row in grid[1:]:
            if not row or self.normalize(row[0]) in existing_keys:
                continue
            cur = [row[0]] + [None] * (len(base_header) - 1)
            cur += [(row[i] if i < len(row) else None) for i, _ in add_cols]
            out.append(cur)
        self._write_grid(base_ws, out)

    def pivot(self, sheet_or_name, group_by=None, value=None, agg="sum", dest_name=None, header_rows=1, workbook=None, **kwargs):
        # Python 집계로 그룹별 요약 표를 새 시트에 만든다(COM PivotTable 보다 안정적).
        if group_by is None:
            group_by = kwargs.get("rows")
        if value is None and "values" in kwargs:
            value = kwargs.get("values")
        if dest_name is None:
            dest_name = kwargs.get("name") or kwargs.get("dest")
        ws = self._ws_of(sheet_or_name, workbook)
        rows = self.rows(ws)
        hr = max(1, int(header_rows or 1))
        header_row = rows[hr - 1] if len(rows) >= hr else []
        data = rows[hr:]
        group_cols = list(group_by) if isinstance(group_by, (list, tuple)) else [group_by]
        gidx = [self._col0(rows, g, hr) for g in group_cols]
        values = list(value) if isinstance(value, (list, tuple)) else [value]
        aggs = list(agg) if isinstance(agg, (list, tuple)) else [agg] * len(values)
        while len(aggs) < len(values):
            aggs.append(aggs[-1] if aggs else "sum")
        aggs = [str(a or "sum").lower() for a in aggs]
        vidxs = [self._col0(rows, v, hr) if v is not None else None for v in values]

        groups = {}
        order = []
        for r in data:
            key = tuple((r[i] if (i is not None and i < len(r)) else "") for i in gidx)
            if key not in groups:
                groups[key] = [[] for _ in values]
                order.append(key)
            for pos, vidx in enumerate(vidxs):
                if vidx is None:
                    groups[key][pos].append(1)
                elif vidx < len(r):
                    groups[key][pos].append(r[vidx])

        def _aggregate(vals, agg_name):
            nums = [n for n in (self._num(v) for v in vals) if n is not None]
            if agg_name == "count":
                return len(vals)
            if agg_name in ("avg", "average", "mean"):
                return (sum(nums) / len(nums)) if nums else 0
            if agg_name == "max":
                return max(nums) if nums else ""
            if agg_name == "min":
                return min(nums) if nums else ""
            return sum(nums)

        out_header = []
        for n, i in enumerate(gidx):
            label = header_row[i] if (i is not None and i < len(header_row)) else ("그룹%d" % (n + 1))
            out_header.append(label)
        for v, agg_name in zip(values, aggs):
            label = str(v) if v is not None else "값"
            out_header.append(label + "_" + (agg_name if agg_name != "average" else "avg"))
        grid = [out_header]
        for key in order:
            grid.append(list(key) + [_aggregate(groups[key][i], aggs[i]) for i in range(len(values))])
        dest_wb = workbook
        if dest_wb is None:
            try:
                dest_wb = ws.Parent
            except Exception:
                dest_wb = self._default_workbook()
        dest = self.add_sheet(dest_name or "피벗요약", workbook=dest_wb)
        self._write_grid(dest, grid)
        self._merge_pivot_grid_into_base(dest_wb, dest.Name, grid)
        return dest


# 스킬에서 import 가능한 안전한 표준 라이브러리만 허용(os/sys/subprocess 등 위험 모듈은 차단).
_SKILL_ALLOWED_IMPORTS = {
    "re", "datetime", "math", "json", "collections", "itertools",
    "functools", "string", "decimal", "statistics", "calendar",
    "textwrap", "unicodedata", "fractions", "random", "operator", "copy",
    "difflib",
}


def _safe_skill_import(name, globals=None, locals=None, fromlist=(), level=0):
    import importlib
    if level and level != 0:
        raise ImportError("relative imports are not allowed in skills")
    root = str(name or "").split(".")[0]
    if root not in _SKILL_ALLOWED_IMPORTS:
        raise ImportError(
            "import of '%s' is not allowed in skills (allowed: %s)"
            % (name, ", ".join(sorted(_SKILL_ALLOWED_IMPORTS)))
        )
    return importlib.import_module(name)


def _safe_python_globals():
    allowed_builtins = {
        "abs": abs, "all": all, "any": any, "bool": bool, "dict": dict, "enumerate": enumerate,
        "float": float, "int": int, "isinstance": isinstance, "len": len, "list": list,
        "max": max, "min": min, "print": print, "range": range, "round": round,
        "set": set, "sorted": sorted, "str": str, "sum": sum, "tuple": tuple,
        "type": type, "zip": zip, "getattr": getattr, "hasattr": hasattr, "iter": iter,
        "next": next, "repr": repr, "abs": abs, "divmod": divmod, "ord": ord, "chr": chr,
        "filter": filter, "map": map, "reversed": reversed, "format": format, "frozenset": frozenset,
        "__import__": _safe_skill_import,
        "Exception": Exception, "RuntimeError": RuntimeError, "ValueError": ValueError,
        "TypeError": TypeError, "KeyError": KeyError, "IndexError": IndexError,
        "AttributeError": AttributeError, "ZeroDivisionError": ZeroDivisionError,
        "StopIteration": StopIteration,
    }
    return {
        "__builtins__": allowed_builtins,
        "datetime": datetime,
        "math": math,
        "re": re,
    }


# =====================================================================
#  순수 Python(openpyxl) 스킬 엔진 — COM/Excel 없이 인프로세스로 실행(빠름).
#  COM 스킬과 동일한 ctx API + ws.Range/Cells/UsedRange/.Value 호환 shim을 제공해
#  기존 스킬 코드/프롬프트가 대부분 그대로 동작하게 한다.
#  주의: 수식 셀을 다시 읽으면 계산값이 아니라 수식 문자열이 나온다(openpyxl 한계).
#       값은 Python에서 계산해 셀에 직접 쓰는 것을 권장.
# =====================================================================
def _opxl_coord(token):
    from openpyxl.utils.cell import coordinate_to_tuple
    row, col = coordinate_to_tuple(str(token).replace("$", "").strip())
    return int(row), int(col)


_OPXL_NO_VALUE = object()


def _opxl_merged_anchor(ws, row, col):
    try:
        for merged in ws.merged_cells.ranges:
            if merged.min_row <= row <= merged.max_row and merged.min_col <= col <= merged.max_col:
                return int(merged.min_row), int(merged.min_col)
    except Exception:
        pass
    return int(row), int(col)


def _opxl_write_cell(ws, row, col, value, redirect_merged=False):
    ar, ac = _opxl_merged_anchor(ws, row, col)
    if (ar, ac) != (int(row), int(col)) and not redirect_merged:
        return False
    ws.cell(row=ar, column=ac, value=value)
    return True


_OPXL_CELL_REF_RE = re.compile(r"(?<![A-Za-z0-9_])(?:'([^']+)'!)?([A-Z]{1,3})([1-9][0-9]*)(?![A-Za-z0-9_])")


def _opxl_col_to_index(col):
    out = 0
    for ch in str(col or "").upper():
        if "A" <= ch <= "Z":
            out = out * 26 + (ord(ch) - 64)
    return out


def _opxl_get_cached_cell_value(cached_ws, row, col):
    if cached_ws is None:
        return None
    try:
        value = cached_ws.cell(row=int(row), column=int(col)).value
        if isinstance(value, str) and value.startswith("="):
            return None
        return value
    except Exception:
        return None


def _opxl_safe_eval_arithmetic(expr):
    allowed_binops = {
        ast.Add: lambda a, b: a + b,
        ast.Sub: lambda a, b: a - b,
        ast.Mult: lambda a, b: a * b,
        ast.Div: lambda a, b: a / b,
        ast.Pow: lambda a, b: a ** b,
        ast.Mod: lambda a, b: a % b,
    }
    allowed_unary = {
        ast.UAdd: lambda a: +a,
        ast.USub: lambda a: -a,
    }

    def _eval(node):
        if isinstance(node, ast.Expression):
            return _eval(node.body)
        if isinstance(node, ast.Constant):
            if isinstance(node.value, (int, float, str, bool)) or node.value is None:
                return node.value
            raise ValueError("unsupported constant")
        if isinstance(node, ast.BinOp) and type(node.op) in allowed_binops:
            return allowed_binops[type(node.op)](_eval(node.left), _eval(node.right))
        if isinstance(node, ast.UnaryOp) and type(node.op) in allowed_unary:
            return allowed_unary[type(node.op)](_eval(node.operand))
        if isinstance(node, ast.Compare):
            left = _eval(node.left)
            for op, comparator in zip(node.ops, node.comparators):
                right = _eval(comparator)
                if isinstance(op, ast.Eq):
                    ok = left == right
                elif isinstance(op, ast.NotEq):
                    ok = left != right
                elif isinstance(op, ast.Lt):
                    ok = left < right
                elif isinstance(op, ast.LtE):
                    ok = left <= right
                elif isinstance(op, ast.Gt):
                    ok = left > right
                elif isinstance(op, ast.GtE):
                    ok = left >= right
                else:
                    raise ValueError("unsupported comparison")
                if not ok:
                    return False
                left = right
            return True
        raise ValueError("unsupported expression")

    return _eval(ast.parse(expr, mode="eval"))


def _opxl_split_top_level_args(text):
    args = []
    cur = []
    depth = 0
    in_quote = False
    quote_ch = ""
    for ch in str(text or ""):
        if in_quote:
            cur.append(ch)
            if ch == quote_ch:
                in_quote = False
            continue
        if ch in ("'", '"'):
            in_quote = True
            quote_ch = ch
            cur.append(ch)
            continue
        if ch == "(":
            depth += 1
            cur.append(ch)
            continue
        if ch == ")":
            depth = max(0, depth - 1)
            cur.append(ch)
            continue
        if ch == "," and depth == 0:
            args.append("".join(cur).strip())
            cur = []
            continue
        cur.append(ch)
    args.append("".join(cur).strip())
    return args


def _opxl_range_values(ws, token, cached_ws=None, seen=None):
    token = str(token or "").replace("$", "").strip()
    sheet_name = None
    if "!" in token:
        sheet_part, token = token.split("!", 1)
        sheet_name = sheet_part.strip("'")
    if ":" not in token:
        match = re.fullmatch(r"([A-Z]{1,3})([1-9][0-9]*)", token, re.I)
        if not match:
            return []
        col = _opxl_col_to_index(match.group(1))
        row = int(match.group(2))
        target_ws = ws.parent[sheet_name] if sheet_name else ws
        target_cached = cached_ws.parent[sheet_name] if (sheet_name and cached_ws is not None and sheet_name in cached_ws.parent.sheetnames) else cached_ws
        return [_opxl_display_cell_value(target_ws, row, col, target_cached, seen)]
    left, right = token.split(":", 1)
    m1 = re.fullmatch(r"([A-Z]{1,3})([1-9][0-9]*)", left, re.I)
    m2 = re.fullmatch(r"([A-Z]{1,3})([1-9][0-9]*)", right, re.I)
    if not m1 or not m2:
        return []
    r1, c1 = int(m1.group(2)), _opxl_col_to_index(m1.group(1))
    r2, c2 = int(m2.group(2)), _opxl_col_to_index(m2.group(1))
    target_ws = ws.parent[sheet_name] if sheet_name else ws
    target_cached = cached_ws.parent[sheet_name] if (sheet_name and cached_ws is not None and sheet_name in cached_ws.parent.sheetnames) else cached_ws
    values = []
    for row in range(min(r1, r2), max(r1, r2) + 1):
        for col in range(min(c1, c2), max(c1, c2) + 1):
            values.append(_opxl_display_cell_value(target_ws, row, col, target_cached, seen))
    return values


def _opxl_numeric_values(values):
    nums = []
    for value in values:
        if isinstance(value, bool) or value in (None, ""):
            continue
        if isinstance(value, (int, float)):
            nums.append(float(value))
            continue
        try:
            nums.append(float(str(value).replace(",", "")))
        except Exception:
            continue
    return nums


def _opxl_eval_formula(ws, formula, cached_ws=None, seen=None):
    expr = str(formula or "")
    if expr.startswith("="):
        expr = expr[1:]
    expr = expr.strip()
    if not expr:
        return None
    upper = expr.upper()
    for fn in ("SUM", "AVERAGE", "COUNTIF", "IFERROR"):
        prefix = fn + "("
        if upper.startswith(prefix) and expr.endswith(")"):
            args = _opxl_split_top_level_args(expr[len(prefix):-1])
            if fn == "SUM":
                values = []
                for arg in args:
                    values.extend(_opxl_range_values(ws, arg, cached_ws, seen) if ":" in arg or re.fullmatch(r"(?:'[^']+'!)?[A-Z]{1,3}[1-9][0-9]*", arg.replace("$", ""), re.I) else [_opxl_eval_formula(ws, "=" + arg, cached_ws, seen)])
                nums = _opxl_numeric_values(values)
                return sum(nums)
            if fn == "AVERAGE":
                values = []
                for arg in args:
                    values.extend(_opxl_range_values(ws, arg, cached_ws, seen) if ":" in arg or re.fullmatch(r"(?:'[^']+'!)?[A-Z]{1,3}[1-9][0-9]*", arg.replace("$", ""), re.I) else [_opxl_eval_formula(ws, "=" + arg, cached_ws, seen)])
                nums = _opxl_numeric_values(values)
                return (sum(nums) / len(nums)) if nums else 0
            if fn == "COUNTIF":
                if len(args) < 2:
                    raise ValueError("COUNTIF requires range and criterion")
                values = _opxl_range_values(ws, args[0], cached_ws, seen)
                criterion = _opxl_eval_formula(ws, "=" + args[1], cached_ws, seen)
                return sum(1 for value in values if value == criterion)
            if fn == "IFERROR":
                if len(args) < 2:
                    raise ValueError("IFERROR requires value and fallback")
                try:
                    return _opxl_eval_formula(ws, "=" + args[0], cached_ws, seen)
                except Exception:
                    return _opxl_eval_formula(ws, "=" + args[1], cached_ws, seen)

    def _replace_cell(match):
        sheet_name, col_text, row_text = match.groups()
        row = int(row_text)
        col = _opxl_col_to_index(col_text)
        target_ws = ws.parent[sheet_name] if sheet_name else ws
        target_cached = cached_ws.parent[sheet_name] if (sheet_name and cached_ws is not None and sheet_name in cached_ws.parent.sheetnames) else cached_ws
        value = _opxl_display_cell_value(target_ws, row, col, target_cached, seen)
        if value in (None, ""):
            value = 0
        return repr(value) if isinstance(value, str) else str(value)

    py_expr = _OPXL_CELL_REF_RE.sub(_replace_cell, expr.replace("$", ""))
    py_expr = py_expr.replace("^", "**")
    return _opxl_safe_eval_arithmetic(py_expr)


def _opxl_display_cell_value(ws, row, col, cached_ws=None, seen=None):
    row, col = int(row), int(col)
    seen = seen or set()
    key = (id(ws), row, col)
    if key in seen:
        raise RuntimeError(f"circular formula reference at {ws.title}!{row},{col}")
    value = ws.cell(row=row, column=col).value
    if not (isinstance(value, str) and value.startswith("=")):
        return value
    cached = _opxl_get_cached_cell_value(cached_ws, row, col)
    if cached is not None:
        return cached
    seen.add(key)
    try:
        return _opxl_eval_formula(ws, value, cached_ws, seen)
    except Exception as err:
        coord = f"{ws.title}!{ws.cell(row=row, column=col).coordinate}"
        raise RuntimeError(f"formula value unavailable for {coord}: {value} ({err})")
    finally:
        seen.discard(key)


class _OpxlCount:
    def __init__(self, count):
        self.Count = int(count)


class _OpxlRange:
    """openpyxl 워크시트 위의 직사각 범위(또는 단일 셀). COM Range.Value 시맨틱을 흉내낸다."""
    def __init__(self, ws, r1, c1, r2, c2):
        self._ws = ws
        self._r1, self._c1 = int(r1), int(c1)
        self._r2, self._c2 = int(r2), int(c2)

    @property
    def _single(self):
        return self._r1 == self._r2 and self._c1 == self._c2

    def _get_value(self):
        if self._single:
            return self._ws.cell(row=self._r1, column=self._c1).value
        out = []
        for r in range(self._r1, self._r2 + 1):
            out.append(tuple(self._ws.cell(row=r, column=c).value for c in range(self._c1, self._c2 + 1)))
        return tuple(out)

    def _set_value(self, value):
        if self._single and not isinstance(value, (list, tuple)):
            _opxl_write_cell(self._ws, self._r1, self._c1, value, redirect_merged=True)
            return
        if isinstance(value, (list, tuple)) and value and not isinstance(value[0], (list, tuple)):
            value = [value]  # 1행 그리드로 취급
        rows = list(value) if isinstance(value, (list, tuple)) else [[value]]
        for i, r in enumerate(range(self._r1, self._r2 + 1)):
            row = rows[i] if i < len(rows) else None
            if row is None:
                if self._single or not isinstance(value, (list, tuple)):
                    self._ws.cell(row=r, column=self._c1, value=value)
                continue
            if not isinstance(row, (list, tuple)):
                row = [row]
            for j, c in enumerate(range(self._c1, self._c2 + 1)):
                if j < len(row):
                    _opxl_write_cell(self._ws, r, c, row[j], redirect_merged=False)

    Value = property(_get_value, _set_value)
    Value2 = property(_get_value, _set_value)

    @property
    def Row(self):
        return self._r1

    @property
    def Column(self):
        return self._c1

    @property
    def Rows(self):
        return _OpxlCount(self._r2 - self._r1 + 1)

    @property
    def Columns(self):
        return _OpxlCount(self._c2 - self._c1 + 1)

    def Select(self):
        return self


class _OpxlRowProxy:
    def __init__(self, sheet_proxy, row_idx):
        self._sheet_proxy = sheet_proxy
        self._ws = sheet_proxy._ws
        self._row_idx = int(row_idx)

    @property
    def values(self):
        self._sheet_proxy.flush_pending_rows()
        max_col = int(self._ws.max_column or 1)
        return [self._ws.cell(row=self._row_idx, column=c).value for c in range(1, max_col + 1)]

    @values.setter
    def values(self, row_values):
        self._sheet_proxy._set_pending_row_values(self._row_idx, list(row_values or []))

    def clear(self):
        self._sheet_proxy._set_pending_row_clear(self._row_idx)
        return self


class _OpxlFormulaString(str):
    def __new__(cls, value, row, col, ws=None):
        obj = str.__new__(cls, value)
        obj._b2b_origin_row = int(row)
        obj._b2b_origin_col = int(col)
        obj._b2b_origin_ws = ws
        return obj

    def replace(self, old, new, count=-1):
        if count == -1:
            value = super().replace(old, new)
        else:
            value = super().replace(old, new, count)
        return _OpxlFormulaString(
            value,
            self._b2b_origin_row,
            self._b2b_origin_col,
            getattr(self, "_b2b_origin_ws", None),
        )


class _OpxlCopiedInt(int):
    def __new__(cls, value, row, col, ws=None):
        obj = int.__new__(cls, value)
        obj._b2b_origin_row = int(row)
        obj._b2b_origin_col = int(col)
        obj._b2b_origin_ws = ws
        return obj


class _OpxlCopiedFloat(float):
    def __new__(cls, value, row, col, ws=None):
        obj = float.__new__(cls, value)
        obj._b2b_origin_row = int(row)
        obj._b2b_origin_col = int(col)
        obj._b2b_origin_ws = ws
        return obj


def _opxl_coord_from_row_col(row, col):
    from openpyxl.utils import get_column_letter
    return f"{get_column_letter(int(col))}{int(row)}"


def _opxl_translate_formula(value, dest_row, dest_col):
    if not (isinstance(value, _OpxlFormulaString) and str(value).startswith("=")):
        return value
    try:
        origin = _opxl_coord_from_row_col(value._b2b_origin_row, value._b2b_origin_col)
        dest = _opxl_coord_from_row_col(dest_row, dest_col)
        if origin == dest:
            return str(value)
        from openpyxl.formula.translate import Translator
        return Translator(str(value), origin=origin).translate_formula(dest)
    except Exception:
        return str(value)


def _opxl_unwrap_copied_value(value):
    if isinstance(value, _OpxlFormulaString):
        return str(value)
    if isinstance(value, _OpxlCopiedInt):
        return int(value)
    if isinstance(value, _OpxlCopiedFloat):
        return float(value)
    return value


def _opxl_copied_source(value):
    ws = getattr(value, "_b2b_origin_ws", None)
    if ws is None:
        return None
    try:
        return ws, int(value._b2b_origin_row), int(value._b2b_origin_col)
    except Exception:
        return None


def _opxl_ranges_overlap(a, b):
    return not (a.max_row < b.min_row or a.min_row > b.max_row or a.max_col < b.min_col or a.min_col > b.max_col)


def _opxl_unmerge_overlapping(ws, target_range):
    try:
        for existing in list(ws.merged_cells.ranges):
            if _opxl_ranges_overlap(existing, target_range):
                ws.unmerge_cells(str(existing))
    except Exception:
        pass


def _opxl_copy_cell_presentation(src_ws, src_row, src_col, dst_ws, dst_row, dst_col):
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.cell_range import CellRange

    src = src_ws.cell(row=int(src_row), column=int(src_col))
    dst = dst_ws.cell(row=int(dst_row), column=int(dst_col))
    if getattr(src, "has_style", False):
        dst._style = copy.copy(src._style)
    if src.number_format:
        dst.number_format = src.number_format
    if src.hyperlink:
        dst._hyperlink = copy.copy(src.hyperlink)
    if src.comment:
        dst.comment = copy.copy(src.comment)

    try:
        src_dim = src_ws.column_dimensions.get(get_column_letter(int(src_col)))
        if src_dim and src_dim.width:
            dst_ws.column_dimensions[get_column_letter(int(dst_col))].width = src_dim.width
    except Exception:
        pass
    try:
        src_height = src_ws.row_dimensions[int(src_row)].height
        if src_height:
            dst_ws.row_dimensions[int(dst_row)].height = src_height
    except Exception:
        pass

    try:
        for merged in list(src_ws.merged_cells.ranges):
            if not (merged.min_row <= int(src_row) <= merged.max_row and merged.min_col <= int(src_col) <= merged.max_col):
                continue
            if int(src_row) != merged.min_row or int(src_col) != merged.min_col:
                return
            row_delta = int(dst_row) - int(src_row)
            col_delta = int(dst_col) - int(src_col)
            target = CellRange(
                min_col=merged.min_col + col_delta,
                min_row=merged.min_row + row_delta,
                max_col=merged.max_col + col_delta,
                max_row=merged.max_row + row_delta,
            )
            if target.min_row < 1 or target.min_col < 1:
                continue
            _opxl_unmerge_overlapping(dst_ws, target)
            dst_ws.merge_cells(str(target))
            break
    except Exception:
        pass


class _OpxlCellProxy:
    def __init__(self, ws, row, col):
        object.__setattr__(self, "_ws", ws)
        object.__setattr__(self, "_row", int(row))
        object.__setattr__(self, "_col", int(col))

    @property
    def _cell(self):
        return self._ws.cell(row=self._row, column=self._col)

    def __getattr__(self, name):
        return getattr(self._cell, name)

    def __setattr__(self, key, value):
        if key in ("_ws", "_row", "_col"):
            object.__setattr__(self, key, value)
        elif key == "value":
            type(self).value.fset(self, value)
        else:
            setattr(self._cell, key, value)

    @property
    def value(self):
        value = self._cell.value
        if isinstance(value, str):
            return _OpxlFormulaString(value, self._row, self._col, self._ws)
        if isinstance(value, bool):
            return value
        if isinstance(value, int):
            return _OpxlCopiedInt(value, self._row, self._col, self._ws)
        if isinstance(value, float):
            return _OpxlCopiedFloat(value, self._row, self._col, self._ws)
        return value

    @value.setter
    def value(self, value):
        ar, ac = _opxl_merged_anchor(self._ws, self._row, self._col)
        if (ar, ac) != (self._row, self._col) and value in (None, ""):
            return
        src = _opxl_copied_source(value)
        if src:
            _opxl_copy_cell_presentation(src[0], src[1], src[2], self._ws, ar, ac)
        translated = _opxl_translate_formula(value, ar, ac)
        self._ws.cell(row=ar, column=ac).value = _opxl_unwrap_copied_value(translated)


class OpenpyxlWorksheetProxy:
    """openpyxl Worksheet 래퍼. COM 풍의 Range/Cells/UsedRange/.Name 을 제공하고
    그 외 속성/메서드(cell, insert_cols, append, max_row 등)는 openpyxl 로 위임한다."""
    def __init__(self, ws):
        object.__setattr__(self, "_ws", ws)

    def __getattr__(self, name):
        return getattr(object.__getattribute__(self, "_ws"), name)

    def __setattr__(self, key, value):
        if key == "_ws":
            object.__setattr__(self, key, value)
        elif key == "Name":
            self._ws.title = value
        else:
            setattr(self._ws, key, value)

    @property
    def Name(self):
        return self._ws.title

    @property
    def Parent(self):
        return self._ws.parent

    def _pending_rows(self):
        pending = getattr(self._ws, "_b2b_pending_rows", None)
        if pending is None:
            pending = {}
            setattr(self._ws, "_b2b_pending_rows", pending)
        return pending

    def _set_pending_row_values(self, row_idx, values):
        self._pending_rows()[int(row_idx)] = list(values or [])

    def _set_pending_row_clear(self, row_idx):
        self._pending_rows()[int(row_idx)] = None

    def flush_pending_rows(self):
        pending = getattr(self._ws, "_b2b_pending_rows", None)
        if not pending:
            return self
        items = sorted(pending.items())
        setattr(self._ws, "_b2b_pending_rows", {})
        has_merges = bool(getattr(getattr(self._ws, "merged_cells", None), "ranges", None))
        cell = self._ws.cell

        def is_blank_initial_row():
            try:
                return (
                    int(self._ws.max_row or 1) == 1
                    and int(self._ws.max_column or 1) == 1
                    and self._ws.cell(row=1, column=1).value in (None, "")
                )
            except Exception:
                return False

        current_max_col = int(self._ws.max_column or 1)

        def write_direct(row_idx, values):
            nonlocal current_max_col
            max_col = max(current_max_col, len(values or []))
            for c in range(1, max_col + 1):
                value = values[c - 1] if values is not None and c <= len(values) else None
                if has_merges:
                    _opxl_write_cell(self._ws, row_idx, c, value, redirect_merged=True)
                else:
                    cell(row=row_idx, column=c).value = value
            current_max_col = max_col

        current_max_row = int(self._ws.max_row or 0)
        for row_idx, values in items:
            row_idx = int(row_idx)
            if values is None:
                write_direct(row_idx, None)
                current_max_row = max(current_max_row, row_idx)
                continue
            if row_idx == 1 and is_blank_initial_row():
                write_direct(row_idx, values)
                current_max_row = max(current_max_row, row_idx)
                try:
                    self._ws._current_row = max(int(getattr(self._ws, "_current_row", 0) or 0), 1)
                except Exception:
                    pass
                continue
            if (not has_merges) and row_idx == current_max_row + 1:
                self._ws.append(values)
                current_max_row = row_idx
            else:
                write_direct(row_idx, values)
                current_max_row = max(current_max_row, row_idx)
        return self

    def Cells(self, r, c):
        self.flush_pending_rows()
        return _OpxlRange(self._ws, r, c, r, c)

    def row(self, row_idx):
        return _OpxlRowProxy(self, row_idx)

    def clear(self):
        self.flush_pending_rows()
        max_row = int(self._ws.max_row or 0)
        if max_row > 0:
            self._ws.delete_rows(1, max_row)
        try:
            self._ws._current_row = 0
        except Exception:
            pass
        return self

    def append(self, values):
        self.flush_pending_rows()
        values = list(values or [])
        current_row = int(getattr(self._ws, "_current_row", 0) or 0)
        is_blank_initial_row = False
        if current_row <= 1:
            try:
                is_blank_initial_row = (
                    int(self._ws.max_row or 1) == 1
                    and int(self._ws.max_column or 1) == 1
                    and self._ws.cell(row=1, column=1).value in (None, "")
                )
            except Exception:
                is_blank_initial_row = False
        if is_blank_initial_row:
            cell = self._ws.cell
            for c, value in enumerate(values, start=1):
                cell(row=1, column=c).value = value
            try:
                self._ws._current_row = 1
            except Exception:
                pass
        else:
            self._ws.append(values)
        return self

    def _formula_cells(self):
        self.flush_pending_rows()
        formulas = []
        for row in self._ws.iter_rows():
            for cell in row:
                value = cell.value
                if isinstance(value, str) and value.startswith("="):
                    formulas.append((int(cell.row), int(cell.column), value))
        return formulas

    def _write_translated_formula(self, old_row, old_col, new_row, new_col, formula):
        try:
            wrapped = _OpxlFormulaString(formula, old_row, old_col)
            self._ws.cell(row=int(new_row), column=int(new_col)).value = _opxl_translate_formula(wrapped, new_row, new_col)
        except Exception:
            self._ws.cell(row=int(new_row), column=int(new_col)).value = formula

    def insert_cols(self, idx, amount=1):
        self.flush_pending_rows()
        idx = int(idx)
        amount = max(1, int(amount or 1))
        formulas = self._formula_cells()
        self._ws.insert_cols(idx, amount)
        for row, col, formula in formulas:
            if col >= idx:
                self._write_translated_formula(row, col, row, col + amount, formula)
        return self

    def insert_rows(self, idx, amount=1):
        self.flush_pending_rows()
        idx = int(idx)
        amount = max(1, int(amount or 1))
        formulas = self._formula_cells()
        self._ws.insert_rows(idx, amount)
        for row, col, formula in formulas:
            if row >= idx:
                self._write_translated_formula(row, col, row + amount, col, formula)
        return self

    def delete_cols(self, idx, amount=1):
        self.flush_pending_rows()
        idx = int(idx)
        amount = max(1, int(amount or 1))
        last_deleted = idx + amount - 1
        formulas = self._formula_cells()
        self._ws.delete_cols(idx, amount)
        for row, col, formula in formulas:
            if col > last_deleted:
                self._write_translated_formula(row, col, row, col - amount, formula)
        return self

    def delete_rows(self, idx, amount=1):
        self.flush_pending_rows()
        idx = int(idx)
        amount = max(1, int(amount or 1))
        last_deleted = idx + amount - 1
        formulas = self._formula_cells()
        self._ws.delete_rows(idx, amount)
        for row, col, formula in formulas:
            if row > last_deleted:
                self._write_translated_formula(row, col, row - amount, col, formula)
        return self

    @property
    def UsedRange(self):
        self.flush_pending_rows()
        mr = self._ws.max_row or 1
        mc = self._ws.max_column or 1
        return _OpxlRange(self._ws, 1, 1, mr, mc)

    def cell(self, row=None, column=None, value=_OPXL_NO_VALUE):
        self.flush_pending_rows()
        ar, ac = _opxl_merged_anchor(self._ws, int(row), int(column))
        proxy = _OpxlCellProxy(self._ws, ar, ac)
        if value is not _OPXL_NO_VALUE:
            proxy.value = value
        return proxy

    def Range(self, a1, a2=None):
        self.flush_pending_rows()
        if a2 is not None:
            r1, c1 = a1._r1, a1._c1
            r2, c2 = a2._r1, a2._c1
            return _OpxlRange(self._ws, min(r1, r2), min(c1, c2), max(r1, r2), max(c1, c2))
        s = str(a1).replace("$", "").strip()
        if ":" in s:
            left, right = s.split(":", 1)
            r1, c1 = _opxl_coord(left)
            r2, c2 = _opxl_coord(right)
            return _OpxlRange(self._ws, min(r1, r2), min(c1, c2), max(r1, r2), max(c1, c2))
        r, c = _opxl_coord(s)
        return _OpxlRange(self._ws, r, c, r, c)


class _OpenpyxlSheetsProxy:
    def __init__(self, workbook_proxy):
        self._workbook_proxy = workbook_proxy

    def __call__(self, name=None):
        return self._workbook_proxy.sheet(name)

    def __getitem__(self, name):
        return self._workbook_proxy.sheet(name)

    def __contains__(self, name):
        raw = self._workbook_proxy.raw
        return str(name) in raw.sheetnames

    def __iter__(self):
        raw = self._workbook_proxy.raw
        for name in raw.sheetnames:
            yield OpenpyxlWorksheetProxy(raw[name])

    def __len__(self):
        return len(self._workbook_proxy.raw.sheetnames)

    def add(self, name=None):
        return self._workbook_proxy._ctx.add_sheet(
            name or "Sheet",
            workbook=self._workbook_proxy._ctx._sheet_add_target(self._workbook_proxy),
        )

    @property
    def names(self):
        return list(self._workbook_proxy.raw.sheetnames)


class OpenpyxlWorkbookProxy:
    def __init__(self, ctx, workbook, name=None):
        self._ctx = ctx
        self._workbook = workbook
        self.name = name or ""
        self._sheets_proxy = _OpenpyxlSheetsProxy(self)

    @property
    def raw(self):
        return self._workbook

    def __getattr__(self, name):
        return getattr(self._workbook, name)

    def __getitem__(self, name):
        return self.sheet(name)

    def __contains__(self, name):
        return str(name) in self._workbook.sheetnames

    def sheet(self, name=None):
        return self._ctx.sheet(name, workbook=self)

    @property
    def sheets(self, name=None):
        return self._sheets_proxy

    def add_sheet(self, name):
        return self._ctx.add_sheet(name, workbook=self)

    def sheet_like(self, name=None):
        return self._ctx.sheet_like(name, workbook=self)

    def range(self, sheet_or_name, address):
        return self._ctx.range(sheet_or_name, address, workbook=self)

    def rows(self, sheet_or_name=None):
        return self._ctx.rows(sheet_or_name, workbook=self)

    def iter_rows(self, sheet_or_name=None, start_row=1):
        return self._ctx.iter_rows(sheet_or_name, workbook=self, start_row=start_row)

    def rows_with_index(self, sheet_or_name=None, start_row=1):
        return self._ctx.iter_rows(sheet_or_name, workbook=self, start_row=start_row)

    def display_rows(self, sheet_or_name=None):
        return self._ctx.display_rows(sheet_or_name, workbook=self)

    def value(self, sheet_or_name, row, col):
        return self._ctx.value(sheet_or_name, row, col, workbook=self)

    def display_value(self, sheet_or_name, row, col):
        return self._ctx.display_value(sheet_or_name, row, col, workbook=self)

    def col(self, sheet_or_name, header, header_rows=20):
        return self._ctx.col(sheet_or_name, header, workbook=self, header_rows=header_rows)

    def header_row(self, sheet_or_name=None, header_rows=20):
        return self._ctx.header_row(sheet_or_name, workbook=self, header_rows=header_rows)


class OpenpyxlSkillContext:
    """COM ExcelSkillContext 와 동일한 API를 openpyxl 위에서 제공한다."""
    def __init__(self, output_wb, input_wbs, output_cached_wb=None, output_name=None, active_file_id=None, active_sheet=None, output_cached_path=None):
        self.excel = None
        self._workbook = output_wb
        self._output_cached_wb = output_cached_wb
        self._output_cached_path = output_cached_path
        self._output_cached_tried = False
        self._dirty_workbook_ids = set()  # [성능] 입력 결과 저장/inspect 스킵 판단용
        self.output_name = output_name or "output"
        self.workbook = OpenpyxlWorkbookProxy(self, output_wb, self.output_name)
        self.output = self.workbook
        self.last_output_sheet = None
        self.last_output_address = None
        self._progress = None  # 느린 루프 진행률 콜백(stage, done, total). 실행기가 주입.
        self.inputs = {
            name: OpenpyxlWorkbookProxy(self, wb, name)
            for name, wb in (input_wbs or {}).items()
        }
        self.active_file_id = str(active_file_id or "")
        self.active_sheet_name = str(active_sheet or "")
        self.active_workbook = self._workbook_for_file_id(self.active_file_id) or self.workbook
        self._last_sheet_workbook_raw = self._unwrap_workbook(self.active_workbook)

    def _unwrap_workbook(self, wb):
        return wb.raw if isinstance(wb, OpenpyxlWorkbookProxy) else wb

    def _sheet_add_target(self, owner):
        owner_raw = self._unwrap_workbook(owner)
        recent_raw = getattr(self, "_last_sheet_workbook_raw", None)
        if owner_raw is self._workbook and recent_raw is not None and recent_raw is not self._workbook:
            return recent_raw
        return owner

    def _workbook_for_file_id(self, file_id):
        file_id = str(file_id or "")
        if not file_id:
            return None
        if file_id == "output" or file_id.startswith("output:"):
            return self.workbook
        if file_id.startswith("input:"):
            hint = file_id[6:]
            try:
                return self.workbook_like(hint)
            except Exception:
                return self.inputs.get(hint)
        return None

    def _default_workbook(self):
        return self.active_workbook or self.workbook

    def _is_output_workbook(self, wb):
        return self._unwrap_workbook(wb) is self._workbook

    def _get_output_cached_wb(self):
        # [성능] 지연 로드: ctx.value/display_value 를 쓰는 스킬에서만 data_only 짝 워크북을 연다.
        if self._output_cached_wb is None and not self._output_cached_tried and self._output_cached_path:
            self._output_cached_tried = True
            try:
                self._output_cached_wb = openpyxl_load_workbook_compatible(Path(self._output_cached_path), data_only=True)
            except Exception:
                self._output_cached_wb = None
        return self._output_cached_wb

    def _cached_ws_for(self, ws):
        raw = getattr(ws, "_ws", ws)
        try:
            cached = self._get_output_cached_wb()
            if raw.parent is self._workbook and cached is not None and raw.title in cached.sheetnames:
                return cached[raw.title]
        except Exception:
            pass
        return None

    def normalize(self, value):
        return normalize_text(value)

    def _sheet_names(self, wb):
        return list(self._unwrap_workbook(wb).sheetnames)

    def workbook_like(self, hint=None):
        if not hint:
            return self.workbook
        norm = self.normalize(hint)
        candidates = [(name, wb) for name, wb in self.inputs.items()]
        candidates.append((self.output_name, self.workbook))
        candidates.append(("output", self.workbook))
        for name, wb in candidates:
            if self.normalize(name) == norm:
                return wb
        for name, wb in candidates:
            if norm in self.normalize(name) or self.normalize(name) in norm:
                return wb
        for _, wb in candidates:
            if self._find_sheet_name(wb, hint, allow_single=False):
                return wb
        if len(self.inputs) == 1:
            return next(iter(self.inputs.values()))
        raise RuntimeError(f"workbook not found: {hint}")

    def input(self, hint=None):
        if hint is None:
            if len(self.inputs) == 1:
                return next(iter(self.inputs.values()))
            raise RuntimeError("input workbook hint is required when multiple input files exist")
        return self.workbook_like(hint)

    def _find_sheet_name(self, wb, name=None, allow_single=True):
        raw = self._unwrap_workbook(wb)
        names = list(raw.sheetnames)
        if not names:
            return None
        if not name:
            try:
                return raw.active.title
            except Exception:
                return names[0]
        alias = _resolve_ephemeral_excel_open_sheet_alias(name, names)
        if alias:
            return alias
        norm = self.normalize(name)
        sheet_norm_loose = normalize_sheet_lookup(name)
        for sheet_name in names:
            if self.normalize(sheet_name) == norm:
                return sheet_name
        for sheet_name in names:
            sheet_norm = self.normalize(sheet_name)
            if norm in sheet_norm or sheet_norm in norm:
                return sheet_name
        for sheet_name in names:
            candidate = normalize_sheet_lookup(sheet_name)
            if sheet_norm_loose and (candidate == sheet_norm_loose or sheet_norm_loose in candidate or candidate in sheet_norm_loose):
                return sheet_name
        if allow_single and len(names) == 1:
            return names[0]
        return None

    def sheet(self, name=None, workbook=None):
        default_wb = workbook or self._default_workbook()
        raw = self._unwrap_workbook(default_wb)
        lookup_name = self.active_sheet_name if (name is None and workbook is None and self.active_sheet_name) else name
        allow_single = (not bool(lookup_name)) or (workbook is not None and not self._is_output_workbook(raw))
        sheet_name = self._find_sheet_name(default_wb, lookup_name, allow_single=allow_single)
        if not sheet_name and workbook is None:
            matches = []
            candidates = [self.workbook] + list(self.inputs.values())
            seen = set()
            for candidate in candidates:
                raw_candidate = self._unwrap_workbook(candidate)
                key = str(id(raw_candidate))
                if key in seen:
                    continue
                seen.add(key)
                candidate_sheet = self._find_sheet_name(raw_candidate, lookup_name, allow_single=False)
                if candidate_sheet:
                    matches.append((raw_candidate, candidate_sheet))
            if len(matches) == 1:
                raw, sheet_name = matches[0]
        if not sheet_name and workbook is not None and self._is_output_workbook(raw) and lookup_name:
            matches = []
            for candidate in self.inputs.values():
                raw_candidate = self._unwrap_workbook(candidate)
                candidate_sheet = self._find_sheet_name(candidate, lookup_name, allow_single=False)
                if candidate_sheet:
                    matches.append((raw_candidate, candidate_sheet))
            if len(matches) == 1:
                raw, sheet_name = matches[0]
        if not sheet_name and workbook is None and lookup_name != name:
            sheet_name = self._find_sheet_name(default_wb, name, allow_single=not bool(name))
        if not sheet_name:
            raise RuntimeError(f"sheet not found: {name}")
        ws = OpenpyxlWorksheetProxy(raw[sheet_name])
        self._last_sheet_workbook_raw = raw
        if self._is_output_workbook(raw):
            self.last_output_sheet = ws.Name
        return ws

    def sheet_like(self, name=None, workbook=None):
        return self.sheet(name, workbook)

    def input_sheet(self, sheet_hint=None, file_hint=None):
        workbooks = []
        if file_hint:
            workbooks.append(self.workbook_like(file_hint))
        else:
            workbooks.extend(self.inputs.values())
        for wb in workbooks:
            sheet_name = self._find_sheet_name(wb, sheet_hint, allow_single=True)
            if sheet_name:
                return OpenpyxlWorksheetProxy(self._unwrap_workbook(wb)[sheet_name])
        raise RuntimeError(f"input sheet not found: {sheet_hint}")

    def range(self, sheet_or_name, address, workbook=None):
        ws = sheet_or_name if hasattr(sheet_or_name, "Range") else self.sheet(sheet_or_name, workbook)
        try:
            if self._is_output_workbook(ws.Parent):
                self.last_output_sheet = ws.Name
                self.last_output_address = str(address)
        except Exception:
            pass
        return ws.Range(str(address))

    def _ws_of(self, sheet_or_name, workbook=None):
        return sheet_or_name if hasattr(sheet_or_name, "UsedRange") else self.sheet(sheet_or_name, workbook)

    def flush_pending_rows(self):
        workbooks = [self._workbook]
        for wb in self.inputs.values():
            raw = self._unwrap_workbook(wb)
            if raw not in workbooks:
                workbooks.append(raw)
        for wb in workbooks:
            for ws in list(getattr(wb, "worksheets", []) or []):
                OpenpyxlWorksheetProxy(ws).flush_pending_rows()

    def rows(self, sheet_or_name, workbook=None):
        ws = self._ws_of(sheet_or_name, workbook)
        if hasattr(ws, "flush_pending_rows"):
            ws.flush_pending_rows()
        raw = getattr(ws, "_ws", ws)
        out = []
        for row in raw.iter_rows(values_only=True):
            out.append(list(row))
        # 끝쪽 완전 빈 행 제거(openpyxl max_row 가 과대평가될 수 있음)
        while out and all(v is None or v == "" for v in out[-1]):
            out.pop()
        return out

    def iter_rows(self, sheet_or_name, workbook=None, start_row=1):
        rows = self.rows(sheet_or_name, workbook=workbook)
        min_row = max(1, int(start_row or 1))
        for excel_row, row in enumerate(rows, start=1):
            if excel_row >= min_row:
                yield excel_row, row

    def rows_with_index(self, sheet_or_name, workbook=None, start_row=1):
        return self.iter_rows(sheet_or_name, workbook=workbook, start_row=start_row)

    def value(self, sheet_or_name, row, col, workbook=None):
        """Return the displayed/calculated value for one cell.

        Use this for "값만 복사" from formula cells. Plain ws.cell(...).value keeps
        the formula string so formula-preserving copy still works.
        """
        ws = self._ws_of(sheet_or_name, workbook)
        if hasattr(ws, "flush_pending_rows"):
            ws.flush_pending_rows()
        raw = getattr(ws, "_ws", ws)
        return _opxl_display_cell_value(raw, int(row), int(col), self._cached_ws_for(raw))

    def display_value(self, sheet_or_name, row, col, workbook=None):
        return self.value(sheet_or_name, row, col, workbook=workbook)

    def display_rows(self, sheet_or_name, workbook=None):
        ws = self._ws_of(sheet_or_name, workbook)
        if hasattr(ws, "flush_pending_rows"):
            ws.flush_pending_rows()
        raw = getattr(ws, "_ws", ws)
        cached_ws = self._cached_ws_for(raw)
        out = []
        for r_idx in range(1, (raw.max_row or 0) + 1):
            row = []
            for c_idx in range(1, (raw.max_column or 0) + 1):
                row.append(_opxl_display_cell_value(raw, r_idx, c_idx, cached_ws))
            out.append(row)
        while out and all(v is None or v == "" for v in out[-1]):
            out.pop()
        return out

    def col(self, sheet_or_name, header, workbook=None, header_rows=20):
        rows = self.rows(sheet_or_name, workbook)
        target = self.normalize(header)
        for row in rows[:header_rows]:
            for c_idx, value in enumerate(row, start=1):
                if self.normalize(value) == target:
                    return ExcelColumnNumber(c_idx)
        for row in rows[:header_rows]:
            for c_idx, value in enumerate(row, start=1):
                if target and target in self.normalize(value):
                    return ExcelColumnNumber(c_idx)
        raise RuntimeError(f"column not found: {header}")

    def header_row(self, sheet_or_name=None, workbook=None, header_rows=20):
        rows = self.rows(sheet_or_name, workbook)
        best_idx = 1
        best_score = -1
        for idx, row in enumerate(rows[:header_rows], start=1):
            score = sum(1 for value in row if value not in (None, ""))
            if score > best_score:
                best_idx, best_score = idx, score
        return best_idx

    def data_start_row(self, sheet_or_name=None, workbook=None, header_rows=20):
        return self.header_row(sheet_or_name, workbook, header_rows) + 1

    def _col0(self, rows, name_or_idx, header_rows=20):
        if isinstance(name_or_idx, (int, ExcelColumnNumber)):
            return max(0, int(name_or_idx) - 1)
        target = self.normalize(name_or_idx)
        scan = rows[:header_rows] if header_rows else rows
        for row in scan:
            for i, v in enumerate(row or []):
                if self.normalize(v) == target:
                    return i
        for row in scan:
            for i, v in enumerate(row or []):
                if target and target in self.normalize(v):
                    return i
        return None

    def add_sheet(self, name, workbook=None):
        wb = self._unwrap_workbook(workbook or self._default_workbook())
        base = re.sub(r"[\[\]:*?/\\]", "_", str(name) or "Sheet")[:31] or "Sheet"
        existing = {self.normalize(n) for n in wb.sheetnames}
        final = base
        idx = 1
        while self.normalize(final) in existing:
            idx += 1
            suffix = "_" + str(idx)
            final = (base[: max(1, 31 - len(suffix))] + suffix)
        raw_ws = wb.create_sheet(title=final)
        try:
            self._dirty_workbook_ids.add(id(wb))
        except Exception:
            pass
        ws = OpenpyxlWorksheetProxy(raw_ws)
        if self._is_output_workbook(wb):
            self.last_output_sheet = ws.Name
        return ws

    def _write_grid(self, ws, grid, start_row=1, start_col=1):
        if not grid:
            return ws
        raw = getattr(ws, "_ws", ws)
        try:
            self._dirty_workbook_ids.add(id(raw.parent))
        except Exception:
            pass
        # 병합 셀 유무를 한 번만 확인. 병합 없는 시트(작업/스크래치 시트 대부분)는 셀마다
        # merged-anchor 스캔(_opxl_write_cell)을 건너뛰고 직접 써서 대용량에서 크게 빨라진다.
        try:
            has_merges = bool(raw.merged_cells.ranges)
        except Exception:
            has_merges = True  # 알 수 없으면 안전(기존) 경로
        total = len(grid)
        cell = raw.cell
        _apply_openpyxl_text_format_for_long_digit_columns(raw, grid, start_row, start_col)
        report = self._progress if (total > 20000 and callable(self._progress)) else None
        for i, row in enumerate(grid):
            R = start_row + i
            if has_merges:
                for j, value in enumerate(row or []):
                    _opxl_write_cell(raw, R, start_col + j, value, redirect_merged=False)
            else:
                for j, value in enumerate(row or []):
                    cell(row=R, column=start_col + j, value=value)
            if report is not None and (i % 20000 == 0):
                try: report("쓰는 중", i, total)
                except Exception: pass
        if report is not None:
            try: report("쓰는 중", total, total)
            except Exception: pass
        return ws

    def write_grid(self, ws, grid, start_row=1, start_col=1):
        ws = self._ws_of(ws)
        self._write_grid(ws, grid, start_row, start_col)
        if self._is_output_workbook(ws.Parent):
            self.last_output_sheet = ws.Name
        return ws

    def set_range(self, sheet_or_name, address, grid, workbook=None):
        ws = self._ws_of(sheet_or_name, workbook)
        if not grid:
            return ws
        r0, c0 = _opxl_coord(str(address).split(":")[0])
        self._write_grid(ws, grid, start_row=r0, start_col=c0)
        if self._is_output_workbook(ws.Parent):
            self.last_output_sheet = ws.Name
            self.last_output_address = str(address)
        return ws

    def sort(self, sheet_or_name, by, ascending=True, header=True, workbook=None):
        # 다중키 지원: by 는 단일 컬럼명/인덱스 또는 컬럼 리스트. ascending 도 bool 또는 리스트.
        ws = self._ws_of(sheet_or_name, workbook)
        rows = self.rows(ws)
        keys = list(by) if isinstance(by, (list, tuple)) else [by]
        rels = []
        for k in keys:
            rel = self._col0(rows, k)
            if rel is None:
                raise RuntimeError("sort: column not found: %r" % (k,))
            rels.append(rel)
        asc_list = list(ascending) if isinstance(ascending, (list, tuple)) else [ascending] * len(rels)
        while len(asc_list) < len(rels):
            asc_list.append(asc_list[-1] if asc_list else True)
        hdr_count = 1 if header else 0
        head = rows[:hdr_count]
        body = rows[hdr_count:]

        def _cellkey(r, rel):
            v = r[rel] if rel < len(r) else None
            num = self._num(v)
            return (0, num) if num is not None else (1, self.normalize(v))

        if len(set(bool(a) for a in asc_list)) <= 1:
            rev = not bool(asc_list[0]) if asc_list else False
            body.sort(key=lambda r: tuple(_cellkey(r, rel) for rel in rels), reverse=rev)
        else:
            # 키별 정렬 방향이 섞이면 안정정렬을 마지막 키부터 적용
            for i in range(len(rels) - 1, -1, -1):
                body.sort(key=lambda r, rel=rels[i]: _cellkey(r, rel), reverse=not bool(asc_list[i]))

        # 정렬은 행 수를 보존하므로, 셀 단위 전체 clear 루프(느림) 대신 직사각형 그리드로 한 번에 재기록.
        max_col = max((len(r) for r in rows), default=0)
        grid = list(head) + body
        padded = [list(r) + [None] * (max_col - len(r)) for r in grid]
        raw = getattr(ws, "_ws", ws)
        self._write_grid(ws, padded)
        # 혹시 기존 시트가 새 그리드보다 길면 그 잔여행만 비운다(보통 없음).
        old_max = raw.max_row or 0
        if old_max > len(padded):
            for r_idx in range(len(padded) + 1, old_max + 1):
                for c_idx in range(1, max_col + 1):
                    raw.cell(row=r_idx, column=c_idx, value=None)
        if self._is_output_workbook(ws.Parent):
            self.last_output_sheet = ws.Name
        return ws

    def filter_to_sheet(self, sheet_or_name, predicate, dest_name, header_rows=1, workbook=None):
        ws = self._ws_of(sheet_or_name, workbook)
        rows = self.rows(ws)
        # [0.5.15] COM UsedRange 는 '첫 사용열'부터 시작 → A열이 비면 행 인덱스가 밀려 predicate 의 절대 열
        # 인덱스가 어긋난다. 선두 빈 열 수만큼 None 패딩해 'A열=index0' 으로 맞춘다(openpyxl 은 이미 A1 기준
        # 이라 lead=0 → 무영향). 두 엔진의 filter_to_sheet 열 인덱싱을 일치시킨다.
        try:
            _lead = int(ws.UsedRange.Column) - 1
        except Exception:
            _lead = 0
        if _lead > 0:
            rows = [[None] * _lead + list(r) for r in rows]
        hr = max(0, int(header_rows or 0))
        header = rows[:hr]
        matched = []
        body = rows[hr:]
        total = len(body)
        report = getattr(self, "_progress", None)
        report = report if (total > 20000 and callable(report)) else None
        for k, r in enumerate(body):
            try:
                if predicate(r):
                    matched.append(r)
            except Exception:
                continue
            if report is not None and (k % 20000 == 0):
                try: report("거르는 중", k, total)
                except Exception: pass
        dest_wb = workbook
        if dest_wb is None:
            try:
                dest_wb = ws.Parent
            except Exception:
                dest_wb = self._default_workbook()
        dest = self.add_sheet(dest_name, workbook=dest_wb)
        self._write_grid(dest, list(header) + matched)
        return dest

    def _merge_pivot_grid_into_base(self, workbook, dest_name, grid):
        name = str(dest_name or "")
        if not name or name.endswith("_피벗") or "_" not in name or not grid or len(grid[0]) < 2:
            return
        prefix = name.rsplit("_", 1)[0]
        base_names = [prefix + "_피벗", prefix + "_pivot"]
        base_ws = None
        for base_name in base_names:
            try:
                base_ws = self.sheet(base_name, workbook=workbook)
                break
            except Exception:
                base_ws = None
        if base_ws is None:
            return
        if self.normalize(base_ws.Name) == self.normalize(name):
            return
        base_rows = self.rows(base_ws)
        if not base_rows:
            return
        base_header = list(base_rows[0] or [])
        src_header = list(grid[0] or [])
        add_cols = []
        for src_idx, label in enumerate(src_header[1:], start=1):
            if not any(self.normalize(label) == self.normalize(h) for h in base_header):
                add_cols.append((src_idx, label))
        if not add_cols:
            return
        out = [base_header + [label for _, label in add_cols]]
        key_to_values = {}
        for row in grid[1:]:
            if not row:
                continue
            key_to_values[self.normalize(row[0])] = row
        for row in base_rows[1:]:
            cur = list(row or [])
            key = self.normalize(cur[0] if cur else "")
            src = key_to_values.get(key)
            cur += [(src[i] if src is not None and i < len(src) else None) for i, _ in add_cols]
            out.append(cur)
        existing_keys = {self.normalize((row or [""])[0]) for row in base_rows[1:]}
        for row in grid[1:]:
            if not row or self.normalize(row[0]) in existing_keys:
                continue
            cur = [row[0]] + [None] * (len(base_header) - 1)
            cur += [(row[i] if i < len(row) else None) for i, _ in add_cols]
            out.append(cur)
        self._write_grid(base_ws, out)

    @staticmethod
    def _num(v):
        if isinstance(v, bool):
            return None
        if isinstance(v, (int, float)):
            return float(v)
        if isinstance(v, str):
            s = v.strip().replace(",", "")
            try:
                return float(s)
            except ValueError:
                return None
        return None

    def pivot(self, sheet_or_name, group_by=None, value=None, agg="sum", dest_name=None, header_rows=1, workbook=None, **kwargs):
        if group_by is None:
            group_by = kwargs.get("rows")
        if value is None and "values" in kwargs:
            value = kwargs.get("values")
        if dest_name is None:
            dest_name = kwargs.get("name") or kwargs.get("dest")
        ws = self._ws_of(sheet_or_name, workbook)
        rows = self.rows(ws)
        hr = max(1, int(header_rows or 1))
        header_row = rows[hr - 1] if len(rows) >= hr else []
        data = rows[hr:]
        group_cols = list(group_by) if isinstance(group_by, (list, tuple)) else [group_by]
        gidx = [self._col0(rows, g, hr) for g in group_cols]
        values = list(value) if isinstance(value, (list, tuple)) else [value]
        aggs = list(agg) if isinstance(agg, (list, tuple)) else [agg] * len(values)
        while len(aggs) < len(values):
            aggs.append(aggs[-1] if aggs else "sum")
        aggs = [str(a or "sum").lower() for a in aggs]
        vidxs = [self._col0(rows, v, hr) if v is not None else None for v in values]

        groups = {}
        order = []
        for r in data:
            key = tuple((r[i] if (i is not None and i < len(r)) else "") for i in gidx)
            if key not in groups:
                groups[key] = [[] for _ in values]
                order.append(key)
            for pos, vidx in enumerate(vidxs):
                if vidx is None:
                    groups[key][pos].append(1)
                elif vidx < len(r):
                    groups[key][pos].append(r[vidx])

        def _aggregate(vals, agg_name):
            nums = [n for n in (self._num(v) for v in vals) if n is not None]
            if agg_name == "count":
                return len(vals)
            if agg_name in ("avg", "average", "mean"):
                return (sum(nums) / len(nums)) if nums else 0
            if agg_name == "max":
                return max(nums) if nums else ""
            if agg_name == "min":
                return min(nums) if nums else ""
            return sum(nums)

        out_header = []
        for n, i in enumerate(gidx):
            label = header_row[i] if (i is not None and i < len(header_row)) else ("그룹%d" % (n + 1))
            out_header.append(label)
        for v, agg_name in zip(values, aggs):
            label = str(v) if v is not None else "값"
            out_header.append(label + "_" + (agg_name if agg_name != "average" else "avg"))
        grid = [out_header]
        for key in order:
            grid.append(list(key) + [_aggregate(groups[key][i], aggs[i]) for i in range(len(values))])
        dest_wb = workbook
        if dest_wb is None:
            try:
                dest_wb = ws.Parent
            except Exception:
                dest_wb = self._default_workbook()
        dest = self.add_sheet(dest_name or "피벗요약", workbook=dest_wb)
        self._write_grid(dest, grid)
        self._merge_pivot_grid_into_base(dest_wb, dest.Name, grid)
        return dest


# ===== 헬퍼 인자 이름 관용 처리 ==========================================================
# [실측 2026-08-06] 사용자 제보: 피벗 생성 스킬이 `ctx.pivot(..., header_row=2)` 로 만들어져
#   "pivot() got an unexpected keyword argument 'header_row'" 로 통째로 실패했다.
#
#   진짜 원인은 모델이 아니라 우리 API 다. 같은 뜻(헤더가 몇 번째 행인지)을 두 이름으로 쓴다:
#     header_row  (단수) : find_header, apply_filter, lookup, dedupe, add_total_row, split_column ...
#     header_rows (복수) : pivot, native_pivot, filter_to_sheet ...
#   단수 쪽이 훨씬 많아서, 그 습관대로 pivot 에 쓰면 터졌다. 프롬프트로 "pivot 은 복수"를 외우게
#   하는 건 임시방편이라(그 힌트는 filter_to_sheet 용으로만 있었다), 아예 **둘 다 받도록** 한다.
#
#   덤으로, 정말 없는 옵션을 준 경우에도 원시 TypeError 대신 '쓸 수 있는 옵션 목록'을 알려준다.
#   자동복구가 그 메시지를 보고 고칠 수 있어야 한 번에 끝난다.

_HEADER_ALIAS_PAIR = ("header_row", "header_rows")


def _wrap_ctx_helper_kwargs(fn, primary, other, has_varkw, allowed):
    @functools.wraps(fn)
    def wrapper(*args, **kwargs):
        # ① header_row ↔ header_rows: 어느 쪽으로 불러도 받는다(같은 뜻).
        if primary and other in kwargs:
            value = kwargs.pop(other)
            kwargs.setdefault(primary, value)
        # ② 없는 옵션은 '무엇을 쓸 수 있는지' 까지 말해 주는 오류로 바꾼다.
        if not has_varkw:
            unknown = [k for k in kwargs if k not in allowed]
            if unknown:
                raise PythonComSkillError(
                    "%s() 에 없는 옵션 %s 를 넘겼습니다. 쓸 수 있는 옵션: %s"
                    % (fn.__name__, ", ".join(sorted(unknown)), ", ".join(sorted(allowed)) or "(없음)")
                )
        return fn(*args, **kwargs)
    wrapper._b2b_kwarg_tolerant = True       # 검증/테스트에서 '내가 씌운 것'만 세기 위한 표식
    return wrapper


def _install_ctx_kwarg_tolerance(*classes):
    """ctx 클래스의 공개 메서드를 훑어 header_row/header_rows 를 서로 받아 주도록 감싼다.
    두 이름 중 하나를 가진 메서드만 감싸므로 나머지 동작에는 영향이 없다."""
    for cls in classes:
        for name, member in list(vars(cls).items()):
            if name.startswith("_") and name != "_pivot_value_table":
                continue
            if not inspect.isfunction(member):
                continue
            try:
                sig = inspect.signature(member)
            except (TypeError, ValueError):
                continue
            params = sig.parameters
            has_varkw = any(p.kind == inspect.Parameter.VAR_KEYWORD for p in params.values())
            primary = None
            for cand in _HEADER_ALIAS_PAIR:
                if cand in params:
                    primary = cand
                    break
            if primary is None:
                continue                     # 헤더 인자가 없는 메서드는 건드리지 않는다
            other = _HEADER_ALIAS_PAIR[1] if primary == _HEADER_ALIAS_PAIR[0] else _HEADER_ALIAS_PAIR[0]
            allowed = {
                p.name for p in params.values()
                if p.kind in (inspect.Parameter.POSITIONAL_OR_KEYWORD, inspect.Parameter.KEYWORD_ONLY)
                and p.name != "self"
            }
            allowed.add(other)               # 별칭도 '쓸 수 있는 옵션'으로 안내
            setattr(cls, name, _wrap_ctx_helper_kwargs(member, primary, other, has_varkw, allowed))


_install_ctx_kwarg_tolerance(PythonComSkillContext, ExcelSkillContext, OpenpyxlSkillContext)


def _run_openpyxl_python_pipeline_impl(payload, job_id=None):
    _pp0 = time.perf_counter(); _pp = {"mode": "openpyxl"}  # F8 패널용
    if openpyxl is None:
        raise RuntimeError("openpyxl 이 설치되어 있지 않습니다(순수 Python 엔진 사용 불가).")
    output_item = payload.get("output") or {}
    if not output_item.get("backendWorkbookId"):
        raise RuntimeError("Python Excel skills require an output workbook.")
    input_items = payload.get("inputs", [])
    output_wb_record = get_workbook_or_raise(output_item.get("backendWorkbookId"))
    input_wb_records = [get_workbook_or_raise(item.get("backendWorkbookId")) for item in input_items]
    active_steps = [s for s in (payload.get("pipeline") or []) if not (s and s.get("enabled") is False)]
    python_steps = [s for s in active_steps if is_python_pipeline_step(s)]
    if len(python_steps) != len(active_steps):
        raise RuntimeError("순수 Python 엔진은 한 실행에서 JavaScript 단계와 섞을 수 없습니다.")

    update_pipeline_job(job_id, {
        "stage": "openpyxl 엔진 준비 중",
        "currentStep": 0,
        "completedSteps": 0,
        "totalSteps": len(python_steps),
        "stepRunning": True,
    })

    output_path = output_wb_record["path"]
    output_path_norm = str(Path(output_path).resolve()).lower()
    # 출력: data_only=False 로 열어 수식을 보존하고 값을 쓴다. 읽기는 쓴 값이 그대로 반영된다(read-after-write).
    output_wb = openpyxl_load_workbook_compatible(Path(output_path), data_only=False)
    # 값만 복사에서 기존 수식 셀의 표시값을 읽기 위한 짝 워크북 — [성능] 지연 로드.
    # 풀 파싱이 대용량에서 수십 초라, ctx.value/display_value 를 실제로 쓰는 스킬에서만 로드한다.
    output_cached_wb = None  # OpenpyxlSkillContext 가 output_cached_path 로 필요 시 로드

    # 입력: data_only=True 로 열어 수식의 "계산된 값"을 읽는다(Excel 이 저장해둔 캐시값).
    # openpyxl 엔진에서 입력은 읽기 전용으로 취급한다(저장하면 수식이 사라지므로). 입력 편집은 Excel 엔진 사용.
    input_wbs = {}
    for item, rec in zip(input_items, input_wb_records):
        name = item.get("name") or rec["name"]
        path_norm = str(Path(rec["path"]).resolve()).lower()
        if path_norm == output_path_norm:
            input_wbs[name] = output_wb
            continue
        wb = openpyxl_load_workbook_compatible(Path(rec["path"]), data_only=True)
        input_wbs[name] = wb

    output_name = output_item.get("name") or output_wb_record["name"]
    current = payload.get("current") or {}
    ctx = OpenpyxlSkillContext(
        output_wb,
        input_wbs,
        output_cached_wb=output_cached_wb,
        output_cached_path=str(output_path),
        output_name=output_name,
        active_file_id=current.get("fileId"),
        active_sheet=current.get("sheet"),
    )
    _pp["loadMs"] = round((time.perf_counter() - _pp0) * 1000, 1); _pp_steps = time.perf_counter()
    for idx, step in enumerate(python_steps, start=1):
        raise_if_pipeline_cancelled(job_id)  # 협조적 취소(스텝 경계)
        update_pipeline_job(job_id, {
            "stage": f"Python(openpyxl) Step {idx}/{len(python_steps)} 실행 중",
            "currentStep": idx,
            "completedSteps": idx - 1,
            "stepRunning": True,
            "errorInfo": None,
        })
        # 느린 루프(대용량 정렬/필터/쓰기)에서 행 진행률을 stage 로 노출(프론트가 폴링해 표시).
        ctx._progress = lambda stage, done, total, _i=idx, _n=len(python_steps): update_pipeline_job(job_id, {
            "stage": f"Python(openpyxl) Step {_i}/{_n} — {stage} {done:,}/{total:,}행",
            "currentStep": _i,
            "completedSteps": _i - 1,
            "stepRunning": True,
        })
        original_code = str(step.get("code") or "")
        code = normalize_python_pipeline_code(original_code)
        namespace = _safe_python_globals()
        try:
            stage_label = "compile"
            exec(compile(code, f"<pipeline_step_{idx}>", "exec"), namespace, namespace)
            stage_label = "lookup transform"
            transform = namespace.get("transform")
            if not callable(transform):
                raise RuntimeError("Python step must define def transform(ctx):")
            stage_label = "transform"
            transform(ctx)
            ctx.flush_pending_rows()
        except Exception as err:
            _cause, _guide = _pipeline_error_guide(str(err), original_code)
            raise PipelineExecutionError({
                "stepIdx": idx - 1,
                "stepId": step.get("id"),
                "description": step.get("description"),
                "code": original_code,
                "normalizedCode": code,
                "language": step.get("language") or "python",
                "message": f"{_cause}\n💡 이렇게 요청해 보세요: {_guide}\n(자세히: {stage_label} 단계 — {err})",
                "cause": _cause,
                "promptGuide": _guide,
                "rawError": f"{stage_label}: {err}",
                "stack": repr(err),
            })

    _pp["stepsMs"] = round((time.perf_counter() - _pp_steps) * 1000, 1); _pp_save = time.perf_counter()
    update_pipeline_job(job_id, {
        "stage": "결과 저장 중",
        "currentStep": len(python_steps),
        "completedSteps": len(python_steps),
        "stepRunning": False,
    })
    BACKEND_DIR.mkdir(parents=True, exist_ok=True)
    ctx.flush_pending_rows()

    # 출력 저장. openpyxl 은 수식을 계산하지 않으므로(셀에 캐시값 없음), Excel 이 열 때 전체 재계산하도록
    # fullCalcOnLoad 를 켠다. → 미러(실제 Excel)와 다운로드 파일에서 수식이 새 값으로 보인다.
    try:
        output_wb.calculation.fullCalcOnLoad = True
    except Exception:
        pass
    original_name = output_item.get("name") or output_wb_record["name"]
    safe_name = Path(str(original_name)).name
    if not Path(safe_name).suffix:
        safe_name += ".xlsx"
    result_path = BACKEND_DIR / f"{uuid.uuid4().hex}_result_{safe_name}"
    output_wb.save(str(result_path))

    # 입력 워크북도 파이프라인 안에서 새 시트/정렬/필터 결과가 만들어질 수 있다.
    # 3.7 JS 실행기와 같은 동작을 위해 수정된 입력 결과도 저장/다운로드 대상으로 노출한다.
    input_previews = {}
    input_download_urls = {}
    # [성능] 입력은 정책상 읽기 전용 — 스킬이 실제로 건드린 입력만 저장/inspect 한다.
    # (ctx 변이 헬퍼의 dirty 마킹 + 코드 정규식 휴리스틱. B2B_ALWAYS_SAVE_INPUTS=1 로 기존 동작 복원)
    _code_all = chr(10).join(str(s.get("code") or "") for s in python_steps)
    _inputs_maybe_written = bool(
        os.environ.get("B2B_ALWAYS_SAVE_INPUTS") == "1"
        or (re.search(r"ctx\.input", _code_all) and re.search(r"\.value\s*=|insert_(?:rows|cols)|delete_(?:rows|cols)|merge_cells|\.append\s*\(", _code_all))
    )
    for item, rec in zip(input_items, input_wb_records):
        name = item.get("name") or rec["name"]
        wb = input_wbs.get(name)
        if wb is None or wb is output_wb:
            continue
        if not _inputs_maybe_written and id(wb) not in getattr(ctx, "_dirty_workbook_ids", set()):
            continue  # 변경 흔적 없는 입력: 저장(수 초)+inspect(재파싱) 생략
        safe_input_name = Path(str(name)).name
        if not Path(safe_input_name).suffix:
            safe_input_name += ".xlsx"
        input_result_path = BACKEND_DIR / f"{uuid.uuid4().hex}_result_{safe_input_name}"
        try:
            wb.save(str(input_result_path))
        except Exception as err:
            raise RuntimeError(f"입력 결과 저장 실패({name}): {err}") from err
        input_result_id = uuid.uuid4().hex
        RESULTS[input_result_id] = {
            "path": str(input_result_path),
            "name": input_result_path.name,
            "created": time.time(),
        }
        inspected_input = inspect_workbook(input_result_path)
        input_previews[name] = inspected_input.get("sheets") or {}
        try:
            update_workbook_current_cache(rec, rows_only_sheets(input_previews[name]))
        except Exception:
            pass
        input_download_urls[f"input:{name}"] = f"/api/workbooks/download/{input_result_id}"

    output_file_id = current.get("outputFileId") or "output:0"
    result_id = uuid.uuid4().hex
    RESULTS[result_id] = {"path": str(result_path), "name": result_path.name, "created": time.time()}
    inspected = inspect_workbook(result_path)
    result_output = inspected.get("sheets") or {}
    update_workbook_current_cache(output_wb_record, rows_only_sheets(result_output))
    previews = build_result_previews(input_previews, result_output, current, {}, [])
    download_urls = dict(input_download_urls)
    download_urls[output_file_id] = f"/api/workbooks/download/{result_id}"
    # 결과 응답으로 마지막 작업 시트를 넘기면 프런트/미러가 사용자의 현재 탭을 바꾸기 쉽다.
    active_output_sheet = None
    active_output_address = None
    _pp["saveInspectMs"] = round((time.perf_counter() - _pp_save) * 1000, 1)
    _pp["totalServerMs"] = round((time.perf_counter() - _pp0) * 1000, 1)
    return {
        "ok": True,
        "pythonExcel": True,
        "engine": "openpyxl",
        "debugTimings": _pp,
        "snapshotHit": False,
        "snapshotStep": 0,
        "diffId": None,
        "diffs": {},
        "forcedValueCells": [],
        "downloadId": result_id,
        "downloadUrl": f"/api/workbooks/download/{result_id}",
        "downloadUrls": download_urls,
        "files": previews,
        "activeSheet": active_output_sheet,
        "activeAddress": active_output_address,
    }


def run_openpyxl_python_pipeline_payload(payload, job_id=None):
    # COM/STA 가 필요 없으므로 잡 스레드에서 바로 실행한다.
    return _run_openpyxl_python_pipeline_impl(payload, job_id=job_id)


def _open_excel_workbook_for_skill(app, path, read_only=False, intended_name=None):
    _park_excel_app_offscreen(app)
    wb, temp_path = excel_workbooks_open(app, path, read_only=read_only, intended_name=intended_name or path)
    _park_excel_app_offscreen(app)
    _hide_excel_app_window(app)
    return wb, temp_path


def _get_python_skill_app():
    # 매 적용마다 Excel 을 새로 띄우고 Quit 하던 비용(콜드스타트 1~3초)을 없애기 위해
    # 숨김 Excel 인스턴스를 한 번만 만들어 재사용한다. 죽었으면 다시 만든다.
    # 반드시 EXCEL_QUEUE STA 워커 스레드에서만 호출된다(excel_call 경유).
    global PYTHON_SKILL_APP, PYTHON_SKILL_APP_PID, PYTHON_SKILL_APP_LAST_USED
    app = PYTHON_SKILL_APP
    PYTHON_SKILL_APP_LAST_USED = time.time()
    if app is not None:
        try:
            _ = app.Workbooks.Count  # 살아있는지 확인
            return app
        except Exception:
            PYTHON_SKILL_APP = None
            PYTHON_SKILL_APP_PID = None
    app = win32com.client.DispatchEx("Excel.Application")
    _track_spawned_excel_app(app)  # [0.5.2 이식] 고아 Excel 추적
    app.Visible = False
    for attr, value in (("DisplayAlerts", False), ("EnableEvents", False), ("AskToUpdateLinks", False)):
        try:
            setattr(app, attr, value)
        except Exception:
            pass
    _hide_excel_app_window(app)
    PYTHON_SKILL_APP = app
    try:
        PYTHON_SKILL_APP_PID = _excel_process_id(app)
    except Exception:
        PYTHON_SKILL_APP_PID = None
    _perf_trace("excel.python_skill.spawned", pid=PYTHON_SKILL_APP_PID)
    return app


def _quit_python_skill_app():
    global PYTHON_SKILL_APP, PYTHON_SKILL_APP_PID, PYTHON_SKILL_APP_LAST_USED
    app = PYTHON_SKILL_APP
    pid = PYTHON_SKILL_APP_PID
    PYTHON_SKILL_APP = None
    PYTHON_SKILL_APP_PID = None
    PYTHON_SKILL_APP_LAST_USED = 0.0
    if app is None:
        return
    try:
        app.Quit()
    except Exception:
        pass
    if pid:
        deadline = time.time() + 1.5
        while time.time() < deadline and _is_pid_alive(pid):
            time.sleep(0.1)
        if _is_pid_alive(pid):
            _perf_trace("excel.python_skill.force_kill_after_quit", pid=pid)
            _force_kill_pid(pid)


def _maybe_quit_idle_python_skill_app():
    """Run only on the Excel COM STA worker. Do not call from HTTP threads."""
    global PYTHON_SKILL_APP_REAP_CHECK_AT
    now = time.time()
    if now - float(PYTHON_SKILL_APP_REAP_CHECK_AT or 0) < 30:
        return
    PYTHON_SKILL_APP_REAP_CHECK_AT = now
    if PYTHON_SKILL_APP is None:
        return
    last_used = float(PYTHON_SKILL_APP_LAST_USED or 0)
    if last_used <= 0 or now - last_used < PYTHON_SKILL_APP_IDLE_TTL_SECONDS:
        return
    pid = PYTHON_SKILL_APP_PID
    _perf_trace(
        "excel.python_skill.idle_quit",
        pid=pid,
        idleSeconds=round(now - last_used, 1),
        ttlSeconds=PYTHON_SKILL_APP_IDLE_TTL_SECONDS,
    )
    _quit_python_skill_app()


def _workbook_fingerprint(wb_record):
    path = Path(wb_record["path"])
    try:
        stat = path.stat()
        stamp = [str(path.resolve()).lower(), stat.st_size, int(stat.st_mtime_ns)]
    except OSError:
        stamp = [str(path).lower(), 0, 0]
    return {
        "id": wb_record.get("id"),
        "name": wb_record.get("name"),
        "stamp": stamp,
    }


def _step_signature(step):
    language = step.get("language") or ("python" if is_python_pipeline_step(step) else "javascript")
    code = step.get("code") or ""
    if str(language).lower() in ("python", "py") or is_python_pipeline_step(step):
        code = normalize_python_pipeline_code(code)
    return {
        "id": step.get("id"),
        "language": language,
        "enabled": step.get("enabled") is not False,
        "code": code,
        "description": step.get("description") or "",
    }


def _pipeline_snapshot_key(input_items, input_wbs, output_item, output_wb_record, steps_prefix):
    payload = {
        "version": 2,
        "inputs": [
            {
                "name": item.get("name") or wb["name"],
                "workbook": _workbook_fingerprint(wb),
            }
            for item, wb in zip(input_items, input_wbs)
        ],
        "output": {
            "name": output_item.get("name") or output_wb_record["name"],
            "workbook": _workbook_fingerprint(output_wb_record),
        },
        "steps": [_step_signature(step) for step in steps_prefix],
    }
    raw = json.dumps(payload, ensure_ascii=False, sort_keys=True, separators=(",", ":")).encode("utf-8")
    return hashlib.sha256(raw).hexdigest()


def _snapshot_path(snapshot, role, name):
    if not snapshot:
        return None
    return snapshot.get("files", {}).get(f"{role}:{name}")


def _snapshot_files_exist(snapshot):
    if not snapshot:
        return False
    return all(Path(path).exists() for path in (snapshot.get("files") or {}).values())


def _find_best_pipeline_snapshot(input_items, input_wbs, output_item, output_wb_record, active_steps):
    best = None
    for prefix_len in range(len(active_steps), 0, -1):
        key = _pipeline_snapshot_key(input_items, input_wbs, output_item, output_wb_record, active_steps[:prefix_len])
        snapshot = PIPELINE_STEP_SNAPSHOTS.get(key)
        if snapshot and _snapshot_files_exist(snapshot):
            best = (prefix_len, key, snapshot)
            break
    return best


def _cleanup_pipeline_step_snapshots():
    _cleanup_pipeline_snapshots_by_limits()


# ===== 라이브 최종상태 스냅샷(엔진 무관) — 새로고침 후 재실행 대신 이 파일로 연다 =====

LIVE_FINAL_SNAPSHOT_DIRNAME = "live_final_snapshots"


def _live_final_snapshot_key(wb_record, state_sig):
    """키 = 원본 파일 지문 + 클라가 계산한 파이프라인 상태 서명.
    원본 지문에 크기·mtime 이 들어가는데, 라이브는 작업복사본을 쓰고 원본을 저장하지 않으므로
    (Close(SaveChanges=False)) 새로고침 뒤에도 지문이 그대로다 → 같은 키가 재현된다.
    상태 서명은 백엔드가 해석하지 않는다 — 저장 때와 조회 때 클라가 같은 규칙으로 만들기만 하면 된다."""
    payload = {
        "version": 1,
        "workbook": _workbook_fingerprint(wb_record),
        "stateSig": str(state_sig or ""),
    }
    raw = json.dumps(payload, ensure_ascii=False, sort_keys=True, separators=(",", ":")).encode("utf-8")
    return hashlib.sha256(raw).hexdigest()


def _live_final_snapshot_stats():
    root = (BACKEND_DIR / LIVE_FINAL_SNAPSHOT_DIRNAME).resolve()
    total = 0
    files = 0
    missing = 0
    for snap in list(LIVE_FINAL_SNAPSHOTS.values()):
        try:
            path = Path(snap.get("path") or "")
            if root not in path.resolve().parents:
                continue
            if path.exists():
                total += path.stat().st_size
                files += 1
            else:
                missing += 1
        except Exception:
            missing += 1
    return {"count": len(LIVE_FINAL_SNAPSHOTS), "files": files, "missingFiles": missing, "bytes": total}


def _cleanup_live_final_snapshots():
    """오래된 것부터 정리(개수·용량 한도는 스텝 스냅샷과 공유). 지워졌으면 조회가 실패하고
    호출부는 '평소대로 전체 재실행'으로 폴백하므로, 정리가 기능을 깨지 않는다."""
    root = (BACKEND_DIR / LIVE_FINAL_SNAPSHOT_DIRNAME).resolve()
    ordered = sorted(LIVE_FINAL_SNAPSHOTS.items(), key=lambda item: item[1].get("created", 0))
    removed = 0
    while ordered:
        stats = _live_final_snapshot_stats()
        if stats["count"] <= MAX_PIPELINE_STEP_SNAPSHOTS and stats["bytes"] <= HOUSEKEEPING_SNAPSHOT_MAX_BYTES:
            break
        key, snap = ordered.pop(0)
        LIVE_FINAL_SNAPSHOTS.pop(key, None)
        try:
            path = Path(snap.get("path") or "").resolve()
            if root in path.parents:
                path.unlink(missing_ok=True)
                shutil.rmtree(path.parent, ignore_errors=True)
        except Exception:
            pass
        removed += 1
    return removed


def _save_live_final_snapshot(wb_record, state_sig, src_path, move=False, link=False):
    """이미 디스크에 있는 결과 파일(src_path)을 최종상태 사본으로 등록한다. 모드 3가지:
      move=True : 옮긴다 — VBA 전체실행의 임시 결과 파일(어차피 rmtree 로 버려질 것).
      link=True : 그 자리를 가리키기만 한다 — Python 경로의 스텝 스냅샷(이미 저장돼 있고 수명도 관리됨).
                  그쪽이 정리로 지워지면 조회가 없음으로 떨어져 평소대로 전체 재실행이 된다(안전).
      기본      : 복사.
    실패해도 절대 예외를 올리지 않는다 — 이건 '있으면 빠른' 부가기능이지 실행 성패와 무관하다."""
    if not wb_record or not state_sig:
        return None
    try:
        src = Path(src_path)
        if not src.exists():
            return None
        key = _live_final_snapshot_key(wb_record, state_sig)
        if link:
            dest = src
        else:
            dest_dir = BACKEND_DIR / LIVE_FINAL_SNAPSHOT_DIRNAME / key
            dest_dir.mkdir(parents=True, exist_ok=True)
            dest = dest_dir / (Path(wb_record.get("name") or src.name).name or src.name)
            if move:
                try:
                    shutil.move(str(src), str(dest))
                except Exception:
                    shutil.copy2(str(src), str(dest))
            else:
                shutil.copy2(str(src), str(dest))
        LIVE_FINAL_SNAPSHOTS[key] = {
            "key": key,
            "path": str(dest),
            "workbookId": wb_record.get("id"),
            "name": wb_record.get("name"),
            "created": time.time(),
        }
        _cleanup_live_final_snapshots()
        return LIVE_FINAL_SNAPSHOTS[key]
    except Exception as err:
        _warn_excel_nonfatal("live final snapshot", err)
        return None


def _find_live_final_snapshot(wb_record, state_sig):
    if not wb_record or not state_sig:
        return None
    snap = LIVE_FINAL_SNAPSHOTS.get(_live_final_snapshot_key(wb_record, state_sig))
    if not snap:
        return None
    try:
        if not Path(snap["path"]).exists():
            LIVE_FINAL_SNAPSHOTS.pop(snap.get("key"), None)   # 파일만 사라진 유령 항목 제거
            return None
    except Exception:
        return None
    return snap


def _save_pipeline_step_snapshot(key, step_idx, app, output_wb, input_wb_by_name, input_stable_src=None):
    snapshot_dir = BACKEND_DIR / "pipeline_step_snapshots" / key
    snapshot_dir.mkdir(parents=True, exist_ok=True)
    files = {}
    output_path = snapshot_dir / "output.xlsx"
    output_wb.SaveCopyAs(str(output_path))
    files["output:output"] = str(output_path)
    input_stable_src = input_stable_src or {}
    for name, wb in input_wb_by_name.items():
        # 스킬이 수정하지 않은 입력(wb.Saved == True)은 안정적인 원본 경로를 그대로 참조해
        # 전체 복사를 건너뛴다(여러 입력 파일일 때 큰 속도 차이).
        stable = input_stable_src.get(name)
        try:
            unmodified = bool(stable) and bool(wb.Saved) and Path(stable).exists()
        except Exception:
            unmodified = False
        if unmodified:
            files[f"input:{name}"] = str(stable)
            continue
        safe_name = re.sub(r"[^0-9A-Za-z_.-]+", "_", str(name))[:80] or "input"
        input_path = snapshot_dir / f"input_{safe_name}.xlsx"
        wb.SaveCopyAs(str(input_path))
        files[f"input:{name}"] = str(input_path)
    PIPELINE_STEP_SNAPSHOTS[key] = {
        "key": key,
        "stepIdx": step_idx,
        "created": time.time(),
        "files": files,
    }
    _cleanup_pipeline_step_snapshots()
    return PIPELINE_STEP_SNAPSHOTS[key]


def _warn_excel_nonfatal(stage, err):
    print(f"[excel pipeline] non-fatal {stage} failed: {err!r}", file=sys.stderr)


def _safe_activate_excel_app(app):
    try:
        app.Visible = True
    except Exception:
        pass
    for attr in ("ActiveWindow", "Windows"):
        try:
            target = getattr(app, attr)
            if attr == "Windows":
                target = target(1)
            target.Activate()
            return True
        except Exception:
            pass
    try:
        app.Activate()
        return True
    except Exception:
        return False


def _hide_excel_hwnd(hwnd):
    if win32gui is None or win32con is None:
        return
    try:
        hwnd = int(hwnd)
    except Exception:
        return
    try:
        if not hwnd or not win32gui.IsWindow(hwnd):
            return
        flags = (
            getattr(win32con, "SWP_NOACTIVATE", 0x0010) |
            getattr(win32con, "SWP_NOOWNERZORDER", 0x0200) |
            getattr(win32con, "SWP_HIDEWINDOW", 0x0080)
        )
        win32gui.SetWindowPos(
            hwnd,
            getattr(win32con, "HWND_BOTTOM", 1),
            -32000,
            -32000,
            1,
            1,
            flags,
        )
        win32gui.ShowWindow(hwnd, getattr(win32con, "SW_HIDE", 0))
    except Exception:
        pass


def _hide_excel_windows_for_pid(pid):
    if win32gui is None or win32process is None:
        return
    try:
        target_pid = int(pid or 0)
    except Exception:
        return
    if not target_pid:
        return

    def visit(hwnd, _):
        try:
            _tid, window_pid = win32process.GetWindowThreadProcessId(hwnd)
            if int(window_pid or 0) == target_pid:
                _hide_excel_hwnd(hwnd)
        except Exception:
            pass
        return True

    try:
        win32gui.EnumWindows(visit, None)
    except Exception:
        pass


def _workbook_identity(wb):
    try:
        return str(Path(wb.FullName).resolve()).lower()
    except Exception:
        try:
            return str(wb.FullName).lower()
        except Exception:
            try:
                return str(wb.Name).lower()
            except Exception:
                return ""


def _hide_workbook_windows(wb):
    try:
        count = int(wb.Windows.Count)
    except Exception:
        count = 0
    for idx in range(1, count + 1):
        try:
            win = wb.Windows(idx)
        except Exception:
            continue
        try:
            win.Visible = False
        except Exception:
            pass
        try:
            _hide_excel_hwnd(int(win.Hwnd))
        except Exception:
            pass


def _hide_non_target_workbook_windows(app, target_wb):
    target_id = _workbook_identity(target_wb)
    try:
        workbooks = list(app.Workbooks)
    except Exception:
        workbooks = []
    for wb in workbooks:
        if target_id and _workbook_identity(wb) == target_id:
            continue
        _hide_workbook_windows(wb)


def _park_excel_app_offscreen(app):
    try:
        app.Visible = False
    except Exception:
        pass
    for _ in range(2):
        for attr, value in (
            ("Left", -32000),
            ("Top", -32000),
            ("Width", 10),
            ("Height", 10),
        ):
            try:
                setattr(app, attr, value)
            except Exception:
                pass
    try:
        _hide_excel_hwnd(app.Hwnd)
    except Exception:
        pass


def _hide_excel_app_window(app):
    _park_excel_app_offscreen(app)
    try:
        app.Visible = False
    except Exception:
        pass
    try:
        _hide_excel_hwnd(app.Hwnd)
    except Exception:
        pass


def _prepare_vba_macro_run_window_state(session, app, wb):
    """Put Excel into the same non-visible state that reliably allows Application.Run.

    Native frame mode parks only the HWND offscreen for visual reasons, leaving
    Application.Visible and the workbook window visible. On some Office builds that
    state refuses both direct injected macros and temp .xlsm runner macros. Macro
    execution should not depend on the UI hide path, so the server forces the
    session into a true hidden state immediately before running VBA and the normal
    finally path restores the live window afterward.
    """
    try:
        session["hidden"] = True
        session["lastNativePositionKey"] = ""
    except Exception:
        pass
    try:
        _hide_workbook_windows(wb)
    except Exception:
        pass
    try:
        _hide_excel_app_window(app)
    except Exception:
        pass
    try:
        pid = session.get("pid") or _excel_process_id(app)
        _hide_excel_windows_for_pid(pid)
    except Exception:
        pass


def _prepare_excel_session_for_close(app, wb=None):
    """Close/Quit 직전에 Excel 이 빈 회색 top-level 창을 복원하지 못하게 먼저 숨긴다."""
    try:
        app.DisplayAlerts = False
    except Exception:
        pass
    try:
        app.ScreenUpdating = False
    except Exception:
        pass
    try:
        app.Interactive = False
    except Exception:
        pass
    if wb is not None:
        try:
            _hide_workbook_windows(wb)
        except Exception:
            pass
    try:
        _hide_excel_app_window(app)
    except Exception:
        pass


def _start_excel_hide_guard(app, enabled=True):
    if not enabled or win32gui is None or win32con is None:
        return None
    try:
        hwnd = int(app.Hwnd)
    except Exception:
        return None
    if not hwnd:
        return None
    stop_event = threading.Event()

    def guard():
        while not stop_event.wait(0.02):
            _hide_excel_hwnd(hwnd)

    thread = threading.Thread(target=guard, name="b2b-excel-hide-guard", daemon=True)
    thread.start()
    return stop_event


def _safe_excel_calculate(app):
    for method_name in ("CalculateFull", "Calculate"):
        try:
            method = getattr(app, method_name)
            method()
            return True
        except Exception as err:
            last_err = err
    _warn_excel_nonfatal("calculate", last_err)
    return False


def _copy_source_workbook_into_target(app, target_wb, source_path):
    source_path = Path(source_path)
    if not source_path.exists():
        raise RuntimeError(f"snapshot source not found: {source_path}")
    open_path = source_path
    temp_copy = None
    source_temp_path = None
    try:
        # [e2e 수정] 같은 '파일명(베이스네임)'의 워크북이 이 앱에 이미 열려 있으면, 알림 억제 상태의
        # Workbooks.Open 이 예외 대신 None 을 돌려준다(경로가 달라도 이름만 같으면 발생).
        # 리셋 소스는 항상 UUID 임시사본으로 복사해 연다 — 이름 충돌을 원천 차단(복사 비용은 로컬 1회).
        temp_copy = BACKEND_DIR / f"live_reset_{uuid.uuid4().hex}{source_path.suffix or '.xlsx'}"
        shutil.copy2(source_path, temp_copy)
        open_path = temp_copy

        _park_excel_app_offscreen(app)
        source_wb, source_temp_path = excel_workbooks_open(app, open_path, read_only=True, intended_name=source_path.name)
        _park_excel_app_offscreen(app)
        _hide_excel_app_window(app)
        try:
            source_wb.Windows(1).Visible = False
        except Exception:
            pass
        try:
            app.DisplayAlerts = False
            placeholder = target_wb.Worksheets.Add(Before=target_wb.Worksheets(1))
            placeholder_name = "__B2B_PIPELINE_SWAP__"
            suffix = 1
            existing = set(_excel_names(target_wb.Worksheets))
            while placeholder_name in existing:
                suffix += 1
                placeholder_name = f"__B2B_PIPELINE_SWAP_{suffix}__"
            placeholder.Name = placeholder_name

            for ws in list(target_wb.Worksheets):
                if ws.Name != placeholder_name:
                    ws.Delete()

            for idx in range(1, source_wb.Worksheets.Count + 1):
                source_wb.Worksheets(idx).Copy(Before=placeholder)
            placeholder.Delete()
            # [원본복원 검증] 복사 후 대상 워크북이 원본의 모든 시트를 실제로 갖췄는지 확인한다.
            # 저사양 COM 등으로 일부 시트 복사가 누락되면, 조용히 깨진 상태(예: 원본 Sheet1 누락)로 남아
            # 다음 스텝이 "시트 못 찾음"으로 엉뚱하게 터진다 → 여기서 원인을 명확히 알리고 중단한다.
            try:
                _src_names = set(_excel_names(source_wb.Worksheets))
                _dst_names = set(_excel_names(target_wb.Worksheets))
            except Exception:
                _src_names = None
                _dst_names = None
            if _src_names and not _src_names.issubset(_dst_names or set()):
                raise RuntimeError(
                    "원본 복원 실패: 워크북이 원본 상태로 완전히 복구되지 않았습니다 "
                    f"(원본 시트={sorted(_src_names)}, 현재 시트={sorted(_dst_names or set())}). "
                    "이전 실행의 중간 상태가 남았을 수 있으니 전체실행으로 다시 실행해 주세요."
                )
        finally:
            source_wb.Close(SaveChanges=False)
    finally:
        if temp_copy:
            try:
                Path(temp_copy).unlink(missing_ok=True)
            except Exception:
                pass
        if source_temp_path and str(source_temp_path) != str(temp_copy):
            try:
                Path(source_temp_path).unlink(missing_ok=True)
            except Exception:
                pass


def _python_step_sig(step):
    # 라이브 미러에 이미 적용된 단계와 새 요청을 비교하기 위한 안정적 시그니처.
    # id + 정규화된 코드가 같으면 같은 단계로 본다(코드 편집 시 시그니처가 달라짐).
    code = normalize_python_pipeline_code(str((step or {}).get("code") or ""))
    # [혼합 호환] 같은 코드라도 언어(vba/python)가 다르면 다른 단계 — 시그니처에 언어 포함.
    raw = (str((step or {}).get("id") or "") + "\x00" + str((step or {}).get("language") or "") + "\x00" + code).encode("utf-8")
    return hashlib.sha1(raw).hexdigest()


def _excel_output_preview_sheets(wb):
    # 라이브 미러 워크북에서 미리보기+수식 메타데이터를 COM 으로 직접 읽는다(결과 파일 저장 없이).
    data = {}
    for name in _excel_collection_names(wb.Worksheets):
        try:
            ws = wb.Worksheets(name)
            used = ws.UsedRange
            max_row = max(0, int(used.Rows.Count))
            max_col = max(0, int(used.Columns.Count))
            rows = min(max_row, PREVIEW_ROWS)
            cols = min(max_col, PREVIEW_COLS or 256)
        except Exception:
            data[name] = {
                "rows": [],
                "formulas": {},
                "originalFormulaValues": {},
                "formats": [],
                "maxRow": 0,
                "maxCol": 0,
            }
            continue
        if rows <= 0 or cols <= 0:
            data[name] = {
                "rows": [],
                "formulas": {},
                "originalFormulaValues": {},
                "formats": [],
                "maxRow": max_row,
                "maxCol": max_col,
            }
            continue
        try:
            rng = ws.Range(ws.Cells(1, 1), ws.Cells(rows, cols))
            values = _range_matrix(rng.Value)
            formulas_matrix = _range_matrix(rng.Formula)
            formats_matrix = _range_matrix(rng.NumberFormat)
            out_rows = []
            formulas = {}
            original_formula_values = {}
            formats = []
            for r_idx in range(rows):
                value_row = values[r_idx] if r_idx < len(values) else []
                formula_row = formulas_matrix[r_idx] if r_idx < len(formulas_matrix) else []
                format_row = formats_matrix[r_idx] if r_idx < len(formats_matrix) else []
                out_row = []
                out_format_row = []
                for c_idx in range(cols):
                    value = value_row[c_idx] if c_idx < len(value_row) else ""
                    formula_value = formula_row[c_idx] if c_idx < len(formula_row) else value
                    formula_text = _com_scalar(formula_value)
                    json_value = cell_to_json(value)
                    out_row.append(json_value)
                    fmt = format_row[c_idx] if c_idx < len(format_row) else ""
                    out_format_row.append(str(fmt or ""))
                    if isinstance(formula_text, str) and formula_text.startswith("="):
                        address = f"{_col_letter(c_idx + 1)}{r_idx + 1}"
                        formulas[address] = formula_text
                        original_formula_values[address] = json_value
                out_rows.append(out_row)
                formats.append(out_format_row)
            data[name] = {
                "rows": out_rows,
                "formulas": formulas,
                "originalFormulaValues": original_formula_values,
                "formats": formats,
                "maxRow": max_row,
                "maxCol": max_col,
            }
        except Exception:
            data[name] = {
                "rows": [],
                "formulas": {},
                "originalFormulaValues": {},
                "formats": [],
                "maxRow": max_row,
                "maxCol": max_col,
            }
    return data


def _result_from_workbook_files(output_path, input_paths_by_name, output_item, output_wb_record, input_wb_records, payload, resume_from):
    # Excel 실행 없이 저장된 파일(스냅샷/원본)에서 미리보기 + 다운로드를 구성한다.
    # liveApplied=False 로 반환 → 프런트가 활성 미러를 이 파일로 교체(replace)해 빠르게 반영.
    current = payload.get("current") or {}
    output_file_id = current.get("outputFileId") or "output:0"
    download_urls = {}
    download_id = None

    out_path = Path(output_path)
    result_output = {}
    if out_path.exists():
        inspected = inspect_workbook(out_path)
        result_output = inspected.get("sheets") or {}
        download_id = uuid.uuid4().hex
        RESULTS[download_id] = {"path": str(out_path), "name": out_path.name, "created": time.time()}
        download_urls[output_file_id] = f"/api/workbooks/download/{download_id}"
        update_workbook_current_cache(output_wb_record, rows_only_sheets(result_output))

    rec_by_name = {}
    for item, rec in zip(payload.get("inputs", []), input_wb_records):
        rec_by_name[item.get("name") or rec["name"]] = rec
    input_previews = {}
    for name, path in (input_paths_by_name or {}).items():
        ip = Path(path)
        if not ip.exists():
            continue
        inspected_in = inspect_workbook(ip)
        input_previews[name] = inspected_in.get("sheets") or {}
        rid = uuid.uuid4().hex
        RESULTS[rid] = {"path": str(ip), "name": ip.name, "created": time.time()}
        download_urls["input:" + name] = f"/api/workbooks/download/{rid}"
        rec = rec_by_name.get(name)
        if rec is not None:
            update_workbook_current_cache(rec, rows_only_sheets(input_previews[name]))

    previews = build_result_previews(input_previews, result_output, current, {}, [])
    return {
        "ok": True,
        "pythonExcel": True,
        "liveApplied": False,
        "snapshotHit": True,
        "snapshotStep": resume_from,
        "diffId": None,
        "diffs": {},
        "forcedValueCells": [],
        "downloadId": download_id,
        "downloadUrl": download_urls.get(output_file_id),
        "downloadUrls": download_urls,
        "files": previews,
        "activeSheet": None,
        "activeAddress": None,
    }


def _worker_step_target_wb(step, input_wb_by_name, output_wb):
    """[혼합 호환] 워커에서 VBA/COM-bulk 스텝의 기준 워크북 결정:
    프론트가 첨부한 targetFileName → 입력 워크북 이름 매칭, 없으면 출력 워크북."""
    name = str((step or {}).get("targetFileName") or "").strip()
    if name:
        if name in input_wb_by_name:
            return input_wb_by_name[name]
        low = name.lower()
        for k, wb in input_wb_by_name.items():
            if str(k).lower() == low:
                return wb
    return output_wb


def _run_excel_python_pipeline_impl(payload, job_id=None):
    if not excel_available():
        raise RuntimeError("Microsoft Excel COM automation is not available. Excel and pywin32 are required.")
    output_item = payload.get("output") or {}
    if not output_item.get("backendWorkbookId"):
        raise RuntimeError("Python Excel skills require an output workbook.")

    input_items = payload.get("inputs", [])
    # [새로고침 즉시복원] 클라가 '원본부터 전체 적용'일 때만 보낸다(부분/이어실행이면 없음).
    state_sig = payload.get("stateSig")
    output_wb_record = get_workbook_or_raise(output_item.get("backendWorkbookId"))
    input_wb_records = [get_workbook_or_raise(item.get("backendWorkbookId")) for item in input_items]
    active_steps = [s for s in (payload.get("pipeline") or []) if not (s and s.get("enabled") is False)]
    # [혼합 호환] VBA 스텝도 같은 체인에서 실행한다(전역 순서 보존). JS(레거시)만 불가.
    python_steps = [s for s in active_steps if is_python_pipeline_step(s) or is_vba_pipeline_step(s)]
    if len(python_steps) != len(active_steps):
        raise RuntimeError("Excel 실행기는 Python/VBA 스텝만 실행할 수 있습니다(JavaScript 스텝은 다시 생성해 주세요).")
    cached_prefix = _find_best_pipeline_snapshot(input_items, input_wb_records, output_item, output_wb_record, python_steps)
    resume_from = cached_prefix[0] if cached_prefix else 0
    resume_snapshot = cached_prefix[2] if cached_prefix else None

    # 빠른 경로: 실행할 단계가 없으면(전부 캐시됨 또는 전부 OFF) Excel을 돌리지 않고
    # 저장된 스냅샷(없으면 원본) 파일로 결과를 즉시 만들어 반환한다(토글 ON/OFF 속도 개선).
    if resume_from >= len(python_steps):
        try:
            out_path = _snapshot_path(resume_snapshot, "output", "output") or output_wb_record["path"]
            in_paths = {}
            for item, rec in zip(input_items, input_wb_records):
                nm = item.get("name") or rec["name"]
                in_paths[nm] = _snapshot_path(resume_snapshot, "input", nm) or rec["path"]
            return _result_from_workbook_files(out_path, in_paths, output_item, output_wb_record, input_wb_records, payload, resume_from)
        except Exception as err:
            _warn_excel_nonfatal("snapshot fast result", err)

    # 단계별 서버 타이밍(F8 디버그 표시용)
    _t0 = time.perf_counter()
    _perf = {"resetMs": 0.0, "openMs": 0.0, "stepsMs": 0.0}

    live_excel_id = (payload.get("current") or {}).get("outputExcelId")
    live_session = None
    if live_excel_id:
        try:
            live_session = get_excel_session(live_excel_id)
        except Exception:
            live_session = None

    update_pipeline_job(job_id, {
        "stage": "Excel Python 실행 준비 중",
        "currentStep": 0,
        "completedSteps": 0,
        "totalSteps": len(python_steps),
        "stepRunning": True,
    })

    if live_session:
        try:
            app, output_wb = session_workbook(live_session)
            # 적용 중에는 미러 창을 보이게 하지 않는다(프런트가 미러를 숨기고 로딩을 표시).
            # 보이게/활성화하면 여러 Excel 창이 적용 중에 앞으로 튀어나온다. 완료 후 프런트가 재배치/표시.
        except Exception:
            live_session = None
    if not live_session:
        app = _get_python_skill_app()  # 재사용 (매번 새로 띄우지 않음)
        output_wb = None

    is_live = bool(live_session)
    # 라이브 미러가 이미 prefix 상태(=resume 지점까지 적용됨)이면 출력 리셋을 생략하고
    # 새 단계만 그 위에 바로 실행한다(빠른 추가). 편집/삽입/삭제로 prefix 가 어긋나면 리셋+재실행.
    desired_sigs = [_python_step_sig(s) for s in python_steps]
    applied_sigs = live_session.get("appliedStepSigs") if is_live else None
    fast_prefix = (
        is_live
        and applied_sigs is not None
        and len(applied_sigs) == resume_from
        and list(applied_sigs) == desired_sigs[:resume_from]
    )
    hide_guard = _start_excel_hide_guard(app, enabled=not live_session)
    live_read_only_mirror = bool(live_session and live_session.get("readOnlyMirror"))
    app.DisplayAlerts = False
    app.EnableEvents = False
    restore_screen_updating = None
    restore_calculation = None
    if live_session:
        try:
            restore_screen_updating = bool(app.ScreenUpdating)
            app.ScreenUpdating = False
        except Exception:
            restore_screen_updating = None
    else:
        # 워커(숨김) 경로 속도 개선: 화면 갱신을 끈다. (계산 모드는 워크북 오픈 후 설정)
        try:
            restore_screen_updating = bool(app.ScreenUpdating)
            app.ScreenUpdating = False
        except Exception:
            restore_screen_updating = None
    try:
        app.AskToUpdateLinks = False
    except Exception:
        pass

    opened = []
    ctx = None
    active_output_sheet = None
    active_output_address = None
    input_previews = {}        # 다중 입력 지원: 수정된 입력 파일 미리보기 {name: sheets}
    input_download_urls = {}   # {"input:<name>": downloadUrl}
    try:
        output_base_path = _snapshot_path(resume_snapshot, "output", "output") or output_wb_record["path"]
        if live_session:
            if live_read_only_mirror:
                try:
                    _protect_workbook_for_read_only_mirror(output_wb, False)
                except Exception as err:
                    _warn_excel_nonfatal("unprotect read-only mirror before reset", err)
            if not fast_prefix:
                # 미러 상태가 요청 prefix 와 다르면 기준 상태로 리셋 후 전체 재실행.
                _t_reset = time.perf_counter()
                _copy_source_workbook_into_target(app, output_wb, output_base_path)
                _perf["resetMs"] = (time.perf_counter() - _t_reset) * 1000
            if live_read_only_mirror:
                try:
                    _protect_workbook_for_read_only_mirror(output_wb, False)
                except Exception as err:
                    _warn_excel_nonfatal("unprotect read-only mirror after reset", err)
        else:
            output_wb, output_temp_path = _open_excel_workbook_for_skill(
                app,
                Path(output_base_path),
                read_only=False,
                intended_name=output_item.get("name") or output_wb_record["name"],
            )
            opened.append((output_wb, output_temp_path))

        input_wbs = {}
        input_wb_by_name = {}
        # 스냅샷 저장 시, 스킬이 수정하지 않은 입력은 안정적인 원본 경로를 그대로 참조해
        # 전체 복사(SaveCopyAs)를 건너뛴다(여러 입력 파일일 때 큰 속도 차이). {name: 원본경로}
        input_stable_src = {}
        output_path_norm = str(Path(output_wb_record["path"]).resolve()).lower()
        for item, wb_record in zip(input_items, input_wb_records):
            name = item.get("name") or wb_record["name"]
            path_norm = str(Path(wb_record["path"]).resolve()).lower()
            snapshot_input_path = _snapshot_path(resume_snapshot, "input", name)
            if snapshot_input_path:
                wb, temp_path = _open_excel_workbook_for_skill(app, Path(snapshot_input_path), read_only=False, intended_name=name)
                opened.append((wb, temp_path))
                input_wbs[name] = wb
                input_wb_by_name[name] = wb
                continue
            if path_norm == output_path_norm and not resume_snapshot:
                input_wbs[name] = output_wb
                input_wb_by_name[name] = output_wb
                continue
            wb, temp_path = _open_excel_workbook_for_skill(app, Path(wb_record["path"]), read_only=False, intended_name=name)
            opened.append((wb, temp_path))
            input_wbs[name] = wb
            input_wb_by_name[name] = wb
            input_stable_src[name] = wb_record["path"]  # 원본(업로드 파일) = 안정적, 미수정 시 참조 가능

        # 큰 출력은 중간 단계 스냅샷을 건너뛰고 마지막 단계만 저장한다(속도).
        try:
            _out_size = Path(output_wb_record["path"]).stat().st_size
        except OSError:
            _out_size = 0
        snapshot_intermediate = _out_size < SNAPSHOT_INTERMEDIATE_MAX_BYTES

        # 워크북이 모두 열린 뒤 자동 재계산을 끈다(워커 경로). 단계마다 명시적으로 계산. (finally 복구)
        if not live_session:
            try:
                restore_calculation = app.Calculation
                app.Calculation = -4135  # xlCalculationManual
            except Exception:
                restore_calculation = None

        # 워크북 열기/리셋까지 걸린 시간(리셋 제외분 = 순수 open).
        _perf["openMs"] = max(0.0, (time.perf_counter() - _t0) * 1000 - _perf["resetMs"])
        _t_steps = time.perf_counter()

        output_name = output_item.get("name") or output_wb_record["name"]
        current = payload.get("current") or {}
        ctx = ExcelSkillContext(
            app,
            output_wb,
            input_wbs,
            output_name=output_name,
            active_file_id=current.get("fileId"),
            active_sheet=current.get("sheet"),
        )
        for idx, step in enumerate(python_steps[resume_from:], start=resume_from + 1):
            raise_if_pipeline_cancelled(job_id)  # 협조적 취소(스텝 경계)
            update_pipeline_job(job_id, {
                "stage": f"Excel Python Step {idx}/{len(python_steps)} 실행 중",
                "currentStep": idx,
                "completedSteps": idx - 1,
                "stepRunning": True,
                "errorInfo": None,
            })
            original_code = str(step.get("code") or "")
            code = normalize_python_pipeline_code(original_code)
            # [혼합 호환] 스텝별 엔진 디스패치 — 같은 체인(같은 워크북 상태)에서 순서대로 실행:
            #   vba          → 숨김 워커 Excel 에 VBA 주입 실행
            #   COM-bulk     → 라이브와 동일한 벌크 ctx 엔진(_exec_python_com_skill, AST 게이트 포함)
            #   레거시 python → 기존 ExcelSkillContext(transform(ctx))
            _step_lang = str(step.get("language") or "").lower()
            namespace = _safe_python_globals()
            try:
                if _step_lang == "vba" or is_vba_pipeline_step(step):
                    stage_label = "vba"
                    _twb = _worker_step_target_wb(step, input_wb_by_name, output_wb)
                    try:
                        _twb.Activate()
                    except Exception:
                        pass
                    _inject_and_run_vba(app, _twb, original_code, VBA_SKILL_ENTRY)
                elif not python_step_uses_legacy_dialect(original_code):
                    stage_label = "com-bulk"
                    _twb = _worker_step_target_wb(step, input_wb_by_name, output_wb)
                    _exec_python_com_skill(
                        app,
                        _twb,
                        None,
                        original_code,
                        skip_static=bool(step.get("trustedStatic") is True),
                        timeout_s=_step_extended_timeout_s(step),
                    )
                else:
                    stage_label = "compile"
                    exec(compile(code, f"<pipeline_step_{idx}>", "exec"), namespace, namespace)
                    stage_label = "lookup transform"
                    transform = namespace.get("transform")
                    if not callable(transform):
                        raise RuntimeError("Python step must define def transform(ctx):")
                    stage_label = "transform"
                    transform(ctx)
            except Exception as err:
                _cause, _guide = _pipeline_error_guide(str(err), original_code)
                raise PipelineExecutionError({
                    "stepIdx": idx - 1,
                    "stepId": step.get("id"),
                    "description": step.get("description"),
                    "code": original_code,
                    "normalizedCode": code,
                    "language": step.get("language") or "python",
                    "message": f"{_cause}\n💡 이렇게 요청해 보세요: {_guide}\n(자세히: {stage_label} 단계 — {err})",
                    "cause": _cause,
                    "promptGuide": _guide,
                    "rawError": f"{stage_label}: {err}",
                    "stack": repr(err),
                })
            _safe_excel_calculate(app)
            if not live_session:
                _hide_excel_app_window(app)
            is_last_step = (idx == len(python_steps))
            if is_last_step or snapshot_intermediate:
                try:
                    snapshot_key = _pipeline_snapshot_key(
                        input_items,
                        input_wb_records,
                        output_item,
                        output_wb_record,
                        python_steps[:idx],
                    )
                    _saved = _save_pipeline_step_snapshot(snapshot_key, idx, app, output_wb, input_wb_by_name, input_stable_src)
                    # [새로고침 즉시복원] 마지막 스텝까지 끝났으면 그 사본이 곧 '최종 상태'다.
                    # 방금 저장한 파일을 그대로 가리키기만 한다 — 추가 저장/복사가 없다(VBA 경로와 대칭).
                    if is_last_step and state_sig and _saved:
                        _files = _saved.get("files") or {}
                        for _item, _rec in zip(input_items, input_wb_records):
                            _nm = _item.get("name") or _rec["name"]
                            _p = _files.get(f"input:{_nm}")
                            if _p:
                                _save_live_final_snapshot(_rec, state_sig, _p, link=True)
                        _op = _files.get("output:output")
                        if _op and output_wb_record:
                            _save_live_final_snapshot(output_wb_record, state_sig, _op, link=True)
                except Exception as err:
                    _warn_excel_nonfatal("pipeline snapshot", err)

        _perf["stepsMs"] = (time.perf_counter() - _t_steps) * 1000
        # 적용 완료 후 마지막으로 쓴 시트/셀을 강제로 Activate/Select 하지 않는다.
        # 사용자가 입력 시트나 채팅창을 보고 있던 상태를 깨면 셀 선택/포커스가 튀는 문제가 생긴다.
        # (frame 모드 철학과 동일 — 적용이 사용자의 보기 상태를 바꾸지 않는다)
        active_output_sheet = None
        active_output_address = None
        if live_session and not live_session.get("liveEditable"):
            # 라이브 세션 폴은 스냅샷 diff 를 쓰지 않으므로(liveSelectionOnly)
            # 전 시트 재스냅샷(거대 파일에서 수 초)은 비라이브 세션에만 의미가 있다.
            try:
                refresh_excel_session_snapshots(live_session, output_wb)
            except Exception as err:
                _warn_excel_nonfatal("refresh live snapshots", err)

        # 다중 입력 지원: 스킬이 읽거나 수정한 입력 파일도 미리보기/다운로드에 반영한다.
        # 다운로드 소스는 가능하면 단계 스냅샷(이미 저장됨) 사본을 재사용해 추가 COM 저장을 피한다.
        input_record_by_name = {}
        for item, wb_record in zip(input_items, input_wb_records):
            input_record_by_name[item.get("name") or wb_record["name"]] = wb_record
        final_snapshot = PIPELINE_STEP_SNAPSHOTS.get(
            _pipeline_snapshot_key(input_items, input_wb_records, output_item, output_wb_record, python_steps)
        )
        for name, wb in input_wb_by_name.items():
            try:
                input_previews[name] = _excel_output_preview_sheets(wb)
            except Exception as err:
                _warn_excel_nonfatal(f"input preview {name}", err)
                continue
            try:
                rec = input_record_by_name.get(name)
                if rec is not None:
                    update_workbook_current_cache(rec, rows_only_sheets(input_previews[name]))
                snap_in = _snapshot_path(final_snapshot, "input", name)
                if snap_in and Path(snap_in).exists():
                    src_path = snap_in  # 단계 스냅샷(이미 저장됨) 재사용 → 복사/추가 저장 없음
                else:
                    BACKEND_DIR.mkdir(parents=True, exist_ok=True)
                    safe_in = Path(str(name)).name or "input.xlsx"
                    if not Path(safe_in).suffix:
                        safe_in += ".xlsx"
                    dest = BACKEND_DIR / f"{uuid.uuid4().hex}_result_{safe_in}"
                    wb.SaveCopyAs(str(dest))
                    src_path = str(dest)
                input_result_id = uuid.uuid4().hex
                RESULTS[input_result_id] = {"path": str(src_path), "name": Path(src_path).name, "created": time.time()}
                input_download_urls["input:" + name] = f"/api/workbooks/download/{input_result_id}"
            except Exception as err:
                _warn_excel_nonfatal(f"input result save {name}", err)

        if is_live:
            # 라이브 미러 = 적용된 워크북. 적용 단계 시그니처를 기록해 다음 추가가 빠른 경로를 타게 한다.
            live_session["appliedStepSigs"] = desired_sigs
        result_path = None
        if not is_live:
            update_pipeline_job(job_id, {
                "stage": "Excel 결과 저장 중",
                "currentStep": len(python_steps),
                "completedSteps": len(python_steps),
                "stepRunning": False,
            })
            BACKEND_DIR.mkdir(parents=True, exist_ok=True)
            original_name = output_item.get("name") or output_wb_record["name"]
            safe_name = Path(str(original_name)).name
            if not Path(safe_name).suffix:
                safe_name += ".xlsx"
            result_path = BACKEND_DIR / f"{uuid.uuid4().hex}_result_{safe_name}"
            _t_save = time.perf_counter()
            output_wb.SaveCopyAs(str(result_path))
            _perf["saveResultMs"] = (time.perf_counter() - _t_save) * 1000
    finally:
        if live_session and live_read_only_mirror:
            try:
                _protect_workbook_for_read_only_mirror(output_wb, True)
                _configure_excel_grid_window(app, output_wb)
            except Exception as err:
                _warn_excel_nonfatal("restore read-only mirror", err)
        for wb, temp_path in reversed(opened):
            try:
                wb.Close(SaveChanges=False)
            except Exception:
                pass
            if temp_path:
                try:
                    Path(temp_path).unlink(missing_ok=True)
                except Exception:
                    pass
        if not live_session:
            # 워커 설정 복구(다음 재사용 대비) 후 숨긴 채로 유지. (열린 워크북은 위 opened 루프에서 모두 닫힘)
            if restore_calculation is not None:
                try:
                    app.Calculation = restore_calculation
                except Exception:
                    pass
            if restore_screen_updating is not None:
                try:
                    app.ScreenUpdating = restore_screen_updating
                except Exception:
                    pass
            try:
                _hide_excel_app_window(app)
            except Exception:
                pass
        elif restore_screen_updating is not None:
            try:
                app.ScreenUpdating = restore_screen_updating
            except Exception:
                pass
        if hide_guard:
            hide_guard.set()

    output_file_id = (payload.get("current") or {}).get("outputFileId") or "output:0"
    current = payload.get("current") or {}
    if is_live:
        # 결과 파일을 저장하지 않는다(다운로드 시점에 라이브 미러를 저장). 미리보기는 COM 으로 읽음.
        try:
            result_output = _excel_output_preview_sheets(output_wb)
        except Exception as err:
            _warn_excel_nonfatal("live output preview", err)
            result_output = {}
        update_workbook_current_cache(output_wb_record, rows_only_sheets(result_output))
        previews = build_result_previews(input_previews, result_output, current, {}, [])
        _perf["totalServerMs"] = (time.perf_counter() - _t0) * 1000
        _perf["finalizeMs"] = max(0.0, _perf["totalServerMs"] - _perf["openMs"] - _perf["resetMs"] - _perf["stepsMs"])
        _perf["mode"] = "live"
        return {
            "ok": True,
            "pythonExcel": True,
            "liveApplied": True,
            "snapshotHit": bool(resume_from),
            "snapshotStep": resume_from,
            "diffId": None,
            "diffs": {},
            "forcedValueCells": [],
            "downloadId": None,
            "downloadUrl": None,
            "downloadUrls": dict(input_download_urls),
            "files": previews,
            "activeSheet": active_output_sheet,
            "activeAddress": active_output_address,
            "debugTimings": _perf,
        }
    result_id = uuid.uuid4().hex
    RESULTS[result_id] = {
        "path": str(result_path),
        "name": result_path.name,
        "created": time.time(),
    }
    _t_inspect = time.perf_counter()
    inspected = inspect_workbook(result_path)
    _perf["inspectMs"] = (time.perf_counter() - _t_inspect) * 1000
    result_output = inspected.get("sheets") or {}
    update_workbook_current_cache(output_wb_record, rows_only_sheets(result_output))
    previews = build_result_previews(input_previews, result_output, current, {}, [])
    download_urls = dict(input_download_urls)
    download_urls[output_file_id] = f"/api/workbooks/download/{result_id}"
    _perf["totalServerMs"] = (time.perf_counter() - _t0) * 1000
    _perf["finalizeMs"] = max(0.0, _perf["totalServerMs"] - _perf["openMs"] - _perf["resetMs"] - _perf["stepsMs"])
    _perf["mode"] = "worker-hidden"
    return {
        "ok": True,
        "pythonExcel": True,
        "snapshotHit": bool(resume_from),
        "snapshotStep": resume_from,
        "debugTimings": _perf,
        "diffId": None,
        "diffs": {},
        "forcedValueCells": [],
        "downloadId": result_id,
        "downloadUrl": f"/api/workbooks/download/{result_id}",
        "downloadUrls": download_urls,
        "files": previews,
        "activeSheet": active_output_sheet,
        "activeAddress": active_output_address,
    }


def run_excel_python_pipeline_payload(payload, job_id=None):
    return excel_call(_run_excel_python_pipeline_impl, payload, job_id, timeout=600)


def run_backend_pipeline_payload(payload, job_id=None):
    update_pipeline_job(job_id, {"stage": "입력 파일 읽는 중", "currentStep": 0})
    inputs = {}
    for idx, item in enumerate(payload.get("inputs", []), start=1):
        wb = get_workbook_or_raise(item.get("backendWorkbookId"))
        update_pipeline_job(job_id, {"stage": f"입력 파일 읽는 중 ({idx}/{len(payload.get('inputs', []))})"})
        inputs[item.get("name") or wb["name"]] = get_workbook_aoa_for_run(wb)

    update_pipeline_job(job_id, {"stage": "출력 템플릿 읽는 중"})
    output_item = payload.get("output") or {}
    output_wb = get_workbook_or_raise(output_item.get("backendWorkbookId")) if output_item.get("backendWorkbookId") else None
    output = get_workbook_aoa_for_run(output_wb) if output_wb else {}

    total_steps = len([s for s in payload.get("pipeline", []) if not (s and s.get("enabled") is False)])
    update_pipeline_job(job_id, {"stage": "스킬 실행 중", "currentStep": 0, "totalSteps": total_steps})
    result = run_js_pipeline_with_node({
        "inputs": inputs,
        "output": output,
        "pipeline": payload.get("pipeline", []),
    }, job_id=job_id)
    result_inputs = result.get("inputs") or inputs
    result_output = result.get("output") or output
    forced_value_cells = result.get("forcedValueCells") or []
    current = payload.get("current") or {}

    update_pipeline_job(job_id, {
        "stage": "diff 계산 중",
        "currentStep": total_steps,
        "completedSteps": total_steps,
        "totalSteps": total_steps,
        "stepRunning": False,
    })
    diffs = build_pipeline_diffs(inputs, output, result_inputs, result_output, current)
    diff_id = uuid.uuid4().hex
    DIFFS[diff_id] = {
        "id": diff_id,
        "created": time.time(),
        "diffs": diffs,
        "current": current,
    }

    update_pipeline_job(job_id, {"stage": "다운로드 준비 중", "currentStep": total_steps})
    download_urls = {}
    input_items = payload.get("inputs", [])
    for item in input_items:
        wb = get_workbook_or_raise(item.get("backendWorkbookId"))
        input_name = item.get("name") or wb["name"]
        if input_name not in result_inputs:
            continue
        input_download_id = uuid.uuid4().hex
        RESULTS[input_download_id] = {
            "template_path": str(wb["path"]),
            "sheets": result_inputs[input_name],
            "forced_value_cells": [cell for cell in forced_value_cells if cell.get("fileId") == "input:" + input_name],
            "name": f"result_{wb['name']}",
            "created": time.time(),
        }
        download_urls["input:" + input_name] = f"/api/workbooks/download/{input_download_id}"

    download_id = None
    if output_wb:
        output_file_id = (payload.get("current") or {}).get("outputFileId") or "output:0"
        download_id = uuid.uuid4().hex
        RESULTS[download_id] = {
            "template_path": str(output_wb["path"]),
            "sheets": result_output,
            "forced_value_cells": [cell for cell in forced_value_cells if cell.get("fileId") == output_file_id],
            "name": f"result_{output_wb['name']}",
            "created": time.time(),
        }
        download_urls[output_file_id] = f"/api/workbooks/download/{download_id}"

    update_pipeline_job(job_id, {"stage": "미리보기 생성 중"})
    previews = build_result_previews(result_inputs, result_output, current, diffs, forced_value_cells)
    return {
        "ok": True,
        "diffId": diff_id,
        "diffs": diffs,
        "forcedValueCells": forced_value_cells,
        "downloadId": download_id,
        "downloadUrl": f"/api/workbooks/download/{download_id}" if download_id else None,
        "downloadUrls": download_urls,
        "files": previews,
    }


def inspect_workbook(path):
    if is_csv_path(path):
        return inspect_csv_workbook(path)
    try:
        wb = openpyxl_load_workbook_compatible(path, read_only=True, data_only=False)
        # [값 로드 미루기 0.7.2.1 / 2026-08-06] 예전엔 여기서 값용(data_only=True)까지 곧바로 열었다.
        #   openpyxl 은 '수식 글자'와 '계산된 값' 중 하나만 줄 수 있어(열 때 정하는 옵션) 둘 다
        #   필요하면 같은 파일을 두 번 열어야 한다. 그런데 값용은 아래에서 **수식 셀에만** 쓰인다.
        #   청구내역처럼 수식이 하나도 없는 파일에서는 그 로드가 통째로 낭비였다
        #   (실측: 47MB 파일 업로드 13.5초 중 6.6초 = 49%).
        #   그래서 '건너뛰기'가 아니라 '미루기'로 바꾼다 — 수식 셀을 처음 만나는 순간 연다.
        #   수식이 있는 파일은 결국 열게 되므로 결과·동작이 예전과 완전히 같다.
        cached_wb = None
    except Exception as err:
        if excel_available():
            # [실제 시트명 보존] 위장 파일(확장자 .xlsx, 내용 OLE/HTML)은 openpyxl 이 못 읽어 이 분기로 온다.
            # COM 검사가 일시 실패(업로드 순간 Excel 바쁨/spawn 실패)하면 폴백이 시트명을 지어내
            # '<32hex>_파일명' 이 매핑 UI 에 그대로 떴다(고객 화면 실측 — 실제 시트명은 'Sheet1').
            # 일시 실패가 대부분이므로 짧게 1회 재시도해 진짜 시트명을 최대한 확보한다.
            #
            # [CoInitialize 누락 수정] 업로드/재검사는 HTTP 워커 스레드에서 실행되는데 그 스레드는
            # CoInitialize 를 부른 적이 없다 → DispatchEx 가 항상 "CoInitialize가 호출되지 않았습니다"
            # (-2147221008)로 죽어, '일시 실패'가 아니라 위장 파일이면 100% 폴백으로 떨어졌다.
            # 그래서 매핑 UI 에 실제 시트명 대신 파일명이 뜨고, 그걸 채택한 스킬이 step1 부터
            # "시트를 찾을 수 없음"으로 실패했다(한전 위장 xlsx 실측). COM 은 CoInitialize 를 보유한
            # 전용 워커(b2b-excel-com)에서만 돌려야 한다 → excel_call 로 마샬링한다.
            # (워커 스레드 자신이 호출한 경우엔 큐에 넣으면 자기 자신을 기다려 데드락 → 직접 실행.)
            def _inspect_via_com():
                in_worker = EXCEL_THREAD is not None and threading.current_thread() is EXCEL_THREAD
                if in_worker:
                    return inspect_workbook_with_excel(path, source_error=err)
                return excel_call(inspect_workbook_with_excel, path, source_error=err, timeout=180)

            last_excel_err = None
            for _attempt in range(2):
                try:
                    return _inspect_via_com()
                except Exception as excel_err:
                    last_excel_err = excel_err
                    time.sleep(1.2)
            return inspect_workbook_fallback(path, f"{err}; excel: {last_excel_err}")
        return inspect_workbook_fallback(path, err)
    try:
        sheets = {}
        # 값용 워크북은 '수식 셀을 처음 만났을 때' 딱 한 번 연다. 그 뒤로는 재사용한다.
        # read_only 시트는 한 번만 훑을 수 있으므로, 시트별 미리보기 격자를 통째로 떠서 들고 있는다
        # (예전 코드가 next(cached_rows) 로 한 줄씩 맞춰 읽던 것과 같은 범위·같은 순서).
        cached_grids = {}
        _cached_load_ms = [0.0]          # 값용 로드에 실제로 쓴 시간(0 이면 안 열었다는 뜻)

        def _cached_grid(sheet_title):
            nonlocal cached_wb
            if sheet_title in cached_grids:
                return cached_grids[sheet_title]
            if cached_wb is None:
                _t0 = time.perf_counter()
                cached_wb = openpyxl_load_workbook_compatible(path, read_only=True, data_only=True)
                _cached_load_ms[0] = (time.perf_counter() - _t0) * 1000
            if sheet_title in cached_wb.sheetnames:
                cws = cached_wb[sheet_title]
                grid = [list(r) for r in cws.iter_rows(max_row=PREVIEW_ROWS, max_col=PREVIEW_COLS)]
            else:
                grid = []          # 값용에 그 시트가 없으면 예전처럼 '빈 줄' 취급
            cached_grids[sheet_title] = grid
            return grid

        for ws in wb.worksheets:
            rows = []
            formulas = {}
            original_formula_values = {}
            formats = []
            for row_idx, row in enumerate(ws.iter_rows(max_row=PREVIEW_ROWS, max_col=PREVIEW_COLS), start=1):
                values = []
                format_row = []
                for cell_idx, cell in enumerate(row):
                    if cell.data_type == "f":
                        grid = _cached_grid(ws.title)                      # ← 여기서 처음으로 값용 로드
                        cached_row = grid[row_idx - 1] if row_idx - 1 < len(grid) else []
                        cached_value = cached_row[cell_idx].value if cell_idx < len(cached_row) else None
                        json_cached = cell_to_json(cached_value)
                        values.append(json_cached if json_cached is not None else "")
                        formulas[cell.coordinate] = cell.value if str(cell.value).startswith("=") else "=" + str(cell.value)
                        if json_cached is not None:
                            original_formula_values[cell.coordinate] = json_cached
                    else:
                        values.append(cell_to_json(cell.value))
                    format_row.append(cell.number_format if cell.number_format else "")
                rows.append(values)
                formats.append(format_row)
            sheets[ws.title] = {
                "rows": rows,
                "formulas": formulas,
                "originalFormulaValues": original_formula_values,
                "formats": formats,
                "maxRow": ws.max_row or len(rows),
                "maxCol": ws.max_column or (max((len(r) for r in rows), default=0)),
            }
        # [계측] 값용 로드를 실제로 했는지 / 얼마나 걸렸는지 — '미루기'가 현장에서 얼마나
        # 이득인지 로그만으로 알 수 있게 남긴다(수식 없는 파일이면 cachedLoadMs=0).
        try:
            _vba_trace(
                "inspect.openpyxl",
                name=Path(str(path)).name,
                sheets=len(sheets),
                formulaCells=sum(len((s or {}).get("formulas") or {}) for s in sheets.values()),
                cachedLoaded=cached_wb is not None,
                cachedLoadMs=round(_cached_load_ms[0]),
            )
        except Exception:
            pass
        return {"sheetNames": wb.sheetnames, "sheets": sheets}
    finally:
        wb.close()
        if cached_wb is not None:
            cached_wb.close()


def inspect_workbook_fallback(path, err=None):
    # [실제 시트명 보존] 끝내 검사가 안 되면 시트명을 알 수 없다 — 저장 파일명 stem 을 그대로 쓰면
    # 내부 해시 접두('<32hex>_원본명')가 사용자에게 노출된다. csv_sheet_name 이 그 접두를 벗겨 주므로
    # 최소한 사용자가 알아보는 이름(원본 파일명)으로 표기한다(requiresExcel=True 라 이후 실제 스키마로 대체됨).
    sheet_name = csv_sheet_name(path) or "Sheet1"
    return {
        "sheetNames": [sheet_name],
        "sheets": {
            sheet_name: {
                "rows": [],
                "formulas": {},
                "originalFormulaValues": {},
                "formats": [],
                "maxRow": 0,
                "maxCol": 0,
                "inspectError": str(err) if err else "",
            }
        },
        "inspectError": str(err) if err else "",
        "requiresExcel": True,
    }


def inspect_workbook_with_excel(path, source_error=None):
    if not excel_available():
        return inspect_workbook_fallback(path, source_error)
    app = win32com.client.DispatchEx("Excel.Application")
    _track_spawned_excel_app(app)  # [0.5.2 이식] 고아 Excel 추적
    app.Visible = False
    app.DisplayAlerts = False
    app.EnableEvents = False
    try:
        app.AskToUpdateLinks = False
    except Exception:
        pass
    wb = None
    temp_path = None
    try:
        wb, temp_path = excel_workbooks_open(app, path, read_only=True)
        sheets = {}
        sheet_names = _excel_collection_names(wb.Worksheets)
        for sheet_idx in range(1, int(wb.Worksheets.Count) + 1):
            ws = wb.Worksheets(sheet_idx)
            sheet_name = str(ws.Name)
            try:
                used = ws.UsedRange
                used_rows = int(used.Rows.Count)
                used_cols = int(used.Columns.Count)
            except Exception:
                used_rows = 0
                used_cols = 0
            max_row = max(0, used_rows)
            max_col = max(0, used_cols)
            preview_rows = min(max_row, PREVIEW_ROWS)
            preview_cols = min(max_col, PREVIEW_COLS or 256)
            rows = []
            formulas = {}
            original_formula_values = {}
            formats = []
            if preview_rows and preview_cols:
                rng = ws.Range(ws.Cells(1, 1), ws.Cells(preview_rows, preview_cols))
                values = _range_matrix(rng.Value)
                for r_idx in range(preview_rows):
                    row_values = values[r_idx] if r_idx < len(values) else []
                    out_row = []
                    format_row = []
                    for c_idx in range(preview_cols):
                        value = row_values[c_idx] if c_idx < len(row_values) else ""
                        out_row.append(cell_to_json(value))
                        try:
                            cell = ws.Cells(r_idx + 1, c_idx + 1)
                            format_row.append(str(cell.NumberFormat or ""))
                            if bool(cell.HasFormula):
                                address = str(cell.Address(False, False))
                                formula = str(cell.Formula)
                                formulas[address] = formula if formula.startswith("=") else "=" + formula
                                original_formula_values[address] = cell_to_json(value)
                        except Exception:
                            format_row.append("")
                    rows.append(out_row)
                    formats.append(format_row)
            sheets[sheet_name] = {
                "rows": rows,
                "formulas": formulas,
                "originalFormulaValues": original_formula_values,
                "formats": formats,
                "maxRow": max_row,
                "maxCol": max_col,
            }
        return {
            "sheetNames": sheet_names,
            "sheets": sheets,
            "excelInspected": True,
            "sourceInspectError": str(source_error) if source_error else "",
        }
    finally:
        try:
            if wb is not None:
                wb.Close(SaveChanges=False)
        except Exception:
            pass
        try:
            app.Quit()
        except Exception:
            pass
        if temp_path:
            try:
                Path(temp_path).unlink(missing_ok=True)
            except Exception:
                pass


def load_workbook_aoa(path):
    if is_csv_path(path):
        return load_csv_aoa(path)
    try:
        wb = openpyxl_load_workbook_compatible(path, read_only=True, data_only=False)
    except Exception:
        if excel_available():
            return load_workbook_aoa_with_excel(path)
        raise
    try:
        data = {}
        for ws in wb.worksheets:
            rows = []
            for row in ws.iter_rows():
                values = ["" if cell.data_type == "f" else cell_to_json(cell.value) for cell in row]
                while values and values[-1] in ("", None):
                    values.pop()
                rows.append(values)
            data[ws.title] = rows
        return data
    finally:
        wb.close()


def load_workbook_aoa_with_excel(path):
    app = win32com.client.DispatchEx("Excel.Application")
    _track_spawned_excel_app(app)  # [0.5.2 이식] 고아 Excel 추적
    app.Visible = False
    app.DisplayAlerts = False
    app.EnableEvents = False
    wb = None
    temp_path = None
    try:
        wb, temp_path = excel_workbooks_open(app, path, read_only=True)
        data = {}
        for sheet_idx in range(1, int(wb.Worksheets.Count) + 1):
            ws = wb.Worksheets(sheet_idx)
            try:
                used = ws.UsedRange
                rows = int(used.Rows.Count)
                cols = int(used.Columns.Count)
            except Exception:
                rows = 0
                cols = 0
            if rows <= 0 or cols <= 0:
                data[str(ws.Name)] = []
                continue
            values = _range_matrix(ws.Range(ws.Cells(1, 1), ws.Cells(rows, cols)).Value)
            out = []
            for row in values:
                row_values = [cell_to_json(value) for value in (row or [])]
                while row_values and row_values[-1] in ("", None):
                    row_values.pop()
                out.append(row_values)
            data[str(ws.Name)] = out
        return data
    finally:
        try:
            if wb is not None:
                wb.Close(SaveChanges=False)
        except Exception:
            pass
        try:
            app.Quit()
        except Exception:
            pass
        if temp_path:
            try:
                Path(temp_path).unlink(missing_ok=True)
            except Exception:
                pass


def is_csv_path(path):
    return Path(path).suffix.lower() in (".csv", ".tsv")


def csv_sheet_name(path):
    stem = Path(path).stem or "Sheet1"
    if len(stem) > 33 and stem[32] == "_":
        prefix = stem[:32]
        if all(ch in "0123456789abcdefABCDEF" for ch in prefix):
            stem = stem[33:] or stem
    return stem


def read_csv_text(path):
    raw = Path(path).read_bytes()
    for enc in ("utf-8-sig", "cp949", "euc-kr", "utf-8"):
        try:
            return raw.decode(enc)
        except UnicodeDecodeError:
            continue
    return raw.decode("utf-8", errors="replace")


def read_csv_rows(path, max_rows=None):
    text = read_csv_text(path)
    sample = text[:4096]
    delimiter = "\t" if Path(path).suffix.lower() == ".tsv" else ","
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",\t;|")
    except csv.Error:
        dialect = csv.excel_tab if delimiter == "\t" else csv.excel
    rows = []
    for idx, row in enumerate(csv.reader(text.splitlines(), dialect)):
        if max_rows is not None and idx >= max_rows:
            break
        rows.append(["" if value is None else value for value in row])
    return rows


def inspect_csv_workbook(path):
    rows = read_csv_rows(path, PREVIEW_ROWS)
    sheet = csv_sheet_name(path)
    max_col = max((len(row or []) for row in rows), default=0)
    total_rows = 0
    try:
        with Path(path).open("rb") as f:
            for _ in f:
                total_rows += 1
    except OSError:
        total_rows = len(rows)
    return {
        "sheetNames": [sheet],
        "sheets": {
            sheet: {
                "rows": rows,
                "formulas": {},
                "originalFormulaValues": {},
                "formats": [],
                "maxRow": total_rows or len(rows),
                "maxCol": max_col,
            }
        },
    }


def load_csv_aoa(path):
    return {csv_sheet_name(path): read_csv_rows(path)}


def write_result_csv(result_path, sheets):
    sheet_name = next(iter((sheets or {}).keys()), csv_sheet_name(result_path))
    rows = (sheets or {}).get(sheet_name) or []
    with Path(result_path).open("w", encoding="utf-8-sig", newline="") as f:
        writer = csv.writer(f)
        for row in rows:
            writer.writerow(["" if value is None else value for value in (row or [])])


def get_workbook_aoa_for_run(wb_record, base_mode="original"):
    with WORKBOOK_CACHE_LOCK:
        if base_mode == "current" and wb_record.get("current_aoa_cache") is not None:
            wb_record["aoa_cache_hits"] = int(wb_record.get("aoa_cache_hits") or 0) + 1
            return wb_record["current_aoa_cache"]
        cached = wb_record.get("aoa_cache")
        if cached is not None:
            wb_record["aoa_cache_hits"] = int(wb_record.get("aoa_cache_hits") or 0) + 1
            return cached

    if base_mode == "current" and wb_record.get("node_current"):
        sheets = export_node_worker_workbook(wb_record["id"])
        with WORKBOOK_CACHE_LOCK:
            wb_record["current_aoa_cache"] = sheets
            wb_record["aoa_cache_hits"] = int(wb_record.get("aoa_cache_hits") or 0) + 1
        return sheets

    loaded = load_workbook_aoa(Path(wb_record["path"]))
    with WORKBOOK_CACHE_LOCK:
        if wb_record.get("aoa_cache") is None:
            wb_record["aoa_cache"] = loaded
            wb_record["aoa_cache_created"] = time.time()
        else:
            loaded = wb_record["aoa_cache"]
        wb_record["aoa_cache_hits"] = int(wb_record.get("aoa_cache_hits") or 0) + 1
        return loaded


def update_workbook_current_cache(wb_record, sheets):
    if not wb_record or sheets is None:
        return
    with WORKBOOK_CACHE_LOCK:
        wb_record["current_aoa_cache"] = sheets
        wb_record["current_aoa_cache_created"] = time.time()


def _sheet_payload_rows(sheet_payload):
    if isinstance(sheet_payload, dict) and "rows" in sheet_payload:
        return sheet_payload.get("rows") or []
    return sheet_payload or []


def rows_only_sheets(sheets):
    return {
        name: _sheet_payload_rows(sheet_payload)
        for name, sheet_payload in (sheets or {}).items()
    }


def sheet_formula_maps(sheets, key):
    out = {}
    for name, sheet_payload in (sheets or {}).items():
        if isinstance(sheet_payload, dict) and key in sheet_payload:
            out[name] = sheet_payload.get(key) or {}
        else:
            out[name] = {}
    return out


def sheet_format_maps(sheets):
    out = {}
    for name, sheet_payload in (sheets or {}).items():
        if isinstance(sheet_payload, dict) and "formats" in sheet_payload:
            out[name] = sheet_payload.get("formats") or []
        else:
            out[name] = []
    return out


def write_result_workbook(template_path, result_path, sheets, forced_value_cells=None):
    if is_csv_path(template_path):
        write_result_csv(result_path, sheets)
        return
    forced_cells = {
        (str(cell.get("sheetName") or ""), int(cell.get("r") or 0) + 1, int(cell.get("c") or 0) + 1)
        for cell in (forced_value_cells or [])
        if isinstance(cell, dict) and cell.get("sheetName")
    }
    wb = openpyxl_load_workbook_compatible(template_path)
    try:
        for sheet_name, rows in (sheets or {}).items():
            ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb.create_sheet(sheet_name)
            for r_idx, row in enumerate(rows or [], start=1):
                for c_idx, value in enumerate(row or [], start=1):
                    cell = ws.cell(row=r_idx, column=c_idx)
                    force_value = (sheet_name, r_idx, c_idx) in forced_cells
                    if cell.data_type == "f" and (value == "" or value is None) and not force_value:
                        continue
                    cell.value = value
        wb.save(result_path)
    finally:
        wb.close()


def cell_to_json(value):
    if value is None:
        return ""
    if hasattr(value, "isoformat"):
        return value.isoformat()
    return value


def recover_workbook_record(workbook_id):
    # 서버가 재시작되면 메모리 WORKBOOKS 가 비지만 업로드 파일(BACKEND_DIR/{id}_{name})은 디스크에 남는다.
    # 프런트의 기존 workbookId 로 요청이 오면 디스크에서 찾아 재등록해 "workbook not found" 를 막는다.
    if not workbook_id:
        return None
    rec = WORKBOOKS.get(workbook_id)
    if rec:
        return rec
    try:
        matches = sorted(BACKEND_DIR.glob(f"{workbook_id}_*"))
    except Exception:
        matches = []
    for path in matches:
        try:
            if not path.is_file():
                continue
        except Exception:
            continue
        name = path.name[len(workbook_id) + 1:] or path.name
        rec = {
            "id": workbook_id,
            "name": name,
            "path": str(path),
            "created": time.time(),
            "aoa_cache": None,
            "current_aoa_cache": None,
            "aoa_cache_created": None,
            "aoa_cache_hits": 0,
        }
        WORKBOOKS[workbook_id] = rec
        return rec
    return None


def get_workbook_or_raise(workbook_id):
    wb = recover_workbook_record(workbook_id)
    if not wb:
        raise ValueError(f"backend workbook not found: {workbook_id}")
    return wb


def run_js_pipeline_with_node(payload, job_id=None):
    progress_path = None
    if job_id:
        progress_file = tempfile.NamedTemporaryFile(prefix=f"b2b_pipeline_{job_id}_", suffix=".json", delete=False)
        progress_path = progress_file.name
        progress_file.close()
        payload = dict(payload)
        payload["progressPath"] = progress_path
    runner = r"""
const fs = require("fs");
const payload = JSON.parse(fs.readFileSync(0, "utf8"));
const progressPath = payload.progressPath || "";
function writeProgress(info) {
  if (!progressPath) return;
  try {
    fs.writeFileSync(progressPath, JSON.stringify(info), "utf8");
  } catch (err) {}
}
const inputs = payload.inputs || {};
const output = payload.output || {};
function normalizeText(v) { return String(v ?? "").trim().toLowerCase().replace(/\s+/g, ""); }
function includesNormalizedText(v, s) { return normalizeText(v).includes(normalizeText(s)); }
function equalsNormalizedText(v, s) { return normalizeText(v) === normalizeText(s); }
function replaceNormalizedText(v) { return String(v ?? ""); }
function similarity(a, b) { a = normalizeText(a); b = normalizeText(b); if (!a || !b) return 0; return a === b ? 1 : (a.includes(b) || b.includes(a) ? 0.8 : 0); }
function headerRowIndex(sheetAoA) {
  let best = 0, bestScore = -1;
  for (let r = 0; r < Math.min((sheetAoA || []).length, 30); r++) {
    const row = sheetAoA[r] || [];
    const score = row.filter(v => String(v ?? "").trim()).length;
    if (score > bestScore) { best = r; bestScore = score; }
  }
  return best;
}
function dataStartRowIndex(sheetAoA) { return headerRowIndex(sheetAoA) + 1; }
function excelRowToIndex(n) { return Math.max(0, Number(n) - 1); }
function col(sheetAoA, name) {
  const h = headerRowIndex(sheetAoA);
  const row = sheetAoA[h] || [];
  const target = normalizeText(name);
  let fallback = -1;
  for (let i = 0; i < row.length; i++) {
    const cur = normalizeText(row[i]);
    if (cur === target) return i;
    if (fallback < 0 && cur && (cur.includes(target) || target.includes(cur))) fallback = i;
  }
  return fallback;
}
function findColumnGlobal(inputsMap, name) {
  const hits = [];
  Object.entries(inputsMap || {}).forEach(([file, sheets]) => {
    Object.entries(sheets || {}).forEach(([sheet, aoa]) => {
      const colIdx = col(aoa, name);
      if (colIdx >= 0) hits.push({ file, sheet, colIdx });
    });
  });
  return hits;
}
function fuzzyGetKey(target, prop) {
  if (!target || typeof target !== "object") return null;
  if (Object.prototype.hasOwnProperty.call(target, prop)) return prop;
  const keys = Object.keys(target);
  if (keys.length === 1) return keys[0];
  const wanted = normalizeText(prop);
  let best = null;
  for (const key of keys) {
    const cur = normalizeText(key);
    if (cur === wanted || (cur && wanted && (cur.includes(wanted) || wanted.includes(cur)))) {
      best = key;
      break;
    }
  }
  return best;
}
function fuzzyProxy(target) {
  if (!target || typeof target !== "object") return target;
  return new Proxy(target, {
    get(t, prop) {
      if (typeof prop === "symbol" || prop in t) return t[prop];
      const key = fuzzyGetKey(t, String(prop));
      return key ? t[key] : undefined;
    },
    set(t, prop, value) {
      if (typeof prop === "symbol" || Object.prototype.hasOwnProperty.call(t, prop)) {
        t[prop] = value;
        return true;
      }
      const key = fuzzyGetKey(t, String(prop));
      if (key && normalizeText(key) === normalizeText(prop)) t[key] = value;
      else t[prop] = value;
      return true;
    },
  });
}
let activeForcedValueCells = {};
let activeClearedValueCells = {};
let activeSheetProxyCache = new WeakMap();
let activeRowProxyCache = new WeakMap();
const activeOutputFileId = (payload.current && payload.current.outputFileId) || "output:0";
function forcedCellKey(fileId, sheetName, r, c) { return `${fileId}\u0000${sheetName}\u0000${r}\u0000${c}`; }
function addForcedValueCell(fileId, sheetName, r, c, value) {
  if (!fileId || !sheetName) return;
  activeForcedValueCells[forcedCellKey(fileId, sheetName, r, c)] = { fileId, sheetName, r, c, value };
}
function trackClearThenSet(fileId, sheetName, r, c, value) {
  if (!fileId || !sheetName) return;
  const key = forcedCellKey(fileId, sheetName, r, c);
  if (value === "") {
    activeClearedValueCells[key] = true;
    return;
  }
  if (activeClearedValueCells[key]) {
    delete activeClearedValueCells[key];
    addForcedValueCell(fileId, sheetName, r, c, value);
  }
}
function trackedRowProxy(row, fileId, sheetName, r) {
  if (!row || typeof row !== "object") return row;
  const key = `${fileId}\u0000${sheetName}\u0000${r}`;
  let cached = activeRowProxyCache.get(row);
  if (cached && cached[key]) return cached[key];
  if (!cached) { cached = {}; activeRowProxyCache.set(row, cached); }
  cached[key] = new Proxy(row, {
    set(target, prop, value) {
      target[prop] = value;
      const c = Number(prop);
      if (Number.isInteger(c) && c >= 0) trackClearThenSet(fileId, sheetName, r, c, value);
      return true;
    },
  });
  return cached[key];
}
function trackedSheetProxy(sheet, fileId, sheetName) {
  if (!sheet || typeof sheet !== "object") return sheet;
  const key = `${fileId}\u0000${sheetName}`;
  let cached = activeSheetProxyCache.get(sheet);
  if (cached && cached[key]) return cached[key];
  if (!cached) { cached = {}; activeSheetProxyCache.set(sheet, cached); }
  cached[key] = new Proxy(sheet, {
    get(target, prop) {
      const value = target[prop];
      const r = Number(prop);
      if (Number.isInteger(r) && r >= 0 && Array.isArray(value)) return trackedRowProxy(value, fileId, sheetName, r);
      return value;
    },
    set(target, prop, value) { target[prop] = value; return true; },
  });
  return cached[key];
}
function trackedSheetsProxy(sheets, fileId) {
  if (!sheets || typeof sheets !== "object") return sheets;
  return new Proxy(sheets, {
    get(target, prop) {
      if (typeof prop === "symbol") return target[prop];
      const key = Object.prototype.hasOwnProperty.call(target, prop) ? prop : fuzzyGetKey(target, String(prop));
      return key ? trackedSheetProxy(target[key], fileId, String(key)) : undefined;
    },
    set(target, prop, value) {
      if (typeof prop === "symbol" || Object.prototype.hasOwnProperty.call(target, prop)) target[prop] = value;
      else {
        const key = fuzzyGetKey(target, String(prop));
        if (key && normalizeText(key) === normalizeText(prop)) target[key] = value;
        else target[prop] = value;
      }
      return true;
    },
  });
}
const wrappedInputs = {};
Object.entries(inputs).forEach(([fileName, sheets]) => { wrappedInputs[fileName] = trackedSheetsProxy(sheets, `input:${fileName}`); });
const proxiedInputs = fuzzyProxy(wrappedInputs);
const proxiedOutput = trackedSheetsProxy(output, activeOutputFileId);
function findInputBySheet(inputsMap, sheetName, options) {
  options = options || {};
  const target = normalizeText(sheetName);
  const preferredFile = options.preferredFile ? normalizeText(options.preferredFile) : "";
  const matches = [];
  Object.entries(inputsMap || {}).forEach(([fileName, sheets]) => {
    Object.entries(sheets || {}).forEach(([sn, sheet]) => {
      if (normalizeText(sn) === target) matches.push({ fileName, file: sheets, sheetName: sn, sheet });
    });
  });
  if (!matches.length) return null;
  if (preferredFile) {
    const preferred = matches.find(item => normalizeText(item.fileName).includes(preferredFile) || preferredFile.includes(normalizeText(item.fileName)));
    if (preferred) return preferred;
  }
  return matches[0];
}
function resolveTargetSheets(fileRef) {
  if (fileRef === "output") return output;
  let key = String(fileRef || "");
  if (key.startsWith("input:")) key = key.slice(6);
  if (Object.prototype.hasOwnProperty.call(inputs, key)) return inputs[key];
  const fuzzyKey = fuzzyGetKey(inputs, key);
  return fuzzyKey ? inputs[fuzzyKey] : null;
}
function insertColumns(fileRef, sheetName, atColIdx, count) {
  const file = resolveTargetSheets(fileRef);
  if (!file) throw new Error(`insertColumns: file not found: ${fileRef}`);
  const sheet = file[sheetName];
  if (!sheet) throw new Error(`insertColumns: sheet not found: ${sheetName}`);
  const at = Math.max(0, Number(atColIdx) || 0);
  const n = Math.max(0, Number(count) || 0);
  if (!n) return;
  for (let r = 0; r < sheet.length; r++) {
    if (!sheet[r]) sheet[r] = [];
    while (sheet[r].length < at) sheet[r].push("");
    sheet[r].splice(at, 0, ...new Array(n).fill(""));
  }
}
function copyColumns(fileRef, sheetName, srcStart, srcCount, destStart) {
  const file = resolveTargetSheets(fileRef);
  if (!file) throw new Error(`copyColumns: file not found: ${fileRef}`);
  const sheet = file[sheetName];
  if (!sheet) throw new Error(`copyColumns: sheet not found: ${sheetName}`);
  const src = Math.max(0, Number(srcStart) || 0);
  const n = Math.max(0, Number(srcCount) || 0);
  const dest = Math.max(0, Number(destStart) || 0);
  if (!n) return;
  for (let r = 0; r < sheet.length; r++) {
    if (!sheet[r]) sheet[r] = [];
    const values = [];
    for (let c = 0; c < n; c++) values.push(sheet[r][src + c] !== undefined ? sheet[r][src + c] : "");
    while (sheet[r].length < dest) sheet[r].push("");
    for (let c = 0; c < n; c++) sheet[r][dest + c] = values[c];
  }
}
function deleteColumns(fileRef, sheetName, atColIdx, count) {
  const file = resolveTargetSheets(fileRef);
  if (!file) throw new Error(`deleteColumns: file not found: ${fileRef}`);
  const sheet = file[sheetName];
  if (!sheet) throw new Error(`deleteColumns: sheet not found: ${sheetName}`);
  const at = Math.max(0, Number(atColIdx) || 0);
  const n = Math.max(0, Number(count) || 0);
  if (!n) return;
  for (let r = 0; r < sheet.length; r++) {
    if (sheet[r]) sheet[r].splice(at, n);
  }
}
function shiftFormulaText(v) { return v; }
function fileIdForSetCellTarget(fileRef) {
  if (fileRef === "output") return activeOutputFileId;
  let key = String(fileRef || "");
  if (key.startsWith("input:")) return key;
  if (Object.prototype.hasOwnProperty.call(inputs, key)) return `input:${key}`;
  return "";
}
function setCellValue(fileRef, sheetName, r, c, value) {
  const file = resolveTargetSheets(fileRef);
  if (!file) throw new Error(`setCellValue: file not found: ${fileRef}`);
  if (!file[sheetName]) file[sheetName] = [];
  const rowIdx = Math.max(0, Number(r) || 0);
  const colIdx = Math.max(0, Number(c) || 0);
  if (!file[sheetName][rowIdx]) file[sheetName][rowIdx] = [];
  file[sheetName][rowIdx][colIdx] = value;
  addForcedValueCell(fileIdForSetCellTarget(fileRef), sheetName, rowIdx, colIdx, value);
  return value;
}
let activeStepIndex = 0;
const totalSteps = (payload.pipeline || []).filter(step => !(step && step.enabled === false)).length;
for (const step of payload.pipeline || []) {
  if (step && step.enabled === false) continue;
  writeProgress({
    stage: "스킬 실행 중",
    currentStep: activeStepIndex,
    completedSteps: activeStepIndex,
    totalSteps,
    stepRunning: true,
    stepDescription: (step && step.description) || `Step ${activeStepIndex + 1}`
  });
  try {
    const code = String((step && step.code) || "");
    const fn = new Function("inputs", "output", "col", "findColumnGlobal", "findInputBySheet", "similarity", "normalizeText", "replaceNormalizedText", "includesNormalizedText", "equalsNormalizedText", "headerRowIndex", "dataStartRowIndex", "excelRowToIndex", "insertColumns", "copyColumns", "deleteColumns", "shiftFormulaText", "setCellValue",
      code + "\nreturn typeof transform === 'function' ? transform(inputs, output) : { inputs, output };");
    const result = fn(proxiedInputs, proxiedOutput, col, findColumnGlobal, findInputBySheet, similarity, normalizeText, replaceNormalizedText, includesNormalizedText, equalsNormalizedText, headerRowIndex, dataStartRowIndex, excelRowToIndex, insertColumns, copyColumns, deleteColumns, shiftFormulaText, setCellValue);
    if (result && typeof result === "object" && !Array.isArray(result)) {
      if (result.inputs && result.inputs !== proxiedInputs && typeof result.inputs === "object") Object.assign(inputs, result.inputs);
      if (result.output && result.output !== proxiedOutput && typeof result.output === "object") Object.assign(output, result.output);
    }
  } catch (err) {
    const info = {
      stepIdx: activeStepIndex,
      stepId: step && step.id || null,
      description: step && step.description || "",
      code: step && step.code || "",
      message: err && err.message || String(err),
      stack: err && err.stack || "",
    };
    writeProgress({
      stage: "오류",
      currentStep: activeStepIndex,
      completedSteps: activeStepIndex,
      totalSteps,
      stepRunning: false,
      stepDescription: info.description || `Step ${activeStepIndex + 1}`,
      errorInfo: info,
      error: info.message,
    });
    process.stderr.write(JSON.stringify({ errorInfo: info, error: info.message }));
    process.exit(1);
  }
  activeStepIndex += 1;
  writeProgress({
    stage: "스킬 실행 중",
    currentStep: activeStepIndex,
    completedSteps: activeStepIndex,
    totalSteps,
    stepRunning: false,
    stepDescription: (step && step.description) || `Step ${activeStepIndex}`
  });
}
process.stdout.write(JSON.stringify({ inputs, output, forcedValueCells: Object.values(activeForcedValueCells) }));
"""
    stdout_file = tempfile.NamedTemporaryFile(prefix="b2b_pipeline_stdout_", suffix=".json", delete=False)
    stderr_file = tempfile.NamedTemporaryFile(prefix="b2b_pipeline_stderr_", suffix=".txt", delete=False)
    stdout_path = stdout_file.name
    stderr_path = stderr_file.name
    stdout_file.close()
    stderr_file.close()
    stdout_handle = open(stdout_path, "w+", encoding="utf-8")
    stderr_handle = open(stderr_path, "w+", encoding="utf-8")
    node_path = node_executable()
    if not node_path:
        raise RuntimeError("node runtime is not available")
    proc = subprocess.Popen(
        [node_path, "-e", runner],
        stdin=subprocess.PIPE,
        stdout=stdout_handle,
        stderr=stderr_handle,
        text=True,
        encoding="utf-8",
        **hidden_subprocess_kwargs(),
    )
    stdout = ""
    stderr = ""
    try:
        try:
            proc.stdin.write(json.dumps(payload, ensure_ascii=False))
            proc.stdin.close()
        except Exception:
            pass
        last_progress = None
        started = time.time()
        while proc.poll() is None:
            if time.time() - started > 300:
                proc.kill()
                raise TimeoutError("node pipeline timed out")
            if progress_path and os.path.exists(progress_path):
                try:
                    progress_mtime = os.path.getmtime(progress_path)
                    if progress_mtime != last_progress:
                        last_progress = progress_mtime
                        with open(progress_path, "r", encoding="utf-8") as f:
                            progress = json.load(f)
                        update_pipeline_job(job_id, progress)
                except Exception:
                    pass
            time.sleep(0.2)
        stdout_handle.seek(0)
        stderr_handle.seek(0)
        stdout = stdout_handle.read()
        stderr = stderr_handle.read()
    finally:
        try:
            stdout_handle.close()
            stderr_handle.close()
        except Exception:
            pass
        for temp_path in (stdout_path, stderr_path):
            try:
                os.unlink(temp_path)
            except OSError:
                pass
        if progress_path:
            try:
                os.unlink(progress_path)
            except OSError:
                pass
    if proc.returncode != 0:
        message = (stderr or "").strip() or "node pipeline failed"
        try:
            parsed_error = json.loads(message)
            info = parsed_error.get("errorInfo") or {}
            raise PipelineExecutionError(parsed_error.get("error") or info.get("message") or "node pipeline failed", info)
        except PipelineExecutionError:
            raise
        except Exception:
            pass
        raise RuntimeError(message)
    return json.loads(stdout)


def sheet_dimensions(sheets):
    dimensions = {}
    for name, sheet_payload in (sheets or {}).items():
        rows = _sheet_payload_rows(sheet_payload)
        if isinstance(sheet_payload, dict) and "rows" in sheet_payload:
            max_row = int(sheet_payload.get("maxRow") or len(rows or []))
            max_col = int(sheet_payload.get("maxCol") or max((len(row or []) for row in (rows or [])), default=0))
            preview_rows = min(len(rows or []), PREVIEW_ROWS)
            preview_cols = max((len(row or []) for row in (rows or [])[:PREVIEW_ROWS]), default=0)
            dimensions[name] = {
                "maxRow": max_row,
                "maxCol": max_col,
                "previewRows": preview_rows,
                "previewCols": preview_cols,
            }
            continue
        dimensions[name] = {
            "maxRow": len(rows or []),
            "maxCol": max((len(row or []) for row in (rows or [])), default=0),
            "previewRows": min(len(rows or []), PREVIEW_ROWS),
            "previewCols": max((len(row or []) for row in (rows or [])[:PREVIEW_ROWS]), default=0),
        }
    return dimensions


def build_result_previews(inputs, output, current, diffs=None, forced_value_cells=None):
    diffs = diffs or {}
    forced_value_cells = forced_value_cells or []
    files = []
    for name, sheets in (inputs or {}).items():
        file_id = "input:" + name
        files.append({
            "fileId": file_id,
            "name": name,
            "sheetNames": list((sheets or {}).keys()),
            "sheets": preview_sheets(sheets),
            "forcedValueCells": [cell for cell in forced_value_cells if cell.get("fileId") == file_id],
            "formulas": sheet_formula_maps(sheets, "formulas"),
            "originalFormulaValues": sheet_formula_maps(sheets, "originalFormulaValues"),
            "formats": sheet_format_maps(sheets),
            "dimensions": sheet_dimensions(sheets),
            "diff": diffs.get(file_id),
        })
    if output:
        output_file_id = current.get("outputFileId") or "output:0"
        files.append({
            "fileId": output_file_id,
            "name": "output",
            "sheetNames": list((output or {}).keys()),
            "sheets": preview_sheets(output),
            "forcedValueCells": [cell for cell in forced_value_cells if cell.get("fileId") == output_file_id],
            "formulas": sheet_formula_maps(output, "formulas"),
            "originalFormulaValues": sheet_formula_maps(output, "originalFormulaValues"),
            "formats": sheet_format_maps(output),
            "dimensions": sheet_dimensions(output),
            "diff": diffs.get(output_file_id),
        })
    return files


def preview_sheets(sheets):
    def preview_row(row):
        values = list(row or [])
        return values if PREVIEW_COLS is None else values[:PREVIEW_COLS]
    return {
        name: [preview_row(row) for row in _sheet_payload_rows(sheet_payload)[:PREVIEW_ROWS]]
        for name, sheet_payload in (sheets or {}).items()
    }


def diff_value(value):
    if value is None:
        return ""
    if hasattr(value, "isoformat"):
        return value.isoformat()
    return value


def compute_sheet_diff(before_rows, after_rows, max_cells=MAX_DIFF_CELLS_PER_SHEET):
    before_rows = before_rows or []
    after_rows = after_rows or []
    max_rows = max(len(before_rows), len(after_rows))
    cells = []
    changed_count = 0
    truncated = False
    for r_idx in range(max_rows):
        before_row = before_rows[r_idx] if r_idx < len(before_rows) and before_rows[r_idx] else []
        after_row = after_rows[r_idx] if r_idx < len(after_rows) and after_rows[r_idx] else []
        if before_row == after_row:
            continue
        max_cols = max(len(before_row), len(after_row))
        for c_idx in range(max_cols):
            before_value = diff_value(before_row[c_idx] if c_idx < len(before_row) else "")
            after_value = diff_value(after_row[c_idx] if c_idx < len(after_row) else "")
            if before_value == after_value:
                continue
            changed_count += 1
            if len(cells) < max_cells:
                cells.append({"r": r_idx, "c": c_idx, "value": after_value})
            else:
                truncated = True
                return {"cells": cells, "changedCount": changed_count, "truncated": truncated}
    return {"cells": cells, "changedCount": changed_count, "truncated": truncated}


def compute_workbook_diff(before_sheets, after_sheets):
    before_sheets = before_sheets or {}
    after_sheets = after_sheets or {}
    sheet_diffs = {}
    total_changed = 0
    truncated = False
    for sheet_name in sorted(set(before_sheets.keys()) | set(after_sheets.keys())):
        diff = compute_sheet_diff(before_sheets.get(sheet_name), after_sheets.get(sheet_name))
        if diff["changedCount"] or sheet_name not in before_sheets or sheet_name not in after_sheets:
            sheet_diffs[sheet_name] = diff
            total_changed += diff["changedCount"]
            truncated = truncated or diff["truncated"]
    return {"sheets": sheet_diffs, "changedCount": total_changed, "truncated": truncated}


def build_pipeline_diffs(before_inputs, before_output, after_inputs, after_output, current):
    diffs = {}
    for name in sorted(set((before_inputs or {}).keys()) | set((after_inputs or {}).keys())):
        file_id = "input:" + name
        diffs[file_id] = compute_workbook_diff((before_inputs or {}).get(name), (after_inputs or {}).get(name))
    output_file_id = (current or {}).get("outputFileId") or "output:0"
    if before_output or after_output:
        diffs[output_file_id] = compute_workbook_diff(before_output, after_output)
    return diffs


def ensure_node_worker():
    global NODE_WORKER, NODE_WORKER_SCRIPT_MTIME
    worker_path = app_base_dir() / "scripts" / "backend-pipeline-worker.js"
    worker_mtime = worker_path.stat().st_mtime if worker_path.exists() else None
    if NODE_WORKER and NODE_WORKER.poll() is None:
        if NODE_WORKER_SCRIPT_MTIME == worker_mtime:
            return NODE_WORKER
        try:
            NODE_WORKER.kill()
        except Exception:
            pass
        NODE_WORKER = None
        NODE_WORKER_READY.clear()
    worker_path = app_base_dir() / "scripts" / "backend-pipeline-worker.js"
    if not worker_path.exists():
        raise RuntimeError(f"backend worker not found: {worker_path}")
    node_path = node_executable()
    if not node_path:
        raise RuntimeError("node runtime is not available")
    NODE_WORKER = subprocess.Popen(
        [node_path, str(worker_path)],
        stdin=subprocess.PIPE,
        stdout=subprocess.PIPE,
        stderr=subprocess.DEVNULL,
        text=True,
        encoding="utf-8",
        bufsize=1,
        **hidden_subprocess_kwargs(),
    )
    NODE_WORKER_SCRIPT_MTIME = worker_mtime
    NODE_WORKER_READY.clear()
    return NODE_WORKER


def node_worker_command(command):
    command = dict(command)
    command["id"] = command.get("id") or uuid.uuid4().hex
    with NODE_WORKER_LOCK:
        worker = ensure_node_worker()
        line = json.dumps(command, ensure_ascii=False, separators=(",", ":")) + "\n"
        try:
            worker.stdin.write(line)
            worker.stdin.flush()
            response_line = worker.stdout.readline()
        except Exception:
            try:
                worker.kill()
            except Exception:
                pass
            raise
        if not response_line:
            raise RuntimeError("backend worker stopped")
        response = json.loads(response_line)
        if not response.get("ok"):
            info = response.get("errorInfo") or {}
            if info:
                raise PipelineExecutionError(response.get("error") or info.get("message") or "backend worker failed", info)
            raise RuntimeError(response.get("error") or "backend worker failed")
        return response


def ensure_worker_workbook(wb_record):
    if not wb_record:
        return
    workbook_id = wb_record["id"]
    if workbook_id in NODE_WORKER_READY:
        return
    sheets = get_workbook_aoa_for_run(wb_record, "original")
    current_sheets = wb_record.get("current_aoa_cache")
    node_worker_command({
        "type": "initWorkbook",
        "workbookId": workbook_id,
        "sheets": sheets,
        "currentSheets": current_sheets,
    })
    NODE_WORKER_READY.add(workbook_id)


def export_node_worker_workbook(workbook_id):
    response = node_worker_command({
        "type": "exportWorkbook",
        "payload": {"workbookId": workbook_id},
    })
    return response.get("sheets") or {}


def run_backend_pipeline_payload_with_worker(payload, job_id=None):
    debug_started = time.perf_counter()
    timings = {}
    if os.environ.get("B2B_DISABLE_NODE_WORKER") == "1":
        raise RuntimeError("node worker disabled")
    input_items = payload.get("inputs", [])
    output_item = payload.get("output") or {}
    input_wbs = []
    stage_started = time.perf_counter()
    for idx, item in enumerate(input_items, start=1):
        wb = get_workbook_or_raise(item.get("backendWorkbookId"))
        update_pipeline_job(job_id, {"stage": f"Node 캐시 준비 중 ({idx}/{len(input_items)})"})
        ensure_worker_workbook(wb)
        input_wbs.append((item, wb))
    timings["prepareInputsMs"] = round((time.perf_counter() - stage_started) * 1000, 2)

    output_wb = None
    stage_started = time.perf_counter()
    if output_item.get("backendWorkbookId"):
        update_pipeline_job(job_id, {"stage": "Node 출력 캐시 준비 중"})
        output_wb = get_workbook_or_raise(output_item.get("backendWorkbookId"))
        ensure_worker_workbook(output_wb)
    timings["prepareOutputMs"] = round((time.perf_counter() - stage_started) * 1000, 2)

    active_steps = [s for s in payload.get("pipeline", []) if not (s and s.get("enabled") is False)]
    total_steps = len(active_steps)
    update_pipeline_job(job_id, {
        "stage": "Node 캐시에서 스킬 실행 중",
        "currentStep": 0,
        "completedSteps": 0,
        "totalSteps": total_steps,
        "stepRunning": True,
    })

    worker_payload = {
        "inputs": [
            {
                "name": item.get("name") or wb["name"],
                "backendWorkbookId": wb["id"],
            }
            for item, wb in input_wbs
        ],
        "output": {
            "name": output_item.get("name") or output_wb["name"],
            "backendWorkbookId": output_wb["id"],
        } if output_wb else None,
        "pipeline": payload.get("pipeline", []),
        "baseMode": payload.get("baseMode") or "original",
        "current": payload.get("current") or {},
    }
    stage_started = time.perf_counter()
    response = node_worker_command({
        "type": "runPipeline",
        "payload": worker_payload,
    })
    timings["workerRunAndPreviewMs"] = round((time.perf_counter() - stage_started) * 1000, 2)
    timings["workerCacheHit"] = bool(response.get("cacheHit"))

    update_pipeline_job(job_id, {
        "stage": "미리보기 반영 중",
        "currentStep": total_steps,
        "completedSteps": total_steps,
        "totalSteps": total_steps,
        "stepRunning": False,
    })

    files = response.get("files") or []
    diffs = response.get("diffs") or {}
    forced_value_cells = response.get("forcedValueCells") or []
    if not forced_value_cells:
        for file_result in files:
            forced_value_cells.extend(file_result.get("forcedValueCells") or [])
    stage_started = time.perf_counter()
    diff_id = uuid.uuid4().hex
    DIFFS[diff_id] = {
        "id": diff_id,
        "created": time.time(),
        "diffs": diffs,
        "current": payload.get("current") or {},
    }
    timings["storeDiffMs"] = round((time.perf_counter() - stage_started) * 1000, 2)

    stage_started = time.perf_counter()
    download_urls = {}
    input_wb_by_name = {item.get("name") or wb["name"]: wb for item, wb in input_wbs}
    for file_result in files:
        file_id = file_result.get("fileId")
        worker_workbook_id = file_result.get("workerWorkbookId")
        if not file_id or not worker_workbook_id:
            continue
        if file_id.startswith("input:"):
            input_name = file_id[6:]
            wb = input_wb_by_name.get(input_name) or get_workbook_or_raise(worker_workbook_id)
            wb["node_current"] = True
            result_id = uuid.uuid4().hex
            RESULTS[result_id] = {
                "template_path": str(wb["path"]),
                "workerWorkbookId": worker_workbook_id,
                "forced_value_cells": file_result.get("forcedValueCells") or [],
                "name": f"result_{wb['name']}",
                "created": time.time(),
            }
            download_urls[file_id] = f"/api/workbooks/download/{result_id}"
        elif output_wb:
            output_wb["node_current"] = True
            result_id = uuid.uuid4().hex
            RESULTS[result_id] = {
                "template_path": str(output_wb["path"]),
                "workerWorkbookId": worker_workbook_id,
                "forced_value_cells": file_result.get("forcedValueCells") or [],
                "name": f"result_{output_wb['name']}",
                "created": time.time(),
            }
            download_urls[file_id] = f"/api/workbooks/download/{result_id}"
    timings["downloadRegistrationMs"] = round((time.perf_counter() - stage_started) * 1000, 2)
    timings["totalServerMs"] = round((time.perf_counter() - debug_started) * 1000, 2)

    return {
        "ok": True,
        "worker": True,
        "cacheHit": bool(response.get("cacheHit")),
        "debugTimings": timings,
        "diffId": diff_id,
        "diffs": diffs,
        "downloadId": None,
        "downloadUrl": None,
        "downloadUrls": download_urls,
        "forcedValueCells": forced_value_cells,
        "files": files,
    }


# openpyxl 저장/읽기로 손상·오작동할 수 있는 요소의 zip 내부 경로.
_OPXL_UNSAFE_OBJECT_LABELS = {
    "xl/charts/": "차트",
    "xl/drawings/": "도형/이미지",
    "xl/media/": "이미지",
    "xl/pivotTables/": "피벗테이블",
    "xl/slicers/": "슬라이서",
    "xl/timelines/": "타임라인",
}


def _xlsx_object_reason(path):
    """openpyxl 로 다시 저장하면 유실되는 '객체'가 있으면 사유, 없으면 빈 문자열.
    (차트/이미지/피벗/슬라이서/타임라인/매크로) — 출력 저장 시에만 문제가 된다.
    수식은 트리거가 아니다: 입력은 계산값으로 읽고, 출력은 수식을 보존(Excel 이 재계산)하기 때문."""
    p = Path(path)
    try:
        if not zipfile.is_zipfile(p):
            return ""
        with zipfile.ZipFile(p) as z:
            for n in z.namelist():
                if n == "xl/vbaProject.bin":
                    return "매크로(VBA)"
                for prefix, label in _OPXL_UNSAFE_OBJECT_LABELS.items():
                    if n.startswith(prefix):
                        return label
    except Exception:
        return ""
    return ""


def _xlsx_has_merged_cells(path):
    p = Path(path)
    try:
        if not zipfile.is_zipfile(p):
            return False
        with zipfile.ZipFile(p) as z:
            for n in z.namelist():
                if not (n.startswith("xl/worksheets/") and n.endswith(".xml")):
                    continue
                try:
                    data = z.read(n, 2000000)
                except TypeError:
                    data = z.read(n)
                if b"<mergeCell" in data or b"<mergeCells" in data:
                    return True
    except Exception:
            return False
    return False


def _xlsx_has_formulas(path):
    p = Path(path)
    try:
        if not zipfile.is_zipfile(p):
            return False
        with zipfile.ZipFile(p) as z:
            for n in z.namelist():
                if not (n.startswith("xl/worksheets/") and n.endswith(".xml")):
                    continue
                try:
                    data = z.read(n, 2000000)
                except TypeError:
                    data = z.read(n)
                if b"<f" in data:
                    return True
    except Exception:
        return False
    return False


def _python_step_requests_excel_com(step):
    code = normalize_python_pipeline_code(str((step or {}).get("code") or ""))
    if re.search(r"B2B_(?:ENGINE_FALLBACK|FORCE_ENGINE)\s*:\s*excel[-_ ]?com", code, re.I):
        return "명시적 Excel COM 요청"
    return ""


def _python_step_has_structural_or_format_operation(step):
    code = normalize_python_pipeline_code(str((step or {}).get("code") or ""))
    patterns_ci = [
        r"\binsert_cols\s*\(",
        r"\binsert_rows\s*\(",
        r"\bdelete_cols\s*\(",
        r"\bdelete_rows\s*\(",
        r"\bmove_range\s*\(",
        r"\bmerge_cells\s*\(",
        r"\bunmerge_cells\s*\(",
        r"\bcopy_worksheet\s*\(",
        r"\.Copy\s*\(",
        r"\.PasteSpecial\s*\(",
        r"\.EntireColumn\b",
        r"\.EntireRow\b",
        r"\.Insert\b",
        r"\.Delete\b",
    ]
    if any(re.search(pat, code, re.I) for pat in patterns_ci):
        return True
    # COM-only properties must stay case-sensitive so ctx.rows(...) is not mistaken for ws.Rows(...).
    patterns_case_sensitive = [
        r"\.Columns\s*\(",
        r"\.Rows\s*\(",
    ]
    return any(re.search(pat, code) for pat in patterns_case_sensitive)


def _python_step_has_values_only_formula_copy_risk(step):
    code = normalize_python_pipeline_code(str((step or {}).get("code") or ""))
    text = "\n".join([
        str((step or {}).get("description") or ""),
        str((step or {}).get("prompt") or ""),
        code,
    ])
    if not re.search(r"(값만|보이는\s*값|계산(?:된)?\s*값|수식\s*(?:말고|빼고|제외|없이)|values?\s*only|paste\s*values?)", text, re.I):
        return False
    # ctx.value/display_value/display_rows 는 openpyxl 에서도 수식 표시값을 계산/조회하는 안전 경로.
    if re.search(r"\bctx\.(?:value|display_value|display_rows)\s*\(", code):
        return False
    if re.search(r"\.value\b", code, re.I) and (re.search(r"\.value\s*=", code, re.I) or re.search(r"\bctx\.(?:write_grid|set_range)\s*\(", code)):
        return True
    if re.search(r"\bctx\.rows\s*\(", code) and re.search(r"\bctx\.(?:write_grid|set_range)\s*\(", code):
        return True
    return False


def _pipeline_payload_needs_com(payload):
    """openpyxl 엔진이 안전하지 않으면 사유 문자열을 반환(없으면 "").
    - 출력에 객체(차트/이미지/피벗/매크로)가 있으면 저장 시 유실 → COM.
    - 출력/입력 중 CSV 가 있으면 openpyxl 로 못 여므로 → COM.
    - 병합셀이 있는 출력에서 구조 변경/서식 복붙은 Excel 방식 보정이 필요 → COM.
    - 수식은 트리거 아님(입력=계산값 읽기, 출력=수식 보존+Excel 재계산)."""
    active_steps = [s for s in (payload.get("pipeline") or []) if not (s and s.get("enabled") is False)]
    python_steps = [s for s in active_steps if is_python_pipeline_step(s)]
    # [혼합 호환] VBA 스텝 또는 라이브 COM 전용(ctx 벌크) 스텝은 openpyxl 로 실행 불가 — Excel 엔진으로.
    if any(is_vba_pipeline_step(s) for s in active_steps):
        return "VBA 스텝 포함"
    if any(is_python_pipeline_step(s) and not python_step_uses_legacy_dialect(str(s.get("code") or "")) for s in active_steps):
        return "라이브 COM 전용(ctx 벌크) 스텝 포함"
    for step in python_steps:
        reason = _python_step_requests_excel_com(step)
        if reason:
            return reason
    out = payload.get("output") or {}
    out_wid = out.get("backendWorkbookId")
    output_has_merged_cells = False
    output_has_formulas = False
    if out_wid:
        try:
            rec = get_workbook_or_raise(out_wid)
            if is_csv_path(rec["path"]):
                return f"{out.get('name') or rec.get('name') or '출력'}: CSV"
            reason = _xlsx_object_reason(rec["path"])
            if reason:
                return f"{out.get('name') or rec.get('name') or '출력'}: {reason}"
            output_has_merged_cells = _xlsx_has_merged_cells(rec["path"])
            output_has_formulas = _xlsx_has_formulas(rec["path"])
        except Exception:
            pass
    if output_has_formulas and any(_python_step_has_values_only_formula_copy_risk(s) for s in python_steps):
        return "수식 셀 값만 복사: Excel 계산값 필요"
    if output_has_merged_cells and any(_python_step_has_structural_or_format_operation(s) for s in python_steps):
        return "병합셀 포함 파일의 구조 변경/서식 복사"
    for it in (payload.get("inputs") or []):
        wid = it.get("backendWorkbookId")
        if not wid:
            continue
        try:
            rec = get_workbook_or_raise(wid)
            if is_csv_path(rec["path"]):
                return f"{it.get('name') or rec.get('name') or '입력'}: CSV"
        except Exception:
            continue
    return ""


def run_backend_pipeline_payload(payload, job_id=None):
    # [혼합 호환] VBA 스텝이 섞인 파이프라인도 Excel 워커 경로에서 실행한다(노드 워커 불가).
    if pipeline_has_python(payload) or _pipeline_payload_has_vba(payload):
        # 엔진 선택: 기본 "python"(openpyxl, COM 없이 인프로세스 — 빠름) / 보조 "excel"(COM Python).
        engine = str(payload.get("engine") or "python").lower()
        if engine in ("python", "openpyxl") and openpyxl is not None:
            # 안전장치: 차트/이미지/피벗/매크로/CSV/병합셀 구조변경 등이 있으면 객체 유실·계산오류를 막기 위해
            # 이 실행만 자동으로 Excel(COM) 엔진으로 전환한다.
            com_reason = _pipeline_payload_needs_com(payload)
            if com_reason:
                update_pipeline_job(job_id, {"stage": f"호환성 보호: Excel 엔진으로 전환 ({com_reason})"})
                res = run_excel_python_pipeline_payload(payload, job_id=job_id)
                if isinstance(res, dict):
                    res["engineFallback"] = "excel"
                    res["engineFallbackReason"] = com_reason
                return res
            return run_openpyxl_python_pipeline_payload(payload, job_id=job_id)
        return run_excel_python_pipeline_payload(payload, job_id=job_id)
    try:
        return run_backend_pipeline_payload_with_worker(payload, job_id=job_id)
    except PipelineExecutionError:
        raise
    except Exception as worker_err:
        NODE_WORKER_READY.clear()
        update_pipeline_job(job_id, {"stage": f"Node worker fallback: {worker_err}"})

    input_items = payload.get("inputs", [])
    base_mode = payload.get("baseMode") or "original"
    update_pipeline_job(job_id, {
        "stage": "입력 파일 읽는 중",
        "currentStep": 0,
        "completedSteps": 0,
        "stepRunning": False,
    })
    inputs = {}
    for idx, item in enumerate(input_items, start=1):
        wb = get_workbook_or_raise(item.get("backendWorkbookId"))
        update_pipeline_job(job_id, {"stage": f"입력 파일 읽는 중 ({idx}/{len(input_items)})"})
        inputs[item.get("name") or wb["name"]] = get_workbook_aoa_for_run(wb, base_mode)

    update_pipeline_job(job_id, {"stage": "출력 템플릿 읽는 중"})
    output_item = payload.get("output") or {}
    output_wb = get_workbook_or_raise(output_item.get("backendWorkbookId")) if output_item.get("backendWorkbookId") else None
    output = get_workbook_aoa_for_run(output_wb, base_mode) if output_wb else {}

    active_steps = [s for s in payload.get("pipeline", []) if not (s and s.get("enabled") is False)]
    total_steps = len(active_steps)
    update_pipeline_job(job_id, {
        "stage": "스킬 실행 중",
        "currentStep": 0,
        "completedSteps": 0,
        "totalSteps": total_steps,
        "stepRunning": False,
    })
    result = run_js_pipeline_with_node({
        "inputs": inputs,
        "output": output,
        "pipeline": payload.get("pipeline", []),
    }, job_id=job_id)
    result_inputs = result.get("inputs") or inputs
    result_output = result.get("output") or output
    forced_value_cells = result.get("forcedValueCells") or []
    current = payload.get("current") or {}

    update_pipeline_job(job_id, {
        "stage": "diff 계산 중",
        "currentStep": total_steps,
        "completedSteps": total_steps,
        "totalSteps": total_steps,
        "stepRunning": False,
    })
    diffs = build_pipeline_diffs(inputs, output, result_inputs, result_output, current)
    diff_id = uuid.uuid4().hex
    DIFFS[diff_id] = {
        "id": diff_id,
        "created": time.time(),
        "diffs": diffs,
        "current": current,
    }

    update_pipeline_job(job_id, {
        "stage": "다운로드 준비 중",
        "currentStep": total_steps,
        "completedSteps": total_steps,
        "totalSteps": total_steps,
        "stepRunning": False,
    })
    download_urls = {}
    for item in input_items:
        wb = get_workbook_or_raise(item.get("backendWorkbookId"))
        input_name = item.get("name") or wb["name"]
        if input_name not in result_inputs:
            continue
        update_workbook_current_cache(wb, result_inputs[input_name])
        input_download_id = uuid.uuid4().hex
        RESULTS[input_download_id] = {
            "template_path": str(wb["path"]),
            "sheets": result_inputs[input_name],
            "forced_value_cells": [cell for cell in forced_value_cells if cell.get("fileId") == "input:" + input_name],
            "name": f"result_{wb['name']}",
            "created": time.time(),
        }
        download_urls["input:" + input_name] = f"/api/workbooks/download/{input_download_id}"

    download_id = None
    if output_wb:
        output_file_id = (payload.get("current") or {}).get("outputFileId") or "output:0"
        update_workbook_current_cache(output_wb, result_output)
        download_id = uuid.uuid4().hex
        RESULTS[download_id] = {
            "template_path": str(output_wb["path"]),
            "sheets": result_output,
            "forced_value_cells": [cell for cell in forced_value_cells if cell.get("fileId") == output_file_id],
            "name": f"result_{output_wb['name']}",
            "created": time.time(),
        }
        download_urls[output_file_id] = f"/api/workbooks/download/{download_id}"

    update_pipeline_job(job_id, {
        "stage": "미리보기 생성 중",
        "currentStep": total_steps,
        "completedSteps": total_steps,
        "totalSteps": total_steps,
        "stepRunning": False,
    })
    previews = build_result_previews(result_inputs, result_output, current, diffs, forced_value_cells)
    return {
        "ok": True,
        "diffId": diff_id,
        "diffs": diffs,
        "forcedValueCells": forced_value_cells,
        "downloadId": download_id,
        "downloadUrl": f"/api/workbooks/download/{download_id}" if download_id else None,
        "downloadUrls": download_urls,
        "files": previews,
    }


if __name__ == "__main__":
    # [시작 유지관리] 트레이스 리셋 + 이전 실행 잔재 정리 + 런타임 유지관리 루프는 모두
    # start_runtime_maintenance_threads() 안으로 통합됐다(멱등). 프로즌 exe 진입점(launch_b2b.py)도
    # 이 함수를 호출하므로 진입점이 갈려도 정리가 누락되지 않는다.
    start_runtime_maintenance_threads()
    with B2BThreadingTCPServer((HOST, PORT), B2BHandler) as httpd:
        print(f"B2B serving on http://{HOST}:{PORT}")
        print(f"Proxying /v1/* to {VLLM_BASE}/v1/*")
        httpd.serve_forever()
