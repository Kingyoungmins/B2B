#!/usr/bin/env python3
"""실 test_data(`test_data/*.xlsx`) 기반 스키마 요약 생성기.

`vba_regression_runner.py` 의 Sonnet 검수(`schema_summary` 인자)에 주입할,
"이 워크북에 실제로 존재하는 시트/헤더/수식셀/병합/표"를 그대로 반영한
요약 문자열을 만든다. 기존 `mock_schema_summary()` 의 가상 데이터를 대체한다.

개발/품질평가 전용(exe 패키징 대상 아님). openpyxl + stdlib 만 사용.
파일이 변경/재생성되어도(`make_test_files.py`) 읽는 시점 구조를 그대로 반영한다.
"""

from __future__ import annotations

from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.utils import get_column_letter

# 레포 루트 = parents[2] (vba_regression → tests → repo root). 러너와 동일 규칙.
ROOT = Path(__file__).resolve().parents[2]
TEST_DATA_DIR = ROOT / "test_data"

# 입력/출력 역할 구분(파일명 접두사 기준). 시연 시나리오상 input_* 는 소스,
# output_* 는 사용자가 채워가는 대상 템플릿이다.
INPUT_PREFIX = "input_"
OUTPUT_PREFIX = "output_"

# 한 시트에 표가 여러 개인지 감지할 때 쓰는, "표 사이 빈 행" 최소 개수.
TABLE_GAP_ROWS = 1
MAX_SAMPLE_ROWS = 2          # 헤더 아래 미리보기 행 수
MAX_FORMULA_EXAMPLES = 4     # 시트당 노출할 수식 예시 수
MAX_COLS = 12                # 한 행에서 보여줄 최대 열 수


def _xlsx_files() -> list[Path]:
    if not TEST_DATA_DIR.is_dir():
        raise RuntimeError(
            f"test_data 디렉토리를 찾을 수 없습니다: {TEST_DATA_DIR}\n"
            "make_test_files.py 로 생성하거나 경로를 확인하세요."
        )
    files = sorted(TEST_DATA_DIR.glob("*.xlsx"))
    if not files:
        raise RuntimeError(f"test_data 에 .xlsx 가 없습니다: {TEST_DATA_DIR}")
    return files


def _cell_repr(value: Any) -> str:
    """셀 값을 요약 문자열로. 수식은 '=' 그대로, 긴 텍스트는 절단."""
    if value is None:
        return ""
    if isinstance(value, str) and value.startswith("="):
        return value  # 수식은 그대로 노출(=B4-C4 등)
    s = str(value)
    return s if len(s) <= 16 else s[:16] + "…"


def _row_values(ws, r: int, max_col: int) -> list[str]:
    out = []
    for c in range(1, min(max_col, MAX_COLS) + 1):
        out.append(_cell_repr(ws.cell(row=r, column=c).value))
    # 우측 빈 칸 제거
    while out and out[-1] == "":
        out.pop()
    return out


def _formula_cells(ws) -> list[tuple[str, str]]:
    cells = []
    for row in ws.iter_rows():
        for c in row:
            if isinstance(c.value, str) and c.value.startswith("="):
                cells.append((c.coordinate, c.value))
    return cells


def _first_nonempty_row(ws) -> int:
    """헤더로 볼 첫 데이터 행(제목/병합 타이틀을 건너뛰기 위해 2칸 이상 채워진 첫 행)."""
    for r in range(1, min(ws.max_row, 12) + 1):
        filled = sum(1 for c in range(1, min(ws.max_column, MAX_COLS) + 1)
                     if ws.cell(row=r, column=c).value not in (None, ""))
        if filled >= 2:
            return r
    return 1


def _row_fill(ws, r: int) -> int:
    return sum(1 for c in range(1, min(ws.max_column, MAX_COLS) + 1)
               if ws.cell(row=r, column=c).value not in (None, ""))


def _detect_table_starts(ws) -> list[int]:
    """한 시트 안 여러 표의 헤더 시작 행을 추정.

    표 사이는 빈 행(들)으로 분리되며, 각 표는 [제목행(1칸)] → 헤더행(2칸+) →
    데이터 순으로 올 수 있다. 따라서 "헤더로 보이는 행(2칸+)이고, 그 위가 빈 행이거나
    (1칸짜리 제목행 + 그 위가 빈 행)" 인 지점을 표 시작으로 본다.
    """
    starts = []
    seen_data = False
    for r in range(1, ws.max_row + 1):
        filled = _row_fill(ws, r)
        if filled < 2:
            continue
        above = _row_fill(ws, r - 1) if r > 1 else 0
        above2 = _row_fill(ws, r - 2) if r > 2 else 0
        # 직전이 빈 행 → 표 시작. 또는 직전이 1칸(제목)이고 그 위가 빈 행 → 표 시작.
        is_start = (r == 1) or (above == 0) or (above == 1 and above2 == 0)
        if is_start and (starts or seen_data or r > 1):
            starts.append(r)
        seen_data = True
    return starts


def _describe_sheet(ws) -> list[str]:
    lines: list[str] = []
    last_row = ws.max_row
    merged = [str(m) for m in ws.merged_cells.ranges]
    header_row = _first_nonempty_row(ws)
    headers = _row_values(ws, header_row, ws.max_column)

    lines.append(f'시트 "{ws.title}": 전체 {last_row}행, {ws.max_column}열' +
                 (f" · 병합 {merged}" if merged else ""))
    if headers:
        loc = "" if header_row == 1 else f"(행 {header_row}) "
        lines.append(f"  헤더 {loc}: {headers}")

    # 표가 여러 개면(빈 행으로 분리) 표 시작들을 노출 — 다중 표 인식 회귀 트리거.
    table_starts = _detect_table_starts(ws)
    if len(table_starts) >= 2:
        labels = []
        for s in table_starts:
            # 표 시작 직전 행에 제목이 있으면 같이 보여줌(예: '■ 상반기 집계')
            title = ws.cell(row=s - 1, column=1).value if s > 1 else None
            head = _row_values(ws, s, ws.max_column)
            tag = f"행{s}: {head}"
            if isinstance(title, str) and title.strip():
                tag = f"행{s} ({_cell_repr(title)}): {head}"
            labels.append(tag)
        lines.append(f"  ⚠️ 표 {len(table_starts)}개 감지 — " + " / ".join(labels))

    # 헤더 아래 샘플 행
    for r in range(header_row + 1, min(header_row + MAX_SAMPLE_ROWS, last_row) + 1):
        vals = _row_values(ws, r, ws.max_column)
        if vals:
            lines.append(f"  행 {r}: {vals}")

    # 수식 셀 — 좌표+식 그대로(수식 보존 채점의 핵심 근거).
    fcells = _formula_cells(ws)
    if fcells:
        examples = ", ".join(f"{coord}={f}" for coord, f in fcells[:MAX_FORMULA_EXAMPLES])
        more = f" …외 {len(fcells) - MAX_FORMULA_EXAMPLES}개" if len(fcells) > MAX_FORMULA_EXAMPLES else ""
        lines.append(f"  수식 셀 {len(fcells)}개: {examples}{more}")
        # 마지막 수식(합계/평균 행)도 노출
        if len(fcells) > MAX_FORMULA_EXAMPLES:
            tail = ", ".join(f"{coord}={f}" for coord, f in fcells[-2:])
            lines.append(f"    (마지막: {tail})")
    return lines


def build_real_schema_summary() -> str:
    """실 test_data 구조를 mock_schema_summary() 와 같은 포맷의 문자열로."""
    files = _xlsx_files()
    inputs = [f for f in files if f.name.startswith(INPUT_PREFIX)]
    outputs = [f for f in files if f.name.startswith(OUTPUT_PREFIX)]
    others = [f for f in files if f not in inputs and f not in outputs]

    parts: list[str] = ["## 입력 파일 목록 (수정 가능)"]
    for f in inputs + others:
        wb = openpyxl.load_workbook(f, data_only=False, read_only=False)
        parts.append(f"\n### {f.name}")
        for ws in wb.worksheets:
            parts.extend("  " + ln if not ln.startswith("시트") else ln
                         for ln in _describe_sheet(ws))
        wb.close()

    # 출력 템플릿 — 사용자가 채워가는 대상. 수식/병합 보존 채점의 기준.
    if outputs:
        parts.append("\n## 출력 템플릿 (원본, 수정 가능)")
        primary = outputs[0]
        for f in outputs:
            wb = openpyxl.load_workbook(f, data_only=False, read_only=False)
            parts.append(f"\n### {f.name}")
            for ws in wb.worksheets:
                parts.extend("  " + ln if not ln.startswith("시트") else ln
                             for ln in _describe_sheet(ws))
            wb.close()

        # 기본 대상/선택 셀 안내(러너의 mock 과 동일한 역할 — 멀티/단일 시트 판정용).
        wb = openpyxl.load_workbook(primary, data_only=False, read_only=False)
        first_sheet = wb.sheetnames[0]
        parts.append(
            "\n## 사용자가 현재 보고 있는 대상 (명령에 파일/시트가 없을 때 기본 대상)\n"
            f'[출력] 파일: "{primary.name}"\n'
            f'현재 활성 시트: "{first_sheet}"\n'
            "사용자가 파일/시트를 명시하지 않으면 이 시트를 기본 대상으로 사용하세요.\n"
            "'여기'·'선택한 셀'·'이 셀'은 현재 선택 위치를 의미합니다."
        )
        wb.close()

    return "\n".join(parts) + "\n"


if __name__ == "__main__":
    print(build_real_schema_summary())
