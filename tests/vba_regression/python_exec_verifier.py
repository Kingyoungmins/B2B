#!/usr/bin/env python3
"""생성된 Python 스킬(`def transform(ctx):`)을 실제로 실행해 결과를 검증한다.

개발/품질평가 전용(exe 패키징 대상 아님). 평가 기준 엔진은 **openpyxl 인프로세스**
(COM 아님)이며, 이 파일은 `ver0.4.8` 의 순수 Python 스킬 엔진을 그대로 가져와
(provenance 아래) 샌드박스에서 `transform(ctx)` 를 돌린 뒤, 케이스/변형의 선택적
`assert` 블록을 결과 워크북에 대조한다.

  Provenance: vendored from `git show ver0.4.8:serve_b2b.py`
    - normalize_text / normalize_python_pipeline_code
    - _opxl_coord / _OpxlCount / _OpxlRange
    - OpenpyxlWorksheetProxy / OpenpyxlWorkbookProxy / OpenpyxlSkillContext
    - _SKILL_ALLOWED_IMPORTS / _safe_skill_import / _safe_python_globals
  exe 런타임(serve_b2b.py)과 동일 수준의 샌드박스만 재현한다(그 이상 권한 없음).

중요한 한계(설계를 좌우):
- openpyxl 은 수식을 **재계산하지 않는다**. 출력 워크북은 data_only=False 로 열기 때문에
  수식 셀을 읽으면 **수식 문자열**('=B4-C4')이 그대로 나온다. 따라서 exec 검증은
  Excel 이 다시 계산한 '숫자'를 확인할 수 없다 → assert 는 (a) 수식 문자열 보존,
  (b) Python 으로 계산 가능한 입력값, (c) 숨김/병합/시트추가 같은 구조 변화만 본다.
"""

from __future__ import annotations

import io
import threading
import traceback
from pathlib import Path
from typing import Any, Callable

try:
    import openpyxl  # noqa: F401  (있으면 사용, 없으면 graceful skip)
    from openpyxl.utils.cell import coordinate_to_tuple, get_column_letter
    _OPENPYXL_OK = True
except Exception:  # pragma: no cover - 환경에 openpyxl 이 없을 때
    openpyxl = None  # type: ignore
    _OPENPYXL_OK = False


# 레포 루트 = parents[2] (vba_regression → tests → repo root). 러너/스키마와 동일 규칙.
ROOT = Path(__file__).resolve().parents[2]
DEFAULT_TEST_DATA_DIR = ROOT / "test_data"
INPUT_PREFIX = "input_"
OUTPUT_PREFIX = "output_"
DEFAULT_EXEC_TIMEOUT = 20  # 초. 무한 루프/과도 연산 방지 워치독.


# =====================================================================
#  --- vendored: ver0.4.8 의 순수 Python(openpyxl) 스킬 엔진 ---
#  COM 스킬과 동일한 ctx API + ws.Range/Cells/UsedRange/.Value 호환 shim.
# =====================================================================
def normalize_text(value):
    return "".join(str(value or "").lower().split())


def normalize_python_pipeline_code(code):
    import re
    text = str(code or "").replace("﻿", "")
    fence = re.search(r"```(?:python|py)?\s*\n([\s\S]*?)```", text, re.I)
    if fence:
        text = fence.group(1)
    lines = text.replace("\r\n", "\n").replace("\r", "\n").split("\n")
    normalized = []
    seen_python = False
    for line in lines:
        stripped = line.strip()
        if not seen_python:
            if not stripped:
                continue
            if stripped.startswith("```"):
                continue
            if stripped.startswith("//"):
                continue
            if re.match(r"^(제목|설명|설명문|title|description)\s*[:：]", stripped, re.I):
                continue
        if re.match(r"^(def|import|from|class|@)\b", stripped):
            seen_python = True
        if stripped.startswith("//"):
            line = line[: len(line) - len(line.lstrip())] + "#" + stripped[2:]
        normalized.append(line)
    return "\n".join(normalized).strip() + "\n"


def _opxl_coord(token):
    row, col = coordinate_to_tuple(str(token).replace("$", "").strip())
    return int(row), int(col)


class _OpxlCount:
    def __init__(self, count):
        self.Count = int(count)


class _OpxlRange:
    """openpyxl 워크시트 위의 직사각 범위(또는 단일 셀). COM Range.Value 시맨틱을 흉내낸다."""

    def __init__(self, ws, r1, c1, r2, c2):
        self._ws = ws
        self._r1, self._c1 = int(r1), int(c1)
        self._r2, self._c2 = int(r2), int(c2)

    @property
    def _single(self):
        return self._r1 == self._r2 and self._c1 == self._c2

    def _get_value(self):
        if self._single:
            return self._ws.cell(row=self._r1, column=self._c1).value
        out = []
        for r in range(self._r1, self._r2 + 1):
            out.append(tuple(self._ws.cell(row=r, column=c).value for c in range(self._c1, self._c2 + 1)))
        return tuple(out)

    def _set_value(self, value):
        if self._single and not isinstance(value, (list, tuple)):
            self._ws.cell(row=self._r1, column=self._c1, value=value)
            return
        if isinstance(value, (list, tuple)) and value and not isinstance(value[0], (list, tuple)):
            value = [value]  # 1행 그리드로 취급
        rows = list(value) if isinstance(value, (list, tuple)) else [[value]]
        for i, r in enumerate(range(self._r1, self._r2 + 1)):
            row = rows[i] if i < len(rows) else None
            if row is None:
                if self._single or not isinstance(value, (list, tuple)):
                    self._ws.cell(row=r, column=self._c1, value=value)
                continue
            if not isinstance(row, (list, tuple)):
                row = [row]
            for j, c in enumerate(range(self._c1, self._c2 + 1)):
                if j < len(row):
                    self._ws.cell(row=r, column=c, value=row[j])

    Value = property(_get_value, _set_value)
    Value2 = property(_get_value, _set_value)

    @property
    def Row(self):
        return self._r1

    @property
    def Column(self):
        return self._c1

    @property
    def Rows(self):
        return _OpxlCount(self._r2 - self._r1 + 1)

    @property
    def Columns(self):
        return _OpxlCount(self._c2 - self._c1 + 1)

    def Select(self):
        return self


class OpenpyxlWorksheetProxy:
    """openpyxl Worksheet 래퍼. COM 풍의 Range/Cells/UsedRange/.Name 을 제공하고
    그 외 속성/메서드(cell, insert_cols, append, max_row 등)는 openpyxl 로 위임한다."""

    def __init__(self, ws):
        object.__setattr__(self, "_ws", ws)

    def __getattr__(self, name):
        return getattr(object.__getattribute__(self, "_ws"), name)

    def __setattr__(self, key, value):
        if key == "_ws":
            object.__setattr__(self, key, value)
        elif key == "Name":
            self._ws.title = value
        else:
            setattr(self._ws, key, value)

    @property
    def Name(self):
        return self._ws.title

    @property
    def Parent(self):
        return self._ws.parent

    def Cells(self, r, c):
        return _OpxlRange(self._ws, r, c, r, c)

    @property
    def UsedRange(self):
        mr = self._ws.max_row or 1
        mc = self._ws.max_column or 1
        return _OpxlRange(self._ws, 1, 1, mr, mc)

    def Range(self, a1, a2=None):
        if a2 is not None:
            r1, c1 = a1._r1, a1._c1
            r2, c2 = a2._r1, a2._c1
            return _OpxlRange(self._ws, min(r1, r2), min(c1, c2), max(r1, r2), max(c1, c2))
        s = str(a1).replace("$", "").strip()
        if ":" in s:
            left, right = s.split(":", 1)
            r1, c1 = _opxl_coord(left)
            r2, c2 = _opxl_coord(right)
            return _OpxlRange(self._ws, min(r1, r2), min(c1, c2), max(r1, r2), max(c1, c2))
        r, c = _opxl_coord(s)
        return _OpxlRange(self._ws, r, c, r, c)


class OpenpyxlWorkbookProxy:
    def __init__(self, ctx, workbook, name=None):
        self._ctx = ctx
        self._workbook = workbook
        self.name = name or ""

    @property
    def raw(self):
        return self._workbook

    def __getattr__(self, name):
        return getattr(self._workbook, name)

    def sheet(self, name=None):
        return self._ctx.sheet(name, workbook=self)

    def sheet_like(self, name=None):
        return self._ctx.sheet_like(name, workbook=self)

    def range(self, sheet_or_name, address):
        return self._ctx.range(sheet_or_name, address, workbook=self)

    def rows(self, sheet_or_name=None):
        return self._ctx.rows(sheet_or_name, workbook=self)

    def col(self, sheet_or_name, header, header_rows=20):
        return self._ctx.col(sheet_or_name, header, workbook=self, header_rows=header_rows)

    def header_row(self, sheet_or_name=None, header_rows=20):
        return self._ctx.header_row(sheet_or_name, workbook=self, header_rows=header_rows)


class OpenpyxlSkillContext:
    """COM ExcelSkillContext 와 동일한 API를 openpyxl 위에서 제공한다."""

    def __init__(self, output_wb, input_wbs):
        self.excel = None
        self._workbook = output_wb
        self.workbook = OpenpyxlWorkbookProxy(self, output_wb, "output")
        self.output = self.workbook
        self.last_output_sheet = None
        self.last_output_address = None
        self.inputs = {
            name: OpenpyxlWorkbookProxy(self, wb, name)
            for name, wb in (input_wbs or {}).items()
        }

    def _unwrap_workbook(self, wb):
        return wb.raw if isinstance(wb, OpenpyxlWorkbookProxy) else wb

    def _is_output_workbook(self, wb):
        return self._unwrap_workbook(wb) is self._workbook

    def normalize(self, value):
        return normalize_text(value)

    def _sheet_names(self, wb):
        return list(self._unwrap_workbook(wb).sheetnames)

    def workbook_like(self, hint=None):
        if not hint:
            return self.workbook
        norm = self.normalize(hint)
        candidates = [(name, wb) for name, wb in self.inputs.items()]
        candidates.append(("output", self.workbook))
        for name, wb in candidates:
            if self.normalize(name) == norm:
                return wb
        for name, wb in candidates:
            if norm in self.normalize(name) or self.normalize(name) in norm:
                return wb
        for _, wb in candidates:
            if self._find_sheet_name(wb, hint, allow_single=False):
                return wb
        if len(self.inputs) == 1:
            return next(iter(self.inputs.values()))
        raise RuntimeError(f"workbook not found: {hint}")

    def input(self, hint=None):
        if hint is None:
            if len(self.inputs) == 1:
                return next(iter(self.inputs.values()))
            raise RuntimeError("input workbook hint is required when multiple input files exist")
        return self.workbook_like(hint)

    def _find_sheet_name(self, wb, name=None, allow_single=True):
        raw = self._unwrap_workbook(wb)
        names = list(raw.sheetnames)
        if not names:
            return None
        if not name:
            try:
                return raw.active.title
            except Exception:
                return names[0]
        norm = self.normalize(name)
        for sheet_name in names:
            if self.normalize(sheet_name) == norm:
                return sheet_name
        for sheet_name in names:
            sheet_norm = self.normalize(sheet_name)
            if norm in sheet_norm or sheet_norm in norm:
                return sheet_name
        if allow_single and len(names) == 1:
            return names[0]
        return None

    def sheet(self, name=None, workbook=None):
        raw = self._unwrap_workbook(workbook or self.workbook)
        sheet_name = self._find_sheet_name(workbook or self.workbook, name)
        if not sheet_name:
            raise RuntimeError(f"sheet not found: {name}")
        ws = OpenpyxlWorksheetProxy(raw[sheet_name])
        if self._is_output_workbook(raw):
            self.last_output_sheet = ws.Name
        return ws

    def sheet_like(self, name=None, workbook=None):
        return self.sheet(name, workbook)

    def input_sheet(self, sheet_hint=None, file_hint=None):
        workbooks = []
        if file_hint:
            workbooks.append(self.workbook_like(file_hint))
        else:
            workbooks.extend(self.inputs.values())
        for wb in workbooks:
            sheet_name = self._find_sheet_name(wb, sheet_hint, allow_single=True)
            if sheet_name:
                return OpenpyxlWorksheetProxy(self._unwrap_workbook(wb)[sheet_name])
        raise RuntimeError(f"input sheet not found: {sheet_hint}")

    def range(self, sheet_or_name, address, workbook=None):
        ws = sheet_or_name if hasattr(sheet_or_name, "Range") else self.sheet(sheet_or_name, workbook)
        try:
            if self._is_output_workbook(ws.Parent):
                self.last_output_sheet = ws.Name
                self.last_output_address = str(address)
        except Exception:
            pass
        return ws.Range(str(address))

    def _ws_of(self, sheet_or_name, workbook=None):
        return sheet_or_name if hasattr(sheet_or_name, "UsedRange") else self.sheet(sheet_or_name, workbook)

    def rows(self, sheet_or_name, workbook=None):
        ws = self._ws_of(sheet_or_name, workbook)
        raw = getattr(ws, "_ws", ws)
        out = []
        for row in raw.iter_rows(values_only=True):
            out.append(list(row))
        while out and all(v is None or v == "" for v in out[-1]):
            out.pop()
        return out

    def col(self, sheet_or_name, header, workbook=None, header_rows=20):
        rows = self.rows(sheet_or_name, workbook)
        target = self.normalize(header)
        for row in rows[:header_rows]:
            for c_idx, value in enumerate(row, start=1):
                if self.normalize(value) == target:
                    return c_idx
        for row in rows[:header_rows]:
            for c_idx, value in enumerate(row, start=1):
                if target and target in self.normalize(value):
                    return c_idx
        return -1

    def header_row(self, sheet_or_name=None, workbook=None, header_rows=20):
        rows = self.rows(sheet_or_name, workbook)
        best_idx = 1
        best_score = -1
        for idx, row in enumerate(rows[:header_rows], start=1):
            score = sum(1 for value in row if value not in (None, ""))
            if score > best_score:
                best_idx, best_score = idx, score
        return best_idx

    def data_start_row(self, sheet_or_name=None, workbook=None, header_rows=20):
        return self.header_row(sheet_or_name, workbook, header_rows) + 1

    def _col0(self, rows, name_or_idx, header_rows=20):
        if isinstance(name_or_idx, int):
            return max(0, name_or_idx - 1)
        target = self.normalize(name_or_idx)
        scan = rows[:header_rows] if header_rows else rows
        for row in scan:
            for i, v in enumerate(row or []):
                if self.normalize(v) == target:
                    return i
        for row in scan:
            for i, v in enumerate(row or []):
                if target and target in self.normalize(v):
                    return i
        return None

    def add_sheet(self, name, workbook=None):
        wb = self._unwrap_workbook(workbook or self.workbook)
        base = (str(name) or "Sheet")[:31]
        existing = {self.normalize(n) for n in wb.sheetnames}
        final = base
        idx = 1
        while self.normalize(final) in existing:
            idx += 1
            suffix = "_" + str(idx)
            final = base[: max(1, 31 - len(suffix))] + suffix
        raw_ws = wb.create_sheet(title=final)
        ws = OpenpyxlWorksheetProxy(raw_ws)
        if self._is_output_workbook(wb):
            self.last_output_sheet = ws.Name
        return ws

    def _write_grid(self, ws, grid, start_row=1, start_col=1):
        if not grid:
            return ws
        raw = getattr(ws, "_ws", ws)
        for i, row in enumerate(grid):
            for j, value in enumerate(row or []):
                raw.cell(row=start_row + i, column=start_col + j, value=value)
        return ws

    def write_grid(self, ws, grid, start_row=1, start_col=1):
        ws = self._ws_of(ws)
        self._write_grid(ws, grid, start_row, start_col)
        if self._is_output_workbook(ws.Parent):
            self.last_output_sheet = ws.Name
        return ws

    def set_range(self, sheet_or_name, address, grid, workbook=None):
        ws = self._ws_of(sheet_or_name, workbook)
        if not grid:
            return ws
        r0, c0 = _opxl_coord(str(address).split(":")[0])
        self._write_grid(ws, grid, start_row=r0, start_col=c0)
        if self._is_output_workbook(ws.Parent):
            self.last_output_sheet = ws.Name
            self.last_output_address = str(address)
        return ws

    def sort(self, sheet_or_name, by, ascending=True, header=True, workbook=None):
        ws = self._ws_of(sheet_or_name, workbook)
        rows = self.rows(ws)
        rel = self._col0(rows, by)
        if rel is None:
            raise RuntimeError("sort: column not found: %r" % (by,))
        hdr_count = 1 if header else 0
        head = rows[:hdr_count]
        body = rows[hdr_count:]

        def _key(r):
            v = r[rel] if rel < len(r) else None
            num = self._num(v)
            return (0, num) if num is not None else (1, self.normalize(v))

        body.sort(key=_key, reverse=not ascending)
        raw = getattr(ws, "_ws", ws)
        max_col = max((len(r) for r in rows), default=0)
        for r_idx in range(1, (raw.max_row or 0) + 1):
            for c_idx in range(1, max_col + 1):
                raw.cell(row=r_idx, column=c_idx, value=None)
        self._write_grid(ws, list(head) + body)
        if self._is_output_workbook(ws.Parent):
            self.last_output_sheet = ws.Name
        return ws

    def filter_to_sheet(self, sheet_or_name, predicate, dest_name, header_rows=1, workbook=None):
        ws = self._ws_of(sheet_or_name, workbook)
        rows = self.rows(ws)
        hr = max(0, int(header_rows or 0))
        header = rows[:hr]
        matched = []
        for r in rows[hr:]:
            try:
                if predicate(r):
                    matched.append(r)
            except Exception:
                continue
        dest = self.add_sheet(dest_name, workbook=workbook or self.workbook)
        self._write_grid(dest, list(header) + matched)
        return dest

    @staticmethod
    def _num(v):
        if isinstance(v, bool):
            return None
        if isinstance(v, (int, float)):
            return float(v)
        if isinstance(v, str):
            s = v.strip().replace(",", "")
            try:
                return float(s)
            except ValueError:
                return None
        return None

    def pivot(self, sheet_or_name, group_by, value=None, agg="sum", dest_name=None, header_rows=1, workbook=None):
        ws = self._ws_of(sheet_or_name, workbook)
        rows = self.rows(ws)
        hr = max(1, int(header_rows or 1))
        header_row = rows[hr - 1] if len(rows) >= hr else []
        data = rows[hr:]
        group_cols = list(group_by) if isinstance(group_by, (list, tuple)) else [group_by]
        gidx = [self._col0(rows, g, hr) for g in group_cols]
        vidx = self._col0(rows, value, hr) if value is not None else None
        agg = str(agg or "sum").lower()

        groups: dict[Any, list] = {}
        order: list = []
        for r in data:
            key = tuple((r[i] if (i is not None and i < len(r)) else "") for i in gidx)
            if key not in groups:
                groups[key] = []
                order.append(key)
            if vidx is not None and vidx < len(r):
                groups[key].append(r[vidx])

        def _aggregate(vals):
            nums = [n for n in (self._num(v) for v in vals) if n is not None]
            if agg == "count":
                return len(vals)
            if agg in ("avg", "average", "mean"):
                return (sum(nums) / len(nums)) if nums else 0
            if agg == "max":
                return max(nums) if nums else ""
            if agg == "min":
                return min(nums) if nums else ""
            return sum(nums)

        out_header = []
        for n, i in enumerate(gidx):
            label = header_row[i] if (i is not None and i < len(header_row)) else ("그룹%d" % (n + 1))
            out_header.append(label)
        value_label = (str(value) if value is not None else "값") + "_" + (agg if agg != "average" else "avg")
        out_header.append(value_label)
        grid = [out_header]
        for key in order:
            grid.append(list(key) + [_aggregate(groups[key])])
        dest = self.add_sheet(dest_name or "피벗요약", workbook=workbook or self.workbook)
        self._write_grid(dest, grid)
        return dest


# 스킬에서 import 가능한 안전한 표준 라이브러리만 허용(os/sys/subprocess 등은 차단).
_SKILL_ALLOWED_IMPORTS = {
    "re", "datetime", "math", "json", "collections", "itertools",
    "functools", "string", "decimal", "statistics", "calendar",
    "textwrap", "unicodedata", "fractions", "random", "operator", "copy",
}


def _safe_skill_import(name, globals=None, locals=None, fromlist=(), level=0):
    import importlib
    if level and level != 0:
        raise ImportError("relative imports are not allowed in skills")
    root = str(name or "").split(".")[0]
    if root not in _SKILL_ALLOWED_IMPORTS:
        raise ImportError(
            "import of '%s' is not allowed in skills (allowed: %s)"
            % (name, ", ".join(sorted(_SKILL_ALLOWED_IMPORTS)))
        )
    return importlib.import_module(name)


def _safe_python_globals():
    import datetime as _datetime
    import math as _math
    import re as _re
    allowed_builtins = {
        "abs": abs, "all": all, "any": any, "bool": bool, "dict": dict, "enumerate": enumerate,
        "float": float, "int": int, "isinstance": isinstance, "len": len, "list": list,
        "max": max, "min": min, "print": print, "range": range, "round": round,
        "set": set, "sorted": sorted, "str": str, "sum": sum, "tuple": tuple,
        "type": type, "zip": zip, "getattr": getattr, "hasattr": hasattr, "iter": iter,
        "next": next, "repr": repr, "divmod": divmod, "ord": ord, "chr": chr,
        "filter": filter, "map": map, "reversed": reversed, "format": format, "frozenset": frozenset,
        "__import__": _safe_skill_import,
        "Exception": Exception, "RuntimeError": RuntimeError, "ValueError": ValueError,
        "TypeError": TypeError, "KeyError": KeyError, "IndexError": IndexError,
        "AttributeError": AttributeError, "ZeroDivisionError": ZeroDivisionError,
        "StopIteration": StopIteration,
    }
    return {
        "__builtins__": allowed_builtins,
        "datetime": _datetime,
        "math": _math,
        "re": _re,
    }


# =====================================================================
#  --- verifier: test_data 로드 → 샌드박스 exec → assert 대조 ---
# =====================================================================
def _load_test_workbooks(test_data_dir: Path):
    """입력(input_*) 은 data_only=True(계산값), 출력(output_*) 은 data_only=False(수식 보존)
    로 메모리 사본을 만들어 ver4.x 엔진과 동일하게 ctx 를 구성한다. 디스크엔 쓰지 않는다."""
    files = sorted(test_data_dir.glob("*.xlsx"))
    if not files:
        raise RuntimeError(f"test_data 에 .xlsx 가 없습니다: {test_data_dir}")
    outputs = [f for f in files if f.name.startswith(OUTPUT_PREFIX)]
    inputs = [f for f in files if f.name.startswith(INPUT_PREFIX)]
    others = [f for f in files if f not in outputs and f not in inputs]
    if not outputs:
        # output_ 접두사가 없으면 첫 파일을 출력으로 간주(스키마 요약과 동일한 휴리스틱).
        outputs = [files[0]]
        inputs = [f for f in files if f is not files[0]]

    def _load(path: Path, data_only: bool):
        data = path.read_bytes()
        return openpyxl.load_workbook(io.BytesIO(data), data_only=data_only)

    output_wb = _load(outputs[0], data_only=False)
    input_wbs: dict[str, Any] = {}
    for f in inputs + others:
        input_wbs[f.name] = _load(f, data_only=True)
    # 출력 파일도 이름으로 참조될 수 있게 inputs 에 같은 객체로 노출(ctx.input("청구서") 등).
    input_wbs.setdefault(outputs[0].name, output_wb)
    return output_wb, input_wbs, outputs[0].name


def _run_with_timeout(fn: Callable[[], Any], timeout: int) -> tuple[bool, Any]:
    """fn 을 데몬 스레드에서 돌리고 timeout 초 안에 끝나지 않으면 (False, TimeoutError) 반환.
    스레드를 강제 종료할 수는 없으나(파이썬 한계), 평가 진행은 막지 않는다."""
    box: dict[str, Any] = {}

    def _target():
        try:
            box["result"] = fn()
        except BaseException as err:  # noqa: BLE001 — exec 내부 예외 보존
            box["error"] = err
            box["tb"] = traceback.format_exc()

    th = threading.Thread(target=_target, daemon=True)
    th.start()
    th.join(timeout)
    if th.is_alive():
        return False, TimeoutError(f"exec exceeded {timeout}s")
    if "error" in box:
        err = box["error"]
        setattr(err, "_tb", box.get("tb"))
        raise err
    return True, box.get("result")


# ---- assert 블록 평가 헬퍼 ----
def _split_ref(ref: str) -> tuple[str | None, str]:
    """'회사별요약!D4' → ('회사별요약','D4'). 시트 미지정이면 (None, addr)."""
    s = str(ref)
    if "!" in s:
        sheet, addr = s.split("!", 1)
        return sheet.strip(), addr.strip()
    return None, s.strip()


def _resolve_ws(ctx: OpenpyxlSkillContext, sheet: str | None):
    return ctx._unwrap_workbook(ctx.workbook) if sheet is None else \
        ctx._unwrap_workbook(ctx.workbook)[ctx._find_sheet_name(ctx.workbook, sheet)]


def _cell_value(ctx: OpenpyxlSkillContext, ref: str):
    sheet, addr = _split_ref(ref)
    raw = _resolve_ws(ctx, sheet)
    r, c = _opxl_coord(addr)
    return raw.cell(row=r, column=c).value


def _num(v):
    return OpenpyxlSkillContext._num(v)


def _check_asserts(ctx: OpenpyxlSkillContext, baseline: dict[str, Any], spec: dict[str, Any]) -> dict[str, Any]:
    """variant 의 assert 블록을 결과 워크북에 대조한다. {ok, checks:[{name,ok,detail}]}."""
    checks: list[dict[str, Any]] = []

    def add(name, ok, detail=""):
        checks.append({"name": name, "ok": bool(ok), "detail": detail})

    # 1) expect_cells: {"월별실적!B4": {"value": 31139[, "approx": 0.01]}} 또는 {"text":"3월"}
    for ref, exp in (spec.get("expect_cells") or {}).items():
        actual = _cell_value(ctx, ref)
        if "value" in exp:
            want = exp["value"]
            an, wn = _num(actual), _num(want)
            if an is not None and wn is not None:
                tol = abs(wn) * float(exp.get("approx", 0)) if exp.get("approx") else 0
                ok = abs(an - wn) <= tol
            else:
                ok = actual == want
            add(f"cell {ref}=={want!r}", ok, f"actual={actual!r}")
        elif "text" in exp:
            ok = normalize_text(actual) == normalize_text(exp["text"])
            add(f"cell {ref} text=={exp['text']!r}", ok, f"actual={actual!r}")
        elif "is_formula" in exp:
            isf = isinstance(actual, str) and actual.startswith("=")
            ok = isf == bool(exp["is_formula"])
            add(f"cell {ref} is_formula=={exp['is_formula']}", ok, f"actual={actual!r}")

    # 2) expect_formula_preserved: ["회사별요약!D4", ...] — 여전히 '=' 로 시작하는 수식 문자열
    for ref in spec.get("expect_formula_preserved") or []:
        actual = _cell_value(ctx, ref)
        ok = isinstance(actual, str) and actual.startswith("=")
        add(f"formula preserved {ref}", ok, f"actual={actual!r}")

    # 3) expect_hidden_cols / expect_hidden_rows
    hc = spec.get("expect_hidden_cols")
    if hc is not None:
        sheet = spec.get("assert_sheet")
        raw = _resolve_ws(ctx, sheet) if sheet else ctx._unwrap_workbook(ctx.workbook).active
        ok = all(bool(raw.column_dimensions[col].hidden) for col in hc)
        shown = {col: bool(raw.column_dimensions[col].hidden) for col in hc}
        add(f"hidden cols {hc}", ok, f"{shown}")
    hr = spec.get("expect_hidden_rows")
    if hr is not None:
        sheet = spec.get("assert_sheet")
        raw = _resolve_ws(ctx, sheet) if sheet else ctx._unwrap_workbook(ctx.workbook).active
        ok = all(bool(raw.row_dimensions[int(rn)].hidden) for rn in hr)
        shown = {int(rn): bool(raw.row_dimensions[int(rn)].hidden) for rn in hr}
        add(f"hidden rows {hr}", ok, f"{shown}")

    # 4) expect_merged / expect_unmerged: ["A1:E1"] (assert_sheet 기준)
    sheet = spec.get("assert_sheet")
    if spec.get("expect_merged") is not None or spec.get("expect_unmerged") is not None:
        raw = _resolve_ws(ctx, sheet) if sheet else ctx._unwrap_workbook(ctx.workbook).active
        merged = {str(m) for m in raw.merged_cells.ranges}
        for rng in spec.get("expect_merged") or []:
            add(f"merged {rng}", rng in merged, f"merged={sorted(merged)}")
        for rng in spec.get("expect_unmerged") or []:
            add(f"unmerged {rng}", rng not in merged, f"merged={sorted(merged)}")

    # 5) expect_sheet_added: "이름" (실행 전 없던 시트가 생겼는지)
    added = spec.get("expect_sheet_added")
    if added:
        names = {normalize_text(n) for n in ctx._unwrap_workbook(ctx.workbook).sheetnames}
        ok = normalize_text(added) in names and normalize_text(added) not in baseline["sheet_names"]
        add(f"sheet added {added!r}", ok, f"sheets={ctx._unwrap_workbook(ctx.workbook).sheetnames}")

    # 6) expect_no_change_to: ["회사별요약!D4", ...] — baseline 과 동일해야(불변)
    for ref in spec.get("expect_no_change_to") or []:
        before = baseline["cells"].get(ref, "__missing__")
        after = _cell_value(ctx, ref)
        add(f"unchanged {ref}", before == after, f"before={before!r} after={after!r}")

    ok_all = all(c["ok"] for c in checks)
    return {"ok": ok_all, "checks": checks}


def _snapshot_baseline(test_data_dir: Path, spec: dict[str, Any]) -> dict[str, Any]:
    """assert 대조에 필요한 실행 전 상태(불변 검증용 셀값, 시트 목록)를 미리 떠둔다."""
    out_wb, in_wbs, _ = _load_test_workbooks(test_data_dir)
    ctx = OpenpyxlSkillContext(out_wb, in_wbs)
    cells: dict[str, Any] = {}
    for ref in spec.get("expect_no_change_to") or []:
        try:
            cells[ref] = _cell_value(ctx, ref)
        except Exception:
            cells[ref] = "__error__"
    return {
        "cells": cells,
        "sheet_names": {normalize_text(n) for n in out_wb.sheetnames},
    }


def collect_assert_spec(case: dict[str, Any], variant: dict[str, Any]) -> dict[str, Any]:
    """case/variant 의 assert 블록을 합친다(variant 우선). 둘 다 없으면 빈 dict."""
    spec: dict[str, Any] = {}
    spec.update(case.get("assert") or {})
    spec.update(variant.get("assert") or {})
    return spec


def verify(
    code: str,
    case: dict[str, Any],
    variant: dict[str, Any],
    test_data_dir: "str | Path | None" = None,
    timeout: int = DEFAULT_EXEC_TIMEOUT,
) -> dict[str, Any]:
    """생성 코드를 openpyxl 엔진으로 실행하고 assert 블록을 대조한다.

    반환:
      {available, ok, ran, matches_expected, has_asserts, error, traceback, asserts}
        available        — openpyxl 사용 가능 여부(False 면 검증 skip)
        ran              — transform(ctx) 가 예외 없이 끝났는가
        has_asserts      — variant/case 에 assert 블록이 있는가
        matches_expected — assert 가 있으면 그 결과, 없으면 None(실행만 확인)
        ok               — 종합(실행 성공 & (assert 없음 or 통과)). expect_raises 처리 포함.
        asserts          — _check_asserts 의 상세
    """
    spec = collect_assert_spec(case, variant)
    expect_raises = bool(spec.get("expect_raises"))
    result: dict[str, Any] = {
        "available": _OPENPYXL_OK,
        "ok": None,
        "ran": False,
        "matches_expected": None,
        "has_asserts": bool(spec),
        "error": None,
        "traceback": None,
        "asserts": None,
    }
    if not _OPENPYXL_OK:
        result["error"] = "openpyxl 미설치 — exec 검증 skip"
        return result

    tdir = Path(test_data_dir) if test_data_dir else DEFAULT_TEST_DATA_DIR
    try:
        baseline = _snapshot_baseline(tdir, spec)
        out_wb, in_wbs, _ = _load_test_workbooks(tdir)
    except Exception as err:  # 데이터 로드 실패는 검증 불가로 보고(케이스 FAIL 아님).
        result["error"] = f"test_data 로드 실패: {err}"
        return result

    ctx = OpenpyxlSkillContext(out_wb, in_wbs)
    normalized = normalize_python_pipeline_code(code)
    ns = _safe_python_globals()

    def _do():
        exec(compile(normalized, "<python_skill>", "exec"), ns, ns)
        fn = ns.get("transform")
        if not callable(fn):
            raise RuntimeError("Python step must define def transform(ctx):")
        return fn(ctx)

    raised_err: BaseException | None = None
    try:
        ok_run, _ = _run_with_timeout(_do, timeout)
        if not ok_run:
            result["error"] = f"timeout: {timeout}s 내 미완료"
            result["ok"] = False
            return result
        result["ran"] = True
    except BaseException as err:  # noqa: BLE001 — 실행 예외를 케이스 결과로 격리
        raised_err = err
        result["error"] = f"{type(err).__name__}: {err}"
        result["traceback"] = getattr(err, "_tb", None) or traceback.format_exc()
        result["ran"] = False

    # expect_raises: 실패(예외)를 '드러내야' 정상인 케이스(미적용 보고 방지 등).
    if expect_raises:
        result["ok"] = raised_err is not None
        result["matches_expected"] = result["ok"]
        result["asserts"] = {"ok": result["ok"], "checks": [
            {"name": "expect_raises", "ok": result["ok"],
             "detail": result.get("error") or "예외 없이 종료(실패를 드러내지 못함)"}
        ]}
        return result

    if raised_err is not None:
        result["ok"] = False
        return result

    if spec:
        try:
            asserts = _check_asserts(ctx, baseline, spec)
        except Exception as err:  # assert 평가 자체 오류
            result["error"] = f"assert 평가 오류: {err}"
            result["traceback"] = traceback.format_exc()
            result["ok"] = False
            return result
        result["asserts"] = asserts
        result["matches_expected"] = asserts["ok"]
        result["ok"] = asserts["ok"]
    else:
        # assert 없음 → '실행만' 성공으로 간주(상세 판정은 Sonnet 에 위임).
        result["ok"] = True

    return result


if __name__ == "__main__":
    # 간단 자가진단: 손으로 쓴 transform 으로 verify 동작 확인.
    import json
    sample = """
제목: 매출 건수 합계를 월별실적 B4에 값으로
def transform(ctx):
    ws_in = ctx.input("매출").sheet("매출")
    rows = ctx.rows(ws_in)
    total = sum(r[2] for r in rows[1:] if isinstance(r[2], (int, float)))
    out = ctx.sheet("월별실적")
    out.Range("B4").Value = total
"""
    case = {"id": "selfcheck", "assert": {}}
    variant = {"id": "v1", "assert": {"expect_cells": {"월별실적!B4": {"value": 31139}}}}
    print(json.dumps(verify(sample, case, variant), ensure_ascii=False, indent=2))
