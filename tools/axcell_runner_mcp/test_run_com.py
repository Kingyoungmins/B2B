# -*- coding: utf-8 -*-
"""[0.8.2 호환 검증] runner_core.run — 실제 Excel COM 으로 0.8.2 엔진과 끝까지 돈다.

test_core.py 는 Excel 없이 도는 검사(check/매핑/MCP 프로토콜)만 있다. 여기는 나머지:
  1) python + VBA + 교차파일 스텝이 0.8.2 엔진(_exec_python_com_skill/_inject_and_run_vba)으로 돈다
  2) @@FILE_n@@ 핸들 복원 + 다른 달 파일명 자동 매핑(안정키)이 실행 경로에서도 성립
  3) 원본은 절대 수정되지 않는다(작업 사본), 출력은 out_dir 에만
  4) 이벤트 콜백(open/step/saved) 순서와 개수
  5) 전체 열(D:D) 교차 복사도 0.8.2 클램프 덕에 빠르다
"""
import io
import json
import sys
import tempfile
import shutil
import time
import zipfile
from pathlib import Path

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")
HERE = Path(__file__).resolve().parent
sys.path.insert(0, str(HERE))
import openpyxl

from axcell_runner import runner_core as core

fails = []


def check(name, cond, detail=""):
    print(("  PASS  " if cond else "  FAIL  ") + name + (("  -> " + str(detail)[:250]) if not cond else ""))
    if not cond:
        fails.append(name)


work = Path(tempfile.mkdtemp(prefix="axr_run_"))
in_dir = work / "inputs"; in_dir.mkdir()

# 입력: 스킬은 4월 파일명을 기억하지만 디렉토리엔 5월 파일 — 안정키 매핑이 이어줘야 한다
wb = openpyxl.Workbook(); ws = wb.active; ws.title = "매출"
for i in range(1, 31):
    ws.cell(row=i, column=4, value=1000 + i)          # D1..D30
wb.save(str(in_dir / "input_매출_2026_5월.xlsx"))
w2 = openpyxl.Workbook(); ws2 = w2.active; ws2.title = "원가"
ws2["A1"] = "원본"
w2.save(str(in_dir / "input_원가_2026_5월.xlsx"))
orig_bytes = (in_dir / "input_매출_2026_5월.xlsx").read_bytes()

PY1 = ("def transform(ctx):\n"
       "    ctx.write('매출', 'F1', [['py-ok']])\n")
VBA2 = ("Sub B2BSkill()\n"
        "    ThisWorkbook.Worksheets(1).Range(\"A1\").Value = \"vba-ok\"\n"
        "End Sub\n")
PY3 = ("def transform(ctx):\n"                       # 교차파일 + 전체 열(0.8.2 클램프 검증)
       "    ctx.copy('input_매출_2026_4월.xlsx!매출', 'D:D', '원가', 'B1')\n")

manifest = {
    "name": "호환검증스킬",
    "requiredFiles": [
        {"handle": "@@FILE_1@@", "name": "input_매출_2026_4월.xlsx"},
        {"handle": "@@FILE_2@@", "name": "input_원가_2026_4월.xlsx"},
    ],
    "pipeline": [
        {"id": "s1", "language": "python", "code": PY1.replace("매출', 'F1", "매출', 'F1"),
         "targetFileId": "input:input_매출_2026_4월.xlsx", "enabled": True},
        {"id": "s2", "language": "vba", "code": VBA2,
         "targetFileId": "input:input_원가_2026_4월.xlsx", "enabled": True},
        {"id": "s3", "language": "python", "code": PY3.replace("input_매출_2026_4월.xlsx", "@@FILE_1@@"),
         "targetFileId": "input:input_원가_2026_4월.xlsx", "enabled": True},
        {"id": "s4", "language": "python", "code": "def transform(ctx):\n    pass\n",
         "targetFileId": "", "enabled": False},          # 꺼진 스텝 — 건너뛰어야 함
    ],
}
zip_path = work / "스킬.zip"
with zipfile.ZipFile(zip_path, "w") as z:
    z.writestr("호환검증스킬.logic.json", json.dumps(manifest, ensure_ascii=False))

print("[0] check_inputs — 다른 달 파일명도 매핑된다")
chk = core.check_inputs(str(zip_path), str(in_dir))
check("매핑 ok", chk["ok"], chk)
check("켜진 스텝만 센다(3)", chk["total_steps"] == 3, chk["total_steps"])
check("두 언어 모두 감지", chk["languages"] == ["python", "vba"], chk["languages"])

print("[1] run — 실제 Excel 로 끝까지")
events = []
t0 = time.perf_counter()
out = core.run(str(zip_path), str(in_dir), str(work / "out"), on_event=events.append)
el = time.perf_counter() - t0
check("완료 ok", out.get("ok") is True, out)
check("출력 2파일", sorted(out.get("files") or []) ==
      ["input_매출_2026_5월.xlsx", "input_원가_2026_5월.xlsx"], out.get("files"))
print("      실행 %.1f초 / 이벤트 %d개" % (el, len(events)))

print("[2] 결과 값 — 세 스텝 전부 반영됐다")
r1 = openpyxl.load_workbook(str(work / "out" / "input_매출_2026_5월.xlsx"))
r2 = openpyxl.load_workbook(str(work / "out" / "input_원가_2026_5월.xlsx"))
check("1스텝(python)", r1["매출"]["F1"].value == "py-ok", r1["매출"]["F1"].value)
check("2스텝(VBA)", r2["원가"]["A1"].value == "vba-ok", r2["원가"]["A1"].value)
check("3스텝(교차 D:D→B) 첫 값", r2["원가"]["B1"].value == 1001, r2["원가"]["B1"].value)
check("3스텝 끝 값(30행)", r2["원가"]["B30"].value == 1030, r2["원가"]["B30"].value)
check("클램프 — 31행 밖은 비었다", r2["원가"]["B31"].value in (None, ""), r2["원가"]["B31"].value)

print("[3] 원본 무수정 + 이벤트 계약")
check("원본 그대로", (in_dir / "input_매출_2026_5월.xlsx").read_bytes() == orig_bytes)
types = [e["type"] for e in events]
check("open 2회 → step 3회 → saved 1회", types.count("open") == 2 and types.count("step") == 3
      and types[-1] == "saved", types)
check("step 이벤트에 언어가 실린다", [e.get("language") for e in events if e["type"] == "step"]
      == ["python", "vba", "python"], [e.get("language") for e in events if e["type"] == "step"])

print("[4] package_outputs")
pkg = core.package_outputs(str(work / "out"), str(work / "결과.zip"),
                           expect_files=["input_원가_2026_5월.xlsx"])
check("zip ok", pkg.get("ok") is True, pkg)

shutil.rmtree(work, ignore_errors=True)
print("")
print("RESULT: ALL PASS" if not fails else "RESULT: %d FAIL -> %s" % (len(fails), fails))
sys.exit(0 if not fails else 1)
