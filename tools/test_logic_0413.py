#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Headless regression harness for the old 2026-05-26 logic conversation.

The harness does not open the browser. It copies the real workbooks to a run
directory, asks the configured OpenAI-compatible vLLM to regenerate Python
skills from the old user prompts, then executes the generated pipeline through
the current 0.4.13 backend runner.
"""

from __future__ import annotations

import argparse
import ast
import datetime as _dt
import json
import os
from pathlib import Path
import re
import shutil
import sys
import threading
import time
import urllib.error
import urllib.request
import uuid

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

import serve_b2b  # noqa: E402


MODEL = "Qwen3.5-27B-FP8"
API_KEY = "7365676d"
DEFAULT_LLM_BASES = [
    "http://localhost:8016/v1",
    "http://127.0.0.1:8016/v1",
    "http://192.168.219.105:8016/v1",
    "http://canvas-ns-1727666527880704.mng.ip.violet.uplus.co.kr/v1",
]

MAIN_INPUT = "input_202602_SS001643_ENTR_BY_STACC_001.xlsx"
EXPECTED_FILES = [
    MAIN_INPUT,
    "input_KGM월별정산_리스트_토레스.xlsx",
    "input_교체된 CCU 목록2026-02-12.xlsx",
    "input_테스트 차량 목록 - LG U  전달2026-03.xlsx",
    "output_02월 검증파일.xlsx",
    "output_03월 검증파일.xlsx",
]

# Representative final prompts from the old conversation. Some early prompts
# were corrected later, so the list uses the later corrected version where the
# old chat had both.
DEFAULT_PAIR_INDEXES = [1, 2, 3, 4, 9, 6, 8, 10, 16, 17, 18, 19, 22, 23, 24, 25, 26, 27]


class ThinkLoopDetected(RuntimeError):
    pass


def _safe_stdout() -> None:
    try:
        sys.stdout.reconfigure(encoding="utf-8")
        sys.stderr.reconfigure(encoding="utf-8")
    except Exception:
        pass


def _read_text(path: Path) -> str:
    return path.read_text(encoding="utf-8")


def find_data_dir() -> Path:
    downloads = Path.home() / "Downloads"
    for path in downloads.iterdir():
        if not path.is_dir():
            continue
        if all((path / name).exists() for name in EXPECTED_FILES):
            return path
    raise FileNotFoundError("Could not find the real workbook set under Downloads")


def find_logic_file() -> Path:
    downloads = Path.home() / "Downloads"
    matches = sorted(downloads.rglob("logic_2026-05-26-05-49.logic.json"))
    if not matches:
        raise FileNotFoundError("logic_2026-05-26-05-49.logic.json not found under Downloads")
    return matches[0]


def extract_prompt_pairs(logic_path: Path) -> list[dict]:
    data = json.loads(_read_text(logic_path))
    pairs = []
    last_user = ""
    for msg in data.get("chatHistory") or []:
        role = msg.get("role")
        content = msg.get("content") or ""
        if role == "user":
            last_user = content
            continue
        if role != "assistant":
            continue
        if "```" not in content:
            continue
        if not any(token in content for token in ("function transform", "def transform", "transform(")):
            continue
        title = ""
        for line in content.splitlines():
            s = line.strip()
            if s.startswith("제목:") or s.startswith("제목："):
                title = re.split(r"[:：]", s, maxsplit=1)[-1].strip()
                break
        pairs.append({"user": last_user, "oldTitle": title, "oldAssistant": content})
    return pairs


def selected_prompts(pairs: list[dict], indexes: list[int]) -> list[dict]:
    out = []
    for seq, idx in enumerate(indexes, start=1):
        if idx < 1 or idx > len(pairs):
            raise IndexError(f"prompt pair index out of range: {idx}")
        item = dict(pairs[idx - 1])
        item["pairIndex"] = idx
        item["seq"] = seq
        out.append(item)
    return out


def col_letter(n: int) -> str:
    s = ""
    while n:
        n, r = divmod(n - 1, 26)
        s = chr(65 + r) + s
    return s or "A"


def workbook_schema_preview(path: Path, max_sheets: int = 8, max_rows: int = 10, max_cols: int = 25) -> str:
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        out = [f"### {path.name}"]
        out.append(f"Sheets: {', '.join(wb.sheetnames[:max_sheets])}")
        for ws in wb.worksheets[:max_sheets]:
            out.append(f'- Sheet "{ws.title}": {ws.max_row} rows x {ws.max_column} cols')
            rows = []
            for i, row in enumerate(ws.iter_rows(max_row=max_rows, max_col=min(max_cols, ws.max_column), values_only=True), start=1):
                rows.append(row)
                nonempty = sum(1 for v in row if v not in (None, ""))
                values = ", ".join("" if v is None else str(v)[:42] for v in row)
                out.append(f"  row {i} nonempty={nonempty}: [{values}]")
            if rows:
                best_idx = max(range(len(rows)), key=lambda i: sum(1 for v in rows[i] if v not in (None, "")))
                header = rows[best_idx]
                header_map = [
                    f"{col_letter(i + 1)}={str(v)[:32]}"
                    for i, v in enumerate(header)
                    if v not in (None, "")
                ]
                if header_map:
                    out.append(f"  header candidate row {best_idx + 1}: " + " | ".join(header_map))
        return "\n".join(out)
    finally:
        wb.close()


def build_schema(data_dir: Path) -> str:
    parts = []
    for name in EXPECTED_FILES:
        if name.startswith("output_03"):
            continue
        parts.append(workbook_schema_preview(data_dir / name))
    return "\n\n".join(parts)


def build_system(schema: str) -> str:
    sys_head_path = ROOT / "_tmp_py_sys_head.txt"
    engine_note_path = ROOT / "_tmp_py_engine_note.txt"
    if sys_head_path.exists():
        head = _read_text(sys_head_path)
    else:
        head = (
            "Generate Python workbook automation. Return exactly one fenced ```python code block.\n"
            "Required signature: def transform(ctx): ...\n"
            "Use ctx.sheet() for the active workbook/sheet. Use ctx.input(file hint).sheet(sheet hint) for explicit input workbooks and ctx.workbook for explicit output workbook writes.\n"
            "ctx.rows(ws) returns plain values, not cell objects. Use enumerate or ctx.iter_rows for Excel row numbers.\n"
        )
    note = _read_text(engine_note_path) if engine_note_path.exists() else ""
    overrides = """

## 0.4.13 test harness safety overrides
- ctx.col(ws, "header") raises `RuntimeError("column not found: ...")` if the header is missing. Do not use negative fallback indexes.
- If the user references an output workbook path/name, use ctx.sheet(...) / ctx.workbook. Do not open the output workbook with ctx.input(...).
- New/intermediate sheets should follow the active/source workbook: ctx.sheet() defaults to the active file/sheet, and ctx.filter_to_sheet(ws, ...) / ctx.pivot(ws, ...) create their result in ws.Parent unless workbook=... is explicitly supplied.
- Never import win32com/win32com.client in generated skills. Excel COM fallback already provides COM-compatible ctx/workbook/worksheet/range objects.
- A schema preview is only a preview. For copy/sort/filter/pivot operations, use the actual used row/column bounds rather than the sampled rows shown in this prompt.
- If multiple aggregates are needed for the same group, create one combined grouped table instead of separate pivot sheets. Prefer `ctx.pivot(ws, group_by=..., value=[...], agg=[...])`.
- For a selected monthly block in the output summary, copy the full bounded block including title/header/blank separator/all data rows, not only the first visible data rows.
- Preserve compact Korean date formatting in replacements. Do not create new spaces such as "2026 년 03 월" or "03 월"; write "2026년 03월", "03월", and "3월".
- The main input workbook can be large. Read used rows once, compute in Python, and write grids or explicit cells only.
- Input workbooks may be modified during the pipeline, and modified input workbooks are saved as downloadable results. If a user asks to create/filter/sort an intermediate sheet in an input workbook, write it to that input workbook and let later steps read the same updated sheet.
- When a prompt includes a 3.7 JS reference implementation, treat it as the behavioral oracle. Python may use different sheet names only when the user did not specify them, but row filters, exact-vs-contains matching, sort-key order, deduplication rules, aggregation targets, and output cell values must match the reference.
"""
    return head + "\n\n## 현재 파일 스키마\n" + schema + note + overrides


def post_json(url: str, payload: dict, timeout: int) -> dict:
    raw = json.dumps(payload, ensure_ascii=False).encode("utf-8")
    req = urllib.request.Request(
        url,
        data=raw,
        method="POST",
        headers={
            "Content-Type": "application/json",
            "Api-Key": API_KEY,
            "Authorization": f"Bearer {API_KEY}",
        },
    )
    try:
        with urllib.request.urlopen(req, timeout=timeout) as resp:
            return json.loads(resp.read().decode("utf-8", errors="replace"))
    except urllib.error.HTTPError as err:
        body = err.read().decode("utf-8", errors="replace")
        raise RuntimeError(f"HTTP {err.code}: {body[:500]}") from err


def _looks_like_repeated_reasoning(text: str) -> bool:
    compact = re.sub(r"\s+", " ", str(text or "")).strip()
    if len(compact) < 8000:
        return False
    tail = compact[-3200:]
    for size in (160, 240, 320, 480):
        unit = tail[-size:]
        if len(unit.strip()) >= size * 0.8 and tail.count(unit) >= 3:
            return True
    chunks = [tail[i:i + 400] for i in range(0, len(tail), 400)]
    chunks = [c for c in chunks if len(c) >= 300]
    if len(chunks) >= 6 and len(set(chunks[-6:])) <= 2:
        return True
    return len(compact) >= 24000


def post_openai_stream(url: str, payload: dict, timeout: int, progress_prefix: str = "") -> tuple[str, str]:
    raw = json.dumps(payload, ensure_ascii=False).encode("utf-8")
    req = urllib.request.Request(
        url,
        data=raw,
        method="POST",
        headers={
            "Content-Type": "application/json",
            "Accept": "text/event-stream",
            "Api-Key": API_KEY,
            "Authorization": f"Bearer {API_KEY}",
        },
    )
    content = []
    reasoning = []
    last_print = 0.0
    last_kind = ""
    try:
        with urllib.request.urlopen(req, timeout=timeout) as resp:
            buffer = ""
            while True:
                chunk = resp.read(8192)
                if not chunk:
                    break
                buffer += chunk.decode("utf-8", errors="replace")
                while "\n" in buffer:
                    line, buffer = buffer.split("\n", 1)
                    line = line.strip()
                    if not line or line.startswith(":") or not line.startswith("data:"):
                        continue
                    data = line[5:].strip()
                    if data == "[DONE]":
                        if progress_prefix:
                            print()
                        return "".join(content), "".join(reasoning)
                    try:
                        parsed = json.loads(data)
                    except Exception:
                        continue
                    choice = (parsed.get("choices") or [{}])[0]
                    delta_obj = choice.get("delta") or {}
                    reasoning_delta = (
                        delta_obj.get("reasoning_content")
                        or delta_obj.get("reasoning")
                        or delta_obj.get("reasoningContent")
                        or choice.get("reasoning_content")
                        or choice.get("reasoning")
                        or ""
                    )
                    if reasoning_delta:
                        reasoning.append(reasoning_delta)
                        if _looks_like_repeated_reasoning("".join(reasoning)) and not "".join(content).strip():
                            raise ThinkLoopDetected("reasoning stream appears repetitive; retrying without think")
                        if progress_prefix:
                            now = time.time()
                            if now - last_print >= 0.25 or last_kind != "t":
                                print("t", end="", flush=True)
                                last_print = now
                                last_kind = "t"
                        continue
                    delta = (
                        delta_obj.get("content")
                        or parsed.get("text")
                        or choice.get("text")
                        or ""
                    )
                    if not delta:
                        continue
                    content.append(delta)
                    if progress_prefix:
                        now = time.time()
                        if now - last_print >= 0.25 or last_kind != ".":
                            print(".", end="", flush=True)
                            last_print = now
                            last_kind = "."
            if progress_prefix:
                print()
            return "".join(content), "".join(reasoning)
    except urllib.error.HTTPError as err:
        body = err.read().decode("utf-8", errors="replace")
        raise RuntimeError(f"HTTP {err.code}: {body[:500]}") from err


def _llm_payload(system: str, messages: list[dict], stream: bool, think: bool) -> dict:
    return {
        "model": MODEL,
        "messages": [{"role": "system", "content": system}, *messages],
        "max_tokens": 4096,
        "temperature": 0.2,
        "stream": bool(stream),
        "chat_template_kwargs": {"enable_thinking": bool(think)},
    }


def call_llm(system: str, messages: list[dict], bases: list[str], timeout: int, stream: bool = True, think: bool = False) -> tuple[str, str]:
    errors = []
    for base in bases:
        url = base.rstrip("/") + "/chat/completions"
        use_think = bool(think)
        retried_without_think = False
        try:
            while True:
                payload = _llm_payload(system, messages, stream=stream, think=use_think)
                try:
                    if stream:
                        content, reasoning = post_openai_stream(url, payload, timeout=timeout, progress_prefix=".")
                        if reasoning:
                            print(f"  [think streamed {len(reasoning)} chars]", flush=True)
                        if use_think and reasoning and not str(content or "").strip():
                            raise ThinkLoopDetected("reasoning ended without final content")
                    else:
                        data = post_json(url, payload, timeout=timeout)
                        content = (data.get("choices") or [{}])[0].get("message", {}).get("content") or ""
                    break
                except ThinkLoopDetected as err:
                    if not use_think or retried_without_think:
                        raise
                    print(f"\n  [think 반복 감지: {err}; no-think 재시도]", flush=True)
                    use_think = False
                    retried_without_think = True
            if not content:
                raise RuntimeError("empty LLM response")
            return content, url
        except Exception as err:
            errors.append(f"{url}: {err}")
    raise RuntimeError("No vLLM endpoint responded:\n" + "\n".join(errors))


def extract_python_code(reply: str) -> str:
    match = re.search(r"```(?:python|py)?\s*([\s\S]*?)```", reply)
    code = (match.group(1) if match else reply).strip()
    if not re.search(r"^\s*def\s+transform\s*\(\s*ctx\s*\)\s*:", code, re.M):
        raise RuntimeError("generated response does not define def transform(ctx)")
    ast.parse(code)
    return code


def generate_steps(prompts: list[dict], system: str, bases: list[str], timeout: int, run_dir: Path, stream: bool, think: bool) -> tuple[list[dict], list[dict]]:
    messages: list[dict] = []
    steps: list[dict] = []
    reports: list[dict] = []
    for item in prompts:
        user = item["user"]
        old_reference = item.get("oldAssistant") or ""
        if old_reference:
            user = (
                user
                + "\n\n## 3.7 JS 기준 구현(정답 기준)\n"
                + "아래 JavaScript는 같은 요청에 대해 3.7에서 실제로 사용되어 정답 엑셀을 만든 구현입니다. "
                + "시트명이 사용자가 명시하지 않은 경우 완전히 같을 필요는 없지만, 필터 조건, 정렬 키 순서, 중복제거 규칙, 집계 대상, 출력 셀 값은 반드시 같은 의미가 되도록 Python으로 변환하세요.\n"
                + old_reference
            )
        seq = item["seq"]
        messages.append({"role": "user", "content": user})
        print(f"[gen {seq:02d}] pair={item['pairIndex']} prompt={one_line(user)[:100]}")
        started = time.time()
        try:
            reply, endpoint = call_llm(system, messages, bases=bases, timeout=timeout, stream=stream, think=think)
            code = extract_python_code(reply)
            status = "ok"
            error = ""
        except Exception as err:
            reply = ""
            endpoint = ""
            code = ""
            status = "error"
            error = str(err)
            print(f"  -> ERROR {error[:300]}")
            reports.append({
                "seq": seq,
                "pairIndex": item["pairIndex"],
                "status": status,
                "error": error,
                "elapsedSec": round(time.time() - started, 2),
            })
            break
        messages.append({"role": "assistant", "content": reply})
        step_id = f"logic0413_s{seq:02d}"
        steps.append({
            "id": step_id,
            "description": f"logic pair {item['pairIndex']}",
            "prompt": user,
            "language": "python",
            "enabled": True,
            "code": code,
        })
        (run_dir / f"step_{seq:02d}_pair_{item['pairIndex']}.py").write_text(code, encoding="utf-8")
        (run_dir / f"reply_{seq:02d}_pair_{item['pairIndex']}.txt").write_text(reply, encoding="utf-8")
        reports.append({
            "seq": seq,
            "pairIndex": item["pairIndex"],
            "status": status,
            "endpoint": endpoint,
            "elapsedSec": round(time.time() - started, 2),
            "codeHead": "\n".join(code.splitlines()[:10]),
        })
        print(f"  -> ok {reports[-1]['elapsedSec']}s via {endpoint}")
    return steps, reports


def one_line(text: str) -> str:
    return " ".join(str(text or "").split())


def copy_workbooks(data_dir: Path, run_dir: Path) -> dict[str, Path]:
    workbooks_dir = run_dir / "workbooks"
    workbooks_dir.mkdir(parents=True, exist_ok=True)
    copied = {}
    for name in EXPECTED_FILES:
        src = data_dir / name
        dst = workbooks_dir / name
        shutil.copy2(src, dst)
        copied[name] = dst
    return copied


def safe_filename(text: str) -> str:
    name = re.sub(r'[<>:"/\\|?*\x00-\x1f]+', "_", str(text or "file")).strip()
    return name[:180] or "file"


def register_workbook(path: Path, name: str) -> str:
    wid = uuid.uuid4().hex
    serve_b2b.WORKBOOKS[wid] = {
        "id": wid,
        "name": name,
        "path": str(path),
        "created": time.time(),
        "aoa_cache": None,
        "current_aoa_cache": None,
        "aoa_cache_created": None,
        "aoa_cache_hits": 0,
    }
    return wid


def execute_pipeline(steps: list[dict], copied: dict[str, Path], run_dir: Path) -> dict:
    backend_dir = run_dir / "backend"
    backend_dir.mkdir(parents=True, exist_ok=True)
    serve_b2b.BACKEND_DIR = backend_dir
    serve_b2b.WORKBOOKS.clear()
    serve_b2b.RESULTS.clear()
    serve_b2b.PIPELINE_JOBS.clear()

    output_name = "output_02월 검증파일.xlsx"
    output_id = register_workbook(copied[output_name], output_name)
    input_items = []
    for name in EXPECTED_FILES:
        if not name.startswith("input_"):
            continue
        input_items.append({
            "backendWorkbookId": register_workbook(copied[name], name),
            "name": name,
        })
    payload = {
        "engine": "python",
        "inputs": input_items,
        "output": {"backendWorkbookId": output_id, "name": output_name},
        "pipeline": steps,
        "current": {"outputFileId": "output:0"},
    }
    started = time.time()
    stop_monitor = threading.Event()

    def monitor_job():
        last = ""
        while not stop_monitor.wait(15):
            job = serve_b2b.PIPELINE_JOBS.get("logic0413") or {}
            msg = (
                f"  [backend] stage={job.get('stage')!r} "
                f"step={job.get('currentStep')}/{job.get('totalSteps')} "
                f"elapsed={round(time.time() - started, 1)}s"
            )
            if msg != last:
                print(msg, flush=True)
                last = msg

    monitor = threading.Thread(target=monitor_job, daemon=True)
    monitor.start()
    try:
        result = serve_b2b.run_backend_pipeline_payload(payload, job_id="logic0413")
        stop_monitor.set()
        result["elapsedSec"] = round(time.time() - started, 2)
        rid = result.get("downloadId")
        if rid and rid in serve_b2b.RESULTS:
            result_path = Path(serve_b2b.RESULTS[rid]["path"])
            final_path = run_dir / "result_output_02월_검증파일.xlsx"
            shutil.copy2(result_path, final_path)
            result["savedResultPath"] = str(final_path)
        saved_downloads = {}
        for file_id, url in (result.get("downloadUrls") or {}).items():
            result_id = str(url).rstrip("/").rsplit("/", 1)[-1]
            rec = serve_b2b.RESULTS.get(result_id)
            if not rec or not rec.get("path"):
                continue
            out_name = safe_filename(f"{file_id.replace(':', '_')}_{rec.get('name') or result_id}")
            if not Path(out_name).suffix:
                out_name += ".xlsx"
            out_path = run_dir / out_name
            shutil.copy2(Path(rec["path"]), out_path)
            saved_downloads[file_id] = str(out_path)
        result["savedDownloadPaths"] = saved_downloads
        return result
    except serve_b2b.PipelineExecutionError as err:
        stop_monitor.set()
        return {
            "ok": False,
            "elapsedSec": round(time.time() - started, 2),
            "error": str(err),
            "errorInfo": err.info,
        }
    except Exception as err:
        stop_monitor.set()
        return {
            "ok": False,
            "elapsedSec": round(time.time() - started, 2),
            "error": str(err),
        }


def find_summary_sheet(wb) -> str:
    for name in wb.sheetnames:
        if "요약" in name or "중고차" in name:
            return name
    return wb.sheetnames[0]


def cell_values(ws, addresses: list[str]) -> dict[str, object]:
    out = {}
    for addr in addresses:
        out[addr] = ws[addr].value
    return out


def compare_result(result_path: Path, expected_path: Path) -> dict:
    if not result_path or not result_path.exists():
        return {"ok": False, "error": "result file missing"}
    addresses = [
        "A1", "G1",
        "B5", "C5", "D5", "B6", "C6", "D6", "B7", "C7", "D7", "B8", "C8", "D8",
        "B30", "C30", "D30", "B31", "C31", "D31", "B32", "C32", "D32",
        "B61", "C61", "D61", "B62", "C62", "D62", "B63", "C63", "D63",
        "B85", "C85", "D85",
        "B99", "C99", "D99", "B100", "C100", "D100", "B101", "C101", "D101",
    ]
    res_wb = load_workbook(result_path, data_only=True)
    exp_wb = load_workbook(expected_path, data_only=True)
    try:
        res_ws = res_wb[find_summary_sheet(res_wb)]
        exp_ws = exp_wb[find_summary_sheet(exp_wb)]
        got = cell_values(res_ws, addresses)
        exp = cell_values(exp_ws, addresses)
        mismatches = []
        for addr in addresses:
            if normalize_compare_value(got.get(addr)) != normalize_compare_value(exp.get(addr)):
                mismatches.append({"cell": addr, "got": got.get(addr), "expected": exp.get(addr)})
        return {
            "ok": len(mismatches) == 0,
            "resultSheet": res_ws.title,
            "expectedSheet": exp_ws.title,
            "resultShape": [res_ws.max_row, res_ws.max_column],
            "expectedShape": [exp_ws.max_row, exp_ws.max_column],
            "checkedCells": len(addresses),
            "mismatchCount": len(mismatches),
            "mismatches": mismatches[:80],
            "title": got.get("A1"),
        }
    finally:
        res_wb.close()
        exp_wb.close()


def normalize_compare_value(value):
    if isinstance(value, float):
        return round(value, 6)
    return value


def find_expected_result_file(expected_dir: Path, kind: str) -> Path | None:
    if not expected_dir or not expected_dir.exists():
        return None
    patterns = {
        "output": ["*output_02*.xlsx", "*검증파일*.xlsx"],
        "main_input": ["*202602_SS001643_ENTR_BY_STACC_001*.xlsx"],
    }.get(kind, ["*.xlsx"])
    for pattern in patterns:
        matches = sorted(expected_dir.glob(pattern))
        if matches:
            return matches[0]
    return None


def find_saved_download(result: dict, file_name_part: str) -> Path | None:
    for _, path_text in (result.get("savedDownloadPaths") or {}).items():
        path = Path(path_text)
        if file_name_part in path.name and path.exists():
            return path
    return None


def normalize_header(text) -> str:
    return re.sub(r"\s+", "", str(text or "")).lower()


def table_rows_by_first_col(path: Path, sheet_name: str) -> dict[str, tuple]:
    wb = load_workbook(path, data_only=True, read_only=True)
    try:
        if sheet_name not in wb.sheetnames:
            return {}
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        out = {}
        for row in rows[1:]:
            if not row or row[0] in (None, ""):
                continue
            out[str(row[0])] = tuple(row[:4])
        return out
    finally:
        wb.close()


def find_sheet_with_headers(path: Path, required_headers: list[str]) -> str | None:
    required = [normalize_header(x) for x in required_headers]
    wb = load_workbook(path, data_only=True, read_only=True)
    try:
        for ws in wb.worksheets:
            rows = list(ws.iter_rows(max_row=5, values_only=True))
            for row in rows:
                headers = {normalize_header(v) for v in (row or []) if v not in (None, "")}
                if all(any(req in h or h in req for h in headers) for req in required):
                    return ws.title
        return None
    finally:
        wb.close()


def find_mvno_summary_sheet_by_safe_count(path: Path, min_count: float, max_count: float) -> str | None:
    wb = load_workbook(path, data_only=True, read_only=True)
    try:
        for ws in wb.worksheets:
            rows = list(ws.iter_rows(max_row=20, max_col=8, values_only=True))
            if not rows:
                continue
            header = [normalize_header(v) for v in rows[0]]
            if not any("mvno" in h for h in header):
                continue
            if not any("건수" in h or "count" in h for h in header):
                continue
            for row in rows[1:]:
                if not row or str(row[0]) != "안전제일":
                    continue
                try:
                    count = float(row[1])
                except Exception:
                    continue
                if min_count <= count <= max_count:
                    return ws.title
        return None
    finally:
        wb.close()


def sheet_shape(path: Path, sheet_name: str) -> list[int] | None:
    wb = load_workbook(path, data_only=True, read_only=True)
    try:
        if sheet_name not in wb.sheetnames:
            return None
        ws = wb[sheet_name]
        return [ws.max_row, ws.max_column]
    finally:
        wb.close()


def compare_main_input(result_path: Path | None, expected_path: Path | None) -> dict:
    if not result_path or not result_path.exists():
        return {"ok": False, "error": "result main input file missing"}
    if not expected_path or not expected_path.exists():
        return {"ok": False, "error": "expected main input file missing"}

    expected_pivot = "피벗_MVNO상품명"
    expected_test = "테스트차량_MVNO집계"
    got_pivot = find_mvno_summary_sheet_by_safe_count(result_path, 100000, 200000)
    got_test = find_mvno_summary_sheet_by_safe_count(result_path, 500, 700)
    if not got_pivot:
        got_pivot = find_sheet_with_headers(result_path, ["MVNO상품명", "건수", "수납금액", "가입자당단가"])

    mismatches = []
    expected_tables = [
        ("pivot", expected_pivot, got_pivot),
        ("test_vehicle", expected_test, got_test),
    ]
    for label, exp_sheet, got_sheet in expected_tables:
        exp_rows = table_rows_by_first_col(expected_path, exp_sheet)
        got_rows = table_rows_by_first_col(result_path, got_sheet) if got_sheet else {}
        if not got_sheet:
            mismatches.append({"table": label, "error": "matching result sheet not found", "expectedSheet": exp_sheet})
            continue
        for key, exp_row in exp_rows.items():
            got_row = got_rows.get(key)
            if tuple(normalize_compare_value(v) for v in (got_row or ())) != tuple(normalize_compare_value(v) for v in exp_row):
                mismatches.append({"table": label, "key": key, "got": got_row, "expected": exp_row, "gotSheet": got_sheet, "expectedSheet": exp_sheet})

    return {
        "ok": not mismatches,
        "resultPath": str(result_path),
        "expectedPath": str(expected_path),
        "matchedSheets": {"pivot": got_pivot, "testVehicle": got_test},
        "expectedShapes": {
            "raw": sheet_shape(expected_path, "202602_SS001643_ENTR_BY_STACC_0"),
            "safe": sheet_shape(expected_path, "안전제일_정렬"),
            "aio": sheet_shape(expected_path, "인포콘올인원_정렬"),
        },
        "mismatchCount": len(mismatches),
        "mismatches": mismatches[:80],
    }


def parse_indexes(raw: str) -> list[int]:
    if not raw:
        return DEFAULT_PAIR_INDEXES
    indexes = []
    for part in raw.split(","):
        part = part.strip()
        if not part:
            continue
        if "-" in part:
            a, b = [int(x) for x in part.split("-", 1)]
            indexes.extend(range(a, b + 1))
        else:
            indexes.append(int(part))
    return indexes


def main(argv: list[str] | None = None) -> int:
    _safe_stdout()
    parser = argparse.ArgumentParser()
    parser.add_argument("--pairs", default="", help="1-based chat pair indexes, e.g. 1,2,3 or 1-18")
    parser.add_argument("--no-execute", action="store_true")
    parser.add_argument("--llm-base", action="append", default=[], help="OpenAI-compatible base URL ending in /v1")
    parser.add_argument("--llm-timeout", type=int, default=180)
    parser.add_argument("--no-stream", action="store_true", help="Disable OpenAI-compatible SSE streaming")
    parser.add_argument("--think", action="store_true", help="Request model reasoning; repetitive reasoning auto-retries without think")
    parser.add_argument("--reuse-steps", default="", help="Directory containing step_*.py files to execute without LLM generation")
    parser.add_argument("--max-steps", type=int, default=0, help="Execute only the first N generated/reused steps")
    parser.add_argument(
        "--expected-dir",
        type=Path,
        default=ROOT.parent / "B2B_ver3.7" / "3.7_result",
        help="Directory containing 3.7 baseline result xlsx files",
    )
    args = parser.parse_args(argv)

    data_dir = find_data_dir()
    logic_path = find_logic_file()
    stamp = _dt.datetime.now().strftime("%Y%m%d_%H%M%S")
    run_dir = ROOT / "test_runs" / f"logic0413_{stamp}"
    run_dir.mkdir(parents=True, exist_ok=True)

    pairs = extract_prompt_pairs(logic_path)
    prompts = selected_prompts(pairs, parse_indexes(args.pairs))
    (run_dir / "selected_prompts.json").write_text(json.dumps(prompts, ensure_ascii=False, indent=2), encoding="utf-8")

    print(f"data_dir={data_dir}")
    print(f"logic={logic_path}")
    print(f"run_dir={run_dir}")
    print(f"selected_pairs={[p['pairIndex'] for p in prompts]}")

    copied = copy_workbooks(data_dir, run_dir)
    steps = []
    generation_report = []
    if args.reuse_steps:
        reuse = Path(args.reuse_steps)
        for idx, path in enumerate(sorted(reuse.glob("step_*.py")), start=1):
            code = _read_text(path)
            ast.parse(code)
            steps.append({
                "id": f"reuse_{idx:02d}",
                "description": path.name,
                "language": "python",
                "enabled": True,
                "code": code,
            })
        print(f"reused {len(steps)} generated steps from {reuse}")
    else:
        schema = build_schema(data_dir)
        (run_dir / "schema_preview.txt").write_text(schema, encoding="utf-8")
        system = build_system(schema)
        (run_dir / "system_prompt.txt").write_text(system, encoding="utf-8")
        bases = args.llm_base or DEFAULT_LLM_BASES
        steps, generation_report = generate_steps(
            prompts,
            system,
            bases=bases,
            timeout=args.llm_timeout,
            run_dir=run_dir,
            stream=not args.no_stream,
            think=args.think,
        )
        (run_dir / "generation_report.json").write_text(json.dumps(generation_report, ensure_ascii=False, indent=2), encoding="utf-8")

    if args.max_steps and args.max_steps > 0:
        steps = steps[:args.max_steps]
        prompts = prompts[:args.max_steps]
        print(f"truncated to first {len(steps)} step(s)")

    if args.no_execute:
        print("generation only; execution skipped")
        return 0 if steps else 2
    if not steps or len(steps) != len(prompts):
        print("execution skipped because generation did not produce every selected step")
        return 2

    result = execute_pipeline(steps, copied, run_dir)
    (run_dir / "execution_result.json").write_text(json.dumps(result, ensure_ascii=False, indent=2, default=str), encoding="utf-8")
    print("execution:", "ok" if result.get("ok") else "error", "elapsed", result.get("elapsedSec"))
    if not result.get("ok"):
        print(json.dumps(result.get("errorInfo") or result, ensure_ascii=False, indent=2, default=str)[:2000])
        return 3

    result_path = Path(result.get("savedResultPath") or "")
    expected_output = find_expected_result_file(args.expected_dir, "output") or copied["output_03월 검증파일.xlsx"]
    compare = compare_result(result_path, expected_output)
    (run_dir / "comparison.json").write_text(json.dumps(compare, ensure_ascii=False, indent=2, default=str), encoding="utf-8")
    print("comparison:", "ok" if compare.get("ok") else f"{compare.get('mismatchCount')} mismatches")
    if compare.get("mismatches"):
        print(json.dumps(compare["mismatches"][:10], ensure_ascii=False, indent=2, default=str))

    expected_main = (
        find_expected_result_file(args.expected_dir, "main_input")
        or ROOT.parent / "B2B_ver3.7" / "test_runs" / "step9_check" / "results" / "result_input_202602_SS001643_ENTR_BY_STACC_001.xlsx"
    )
    result_main = find_saved_download(result, MAIN_INPUT)
    input_compare = compare_main_input(result_main, expected_main)
    (run_dir / "main_input_comparison.json").write_text(json.dumps(input_compare, ensure_ascii=False, indent=2, default=str), encoding="utf-8")
    print("main input comparison:", "ok" if input_compare.get("ok") else f"{input_compare.get('mismatchCount')} mismatches")
    if input_compare.get("mismatches"):
        print(json.dumps(input_compare["mismatches"][:10], ensure_ascii=False, indent=2, default=str))
    return 0 if compare.get("ok") and input_compare.get("ok") else 4


if __name__ == "__main__":
    raise SystemExit(main())
