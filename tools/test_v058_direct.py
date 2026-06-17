#!/usr/bin/env python3
"""Direct ver0.5.8 regression probe using the local OpenAI-compatible vLLM.

This script is intentionally not part of the packaged runtime. It:
- builds a compact schema from test_data/input_v058_regression_source.xlsx and
  test_data/output_v058_regression_target.xlsx
- asks the local vLLM to generate Python COM/VBA skills for the 0.5.8 scenarios
- statically checks the generated code for the patched routing/value-copy rules
- executes the generated code against copied workbooks through Excel COM when safe
- writes a JSON report under test_runs/v058_direct_*/
"""

from __future__ import annotations

import argparse
import datetime as dt
import json
import os
import re
import shutil
import sys
import time
import urllib.error
import urllib.request
from pathlib import Path
from typing import Any

import openpyxl

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))
TEST_DATA = ROOT / "test_data"
SOURCE_XLSX = TEST_DATA / "input_v058_regression_source.xlsx"
TARGET_XLSX = TEST_DATA / "output_v058_regression_target.xlsx"
DEFAULT_BASE_URL = "http://192.168.219.111:8000/v1"
DEFAULT_MODEL = "Qwen/Qwen3.6-27B-FP8"
DEFAULT_API_KEY = "khkim"


def read_text(path: Path) -> str:
    return path.read_text(encoding="utf-8-sig")


def extract_js_template_constant(source: str, name: str) -> str:
    needle = f"const {name} = `"
    start = source.find(needle)
    if start < 0:
        raise RuntimeError(f"{name} not found")
    i = start + len(needle)
    out: list[str] = []
    escaped = False
    while i < len(source):
        ch = source[i]
        if escaped:
            out.append("\\" + ch)
            escaped = False
        elif ch == "\\":
            escaped = True
        elif ch == "`":
            return "".join(out).replace("\\`", "`")
        else:
            out.append(ch)
        i += 1
    raise RuntimeError(f"{name} template is not closed")


def extract_code(text: str) -> str:
    m = re.search(r"```(?:python|py|vba|vb)?\s*\n([\s\S]*?)```", text or "", re.I)
    return (m.group(1) if m else text or "").strip()


def build_schema_summary() -> str:
    lines: list[str] = [
        "## 현재 업로드 파일/시트/컬럼 스키마",
        "",
        f"### 입력 파일: {SOURCE_XLSX.name}",
    ]
    wb = openpyxl.load_workbook(SOURCE_XLSX, data_only=False)
    for ws in wb.worksheets:
        lines.append(f'- 시트 "{ws.title}" ({ws.max_row}행 x {ws.max_column}열)')
        header = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
        lines.append(f"  - 1행 헤더: {header}")
        sample = []
        for r in range(2, min(ws.max_row, 4) + 1):
            sample.append([ws.cell(r, c).value for c in range(1, min(ws.max_column, 8) + 1)])
        if sample:
            lines.append(f"  - 샘플: {sample}")
        formulas = []
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and cell.value.startswith("="):
                    formulas.append(f"{cell.coordinate}={cell.value}")
        if formulas:
            lines.append(f"  - 수식 예: {formulas[:6]}")
    wb.close()

    lines += ["", f"### 출력/대상 파일: {TARGET_XLSX.name}"]
    wb = openpyxl.load_workbook(TARGET_XLSX, data_only=False)
    for ws in wb.worksheets:
        lines.append(f'- 시트 "{ws.title}" ({ws.max_row}행 x {ws.max_column}열)')
        header = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
        lines.append(f"  - 1행: {header}")
        formulas = []
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and cell.value.startswith("="):
                    formulas.append(f"{cell.coordinate}={cell.value}")
        if formulas:
            lines.append(f"  - 수식 예: {formulas[:8]}")
    wb.close()
    return "\n".join(lines)


def build_system_prompt(engine: str, schema: str) -> str:
    schema_js = read_text(ROOT / "scripts" / "file-schema.js")
    if engine == "python":
        prompt = extract_js_template_constant(schema_js, "PYTHON_COM_SYSTEM_PROMPT")
    elif engine == "vba":
        prompt = extract_js_template_constant(schema_js, "VBA_SYSTEM_PROMPT")
    else:
        raise ValueError(engine)
    return prompt + "\n\n" + schema


def call_vllm(base_url: str, api_key: str, model: str, system: str, user: str, timeout: int) -> str:
    payload = {
        "model": model,
        "messages": [
            {"role": "system", "content": system},
            {"role": "user", "content": user},
        ],
        "max_tokens": 4096,
        "temperature": 0.7,
        "top_p": 0.8,
        "top_k": 20,
        "presence_penalty": 0.5,
        "stream": False,
        "chat_template_kwargs": {"enable_thinking": False},
    }
    req = urllib.request.Request(
        base_url.rstrip("/") + "/chat/completions",
        data=json.dumps(payload, ensure_ascii=False).encode("utf-8"),
        method="POST",
        headers={
            "Content-Type": "application/json",
            "Api-Key": api_key,
            "Authorization": f"Bearer {api_key}",
        },
    )
    with urllib.request.urlopen(req, timeout=timeout) as resp:
        raw = resp.read().decode("utf-8", errors="replace")
    data = json.loads(raw)
    return data.get("choices", [{}])[0].get("message", {}).get("content", "") or ""


def static_python_values_only(code: str) -> list[str]:
    failures: list[str] = []
    if not re.search(r"def\s+transform\s*\(\s*ctx\s*\)\s*:", code):
        failures.append("missing def transform(ctx)")
    if re.search(r"\bctx\.copy\s*\(", code):
        failures.append("values-only request used ctx.copy")
    if not re.search(r"\b(?:ctx|\w+)\.(?:book\([^)]*\)\.)?read\s*\(", code):
        failures.append("values-only request did not read values")
    if not re.search(r"\b(?:ctx|\w+)\.(?:book\([^)]*\)\.)?write\s*\(", code):
        failures.append("values-only request did not write values")
    if "overwrite_formulas=True" not in code:
        failures.append("values-only write did not set overwrite_formulas=True")
    return failures


def static_vba_common(code: str) -> list[str]:
    failures: list[str] = []
    if not re.search(r"^\s*Sub\s+B2BSkill\s*\(\s*\)", code, re.I | re.M):
        failures.append("missing Sub B2BSkill()")
    if not re.search(r"\bEnd\s+Sub\b", code, re.I):
        failures.append("missing End Sub")
    blocked = [
        (r"\bOn\s+Error\s+Resume\s+Next\b", "On Error Resume Next"),
        (r"\bMsgBox\s*(?:\(|\s)", "MsgBox"),
        (r"\bInputBox\s*(?:\(|\s)", "InputBox"),
        (r"\bShell\s*(?:\(|\s)", "Shell"),
        (r"\bWorkbooks\s*\.\s*Open\b", "Workbooks.Open"),
        (r"\bApplication\s*\.\s*Quit\b", "Application.Quit"),
        (r"\.(?:Save|SaveAs|SaveCopyAs|Close)\b", "Save/Close"),
    ]
    for pattern, label in blocked:
        if re.search(pattern, code, re.I):
            failures.append(f"forbidden {label}")
    scalar_decls: dict[str, str] = {}
    for m in re.finditer(r"(?:\bDim|,)\s+([A-Za-z_][A-Za-z0-9_]*)\s+As\s+(String|Long|Integer|Double|Currency|Single|Date|Boolean)\b", code, re.I):
        scalar_decls[m.group(1).lower()] = m.group(2)
    for m in re.finditer(r"\bFor\s+Each\s+([A-Za-z_][A-Za-z0-9_]*)\s+In\b", code, re.I):
        var_name = m.group(1)
        scalar_type = scalar_decls.get(var_name.lower())
        if scalar_type:
            failures.append(f"For Each control variable {var_name} is declared As {scalar_type}; use Variant/Object")
    return failures


def static_vba_pivot_phone(code: str) -> list[str]:
    failures = static_vba_common(code)
    if not re.search(r"\.Text\b", code, re.I):
        failures.append("pivot caller key does not use .Text")
    if not re.search(r"NumberFormat\s*=\s*[\"']@", code, re.I):
        failures.append("pivot output key column does not set text NumberFormat")
    if re.search(r"\b(?:New\s+Collection|As\s+New\s+Collection|As\s+Collection)\b", code, re.I):
        failures.append("pivot unique labels use Collection; use Scripting.Dictionary.Exists instead")
    fmt = re.search(r"(?:Columns\s*\(\s*(?:1|[\"']A[\"'])\s*\)|Range\s*\(\s*[\"']A:A[\"']\s*\))\s*\.\s*NumberFormat\s*=\s*[\"']@", code, re.I)
    data_write = re.search(r"Cells\s*\(\s*(?:outRow|outR|rowNo|rIdx|kIdx\s*\+\s*\d+)\s*,\s*1\s*\)\s*\.\s*(?:Value|Value2)\s*=", code, re.I)
    if fmt and data_write and fmt.start() > data_write.start():
        failures.append("pivot output key column text NumberFormat is applied after writing data")
    return failures


def static_vba_sheet_copy(code: str) -> list[str]:
    failures = static_vba_common(code)
    if not re.search(r"\.Copy\b", code, re.I):
        failures.append("sheet-copy request did not use Worksheet.Copy")
    if not re.search(r"전체복사_복사본", code):
        failures.append("sheet-copy request did not name 전체복사_복사본")
    return failures


def static_vba_condition(code: str) -> list[str]:
    failures = static_vba_common(code)
    if re.search(r"\bval\w*\s+Is\s+Nothing\b", code, re.I):
        failures.append("cell value Variant uses Is Nothing; use IsEmpty/Len instead")
    uses_h = re.search(r"\bH\b|Cells\s*\([^,\)]*,\s*8\s*\)|Columns\s*\(\s*[\"']H", code, re.I) or "\uc11c\ube44\uc2a4" in code
    uses_q = re.search(r"\bQ\b|Cells\s*\([^,\)]*,\s*17\s*\)|Columns\s*\(\s*[\"']Q", code, re.I) or "\uc2dc\uac04" in code
    uses_s = re.search(r"\bS\b|Cells\s*\([^,\)]*,\s*19\s*\)|Columns\s*\(\s*[\"']S", code, re.I) or "\ucd08\ud658\uc0b0\uacb0\uacfc" in code
    if not uses_h:
        failures.append("condition code does not appear to use H column")
    if not uses_q:
        failures.append("condition code does not appear to use Q column")
    if not uses_s:
        failures.append("condition code does not appear to use S column")
    if not re.search(r"Split\s*\(|TimeValue\s*\(|DateDiff\s*\(|InStr\s*\([^)]*:", code, re.I):
        failures.append("time-to-seconds code does not handle hh:mm:ss colon text")
    if not re.search(r"86400|TimeValue\s*\(|DateDiff\s*\(", code, re.I):
        failures.append("time-to-seconds code does not handle Excel time serials")
    if re.search(r"Range\s*\([^)]*hCol[^)]*qCol[^)]*\)\.Value", code, re.I) and re.search(r"dataArr\s*\(\s*r\s*,\s*2\s*\)", code, re.I):
        failures.append("H:Q array uses dataArr(r,2) for Q; Q is the 10th relative column")
    return failures


def build_repair_prompt(original_prompt: str, code: str, failures: list[str], engine: str) -> str:
    lang = "python" if engine == "python" else "vba"
    extra = ""
    if engine == "vba":
        extra = "\n".join([
            "",
            "## VBA extra strict rules",
            "- If `On Error Resume Next` appears even once, the code fails. Do not suppress errors.",
            "- For duplicate removal, do not rely on Collection.Add errors; use `Scripting.Dictionary.Exists`.",
            "- Keep all `Dim` declarations near the top of Sub, not inside If/For/Else blocks.",
            "- `For Each item In dict.Keys` needs `item As Variant`; do not reuse a `String` variable as the loop control variable.",
            "- For time-to-seconds conversion, handle colon text like `01:02:03` with Split/InStr/TimeValue/DateDiff, and handle Excel serial time with `* 86400`.",
            "- Find workbooks/sheets explicitly with `For Each wb/sh`; raise Err.Raise when missing.",
        ])
    return "\n".join([
        extra,
        f"방금 생성한 {lang.upper()} 코드가 0.5.8 직접 테스트 정적 검사에서 막혔습니다.",
        "원래 요청을 그대로 만족하되 아래 실패 사유를 모두 고쳐 다시 작성하세요.",
        "",
        "## 원래 요청",
        original_prompt,
        "",
        "## 실패 사유",
        "\n".join(f"- {f}" for f in failures),
        "",
        "## 이전 코드",
        f"```{lang}",
        code,
        "```",
        "",
        f"반드시 하나의 ```{lang} 코드 블록만 출력하세요.",
        "/no_think",
    ])


def execute_excel(run_dir: Path, codes: dict[str, str]) -> dict[str, Any]:
    import win32com.client
    import serve_b2b

    src = run_dir / SOURCE_XLSX.name
    dst = run_dir / TARGET_XLSX.name
    app = win32com.client.DispatchEx("Excel.Application")
    app.Visible = False
    app.DisplayAlerts = False
    result: dict[str, Any] = {"executed": {}, "checks": {}}
    wb_src = wb_dst = None
    try:
        wb_src, _ = serve_b2b.excel_workbooks_open(app, src, read_only=False)
        wb_dst, _ = serve_b2b.excel_workbooks_open(app, dst, read_only=False)

        if codes.get("values_python"):
            try:
                serve_b2b._exec_python_com_skill(app, wb_dst, None, codes["values_python"])
                result["executed"]["values_python"] = "ok"
            except Exception as exc:  # noqa: BLE001
                result["executed"]["values_python"] = f"FAIL: {exc}"

        for key, wb in [
            ("condition_vba", wb_src),
            ("pivot_vba", wb_src),
            ("sheet_copy_vba", wb_dst),
        ]:
            if not codes.get(key):
                continue
            try:
                serve_b2b._inject_and_run_vba(app, wb, codes[key], "B2BSkill")
                result["executed"][key] = "ok"
            except Exception as exc:  # noqa: BLE001
                result["executed"][key] = f"FAIL: {exc}"

        try:
            app.CalculateFull()
        except Exception:
            app.Calculate()

        # Save a copy of the actually mutated workbooks for manual inspection.
        wb_src.Save()
        wb_dst.Save()

        ws = wb_dst.Worksheets("값복사_대상")
        vals = [ws.Range(f"H{r}").Value for r in range(2, 7)]
        formulas = [ws.Range(f"H{r}").Formula for r in range(2, 7)]
        result["checks"]["values_h2_h6"] = vals
        result["checks"]["values_h2_h6_formulas"] = formulas
        result["checks"]["values_pass"] = vals == [3600, 4000, 11000, 770, 9200] and all(
            not str(f).startswith("=") for f in formulas
        )

        ws = wb_src.Worksheets("조건문_원본")
        vals = [ws.Range(f"S{r}").Value for r in range(2, 9)]
        result["checks"]["condition_s2_s8"] = vals
        result["checks"]["condition_pass"] = vals == [3723, None, 45, None, 8110, None, 125]

        pivot_pass = False
        try:
            ws = None
            for _wb in (wb_src, wb_dst):
                try:
                    ws = _wb.Worksheets("피벗_결과")
                    break
                except Exception:
                    ws = None
            if ws is None:
                raise RuntimeError("피벗_결과 sheet was not created")
            first_col = [str(ws.Cells(r, 1).Text) for r in range(2, min(8, int(ws.UsedRange.Rows.Count)) + 1)]
            result["checks"]["pivot_first_col_text"] = first_col
            pivot_pass = any(v.startswith("010") for v in first_col) and any(v.startswith("020") for v in first_col)
        except Exception as exc:  # noqa: BLE001
            result["checks"]["pivot_error"] = str(exc)
        result["checks"]["pivot_pass"] = pivot_pass

        copy_pass = False
        try:
            ws = wb_dst.Worksheets("전체복사_복사본")
            copy_pass = bool(ws.Range("A1:D1").MergeCells) and str(ws.Range("D3").Formula).startswith("=")
            result["checks"]["sheet_copy_a1"] = str(ws.Range("A1").Text)
            result["checks"]["sheet_copy_d3_formula"] = str(ws.Range("D3").Formula)
        except Exception as exc:  # noqa: BLE001
            result["checks"]["sheet_copy_error"] = str(exc)
        result["checks"]["sheet_copy_pass"] = copy_pass
    finally:
        try:
            if wb_src is not None:
                wb_src.Close(SaveChanges=True)
        except Exception:
            pass
        try:
            if wb_dst is not None:
                wb_dst.Close(SaveChanges=True)
        except Exception:
            pass
        try:
            app.Quit()
        except Exception:
            pass
    return result


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--base-url", default=os.environ.get("B2B_QWEN_BASE_URL", DEFAULT_BASE_URL))
    ap.add_argument("--api-key", default=os.environ.get("B2B_QWEN_API_KEY", DEFAULT_API_KEY))
    ap.add_argument("--model", default=os.environ.get("B2B_QWEN_MODEL", DEFAULT_MODEL))
    ap.add_argument("--timeout", type=int, default=int(os.environ.get("B2B_QWEN_TIMEOUT", "180")))
    ap.add_argument("--no-excel", action="store_true")
    args = ap.parse_args()

    stamp = dt.datetime.now().strftime("%Y%m%d_%H%M%S")
    run_dir = ROOT / "test_runs" / f"v058_direct_{stamp}"
    run_dir.mkdir(parents=True, exist_ok=True)
    shutil.copy2(SOURCE_XLSX, run_dir / SOURCE_XLSX.name)
    shutil.copy2(TARGET_XLSX, run_dir / TARGET_XLSX.name)

    schema = build_schema_summary()
    py_system = build_system_prompt("python", schema)
    vba_system = build_system_prompt("vba", schema)
    scenarios = [
        {
            "id": "values_python",
            "engine": "python",
            "prompt": (
                "input_v058_regression_source.xlsx의 값복사_수식원본 시트 D2:D6 값을 "
                "output_v058_regression_target.xlsx의 값복사_대상 시트 H2:H6에 값만 복사해서 붙여넣어. "
                "수식은 복사하지 말고 계산된 값만 넣어."
            ),
            "check": static_python_values_only,
        },
        {
            "id": "condition_vba",
            "engine": "vba",
            "prompt": (
                'input_v058_regression_source.xlsx의 조건문_원본 시트에서 H열(서비스)이 "국제" 일 때 '
                "Q열(시간)을 초로 환산하여 S열 동일 행에 입력해줘."
            ),
            "check": static_vba_condition,
        },
        {
            "id": "pivot_vba",
            "engine": "vba",
            "prompt": (
                'input_v058_regression_source.xlsx의 피벗_발신번호 시트에서 발신번호를 행으로, 호유형을 열로, '
                '통화요금의 합계를 값으로 하는 유사 피벗을 새 시트 "피벗_결과"에 만들어줘. VBA로 해줘.'
            ),
            "check": static_vba_pivot_phone,
        },
        {
            "id": "sheet_copy_vba",
            "engine": "vba",
            "prompt": (
                'input_v058_regression_source.xlsx의 전체복사_원본 시트 전체를 '
                'output_v058_regression_target.xlsx에 새 시트 "전체복사_복사본"으로 복사해줘.'
            ),
            "check": static_vba_sheet_copy,
        },
    ]

    report: dict[str, Any] = {
        "runDir": str(run_dir),
        "baseUrl": args.base_url,
        "model": args.model,
        "scenarios": [],
    }
    codes: dict[str, str] = {}
    for sc in scenarios:
        system = py_system if sc["engine"] == "python" else vba_system
        print(f"[LLM] {sc['id']} ({sc['engine']})", flush=True)
        started = time.perf_counter()
        try:
            reply = call_vllm(args.base_url, args.api_key, args.model, system, sc["prompt"], args.timeout)
            code = extract_code(reply)
            (run_dir / f"{sc['id']}_reply.txt").write_text(reply, encoding="utf-8")
            (run_dir / f"{sc['id']}_code.txt").write_text(code, encoding="utf-8")
            failures = sc["check"](code)
            repair_used = False
            if failures:
                print(f"  repair static failures={failures}", flush=True)
                repair_prompt = build_repair_prompt(sc["prompt"], code, failures, sc["engine"])
                repair_reply = call_vllm(args.base_url, args.api_key, args.model, system, repair_prompt, args.timeout)
                repair_code = extract_code(repair_reply)
                repair_failures = sc["check"](repair_code)
                (run_dir / f"{sc['id']}_repair_reply.txt").write_text(repair_reply, encoding="utf-8")
                (run_dir / f"{sc['id']}_repair_code.txt").write_text(repair_code, encoding="utf-8")
                if len(repair_failures) <= len(failures):
                    reply = repair_reply
                    code = repair_code
                    failures = repair_failures
                    repair_used = True
            codes[sc["id"]] = code if not failures else ""
            item = {
                "id": sc["id"],
                "engine": sc["engine"],
                "elapsedSec": round(time.perf_counter() - started, 2),
                "codeChars": len(code),
                "repairUsed": repair_used,
                "staticFailures": failures,
                "staticPass": not failures,
            }
            print(f"  staticPass={item['staticPass']} failures={failures}", flush=True)
        except Exception as exc:  # noqa: BLE001
            item = {
                "id": sc["id"],
                "engine": sc["engine"],
                "elapsedSec": round(time.perf_counter() - started, 2),
                "error": repr(exc),
                "staticPass": False,
            }
            print(f"  ERROR {exc!r}", flush=True)
        report["scenarios"].append(item)

    if not args.no_excel:
        print("[EXCEL] executing statically-passing generated code", flush=True)
        try:
            report["excel"] = execute_excel(run_dir, codes)
            print(json.dumps(report["excel"], ensure_ascii=False, indent=2), flush=True)
        except Exception as exc:  # noqa: BLE001
            report["excel"] = {"error": repr(exc)}
            print(f"  EXCEL ERROR {exc!r}", flush=True)

    report_path = run_dir / "report.json"
    report_path.write_text(json.dumps(report, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"[REPORT] {report_path}", flush=True)

    all_static = all(x.get("staticPass") for x in report["scenarios"])
    excel_obj = report.get("excel") or {}
    excel_checks = excel_obj.get("checks") or {}
    all_excel = args.no_excel or all(
        excel_checks.get(k) is True
        for k in ["values_pass", "condition_pass", "pivot_pass", "sheet_copy_pass"]
        if k in excel_checks
    )
    return 0 if all_static and all_excel else 1


if __name__ == "__main__":
    raise SystemExit(main())
