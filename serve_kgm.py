#!/usr/bin/env python3
import http.server
import json
import os
from pathlib import Path
import shutil
import socketserver
import subprocess
import tempfile
import threading
import time
from urllib.parse import parse_qs, quote, unquote, urlparse
import urllib.error
import urllib.request
import uuid

try:
    import openpyxl
except Exception:
    openpyxl = None


HOST = os.environ.get("KGM_HOST", "127.0.0.1")
PORT = int(os.environ.get("KGM_PORT", "8090"))
VLLM_BASE = os.environ.get(
    "KGM_VLLM_BASE",
    "http://canvas-ns-1727666527880704.mng.ip.violet.uplus.co.kr",
).rstrip("/")
BACKEND_DIR = Path(tempfile.gettempdir()) / "kgm_b2b_backend_v37"
WORKBOOKS = {}
RESULTS = {}
DIFFS = {}
PIPELINE_JOBS = {}
PIPELINE_JOBS_LOCK = threading.Lock()
WORKBOOK_CACHE_LOCK = threading.Lock()
NODE_WORKER_LOCK = threading.Lock()
NODE_WORKER = None
NODE_WORKER_SCRIPT_MTIME = None
NODE_WORKER_READY = set()
PREVIEW_ROWS = 500
PREVIEW_COLS = None
MAX_DIFF_CELLS_PER_SHEET = 5000
APP_BUILD_STAMP = "run-adapter-20260526-4"


class PipelineExecutionError(RuntimeError):
    def __init__(self, message, info=None):
        super().__init__(message)
        self.info = info or {}


class KGMHandler(http.server.SimpleHTTPRequestHandler):
    def log_message(self, format, *args):
        if os.environ.get("KGM_LOG_REQUESTS") == "1":
            super().log_message(format, *args)

    def end_headers(self):
        self.send_header("Access-Control-Allow-Origin", "*")
        self.send_header("Access-Control-Allow-Methods", "GET, POST, OPTIONS")
        self.send_header("Access-Control-Allow-Headers", "authorization, content-type, api-key, x-api-key")
        self.send_header("Access-Control-Allow-Private-Network", "true")
        self.send_header("Cache-Control", "no-store")
        super().end_headers()

    def do_OPTIONS(self):
        self.send_response(200)
        self.end_headers()

    def do_GET(self):
        if self.path == "/api/backend/health":
            app_dir = Path(__file__).resolve().parent
            def file_info(relative_path):
                path = app_dir / relative_path
                if not path.exists():
                    return None
                return {
                    "path": str(path),
                    "mtime": path.stat().st_mtime,
                }
            self.send_json({
                "ok": True,
                "mode": "python-backend-workbooks",
                "buildStamp": APP_BUILD_STAMP,
                "pid": os.getpid(),
                "cwd": os.getcwd(),
                "serverFile": str(Path(__file__).resolve()),
                "appDir": str(app_dir),
                "openpyxl": bool(openpyxl),
                "node": bool(shutil.which("node")),
                "files": {
                    "index.html": file_info("index.html"),
                    "scripts/config.js": file_info("scripts/config.js"),
                    "scripts/excel-viewer.js": file_info("scripts/excel-viewer.js"),
                    "scripts/backend-workbooks.js": file_info("scripts/backend-workbooks.js"),
                    "scripts/backend-pipeline-worker.js": file_info("scripts/backend-pipeline-worker.js"),
                },
            })
            return
        if self.path.startswith("/api/workbooks/download/"):
            self.handle_backend_download()
            return
        if self.path.startswith("/api/pipeline/status/"):
            self.handle_pipeline_status()
            return
        if self.path.startswith("/api/diff/"):
            self.handle_cached_diff()
            return
        if self.path.startswith("/v1/"):
            self.proxy()
            return
        super().do_GET()

    def do_POST(self):
        if self.path.startswith("/api/workbooks/upload"):
            self.handle_workbook_upload()
            return
        if self.path == "/api/pipeline/run":
            self.handle_backend_pipeline_run()
            return
        if self.path == "/api/pipeline/start":
            self.handle_backend_pipeline_start()
            return
        if self.path == "/api/diff/current-view":
            self.handle_current_view_diff()
            return
        if self.path.startswith("/v1/"):
            self.proxy()
            return
        self.send_error(404)

    def read_json_body(self):
        length = int(self.headers.get("content-length") or 0)
        if not length:
            return {}
        raw = self.rfile.read(length)
        return json.loads(raw.decode("utf-8"))

    def send_json(self, payload, status=200):
        body = json.dumps(payload, ensure_ascii=False).encode("utf-8")
        self.send_response(status)
        self.send_header("content-type", "application/json; charset=utf-8")
        self.send_header("content-length", str(len(body)))
        self.end_headers()
        self.wfile.write(body)

    def handle_current_view_diff(self):
        try:
            payload = self.read_json_body()
            before = payload.get("before") or {}
            after = payload.get("after") or {}
            before_cells = {
                f"{cell.get('r')}:{cell.get('c')}": cell.get("value")
                for cell in before.get("cells", [])
            }
            changes = []
            for cell in after.get("cells", []):
                key = f"{cell.get('r')}:{cell.get('c')}"
                value = cell.get("value")
                if before_cells.get(key) != value:
                    changes.append({
                        "r": cell.get("r"),
                        "c": cell.get("c"),
                        "value": value,
                    })
            self.send_json({
                "ok": True,
                "fileId": after.get("fileId"),
                "sheet": after.get("sheet"),
                "changes": changes,
            })
        except Exception as err:
            self.send_json({"ok": False, "error": str(err)}, status=400)

    def handle_cached_diff(self):
        diff_id = self.path.rsplit("/", 1)[-1]
        diff = DIFFS.get(diff_id)
        if not diff:
            self.send_json({"ok": False, "error": "diff not found"}, status=404)
            return
        self.send_json({"ok": True, **diff})

    def handle_workbook_upload(self):
        if openpyxl is None:
            self.send_json({"ok": False, "error": "openpyxl is not available"}, status=500)
            return
        parsed = urlparse(self.path)
        qs = parse_qs(parsed.query)
        raw_name = qs.get("name", ["workbook.xlsx"])[0]
        name = Path(unquote(raw_name)).name or "workbook.xlsx"
        length = int(self.headers.get("content-length") or 0)
        if length <= 0:
            self.send_json({"ok": False, "error": "empty upload"}, status=400)
            return
        BACKEND_DIR.mkdir(parents=True, exist_ok=True)
        workbook_id = uuid.uuid4().hex
        path = BACKEND_DIR / f"{workbook_id}_{name}"
        with path.open("wb") as f:
            remaining = length
            while remaining > 0:
                chunk = self.rfile.read(min(1024 * 1024, remaining))
                if not chunk:
                    break
                f.write(chunk)
                remaining -= len(chunk)
        meta = inspect_workbook(path)
        WORKBOOKS[workbook_id] = {
            "id": workbook_id,
            "name": name,
            "path": str(path),
            "created": time.time(),
            "aoa_cache": None,
            "current_aoa_cache": None,
            "aoa_cache_created": None,
            "aoa_cache_hits": 0,
        }
        self.send_json({"ok": True, "workbookId": workbook_id, "name": name, "meta": meta})

    def handle_backend_pipeline_run(self):
        if openpyxl is None:
            self.send_json({"ok": False, "error": "openpyxl is not available"}, status=500)
            return
        if not shutil.which("node"):
            self.send_json({"ok": False, "error": "node runtime is not available"}, status=500)
            return
        payload = self.read_json_body()
        try:
            self.send_json(run_backend_pipeline_payload(payload))
        except Exception as err:
            self.send_json({"ok": False, "error": str(err)}, status=400)

    def handle_backend_pipeline_start(self):
        payload = self.read_json_body()
        job_id = uuid.uuid4().hex
        total_steps = len([s for s in payload.get("pipeline", []) if not (s and s.get("enabled") is False)])
        update_pipeline_job(job_id, {
            "ok": True,
            "jobId": job_id,
            "status": "running",
            "stage": "준비 중",
            "currentStep": 0,
            "totalSteps": total_steps,
            "created": time.time(),
        })

        def worker():
            try:
                result = run_backend_pipeline_payload(payload, job_id=job_id)
                result.update({"ok": True, "jobId": job_id, "status": "done", "stage": "완료"})
                update_pipeline_job(job_id, result)
            except PipelineExecutionError as err:
                update_pipeline_job(job_id, {
                    "ok": False,
                    "jobId": job_id,
                    "status": "error",
                    "stage": "?ㅻ쪟",
                    "error": str(err),
                    "errorInfo": err.info,
                })
            except Exception as err:
                update_pipeline_job(job_id, {
                    "ok": False,
                    "jobId": job_id,
                    "status": "error",
                    "stage": "오류",
                    "error": str(err),
                })

        threading.Thread(target=worker, name=f"kgm-pipeline-{job_id[:8]}", daemon=True).start()
        self.send_json({"ok": True, "jobId": job_id, "status": "running"})

    def handle_pipeline_status(self):
        job_id = self.path.rsplit("/", 1)[-1]
        with PIPELINE_JOBS_LOCK:
            job = dict(PIPELINE_JOBS.get(job_id) or {})
        if not job:
            self.send_json({"ok": False, "error": "pipeline job not found"}, status=404)
            return
        self.send_json(job)

    def handle_backend_download(self):
        result_id = self.path.rsplit("/", 1)[-1]
        result = RESULTS.get(result_id)
        if not result:
            self.send_error(404)
            return
        if "workerWorkbookId" in result and "sheets" not in result:
            result["sheets"] = export_node_worker_workbook(result["workerWorkbookId"])
        if "path" not in result:
            result_path = BACKEND_DIR / f"{result_id}_{result.get('name') or 'result.xlsx'}"
            write_result_workbook(
                Path(result["template_path"]),
                result_path,
                result.get("sheets") or {},
                result.get("forced_value_cells") or [],
            )
            result["path"] = str(result_path)
        path = Path(result["path"])
        if not path.exists():
            self.send_error(404)
            return
        data = path.read_bytes()
        self.send_response(200)
        self.send_header("content-type", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        self.send_header("content-disposition", content_disposition_attachment(path.name))
        self.send_header("content-length", str(len(data)))
        self.end_headers()
        self.wfile.write(data)

    def proxy(self):
        body = None
        if self.command in {"POST", "PUT", "PATCH"}:
            length = int(self.headers.get("content-length") or 0)
            body = self.rfile.read(length) if length else None

        target = VLLM_BASE + self.path
        headers = {}
        for key in ("authorization", "api-key", "content-type", "accept"):
            value = self.headers.get(key)
            if value:
                headers[key] = value
        if body is not None and "content-type" not in headers:
            headers["content-type"] = "application/json"

        req = urllib.request.Request(
            target,
            data=body,
            headers=headers,
            method=self.command,
        )
        try:
            with urllib.request.urlopen(req, timeout=300) as resp:
                self.send_response(resp.status)
                for key, value in resp.headers.items():
                    if key.lower() in {"connection", "transfer-encoding", "content-encoding", "content-length"}:
                        continue
                    self.send_header(key, value)
                self.end_headers()
                while True:
                    chunk = resp.read(8192)
                    if not chunk:
                        break
                    self.wfile.write(chunk)
                    self.wfile.flush()
        except urllib.error.HTTPError as err:
            payload = err.read()
            self.send_response(err.code)
            self.send_header("content-type", err.headers.get("content-type", "text/plain"))
            self.end_headers()
            self.wfile.write(payload)
        except Exception as err:
            payload = f"Proxy error: {err}".encode("utf-8")
            self.send_response(502)
            self.send_header("content-type", "text/plain; charset=utf-8")
            self.send_header("content-length", str(len(payload)))
            self.end_headers()
            self.wfile.write(payload)


def update_pipeline_job(job_id, patch):
    if not job_id:
        return
    with PIPELINE_JOBS_LOCK:
        current = PIPELINE_JOBS.get(job_id, {})
        current.update(patch)
        current["updated"] = time.time()
        PIPELINE_JOBS[job_id] = current


def content_disposition_attachment(filename):
    safe_ascii = "".join(ch if 32 <= ord(ch) < 127 and ch not in '"\\' else "_" for ch in str(filename))
    encoded = quote(str(filename), safe="")
    return f'attachment; filename="{safe_ascii}"; filename*=UTF-8\'\'{encoded}'


def run_backend_pipeline_payload(payload, job_id=None):
    update_pipeline_job(job_id, {"stage": "입력 파일 읽는 중", "currentStep": 0})
    inputs = {}
    for idx, item in enumerate(payload.get("inputs", []), start=1):
        wb = get_workbook_or_raise(item.get("backendWorkbookId"))
        update_pipeline_job(job_id, {"stage": f"입력 파일 읽는 중 ({idx}/{len(payload.get('inputs', []))})"})
        inputs[item.get("name") or wb["name"]] = get_workbook_aoa_for_run(wb)

    update_pipeline_job(job_id, {"stage": "출력 템플릿 읽는 중"})
    output_item = payload.get("output") or {}
    output_wb = get_workbook_or_raise(output_item.get("backendWorkbookId")) if output_item.get("backendWorkbookId") else None
    output = get_workbook_aoa_for_run(output_wb) if output_wb else {}

    total_steps = len([s for s in payload.get("pipeline", []) if not (s and s.get("enabled") is False)])
    update_pipeline_job(job_id, {"stage": "스킬 실행 중", "currentStep": 0, "totalSteps": total_steps})
    result = run_js_pipeline_with_node({
        "inputs": inputs,
        "output": output,
        "pipeline": payload.get("pipeline", []),
    }, job_id=job_id)
    result_inputs = result.get("inputs") or inputs
    result_output = result.get("output") or output
    forced_value_cells = result.get("forcedValueCells") or []
    current = payload.get("current") or {}

    update_pipeline_job(job_id, {
        "stage": "diff 계산 중",
        "currentStep": total_steps,
        "completedSteps": total_steps,
        "totalSteps": total_steps,
        "stepRunning": False,
    })
    diffs = build_pipeline_diffs(inputs, output, result_inputs, result_output, current)
    diff_id = uuid.uuid4().hex
    DIFFS[diff_id] = {
        "id": diff_id,
        "created": time.time(),
        "diffs": diffs,
        "current": current,
    }

    update_pipeline_job(job_id, {"stage": "다운로드 준비 중", "currentStep": total_steps})
    download_urls = {}
    input_items = payload.get("inputs", [])
    for item in input_items:
        wb = get_workbook_or_raise(item.get("backendWorkbookId"))
        input_name = item.get("name") or wb["name"]
        if input_name not in result_inputs:
            continue
        input_download_id = uuid.uuid4().hex
        RESULTS[input_download_id] = {
            "template_path": str(wb["path"]),
            "sheets": result_inputs[input_name],
            "forced_value_cells": [cell for cell in forced_value_cells if cell.get("fileId") == "input:" + input_name],
            "name": f"result_{wb['name']}",
            "created": time.time(),
        }
        download_urls["input:" + input_name] = f"/api/workbooks/download/{input_download_id}"

    download_id = None
    if output_wb:
        output_file_id = (payload.get("current") or {}).get("outputFileId") or "output:0"
        download_id = uuid.uuid4().hex
        RESULTS[download_id] = {
            "template_path": str(output_wb["path"]),
            "sheets": result_output,
            "forced_value_cells": [cell for cell in forced_value_cells if cell.get("fileId") == output_file_id],
            "name": f"result_{output_wb['name']}",
            "created": time.time(),
        }
        download_urls[output_file_id] = f"/api/workbooks/download/{download_id}"

    update_pipeline_job(job_id, {"stage": "미리보기 생성 중"})
    previews = build_result_previews(result_inputs, result_output, current, diffs, forced_value_cells)
    return {
        "ok": True,
        "diffId": diff_id,
        "diffs": diffs,
        "forcedValueCells": forced_value_cells,
        "downloadId": download_id,
        "downloadUrl": f"/api/workbooks/download/{download_id}" if download_id else None,
        "downloadUrls": download_urls,
        "files": previews,
    }


def inspect_workbook(path):
    wb = openpyxl.load_workbook(path, read_only=True, data_only=False)
    cached_wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        sheets = {}
        for ws in wb.worksheets:
            cached_ws = cached_wb[ws.title] if ws.title in cached_wb.sheetnames else None
            cached_rows = cached_ws.iter_rows(max_row=PREVIEW_ROWS, max_col=PREVIEW_COLS) if cached_ws else None
            rows = []
            formulas = {}
            original_formula_values = {}
            formats = []
            for row_idx, row in enumerate(ws.iter_rows(max_row=PREVIEW_ROWS, max_col=PREVIEW_COLS), start=1):
                try:
                    cached_row = next(cached_rows) if cached_rows else []
                except StopIteration:
                    cached_row = []
                values = []
                format_row = []
                for cell_idx, cell in enumerate(row):
                    if cell.data_type == "f":
                        cached_value = cached_row[cell_idx].value if cell_idx < len(cached_row) else None
                        json_cached = cell_to_json(cached_value)
                        values.append(json_cached if json_cached is not None else "")
                        formulas[cell.coordinate] = cell.value if str(cell.value).startswith("=") else "=" + str(cell.value)
                        if json_cached is not None:
                            original_formula_values[cell.coordinate] = json_cached
                    else:
                        values.append(cell_to_json(cell.value))
                    format_row.append(cell.number_format if cell.number_format else "")
                rows.append(values)
                formats.append(format_row)
            sheets[ws.title] = {
                "rows": rows,
                "formulas": formulas,
                "originalFormulaValues": original_formula_values,
                "formats": formats,
                "maxRow": ws.max_row or len(rows),
                "maxCol": ws.max_column or (max((len(r) for r in rows), default=0)),
            }
        return {"sheetNames": wb.sheetnames, "sheets": sheets}
    finally:
        wb.close()
        cached_wb.close()


def load_workbook_aoa(path):
    wb = openpyxl.load_workbook(path, read_only=True, data_only=False)
    try:
        data = {}
        for ws in wb.worksheets:
            rows = []
            for row in ws.iter_rows():
                values = ["" if cell.data_type == "f" else cell_to_json(cell.value) for cell in row]
                while values and values[-1] in ("", None):
                    values.pop()
                rows.append(values)
            data[ws.title] = rows
        return data
    finally:
        wb.close()


def get_workbook_aoa_for_run(wb_record, base_mode="original"):
    with WORKBOOK_CACHE_LOCK:
        if base_mode == "current" and wb_record.get("current_aoa_cache") is not None:
            wb_record["aoa_cache_hits"] = int(wb_record.get("aoa_cache_hits") or 0) + 1
            return wb_record["current_aoa_cache"]
        cached = wb_record.get("aoa_cache")
        if cached is not None:
            wb_record["aoa_cache_hits"] = int(wb_record.get("aoa_cache_hits") or 0) + 1
            return cached

    if base_mode == "current" and wb_record.get("node_current"):
        sheets = export_node_worker_workbook(wb_record["id"])
        with WORKBOOK_CACHE_LOCK:
            wb_record["current_aoa_cache"] = sheets
            wb_record["aoa_cache_hits"] = int(wb_record.get("aoa_cache_hits") or 0) + 1
        return sheets

    loaded = load_workbook_aoa(Path(wb_record["path"]))
    with WORKBOOK_CACHE_LOCK:
        if wb_record.get("aoa_cache") is None:
            wb_record["aoa_cache"] = loaded
            wb_record["aoa_cache_created"] = time.time()
        else:
            loaded = wb_record["aoa_cache"]
        wb_record["aoa_cache_hits"] = int(wb_record.get("aoa_cache_hits") or 0) + 1
        return loaded


def update_workbook_current_cache(wb_record, sheets):
    if not wb_record or sheets is None:
        return
    with WORKBOOK_CACHE_LOCK:
        wb_record["current_aoa_cache"] = sheets
        wb_record["current_aoa_cache_created"] = time.time()


def write_result_workbook(template_path, result_path, sheets, forced_value_cells=None):
    forced_cells = {
        (str(cell.get("sheetName") or ""), int(cell.get("r") or 0) + 1, int(cell.get("c") or 0) + 1)
        for cell in (forced_value_cells or [])
        if isinstance(cell, dict) and cell.get("sheetName")
    }
    wb = openpyxl.load_workbook(template_path)
    try:
        for sheet_name, rows in (sheets or {}).items():
            ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb.create_sheet(sheet_name)
            for r_idx, row in enumerate(rows or [], start=1):
                for c_idx, value in enumerate(row or [], start=1):
                    cell = ws.cell(row=r_idx, column=c_idx)
                    force_value = (sheet_name, r_idx, c_idx) in forced_cells
                    if cell.data_type == "f" and (value == "" or value is None) and not force_value:
                        continue
                    cell.value = value
        wb.save(result_path)
    finally:
        wb.close()


def cell_to_json(value):
    if value is None:
        return ""
    if hasattr(value, "isoformat"):
        return value.isoformat()
    return value


def get_workbook_or_raise(workbook_id):
    wb = WORKBOOKS.get(workbook_id or "")
    if not wb:
        raise ValueError(f"backend workbook not found: {workbook_id}")
    return wb


def run_js_pipeline_with_node(payload, job_id=None):
    progress_path = None
    if job_id:
        progress_file = tempfile.NamedTemporaryFile(prefix=f"kgm_pipeline_{job_id}_", suffix=".json", delete=False)
        progress_path = progress_file.name
        progress_file.close()
        payload = dict(payload)
        payload["progressPath"] = progress_path
    runner = r"""
const fs = require("fs");
const payload = JSON.parse(fs.readFileSync(0, "utf8"));
const progressPath = payload.progressPath || "";
function writeProgress(info) {
  if (!progressPath) return;
  try {
    fs.writeFileSync(progressPath, JSON.stringify(info), "utf8");
  } catch (err) {}
}
const inputs = payload.inputs || {};
const output = payload.output || {};
function normalizeText(v) { return String(v ?? "").trim().toLowerCase().replace(/\s+/g, ""); }
function includesNormalizedText(v, s) { return normalizeText(v).includes(normalizeText(s)); }
function equalsNormalizedText(v, s) { return normalizeText(v) === normalizeText(s); }
function replaceNormalizedText(v) { return String(v ?? ""); }
function similarity(a, b) { a = normalizeText(a); b = normalizeText(b); if (!a || !b) return 0; return a === b ? 1 : (a.includes(b) || b.includes(a) ? 0.8 : 0); }
function headerRowIndex(sheetAoA) {
  let best = 0, bestScore = -1;
  for (let r = 0; r < Math.min((sheetAoA || []).length, 30); r++) {
    const row = sheetAoA[r] || [];
    const score = row.filter(v => String(v ?? "").trim()).length;
    if (score > bestScore) { best = r; bestScore = score; }
  }
  return best;
}
function dataStartRowIndex(sheetAoA) { return headerRowIndex(sheetAoA) + 1; }
function excelRowToIndex(n) { return Math.max(0, Number(n) - 1); }
function col(sheetAoA, name) {
  const h = headerRowIndex(sheetAoA);
  const row = sheetAoA[h] || [];
  const target = normalizeText(name);
  let fallback = -1;
  for (let i = 0; i < row.length; i++) {
    const cur = normalizeText(row[i]);
    if (cur === target) return i;
    if (fallback < 0 && cur && (cur.includes(target) || target.includes(cur))) fallback = i;
  }
  return fallback;
}
function findColumnGlobal(inputsMap, name) {
  const hits = [];
  Object.entries(inputsMap || {}).forEach(([file, sheets]) => {
    Object.entries(sheets || {}).forEach(([sheet, aoa]) => {
      const colIdx = col(aoa, name);
      if (colIdx >= 0) hits.push({ file, sheet, colIdx });
    });
  });
  return hits;
}
function fuzzyGetKey(target, prop) {
  if (!target || typeof target !== "object") return null;
  if (Object.prototype.hasOwnProperty.call(target, prop)) return prop;
  const keys = Object.keys(target);
  if (keys.length === 1) return keys[0];
  const wanted = normalizeText(prop);
  let best = null;
  for (const key of keys) {
    const cur = normalizeText(key);
    if (cur === wanted || (cur && wanted && (cur.includes(wanted) || wanted.includes(cur)))) {
      best = key;
      break;
    }
  }
  return best;
}
function fuzzyProxy(target) {
  if (!target || typeof target !== "object") return target;
  return new Proxy(target, {
    get(t, prop) {
      if (typeof prop === "symbol" || prop in t) return t[prop];
      const key = fuzzyGetKey(t, String(prop));
      return key ? t[key] : undefined;
    },
    set(t, prop, value) {
      if (typeof prop === "symbol" || Object.prototype.hasOwnProperty.call(t, prop)) {
        t[prop] = value;
        return true;
      }
      const key = fuzzyGetKey(t, String(prop));
      if (key && normalizeText(key) === normalizeText(prop)) t[key] = value;
      else t[prop] = value;
      return true;
    },
  });
}
let activeForcedValueCells = {};
let activeClearedValueCells = {};
let activeSheetProxyCache = new WeakMap();
let activeRowProxyCache = new WeakMap();
const activeOutputFileId = (payload.current && payload.current.outputFileId) || "output:0";
function forcedCellKey(fileId, sheetName, r, c) { return `${fileId}\u0000${sheetName}\u0000${r}\u0000${c}`; }
function addForcedValueCell(fileId, sheetName, r, c, value) {
  if (!fileId || !sheetName) return;
  activeForcedValueCells[forcedCellKey(fileId, sheetName, r, c)] = { fileId, sheetName, r, c, value };
}
function trackClearThenSet(fileId, sheetName, r, c, value) {
  if (!fileId || !sheetName) return;
  const key = forcedCellKey(fileId, sheetName, r, c);
  if (value === "") {
    activeClearedValueCells[key] = true;
    return;
  }
  if (activeClearedValueCells[key]) {
    delete activeClearedValueCells[key];
    addForcedValueCell(fileId, sheetName, r, c, value);
  }
}
function trackedRowProxy(row, fileId, sheetName, r) {
  if (!row || typeof row !== "object") return row;
  const key = `${fileId}\u0000${sheetName}\u0000${r}`;
  let cached = activeRowProxyCache.get(row);
  if (cached && cached[key]) return cached[key];
  if (!cached) { cached = {}; activeRowProxyCache.set(row, cached); }
  cached[key] = new Proxy(row, {
    set(target, prop, value) {
      target[prop] = value;
      const c = Number(prop);
      if (Number.isInteger(c) && c >= 0) trackClearThenSet(fileId, sheetName, r, c, value);
      return true;
    },
  });
  return cached[key];
}
function trackedSheetProxy(sheet, fileId, sheetName) {
  if (!sheet || typeof sheet !== "object") return sheet;
  const key = `${fileId}\u0000${sheetName}`;
  let cached = activeSheetProxyCache.get(sheet);
  if (cached && cached[key]) return cached[key];
  if (!cached) { cached = {}; activeSheetProxyCache.set(sheet, cached); }
  cached[key] = new Proxy(sheet, {
    get(target, prop) {
      const value = target[prop];
      const r = Number(prop);
      if (Number.isInteger(r) && r >= 0 && Array.isArray(value)) return trackedRowProxy(value, fileId, sheetName, r);
      return value;
    },
    set(target, prop, value) { target[prop] = value; return true; },
  });
  return cached[key];
}
function trackedSheetsProxy(sheets, fileId) {
  if (!sheets || typeof sheets !== "object") return sheets;
  return new Proxy(sheets, {
    get(target, prop) {
      if (typeof prop === "symbol") return target[prop];
      const key = Object.prototype.hasOwnProperty.call(target, prop) ? prop : fuzzyGetKey(target, String(prop));
      return key ? trackedSheetProxy(target[key], fileId, String(key)) : undefined;
    },
    set(target, prop, value) {
      if (typeof prop === "symbol" || Object.prototype.hasOwnProperty.call(target, prop)) target[prop] = value;
      else {
        const key = fuzzyGetKey(target, String(prop));
        if (key && normalizeText(key) === normalizeText(prop)) target[key] = value;
        else target[prop] = value;
      }
      return true;
    },
  });
}
const wrappedInputs = {};
Object.entries(inputs).forEach(([fileName, sheets]) => { wrappedInputs[fileName] = trackedSheetsProxy(sheets, `input:${fileName}`); });
const proxiedInputs = fuzzyProxy(wrappedInputs);
const proxiedOutput = trackedSheetsProxy(output, activeOutputFileId);
function findInputBySheet(inputsMap, sheetName, options) {
  options = options || {};
  const target = normalizeText(sheetName);
  const preferredFile = options.preferredFile ? normalizeText(options.preferredFile) : "";
  const matches = [];
  Object.entries(inputsMap || {}).forEach(([fileName, sheets]) => {
    Object.entries(sheets || {}).forEach(([sn, sheet]) => {
      if (normalizeText(sn) === target) matches.push({ fileName, file: sheets, sheetName: sn, sheet });
    });
  });
  if (!matches.length) return null;
  if (preferredFile) {
    const preferred = matches.find(item => normalizeText(item.fileName).includes(preferredFile) || preferredFile.includes(normalizeText(item.fileName)));
    if (preferred) return preferred;
  }
  return matches[0];
}
function resolveTargetSheets(fileRef) {
  if (fileRef === "output") return output;
  let key = String(fileRef || "");
  if (key.startsWith("input:")) key = key.slice(6);
  if (Object.prototype.hasOwnProperty.call(inputs, key)) return inputs[key];
  const fuzzyKey = fuzzyGetKey(inputs, key);
  return fuzzyKey ? inputs[fuzzyKey] : null;
}
function insertColumns(fileRef, sheetName, atColIdx, count) {
  const file = resolveTargetSheets(fileRef);
  if (!file) throw new Error(`insertColumns: file not found: ${fileRef}`);
  const sheet = file[sheetName];
  if (!sheet) throw new Error(`insertColumns: sheet not found: ${sheetName}`);
  const at = Math.max(0, Number(atColIdx) || 0);
  const n = Math.max(0, Number(count) || 0);
  if (!n) return;
  for (let r = 0; r < sheet.length; r++) {
    if (!sheet[r]) sheet[r] = [];
    while (sheet[r].length < at) sheet[r].push("");
    sheet[r].splice(at, 0, ...new Array(n).fill(""));
  }
}
function copyColumns(fileRef, sheetName, srcStart, srcCount, destStart) {
  const file = resolveTargetSheets(fileRef);
  if (!file) throw new Error(`copyColumns: file not found: ${fileRef}`);
  const sheet = file[sheetName];
  if (!sheet) throw new Error(`copyColumns: sheet not found: ${sheetName}`);
  const src = Math.max(0, Number(srcStart) || 0);
  const n = Math.max(0, Number(srcCount) || 0);
  const dest = Math.max(0, Number(destStart) || 0);
  if (!n) return;
  for (let r = 0; r < sheet.length; r++) {
    if (!sheet[r]) sheet[r] = [];
    const values = [];
    for (let c = 0; c < n; c++) values.push(sheet[r][src + c] !== undefined ? sheet[r][src + c] : "");
    while (sheet[r].length < dest) sheet[r].push("");
    for (let c = 0; c < n; c++) sheet[r][dest + c] = values[c];
  }
}
function deleteColumns(fileRef, sheetName, atColIdx, count) {
  const file = resolveTargetSheets(fileRef);
  if (!file) throw new Error(`deleteColumns: file not found: ${fileRef}`);
  const sheet = file[sheetName];
  if (!sheet) throw new Error(`deleteColumns: sheet not found: ${sheetName}`);
  const at = Math.max(0, Number(atColIdx) || 0);
  const n = Math.max(0, Number(count) || 0);
  if (!n) return;
  for (let r = 0; r < sheet.length; r++) {
    if (sheet[r]) sheet[r].splice(at, n);
  }
}
function shiftFormulaText(v) { return v; }
function fileIdForSetCellTarget(fileRef) {
  if (fileRef === "output") return activeOutputFileId;
  let key = String(fileRef || "");
  if (key.startsWith("input:")) return key;
  if (Object.prototype.hasOwnProperty.call(inputs, key)) return `input:${key}`;
  return "";
}
function setCellValue(fileRef, sheetName, r, c, value) {
  const file = resolveTargetSheets(fileRef);
  if (!file) throw new Error(`setCellValue: file not found: ${fileRef}`);
  if (!file[sheetName]) file[sheetName] = [];
  const rowIdx = Math.max(0, Number(r) || 0);
  const colIdx = Math.max(0, Number(c) || 0);
  if (!file[sheetName][rowIdx]) file[sheetName][rowIdx] = [];
  file[sheetName][rowIdx][colIdx] = value;
  addForcedValueCell(fileIdForSetCellTarget(fileRef), sheetName, rowIdx, colIdx, value);
  return value;
}
let activeStepIndex = 0;
const totalSteps = (payload.pipeline || []).filter(step => !(step && step.enabled === false)).length;
for (const step of payload.pipeline || []) {
  if (step && step.enabled === false) continue;
  writeProgress({
    stage: "스킬 실행 중",
    currentStep: activeStepIndex,
    completedSteps: activeStepIndex,
    totalSteps,
    stepRunning: true,
    stepDescription: (step && step.description) || `Step ${activeStepIndex + 1}`
  });
  try {
    const code = String((step && step.code) || "");
    const fn = new Function("inputs", "output", "col", "findColumnGlobal", "findInputBySheet", "similarity", "normalizeText", "replaceNormalizedText", "includesNormalizedText", "equalsNormalizedText", "headerRowIndex", "dataStartRowIndex", "excelRowToIndex", "insertColumns", "copyColumns", "deleteColumns", "shiftFormulaText", "setCellValue",
      code + "\nreturn typeof transform === 'function' ? transform(inputs, output) : { inputs, output };");
    const result = fn(proxiedInputs, proxiedOutput, col, findColumnGlobal, findInputBySheet, similarity, normalizeText, replaceNormalizedText, includesNormalizedText, equalsNormalizedText, headerRowIndex, dataStartRowIndex, excelRowToIndex, insertColumns, copyColumns, deleteColumns, shiftFormulaText, setCellValue);
    if (result && typeof result === "object" && !Array.isArray(result)) {
      if (result.inputs && result.inputs !== proxiedInputs && typeof result.inputs === "object") Object.assign(inputs, result.inputs);
      if (result.output && result.output !== proxiedOutput && typeof result.output === "object") Object.assign(output, result.output);
    }
  } catch (err) {
    const info = {
      stepIdx: activeStepIndex,
      stepId: step && step.id || null,
      description: step && step.description || "",
      code: step && step.code || "",
      message: err && err.message || String(err),
      stack: err && err.stack || "",
    };
    writeProgress({
      stage: "?ㅻ쪟",
      currentStep: activeStepIndex,
      completedSteps: activeStepIndex,
      totalSteps,
      stepRunning: false,
      stepDescription: info.description || `Step ${activeStepIndex + 1}`,
      errorInfo: info,
      error: info.message,
    });
    process.stderr.write(JSON.stringify({ errorInfo: info, error: info.message }));
    process.exit(1);
  }
  activeStepIndex += 1;
  writeProgress({
    stage: "스킬 실행 중",
    currentStep: activeStepIndex,
    completedSteps: activeStepIndex,
    totalSteps,
    stepRunning: false,
    stepDescription: (step && step.description) || `Step ${activeStepIndex}`
  });
}
process.stdout.write(JSON.stringify({ inputs, output, forcedValueCells: Object.values(activeForcedValueCells) }));
"""
    stdout_file = tempfile.NamedTemporaryFile(prefix="kgm_pipeline_stdout_", suffix=".json", delete=False)
    stderr_file = tempfile.NamedTemporaryFile(prefix="kgm_pipeline_stderr_", suffix=".txt", delete=False)
    stdout_path = stdout_file.name
    stderr_path = stderr_file.name
    stdout_file.close()
    stderr_file.close()
    stdout_handle = open(stdout_path, "w+", encoding="utf-8")
    stderr_handle = open(stderr_path, "w+", encoding="utf-8")
    proc = subprocess.Popen(
        ["node", "-e", runner],
        stdin=subprocess.PIPE,
        stdout=stdout_handle,
        stderr=stderr_handle,
        text=True,
        encoding="utf-8",
    )
    stdout = ""
    stderr = ""
    try:
        try:
            proc.stdin.write(json.dumps(payload, ensure_ascii=False))
            proc.stdin.close()
        except Exception:
            pass
        last_progress = None
        started = time.time()
        while proc.poll() is None:
            if time.time() - started > 300:
                proc.kill()
                raise TimeoutError("node pipeline timed out")
            if progress_path and os.path.exists(progress_path):
                try:
                    progress_mtime = os.path.getmtime(progress_path)
                    if progress_mtime != last_progress:
                        last_progress = progress_mtime
                        with open(progress_path, "r", encoding="utf-8") as f:
                            progress = json.load(f)
                        update_pipeline_job(job_id, progress)
                except Exception:
                    pass
            time.sleep(0.2)
        stdout_handle.seek(0)
        stderr_handle.seek(0)
        stdout = stdout_handle.read()
        stderr = stderr_handle.read()
    finally:
        try:
            stdout_handle.close()
            stderr_handle.close()
        except Exception:
            pass
        for temp_path in (stdout_path, stderr_path):
            try:
                os.unlink(temp_path)
            except OSError:
                pass
        if progress_path:
            try:
                os.unlink(progress_path)
            except OSError:
                pass
    if proc.returncode != 0:
        message = (stderr or "").strip() or "node pipeline failed"
        try:
            parsed_error = json.loads(message)
            info = parsed_error.get("errorInfo") or {}
            raise PipelineExecutionError(parsed_error.get("error") or info.get("message") or "node pipeline failed", info)
        except PipelineExecutionError:
            raise
        except Exception:
            pass
        raise RuntimeError(message)
    return json.loads(stdout)


def sheet_dimensions(sheets):
    dimensions = {}
    for name, rows in (sheets or {}).items():
        dimensions[name] = {
            "maxRow": len(rows or []),
            "maxCol": max((len(row or []) for row in (rows or [])), default=0),
            "previewRows": min(len(rows or []), PREVIEW_ROWS),
            "previewCols": max((len(row or []) for row in (rows or [])[:PREVIEW_ROWS]), default=0),
        }
    return dimensions


def build_result_previews(inputs, output, current, diffs=None, forced_value_cells=None):
    diffs = diffs or {}
    forced_value_cells = forced_value_cells or []
    files = []
    for name, sheets in (inputs or {}).items():
        file_id = "input:" + name
        files.append({
            "fileId": file_id,
            "name": name,
            "sheetNames": list((sheets or {}).keys()),
            "sheets": preview_sheets(sheets),
            "forcedValueCells": [cell for cell in forced_value_cells if cell.get("fileId") == file_id],
            "formulas": {},
            "formats": {},
            "dimensions": sheet_dimensions(sheets),
            "diff": diffs.get(file_id),
        })
    if output:
        output_file_id = current.get("outputFileId") or "output:0"
        files.append({
            "fileId": output_file_id,
            "name": "output",
            "sheetNames": list((output or {}).keys()),
            "sheets": preview_sheets(output),
            "forcedValueCells": [cell for cell in forced_value_cells if cell.get("fileId") == output_file_id],
            "formulas": {},
            "formats": {},
            "dimensions": sheet_dimensions(output),
            "diff": diffs.get(output_file_id),
        })
    return files


def preview_sheets(sheets):
    def preview_row(row):
        values = list(row or [])
        return values if PREVIEW_COLS is None else values[:PREVIEW_COLS]
    return {
        name: [preview_row(row) for row in (rows or [])[:PREVIEW_ROWS]]
        for name, rows in (sheets or {}).items()
    }


def diff_value(value):
    if value is None:
        return ""
    if hasattr(value, "isoformat"):
        return value.isoformat()
    return value


def compute_sheet_diff(before_rows, after_rows, max_cells=MAX_DIFF_CELLS_PER_SHEET):
    before_rows = before_rows or []
    after_rows = after_rows or []
    max_rows = max(len(before_rows), len(after_rows))
    cells = []
    changed_count = 0
    truncated = False
    for r_idx in range(max_rows):
        before_row = before_rows[r_idx] if r_idx < len(before_rows) and before_rows[r_idx] else []
        after_row = after_rows[r_idx] if r_idx < len(after_rows) and after_rows[r_idx] else []
        if before_row == after_row:
            continue
        max_cols = max(len(before_row), len(after_row))
        for c_idx in range(max_cols):
            before_value = diff_value(before_row[c_idx] if c_idx < len(before_row) else "")
            after_value = diff_value(after_row[c_idx] if c_idx < len(after_row) else "")
            if before_value == after_value:
                continue
            changed_count += 1
            if len(cells) < max_cells:
                cells.append({"r": r_idx, "c": c_idx, "value": after_value})
            else:
                truncated = True
                return {"cells": cells, "changedCount": changed_count, "truncated": truncated}
    return {"cells": cells, "changedCount": changed_count, "truncated": truncated}


def compute_workbook_diff(before_sheets, after_sheets):
    before_sheets = before_sheets or {}
    after_sheets = after_sheets or {}
    sheet_diffs = {}
    total_changed = 0
    truncated = False
    for sheet_name in sorted(set(before_sheets.keys()) | set(after_sheets.keys())):
        diff = compute_sheet_diff(before_sheets.get(sheet_name), after_sheets.get(sheet_name))
        if diff["changedCount"] or sheet_name not in before_sheets or sheet_name not in after_sheets:
            sheet_diffs[sheet_name] = diff
            total_changed += diff["changedCount"]
            truncated = truncated or diff["truncated"]
    return {"sheets": sheet_diffs, "changedCount": total_changed, "truncated": truncated}


def build_pipeline_diffs(before_inputs, before_output, after_inputs, after_output, current):
    diffs = {}
    for name in sorted(set((before_inputs or {}).keys()) | set((after_inputs or {}).keys())):
        file_id = "input:" + name
        diffs[file_id] = compute_workbook_diff((before_inputs or {}).get(name), (after_inputs or {}).get(name))
    output_file_id = (current or {}).get("outputFileId") or "output:0"
    if before_output or after_output:
        diffs[output_file_id] = compute_workbook_diff(before_output, after_output)
    return diffs


def ensure_node_worker():
    global NODE_WORKER, NODE_WORKER_SCRIPT_MTIME
    worker_path = Path(__file__).with_name("scripts") / "backend-pipeline-worker.js"
    worker_mtime = worker_path.stat().st_mtime if worker_path.exists() else None
    if NODE_WORKER and NODE_WORKER.poll() is None:
        if NODE_WORKER_SCRIPT_MTIME == worker_mtime:
            return NODE_WORKER
        try:
            NODE_WORKER.kill()
        except Exception:
            pass
        NODE_WORKER = None
        NODE_WORKER_READY.clear()
    worker_path = Path(__file__).with_name("scripts") / "backend-pipeline-worker.js"
    if not worker_path.exists():
        raise RuntimeError(f"backend worker not found: {worker_path}")
    NODE_WORKER = subprocess.Popen(
        ["node", str(worker_path)],
        stdin=subprocess.PIPE,
        stdout=subprocess.PIPE,
        stderr=subprocess.DEVNULL,
        text=True,
        encoding="utf-8",
        bufsize=1,
    )
    NODE_WORKER_SCRIPT_MTIME = worker_mtime
    NODE_WORKER_READY.clear()
    return NODE_WORKER


def node_worker_command(command):
    command = dict(command)
    command["id"] = command.get("id") or uuid.uuid4().hex
    with NODE_WORKER_LOCK:
        worker = ensure_node_worker()
        line = json.dumps(command, ensure_ascii=False, separators=(",", ":")) + "\n"
        try:
            worker.stdin.write(line)
            worker.stdin.flush()
            response_line = worker.stdout.readline()
        except Exception:
            try:
                worker.kill()
            except Exception:
                pass
            raise
        if not response_line:
            raise RuntimeError("backend worker stopped")
        response = json.loads(response_line)
        if not response.get("ok"):
            info = response.get("errorInfo") or {}
            if info:
                raise PipelineExecutionError(response.get("error") or info.get("message") or "backend worker failed", info)
            raise RuntimeError(response.get("error") or "backend worker failed")
        return response


def ensure_worker_workbook(wb_record):
    if not wb_record:
        return
    workbook_id = wb_record["id"]
    if workbook_id in NODE_WORKER_READY:
        return
    sheets = get_workbook_aoa_for_run(wb_record, "original")
    current_sheets = wb_record.get("current_aoa_cache")
    node_worker_command({
        "type": "initWorkbook",
        "workbookId": workbook_id,
        "sheets": sheets,
        "currentSheets": current_sheets,
    })
    NODE_WORKER_READY.add(workbook_id)


def export_node_worker_workbook(workbook_id):
    response = node_worker_command({
        "type": "exportWorkbook",
        "payload": {"workbookId": workbook_id},
    })
    return response.get("sheets") or {}


def run_backend_pipeline_payload_with_worker(payload, job_id=None):
    debug_started = time.perf_counter()
    timings = {}
    if os.environ.get("KGM_DISABLE_NODE_WORKER") == "1":
        raise RuntimeError("node worker disabled")
    input_items = payload.get("inputs", [])
    output_item = payload.get("output") or {}
    input_wbs = []
    stage_started = time.perf_counter()
    for idx, item in enumerate(input_items, start=1):
        wb = get_workbook_or_raise(item.get("backendWorkbookId"))
        update_pipeline_job(job_id, {"stage": f"Node 캐시 준비 중 ({idx}/{len(input_items)})"})
        ensure_worker_workbook(wb)
        input_wbs.append((item, wb))
    timings["prepareInputsMs"] = round((time.perf_counter() - stage_started) * 1000, 2)

    output_wb = None
    stage_started = time.perf_counter()
    if output_item.get("backendWorkbookId"):
        update_pipeline_job(job_id, {"stage": "Node 출력 캐시 준비 중"})
        output_wb = get_workbook_or_raise(output_item.get("backendWorkbookId"))
        ensure_worker_workbook(output_wb)
    timings["prepareOutputMs"] = round((time.perf_counter() - stage_started) * 1000, 2)

    active_steps = [s for s in payload.get("pipeline", []) if not (s and s.get("enabled") is False)]
    total_steps = len(active_steps)
    update_pipeline_job(job_id, {
        "stage": "Node 캐시에서 스킬 실행 중",
        "currentStep": 0,
        "completedSteps": 0,
        "totalSteps": total_steps,
        "stepRunning": True,
    })

    worker_payload = {
        "inputs": [
            {
                "name": item.get("name") or wb["name"],
                "backendWorkbookId": wb["id"],
            }
            for item, wb in input_wbs
        ],
        "output": {
            "name": output_item.get("name") or output_wb["name"],
            "backendWorkbookId": output_wb["id"],
        } if output_wb else None,
        "pipeline": payload.get("pipeline", []),
        "baseMode": payload.get("baseMode") or "original",
        "current": payload.get("current") or {},
    }
    stage_started = time.perf_counter()
    response = node_worker_command({
        "type": "runPipeline",
        "payload": worker_payload,
    })
    timings["workerRunAndPreviewMs"] = round((time.perf_counter() - stage_started) * 1000, 2)
    timings["workerCacheHit"] = bool(response.get("cacheHit"))

    update_pipeline_job(job_id, {
        "stage": "미리보기 반영 중",
        "currentStep": total_steps,
        "completedSteps": total_steps,
        "totalSteps": total_steps,
        "stepRunning": False,
    })

    files = response.get("files") or []
    diffs = response.get("diffs") or {}
    forced_value_cells = response.get("forcedValueCells") or []
    if not forced_value_cells:
        for file_result in files:
            forced_value_cells.extend(file_result.get("forcedValueCells") or [])
    stage_started = time.perf_counter()
    diff_id = uuid.uuid4().hex
    DIFFS[diff_id] = {
        "id": diff_id,
        "created": time.time(),
        "diffs": diffs,
        "current": payload.get("current") or {},
    }
    timings["storeDiffMs"] = round((time.perf_counter() - stage_started) * 1000, 2)

    stage_started = time.perf_counter()
    download_urls = {}
    input_wb_by_name = {item.get("name") or wb["name"]: wb for item, wb in input_wbs}
    for file_result in files:
        file_id = file_result.get("fileId")
        worker_workbook_id = file_result.get("workerWorkbookId")
        if not file_id or not worker_workbook_id:
            continue
        if file_id.startswith("input:"):
            input_name = file_id[6:]
            wb = input_wb_by_name.get(input_name) or get_workbook_or_raise(worker_workbook_id)
            wb["node_current"] = True
            result_id = uuid.uuid4().hex
            RESULTS[result_id] = {
                "template_path": str(wb["path"]),
                "workerWorkbookId": worker_workbook_id,
                "forced_value_cells": file_result.get("forcedValueCells") or [],
                "name": f"result_{wb['name']}",
                "created": time.time(),
            }
            download_urls[file_id] = f"/api/workbooks/download/{result_id}"
        elif output_wb:
            output_wb["node_current"] = True
            result_id = uuid.uuid4().hex
            RESULTS[result_id] = {
                "template_path": str(output_wb["path"]),
                "workerWorkbookId": worker_workbook_id,
                "forced_value_cells": file_result.get("forcedValueCells") or [],
                "name": f"result_{output_wb['name']}",
                "created": time.time(),
            }
            download_urls[file_id] = f"/api/workbooks/download/{result_id}"
    timings["downloadRegistrationMs"] = round((time.perf_counter() - stage_started) * 1000, 2)
    timings["totalServerMs"] = round((time.perf_counter() - debug_started) * 1000, 2)

    return {
        "ok": True,
        "worker": True,
        "cacheHit": bool(response.get("cacheHit")),
        "debugTimings": timings,
        "diffId": diff_id,
        "diffs": diffs,
        "downloadId": None,
        "downloadUrl": None,
        "downloadUrls": download_urls,
        "forcedValueCells": forced_value_cells,
        "files": files,
    }


def run_backend_pipeline_payload(payload, job_id=None):
    try:
        return run_backend_pipeline_payload_with_worker(payload, job_id=job_id)
    except PipelineExecutionError:
        raise
    except Exception as worker_err:
        NODE_WORKER_READY.clear()
        update_pipeline_job(job_id, {"stage": f"Node worker fallback: {worker_err}"})

    input_items = payload.get("inputs", [])
    base_mode = payload.get("baseMode") or "original"
    update_pipeline_job(job_id, {
        "stage": "입력 파일 읽는 중",
        "currentStep": 0,
        "completedSteps": 0,
        "stepRunning": False,
    })
    inputs = {}
    for idx, item in enumerate(input_items, start=1):
        wb = get_workbook_or_raise(item.get("backendWorkbookId"))
        update_pipeline_job(job_id, {"stage": f"입력 파일 읽는 중 ({idx}/{len(input_items)})"})
        inputs[item.get("name") or wb["name"]] = get_workbook_aoa_for_run(wb, base_mode)

    update_pipeline_job(job_id, {"stage": "출력 템플릿 읽는 중"})
    output_item = payload.get("output") or {}
    output_wb = get_workbook_or_raise(output_item.get("backendWorkbookId")) if output_item.get("backendWorkbookId") else None
    output = get_workbook_aoa_for_run(output_wb, base_mode) if output_wb else {}

    active_steps = [s for s in payload.get("pipeline", []) if not (s and s.get("enabled") is False)]
    total_steps = len(active_steps)
    update_pipeline_job(job_id, {
        "stage": "스킬 실행 중",
        "currentStep": 0,
        "completedSteps": 0,
        "totalSteps": total_steps,
        "stepRunning": False,
    })
    result = run_js_pipeline_with_node({
        "inputs": inputs,
        "output": output,
        "pipeline": payload.get("pipeline", []),
    }, job_id=job_id)
    result_inputs = result.get("inputs") or inputs
    result_output = result.get("output") or output
    forced_value_cells = result.get("forcedValueCells") or []
    current = payload.get("current") or {}

    update_pipeline_job(job_id, {
        "stage": "diff 계산 중",
        "currentStep": total_steps,
        "completedSteps": total_steps,
        "totalSteps": total_steps,
        "stepRunning": False,
    })
    diffs = build_pipeline_diffs(inputs, output, result_inputs, result_output, current)
    diff_id = uuid.uuid4().hex
    DIFFS[diff_id] = {
        "id": diff_id,
        "created": time.time(),
        "diffs": diffs,
        "current": current,
    }

    update_pipeline_job(job_id, {
        "stage": "다운로드 준비 중",
        "currentStep": total_steps,
        "completedSteps": total_steps,
        "totalSteps": total_steps,
        "stepRunning": False,
    })
    download_urls = {}
    for item in input_items:
        wb = get_workbook_or_raise(item.get("backendWorkbookId"))
        input_name = item.get("name") or wb["name"]
        if input_name not in result_inputs:
            continue
        update_workbook_current_cache(wb, result_inputs[input_name])
        input_download_id = uuid.uuid4().hex
        RESULTS[input_download_id] = {
            "template_path": str(wb["path"]),
            "sheets": result_inputs[input_name],
            "forced_value_cells": [cell for cell in forced_value_cells if cell.get("fileId") == "input:" + input_name],
            "name": f"result_{wb['name']}",
            "created": time.time(),
        }
        download_urls["input:" + input_name] = f"/api/workbooks/download/{input_download_id}"

    download_id = None
    if output_wb:
        output_file_id = (payload.get("current") or {}).get("outputFileId") or "output:0"
        update_workbook_current_cache(output_wb, result_output)
        download_id = uuid.uuid4().hex
        RESULTS[download_id] = {
            "template_path": str(output_wb["path"]),
            "sheets": result_output,
            "forced_value_cells": [cell for cell in forced_value_cells if cell.get("fileId") == output_file_id],
            "name": f"result_{output_wb['name']}",
            "created": time.time(),
        }
        download_urls[output_file_id] = f"/api/workbooks/download/{download_id}"

    update_pipeline_job(job_id, {
        "stage": "미리보기 생성 중",
        "currentStep": total_steps,
        "completedSteps": total_steps,
        "totalSteps": total_steps,
        "stepRunning": False,
    })
    previews = build_result_previews(result_inputs, result_output, current, diffs, forced_value_cells)
    return {
        "ok": True,
        "diffId": diff_id,
        "diffs": diffs,
        "forcedValueCells": forced_value_cells,
        "downloadId": download_id,
        "downloadUrl": f"/api/workbooks/download/{download_id}" if download_id else None,
        "downloadUrls": download_urls,
        "files": previews,
    }


if __name__ == "__main__":
    with socketserver.ThreadingTCPServer((HOST, PORT), KGMHandler) as httpd:
        httpd.allow_reuse_address = True
        print(f"KGM serving on http://{HOST}:{PORT}")
        print(f"Proxying /v1/* to {VLLM_BASE}/v1/*")
        httpd.serve_forever()
